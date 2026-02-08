"""Abuse IP sheets writer."""

from __future__ import annotations

import logging
from datetime import datetime
from typing import TYPE_CHECKING, Any

from .base import BaseSheetWriter, get_column_letter
from .config import HEADERS, SHEET_NAMES, STATUS_CODE_TYPES, SheetConfig

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class AbuseIPSheetWriter(BaseSheetWriter):
    """Creates abuse IP list sheet."""

    def write(self) -> None:
        """Create abuse IP list sheet."""
        try:
            abuse_ips_list, abuse_ip_details = self.get_normalized_abuse_ips()
            if not abuse_ips_list:
                return

            # Get matching abuse IPs (those actually in logs)
            actual_client_ips = set(self.data.get("client_ip_counts", {}).keys())
            matching_ips = actual_client_ips.intersection(set(abuse_ips_list))

            if not matching_ips:
                return

            ws = self.create_sheet(SHEET_NAMES.ABUSE_IP_LIST)
            headers = list(HEADERS.ABUSE_IP)
            self.write_header_row(ws, headers)

            # Sort by request count
            client_ip_counts = self.data.get("client_ip_counts", {})
            sorted_ips = sorted(
                matching_ips,
                key=lambda ip: client_ip_counts.get(ip, 0),
                reverse=True,
            )

            self._write_abuse_ip_rows(ws, sorted_ips, abuse_ip_details)
            self._finalize_abuse_sheet(ws, headers, len(matching_ips))

        except Exception as e:
            logger.error(f"Abuse IP 시트 생성 중 오류: {e}")

    def _write_abuse_ip_rows(
        self,
        ws: Worksheet,
        sorted_ips: list[str],
        abuse_ip_details: dict[str, Any],
    ) -> None:
        """Write abuse IP rows."""
        client_ip_counts = self.data.get("client_ip_counts", {})
        country_mapping = self.data.get("ip_country_mapping", {})
        border = self.styles.thin_border

        for row_idx, ip in enumerate(sorted_ips, start=2):
            details = abuse_ip_details.get(ip, {}) if isinstance(abuse_ip_details, dict) else {}
            request_count = client_ip_counts.get(ip, 0)

            # Count (A)
            cell = ws.cell(row=row_idx, column=1, value=request_count)
            cell.border = border
            cell.alignment = self.styles.align_right
            cell.number_format = "#,##0"

            # IP (B)
            cell = ws.cell(row=row_idx, column=2, value=ip)
            cell.border = border
            cell.alignment = self.styles.align_center

            # Country (C)
            country = country_mapping.get(ip, "N/A")
            cell = ws.cell(row=row_idx, column=3, value=country)
            cell.border = border
            cell.alignment = self.styles.align_center

            # ASN (D)
            cell = ws.cell(row=row_idx, column=4, value=details.get("asn", "N/A"))
            cell.border = border
            cell.alignment = self.styles.align_center

            # ISP (E)
            cell = ws.cell(row=row_idx, column=5, value=details.get("isp", "N/A"))
            cell.border = border
            cell.alignment = self.styles.align_center

    def _finalize_abuse_sheet(
        self,
        ws: Worksheet,
        headers: list[str],
        data_count: int,
    ) -> None:
        """Finalize abuse IP sheet."""
        ws.row_dimensions[1].height = SheetConfig.HEADER_ROW_HEIGHT

        if data_count > 0:
            for row_idx in range(2, 2 + data_count):
                ws.row_dimensions[row_idx].height = SheetConfig.DATA_ROW_HEIGHT

            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}{data_count + 1}"

        self._apply_column_widths(ws, headers)
        ws.freeze_panes = ws.cell(row=2, column=1)  # pyright: ignore[reportAttributeAccessIssue]
        ws.sheet_view.zoomScale = SheetConfig.ZOOM_SCALE


class AbuseRequestsSheetWriter(BaseSheetWriter):
    """Creates abuse IP requests and security events analysis sheet.

    Merges:
    - Abuse IP requests (marked as 'Abuse')
    - Security events: Ambiguous/Severe HTTP classification (Desync detection)

    Color coding:
    - Abuse IP → Red (danger)
    - Ambiguous → Orange (warning)
    - Abuse + Ambiguous → Red (Abuse takes priority)
    """

    def write(self) -> None:
        """Create abuse IP requests and security events sheet."""
        try:
            matching_ips = self.get_matching_abuse_ips()

            # Collect abuse requests
            abuse_requests = self._collect_abuse_requests(matching_ips) if matching_ips else []

            # Collect security events (Ambiguous/Severe)
            security_events = self._collect_security_events()

            # Merge: add classification to abuse requests, merge with security events
            all_requests = self._merge_requests(abuse_requests, security_events, matching_ips)

            if not all_requests:
                return

            ws = self.create_sheet(SHEET_NAMES.ABUSE_REQUESTS)
            headers = list(HEADERS.ABUSE_REQUESTS)
            self.write_header_row(ws, headers)

            # Sort by timestamp and validate Excel limits
            sorted_requests = sorted(all_requests, key=self._safe_timestamp_key)
            sorted_requests = self.truncate_data(sorted_requests, SHEET_NAMES.ABUSE_REQUESTS)

            # Validate column count
            _, col_count = self.validate_excel_limits(len(sorted_requests), len(headers), SHEET_NAMES.ABUSE_REQUESTS)
            if col_count < len(headers):
                headers = headers[:col_count]

            self._write_abuse_request_rows(ws, sorted_requests, headers, matching_ips)

            # Finalize
            if sorted_requests:
                last_col = get_column_letter(len(headers))
                ws.auto_filter.ref = f"A1:{last_col}{len(sorted_requests) + 1}"

            self.apply_wrap_text(ws, headers)
            self._apply_column_widths(ws, headers)
            ws.row_dimensions[1].height = SheetConfig.HEADER_ROW_HEIGHT
            ws.freeze_panes = ws.cell(row=2, column=1)  # pyright: ignore[reportAttributeAccessIssue]
            ws.sheet_view.zoomScale = SheetConfig.ZOOM_SCALE

        except Exception as e:
            logger.error(f"악성 IP/보안 이벤트 시트 생성 중 오류: {e}")

    def _collect_abuse_requests(self, abuse_ips: set[str]) -> list[dict[str, Any]]:
        """Collect all requests from abuse IPs."""
        all_logs = []

        for status_type in STATUS_CODE_TYPES:
            if status_type in self.data and isinstance(self.data[status_type], dict):
                full_logs = self.data[status_type].get("full_logs", [])
                if isinstance(full_logs, list):
                    all_logs.extend(full_logs)

        # Filter for abuse IPs
        return [log for log in all_logs if log.get("client_ip", "N/A") in abuse_ips]

    def _collect_security_events(self) -> list[dict[str, Any]]:
        """Collect security events (Ambiguous/Severe classification)."""
        classification_stats = self.data.get("classification_stats", {})
        if not classification_stats:
            return []
        return classification_stats.get("security_events", [])

    def _merge_requests(
        self,
        abuse_requests: list[dict[str, Any]],
        security_events: list[dict[str, Any]],
        abuse_ips: set[str],
    ) -> list[dict[str, Any]]:
        """Merge abuse requests and security events, avoiding duplicates.

        Classification values:
        - "Abuse": AbuseIPDB registered IP
        - "Ambiguous": ALB Ambiguous classification
        - "Severe": ALB Severe classification
        - "Abuse+Ambiguous": Both AbuseIPDB and Ambiguous
        - "Abuse+Severe": Both AbuseIPDB and Severe

        Reason: Only contains ALB classification_reason (empty for pure Abuse)
        """
        merged: list[dict[str, Any]] = []
        seen_keys: set[tuple] = set()

        def make_key(log: dict[str, Any]) -> tuple:
            """Create unique key for deduplication."""
            ts = log.get("timestamp")
            ts_str = ts.strftime("%Y-%m-%d %H:%M:%S") if hasattr(ts, "strftime") else str(ts)
            return (ts_str, log.get("client_ip", ""), log.get("request", "")[:50])

        # Add abuse requests with 'Abuse' classification
        for log in abuse_requests:
            key = make_key(log)
            if key not in seen_keys:
                seen_keys.add(key)
                log_copy = log.copy()
                log_copy["_classification"] = "Abuse"
                log_copy["_reason"] = ""  # No ALB reason for pure abuse
                log_copy["_is_abuse"] = True
                merged.append(log_copy)

        # Add security events (Ambiguous/Severe)
        for event in security_events:
            key = make_key(event)
            client_ip = event.get("client_ip", "")
            is_abuse = client_ip in abuse_ips
            event_class = event.get("classification", "")
            event_reason = event.get("reason", "")

            if key in seen_keys:
                # Already exists from abuse - update to combined classification
                for log in merged:
                    if make_key(log) == key:
                        if event_class in ("Ambiguous", "Severe"):
                            log["_classification"] = f"Abuse+{event_class}"
                            log["_reason"] = event_reason
                        break
            else:
                seen_keys.add(key)
                # Determine classification
                if is_abuse:
                    classification = f"Abuse+{event_class}" if event_class else "Abuse"
                else:
                    classification = event_class

                log_copy = {
                    "timestamp": event.get("timestamp"),
                    "client_ip": client_ip,
                    "target": event.get("target", ""),
                    "target_group_name": event.get("target_group", ""),
                    "http_method": event.get("method", ""),
                    "request": event.get("url", ""),
                    "user_agent": event.get("user_agent", ""),
                    "elb_status_code": event.get("status_code", ""),
                    "target_status_code": event.get("target_status_code", ""),
                    "_classification": classification,
                    "_reason": event_reason,
                    "_is_abuse": is_abuse,
                }
                merged.append(log_copy)

        return merged

    def _safe_timestamp_key(self, entry: dict[str, Any]) -> datetime:
        """Get timestamp for sorting."""
        ts = entry.get("timestamp")
        if isinstance(ts, datetime):
            return ts
        if isinstance(ts, str):
            try:
                return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        return datetime.min

    def _extract_reason(self, reason: str) -> str:
        """Extract clean ALB classification reason.

        Only returns the actual ALB classification_reason, not "AbuseIPDB".

        Examples:
            - "AbuseIPDB" -> ""
            - "AbuseIPDB + Ambiguous: HeaderNameContainsUnprintableCharacter"
              -> "HeaderNameContainsUnprintableCharacter"
            - "HeaderNameContainsUnprintableCharacter"
              -> "HeaderNameContainsUnprintableCharacter"
        """
        if not reason:
            return ""

        # If it contains a colon, extract the part after it
        if ": " in reason:
            return reason.split(": ", 1)[1]

        # If it's just "AbuseIPDB", return empty
        if reason == "AbuseIPDB":
            return ""

        # Otherwise return as-is (it's an ALB classification_reason)
        return reason

    def _write_abuse_request_rows(
        self,
        ws: Worksheet,
        requests: list[dict[str, Any]],
        headers: list[str],
        abuse_ips: set[str],
    ) -> None:
        """Write abuse request and security event rows with color coding."""
        border = self.styles.thin_border
        abuse_fill = self.styles.fill_abuse  # Red for Abuse IP
        warning_fill = self.styles.warning_fill  # Orange for Ambiguous

        for row_idx, log in enumerate(requests, start=2):
            client_ip = log.get("client_ip", "N/A")
            country = self.get_country_code(client_ip)

            target = log.get("target", "")
            target_field = "" if not target or target == "-" else target
            target_group = log.get("target_group_name", "") or ""

            timestamp_str = self.format_timestamp(log.get("timestamp"))
            classification = log.get("_classification", "")
            reason = log.get("_reason", "")
            reason_display = self._extract_reason(reason)
            method = log.get("http_method", "").replace("-", "")
            request = log.get("request", "N/A")
            user_agent = log.get("user_agent", "N/A")
            elb_status = self.convert_status_code(log.get("elb_status_code", "N/A"))
            backend_status = self.convert_status_code(log.get("target_status_code", "N/A"))

            # Determine row fill color (Abuse > Ambiguous)
            is_abuse = log.get("_is_abuse", False) or client_ip in abuse_ips
            is_ambiguous = classification == "Ambiguous"
            row_fill = abuse_fill if is_abuse else (warning_fill if is_ambiguous else None)

            values = [
                timestamp_str,
                client_ip,
                country,
                classification,
                reason_display,
                target_field,
                target_group,
                method,
                request,
                user_agent,
                elb_status,
                backend_status,
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                # Alignment (12 columns)
                if col_idx in (9, 10):  # Request, User Agent - left align
                    cell.alignment = self.styles.align_left
                elif col_idx in (11, 12):  # Status codes - right align
                    cell.alignment = self.styles.align_right
                    if isinstance(value, int):
                        cell.number_format = "0"
                else:  # All others (1-8) - center align
                    cell.alignment = self.styles.align_center

                # Apply row color
                if row_fill:
                    cell.fill = row_fill
