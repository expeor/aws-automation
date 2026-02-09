"""
functions/analyzers/dynamodb/capacity_mode.py - DynamoDB 용량 모드 분석

Provisioned vs On-Demand 용량 모드 최적화 분석

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.aws.pricing.dynamodb import (
    estimate_ondemand_cost,
    get_dynamodb_monthly_cost,
)
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 분석 기간 (일)
ANALYSIS_DAYS = 14

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "dynamodb:ListTables",
        "dynamodb:DescribeTable",
        "cloudwatch:GetMetricStatistics",
    ],
}


class CapacityRecommendation(Enum):
    """DynamoDB 용량 모드 최적화 권장 사항 분류.

    사용률, 쓰로틀링, Provisioned/On-Demand 비용 비교를 기반으로
    최적 용량 모드를 권장한다.
    """

    KEEP_PROVISIONED = "keep_provisioned"
    SWITCH_TO_ONDEMAND = "switch_to_ondemand"
    SWITCH_TO_PROVISIONED = "switch_to_provisioned"
    REDUCE_CAPACITY = "reduce_capacity"
    INCREASE_CAPACITY = "increase_capacity"
    OPTIMAL = "optimal"


@dataclass
class TableCapacityInfo:
    """DynamoDB 테이블 용량 메타데이터 및 CloudWatch 지표 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        table_name: DynamoDB 테이블 이름.
        table_status: 테이블 상태 (예: ACTIVE).
        billing_mode: 용량 모드 (PROVISIONED 또는 PAY_PER_REQUEST).
        item_count: 테이블 아이템 수.
        size_bytes: 테이블 크기 (bytes).
        provisioned_read: 프로비저닝된 읽기 용량 단위 (RCU).
        provisioned_write: 프로비저닝된 쓰기 용량 단위 (WCU).
        last_increase_dt: 마지막 용량 증가 시각.
        last_decrease_dt: 마지막 용량 감소 시각.
        decreases_today: 오늘 용량 감소 횟수.
        avg_consumed_read: 14일 평균 소비 읽기 용량 (RCU).
        avg_consumed_write: 14일 평균 소비 쓰기 용량 (WCU).
        max_consumed_read: 14일 최대 소비 읽기 용량 (RCU).
        max_consumed_write: 14일 최대 소비 쓰기 용량 (WCU).
        throttled_read: 14일간 읽기 쓰로틀링 횟수.
        throttled_write: 14일간 쓰기 쓰로틀링 횟수.
        created_at: 테이블 생성 시각.
    """

    account_id: str
    account_name: str
    region: str
    table_name: str
    table_status: str
    billing_mode: str  # PROVISIONED or PAY_PER_REQUEST
    item_count: int
    size_bytes: int
    # Provisioned 설정
    provisioned_read: int = 0
    provisioned_write: int = 0
    last_increase_dt: datetime | None = None
    last_decrease_dt: datetime | None = None
    decreases_today: int = 0
    # CloudWatch 지표 (평균)
    avg_consumed_read: float = 0.0
    avg_consumed_write: float = 0.0
    max_consumed_read: float = 0.0
    max_consumed_write: float = 0.0
    throttled_read: float = 0.0
    throttled_write: float = 0.0
    created_at: datetime | None = None

    @property
    def size_mb(self) -> float:
        """테이블 크기를 MB 단위로 반환한다.

        Returns:
            테이블 크기 (MB).
        """
        return self.size_bytes / (1024 * 1024)

    @property
    def read_utilization(self) -> float:
        """읽기 용량 사용률을 백분율로 반환한다.

        Returns:
            읽기 사용률 (%). provisioned_read가 0이면 0.
        """
        if self.provisioned_read <= 0:
            return 0
        return (self.avg_consumed_read / self.provisioned_read) * 100

    @property
    def write_utilization(self) -> float:
        """쓰기 용량 사용률을 백분율로 반환한다.

        Returns:
            쓰기 사용률 (%). provisioned_write가 0이면 0.
        """
        if self.provisioned_write <= 0:
            return 0
        return (self.avg_consumed_write / self.provisioned_write) * 100

    @property
    def estimated_provisioned_cost(self) -> float:
        """Provisioned 모드 기준 월간 예상 비용.

        Returns:
            Provisioned 모드 월간 예상 비용 (USD).
        """
        storage_gb = self.size_bytes / (1024**3)
        return get_dynamodb_monthly_cost(
            region=self.region,
            billing_mode="PROVISIONED",
            rcu=self.provisioned_read,
            wcu=self.provisioned_write,
            storage_gb=storage_gb,
        )

    @property
    def estimated_ondemand_cost(self) -> float:
        """On-Demand 모드 기준 월간 예상 비용.

        Returns:
            On-Demand 모드 월간 예상 비용 (USD).
        """
        storage_gb = self.size_bytes / (1024**3)
        return estimate_ondemand_cost(
            region=self.region,
            avg_consumed_rcu=self.avg_consumed_read,
            avg_consumed_wcu=self.avg_consumed_write,
            storage_gb=storage_gb,
        )


@dataclass
class TableCapacityFinding:
    """개별 DynamoDB 테이블의 용량 모드 분석 결과.

    Attributes:
        table: 분석 대상 테이블 용량 정보.
        recommendation: 용량 모드 권장 사항.
        reason: 권장 사유 설명 문자열.
        potential_savings: 예상 월간 절감액 (USD).
    """

    table: TableCapacityInfo
    recommendation: CapacityRecommendation
    reason: str
    potential_savings: float = 0.0


@dataclass
class CapacityAnalysisResult:
    """단일 계정/리전의 DynamoDB 용량 모드 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        total_tables: 전체 테이블 수.
        provisioned_tables: Provisioned 모드 테이블 수.
        ondemand_tables: On-Demand 모드 테이블 수.
        optimization_candidates: 최적화 대상 테이블 수.
        potential_savings: 예상 총 월간 절감액 (USD).
        findings: 개별 테이블별 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_tables: int = 0
    provisioned_tables: int = 0
    ondemand_tables: int = 0
    optimization_candidates: int = 0
    potential_savings: float = 0.0
    findings: list[TableCapacityFinding] = field(default_factory=list)


def collect_capacity_info(session, account_id: str, account_name: str, region: str) -> list[TableCapacityInfo]:
    """DynamoDB 테이블 용량 정보 및 CloudWatch 지표를 수집한다.

    ListTables + DescribeTable로 프로비저닝 설정을 수집하고,
    CloudWatch GetMetricStatistics로 14일간 소비 용량/쓰로틀링 지표를 조회한다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        TableCapacityInfo 목록.
    """
    from botocore.exceptions import ClientError

    dynamodb = get_client(session, "dynamodb", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    tables = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        paginator = dynamodb.get_paginator("list_tables")
        for page in paginator.paginate():
            for table_name in page.get("TableNames", []):
                try:
                    desc = dynamodb.describe_table(TableName=table_name)
                    t = desc.get("Table", {})

                    billing = t.get("BillingModeSummary", {})
                    billing_mode = billing.get("BillingMode", "PROVISIONED")

                    throughput = t.get("ProvisionedThroughput", {})

                    table = TableCapacityInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        table_name=table_name,
                        table_status=t.get("TableStatus", ""),
                        billing_mode=billing_mode,
                        item_count=t.get("ItemCount", 0),
                        size_bytes=t.get("TableSizeBytes", 0),
                        provisioned_read=throughput.get("ReadCapacityUnits", 0),
                        provisioned_write=throughput.get("WriteCapacityUnits", 0),
                        last_increase_dt=throughput.get("LastIncreaseDateTime"),
                        last_decrease_dt=throughput.get("LastDecreaseDateTime"),
                        decreases_today=throughput.get("NumberOfDecreasesToday", 0),
                        created_at=t.get("CreationDateTime"),
                    )

                    # CloudWatch 지표 조회
                    try:
                        # ConsumedReadCapacityUnits
                        read_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ConsumedReadCapacityUnits",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Average", "Maximum"],
                        )
                        if read_resp.get("Datapoints"):
                            table.avg_consumed_read = sum(d["Average"] for d in read_resp["Datapoints"]) / len(
                                read_resp["Datapoints"]
                            )
                            table.max_consumed_read = max(d["Maximum"] for d in read_resp["Datapoints"])

                        # ConsumedWriteCapacityUnits
                        write_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ConsumedWriteCapacityUnits",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Average", "Maximum"],
                        )
                        if write_resp.get("Datapoints"):
                            table.avg_consumed_write = sum(d["Average"] for d in write_resp["Datapoints"]) / len(
                                write_resp["Datapoints"]
                            )
                            table.max_consumed_write = max(d["Maximum"] for d in write_resp["Datapoints"])

                        # ThrottledRequests (Read)
                        read_throttle = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ReadThrottledRequests",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if read_throttle.get("Datapoints"):
                            table.throttled_read = sum(d["Sum"] for d in read_throttle["Datapoints"])

                        # ThrottledRequests (Write)
                        write_throttle = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="WriteThrottledRequests",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if write_throttle.get("Datapoints"):
                            table.throttled_write = sum(d["Sum"] for d in write_throttle["Datapoints"])

                    except ClientError:
                        pass

                    tables.append(table)

                except ClientError:
                    continue

    except ClientError:
        pass

    return tables


def analyze_capacity(
    tables: list[TableCapacityInfo], account_id: str, account_name: str, region: str
) -> CapacityAnalysisResult:
    """DynamoDB 테이블의 용량 모드를 분석하여 최적화 권장 사항을 도출한다.

    Provisioned/On-Demand 비용을 비교하고, 사용률과 쓰로틀링을 분석하여
    모드 전환, 용량 축소/증가, 또는 최적 상태를 판별한다.

    Args:
        tables: 분석할 테이블 용량 정보 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        용량 분석 결과 집계 (CapacityAnalysisResult).
    """
    result = CapacityAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_tables=len(tables),
    )

    for table in tables:
        if table.billing_mode == "PAY_PER_REQUEST":
            result.ondemand_tables += 1

            # On-Demand → Provisioned 전환 검토
            # 사용량이 일정하고 예측 가능한 경우
            if table.avg_consumed_read > 0 or table.avg_consumed_write > 0:
                savings = table.estimated_ondemand_cost - table.estimated_provisioned_cost
                if savings > 0:
                    result.optimization_candidates += 1
                    result.potential_savings += savings
                    result.findings.append(
                        TableCapacityFinding(
                            table=table,
                            recommendation=CapacityRecommendation.SWITCH_TO_PROVISIONED,
                            reason=f"Provisioned 전환 시 월 ${savings:.2f} 절감 가능",
                            potential_savings=savings,
                        )
                    )
                else:
                    result.findings.append(
                        TableCapacityFinding(
                            table=table,
                            recommendation=CapacityRecommendation.OPTIMAL,
                            reason="On-Demand가 현재 사용 패턴에 적합",
                        )
                    )
            else:
                result.findings.append(
                    TableCapacityFinding(
                        table=table,
                        recommendation=CapacityRecommendation.OPTIMAL,
                        reason="사용량 없음 (삭제 검토 권장)",
                    )
                )

        else:  # PROVISIONED
            result.provisioned_tables += 1

            # Throttling 발생 시 용량 증가 필요
            if table.throttled_read > 0 or table.throttled_write > 0:
                result.findings.append(
                    TableCapacityFinding(
                        table=table,
                        recommendation=CapacityRecommendation.INCREASE_CAPACITY,
                        reason=f"Throttling 발생 (R:{table.throttled_read:.0f}, W:{table.throttled_write:.0f})",
                    )
                )
                continue

            # 사용률 분석
            read_util = table.read_utilization
            write_util = table.write_utilization

            # 매우 낮은 사용률 → On-Demand 전환 검토
            if read_util < 10 and write_util < 10:
                savings = table.estimated_provisioned_cost - table.estimated_ondemand_cost
                if savings > 0:
                    result.optimization_candidates += 1
                    result.potential_savings += savings
                    result.findings.append(
                        TableCapacityFinding(
                            table=table,
                            recommendation=CapacityRecommendation.SWITCH_TO_ONDEMAND,
                            reason=f"저사용 (R:{read_util:.1f}%, W:{write_util:.1f}%) - On-Demand 전환 시 월 ${savings:.2f} 절감",
                            potential_savings=savings,
                        )
                    )
                else:
                    result.findings.append(
                        TableCapacityFinding(
                            table=table,
                            recommendation=CapacityRecommendation.REDUCE_CAPACITY,
                            reason=f"저사용 (R:{read_util:.1f}%, W:{write_util:.1f}%) - 용량 축소 검토",
                        )
                    )
            # 적정 사용률
            elif read_util < 70 and write_util < 70:
                result.findings.append(
                    TableCapacityFinding(
                        table=table,
                        recommendation=CapacityRecommendation.OPTIMAL,
                        reason=f"적정 사용률 (R:{read_util:.1f}%, W:{write_util:.1f}%)",
                    )
                )
            # 높은 사용률
            else:
                result.findings.append(
                    TableCapacityFinding(
                        table=table,
                        recommendation=CapacityRecommendation.INCREASE_CAPACITY,
                        reason=f"높은 사용률 (R:{read_util:.1f}%, W:{write_util:.1f}%) - 용량 증가 검토",
                    )
                )

    return result


def generate_report(results: list[CapacityAnalysisResult], output_dir: str) -> str:
    """DynamoDB 용량 모드 분석 Excel 보고서를 생성한다.

    Summary 시트(계정/리전별 집계)와 Tables 시트(전체 테이블 상세, 권장 사항 포함)를 포함.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        저장된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="Provisioned", width=12, style="number"),
        ColumnDef(header="On-Demand", width=12, style="number"),
        ColumnDef(header="최적화 대상", width=12, style="number"),
        ColumnDef(header="예상 절감액", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_style = Styles.success() if r.potential_savings > 0 else None
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_tables,
                r.provisioned_tables,
                r.ondemand_tables,
                r.optimization_candidates,
                f"${r.potential_savings:,.2f}",
            ],
            style=row_style,
        )
        # Cell-level highlighting for savings
        if r.potential_savings > 0:
            ws = summary_sheet._ws
            ws.cell(row=row_num, column=7).fill = green_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Table Name", width=30),
        ColumnDef(header="Billing Mode", width=15),
        ColumnDef(header="Prov. RCU", width=12, style="number"),
        ColumnDef(header="Prov. WCU", width=12, style="number"),
        ColumnDef(header="Avg R", width=10),
        ColumnDef(header="Avg W", width=10),
        ColumnDef(header="R Util%", width=10),
        ColumnDef(header="W Util%", width=10),
        ColumnDef(header="Prov 비용", width=12),
        ColumnDef(header="OD 비용", width=12),
        ColumnDef(header="권장 사항", width=18, style="center"),
        ColumnDef(header="사유", width=40),
        ColumnDef(header="절감액", width=12),
    ]
    detail_sheet = wb.new_sheet("Tables", detail_columns)

    rec_labels = {
        CapacityRecommendation.SWITCH_TO_ONDEMAND: "On-Demand 전환",
        CapacityRecommendation.SWITCH_TO_PROVISIONED: "Provisioned 전환",
        CapacityRecommendation.REDUCE_CAPACITY: "용량 축소",
        CapacityRecommendation.INCREASE_CAPACITY: "용량 증가",
        CapacityRecommendation.OPTIMAL: "최적",
    }

    for r in results:
        for f in r.findings:
            t = f.table
            row_num = detail_sheet.add_row(
                [
                    t.account_name,
                    t.region,
                    t.table_name,
                    t.billing_mode,
                    t.provisioned_read,
                    t.provisioned_write,
                    f"{t.avg_consumed_read:.1f}",
                    f"{t.avg_consumed_write:.1f}",
                    f"{t.read_utilization:.1f}",
                    f"{t.write_utilization:.1f}",
                    f"${t.estimated_provisioned_cost:.2f}",
                    f"${t.estimated_ondemand_cost:.2f}",
                    rec_labels.get(f.recommendation, f.recommendation.value),
                    f.reason,
                    f"${f.potential_savings:.2f}",
                ],
            )
            # Cell-level highlighting for recommendation
            ws = detail_sheet._ws
            if f.recommendation in (
                CapacityRecommendation.SWITCH_TO_ONDEMAND,
                CapacityRecommendation.SWITCH_TO_PROVISIONED,
            ):
                ws.cell(row=row_num, column=13).fill = green_fill
            elif f.recommendation == CapacityRecommendation.INCREASE_CAPACITY:
                ws.cell(row=row_num, column=13).fill = red_fill
            elif f.recommendation == CapacityRecommendation.REDUCE_CAPACITY:
                ws.cell(row=row_num, column=13).fill = yellow_fill

    return str(wb.save_as(output_dir, "DynamoDB_Capacity"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> CapacityAnalysisResult | None:
    """단일 계정/리전의 DynamoDB 용량 정보를 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 병렬로 실행된다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        용량 분석 결과. 테이블이 없으면 None.
    """
    tables = collect_capacity_info(session, account_id, account_name, region)
    if not tables:
        return None
    return analyze_capacity(tables, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """DynamoDB 용량 모드 분석 도구의 진입점.

    멀티 계정/리전 병렬 수집 후 콘솔 요약 출력, Excel 보고서를 생성한다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 리전, 출력 설정 포함).
    """
    console.print("[bold]DynamoDB 용량 모드 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="dynamodb")
    results: list[CapacityAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_tables = sum(r.total_tables for r in results)
    total_provisioned = sum(r.provisioned_tables for r in results)
    total_ondemand = sum(r.ondemand_tables for r in results)
    total_candidates = sum(r.optimization_candidates for r in results)
    total_savings = sum(r.potential_savings for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체 테이블: {total_tables}개")
    console.print(f"  Provisioned: {total_provisioned}개 / On-Demand: {total_ondemand}개")
    console.print(
        f"최적화 대상: [#FF9900]{total_candidates}개[/#FF9900] (예상 절감: [green]${total_savings:,.2f}/월[/green])"
    )

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("dynamodb", "cost").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
