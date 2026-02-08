"""
functions/analyzers/dynamodb/unused.py - DynamoDB 미사용 테이블 분석

유휴/저사용 DynamoDB 테이블 탐지 (CloudWatch 지표 기반)

탐지 기준:
- 미사용: ConsumedReadCapacityUnits = 0 AND ConsumedWriteCapacityUnits = 0 (7-30일)
- 저사용: Consumed < 30% of Provisioned (AWS nOps 기준)
- https://docs.aws.amazon.com/amazondynamodb/latest/developerguide/metrics-dimensions.html
- https://www.nops.io/underutilized-30-read-write-dynamodb-tables/

CloudWatch 메트릭:
- Namespace: AWS/DynamoDB
- 메트릭: ConsumedReadCapacityUnits, ConsumedWriteCapacityUnits, ThrottledRequests
- Dimension: TableName

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from core.shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from core.shared.aws.pricing.dynamodb import get_dynamodb_monthly_cost
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "dynamodb:ListTables",
        "dynamodb:DescribeTable",
        "cloudwatch:GetMetricStatistics",
    ],
}

# 분석 기간 (일)
ANALYSIS_DAYS = 7
# 저사용 기준: 평균 소비 용량이 프로비저닝의 30% 미만 (AWS nOps 기준)
LOW_USAGE_THRESHOLD_PERCENT = 30


class TableStatus(Enum):
    """DynamoDB 테이블 사용 상태 분류.

    CloudWatch 지표(ConsumedReadCapacityUnits, ConsumedWriteCapacityUnits)를
    기반으로 유휴/저사용/정상 상태를 분류한다.
    """

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"


@dataclass
class TableInfo:
    """DynamoDB 테이블 메타데이터 및 CloudWatch 지표 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        table_name: DynamoDB 테이블 이름.
        table_status: 테이블 상태 (예: ACTIVE).
        billing_mode: 용량 모드 (PROVISIONED 또는 PAY_PER_REQUEST).
        item_count: 테이블 아이템 수.
        size_bytes: 테이블 크기 (bytes).
        provisioned_read: 프로비저닝된 읽기 용량 단위 (RCU).
        provisioned_write: 프로비저닝된 쓰기 용량 단위 (WCU).
        consumed_read: 7일 평균 소비 읽기 용량.
        consumed_write: 7일 평균 소비 쓰기 용량.
        throttled_requests: 7일간 쓰로틀링 발생 횟수.
        created_at: 테이블 생성 시각.
    """

    account_id: str
    account_name: str
    region: str
    table_name: str
    table_status: str
    billing_mode: str  # PROVISIONED or PAY_PER_REQUEST
    item_count: int
    size_bytes: int
    provisioned_read: int = 0
    provisioned_write: int = 0
    # CloudWatch 지표
    consumed_read: float = 0.0
    consumed_write: float = 0.0
    throttled_requests: float = 0.0
    created_at: datetime | None = None

    @property
    def size_mb(self) -> float:
        """테이블 크기를 MB 단위로 반환한다.

        Returns:
            테이블 크기 (MB).
        """
        return self.size_bytes / (1024 * 1024)

    @property
    def estimated_monthly_cost(self) -> float:
        """Pricing 모듈을 사용하여 월간 예상 비용을 계산한다.

        billing_mode에 따라 On-Demand 또는 Provisioned 비용을 계산한다.

        Returns:
            월간 예상 비용 (USD).
        """
        storage_gb = self.size_bytes / (1024**3)

        if self.billing_mode == "PAY_PER_REQUEST":
            # On-Demand: 일평균 소비량 * 30일 * 86400초
            seconds_per_month = 30 * 24 * 3600
            return get_dynamodb_monthly_cost(
                region=self.region,
                billing_mode="PAY_PER_REQUEST",
                read_requests=int(self.consumed_read * seconds_per_month),
                write_requests=int(self.consumed_write * seconds_per_month),
                storage_gb=storage_gb,
            )
        else:
            # Provisioned
            return get_dynamodb_monthly_cost(
                region=self.region,
                billing_mode="PROVISIONED",
                rcu=self.provisioned_read,
                wcu=self.provisioned_write,
                storage_gb=storage_gb,
            )


@dataclass
class TableFinding:
    """개별 DynamoDB 테이블의 분석 결과.

    Attributes:
        table: 분석 대상 테이블 정보.
        status: 분석된 사용 상태.
        recommendation: 권장 조치 사항 문자열.
    """

    table: TableInfo
    status: TableStatus
    recommendation: str


@dataclass
class DynamoDBAnalysisResult:
    """단일 계정/리전의 DynamoDB 미사용 테이블 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        total_tables: 전체 테이블 수.
        unused_tables: 미사용 테이블 수.
        low_usage_tables: 저사용 테이블 수.
        normal_tables: 정상 테이블 수.
        unused_monthly_cost: 미사용 테이블 월간 비용 합계 (USD).
        low_usage_monthly_cost: 저사용 테이블 월간 비용 합계 (USD).
        findings: 개별 테이블별 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_tables: int = 0
    unused_tables: int = 0
    low_usage_tables: int = 0
    normal_tables: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[TableFinding] = field(default_factory=list)


def collect_dynamodb_tables(session, account_id: str, account_name: str, region: str) -> list[TableInfo]:
    """DynamoDB 테이블 목록 수집 및 CloudWatch 메트릭 배치 조회.

    ListTables + DescribeTable로 테이블 메타데이터를 수집한 후,
    GetMetricData API를 사용하여 CloudWatch 지표를 배치로 조회한다.

    최적화:
    - 기존: 테이블당 3 API 호출 -> 최적화: 전체 1-2 API 호출

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        TableInfo 목록.
    """
    from botocore.exceptions import ClientError

    dynamodb = get_client(session, "dynamodb", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    tables: list[TableInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 테이블 목록 수집
        paginator = dynamodb.get_paginator("list_tables")
        for page in paginator.paginate():
            for table_name in page.get("TableNames", []):
                try:
                    # 테이블 상세 정보 조회
                    desc = dynamodb.describe_table(TableName=table_name)
                    t = desc.get("Table", {})

                    billing = t.get("BillingModeSummary", {})
                    billing_mode = billing.get("BillingMode", "PROVISIONED")

                    throughput = t.get("ProvisionedThroughput", {})

                    table = TableInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        table_name=table_name,
                        table_status=t.get("TableStatus", ""),
                        billing_mode=billing_mode,
                        item_count=t.get("ItemCount", 0),
                        size_bytes=t.get("TableSizeBytes", 0),
                        provisioned_read=throughput.get("ReadCapacityUnits", 0),
                        provisioned_write=throughput.get("WriteCapacityUnits", 0),
                        created_at=t.get("CreationDateTime"),
                    )
                    tables.append(table)

                except ClientError as e:
                    category = categorize_error(e)
                    if category == ErrorCategory.ACCESS_DENIED:
                        logger.info(f"DynamoDB 권한 없음: {table_name} ({get_error_code(e)})")
                    elif category != ErrorCategory.NOT_FOUND:
                        logger.warning(f"DynamoDB 테이블 조회 오류: {table_name} ({get_error_code(e)})")
                    continue

        # 2단계: 배치 메트릭 조회
        if tables:
            _collect_dynamodb_metrics_batch(cloudwatch, tables, start_time, now)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"DynamoDB ListTables 권한 없음: {account_name}/{region}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"DynamoDB API 쓰로틀링: {account_name}/{region}")
        else:
            logger.warning(f"DynamoDB 목록 조회 오류: {account_name}/{region} ({get_error_code(e)})")

    return tables


def _collect_dynamodb_metrics_batch(
    cloudwatch,
    tables: list[TableInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """DynamoDB 테이블의 CloudWatch 메트릭을 배치로 수집한다.

    ConsumedReadCapacityUnits, ConsumedWriteCapacityUnits, ThrottledRequests
    메트릭을 GetMetricData API로 한 번에 조회하여 각 테이블에 매핑한다.

    최적화:
    - 기존: 테이블당 3 API 호출 (read, write, throttle)
    - 최적화: 전체 테이블 1-2 API 호출 (500개 단위 자동 분할)

    Args:
        cloudwatch: CloudWatch boto3 client.
        tables: 메트릭을 수집할 DynamoDB 테이블 목록.
        start_time: 조회 시작 시각 (UTC).
        end_time: 조회 종료 시각 (UTC).
    """
    from botocore.exceptions import ClientError

    # 쿼리 생성
    queries: list[MetricQuery] = []
    for table in tables:
        safe_id = sanitize_metric_id(table.table_name)
        dimensions = {"TableName": table.table_name}

        # ConsumedReadCapacityUnits (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_read",
                namespace="AWS/DynamoDB",
                metric_name="ConsumedReadCapacityUnits",
                dimensions=dimensions,
                stat="Sum",
            )
        )

        # ConsumedWriteCapacityUnits (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_write",
                namespace="AWS/DynamoDB",
                metric_name="ConsumedWriteCapacityUnits",
                dimensions=dimensions,
                stat="Sum",
            )
        )

        # ThrottledRequests (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_throttle",
                namespace="AWS/DynamoDB",
                metric_name="ThrottledRequests",
                dimensions=dimensions,
                stat="Sum",
            )
        )

    try:
        # 배치 조회 (내장 pagination + retry)
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        # 결과 매핑 - 일평균 계산
        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for table in tables:
            safe_id = sanitize_metric_id(table.table_name)

            # 일평균 소비 용량 (Sum / days)
            table.consumed_read = results.get(f"{safe_id}_read", 0.0) / days
            table.consumed_write = results.get(f"{safe_id}_write", 0.0) / days
            # 총 쓰로틀 횟수
            table.throttled_requests = results.get(f"{safe_id}_throttle", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_tables(tables: list[TableInfo], account_id: str, account_name: str, region: str) -> DynamoDBAnalysisResult:
    """DynamoDB 테이블을 CloudWatch 지표 기준으로 분석하여 유휴/저사용을 판별한다.

    미사용: 읽기/쓰기 소비 용량이 모두 0.
    저사용: Provisioned 모드에서 사용량이 프로비저닝의 30% 미만 (AWS nOps 기준).

    Args:
        tables: 분석할 DynamoDB 테이블 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        분석 결과 집계 (DynamoDBAnalysisResult).
    """
    result = DynamoDBAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_tables=len(tables),
    )

    for table in tables:
        # 미사용: 읽기/쓰기 모두 0
        if table.consumed_read == 0 and table.consumed_write == 0:
            result.unused_tables += 1
            result.unused_monthly_cost += table.estimated_monthly_cost
            result.findings.append(
                TableFinding(
                    table=table,
                    status=TableStatus.UNUSED,
                    recommendation=f"읽기/쓰기 없음 - 삭제 검토 (${table.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: Provisioned 모드에서 사용량이 프로비저닝의 10% 미만
        if table.billing_mode == "PROVISIONED":
            read_usage = (table.consumed_read / table.provisioned_read * 100) if table.provisioned_read > 0 else 0
            write_usage = (table.consumed_write / table.provisioned_write * 100) if table.provisioned_write > 0 else 0

            if read_usage < LOW_USAGE_THRESHOLD_PERCENT and write_usage < LOW_USAGE_THRESHOLD_PERCENT:
                result.low_usage_tables += 1
                result.low_usage_monthly_cost += table.estimated_monthly_cost
                result.findings.append(
                    TableFinding(
                        table=table,
                        status=TableStatus.LOW_USAGE,
                        recommendation=f"저사용 (R:{read_usage:.1f}%, W:{write_usage:.1f}%) - On-Demand 전환 또는 용량 축소 검토",
                    )
                )
                continue

        result.normal_tables += 1
        result.findings.append(
            TableFinding(
                table=table,
                status=TableStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[DynamoDBAnalysisResult], output_dir: str) -> str:
    """DynamoDB 미사용 테이블 분석 Excel 보고서를 생성한다.

    Summary 시트(계정/리전별 집계)와 Tables 시트(미사용/저사용 테이블 상세)를 포함.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        저장된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_style = None
        if r.unused_tables > 0:
            row_style = Styles.danger()
        elif r.low_usage_tables > 0:
            row_style = Styles.warning()
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_tables,
                r.unused_tables,
                r.low_usage_tables,
                r.normal_tables,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
            ],
            style=row_style,
        )
        # Cell-level highlighting
        ws = summary_sheet._ws
        if r.unused_tables > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_tables > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Table Name", width=30),
        ColumnDef(header="Status", width=12, style="center"),
        ColumnDef(header="Billing Mode", width=15),
        ColumnDef(header="Items", width=12, style="number"),
        ColumnDef(header="Size (MB)", width=12),
        ColumnDef(header="Prov. RCU", width=12, style="number"),
        ColumnDef(header="Prov. WCU", width=12, style="number"),
        ColumnDef(header="Consumed R", width=12),
        ColumnDef(header="Consumed W", width=12),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Tables", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != TableStatus.NORMAL:
                t = f.table
                row_style = Styles.danger() if f.status == TableStatus.UNUSED else Styles.warning()
                detail_sheet.add_row(
                    [
                        t.account_name,
                        t.region,
                        t.table_name,
                        f.status.value,
                        t.billing_mode,
                        t.item_count,
                        f"{t.size_mb:.2f}",
                        t.provisioned_read,
                        t.provisioned_write,
                        f"{t.consumed_read:.1f}",
                        f"{t.consumed_write:.1f}",
                        f"${t.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ],
                    style=row_style,
                )

    return str(wb.save_as(output_dir, "DynamoDB_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> DynamoDBAnalysisResult | None:
    """단일 계정/리전의 DynamoDB 테이블을 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 병렬로 실행된다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        분석 결과. 테이블이 없으면 None.
    """
    tables = collect_dynamodb_tables(session, account_id, account_name, region)
    if not tables:
        return None
    return analyze_tables(tables, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """DynamoDB 미사용 테이블 분석 도구의 진입점.

    멀티 계정/리전 병렬 수집 후 콘솔 요약 출력, Excel 보고서를 생성한다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 리전, 출력 설정 포함).
    """
    console.print("[bold]DynamoDB 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="dynamodb")
    results: list[DynamoDBAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_tables for r in results)
    total_low = sum(r.low_usage_tables for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월)"
    )

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("dynamodb", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
