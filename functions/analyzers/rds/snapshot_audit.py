"""
functions/analyzers/rds/snapshot_audit.py - RDS Snapshot 미사용 분석

오래된 수동 RDS/Aurora 스냅샷 탐지

분석 기준:
- 수동 스냅샷만 (자동 스냅샷 제외)
- 14일 이상 경과

비용:
- RDS: $0.02/GB-월
- Aurora: $0.021/GB-월

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.shared.aws.pricing import get_rds_snapshot_monthly_cost
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

logger = logging.getLogger(__name__)
console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "rds:DescribeDBSnapshots",
        "rds:DescribeDBClusterSnapshots",
    ],
}


# =============================================================================
# 상수
# =============================================================================

# 오래된 스냅샷 기준 (일)
OLD_SNAPSHOT_DAYS = 14


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """RDS 스냅샷 사용 상태 분류.

    OLD_SNAPSHOT_DAYS 기준으로 오래된 스냅샷과 최근 스냅샷을 구분한다.
    """

    OLD = "old"  # 오래됨
    NORMAL = "normal"  # 최근


class SnapshotType(Enum):
    """RDS 스냅샷 유형 분류.

    RDS 인스턴스 스냅샷과 Aurora 클러스터 스냅샷을 구분한다.
    """

    RDS = "rds"
    AURORA = "aurora"


class Severity(Enum):
    """스냅샷 분석 심각도 분류.

    스냅샷 크기(allocated_storage_gb)를 기준으로 분류한다.
    500GB 이상이면 HIGH, 100GB 이상이면 MEDIUM, 그 외 LOW.
    """

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class RDSSnapshotInfo:
    """RDS/Aurora 수동 스냅샷 메타데이터.

    Attributes:
        id: 스냅샷 식별자.
        db_identifier: 원본 DB 인스턴스/클러스터 식별자.
        snapshot_type: 스냅샷 유형 (RDS 또는 AURORA).
        engine: DB 엔진 (예: mysql, aurora-postgresql).
        engine_version: DB 엔진 버전.
        status: 스냅샷 상태 (예: available).
        create_time: 스냅샷 생성 시각.
        allocated_storage_gb: 할당된 스토리지 용량 (GB).
        encrypted: 암호화 여부.
        arn: 스냅샷 ARN.
        tags: AWS 태그 딕셔너리 (aws: 접두사 제외).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        monthly_cost: 월간 스토리지 비용 (USD).
    """

    id: str
    db_identifier: str
    snapshot_type: SnapshotType
    engine: str
    engine_version: str
    status: str
    create_time: datetime | None
    allocated_storage_gb: int
    encrypted: bool
    arn: str
    tags: dict[str, str]

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def age_days(self) -> int:
        """스냅샷 생성 후 경과 일수.

        Returns:
            경과 일수. create_time이 없으면 0.
        """
        if not self.create_time:
            return 0
        now = datetime.now(timezone.utc)
        delta = now - self.create_time.replace(tzinfo=timezone.utc)
        return delta.days

    @property
    def is_old(self) -> bool:
        """OLD_SNAPSHOT_DAYS 기준 초과 여부.

        Returns:
            age_days >= OLD_SNAPSHOT_DAYS이면 True.
        """
        return self.age_days >= OLD_SNAPSHOT_DAYS


@dataclass
class RDSSnapshotFinding:
    """개별 스냅샷의 분석 결과.

    Attributes:
        snapshot: 분석 대상 스냅샷 정보.
        usage_status: 사용 상태 (OLD 또는 NORMAL).
        severity: 심각도 (스토리지 크기 기준).
        description: 분석 결과 설명 문자열.
        recommendation: 권장 조치 사항 문자열.
    """

    snapshot: RDSSnapshotInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class RDSSnapshotAnalysisResult:
    """단일 계정/리전의 RDS 스냅샷 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        findings: 개별 스냅샷별 분석 결과 목록.
        total_count: 전체 수동 스냅샷 수.
        old_count: 오래된 스냅샷 수.
        normal_count: 최근 스냅샷 수.
        total_size_gb: 전체 스냅샷 용량 (GB).
        old_size_gb: 오래된 스냅샷 용량 (GB).
        old_monthly_cost: 오래된 스냅샷 월간 비용 (USD).
    """

    account_id: str
    account_name: str
    region: str
    findings: list[RDSSnapshotFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    old_count: int = 0
    normal_count: int = 0

    # 용량/비용
    total_size_gb: int = 0
    old_size_gb: int = 0
    old_monthly_cost: float = 0.0


# =============================================================================
# 수집
# =============================================================================


def collect_rds_snapshots(session, account_id: str, account_name: str, region: str) -> list[RDSSnapshotInfo]:
    """RDS 및 Aurora 수동 스냅샷을 수집한다.

    DescribeDBSnapshots(RDS)와 DescribeDBClusterSnapshots(Aurora) API로
    수동 스냅샷만 필터링하여 수집하고, 각 스냅샷의 태그와 비용을 계산한다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        RDSSnapshotInfo 목록 (available 상태의 수동 스냅샷만).
    """
    from botocore.exceptions import ClientError

    snapshots = []

    try:
        rds = get_client(session, "rds", region_name=region)

        # RDS 인스턴스 스냅샷 (수동)
        try:
            paginator = rds.get_paginator("describe_db_snapshots")
            for page in paginator.paginate(SnapshotType="manual"):
                for data in page.get("DBSnapshots", []):
                    if data.get("Status") != "available":
                        continue

                    snap = _parse_rds_snapshot(data, rds, account_id, account_name, region)
                    if snap:
                        snapshots.append(snap)
        except ClientError as e:
            error_code = e.response.get("Error", {}).get("Code", "Unknown")
            if not is_quiet():
                console.print(f"[yellow]{account_name}/{region} RDS 스냅샷 수집 오류: {error_code}[/yellow]")

        # Aurora 클러스터 스냅샷 (수동)
        try:
            paginator = rds.get_paginator("describe_db_cluster_snapshots")
            for page in paginator.paginate(SnapshotType="manual"):
                for data in page.get("DBClusterSnapshots", []):
                    if data.get("Status") != "available":
                        continue

                    snap = _parse_aurora_snapshot(data, rds, account_id, account_name, region)
                    if snap:
                        snapshots.append(snap)
        except ClientError as e:
            error_code = e.response.get("Error", {}).get("Code", "Unknown")
            if not is_quiet():
                console.print(f"[yellow]{account_name}/{region} Aurora 스냅샷 수집 오류: {error_code}[/yellow]")

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"[yellow]{account_name}/{region} 수집 오류: {error_code}[/yellow]")

    return snapshots


def _parse_rds_snapshot(data: dict, rds, account_id: str, account_name: str, region: str) -> RDSSnapshotInfo | None:
    """RDS 인스턴스 스냅샷 API 응답을 RDSSnapshotInfo로 변환한다.

    Args:
        data: DescribeDBSnapshots API 응답의 단일 스냅샷 딕셔너리.
        rds: RDS boto3 client (태그 조회용).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        파싱된 스냅샷 정보. 파싱 실패 시 None.
    """
    try:
        arn = data.get("DBSnapshotArn", "")
        allocated_storage = data.get("AllocatedStorage", 0)
        monthly_cost = get_rds_snapshot_monthly_cost(region, allocated_storage, is_aurora=False)

        # 태그 조회
        tags = {}
        from botocore.exceptions import ClientError

        try:
            tag_response = rds.list_tags_for_resource(ResourceName=arn)
            tags = {t["Key"]: t["Value"] for t in tag_response.get("TagList", []) if not t["Key"].startswith("aws:")}
        except ClientError:
            pass

        return RDSSnapshotInfo(
            id=data.get("DBSnapshotIdentifier", ""),
            db_identifier=data.get("DBInstanceIdentifier", ""),
            snapshot_type=SnapshotType.RDS,
            engine=data.get("Engine", ""),
            engine_version=data.get("EngineVersion", ""),
            status=data.get("Status", ""),
            create_time=data.get("SnapshotCreateTime"),
            allocated_storage_gb=allocated_storage,
            encrypted=data.get("Encrypted", False),
            arn=arn,
            tags=tags,
            account_id=account_id,
            account_name=account_name,
            region=region,
            monthly_cost=monthly_cost,
        )
    except Exception as e:
        logger.debug("Failed to parse RDS snapshot: %s", e)
        return None


def _parse_aurora_snapshot(data: dict, rds, account_id: str, account_name: str, region: str) -> RDSSnapshotInfo | None:
    """Aurora 클러스터 스냅샷 API 응답을 RDSSnapshotInfo로 변환한다.

    Args:
        data: DescribeDBClusterSnapshots API 응답의 단일 스냅샷 딕셔너리.
        rds: RDS boto3 client (태그 조회용).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        파싱된 스냅샷 정보. 파싱 실패 시 None.
    """
    try:
        arn = data.get("DBClusterSnapshotArn", "")
        allocated_storage = data.get("AllocatedStorage", 0)
        monthly_cost = get_rds_snapshot_monthly_cost(region, allocated_storage, is_aurora=True)

        # 태그 조회
        tags = {}
        from botocore.exceptions import ClientError

        try:
            tag_response = rds.list_tags_for_resource(ResourceName=arn)
            tags = {t["Key"]: t["Value"] for t in tag_response.get("TagList", []) if not t["Key"].startswith("aws:")}
        except ClientError:
            pass

        return RDSSnapshotInfo(
            id=data.get("DBClusterSnapshotIdentifier", ""),
            db_identifier=data.get("DBClusterIdentifier", ""),
            snapshot_type=SnapshotType.AURORA,
            engine=data.get("Engine", ""),
            engine_version=data.get("EngineVersion", ""),
            status=data.get("Status", ""),
            create_time=data.get("SnapshotCreateTime"),
            allocated_storage_gb=allocated_storage,
            encrypted=data.get("StorageEncrypted", False),
            arn=arn,
            tags=tags,
            account_id=account_id,
            account_name=account_name,
            region=region,
            monthly_cost=monthly_cost,
        )
    except Exception as e:
        logger.debug("Failed to parse Aurora snapshot: %s", e)
        return None


# =============================================================================
# 분석
# =============================================================================


def analyze_rds_snapshots(
    snapshots: list[RDSSnapshotInfo], account_id: str, account_name: str, region: str
) -> RDSSnapshotAnalysisResult:
    """수동 스냅샷을 경과 일수 기준으로 분석한다.

    OLD_SNAPSHOT_DAYS 이상 경과한 스냅샷을 오래된 것으로 분류하고,
    크기에 따라 심각도를 결정한다.

    Args:
        snapshots: 분석할 스냅샷 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        분석 결과 집계 (RDSSnapshotAnalysisResult).
    """
    result = RDSSnapshotAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for snapshot in snapshots:
        finding = _analyze_single_snapshot(snapshot)
        result.findings.append(finding)
        result.total_size_gb += snapshot.allocated_storage_gb

        if finding.usage_status == UsageStatus.OLD:
            result.old_count += 1
            result.old_size_gb += snapshot.allocated_storage_gb
            result.old_monthly_cost += snapshot.monthly_cost
        else:
            result.normal_count += 1

    result.total_count = len(snapshots)
    return result


def _analyze_single_snapshot(snapshot: RDSSnapshotInfo) -> RDSSnapshotFinding:
    """개별 스냅샷의 경과 일수와 크기를 기준으로 분석한다.

    Args:
        snapshot: 분석 대상 스냅샷.

    Returns:
        스냅샷 분석 결과 (RDSSnapshotFinding).
    """

    # 최근 생성
    if not snapshot.is_old:
        return RDSSnapshotFinding(
            snapshot=snapshot,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"최근 생성 ({snapshot.age_days}일)",
            recommendation="모니터링",
        )

    # 오래됨
    if snapshot.allocated_storage_gb >= 500:
        severity = Severity.HIGH
    elif snapshot.allocated_storage_gb >= 100:
        severity = Severity.MEDIUM
    else:
        severity = Severity.LOW

    type_str = "Aurora" if snapshot.snapshot_type == SnapshotType.AURORA else "RDS"
    return RDSSnapshotFinding(
        snapshot=snapshot,
        usage_status=UsageStatus.OLD,
        severity=severity,
        description=f"오래된 {type_str} 스냅샷 ({snapshot.age_days}일, {snapshot.allocated_storage_gb}GB, ${snapshot.monthly_cost:.2f}/월)",
        recommendation="필요 여부 검토 후 삭제",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[RDSSnapshotAnalysisResult], output_dir: str) -> str:
    """RDS 스냅샷 미사용 분석 Excel 보고서를 생성한다.

    Summary 시트(총괄 통계)와 Findings 시트(오래된 스냅샷 상세, 비용순 정렬)를 포함.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        저장된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 스타일
    status_fills = {
        UsageStatus.OLD: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # Summary
    totals = {
        "total": sum(r.total_count for r in results),
        "old": sum(r.old_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "total_size": sum(r.total_size_gb for r in results),
        "old_size": sum(r.old_size_gb for r in results),
        "old_cost": sum(r.old_monthly_cost for r in results),
    }

    summary = wb.new_summary_sheet("Summary")
    summary.add_title("RDS Snapshot 미사용 분석 보고서")
    summary.add_item("생성일", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()
    summary.add_section("분석 결과")
    summary.add_item("전체 스냅샷", totals["total"])
    summary.add_item("오래된 스냅샷", totals["old"], highlight="warning" if totals["old"] > 0 else None)
    summary.add_item("최근 스냅샷", totals["normal"])
    summary.add_item("전체 용량 (GB)", totals["total_size"])
    summary.add_item("오래된 용량 (GB)", totals["old_size"], highlight="warning" if totals["old_size"] > 0 else None)
    summary.add_item(
        "오래된 월 비용 ($)", f"${totals['old_cost']:.2f}", highlight="danger" if totals["old_cost"] > 0 else None
    )

    # Findings
    finding_columns = [
        ColumnDef(header="Account", width=15),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Snapshot ID", width=35),
        ColumnDef(header="DB Identifier", width=25),
        ColumnDef(header="Type", width=10, style="center"),
        ColumnDef(header="Usage", width=10, style="center"),
        ColumnDef(header="Engine", width=20),
        ColumnDef(header="Size (GB)", width=12, style="number"),
        ColumnDef(header="Age (days)", width=12, style="number"),
        ColumnDef(header="Monthly Cost ($)", width=15, style="currency"),
        ColumnDef(header="Encrypted", width=10, style="center"),
        ColumnDef(header="Created", width=12),
        ColumnDef(header="Recommendation", width=25),
    ]
    finding_sheet = wb.new_sheet("Findings", finding_columns)

    # 오래된 것만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status == UsageStatus.OLD:
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.snapshot.monthly_cost, reverse=True)

    for f in all_findings:
        snap = f.snapshot
        row_num = finding_sheet.add_row(
            [
                snap.account_name,
                snap.region,
                snap.id,
                snap.db_identifier,
                snap.snapshot_type.value.upper(),
                f.usage_status.value,
                f"{snap.engine} {snap.engine_version}",
                snap.allocated_storage_gb,
                snap.age_days,
                round(snap.monthly_cost, 2),
                "Yes" if snap.encrypted else "No",
                snap.create_time.strftime("%Y-%m-%d") if snap.create_time else "",
                f.recommendation,
            ],
            style=Styles.warning(),
        )

        fill = status_fills.get(f.usage_status)
        if fill:
            finding_sheet._ws.cell(row=row_num, column=6).fill = fill

    return str(wb.save_as(output_dir, "RDS_Snapshot_Unused"))


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> RDSSnapshotAnalysisResult | None:
    """단일 계정/리전의 RDS 수동 스냅샷을 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 병렬로 실행된다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        분석 결과. 스냅샷이 없으면 None.
    """
    snapshots = collect_rds_snapshots(session, account_id, account_name, region)
    if not snapshots:
        return None
    return analyze_rds_snapshots(snapshots, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """RDS Snapshot 미사용 분석 도구의 진입점.

    멀티 계정/리전 병렬 수집 후 콘솔 요약 출력, Excel 보고서를 생성한다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 리전, 출력 설정 포함).
    """
    console.print("[bold]RDS Snapshot 미사용 분석 시작...[/bold]")
    console.print(f"  [dim]기준: {OLD_SNAPSHOT_DAYS}일 이상 오래된 수동 스냅샷[/dim]")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="rds")
    all_results: list[RDSSnapshotAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not all_results:
        console.print("[yellow]분석할 RDS 스냅샷 없음[/yellow]")
        return

    # 요약
    totals = {
        "total": sum(r.total_count for r in all_results),
        "old": sum(r.old_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "total_size": sum(r.total_size_gb for r in all_results),
        "old_size": sum(r.old_size_gb for r in all_results),
        "old_cost": sum(r.old_monthly_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 RDS 스냅샷: {totals['total']}개 ({totals['total_size']}GB)[/bold]")
    if totals["old"] > 0:
        console.print(
            f"  [yellow bold]오래됨: {totals['old']}개 ({totals['old_size']}GB, ${totals['old_cost']:.2f}/월)[/yellow bold]"
        )
    console.print(f"  [green]최근: {totals['normal']}개[/green]")

    # 보고서
    console.print("\n[#FF9900]Excel 보고서 생성 중...[/#FF9900]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("rds", "snapshot").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
