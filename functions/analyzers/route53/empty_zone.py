"""
functions/analyzers/route53/empty_zone.py - 빈 Hosted Zone 탐지

레코드가 없는 미사용 Route53 Hosted Zone 분석

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.aws.pricing import get_hosted_zone_price
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "route53:ListHostedZones",
        "route53:GetHostedZone",
        "route53:ListResourceRecordSets",
    ],
}


class ZoneStatus(Enum):
    """Route53 Hosted Zone 분석 상태.

    레코드 존재 여부 및 유형 기반으로 분류한다.

    Attributes:
        NORMAL: 실제 레코드가 존재하는 정상 Zone.
        EMPTY: 레코드가 전혀 없는 빈 Zone.
        NS_SOA_ONLY: NS/SOA 레코드만 존재하는 실질적 미사용 Zone.
    """

    NORMAL = "normal"
    EMPTY = "empty"
    NS_SOA_ONLY = "ns_soa_only"


@dataclass
class HostedZoneInfo:
    """Route53 Hosted Zone 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        zone_id: Hosted Zone ID.
        name: 도메인 이름.
        is_private: Private Hosted Zone 여부.
        record_count: 레코드 셋 수 (NS/SOA 포함).
        comment: Zone 설명.
        vpcs: Private Zone에 연결된 VPC 목록 (리전:VPC-ID 형식).
        has_real_records: NS/SOA 외 실제 레코드 존재 여부.
    """

    account_id: str
    account_name: str
    zone_id: str
    name: str
    is_private: bool
    record_count: int
    comment: str
    vpcs: list[str] = field(default_factory=list)
    has_real_records: bool = False

    @property
    def monthly_cost(self) -> float:
        """Hosted Zone 월간 비용 (USD).

        Returns:
            첫 25개 기준 Zone 호스팅 가격.
        """
        # 첫 25개 기준 가격 (API에서 조회)
        return get_hosted_zone_price(zone_index=1)


@dataclass
class ZoneFinding:
    """개별 Hosted Zone에 대한 분석 결과.

    Attributes:
        zone: 분석 대상 Zone 정보.
        status: 분석 결과 상태.
        recommendation: 권장 조치 사항 (한글).
    """

    zone: HostedZoneInfo
    status: ZoneStatus
    recommendation: str


@dataclass
class Route53AnalysisResult:
    """단일 계정의 Route53 Hosted Zone 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        total_zones: 전체 Hosted Zone 수.
        empty_zones: 빈 Zone 수.
        ns_soa_only_zones: NS/SOA만 있는 Zone 수.
        private_zones: Private Zone 수.
        public_zones: Public Zone 수.
        wasted_monthly_cost: 미사용 Zone으로 인한 월간 낭비 비용 (USD).
        findings: 개별 Zone 분석 결과 목록.
    """

    account_id: str
    account_name: str
    total_zones: int = 0
    empty_zones: int = 0
    ns_soa_only_zones: int = 0
    private_zones: int = 0
    public_zones: int = 0
    wasted_monthly_cost: float = 0.0
    findings: list[ZoneFinding] = field(default_factory=list)


def collect_hosted_zones(session, account_id: str, account_name: str) -> list[HostedZoneInfo]:
    """Route53 Hosted Zone을 수집한다 (글로벌 서비스).

    Private Zone의 경우 연결된 VPC 정보도 함께 수집하고,
    NS/SOA 외 실제 레코드 존재 여부를 확인한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.

    Returns:
        수집된 Hosted Zone 정보 목록.
    """
    from botocore.exceptions import ClientError

    route53 = get_client(session, "route53")
    zones = []

    paginator = route53.get_paginator("list_hosted_zones")
    for page in paginator.paginate():
        for zone in page.get("HostedZones", []):
            zone_id = zone.get("Id", "").replace("/hostedzone/", "")

            zone_info = HostedZoneInfo(
                account_id=account_id,
                account_name=account_name,
                zone_id=zone_id,
                name=zone.get("Name", ""),
                is_private=zone.get("Config", {}).get("PrivateZone", False),
                record_count=zone.get("ResourceRecordSetCount", 0),
                comment=zone.get("Config", {}).get("Comment", ""),
            )

            # Private zone인 경우 연결된 VPC 조회
            if zone_info.is_private:
                try:
                    hz_detail = route53.get_hosted_zone(Id=zone_id)
                    vpcs = hz_detail.get("VPCs", [])
                    zone_info.vpcs = [f"{v.get('VPCRegion')}:{v.get('VPCId')}" for v in vpcs]
                except ClientError:
                    pass

            # 실제 레코드 존재 여부 확인 (NS, SOA 제외)
            try:
                records = route53.list_resource_record_sets(HostedZoneId=zone_id, MaxItems="100")
                for record in records.get("ResourceRecordSets", []):
                    record_type = record.get("Type", "")
                    if record_type not in ("NS", "SOA"):
                        zone_info.has_real_records = True
                        break
            except ClientError:
                pass

            zones.append(zone_info)

    return zones


def analyze_hosted_zones(zones: list[HostedZoneInfo], account_id: str, account_name: str) -> Route53AnalysisResult:
    """수집된 Hosted Zone을 분석하여 빈 Zone/미사용 Zone을 식별한다.

    완전히 빈 Zone과 NS/SOA만 있는 Zone을 분류하고, 월간 낭비 비용을 산출한다.

    Args:
        zones: 분석 대상 Hosted Zone 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.

    Returns:
        Hosted Zone 분석 결과 집계 객체.
    """
    result = Route53AnalysisResult(
        account_id=account_id,
        account_name=account_name,
        total_zones=len(zones),
    )

    for zone in zones:
        if zone.is_private:
            result.private_zones += 1
        else:
            result.public_zones += 1

        # 완전히 빈 zone (레코드가 0개인 경우 - 드물지만)
        if zone.record_count == 0:
            result.empty_zones += 1
            result.wasted_monthly_cost += zone.monthly_cost
            result.findings.append(
                ZoneFinding(
                    zone=zone,
                    status=ZoneStatus.EMPTY,
                    recommendation="완전히 빈 Zone - 삭제 검토",
                )
            )
            continue

        # NS/SOA만 있는 경우 (실질적으로 미사용)
        if not zone.has_real_records:
            result.ns_soa_only_zones += 1
            result.wasted_monthly_cost += zone.monthly_cost
            result.findings.append(
                ZoneFinding(
                    zone=zone,
                    status=ZoneStatus.NS_SOA_ONLY,
                    recommendation="NS/SOA만 존재 - 사용 여부 확인",
                )
            )
            continue

        result.findings.append(
            ZoneFinding(
                zone=zone,
                status=ZoneStatus.NORMAL,
                recommendation="정상 (레코드 있음)",
            )
        )

    return result


def generate_report(results: list[Route53AnalysisResult], output_dir: str) -> str:
    """Route53 빈 Hosted Zone 분석 결과를 Excel 보고서로 생성한다.

    Summary 시트, Summary Data 시트(계정별 통계), Detail 시트(비정상 Zone 상세)를 포함한다.

    Args:
        results: 계정별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary Sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("Route53 빈 Hosted Zone 보고서")

    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="전체 Zone", width=12, style="number"),
        ColumnDef(header="빈 Zone", width=10, style="number"),
        ColumnDef(header="NS/SOA만", width=12, style="number"),
        ColumnDef(header="Public", width=10, style="number"),
        ColumnDef(header="Private", width=10, style="number"),
        ColumnDef(header="낭비 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.total_zones,
                r.empty_zones,
                r.ns_soa_only_zones,
                r.public_zones,
                r.private_zones,
                f"${r.wasted_monthly_cost:,.2f}",
            ]
        )
        ws = summary_sheet._ws
        if r.empty_zones > 0:
            ws.cell(row=row_num, column=3).fill = red_fill
        if r.ns_soa_only_zones > 0:
            ws.cell(row=row_num, column=4).fill = yellow_fill

    # Detail Sheet
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Zone ID", width=25),
        ColumnDef(header="Domain", width=40),
        ColumnDef(header="Type", width=10),
        ColumnDef(header="레코드수", width=10, style="number"),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="Comment", width=40),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Hosted Zones", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != ZoneStatus.NORMAL:
                zone = f.zone
                detail_sheet.add_row(
                    [
                        zone.account_name,
                        zone.zone_id,
                        zone.name,
                        "Private" if zone.is_private else "Public",
                        zone.record_count,
                        f.status.value,
                        zone.comment,
                        f.recommendation,
                    ]
                )

    return str(wb.save_as(output_dir, "Route53_EmptyZone"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> Route53AnalysisResult | None:
    """단일 계정의 Hosted Zone을 수집하고 분석한다.

    parallel_collect 콜백으로 사용된다. Route53는 글로벌 서비스이므로 region 파라미터는
    무시되며, parallel_collect의 중복 제거가 계정당 한 번만 실행되도록 보장한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 리전 (Route53는 글로벌 서비스이므로 무시됨).

    Returns:
        분석 결과 객체. Zone이 없으면 None.
    """
    zones = collect_hosted_zones(session, account_id, account_name)
    if not zones:
        return None
    return analyze_hosted_zones(zones, account_id, account_name)


def run(ctx: ExecutionContext) -> None:
    """빈 Route53 Hosted Zone 분석 도구의 메인 실행 함수.

    멀티 계정 병렬 수집 후 결과를 집계하고, Excel 보고서를 생성하여 출력 디렉토리에 저장한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 계정/리전 목록, 옵션 등 포함).
    """
    console.print("[bold]Route53 빈 Hosted Zone 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="route53")
    results: list[Route53AnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.empty_zones + r.ns_soa_only_zones for r in results)
    total_cost = sum(r.wasted_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미사용 Hosted Zone: [yellow]{total_unused}개[/yellow] (${total_cost:,.2f}/월)")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("route53", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
