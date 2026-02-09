"""
functions/analyzers/efs/unused.py - EFS 미사용 파일시스템 분석

유휴/미사용 EFS 파일시스템 탐지 (마운트 타겟 및 CloudWatch 지표 기반)

탐지 기준 (CloudFix 기반):
- 미사용: ClientConnections = 0 AND MeteredIOBytes = 0 (30일)
- 마운트없음: mount_target_count = 0
- 빈파일: size < 1MB
- https://docs.aws.amazon.com/efs/latest/ug/efs-metrics.html
- https://cloudfix.com/blog/delete-idle-volumes-save-on-efs/

CloudWatch 메트릭:
- Namespace: AWS/EFS
- 메트릭: ClientConnections, MeteredIOBytes, DataReadIOBytes, DataWriteIOBytes
- Dimension: FileSystemId

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from core.shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 분석 기간 (일) - CloudFix 권장 30일
ANALYSIS_DAYS = 30

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticfilesystem:DescribeFileSystems",
        "elasticfilesystem:DescribeMountTargets",
        "cloudwatch:GetMetricStatistics",
    ],
}


class FileSystemStatus(Enum):
    """EFS 파일시스템 사용 상태 분류

    마운트 타겟, I/O 활동, 파일 크기를 기준으로
    정상, 마운트 없음, I/O 없음, 빈 파일시스템으로 구분합니다.
    """

    NORMAL = "normal"
    NO_MOUNT_TARGET = "no_mount_target"
    NO_IO = "no_io"
    EMPTY = "empty"


@dataclass
class EFSInfo:
    """EFS 파일시스템 정보

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        file_system_id: EFS 파일시스템 ID
        name: Name 태그 값
        lifecycle_state: 라이프사이클 상태 (available, creating 등)
        performance_mode: 성능 모드 (generalPurpose, maxIO)
        throughput_mode: 처리량 모드 (bursting, provisioned, elastic)
        size_bytes: 파일시스템 크기 (바이트)
        mount_target_count: 마운트 타겟 수
        created_at: 생성 시각
        avg_client_connections: 분석 기간 평균 클라이언트 연결 수
        metered_io_bytes: 분석 기간 과금 대상 I/O 바이트 합계
        data_read_bytes: 분석 기간 읽기 I/O 바이트 합계
        data_write_bytes: 분석 기간 쓰기 I/O 바이트 합계
    """

    account_id: str
    account_name: str
    region: str
    file_system_id: str
    name: str
    lifecycle_state: str
    performance_mode: str
    throughput_mode: str
    size_bytes: int
    mount_target_count: int
    created_at: datetime | None
    # CloudWatch 지표
    avg_client_connections: float = 0.0
    metered_io_bytes: float = 0.0  # 과금 대상 I/O (더 정확)
    data_read_bytes: float = 0.0
    data_write_bytes: float = 0.0

    @property
    def size_gb(self) -> float:
        """파일시스템 크기를 GB 단위로 반환

        Returns:
            파일시스템 크기 (GB)
        """
        return self.size_bytes / (1024**3)

    def get_estimated_monthly_cost(self, session=None) -> float:
        """월간 비용 추정 (Pricing API 사용)

        Args:
            session: boto3 Session 객체. None이면 기본 세션 사용.

        Returns:
            추정 월간 비용 (USD)
        """
        from core.shared.aws.pricing.efs import get_efs_monthly_cost

        return get_efs_monthly_cost(self.region, self.size_gb, session=session)

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (후방 호환용)

        Returns:
            추정 월간 비용 (USD)
        """
        return self.get_estimated_monthly_cost()


@dataclass
class EFSFinding:
    """EFS 개별 분석 결과

    Attributes:
        efs: 분석 대상 EFS 파일시스템 정보
        status: 분석된 사용 상태
        recommendation: 권장 조치 사항
    """

    efs: EFSInfo
    status: FileSystemStatus
    recommendation: str


@dataclass
class EFSAnalysisResult:
    """EFS 분석 결과 집계

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        total_filesystems: 전체 EFS 파일시스템 수
        no_mount_target: 마운트 타겟 없는 파일시스템 수
        no_io: I/O 활동 없는 파일시스템 수
        empty: 빈 파일시스템 수 (1MB 미만)
        normal: 정상 파일시스템 수
        unused_monthly_cost: 미사용 파일시스템 합산 월간 비용 (USD)
        findings: 개별 분석 결과 목록
    """

    account_id: str
    account_name: str
    region: str
    total_filesystems: int = 0
    no_mount_target: int = 0
    no_io: int = 0
    empty: int = 0
    normal: int = 0
    unused_monthly_cost: float = 0.0
    findings: list[EFSFinding] = field(default_factory=list)


def collect_efs_filesystems(session, account_id: str, account_name: str, region: str) -> list[EFSInfo]:
    """EFS 파일시스템 수집 (배치 메트릭 최적화)

    파일시스템 목록과 마운트 타겟 수를 수집한 후,
    CloudWatch 메트릭을 배치로 조회합니다.

    최적화:
    - 기존: 파일시스템당 4 API 호출 -> 최적화: 전체 1-2 API 호출

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        CloudWatch 메트릭이 포함된 EFS 파일시스템 정보 목록
    """
    from botocore.exceptions import ClientError

    efs_client = get_client(session, "efs", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    filesystems: list[EFSInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 파일시스템 목록 수집
        paginator = efs_client.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page.get("FileSystems", []):
                fs_id = fs.get("FileSystemId", "")

                # 이름 태그 찾기
                name = ""
                for tag in fs.get("Tags", []):
                    if tag.get("Key") == "Name":
                        name = tag.get("Value", "")
                        break

                # 마운트 타겟 수 확인
                mount_target_count = 0
                try:
                    mt_resp = efs_client.describe_mount_targets(FileSystemId=fs_id)
                    mount_target_count = len(mt_resp.get("MountTargets", []))
                except ClientError as e:
                    category = categorize_error(e)
                    if category != ErrorCategory.NOT_FOUND:
                        logger.debug(f"EFS 마운트 타겟 조회 실패: {fs_id} ({get_error_code(e)})")

                info = EFSInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    file_system_id=fs_id,
                    name=name,
                    lifecycle_state=fs.get("LifeCycleState", ""),
                    performance_mode=fs.get("PerformanceMode", ""),
                    throughput_mode=fs.get("ThroughputMode", ""),
                    size_bytes=fs.get("SizeInBytes", {}).get("Value", 0),
                    mount_target_count=mount_target_count,
                    created_at=fs.get("CreationTime"),
                )
                filesystems.append(info)

        # 2단계: 배치 메트릭 조회
        if filesystems:
            _collect_efs_metrics_batch(cloudwatch, filesystems, start_time, now)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"EFS 권한 없음: {account_name}/{region}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"EFS 조회 오류: {get_error_code(e)}")

    return filesystems


def _collect_efs_metrics_batch(
    cloudwatch,
    filesystems: list[EFSInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """EFS CloudWatch 메트릭 배치 수집

    파일시스템별 ClientConnections, MeteredIOBytes, DataReadIOBytes,
    DataWriteIOBytes를 batch_get_metrics로 일괄 조회하여
    각 EFSInfo 객체에 in-place 업데이트합니다.

    Args:
        cloudwatch: CloudWatch boto3 클라이언트
        filesystems: 메트릭을 채울 EFS 정보 목록 (in-place 업데이트)
        start_time: 메트릭 조회 시작 시각
        end_time: 메트릭 조회 종료 시각
    """
    from botocore.exceptions import ClientError

    queries: list[MetricQuery] = []
    days = (end_time - start_time).days or 1

    for fs in filesystems:
        safe_id = sanitize_metric_id(fs.file_system_id)
        dimensions = {"FileSystemId": fs.file_system_id}

        queries.extend(
            [
                MetricQuery(
                    id=f"{safe_id}_conn",
                    namespace="AWS/EFS",
                    metric_name="ClientConnections",
                    dimensions=dimensions,
                    stat="Average",
                ),
                MetricQuery(
                    id=f"{safe_id}_metered",
                    namespace="AWS/EFS",
                    metric_name="MeteredIOBytes",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_read",
                    namespace="AWS/EFS",
                    metric_name="DataReadIOBytes",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_write",
                    namespace="AWS/EFS",
                    metric_name="DataWriteIOBytes",
                    dimensions=dimensions,
                    stat="Sum",
                ),
            ]
        )

    if not queries:
        return

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        for fs in filesystems:
            safe_id = sanitize_metric_id(fs.file_system_id)
            # Average는 일별 평균의 합을 일수로 나눔
            fs.avg_client_connections = results.get(f"{safe_id}_conn", 0.0) / days
            fs.metered_io_bytes = results.get(f"{safe_id}_metered", 0.0)
            fs.data_read_bytes = results.get(f"{safe_id}_read", 0.0)
            fs.data_write_bytes = results.get(f"{safe_id}_write", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_filesystems(
    filesystems: list[EFSInfo], account_id: str, account_name: str, region: str
) -> EFSAnalysisResult:
    """EFS 파일시스템 사용 상태 분석

    마운트 타겟 존재 여부, 파일 크기, I/O 활동을 기준으로
    각 파일시스템의 사용 상태를 판별하고 미사용 비용을 합산합니다.

    Args:
        filesystems: 분석 대상 EFS 파일시스템 정보 목록
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        상태별 카운트, 미사용 비용, 개별 분석 결과가 포함된 집계 결과
    """
    result = EFSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_filesystems=len(filesystems),
    )

    for fs in filesystems:
        # 마운트 타겟 없음
        if fs.mount_target_count == 0:
            result.no_mount_target += 1
            result.unused_monthly_cost += fs.estimated_monthly_cost
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.NO_MOUNT_TARGET,
                    recommendation=f"마운트 타겟 없음 - 삭제 검토 (${fs.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 빈 파일시스템 (거의 0 바이트)
        if fs.size_bytes < 1024 * 1024:  # 1MB 미만
            result.empty += 1
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.EMPTY,
                    recommendation="빈 파일시스템 - 삭제 검토",
                )
            )
            continue

        # I/O 없음 (MeteredIOBytes 사용 - 과금 대상 I/O)
        if fs.metered_io_bytes == 0 and fs.avg_client_connections == 0:
            result.no_io += 1
            result.unused_monthly_cost += fs.estimated_monthly_cost
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.NO_IO,
                    recommendation=f"30일간 I/O 없음 - 삭제 검토 (${fs.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        result.normal += 1
        result.findings.append(
            EFSFinding(
                efs=fs,
                status=FileSystemStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[EFSAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성

    Summary(계정/리전별 상태 카운트 및 비용)와 FileSystems(미사용 FS 상세)
    2개 시트로 구성된 보고서를 생성합니다.

    Args:
        results: 계정/리전별 분석 결과 목록
        output_dir: 출력 디렉토리 경로

    Returns:
        저장된 Excel 파일 경로
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="마운트없음", width=12, style="number"),
        ColumnDef(header="I/O없음", width=10, style="number"),
        ColumnDef(header="빈FS", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=12),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_filesystems,
                r.no_mount_target,
                r.no_io,
                r.empty,
                r.normal,
                f"${r.unused_monthly_cost:,.2f}",
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.no_mount_target > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.no_io > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="ID", width=22),
        ColumnDef(header="Name", width=25),
        ColumnDef(header="Size", width=12),
        ColumnDef(header="Mount Targets", width=12, style="number"),
        ColumnDef(header="Mode", width=12),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="Avg Conn", width=10),
        ColumnDef(header="Metered I/O", width=12),
        ColumnDef(header="Read/Write", width=15),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=35),
    ]
    detail_sheet = wb.new_sheet("FileSystems", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != FileSystemStatus.NORMAL:
                fs = f.efs
                style = Styles.danger() if f.status == FileSystemStatus.NO_MOUNT_TARGET else Styles.warning()

                detail_sheet.add_row(
                    [
                        fs.account_name,
                        fs.region,
                        fs.file_system_id,
                        fs.name or "-",
                        f"{fs.size_gb:.2f} GB",
                        fs.mount_target_count,
                        fs.throughput_mode,
                        f.status.value,
                        f"{fs.avg_client_connections:.1f}",
                        f"{fs.metered_io_bytes / (1024**2):.1f} MB",
                        f"R:{fs.data_read_bytes / (1024**2):.1f}/W:{fs.data_write_bytes / (1024**2):.1f} MB",
                        f"${fs.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "EFS_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EFSAnalysisResult | None:
    """단일 계정/리전의 EFS 수집 및 분석 (parallel_collect 콜백)

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        EFS 분석 결과. 파일시스템이 없으면 None.
    """
    filesystems = collect_efs_filesystems(session, account_id, account_name, region)
    if not filesystems:
        return None
    return analyze_filesystems(filesystems, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """EFS 미사용 파일시스템 분석

    멀티 계정/리전에서 EFS 파일시스템을 병렬 수집하고,
    마운트 타겟 없음, I/O 없음, 빈 파일시스템을 식별하여
    Excel 보고서를 생성합니다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 계정/리전 선택, 출력 설정 포함)
    """
    console.print("[bold]EFS 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="efs")
    results: list[EFSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.no_mount_target + r.no_io + r.empty for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월)")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("efs", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
