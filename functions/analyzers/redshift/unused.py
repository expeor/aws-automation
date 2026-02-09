"""
functions/analyzers/redshift/unused.py - Redshift 미사용 클러스터 분석

유휴/저사용 Redshift 클러스터 탐지 (CloudWatch 지표 기반)

분석 기준:
- Unused: DatabaseConnections < 1 (7일 평균) AND IOPS < 20/일
- Low Usage: DatabaseConnections < 5 AND CPUUtilization < 5%
- Stopped: cluster_status = "paused"

최적화:
- CloudWatch GetMetricData API 사용 (배치 조회)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 분석 기간 (일)
ANALYSIS_DAYS = 7
# 미사용 기준: 연결 수 평균 < 1
UNUSED_CONNECTION_THRESHOLD = 1
# 저사용 기준: CPU 평균 5% 미만
LOW_USAGE_CPU_THRESHOLD = 5.0
# 저사용 기준: 연결 수 평균 < 5
LOW_USAGE_CONNECTION_THRESHOLD = 5

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "redshift:DescribeClusters",
        "cloudwatch:GetMetricData",
    ],
}


class ClusterStatus(Enum):
    """Redshift 클러스터 사용 상태 분류.

    CloudWatch 지표(DatabaseConnections, CPUUtilization, IOPS)와
    클러스터 상태(paused)를 기반으로 분류한다.
    """

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"
    PAUSED = "paused"


@dataclass
class RedshiftClusterInfo:
    """Redshift 클러스터 메타데이터 및 CloudWatch 지표 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        cluster_id: Redshift 클러스터 식별자.
        node_type: 노드 타입 (예: dc2.large, ra3.xlplus).
        num_nodes: 노드 수.
        status: 클러스터 상태 (예: available, paused).
        db_name: 기본 데이터베이스 이름.
        created_at: 클러스터 생성 시각.
        avg_connections: 7일 평균 DB 연결 수.
        avg_cpu: 7일 평균 CPU 사용률 (%).
        avg_read_iops: 7일 평균 읽기 IOPS.
        avg_write_iops: 7일 평균 쓰기 IOPS.
    """

    account_id: str
    account_name: str
    region: str
    cluster_id: str
    node_type: str
    num_nodes: int
    status: str
    db_name: str
    created_at: datetime | None
    # CloudWatch 지표
    avg_connections: float = 0.0
    avg_cpu: float = 0.0
    avg_read_iops: float = 0.0
    avg_write_iops: float = 0.0

    def get_estimated_monthly_cost(self, session=None) -> float:
        """Pricing API를 사용하여 월간 예상 비용을 계산한다.

        Args:
            session: boto3 Session. None이면 기본 세션 사용.

        Returns:
            월간 예상 비용 (USD).
        """
        from core.shared.aws.pricing.redshift import get_redshift_monthly_cost

        return get_redshift_monthly_cost(
            region=self.region,
            node_type=self.node_type,
            num_nodes=self.num_nodes,
            session=session,
        )

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 예상 비용 (후방 호환용 property).

        Returns:
            월간 예상 비용 (USD).
        """
        return self.get_estimated_monthly_cost()


@dataclass
class RedshiftClusterFinding:
    """개별 Redshift 클러스터의 분석 결과.

    Attributes:
        cluster: 분석 대상 클러스터 정보.
        status: 분석된 사용 상태.
        recommendation: 권장 조치 사항 문자열.
    """

    cluster: RedshiftClusterInfo
    status: ClusterStatus
    recommendation: str


@dataclass
class RedshiftAnalysisResult:
    """단일 계정/리전의 Redshift 미사용 클러스터 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        total_clusters: 전체 클러스터 수.
        unused_clusters: 미사용 클러스터 수.
        low_usage_clusters: 저사용 클러스터 수.
        paused_clusters: 일시정지된 클러스터 수.
        normal_clusters: 정상 클러스터 수.
        unused_monthly_cost: 미사용 클러스터 월간 비용 합계 (USD).
        low_usage_monthly_cost: 저사용 클러스터 월간 비용 합계 (USD).
        paused_monthly_cost: 일시정지된 클러스터 월간 비용 합계 (USD).
        findings: 개별 클러스터별 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_clusters: int = 0
    unused_clusters: int = 0
    low_usage_clusters: int = 0
    paused_clusters: int = 0
    normal_clusters: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    paused_monthly_cost: float = 0.0
    findings: list[RedshiftClusterFinding] = field(default_factory=list)


def collect_redshift_clusters(session, account_id: str, account_name: str, region: str) -> list[RedshiftClusterInfo]:
    """Redshift 클러스터 목록 수집 및 CloudWatch 메트릭 배치 조회.

    DescribeClusters로 클러스터 메타데이터를 수집한 후,
    GetMetricData API로 CloudWatch 지표를 배치 조회한다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        RedshiftClusterInfo 목록.
    """
    from botocore.exceptions import ClientError

    redshift = get_client(session, "redshift", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)

    clusters: list[RedshiftClusterInfo] = []
    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        paginator = redshift.get_paginator("describe_clusters")
        for page in paginator.paginate():
            for c in page.get("Clusters", []):
                cluster_id = c.get("ClusterIdentifier", "")
                cluster = RedshiftClusterInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    cluster_id=cluster_id,
                    node_type=c.get("NodeType", "unknown"),
                    num_nodes=c.get("NumberOfNodes", 1),
                    status=c.get("ClusterStatus", ""),
                    db_name=c.get("DBName", ""),
                    created_at=c.get("ClusterCreateTime"),
                )
                clusters.append(cluster)
    except ClientError:
        pass

    # 배치 메트릭 조회
    if clusters:
        _collect_redshift_metrics_batch(cloudwatch, clusters, start_time, now)

    return clusters


def _collect_redshift_metrics_batch(
    cloudwatch,
    clusters: list[RedshiftClusterInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """Redshift 클러스터의 CloudWatch 메트릭을 배치로 수집한다.

    DatabaseConnections, CPUUtilization, ReadIOPS, WriteIOPS 메트릭을
    GetMetricData API로 한 번에 조회하여 각 클러스터에 매핑한다.

    Args:
        cloudwatch: CloudWatch boto3 client.
        clusters: 메트릭을 수집할 클러스터 목록.
        start_time: 조회 시작 시각 (UTC).
        end_time: 조회 종료 시각 (UTC).
    """
    from botocore.exceptions import ClientError

    metrics_to_fetch = [
        ("DatabaseConnections", "avg_connections"),
        ("CPUUtilization", "avg_cpu"),
        ("ReadIOPS", "avg_read_iops"),
        ("WriteIOPS", "avg_write_iops"),
    ]

    queries = []
    for cluster in clusters:
        safe_id = sanitize_metric_id(cluster.cluster_id)
        for metric_name, _ in metrics_to_fetch:
            metric_key = metric_name.lower()
            queries.append(
                MetricQuery(
                    id=f"{safe_id}_{metric_key}",
                    namespace="AWS/Redshift",
                    metric_name=metric_name,
                    dimensions={"ClusterIdentifier": cluster.cluster_id},
                    stat="Average",
                )
            )

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for cluster in clusters:
            safe_id = sanitize_metric_id(cluster.cluster_id)
            for metric_name, attr_name in metrics_to_fetch:
                metric_key = metric_name.lower()
                value = results.get(f"{safe_id}_{metric_key}", 0.0) / days
                setattr(cluster, attr_name, value)

    except ClientError:
        pass


def analyze_clusters(
    clusters: list[RedshiftClusterInfo], account_id: str, account_name: str, region: str
) -> RedshiftAnalysisResult:
    """Redshift 클러스터를 CloudWatch 지표 기준으로 분석하여 유휴/저사용을 판별한다.

    미사용: 연결 수 < 1 AND IOPS < 20.
    저사용: 연결 수 < 5 AND CPU < 5%.
    일시정지: 클러스터 상태가 paused.

    Args:
        clusters: 분석할 Redshift 클러스터 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        분석 결과 집계 (RedshiftAnalysisResult).
    """
    result = RedshiftAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_clusters=len(clusters),
    )

    for cluster in clusters:
        # Paused 상태 확인
        if cluster.status == "paused":
            result.paused_clusters += 1
            result.paused_monthly_cost += cluster.estimated_monthly_cost
            result.findings.append(
                RedshiftClusterFinding(
                    cluster=cluster,
                    status=ClusterStatus.PAUSED,
                    recommendation=f"일시정지됨 - 장기간 미사용 시 삭제 검토 (${cluster.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 미사용: 연결 수 < 1 AND 낮은 IOPS
        total_iops = cluster.avg_read_iops + cluster.avg_write_iops
        if cluster.avg_connections < UNUSED_CONNECTION_THRESHOLD and total_iops < 20:
            result.unused_clusters += 1
            result.unused_monthly_cost += cluster.estimated_monthly_cost
            result.findings.append(
                RedshiftClusterFinding(
                    cluster=cluster,
                    status=ClusterStatus.UNUSED,
                    recommendation=f"미사용 (연결 {cluster.avg_connections:.1f}, IOPS {total_iops:.1f}) - 삭제 검토 (${cluster.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: 연결 수 < 5 AND CPU < 5%
        if cluster.avg_connections < LOW_USAGE_CONNECTION_THRESHOLD and cluster.avg_cpu < LOW_USAGE_CPU_THRESHOLD:
            result.low_usage_clusters += 1
            result.low_usage_monthly_cost += cluster.estimated_monthly_cost
            result.findings.append(
                RedshiftClusterFinding(
                    cluster=cluster,
                    status=ClusterStatus.LOW_USAGE,
                    recommendation=f"저사용 (연결 {cluster.avg_connections:.1f}, CPU {cluster.avg_cpu:.1f}%) - 다운사이징 검토",
                )
            )
            continue

        result.normal_clusters += 1
        result.findings.append(
            RedshiftClusterFinding(
                cluster=cluster,
                status=ClusterStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[RedshiftAnalysisResult], output_dir: str) -> str:
    """Redshift 미사용 클러스터 분석 Excel 보고서를 생성한다.

    Summary 시트(계정/리전별 집계)와 Clusters 시트(문제 클러스터 상세)를 포함.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        저장된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="일시정지", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_clusters,
                r.unused_clusters,
                r.low_usage_clusters,
                r.paused_clusters,
                r.normal_clusters,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
            ]
        )
        ws = summary_sheet._ws
        if r.unused_clusters > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_clusters > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Cluster ID", width=30),
        ColumnDef(header="Node Type", width=15),
        ColumnDef(header="Nodes", width=8, style="number"),
        ColumnDef(header="Status", width=12),
        ColumnDef(header="Avg Conn", width=10),
        ColumnDef(header="Avg CPU", width=10),
        ColumnDef(header="Avg IOPS", width=10),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="분석상태", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Clusters", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != ClusterStatus.NORMAL:
                c = f.cluster
                style = Styles.danger() if f.status == ClusterStatus.UNUSED else Styles.warning()

                detail_sheet.add_row(
                    [
                        c.account_name,
                        c.region,
                        c.cluster_id,
                        c.node_type,
                        c.num_nodes,
                        c.status,
                        f"{c.avg_connections:.1f}",
                        f"{c.avg_cpu:.1f}%",
                        f"{c.avg_read_iops + c.avg_write_iops:.1f}",
                        f"${c.estimated_monthly_cost:.2f}",
                        f.status.value,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "Redshift_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> RedshiftAnalysisResult | None:
    """단일 계정/리전의 Redshift 클러스터를 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 병렬로 실행된다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        분석 결과. 클러스터가 없으면 None.
    """
    clusters = collect_redshift_clusters(session, account_id, account_name, region)
    if not clusters:
        return None
    return analyze_clusters(clusters, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Redshift 미사용 클러스터 분석 도구의 진입점.

    멀티 계정/리전 병렬 수집 후 콘솔 요약 출력, Excel 보고서를 생성한다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 리전, 출력 설정 포함).
    """
    console.print("[bold]Redshift 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="redshift")
    results: list[RedshiftAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_clusters for r in results)
    total_low = sum(r.low_usage_clusters for r in results)
    total_paused = sum(r.paused_clusters for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월) / "
        f"일시정지: {total_paused}개"
    )

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("redshift", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
