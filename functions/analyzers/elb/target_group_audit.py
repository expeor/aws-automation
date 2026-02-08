"""
functions/analyzers/elb/target_group_audit.py - Target Group 미사용 분석

ELB에 연결되지 않은 Target Group 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticloadbalancing:DescribeTargetGroups",
        "elasticloadbalancing:DescribeTargetHealth",
    ],
}


class TargetGroupStatus(Enum):
    """Target Group 사용 상태 분류

    정상, LB 미연결, 타겟 없음, 전체 비정상 4가지로 구분하여
    미사용 또는 문제 있는 Target Group을 식별합니다.
    """

    NORMAL = "normal"
    UNATTACHED = "unattached"  # LB에 연결 안 됨
    NO_TARGETS = "no_targets"  # 타겟 없음
    ALL_UNHEALTHY = "all_unhealthy"  # 모든 타겟 비정상


@dataclass
class TargetGroupInfo:
    """Target Group 정보

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        arn: Target Group ARN
        name: Target Group 이름
        protocol: 프로토콜 (HTTP, HTTPS, TCP 등). Lambda 타입은 None.
        port: 포트 번호. Lambda 타입은 None.
        target_type: 타겟 유형 (instance, ip, lambda, alb)
        vpc_id: VPC ID. Lambda 타입은 None.
        load_balancer_arns: 연결된 로드밸런서 ARN 목록
        total_targets: 등록된 전체 타겟 수
        healthy_targets: 정상(healthy) 타겟 수
        unhealthy_targets: 비정상 타겟 수
    """

    account_id: str
    account_name: str
    region: str
    arn: str
    name: str
    protocol: str | None
    port: int | None
    target_type: str  # instance, ip, lambda, alb
    vpc_id: str | None
    load_balancer_arns: list[str] = field(default_factory=list)
    total_targets: int = 0
    healthy_targets: int = 0
    unhealthy_targets: int = 0

    @property
    def is_attached(self) -> bool:
        """로드밸런서 연결 여부 확인

        Returns:
            하나 이상의 LB에 연결되어 있으면 True
        """
        return len(self.load_balancer_arns) > 0


@dataclass
class TargetGroupFinding:
    """Target Group 개별 분석 결과

    Attributes:
        tg: 분석 대상 Target Group 정보
        status: 분석된 사용 상태
        recommendation: 권장 조치 사항
    """

    tg: TargetGroupInfo
    status: TargetGroupStatus
    recommendation: str


@dataclass
class TargetGroupAnalysisResult:
    """Target Group 분석 결과 집계

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        total_count: 전체 Target Group 수
        unattached_count: LB 미연결 Target Group 수
        no_targets_count: 타겟 미등록 Target Group 수
        unhealthy_count: 전체 타겟 비정상 Target Group 수
        normal_count: 정상 Target Group 수
        findings: 개별 분석 결과 목록
    """

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    unattached_count: int = 0
    no_targets_count: int = 0
    unhealthy_count: int = 0
    normal_count: int = 0
    findings: list[TargetGroupFinding] = field(default_factory=list)


# =============================================================================
# 수집
# =============================================================================


def collect_target_groups(session, account_id: str, account_name: str, region: str) -> list[TargetGroupInfo]:
    """Target Group 목록 및 타겟 헬스 정보 수집

    ELBv2 API로 Target Group 목록을 페이지네이션으로 수집하고,
    각 Target Group의 타겟 헬스 상태를 조회합니다.

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        Target Group 정보 목록
    """
    from botocore.exceptions import ClientError

    elbv2 = get_client(session, "elbv2", region_name=region)
    target_groups = []

    paginator = elbv2.get_paginator("describe_target_groups")
    for page in paginator.paginate():
        for tg in page.get("TargetGroups", []):
            tg_info = TargetGroupInfo(
                account_id=account_id,
                account_name=account_name,
                region=region,
                arn=tg["TargetGroupArn"],
                name=tg["TargetGroupName"],
                protocol=tg.get("Protocol"),
                port=tg.get("Port"),
                target_type=tg.get("TargetType", "instance"),
                vpc_id=tg.get("VpcId"),
                load_balancer_arns=tg.get("LoadBalancerArns", []),
            )

            # 타겟 상태 조회
            try:
                health_resp = elbv2.describe_target_health(TargetGroupArn=tg["TargetGroupArn"])
                targets = health_resp.get("TargetHealthDescriptions", [])
                tg_info.total_targets = len(targets)
                tg_info.healthy_targets = sum(1 for t in targets if t.get("TargetHealth", {}).get("State") == "healthy")
                tg_info.unhealthy_targets = tg_info.total_targets - tg_info.healthy_targets
            except ClientError:
                pass

            target_groups.append(tg_info)

    return target_groups


# =============================================================================
# 분석
# =============================================================================


def analyze_target_groups(
    target_groups: list[TargetGroupInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> TargetGroupAnalysisResult:
    """Target Group 사용 상태 분석

    LB 연결 여부, 타겟 등록 여부, 타겟 헬스 상태를 기준으로
    각 Target Group의 사용 상태를 판별하고 카운트를 집계합니다.

    Args:
        target_groups: 분석 대상 Target Group 정보 목록
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        상태별 카운트 및 개별 분석 결과가 포함된 집계 결과
    """
    result = TargetGroupAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(target_groups),
    )

    for tg in target_groups:
        # LB에 연결 안 됨
        if not tg.is_attached:
            result.unattached_count += 1
            result.findings.append(
                TargetGroupFinding(
                    tg=tg,
                    status=TargetGroupStatus.UNATTACHED,
                    recommendation="로드밸런서에 연결되지 않음 - 삭제 검토",
                )
            )
            continue

        # 타겟 없음
        if tg.total_targets == 0:
            result.no_targets_count += 1
            result.findings.append(
                TargetGroupFinding(
                    tg=tg,
                    status=TargetGroupStatus.NO_TARGETS,
                    recommendation="등록된 타겟 없음 - 타겟 등록 또는 삭제 검토",
                )
            )
            continue

        # 모든 타겟 비정상
        if tg.healthy_targets == 0 and tg.total_targets > 0:
            result.unhealthy_count += 1
            result.findings.append(
                TargetGroupFinding(
                    tg=tg,
                    status=TargetGroupStatus.ALL_UNHEALTHY,
                    recommendation="모든 타겟 비정상 - 헬스체크 확인 필요",
                )
            )
            continue

        result.normal_count += 1
        result.findings.append(
            TargetGroupFinding(
                tg=tg,
                status=TargetGroupStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[TargetGroupAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성

    Summary(계정/리전별 상태 카운트)와 Target Groups(비정상 TG 상세) 2개 시트로
    구성된 보고서를 생성합니다.

    Args:
        results: 계정/리전별 분석 결과 목록
        output_dir: 출력 디렉토리 경로

    Returns:
        저장된 Excel 파일 경로
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 셀 수준 조건부 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미연결", width=10, style="number"),
        ColumnDef(header="타겟 없음", width=12, style="number"),
        ColumnDef(header="비정상", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_count,
                r.unattached_count,
                r.no_targets_count,
                r.unhealthy_count,
                r.normal_count,
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.unattached_count > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.no_targets_count > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill
        if r.unhealthy_count > 0:
            ws.cell(row=row_num, column=6).fill = red_fill

    # Target Groups 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Name", width=25),
        ColumnDef(header="상태", width=15, style="center"),
        ColumnDef(header="Type", width=12),
        ColumnDef(header="Protocol", width=10, style="center"),
        ColumnDef(header="Port", width=10, style="number"),
        ColumnDef(header="LB 연결", width=10, style="number"),
        ColumnDef(header="Total Targets", width=14, style="number"),
        ColumnDef(header="Healthy", width=10, style="number"),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Target Groups", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != TargetGroupStatus.NORMAL:
                tg = f.tg
                style = None
                if f.status == TargetGroupStatus.UNATTACHED:
                    style = Styles.danger()
                elif f.status == TargetGroupStatus.NO_TARGETS:
                    style = Styles.warning()
                elif f.status == TargetGroupStatus.ALL_UNHEALTHY:
                    style = Styles.danger()

                detail_sheet.add_row(
                    [
                        tg.account_name,
                        tg.region,
                        tg.name,
                        f.status.value,
                        tg.target_type,
                        tg.protocol or "-",
                        tg.port or "-",
                        len(tg.load_balancer_arns),
                        tg.total_targets,
                        tg.healthy_targets,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "TargetGroup_Audit"))


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> TargetGroupAnalysisResult | None:
    """단일 계정/리전의 Target Group 수집 및 분석 (parallel_collect 콜백)

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        Target Group 분석 결과. Target Group이 없으면 None.
    """
    target_groups = collect_target_groups(session, account_id, account_name, region)
    if not target_groups:
        return None
    return analyze_target_groups(target_groups, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Target Group 미사용 분석

    멀티 계정/리전에서 Target Group을 병렬 수집하고,
    LB 미연결, 타겟 미등록, 전체 비정상 상태를 식별하여
    Excel 보고서를 생성합니다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 계정/리전 선택, 출력 설정 포함)
    """
    console.print("[bold]Target Group 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="elbv2")
    results: list[TargetGroupAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_unattached = sum(r.unattached_count for r in results)
    total_no_targets = sum(r.no_targets_count for r in results)
    total_unhealthy = sum(r.unhealthy_count for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미연결: [red]{total_unattached}개[/red]")
    console.print(f"타겟 없음: [yellow]{total_no_targets}개[/yellow]")
    console.print(f"전체 비정상: [red]{total_unhealthy}개[/red]")

    # 보고서 생성
    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("elb", "inventory").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
