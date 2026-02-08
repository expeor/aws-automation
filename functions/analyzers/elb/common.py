"""
functions/analyzers/elb/common.py - ELB 공통 유틸리티

ALB, NLB, CLB, GWLB에서 공유하는 데이터 구조와 헬퍼 함수.

boto3 클라이언트:
    - ALB/NLB/GWLB: elbv2 클라이언트
    - CLB: elb 클라이언트
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from core.shared.aws.pricing import get_elb_monthly_cost


class LBType(Enum):
    """Load Balancer 타입 열거형

    AWS ELB의 4가지 유형을 나타냅니다.
    """

    ALB = "application"
    NLB = "network"
    GWLB = "gateway"
    CLB = "classic"

    @classmethod
    def from_string(cls, value: str) -> LBType:
        """문자열에서 LBType 변환

        Args:
            value: LB 타입 문자열 (application, network, gateway, classic)

        Returns:
            대응하는 LBType. 매칭 실패 시 ALB 반환.
        """
        mapping = {
            "application": cls.ALB,
            "network": cls.NLB,
            "gateway": cls.GWLB,
            "classic": cls.CLB,
        }
        return mapping.get(value, cls.ALB)

    @property
    def display_name(self) -> str:
        """표시용 이름

        Returns:
            ALB, NLB, GWLB, CLB 중 하나
        """
        return {
            LBType.ALB: "ALB",
            LBType.NLB: "NLB",
            LBType.GWLB: "GWLB",
            LBType.CLB: "CLB",
        }.get(self, "Unknown")


class UsageStatus(Enum):
    """Load Balancer의 사용 상태 분류

    타겟 등록 여부와 헬스체크 상태로 판별됩니다.
    """

    UNUSED = "unused"  # 미사용 (타겟 없음)
    UNHEALTHY = "unhealthy"  # 모든 타겟 비정상
    NORMAL = "normal"  # 정상 사용


class Severity(Enum):
    """분석 결과의 심각도 수준"""

    CRITICAL = "critical"
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class TargetGroupInfo:
    """타겟 그룹 정보

    Attributes:
        arn: Target Group ARN
        name: Target Group 이름
        target_type: 타겟 유형 (instance, ip, lambda, alb)
        total_targets: 전체 타겟 수
        healthy_targets: 정상 타겟 수
        unhealthy_targets: 비정상 타겟 수
    """

    arn: str
    name: str
    target_type: str
    total_targets: int
    healthy_targets: int
    unhealthy_targets: int


@dataclass
class LoadBalancerInfo:
    """Load Balancer 정보

    Attributes:
        arn: Load Balancer ARN
        name: Load Balancer 이름
        dns_name: DNS 이름
        lb_type: LB 유형 (application, network, gateway, classic)
        scheme: 스킴 (internet-facing, internal)
        state: LB 상태
        vpc_id: VPC ID
        availability_zones: 가용 영역 리스트
        created_time: 생성 시간
        tags: 사용자 태그 딕셔너리
        target_groups: 연결된 타겟 그룹 리스트 (ALB/NLB/GWLB)
        registered_instances: 등록된 인스턴스 수 (CLB 전용)
        healthy_instances: 정상 인스턴스 수 (CLB 전용)
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        monthly_cost: 추정 월간 고정 비용 (USD)
    """

    arn: str
    name: str
    dns_name: str
    lb_type: str  # application, network, gateway, classic
    scheme: str  # internet-facing, internal
    state: str
    vpc_id: str
    availability_zones: list[str]
    created_time: datetime | None
    tags: dict[str, str]

    # 타겟 그룹 (ALB/NLB/GWLB)
    target_groups: list[TargetGroupInfo] = field(default_factory=list)

    # CLB 전용
    registered_instances: int = 0
    healthy_instances: int = 0

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def lb_type_enum(self) -> LBType:
        """LBType enum으로 변환

        Returns:
            lb_type 문자열에 대응하는 LBType enum 값
        """
        return LBType.from_string(self.lb_type)

    @property
    def total_targets(self) -> int:
        """전체 타겟 수

        Returns:
            CLB는 등록 인스턴스 수, ALB/NLB/GWLB는 타겟 그룹별 합계
        """
        if self.lb_type == "classic":
            return self.registered_instances
        return sum(tg.total_targets for tg in self.target_groups)

    @property
    def healthy_targets(self) -> int:
        """정상 타겟 수

        Returns:
            CLB는 InService 인스턴스 수, ALB/NLB/GWLB는 healthy 타겟 합계
        """
        if self.lb_type == "classic":
            return self.healthy_instances
        return sum(tg.healthy_targets for tg in self.target_groups)


@dataclass
class LBFinding:
    """개별 Load Balancer 분석 결과

    Attributes:
        lb: 분석 대상 Load Balancer 정보
        usage_status: 분석으로 판별된 사용 상태
        severity: 심각도 수준
        description: 상태 설명
        recommendation: 권장 조치 사항
    """

    lb: LoadBalancerInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class LBAnalysisResult:
    """Load Balancer 분석 결과 집계 (계정/리전별)

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        lb_type_filter: 필터링된 LB 타입 (None이면 전체)
        findings: 개별 LB 분석 결과 리스트
        total_count: 전체 LB 수
        unused_count: 미사용 LB 수
        unhealthy_count: 전체 타겟 비정상 LB 수
        normal_count: 정상 LB 수
        unused_monthly_cost: 미사용/unhealthy LB 추정 월간 비용 (USD)
    """

    account_id: str
    account_name: str
    region: str
    lb_type_filter: str | None = None  # 필터링된 LB 타입 (None이면 전체)
    findings: list[LBFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    unhealthy_count: int = 0
    normal_count: int = 0

    # 비용
    unused_monthly_cost: float = 0.0


# =============================================================================
# 수집 함수 - ALB/NLB/GWLB (elbv2)
# =============================================================================


def collect_v2_load_balancers(
    session,
    account_id: str,
    account_name: str,
    region: str,
    lb_type_filter: str | None = None,
) -> list[LoadBalancerInfo]:
    """ALB/NLB/GWLB 목록 수집

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        lb_type_filter: 필터링할 LB 타입 (application, network, gateway)

    Returns:
        Load Balancer 정보 리스트 (타겟 그룹 헬스 포함)
    """
    from botocore.exceptions import ClientError

    from core.parallel import get_client

    load_balancers = []

    try:
        elbv2 = get_client(session, "elbv2", region_name=region)

        # Load Balancers 조회
        paginator = elbv2.get_paginator("describe_load_balancers")
        for page in paginator.paginate():
            for data in page.get("LoadBalancers", []):
                lb_type = data.get("Type", "application")

                # 타입 필터
                if lb_type_filter and lb_type != lb_type_filter:
                    continue

                lb_arn = data.get("LoadBalancerArn", "")

                # 태그 조회
                tags = {}
                try:
                    tag_response = elbv2.describe_tags(ResourceArns=[lb_arn])
                    for tag_desc in tag_response.get("TagDescriptions", []):
                        for t in tag_desc.get("Tags", []):
                            key = t.get("Key", "")
                            if not key.startswith("aws:"):
                                tags[key] = t.get("Value", "")
                except ClientError:
                    pass

                lb = LoadBalancerInfo(
                    arn=lb_arn,
                    name=data.get("LoadBalancerName", ""),
                    dns_name=data.get("DNSName", ""),
                    lb_type=lb_type,
                    scheme=data.get("Scheme", ""),
                    state=data.get("State", {}).get("Code", ""),
                    vpc_id=data.get("VpcId", ""),
                    availability_zones=[az.get("ZoneName", "") for az in data.get("AvailabilityZones", [])],
                    created_time=data.get("CreatedTime"),
                    tags=tags,
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    monthly_cost=get_elb_monthly_cost(region, lb_type),
                )

                # 타겟 그룹 조회
                lb.target_groups = _get_target_groups(elbv2, lb_arn)
                load_balancers.append(lb)

    except ClientError:
        pass

    return load_balancers


def _get_target_groups(elbv2, lb_arn: str) -> list[TargetGroupInfo]:
    """LB에 연결된 타겟 그룹 및 헬스 상태 조회

    Args:
        elbv2: ELBv2 클라이언트
        lb_arn: Load Balancer ARN

    Returns:
        타겟 그룹 정보 리스트
    """
    from botocore.exceptions import ClientError

    target_groups = []

    try:
        response = elbv2.describe_target_groups(LoadBalancerArn=lb_arn)

        for tg in response.get("TargetGroups", []):
            tg_arn = tg.get("TargetGroupArn", "")

            # 타겟 헬스 조회
            healthy = 0
            unhealthy = 0
            total = 0

            try:
                health_response = elbv2.describe_target_health(TargetGroupArn=tg_arn)
                for target in health_response.get("TargetHealthDescriptions", []):
                    total += 1
                    state = target.get("TargetHealth", {}).get("State", "")
                    if state == "healthy":
                        healthy += 1
                    else:
                        unhealthy += 1
            except ClientError:
                pass

            target_groups.append(
                TargetGroupInfo(
                    arn=tg_arn,
                    name=tg.get("TargetGroupName", ""),
                    target_type=tg.get("TargetType", ""),
                    total_targets=total,
                    healthy_targets=healthy,
                    unhealthy_targets=unhealthy,
                )
            )

    except ClientError:
        pass

    return target_groups


# =============================================================================
# 수집 함수 - CLB (elb)
# =============================================================================


def collect_classic_load_balancers(
    session,
    account_id: str,
    account_name: str,
    region: str,
) -> list[LoadBalancerInfo]:
    """Classic Load Balancer 목록 수집

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        CLB 정보 리스트 (인스턴스 헬스 포함)
    """
    from botocore.exceptions import ClientError

    from core.parallel import get_client

    load_balancers = []

    try:
        elb = get_client(session, "elb", region_name=region)

        response = elb.describe_load_balancers()

        for data in response.get("LoadBalancerDescriptions", []):
            lb_name = data.get("LoadBalancerName", "")

            # 태그 조회
            tags = {}
            try:
                tag_response = elb.describe_tags(LoadBalancerNames=[lb_name])
                for tag_desc in tag_response.get("TagDescriptions", []):
                    for t in tag_desc.get("Tags", []):
                        key = t.get("Key", "")
                        if not key.startswith("aws:"):
                            tags[key] = t.get("Value", "")
            except ClientError:
                pass

            # 인스턴스 헬스 조회
            instances = data.get("Instances", [])
            healthy = 0
            try:
                if instances:
                    health_response = elb.describe_instance_health(LoadBalancerName=lb_name)
                    for state in health_response.get("InstanceStates", []):
                        if state.get("State") == "InService":
                            healthy += 1
            except ClientError:
                pass

            lb = LoadBalancerInfo(
                arn=f"arn:aws:elasticloadbalancing:{region}:{account_id}:loadbalancer/{lb_name}",
                name=lb_name,
                dns_name=data.get("DNSName", ""),
                lb_type="classic",
                scheme=data.get("Scheme", ""),
                state="active",  # CLB는 state 없음
                vpc_id=data.get("VPCId", ""),
                availability_zones=data.get("AvailabilityZones", []),
                created_time=data.get("CreatedTime"),
                tags=tags,
                registered_instances=len(instances),
                healthy_instances=healthy,
                account_id=account_id,
                account_name=account_name,
                region=region,
                monthly_cost=get_elb_monthly_cost(region, "classic"),
            )
            load_balancers.append(lb)

    except ClientError:
        pass

    return load_balancers


# =============================================================================
# 분석 함수
# =============================================================================


def analyze_single_lb(lb: LoadBalancerInfo) -> LBFinding:
    """개별 Load Balancer 사용 상태 분석

    Args:
        lb: 분석 대상 Load Balancer 정보

    Returns:
        LB 분석 결과 (상태, 심각도, 권장 조치 포함)
    """

    # 비활성 상태
    if lb.state not in ("active", ""):
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNUSED,
            severity=Severity.LOW,
            description=f"비활성 상태 ({lb.state})",
            recommendation="상태 확인 또는 삭제 검토",
        )

    total = lb.total_targets
    healthy = lb.healthy_targets

    # 타겟 없음
    if total == 0:
        # 타겟 그룹 자체가 없는 경우 (ALB/NLB)
        if lb.lb_type != "classic" and not lb.target_groups:
            return LBFinding(
                lb=lb,
                usage_status=UsageStatus.UNUSED,
                severity=Severity.HIGH,
                description=f"타겟 그룹 없음 (${lb.monthly_cost:.2f}/월)",
                recommendation="타겟 그룹 연결 또는 삭제 검토",
            )

        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNUSED,
            severity=Severity.HIGH,
            description=f"등록된 타겟 없음 (${lb.monthly_cost:.2f}/월)",
            recommendation="타겟 등록 또는 삭제 검토",
        )

    # 모든 타겟 unhealthy
    if healthy == 0:
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNHEALTHY,
            severity=Severity.HIGH,
            description=f"모든 타겟 비정상 ({total}개 unhealthy)",
            recommendation="타겟 헬스체크 확인",
        )

    # 일부 unhealthy
    unhealthy = total - healthy
    if unhealthy > 0:
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.MEDIUM,
            description=f"일부 타겟 비정상 ({healthy}/{total} healthy)",
            recommendation="비정상 타겟 확인",
        )

    # 정상
    return LBFinding(
        lb=lb,
        usage_status=UsageStatus.NORMAL,
        severity=Severity.INFO,
        description=f"정상 ({healthy}개 healthy)",
        recommendation="정상 운영 중",
    )


def analyze_load_balancers(
    load_balancers: list[LoadBalancerInfo],
    account_id: str,
    account_name: str,
    region: str,
    lb_type_filter: str | None = None,
) -> LBAnalysisResult:
    """Load Balancer 미사용 분석

    Args:
        load_balancers: 수집된 LB 정보 리스트
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        lb_type_filter: 필터링된 LB 타입 (None이면 전체)

    Returns:
        계정/리전별 분석 결과 (상태별 LB 수, 비용 포함)
    """
    result = LBAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        lb_type_filter=lb_type_filter,
    )

    for lb in load_balancers:
        finding = analyze_single_lb(lb)
        result.findings.append(finding)

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
            result.unused_monthly_cost += lb.monthly_cost
        elif finding.usage_status == UsageStatus.UNHEALTHY:
            result.unhealthy_count += 1
            result.unused_monthly_cost += lb.monthly_cost  # unhealthy도 낭비
        else:
            result.normal_count += 1

    result.total_count = len(load_balancers)
    return result


# =============================================================================
# Excel 보고서 공통
# =============================================================================


def generate_unused_report(
    results: list[LBAnalysisResult],
    output_dir: str,
    lb_type_name: str = "ELB",
) -> str:
    """미사용 분석 Excel 보고서 생성

    Summary 시트와 미사용/unhealthy LB Findings 시트를 포함합니다.

    Args:
        results: 분석 결과 리스트
        output_dir: 출력 디렉토리 경로
        lb_type_name: 보고서 제목에 사용할 LB 타입 이름 (ELB, ALB, NLB, CLB)

    Returns:
        생성된 Excel 파일 경로
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 상태별 색상 정의 (cell-level 적용용)
    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.UNHEALTHY: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # Summary 시트
    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "unhealthy": sum(r.unhealthy_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "unused_cost": sum(r.unused_monthly_cost for r in results),
    }

    summary = wb.new_summary_sheet("Summary")
    summary.add_title(f"{lb_type_name} 미사용 분석 보고서")
    summary.add_section("분석 결과")
    summary.add_item(f"전체 {lb_type_name}", totals["total"])
    summary.add_item("미사용", totals["unused"], highlight="danger" if totals["unused"] > 0 else None)
    summary.add_item("Unhealthy", totals["unhealthy"], highlight="warning" if totals["unhealthy"] > 0 else None)
    summary.add_item("정상", totals["normal"])
    summary.add_item(
        "미사용 월 비용", f"${totals['unused_cost']:.2f}", highlight="danger" if totals["unused_cost"] > 0 else None
    )

    # Findings 시트
    columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Name", width=30),
        ColumnDef(header="Type", width=10, style="center"),
        ColumnDef(header="Scheme", width=15, style="center"),
        ColumnDef(header="Usage", width=12, style="center"),
        ColumnDef(header="Severity", width=10, style="center"),
        ColumnDef(header="Targets", width=10, style="number"),
        ColumnDef(header="Healthy", width=10, style="number"),
        ColumnDef(header="Monthly Cost ($)", width=15, style="currency"),
        ColumnDef(header="DNS Name", width=50),
        ColumnDef(header="Description", width=30),
        ColumnDef(header="Recommendation", width=25),
    ]
    sheet = wb.new_sheet("Findings", columns)

    # 미사용/unhealthy만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.UNUSED, UsageStatus.UNHEALTHY):
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.lb.monthly_cost, reverse=True)

    for f in all_findings:
        lb = f.lb
        # 행 스타일 결정
        style = Styles.danger() if f.usage_status == UsageStatus.UNUSED else Styles.warning()

        row_num = sheet.add_row(
            [
                lb.account_name,
                lb.region,
                lb.name,
                lb.lb_type.upper(),
                lb.scheme,
                f.usage_status.value,
                f.severity.value,
                lb.total_targets,
                lb.healthy_targets,
                round(lb.monthly_cost, 2),
                lb.dns_name,
                f.description,
                f.recommendation,
            ],
            style=style,
        )

        # Usage 컬럼에 상태별 색상 적용
        fill = status_fills.get(f.usage_status)
        if fill:
            sheet._ws.cell(row=row_num, column=6).fill = fill

    return str(wb.save_as(output_dir, f"{lb_type_name}_Unused"))
