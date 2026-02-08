"""
functions/analyzers/elb/migration_advisor.py - CLB Migration Advisor

Classic Load Balancer(CLB)를 ALB/NLB로 마이그레이션 분석 및 추천

분석 항목:
1. 마이그레이션 적합성 분석
   - CLB 기능 사용 현황
   - ALB/NLB 호환성 체크

2. 추천 타겟 LB 결정
   - HTTP/HTTPS → ALB 추천
   - TCP/UDP → NLB 추천
   - 혼합 → 분리 또는 NLB 추천

3. 비용 비교 분석
   - 현재 CLB 비용
   - 예상 ALB/NLB 비용
   - 절감/증가 예측

4. 마이그레이션 체크리스트
   - 필요한 작업 항목
   - 주의사항
   - 호환성 이슈

AWS 공식 마이그레이션 가이드:
- https://docs.aws.amazon.com/elasticloadbalancing/latest/userguide/migrate-classic-load-balancer.html

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import TYPE_CHECKING, Any

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.shared.aws.pricing import get_elb_monthly_cost
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticloadbalancing:DescribeLoadBalancers",
        "elasticloadbalancing:DescribeLoadBalancerAttributes",
        "elasticloadbalancing:DescribeLoadBalancerPolicies",
    ],
}


class RecommendedTarget(Enum):
    """CLB 마이그레이션 추천 타겟 유형

    리스너 프로토콜 분석을 기반으로 결정됩니다.
    """

    ALB = "ALB"  # Application Load Balancer
    NLB = "NLB"  # Network Load Balancer
    SPLIT = "SPLIT"  # ALB + NLB 분리 권장
    KEEP = "KEEP"  # CLB 유지 권장 (특수 케이스)


class MigrationComplexity(Enum):
    """마이그레이션 복잡도

    호환성 이슈, 리스너 수, 인스턴스 수를 기반으로 결정됩니다.
    """

    SIMPLE = "simple"  # 단순 마이그레이션
    MODERATE = "moderate"  # 중간 복잡도
    COMPLEX = "complex"  # 복잡한 마이그레이션


class CompatibilityStatus(Enum):
    """CLB 기능의 ALB/NLB 호환성 상태"""

    COMPATIBLE = "compatible"  # 완전 호환
    PARTIAL = "partial"  # 부분 호환 (조정 필요)
    INCOMPATIBLE = "incompatible"  # 호환 불가


@dataclass
class ListenerConfig:
    """CLB 리스너 설정

    Attributes:
        protocol: 프론트엔드 프로토콜 (HTTP, HTTPS, TCP, SSL)
        port: 프론트엔드 포트
        instance_protocol: 백엔드 프로토콜
        instance_port: 백엔드 포트
        ssl_certificate_id: SSL 인증서 ARN
    """

    protocol: str  # HTTP, HTTPS, TCP, SSL
    port: int
    instance_protocol: str
    instance_port: int
    ssl_certificate_id: str | None = None


@dataclass
class HealthCheckConfig:
    """CLB 헬스체크 설정

    Attributes:
        target: 헬스체크 대상 (TCP:80, HTTP:80/path)
        interval: 체크 간격 (초)
        timeout: 타임아웃 (초)
        unhealthy_threshold: 비정상 판정 임계값
        healthy_threshold: 정상 판정 임계값
    """

    target: str  # TCP:80, HTTP:80/path
    interval: int
    timeout: int
    unhealthy_threshold: int
    healthy_threshold: int


@dataclass
class CLBInfo:
    """Classic Load Balancer 상세 정보

    Attributes:
        name: CLB 이름
        dns_name: DNS 이름
        scheme: 스킴 (internet-facing, internal)
        vpc_id: VPC ID
        availability_zones: 가용 영역 리스트
        subnets: 서브넷 리스트
        security_groups: Security Group 리스트
        listeners: 리스너 설정 리스트
        health_check: 헬스체크 설정
        instances: 등록된 인스턴스 ID 리스트
        created_time: 생성 시간
        policies: 정책 이름-타입 매핑 딕셔너리
        backend_policies: 백엔드 정책 이름 리스트
        cross_zone_enabled: Cross-Zone Load Balancing 활성화 여부
        connection_draining_enabled: Connection Draining 활성화 여부
        connection_draining_timeout: Connection Draining 타임아웃 (초)
        idle_timeout: Idle 타임아웃 (초)
        access_log_enabled: 액세스 로그 활성화 여부
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
    """

    name: str
    dns_name: str
    scheme: str
    vpc_id: str | None
    availability_zones: list[str]
    subnets: list[str]
    security_groups: list[str]
    listeners: list[ListenerConfig]
    health_check: HealthCheckConfig | None
    instances: list[str]
    created_time: datetime | None

    # 정책
    policies: dict[str, Any] = field(default_factory=dict)
    backend_policies: list[str] = field(default_factory=list)

    # 속성
    cross_zone_enabled: bool = False
    connection_draining_enabled: bool = False
    connection_draining_timeout: int = 300
    idle_timeout: int = 60
    access_log_enabled: bool = False

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    @property
    def has_http_listeners(self) -> bool:
        """HTTP/HTTPS 리스너 존재 여부

        Returns:
            HTTP 또는 HTTPS 프로토콜 리스너가 있으면 True
        """
        return any(listener.protocol in ("HTTP", "HTTPS") for listener in self.listeners)

    @property
    def has_tcp_listeners(self) -> bool:
        """TCP/SSL 리스너 존재 여부

        Returns:
            TCP 또는 SSL 프로토콜 리스너가 있으면 True
        """
        return any(listener.protocol in ("TCP", "SSL") for listener in self.listeners)

    @property
    def is_vpc_based(self) -> bool:
        """VPC 기반 CLB 여부

        Returns:
            VPC ID가 설정되어 있으면 True
        """
        return bool(self.vpc_id)


@dataclass
class CompatibilityIssue:
    """CLB 기능의 ALB/NLB 호환성 이슈

    Attributes:
        feature: 기능 이름
        status: 호환성 상태
        description: 이슈 설명
        workaround: 해결 방법
    """

    feature: str
    status: CompatibilityStatus
    description: str
    workaround: str | None = None


@dataclass
class MigrationRecommendation:
    """개별 CLB 마이그레이션 추천 결과

    Attributes:
        clb: 분석 대상 CLB 정보
        target: 추천 마이그레이션 타겟
        complexity: 마이그레이션 복잡도
        compatibility_issues: 호환성 이슈 리스트
        current_monthly_cost: 현재 CLB 월간 비용 (USD)
        estimated_monthly_cost: 마이그레이션 후 예상 월간 비용 (USD)
        cost_difference: 비용 차이 (양수: 증가, 음수: 절감)
        checklist: 마이그레이션 체크리스트 항목 리스트
        warnings: 경고 사항 리스트
        summary: 마이그레이션 요약 문자열
    """

    clb: CLBInfo
    target: RecommendedTarget
    complexity: MigrationComplexity

    # 호환성
    compatibility_issues: list[CompatibilityIssue] = field(default_factory=list)

    # 비용
    current_monthly_cost: float = 0.0
    estimated_monthly_cost: float = 0.0
    cost_difference: float = 0.0

    # 체크리스트
    checklist: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    # 요약
    summary: str = ""

    @property
    def is_cost_saving(self) -> bool:
        """비용 절감 여부

        Returns:
            마이그레이션 후 비용이 감소하면 True
        """
        return self.cost_difference < 0

    @property
    def overall_status(self) -> CompatibilityStatus:
        """전체 호환성 상태

        Returns:
            INCOMPATIBLE 이슈가 있으면 INCOMPATIBLE,
            PARTIAL 이슈가 있으면 PARTIAL, 그 외 COMPATIBLE
        """
        if any(i.status == CompatibilityStatus.INCOMPATIBLE for i in self.compatibility_issues):
            return CompatibilityStatus.INCOMPATIBLE
        if any(i.status == CompatibilityStatus.PARTIAL for i in self.compatibility_issues):
            return CompatibilityStatus.PARTIAL
        return CompatibilityStatus.COMPATIBLE


@dataclass
class MigrationAnalysisResult:
    """CLB 마이그레이션 분석 결과 집계 (계정/리전별)

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        recommendations: 개별 CLB 마이그레이션 추천 리스트
        total_clbs: 전체 CLB 수
        alb_recommended: ALB 추천 수
        nlb_recommended: NLB 추천 수
        split_recommended: ALB+NLB 분리 추천 수
        keep_recommended: CLB 유지 추천 수
        total_current_cost: 현재 전체 월간 비용 (USD)
        total_estimated_cost: 마이그레이션 후 예상 전체 월간 비용 (USD)
    """

    account_id: str
    account_name: str
    region: str
    recommendations: list[MigrationRecommendation] = field(default_factory=list)

    # 통계
    total_clbs: int = 0
    alb_recommended: int = 0
    nlb_recommended: int = 0
    split_recommended: int = 0
    keep_recommended: int = 0

    # 비용
    total_current_cost: float = 0.0
    total_estimated_cost: float = 0.0


# =============================================================================
# 수집
# =============================================================================


def collect_clb_details(session, account_id: str, account_name: str, region: str) -> list[CLBInfo]:
    """CLB 상세 정보 수집

    리스너, 헬스체크, 속성, 정책 정보를 포함하여 수집합니다.

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        CLB 상세 정보 리스트
    """
    from botocore.exceptions import ClientError

    clbs = []

    try:
        elb = get_client(session, "elb", region_name=region)

        response = elb.describe_load_balancers()

        for data in response.get("LoadBalancerDescriptions", []):
            lb_name = data.get("LoadBalancerName", "")

            # 리스너 파싱
            listeners = []
            for ld in data.get("ListenerDescriptions", []):
                listener_data = ld.get("Listener", {})
                listeners.append(
                    ListenerConfig(
                        protocol=listener_data.get("Protocol", ""),
                        port=listener_data.get("LoadBalancerPort", 0),
                        instance_protocol=listener_data.get("InstanceProtocol", ""),
                        instance_port=listener_data.get("InstancePort", 0),
                        ssl_certificate_id=listener_data.get("SSLCertificateId"),
                    )
                )

            # 헬스체크 파싱
            hc_data = data.get("HealthCheck", {})
            health_check = (
                HealthCheckConfig(
                    target=hc_data.get("Target", ""),
                    interval=hc_data.get("Interval", 30),
                    timeout=hc_data.get("Timeout", 5),
                    unhealthy_threshold=hc_data.get("UnhealthyThreshold", 2),
                    healthy_threshold=hc_data.get("HealthyThreshold", 10),
                )
                if hc_data
                else None
            )

            clb = CLBInfo(
                name=lb_name,
                dns_name=data.get("DNSName", ""),
                scheme=data.get("Scheme", "internet-facing"),
                vpc_id=data.get("VPCId"),
                availability_zones=data.get("AvailabilityZones", []),
                subnets=data.get("Subnets", []),
                security_groups=data.get("SecurityGroups", []),
                listeners=listeners,
                health_check=health_check,
                instances=[i.get("InstanceId", "") for i in data.get("Instances", [])],
                created_time=data.get("CreatedTime"),
                account_id=account_id,
                account_name=account_name,
                region=region,
            )

            # 속성 조회
            try:
                attrs = elb.describe_load_balancer_attributes(LoadBalancerName=lb_name)
                lb_attrs = attrs.get("LoadBalancerAttributes", {})

                clb.cross_zone_enabled = lb_attrs.get("CrossZoneLoadBalancing", {}).get("Enabled", False)

                conn_drain = lb_attrs.get("ConnectionDraining", {})
                clb.connection_draining_enabled = conn_drain.get("Enabled", False)
                clb.connection_draining_timeout = conn_drain.get("Timeout", 300)

                clb.idle_timeout = lb_attrs.get("ConnectionSettings", {}).get("IdleTimeout", 60)
                clb.access_log_enabled = lb_attrs.get("AccessLog", {}).get("Enabled", False)

            except ClientError:
                pass

            # 정책 조회
            try:
                policies = elb.describe_load_balancer_policies(LoadBalancerName=lb_name)
                for policy in policies.get("PolicyDescriptions", []):
                    clb.policies[policy.get("PolicyName", "")] = policy.get("PolicyTypeName", "")
            except ClientError:
                pass

            clbs.append(clb)

    except ClientError as e:
        if "not available" not in str(e).lower():
            error_code = e.response.get("Error", {}).get("Code", "Unknown")
            if not is_quiet():
                console.print(f"    [yellow]{account_name}/{region} CLB 수집 오류: {error_code}[/yellow]")

    return clbs


# =============================================================================
# 분석
# =============================================================================


def analyze_migration(clb: CLBInfo, region: str) -> MigrationRecommendation:
    """개별 CLB 마이그레이션 분석

    타겟 결정, 호환성 분석, 복잡도 판정, 비용 비교, 체크리스트를 생성합니다.

    Args:
        clb: CLB 상세 정보
        region: AWS 리전

    Returns:
        마이그레이션 추천 결과
    """

    # 추천 타겟 결정
    target = _determine_target(clb)

    # 호환성 분석
    issues = _analyze_compatibility(clb, target)

    # 복잡도 결정
    complexity = _determine_complexity(clb, issues)

    # 비용 분석
    current_cost = get_elb_monthly_cost(region, "classic")
    if target == RecommendedTarget.ALB:
        estimated_cost = get_elb_monthly_cost(region, "application")
    elif target == RecommendedTarget.NLB:
        estimated_cost = get_elb_monthly_cost(region, "network")
    elif target == RecommendedTarget.SPLIT:
        estimated_cost = get_elb_monthly_cost(region, "application") + get_elb_monthly_cost(region, "network")
    else:
        estimated_cost = current_cost

    # 체크리스트 생성
    checklist, warnings = _generate_checklist(clb, target, issues)

    # 요약 생성
    summary = _generate_summary(clb, target, complexity)

    return MigrationRecommendation(
        clb=clb,
        target=target,
        complexity=complexity,
        compatibility_issues=issues,
        current_monthly_cost=current_cost,
        estimated_monthly_cost=estimated_cost,
        cost_difference=estimated_cost - current_cost,
        checklist=checklist,
        warnings=warnings,
        summary=summary,
    )


def _determine_target(clb: CLBInfo) -> RecommendedTarget:
    """리스너 프로토콜 기반 추천 마이그레이션 타겟 결정

    Args:
        clb: CLB 상세 정보

    Returns:
        추천 타겟 (ALB, NLB, SPLIT, KEEP)
    """

    has_http = clb.has_http_listeners
    has_tcp = clb.has_tcp_listeners

    # HTTP/HTTPS만 → ALB
    if has_http and not has_tcp:
        return RecommendedTarget.ALB

    # TCP/SSL만 → NLB
    if has_tcp and not has_http:
        return RecommendedTarget.NLB

    # 혼합 → 분리 권장
    if has_http and has_tcp:
        return RecommendedTarget.SPLIT

    # 기타 (리스너 없음 등) → 유지
    return RecommendedTarget.KEEP


def _analyze_compatibility(clb: CLBInfo, target: RecommendedTarget) -> list[CompatibilityIssue]:
    """CLB 기능의 ALB/NLB 호환성 분석

    Args:
        clb: CLB 상세 정보
        target: 추천 마이그레이션 타겟

    Returns:
        호환성 이슈 리스트
    """
    issues = []

    # 1. VPC 기반 여부
    if not clb.is_vpc_based:
        issues.append(
            CompatibilityIssue(
                feature="EC2-Classic",
                status=CompatibilityStatus.INCOMPATIBLE,
                description="EC2-Classic 기반 CLB는 VPC로 먼저 마이그레이션 필요",
                workaround="VPC로 EC2 인스턴스 마이그레이션 후 진행",
            )
        )

    # 2. 프록시 프로토콜 (NLB 타겟 시)
    if target == RecommendedTarget.NLB:
        proxy_policies = [p for p, t in clb.policies.items() if "proxy" in t.lower()]
        if proxy_policies:
            issues.append(
                CompatibilityIssue(
                    feature="Proxy Protocol",
                    status=CompatibilityStatus.COMPATIBLE,
                    description="NLB도 Proxy Protocol v2 지원",
                    workaround=None,
                )
            )

    # 3. 스티키 세션 (ALB 타겟 시)
    if target in (RecommendedTarget.ALB, RecommendedTarget.SPLIT):
        sticky_policies = [p for p, t in clb.policies.items() if "sticky" in t.lower() or "LBCookie" in t]
        if sticky_policies:
            issues.append(
                CompatibilityIssue(
                    feature="Sticky Sessions",
                    status=CompatibilityStatus.PARTIAL,
                    description="ALB는 다른 스티키 세션 메커니즘 사용 (Application-based)",
                    workaround="Target Group 수준에서 스티키니스 재설정 필요",
                )
            )

    # 4. Backend Server Authentication
    backend_auth = [p for p, t in clb.policies.items() if "BackendServer" in t]
    if backend_auth:
        issues.append(
            CompatibilityIssue(
                feature="Backend Server Auth",
                status=CompatibilityStatus.INCOMPATIBLE,
                description="ALB/NLB는 Backend Server Authentication 미지원",
                workaround="애플리케이션 레벨에서 mTLS 구현 또는 Private CA 사용",
            )
        )

    # 5. 커넥션 드레이닝
    if clb.connection_draining_enabled:
        issues.append(
            CompatibilityIssue(
                feature="Connection Draining",
                status=CompatibilityStatus.COMPATIBLE,
                description="ALB/NLB 모두 Deregistration Delay로 지원",
                workaround=None,
            )
        )

    # 6. 헬스체크
    if clb.health_check:
        hc_target = clb.health_check.target
        if hc_target.startswith("TCP:") and target == RecommendedTarget.ALB:
            issues.append(
                CompatibilityIssue(
                    feature="TCP Health Check",
                    status=CompatibilityStatus.PARTIAL,
                    description="ALB는 HTTP/HTTPS 헬스체크만 지원",
                    workaround="HTTP 헬스체크 엔드포인트 추가 필요",
                )
            )

    # 7. 여러 포트 리스너
    unique_instance_ports = set(listener.instance_port for listener in clb.listeners)
    if len(unique_instance_ports) > 1:
        issues.append(
            CompatibilityIssue(
                feature="Multiple Backend Ports",
                status=CompatibilityStatus.PARTIAL,
                description=f"다중 백엔드 포트 사용 ({len(unique_instance_ports)}개)",
                workaround="각 포트별로 별도 Target Group 생성 필요",
            )
        )

    # 8. SSL 정책
    ssl_listeners = [listener for listener in clb.listeners if listener.protocol in ("HTTPS", "SSL")]
    if ssl_listeners:
        issues.append(
            CompatibilityIssue(
                feature="SSL/TLS Termination",
                status=CompatibilityStatus.COMPATIBLE,
                description="SSL 인증서 ACM으로 마이그레이션 권장",
                workaround="IAM 인증서 → ACM 인증서 전환",
            )
        )

    return issues


def _determine_complexity(clb: CLBInfo, issues: list[CompatibilityIssue]) -> MigrationComplexity:
    """마이그레이션 복잡도 결정"""

    # 호환 불가 이슈 → 복잡
    incompatible = sum(1 for i in issues if i.status == CompatibilityStatus.INCOMPATIBLE)
    if incompatible > 0:
        return MigrationComplexity.COMPLEX

    # 부분 호환 이슈 2개 이상 → 중간
    partial = sum(1 for i in issues if i.status == CompatibilityStatus.PARTIAL)
    if partial >= 2:
        return MigrationComplexity.MODERATE

    # 리스너 3개 이상 → 중간
    if len(clb.listeners) >= 3:
        return MigrationComplexity.MODERATE

    # 인스턴스 10개 이상 → 중간
    if len(clb.instances) >= 10:
        return MigrationComplexity.MODERATE

    return MigrationComplexity.SIMPLE


def _generate_checklist(
    clb: CLBInfo, target: RecommendedTarget, issues: list[CompatibilityIssue]
) -> tuple[list[str], list[str]]:
    """마이그레이션 체크리스트 및 경고 생성"""
    checklist = []
    warnings = []

    # 기본 체크리스트
    checklist.append("1. 새 Target Group 생성")
    checklist.append("2. EC2 인스턴스를 Target Group에 등록")

    if target == RecommendedTarget.ALB:
        checklist.append("3. ALB 생성 (동일 서브넷/보안그룹)")
        checklist.append("4. 리스너 및 규칙 설정")
    elif target == RecommendedTarget.NLB:
        checklist.append("3. NLB 생성 (동일 서브넷)")
        checklist.append("4. 리스너 설정")
    elif target == RecommendedTarget.SPLIT:
        checklist.append("3. HTTP/HTTPS용 ALB 생성")
        checklist.append("4. TCP용 NLB 생성")
        warnings.append("두 개의 LB로 분리되어 DNS 설정 변경 필요")

    checklist.append(f"{len(checklist) + 1}. 헬스체크 설정 확인")
    checklist.append(f"{len(checklist) + 1}. DNS 가중치 기반 점진적 전환")
    checklist.append(f"{len(checklist) + 1}. 모니터링 및 롤백 계획 수립")
    checklist.append(f"{len(checklist) + 1}. CLB 삭제 (전환 완료 후)")

    # 이슈별 추가 항목
    for issue in issues:
        if issue.status == CompatibilityStatus.PARTIAL and issue.workaround:
            checklist.append(f"• {issue.feature}: {issue.workaround}")
        if issue.status == CompatibilityStatus.INCOMPATIBLE:
            warnings.append(f"{issue.feature}: {issue.description}")

    # 추가 경고
    if clb.scheme == "internet-facing":
        warnings.append("인터넷 페이싱 LB - DNS 변경 시 다운타임 주의")

    if len(clb.instances) == 0:
        warnings.append("등록된 인스턴스 없음 - 삭제 검토")

    return checklist, warnings


def _generate_summary(clb: CLBInfo, target: RecommendedTarget, complexity: MigrationComplexity) -> str:
    """마이그레이션 요약 생성"""

    target_names = {
        RecommendedTarget.ALB: "Application Load Balancer (ALB)",
        RecommendedTarget.NLB: "Network Load Balancer (NLB)",
        RecommendedTarget.SPLIT: "ALB + NLB 분리",
        RecommendedTarget.KEEP: "CLB 유지",
    }

    complexity_desc = {
        MigrationComplexity.SIMPLE: "단순",
        MigrationComplexity.MODERATE: "중간",
        MigrationComplexity.COMPLEX: "복잡",
    }

    protocols = set(listener.protocol for listener in clb.listeners)

    return (
        f"{clb.name}: {'/'.join(protocols)} 리스너 → "
        f"{target_names[target]} 추천 "
        f"(복잡도: {complexity_desc[complexity]})"
    )


def analyze_all(clbs: list[CLBInfo], region: str, account_id: str, account_name: str) -> MigrationAnalysisResult:
    """전체 CLB 마이그레이션 분석

    Args:
        clbs: CLB 상세 정보 리스트
        region: AWS 리전
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름

    Returns:
        계정/리전별 마이그레이션 분석 결과
    """
    result = MigrationAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_clbs=len(clbs),
    )

    for clb in clbs:
        rec = analyze_migration(clb, region)
        result.recommendations.append(rec)

        # 통계 집계
        if rec.target == RecommendedTarget.ALB:
            result.alb_recommended += 1
        elif rec.target == RecommendedTarget.NLB:
            result.nlb_recommended += 1
        elif rec.target == RecommendedTarget.SPLIT:
            result.split_recommended += 1
        else:
            result.keep_recommended += 1

        result.total_current_cost += rec.current_monthly_cost
        result.total_estimated_cost += rec.estimated_monthly_cost

    return result


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[MigrationAnalysisResult], output_dir: str) -> str:
    """CLB 마이그레이션 분석 Excel 보고서 생성

    Summary, Recommendations, Checklists 3개 시트를 포함합니다.

    Args:
        results: 계정/리전별 분석 결과 리스트
        output_dir: 출력 디렉토리 경로

    Returns:
        생성된 Excel 파일 경로
    """
    from openpyxl.styles import Font, PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 셀 수준 조건부 스타일링용 Fill
    target_fills = {
        RecommendedTarget.ALB: PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        RecommendedTarget.NLB: PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid"),
        RecommendedTarget.SPLIT: PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid"),
        RecommendedTarget.KEEP: PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
    }
    target_font = Font(bold=True, color="FFFFFF")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Total CLBs", width=12, style="number"),
        ColumnDef(header="ALB 추천", width=12, style="number"),
        ColumnDef(header="NLB 추천", width=12, style="number"),
        ColumnDef(header="분리 추천", width=12, style="number"),
        ColumnDef(header="유지 추천", width=12, style="number"),
        ColumnDef(header="현재 월 비용", width=15),
        ColumnDef(header="예상 월 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_clbs,
                r.alb_recommended,
                r.nlb_recommended,
                r.split_recommended,
                r.keep_recommended,
                f"${r.total_current_cost:.2f}",
                f"${r.total_estimated_cost:.2f}",
            ]
        )

    # Recommendations 시트
    rec_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="CLB Name", width=25),
        ColumnDef(header="Scheme", width=15),
        ColumnDef(header="Listeners", width=20),
        ColumnDef(header="Instances", width=12, style="number"),
        ColumnDef(header="추천 타겟", width=12, style="center"),
        ColumnDef(header="복잡도", width=12, style="center"),
        ColumnDef(header="호환성", width=12, style="center"),
        ColumnDef(header="현재 비용($)", width=14, style="number"),
        ColumnDef(header="예상 비용($)", width=14, style="number"),
        ColumnDef(header="차이($)", width=12, style="number"),
        ColumnDef(header="요약", width=50),
    ]
    rec_sheet = wb.new_sheet("Recommendations", rec_columns)

    for result in results:
        for rec in result.recommendations:
            clb = rec.clb
            listener_str = ", ".join(f"{listener.protocol}:{listener.port}" for listener in clb.listeners)

            style = None
            if rec.target == RecommendedTarget.SPLIT:
                style = Styles.warning()

            row_num = rec_sheet.add_row(
                [
                    clb.account_name,
                    clb.region,
                    clb.name,
                    clb.scheme,
                    listener_str,
                    len(clb.instances),
                    rec.target.value,
                    rec.complexity.value,
                    rec.overall_status.value,
                    round(rec.current_monthly_cost, 2),
                    round(rec.estimated_monthly_cost, 2),
                    round(rec.cost_difference, 2),
                    rec.summary,
                ],
                style=style,
            )
            # 추천 타겟 셀에 색상 적용
            ws = rec_sheet._ws
            if rec.target in target_fills:
                ws.cell(row=row_num, column=7).fill = target_fills[rec.target]
                ws.cell(row=row_num, column=7).font = target_font

    # Checklists 시트
    checklist_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="CLB Name", width=25),
        ColumnDef(header="타겟", width=12, style="center"),
        ColumnDef(header="체크리스트", width=60, style="wrap"),
        ColumnDef(header="경고사항", width=50, style="wrap"),
    ]
    checklist_sheet = wb.new_sheet("Checklists", checklist_columns)

    for result in results:
        for rec in result.recommendations:
            style = None
            if rec.warnings:
                style = Styles.warning()

            checklist_sheet.add_row(
                [
                    rec.clb.account_name,
                    rec.clb.region,
                    rec.clb.name,
                    rec.target.value,
                    "\n".join(rec.checklist),
                    "\n".join(rec.warnings) if rec.warnings else "-",
                ],
                style=style,
            )

    return str(wb.save_as(output_dir, "CLB_Migration_Advisor"))


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> MigrationAnalysisResult | None:
    """단일 계정/리전의 CLB 수집 및 마이그레이션 분석 (parallel_collect 콜백)

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        마이그레이션 분석 결과. CLB가 없으면 None.
    """
    clbs = collect_clb_details(session, account_id, account_name, region)
    if not clbs:
        return None
    return analyze_all(clbs, region, account_id, account_name)


def run(ctx: ExecutionContext) -> None:
    """CLB Migration Advisor 실행

    멀티 계정/리전에서 CLB를 병렬 수집하고,
    ALB/NLB 마이그레이션 적합성을 분석하여 Excel 보고서를 생성합니다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 계정/리전 선택, 출력 설정 포함)
    """
    console.print("[bold]CLB Migration Advisor 시작...[/bold]")
    console.print("[dim]Classic Load Balancer → ALB/NLB 마이그레이션 분석[/dim]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="elb")
    all_results: list[MigrationAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not all_results:
        console.print("\n[yellow]분석할 CLB 없음[/yellow]")
        return

    # 전체 요약
    totals = {
        "total": sum(r.total_clbs for r in all_results),
        "alb": sum(r.alb_recommended for r in all_results),
        "nlb": sum(r.nlb_recommended for r in all_results),
        "split": sum(r.split_recommended for r in all_results),
        "current_cost": sum(r.total_current_cost for r in all_results),
        "estimated_cost": sum(r.total_estimated_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 CLB: {totals['total']}개[/bold]")
    console.print(f"  [green]ALB 추천: {totals['alb']}개[/green]")
    console.print(f"  [blue]NLB 추천: {totals['nlb']}개[/blue]")
    if totals["split"] > 0:
        console.print(f"  [yellow]분리 추천: {totals['split']}개[/yellow]")

    cost_diff = totals["estimated_cost"] - totals["current_cost"]
    if cost_diff > 0:
        console.print(f"\n  [yellow]예상 비용 증가: +${cost_diff:.2f}/월[/yellow]")
    else:
        console.print(f"\n  [green]예상 비용 절감: ${abs(cost_diff):.2f}/월[/green]")

    # 보고서 생성
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("elb", "inventory").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
