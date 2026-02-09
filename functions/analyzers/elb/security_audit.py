"""
functions/analyzers/elb/security_audit.py - ELB 보안 감사

ELB의 보안 설정을 종합 분석하여 취약점을 탐지합니다.

검사 항목:
1. SSL/TLS 정책 분석
   - 취약한 TLS 버전 (TLS 1.0, 1.1)
   - 약한 암호화 스위트 (RC4, DES, 3DES 등)
   - 만료 임박 인증서

2. WAF 연결 상태 (ALB만)
   - WAF WebACL 미연결 탐지
   - 인터넷 페이싱 ALB 우선 검사

3. 액세스 로그 설정
   - 액세스 로그 미활성화 탐지
   - 감사 추적 불가 리스크

4. 삭제 보호 (Deletion Protection)
   - 프로덕션 LB 삭제 보호 미설정
   - 실수로 인한 삭제 방지

5. 보안 리스너 설정
   - HTTP→HTTPS 리다이렉트 미설정
   - HTTPS 리스너 없는 인터넷 페이싱 LB

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from typing import TYPE_CHECKING, Any

from rich.console import Console

from core.config import settings
from core.parallel import get_client, is_quiet, parallel_collect
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticloadbalancing:DescribeLoadBalancers",
        "elasticloadbalancing:DescribeLoadBalancerAttributes",
        "elasticloadbalancing:DescribeListeners",
        "elasticloadbalancing:DescribeSSLPolicies",
        "wafv2:GetWebACLForResource",
        "acm:DescribeCertificate",
    ],
}


# =============================================================================
# 중앙 설정에서 보안 정책 로드
# =============================================================================

# core/config.py의 Settings 클래스에서 가져옴
# 회사 정책에 맞게 config.py만 수정하면 됨
VULNERABLE_PROTOCOLS = settings.SECURITY_VULNERABLE_PROTOCOLS
SECURE_PROTOCOLS = settings.SECURITY_SECURE_PROTOCOLS
WEAK_CIPHER_PATTERNS = settings.SECURITY_WEAK_CIPHER_PATTERNS
DEFAULT_MIN_TLS_VERSION = settings.SECURITY_MIN_TLS_VERSION

# 인증서 만료 임계값
CERT_EXPIRY_CRITICAL = settings.SECURITY_CERT_EXPIRY_CRITICAL
CERT_EXPIRY_HIGH = settings.SECURITY_CERT_EXPIRY_HIGH
CERT_EXPIRY_MEDIUM = settings.SECURITY_CERT_EXPIRY_MEDIUM

# 삭제 보호 필수 패턴
DELETION_PROTECTION_PATTERNS = settings.SECURITY_DELETION_PROTECTION_PATTERNS


class SSLPolicyAnalyzer:
    """SSL 정책 동적 분석기

    AWS API를 통해 실제 정책 정보를 조회하여 분석합니다.
    하드코딩 대신 동적으로 프로토콜/암호 스위트를 검사합니다.

    Attributes:
        client: ELBv2 boto3 클라이언트
        min_tls_version: 최소 허용 TLS 버전 (예: "TLSv1.2")
        _policy_cache: 정책 이름별 상세 정보 캐시
    """

    def __init__(self, elbv2_client, min_tls_version: str = DEFAULT_MIN_TLS_VERSION):
        self.client = elbv2_client
        self.min_tls_version = min_tls_version
        self._policy_cache: dict[str, dict] = {}

    def get_policy_details(self, policy_name: str) -> dict | None:
        """AWS API로 SSL 정책 상세 정보 조회

        정책 이름으로 AWS API를 호출하여 프로토콜, 암호 스위트 등
        상세 정보를 가져옵니다. 결과는 내부 캐시에 저장됩니다.

        Args:
            policy_name: SSL 정책 이름 (예: "ELBSecurityPolicy-TLS13-1-2-2021-06")

        Returns:
            정책 상세 정보 딕셔너리. 조회 실패 시 None.
        """
        from botocore.exceptions import ClientError

        if policy_name in self._policy_cache:
            return self._policy_cache[policy_name]

        try:
            response = self.client.describe_ssl_policies(Names=[policy_name])
            policies = response.get("SslPolicies", [])
            if policies:
                policy: dict[str, Any] = policies[0]
                self._policy_cache[policy_name] = policy
                return policy
        except ClientError:
            pass
        return None

    def analyze_policy(self, policy_name: str) -> dict[str, Any]:
        """SSL 정책 보안 취약점 분석

        AWS API로 정책 상세 정보를 조회하여 취약 프로토콜(TLS 1.0/1.1),
        약한 암호 스위트(RC4, DES 등)를 탐지합니다. API 조회 실패 시
        정책 이름 기반 휴리스틱 분석으로 fallback합니다.

        Args:
            policy_name: SSL 정책 이름

        Returns:
            분석 결과 딕셔너리. is_vulnerable, risk_level, issues,
            protocols, weak_ciphers, recommendation 키 포함.
        """
        issues: list[str] = []
        protocols: list[str] = []
        weak_ciphers: list[str] = []
        result: dict[str, Any] = {
            "is_vulnerable": False,
            "risk_level": "info",
            "issues": issues,
            "protocols": protocols,
            "weak_ciphers": weak_ciphers,
            "recommendation": "",
        }

        policy = self.get_policy_details(policy_name)
        if not policy:
            # API 조회 실패 시 정책 이름으로 휴리스틱 판단
            return self._analyze_by_name(policy_name)

        # 프로토콜 분석
        protocols = policy.get("SslProtocols", [])
        result["protocols"] = protocols

        vulnerable_protocols = [p for p in protocols if p in VULNERABLE_PROTOCOLS]
        if vulnerable_protocols:
            result["is_vulnerable"] = True
            result["issues"].append(f"취약 프로토콜: {', '.join(vulnerable_protocols)}")
            result["risk_level"] = "high"

        # TLS 1.2/1.3 없으면 위험
        if not any(p in SECURE_PROTOCOLS for p in protocols):
            result["is_vulnerable"] = True
            result["issues"].append("TLS 1.2/1.3 미지원")
            result["risk_level"] = "critical"

        # 암호화 스위트 분석
        ciphers = policy.get("Ciphers", [])
        for cipher_info in ciphers:
            cipher_name = cipher_info.get("Name", "")
            for weak_pattern in WEAK_CIPHER_PATTERNS:
                if weak_pattern.upper() in cipher_name.upper():
                    result["weak_ciphers"].append(cipher_name)
                    result["is_vulnerable"] = True
                    break

        if result["weak_ciphers"]:
            result["issues"].append(f"약한 암호: {len(result['weak_ciphers'])}개")
            if result["risk_level"] != "critical":
                result["risk_level"] = "medium"

        # 권장사항
        if result["is_vulnerable"]:
            result["recommendation"] = self._get_recommendation(protocols)

        return result

    def _analyze_by_name(self, policy_name: str) -> dict[str, Any]:
        """정책 이름 기반 휴리스틱 분석 (API 실패 시 fallback)

        정책 이름에 포함된 TLS 버전 및 연도 정보를 기반으로
        취약 여부를 판단합니다.

        Args:
            policy_name: SSL 정책 이름

        Returns:
            분석 결과 딕셔너리. analyze_policy()와 동일한 구조.
        """
        issues: list[str] = []
        protocols: list[str] = []
        weak_ciphers: list[str] = []
        result: dict[str, Any] = {
            "is_vulnerable": False,
            "risk_level": "info",
            "issues": issues,
            "protocols": protocols,
            "weak_ciphers": weak_ciphers,
            "recommendation": "",
        }

        name_lower = policy_name.lower()

        # TLS 1.0/1.1 명시된 정책
        if "tls-1-0" in name_lower or "tls-1-1" in name_lower:
            result["is_vulnerable"] = True
            result["risk_level"] = "high"
            result["issues"].append("정책명에 TLS 1.0/1.1 포함")

        # 오래된 정책 (2016년 이전)
        elif any(year in name_lower for year in ["2015", "2014", "2013"]):
            result["is_vulnerable"] = True
            result["risk_level"] = "high"
            result["issues"].append("오래된 정책 (2016년 이전)")

        # TLS 1.2/1.3 명시된 정책은 안전
        elif "tls-1-2" in name_lower or "tls13" in name_lower:
            result["risk_level"] = "info"

        if result["is_vulnerable"]:
            result["recommendation"] = "ELBSecurityPolicy-TLS13-1-2-2021-06 이상 권장"

        return result

    def _get_recommendation(self, current_protocols: list[str]) -> str:
        """현재 프로토콜 구성에 기반한 권장 정책 반환

        Args:
            current_protocols: 현재 지원 중인 TLS 프로토콜 목록

        Returns:
            권장 SSL 정책 이름 및 설명 문자열
        """
        if "TLSv1.3" in current_protocols:
            return "현재 TLS 1.3 지원 중. 취약 프로토콜만 제거 권장"
        elif "TLSv1.2" in current_protocols:
            return "ELBSecurityPolicy-TLS13-1-2-2021-06 (TLS 1.3 + 1.2) 권장"
        else:
            return "ELBSecurityPolicy-TLS-1-2-2017-01 이상 필수 적용"


class RiskLevel(Enum):
    """보안 발견 항목의 위험 수준 분류

    CRITICAL부터 INFO까지 5단계로 구분하며,
    보안 감사 보고서의 우선순위 결정에 사용됩니다.
    """

    CRITICAL = "critical"  # 즉시 조치 필요
    HIGH = "high"  # 높은 위험
    MEDIUM = "medium"  # 중간 위험
    LOW = "low"  # 낮은 위험
    INFO = "info"  # 참고


class FindingCategory(Enum):
    """보안 발견 항목의 카테고리 분류

    SSL/TLS, WAF, 액세스 로그, 삭제 보호, 리스너, 인증서 등
    보안 검사 영역별 분류에 사용됩니다.
    """

    SSL_TLS = "ssl_tls"
    WAF = "waf"
    ACCESS_LOG = "access_log"
    DELETION_PROTECTION = "deletion_protection"
    LISTENER = "listener"
    CERTIFICATE = "certificate"


@dataclass
class SecurityFinding:
    """보안 발견 항목

    Attributes:
        category: 발견 항목 카테고리 (SSL_TLS, WAF, ACCESS_LOG 등)
        risk_level: 위험 수준 (CRITICAL, HIGH, MEDIUM, LOW, INFO)
        title: 발견 항목 제목 (간략 설명)
        description: 상세 설명
        recommendation: 권장 조치 사항
        details: 추가 상세 정보 딕셔너리
    """

    category: FindingCategory
    risk_level: RiskLevel
    title: str
    description: str
    recommendation: str
    details: dict[str, Any] = field(default_factory=dict)


@dataclass
class ListenerInfo:
    """리스너 정보

    Attributes:
        arn: 리스너 ARN (CLB의 경우 빈 문자열)
        protocol: 리스너 프로토콜 (HTTP, HTTPS, TCP, TLS 등)
        port: 리스너 포트 번호
        ssl_policy: SSL 정책 이름. HTTPS/TLS 리스너만 해당.
        certificates: 연결된 인증서 ARN 목록
        default_actions: 기본 액션 설정 목록
    """

    arn: str
    protocol: str
    port: int
    ssl_policy: str | None = None
    certificates: list[str] = field(default_factory=list)
    default_actions: list[dict] = field(default_factory=list)


@dataclass
class LBSecurityInfo:
    """LB 보안 정보

    Attributes:
        arn: 로드밸런서 ARN
        name: 로드밸런서 이름
        lb_type: LB 유형 (application, network, gateway, classic)
        scheme: 스킴 (internet-facing, internal)
        dns_name: DNS 이름
        vpc_id: VPC ID
        state: 현재 상태
        created_time: 생성 시각
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        access_logs_enabled: 액세스 로그 활성화 여부
        access_logs_bucket: 액세스 로그 S3 버킷 이름
        deletion_protection: 삭제 보호 활성화 여부
        waf_web_acl_arn: WAF WebACL ARN. 미연결 시 None.
        listeners: 리스너 정보 목록
        findings: 보안 발견 항목 목록
    """

    # 기본 정보
    arn: str
    name: str
    lb_type: str  # application, network, gateway, classic
    scheme: str  # internet-facing, internal
    dns_name: str
    vpc_id: str
    state: str
    created_time: datetime | None

    # 계정/리전
    account_id: str
    account_name: str
    region: str

    # 보안 속성
    access_logs_enabled: bool = False
    access_logs_bucket: str = ""
    deletion_protection: bool = False
    waf_web_acl_arn: str | None = None

    # 리스너
    listeners: list[ListenerInfo] = field(default_factory=list)

    # 발견 항목
    findings: list[SecurityFinding] = field(default_factory=list)

    @property
    def is_internet_facing(self) -> bool:
        """인터넷 페이싱 여부 확인

        Returns:
            scheme이 "internet-facing"이면 True
        """
        return self.scheme == "internet-facing"

    @property
    def has_https_listener(self) -> bool:
        """HTTPS/TLS 리스너 보유 여부 확인

        Returns:
            HTTPS 또는 TLS 프로토콜 리스너가 하나 이상 있으면 True
        """
        return any(listener.protocol in ("HTTPS", "TLS") for listener in self.listeners)

    @property
    def risk_score(self) -> int:
        """위험 점수 계산 (높을수록 위험)

        발견 항목의 위험 수준별 가중치를 합산합니다.
        CRITICAL=100, HIGH=50, MEDIUM=20, LOW=5.

        Returns:
            합산된 위험 점수
        """
        score = 0
        for f in self.findings:
            if f.risk_level == RiskLevel.CRITICAL:
                score += 100
            elif f.risk_level == RiskLevel.HIGH:
                score += 50
            elif f.risk_level == RiskLevel.MEDIUM:
                score += 20
            elif f.risk_level == RiskLevel.LOW:
                score += 5
        return score


@dataclass
class SecurityAuditResult:
    """보안 감사 결과

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        load_balancers: 분석된 LB 보안 정보 목록
        total_count: 전체 LB 수
        critical_count: CRITICAL 위험 발견 수
        high_count: HIGH 위험 발견 수
        medium_count: MEDIUM 위험 발견 수
        low_count: LOW 위험 발견 수
    """

    account_id: str
    account_name: str
    region: str
    load_balancers: list[LBSecurityInfo] = field(default_factory=list)

    # 통계
    total_count: int = 0
    critical_count: int = 0
    high_count: int = 0
    medium_count: int = 0
    low_count: int = 0


# =============================================================================
# 수집
# =============================================================================


def collect_v2_lb_security(session, account_id: str, account_name: str, region: str) -> list[LBSecurityInfo]:
    """ALB/NLB/GWLB 보안 정보 수집

    ELBv2 API로 로드밸런서, 속성(액세스 로그, 삭제 보호),
    WAF WebACL(ALB만), 리스너 정보를 수집합니다.

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        LB 보안 정보 목록
    """
    from botocore.exceptions import ClientError

    load_balancers = []

    try:
        elbv2 = get_client(session, "elbv2", region_name=region)

        # Load Balancers 조회
        paginator = elbv2.get_paginator("describe_load_balancers")
        for page in paginator.paginate():
            for data in page.get("LoadBalancers", []):
                lb_arn = data.get("LoadBalancerArn", "")

                lb = LBSecurityInfo(
                    arn=lb_arn,
                    name=data.get("LoadBalancerName", ""),
                    lb_type=data.get("Type", "application"),
                    scheme=data.get("Scheme", ""),
                    dns_name=data.get("DNSName", ""),
                    vpc_id=data.get("VpcId", ""),
                    state=data.get("State", {}).get("Code", ""),
                    created_time=data.get("CreatedTime"),
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                )

                # 속성 조회 (액세스 로그, 삭제 보호)
                try:
                    attrs = elbv2.describe_load_balancer_attributes(LoadBalancerArn=lb_arn)
                    for attr in attrs.get("Attributes", []):
                        key = attr.get("Key", "")
                        value = attr.get("Value", "")

                        if key == "access_logs.s3.enabled":
                            lb.access_logs_enabled = value.lower() == "true"
                        elif key == "access_logs.s3.bucket":
                            lb.access_logs_bucket = value
                        elif key == "deletion_protection.enabled":
                            lb.deletion_protection = value.lower() == "true"
                except ClientError:
                    pass

                # WAF WebACL 조회 (ALB만)
                if lb.lb_type == "application":
                    try:
                        wafv2 = get_client(session, "wafv2", region_name=region)
                        waf_resp = wafv2.get_web_acl_for_resource(ResourceArn=lb_arn)
                        if waf_resp.get("WebACL"):
                            lb.waf_web_acl_arn = waf_resp["WebACL"].get("ARN")
                    except ClientError:
                        pass

                # 리스너 조회
                lb.listeners = _get_listeners(elbv2, lb_arn)

                load_balancers.append(lb)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} ELBv2 수집 오류: {error_code}[/yellow]")

    return load_balancers


def _get_listeners(elbv2, lb_arn: str) -> list[ListenerInfo]:
    """지정 LB의 리스너 정보 조회

    Args:
        elbv2: ELBv2 boto3 클라이언트
        lb_arn: 로드밸런서 ARN

    Returns:
        리스너 정보 목록. 조회 실패 시 빈 리스트.
    """
    from botocore.exceptions import ClientError

    listeners = []

    try:
        response = elbv2.describe_listeners(LoadBalancerArn=lb_arn)

        for data in response.get("Listeners", []):
            listener = ListenerInfo(
                arn=data.get("ListenerArn", ""),
                protocol=data.get("Protocol", ""),
                port=data.get("Port", 0),
                ssl_policy=data.get("SslPolicy"),
                certificates=[c.get("CertificateArn", "") for c in data.get("Certificates", [])],
                default_actions=data.get("DefaultActions", []),
            )
            listeners.append(listener)

    except ClientError:
        pass

    return listeners


def collect_classic_lb_security(session, account_id: str, account_name: str, region: str) -> list[LBSecurityInfo]:
    """Classic Load Balancer 보안 정보 수집

    ELB API로 CLB의 속성(액세스 로그), 리스너 정보를 수집합니다.
    CLB는 WAF 및 삭제 보호를 지원하지 않습니다.

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        CLB 보안 정보 목록
    """
    from botocore.exceptions import ClientError

    load_balancers = []

    try:
        elb = get_client(session, "elb", region_name=region)

        response = elb.describe_load_balancers()

        for data in response.get("LoadBalancerDescriptions", []):
            lb_name = data.get("LoadBalancerName", "")

            lb = LBSecurityInfo(
                arn=f"arn:aws:elasticloadbalancing:{region}:{account_id}:loadbalancer/{lb_name}",
                name=lb_name,
                lb_type="classic",
                scheme=data.get("Scheme", ""),
                dns_name=data.get("DNSName", ""),
                vpc_id=data.get("VPCId", ""),
                state="active",
                created_time=data.get("CreatedTime"),
                account_id=account_id,
                account_name=account_name,
                region=region,
            )

            # 속성 조회
            try:
                attrs = elb.describe_load_balancer_attributes(LoadBalancerName=lb_name)
                access_log = attrs.get("LoadBalancerAttributes", {}).get("AccessLog", {})
                lb.access_logs_enabled = access_log.get("Enabled", False)
                lb.access_logs_bucket = access_log.get("S3BucketName", "")
            except ClientError:
                pass

            # 리스너 정보
            for listener_desc in data.get("ListenerDescriptions", []):
                listener_data = listener_desc.get("Listener", {})
                listener = ListenerInfo(
                    arn="",
                    protocol=listener_data.get("Protocol", ""),
                    port=listener_data.get("LoadBalancerPort", 0),
                    ssl_policy=listener_data.get("SSLCertificateId"),
                )
                lb.listeners.append(listener)

            load_balancers.append(lb)

    except ClientError as e:
        if "not available" not in str(e).lower():
            error_code = e.response.get("Error", {}).get("Code", "Unknown")
            if not is_quiet():
                console.print(f"    [yellow]{account_name}/{region} CLB 수집 오류: {error_code}[/yellow]")

    return load_balancers


# =============================================================================
# 분석
# =============================================================================


def analyze_security(lb: LBSecurityInfo, session, region: str) -> None:
    """개별 LB 보안 종합 분석

    SSL/TLS 정책, 인증서, WAF, 액세스 로그, 삭제 보호, 리스너 보안을
    순차적으로 분석하고 발견된 취약점을 lb.findings에 추가합니다.

    Args:
        lb: 분석 대상 LB 보안 정보 (findings가 in-place 업데이트됨)
        session: boto3 Session 객체
        region: AWS 리전
    """
    from botocore.exceptions import ClientError

    # SSL 정책 분석기 생성 (ALB/NLB만, CLB는 None)
    ssl_analyzer: SSLPolicyAnalyzer | None = None
    if lb.lb_type != "classic":
        try:
            elbv2 = get_client(session, "elbv2", region_name=region)
            ssl_analyzer = SSLPolicyAnalyzer(elbv2)
        except ClientError:
            pass

    # 1. SSL/TLS 정책 분석
    _analyze_ssl_policy(lb, ssl_analyzer)

    # 2. 인증서 분석
    _analyze_certificates(lb, session, region)

    # 3. WAF 분석 (ALB만)
    _analyze_waf(lb)

    # 4. 액세스 로그 분석
    _analyze_access_logs(lb)

    # 5. 삭제 보호 분석
    _analyze_deletion_protection(lb)

    # 6. 리스너 보안 분석
    _analyze_listener_security(lb)


def _analyze_ssl_policy(lb: LBSecurityInfo, ssl_analyzer: SSLPolicyAnalyzer | None) -> None:
    """SSL/TLS 정책 분석 (AWS API 기반 동적 분석)

    HTTPS/TLS 리스너의 SSL 정책을 분석하여 취약 프로토콜 및
    약한 암호 스위트를 탐지합니다.

    Args:
        lb: 분석 대상 LB 보안 정보 (findings가 in-place 업데이트됨)
        ssl_analyzer: SSL 정책 분석기. None이면 fallback 휴리스틱 사용.
    """
    for listener in lb.listeners:
        if listener.protocol not in ("HTTPS", "TLS"):
            continue

        policy = listener.ssl_policy
        if not policy:
            continue

        # SSLPolicyAnalyzer로 동적 분석, Fallback: 정책 이름 기반 휴리스틱
        analysis = ssl_analyzer.analyze_policy(policy) if ssl_analyzer else _fallback_policy_analysis(policy)

        if not analysis["is_vulnerable"]:
            continue

        # 위험 수준 매핑
        risk_map = {
            "critical": RiskLevel.CRITICAL,
            "high": RiskLevel.HIGH,
            "medium": RiskLevel.MEDIUM,
            "low": RiskLevel.LOW,
            "info": RiskLevel.INFO,
        }
        risk_level = risk_map.get(analysis["risk_level"], RiskLevel.MEDIUM)

        lb.findings.append(
            SecurityFinding(
                category=FindingCategory.SSL_TLS,
                risk_level=risk_level,
                title="취약한 TLS 정책 사용",
                description=f"리스너 :{listener.port} - {', '.join(analysis['issues'])}",
                recommendation=analysis["recommendation"],
                details={
                    "listener_port": listener.port,
                    "current_policy": policy,
                    "protocols": analysis.get("protocols", []),
                    "weak_ciphers": analysis.get("weak_ciphers", []),
                },
            )
        )


def _fallback_policy_analysis(policy_name: str) -> dict[str, Any]:
    """API 없이 정책 이름 기반 휴리스틱 분석

    SSLPolicyAnalyzer를 사용할 수 없을 때 정책 이름에서
    TLS 버전 및 연도 정보를 파싱하여 취약 여부를 판단합니다.

    Args:
        policy_name: SSL 정책 이름

    Returns:
        분석 결과 딕셔너리. analyze_policy()와 동일한 구조.
    """
    issues: list[str] = []
    protocols: list[str] = []
    weak_ciphers: list[str] = []
    result: dict[str, Any] = {
        "is_vulnerable": False,
        "risk_level": "info",
        "issues": issues,
        "protocols": protocols,
        "weak_ciphers": weak_ciphers,
        "recommendation": "",
    }

    name_lower = policy_name.lower()

    # TLS 1.0/1.1 명시된 정책
    if "tls-1-0" in name_lower or "tls-1-1" in name_lower:
        result["is_vulnerable"] = True
        result["risk_level"] = "high"
        result["issues"].append("TLS 1.0/1.1 정책")
        result["recommendation"] = "ELBSecurityPolicy-TLS-1-2-2017-01 이상 권장"

    # 오래된 정책 (2016년 이전)
    elif any(year in name_lower for year in ["2015", "2014", "2013"]):
        result["is_vulnerable"] = True
        result["risk_level"] = "high"
        result["issues"].append("오래된 정책 (2016년 이전)")
        result["recommendation"] = "최신 TLS 정책으로 업그레이드 권장"

    return result


def _analyze_certificates(lb: LBSecurityInfo, session, region: str) -> None:
    """ACM 인증서 만료 분석

    리스너에 연결된 인증서의 만료일을 확인하고,
    중앙 설정의 임계값에 따라 위험 수준을 분류합니다.

    Args:
        lb: 분석 대상 LB 보안 정보 (findings가 in-place 업데이트됨)
        session: boto3 Session 객체
        region: AWS 리전
    """
    from botocore.exceptions import ClientError

    cert_arns: set[str] = set()
    for listener in lb.listeners:
        cert_arns.update(listener.certificates)

    if not cert_arns:
        return

    try:
        acm = get_client(session, "acm", region_name=region)

        for cert_arn in cert_arns:
            try:
                cert = acm.describe_certificate(CertificateArn=cert_arn)
                cert_detail = cert.get("Certificate", {})

                # 만료일 확인
                not_after = cert_detail.get("NotAfter")
                if not_after:
                    now = datetime.now(timezone.utc)
                    days_until_expiry = (not_after - now).days

                    # 중앙 설정의 임계값 사용
                    if days_until_expiry < 0:
                        lb.findings.append(
                            SecurityFinding(
                                category=FindingCategory.CERTIFICATE,
                                risk_level=RiskLevel.CRITICAL,
                                title="만료된 인증서",
                                description=f"인증서가 {abs(days_until_expiry)}일 전에 만료됨",
                                recommendation="즉시 인증서 갱신 필요",
                                details={
                                    "certificate_arn": cert_arn,
                                    "expired_days_ago": abs(days_until_expiry),
                                    "domain": cert_detail.get("DomainName", ""),
                                },
                            )
                        )
                    elif days_until_expiry <= CERT_EXPIRY_CRITICAL:
                        lb.findings.append(
                            SecurityFinding(
                                category=FindingCategory.CERTIFICATE,
                                risk_level=RiskLevel.CRITICAL,
                                title=f"인증서 만료 임박 ({CERT_EXPIRY_CRITICAL}일 이내)",
                                description=f"인증서가 {days_until_expiry}일 후 만료 예정",
                                recommendation="즉시 인증서 갱신",
                                details={
                                    "certificate_arn": cert_arn,
                                    "days_until_expiry": days_until_expiry,
                                    "domain": cert_detail.get("DomainName", ""),
                                },
                            )
                        )
                    elif days_until_expiry <= CERT_EXPIRY_HIGH:
                        lb.findings.append(
                            SecurityFinding(
                                category=FindingCategory.CERTIFICATE,
                                risk_level=RiskLevel.HIGH,
                                title=f"인증서 만료 임박 ({CERT_EXPIRY_HIGH}일 이내)",
                                description=f"인증서가 {days_until_expiry}일 후 만료 예정",
                                recommendation="인증서 갱신 계획 수립",
                                details={
                                    "certificate_arn": cert_arn,
                                    "days_until_expiry": days_until_expiry,
                                    "domain": cert_detail.get("DomainName", ""),
                                },
                            )
                        )
                    elif days_until_expiry <= CERT_EXPIRY_MEDIUM:
                        lb.findings.append(
                            SecurityFinding(
                                category=FindingCategory.CERTIFICATE,
                                risk_level=RiskLevel.MEDIUM,
                                title=f"인증서 만료 예정 ({CERT_EXPIRY_MEDIUM}일 이내)",
                                description=f"인증서가 {days_until_expiry}일 후 만료 예정",
                                recommendation="인증서 갱신 준비",
                                details={
                                    "certificate_arn": cert_arn,
                                    "days_until_expiry": days_until_expiry,
                                    "domain": cert_detail.get("DomainName", ""),
                                },
                            )
                        )
            except ClientError:
                pass
    except ClientError:
        pass


def _analyze_waf(lb: LBSecurityInfo) -> None:
    """WAF WebACL 연결 상태 분석

    ALB에 WAF WebACL이 연결되어 있는지 확인합니다.
    인터넷 페이싱 ALB는 HIGH, 내부 ALB는 INFO 수준으로 분류합니다.

    Args:
        lb: 분석 대상 LB 보안 정보 (findings가 in-place 업데이트됨)
    """
    if lb.lb_type != "application":
        return

    if lb.is_internet_facing and not lb.waf_web_acl_arn:
        lb.findings.append(
            SecurityFinding(
                category=FindingCategory.WAF,
                risk_level=RiskLevel.HIGH,
                title="WAF 미연결 (인터넷 페이싱)",
                description="인터넷에 노출된 ALB에 WAF WebACL이 연결되지 않음",
                recommendation="AWS WAF WebACL 연결하여 OWASP Top 10 공격 방어",
                details={
                    "scheme": lb.scheme,
                    "dns_name": lb.dns_name,
                },
            )
        )
    elif not lb.is_internet_facing and not lb.waf_web_acl_arn:
        lb.findings.append(
            SecurityFinding(
                category=FindingCategory.WAF,
                risk_level=RiskLevel.INFO,
                title="WAF 미연결 (내부용)",
                description="내부 ALB에 WAF WebACL 미연결",
                recommendation="내부 트래픽도 WAF 보호 검토",
                details={},
            )
        )


def _analyze_access_logs(lb: LBSecurityInfo) -> None:
    """액세스 로그 활성화 여부 분석

    액세스 로그가 비활성화된 LB를 탐지합니다.
    인터넷 페이싱 LB는 HIGH, 내부 LB는 MEDIUM 수준으로 분류합니다.

    Args:
        lb: 분석 대상 LB 보안 정보 (findings가 in-place 업데이트됨)
    """
    if not lb.access_logs_enabled:
        risk = RiskLevel.HIGH if lb.is_internet_facing else RiskLevel.MEDIUM
        lb.findings.append(
            SecurityFinding(
                category=FindingCategory.ACCESS_LOG,
                risk_level=risk,
                title="액세스 로그 비활성화",
                description="액세스 로그가 활성화되지 않아 감사 추적 불가",
                recommendation="S3 버킷에 액세스 로그 활성화",
                details={
                    "scheme": lb.scheme,
                },
            )
        )


def _analyze_deletion_protection(lb: LBSecurityInfo) -> None:
    """삭제 보호 설정 분석

    ALB/NLB/GWLB의 삭제 보호 활성화 여부를 확인합니다.
    프로덕션 패턴과 일치하는 LB는 HIGH, 그 외는 LOW 수준으로 분류합니다.
    CLB는 삭제 보호를 지원하지 않으므로 스킵합니다.

    Args:
        lb: 분석 대상 LB 보안 정보 (findings가 in-place 업데이트됨)
    """
    if lb.lb_type == "classic":
        return  # CLB는 삭제 보호 없음

    if not lb.deletion_protection:
        # 중앙 설정의 패턴으로 프로덕션 판단
        is_production = any(keyword in lb.name.lower() for keyword in DELETION_PROTECTION_PATTERNS)
        risk = RiskLevel.HIGH if is_production else RiskLevel.LOW

        lb.findings.append(
            SecurityFinding(
                category=FindingCategory.DELETION_PROTECTION,
                risk_level=risk,
                title="삭제 보호 미설정",
                description="실수로 인한 삭제 위험" + (" (프로덕션 추정)" if is_production else ""),
                recommendation="삭제 보호 활성화",
                details={
                    "is_production_suspected": is_production,
                    "matched_patterns": DELETION_PROTECTION_PATTERNS,
                },
            )
        )


def _analyze_listener_security(lb: LBSecurityInfo) -> None:
    """리스너 보안 구성 분석

    인터넷 페이싱 LB에 HTTPS 리스너가 없는 경우(CRITICAL)와
    HTTP->HTTPS 리다이렉트가 미설정된 경우(MEDIUM)를 탐지합니다.

    Args:
        lb: 분석 대상 LB 보안 정보 (findings가 in-place 업데이트됨)
    """
    if not lb.listeners:
        return

    has_http = any(listener.protocol == "HTTP" for listener in lb.listeners)
    has_https = any(listener.protocol in ("HTTPS", "TLS") for listener in lb.listeners)

    # HTTP만 있고 HTTPS 없는 인터넷 페이싱 LB
    if lb.is_internet_facing and has_http and not has_https:
        lb.findings.append(
            SecurityFinding(
                category=FindingCategory.LISTENER,
                risk_level=RiskLevel.CRITICAL,
                title="HTTPS 리스너 없음 (인터넷 페이싱)",
                description="인터넷에 노출된 LB가 HTTPS 없이 HTTP만 사용",
                recommendation="HTTPS 리스너 추가 및 인증서 적용",
                details={
                    "protocols": [listener.protocol for listener in lb.listeners],
                },
            )
        )

    # HTTP와 HTTPS 모두 있지만 리다이렉트 미설정
    if has_http and has_https and lb.lb_type == "application":
        http_listeners = [listener for listener in lb.listeners if listener.protocol == "HTTP"]
        for http_listener in http_listeners:
            has_redirect = any(
                action.get("Type") == "redirect" and action.get("RedirectConfig", {}).get("Protocol") == "HTTPS"
                for action in http_listener.default_actions
            )
            if not has_redirect:
                lb.findings.append(
                    SecurityFinding(
                        category=FindingCategory.LISTENER,
                        risk_level=RiskLevel.MEDIUM,
                        title="HTTP→HTTPS 리다이렉트 미설정",
                        description=f"HTTP :{http_listener.port} 리스너에 HTTPS 리다이렉트 없음",
                        recommendation="HTTP 요청을 HTTPS로 리다이렉트하도록 설정",
                        details={
                            "http_port": http_listener.port,
                        },
                    )
                )


def analyze_all(
    load_balancers: list[LBSecurityInfo],
    session,
    region: str,
    account_id: str,
    account_name: str,
) -> SecurityAuditResult:
    """전체 LB 보안 분석

    수집된 모든 LB에 대해 보안 분석을 수행하고
    위험 수준별 카운트를 집계합니다.

    Args:
        load_balancers: 분석 대상 LB 보안 정보 목록
        session: boto3 Session 객체
        region: AWS 리전
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름

    Returns:
        위험 수준별 카운트가 포함된 보안 감사 결과
    """
    result = SecurityAuditResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(load_balancers),
    )

    for lb in load_balancers:
        analyze_security(lb, session, region)
        result.load_balancers.append(lb)

        # 카운트 집계
        for finding in lb.findings:
            if finding.risk_level == RiskLevel.CRITICAL:
                result.critical_count += 1
            elif finding.risk_level == RiskLevel.HIGH:
                result.high_count += 1
            elif finding.risk_level == RiskLevel.MEDIUM:
                result.medium_count += 1
            elif finding.risk_level == RiskLevel.LOW:
                result.low_count += 1

    return result


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[SecurityAuditResult], output_dir: str) -> str:
    """Excel 보안 감사 보고서 생성

    Summary(계정/리전별 위험 수준 카운트), Findings(위험도순 발견 항목),
    Load Balancers(LB별 보안 속성 현황) 3개 시트로 구성된 보고서를 생성합니다.

    Args:
        results: 계정/리전별 보안 감사 결과 목록
        output_dir: 출력 디렉토리 경로

    Returns:
        저장된 Excel 파일 경로
    """
    from openpyxl.styles import Font, PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 셀 수준 조건부 스타일링용 Fill 및 Font
    risk_fills = {
        RiskLevel.CRITICAL: PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid"),
        RiskLevel.HIGH: PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
        RiskLevel.MEDIUM: PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
        RiskLevel.LOW: PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid"),
        RiskLevel.INFO: PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid"),
    }
    risk_fonts = {
        RiskLevel.CRITICAL: Font(bold=True, color="FFFFFF"),
        RiskLevel.HIGH: Font(bold=True, color="FFFFFF"),
        RiskLevel.MEDIUM: Font(color="000000"),
        RiskLevel.LOW: Font(color="FFFFFF"),
        RiskLevel.INFO: Font(color="FFFFFF"),
    }

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Total LBs", width=12, style="number"),
        ColumnDef(header="CRITICAL", width=12, style="number"),
        ColumnDef(header="HIGH", width=12, style="number"),
        ColumnDef(header="MEDIUM", width=12, style="number"),
        ColumnDef(header="LOW", width=12, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_count,
                r.critical_count,
                r.high_count,
                r.medium_count,
                r.low_count,
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.critical_count > 0:
            ws.cell(row=row_num, column=4).fill = risk_fills[RiskLevel.CRITICAL]
            ws.cell(row=row_num, column=4).font = risk_fonts[RiskLevel.CRITICAL]
        if r.high_count > 0:
            ws.cell(row=row_num, column=5).fill = risk_fills[RiskLevel.HIGH]
            ws.cell(row=row_num, column=5).font = risk_fonts[RiskLevel.HIGH]

    # Findings 시트
    findings_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="LB Name", width=25),
        ColumnDef(header="Type", width=10, style="center"),
        ColumnDef(header="Scheme", width=15),
        ColumnDef(header="Category", width=18),
        ColumnDef(header="Risk", width=12, style="center"),
        ColumnDef(header="Title", width=30),
        ColumnDef(header="Description", width=40),
        ColumnDef(header="Recommendation", width=40),
    ]
    findings_sheet = wb.new_sheet("Findings", findings_columns)

    # 모든 Findings 수집 및 정렬
    all_findings = []
    for result in results:
        for lb in result.load_balancers:
            for finding in lb.findings:
                all_findings.append((lb, finding))

    # 위험도순 정렬 (CRITICAL > HIGH > MEDIUM > LOW > INFO)
    risk_order = {
        RiskLevel.CRITICAL: 0,
        RiskLevel.HIGH: 1,
        RiskLevel.MEDIUM: 2,
        RiskLevel.LOW: 3,
        RiskLevel.INFO: 4,
    }
    all_findings.sort(key=lambda x: risk_order.get(x[1].risk_level, 99))

    for lb, finding in all_findings:
        style = None
        if finding.risk_level == RiskLevel.CRITICAL:
            style = Styles.danger()
        elif finding.risk_level == RiskLevel.HIGH:
            style = Styles.error()
        elif finding.risk_level == RiskLevel.MEDIUM:
            style = Styles.warning()

        row_num = findings_sheet.add_row(
            [
                lb.account_name,
                lb.region,
                lb.name,
                lb.lb_type.upper(),
                lb.scheme,
                finding.category.value,
                finding.risk_level.value.upper(),
                finding.title,
                finding.description,
                finding.recommendation,
            ],
            style=style,
        )
        # Risk 셀에 색상 적용
        ws = findings_sheet._ws
        if finding.risk_level in risk_fills:
            ws.cell(row=row_num, column=7).fill = risk_fills[finding.risk_level]
            ws.cell(row=row_num, column=7).font = risk_fonts[finding.risk_level]

    # Load Balancers 시트
    lb_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Name", width=25),
        ColumnDef(header="Type", width=10, style="center"),
        ColumnDef(header="Scheme", width=15),
        ColumnDef(header="State", width=12),
        ColumnDef(header="Risk Score", width=12, style="number"),
        ColumnDef(header="Findings", width=10, style="number"),
        ColumnDef(header="Access Logs", width=12, style="center"),
        ColumnDef(header="Deletion Protection", width=18, style="center"),
        ColumnDef(header="WAF", width=10, style="center"),
        ColumnDef(header="HTTPS", width=10, style="center"),
    ]
    lb_sheet = wb.new_sheet("Load Balancers", lb_columns)

    for result in results:
        for lb in result.load_balancers:
            lb_sheet.add_row(
                [
                    lb.account_name,
                    lb.region,
                    lb.name,
                    lb.lb_type.upper(),
                    lb.scheme,
                    lb.state,
                    lb.risk_score,
                    len(lb.findings),
                    "Yes" if lb.access_logs_enabled else "No",
                    "Yes" if lb.deletion_protection else "No",
                    "Yes" if lb.waf_web_acl_arn else "No",
                    "Yes" if lb.has_https_listener else "No",
                ]
            )

    return str(wb.save_as(output_dir, "ELB_Security_Audit"))


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SecurityAuditResult | None:
    """단일 계정/리전의 ELB 보안 정보 수집 및 분석 (parallel_collect 콜백)

    ALB/NLB/GWLB와 CLB를 모두 수집하고 보안 분석을 수행합니다.

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        보안 감사 결과. LB가 없으면 None.
    """
    v2_lbs = collect_v2_lb_security(session, account_id, account_name, region)
    classic_lbs = collect_classic_lb_security(session, account_id, account_name, region)
    all_lbs = v2_lbs + classic_lbs
    if not all_lbs:
        return None
    return analyze_all(all_lbs, session, region, account_id, account_name)


def run(ctx: ExecutionContext) -> None:
    """ELB 보안 감사 실행

    멀티 계정/리전에서 모든 유형의 ELB를 병렬 수집하고,
    SSL/TLS, WAF, 액세스 로그, 삭제 보호, 인증서 만료 등
    보안 취약점을 분석하여 Excel 보고서를 생성합니다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 계정/리전 선택, 출력 설정 포함)
    """
    console.print("[bold]ELB Security Audit 시작...[/bold]")
    console.print("[dim]SSL/TLS, WAF, 액세스 로그, 삭제 보호, 인증서 만료 분석[/dim]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="elbv2")
    all_results: list[SecurityAuditResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not all_results:
        console.print("\n[yellow]분석할 ELB 없음[/yellow]")
        return

    # 전체 요약
    totals = {
        "total": sum(r.total_count for r in all_results),
        "critical": sum(r.critical_count for r in all_results),
        "high": sum(r.high_count for r in all_results),
        "medium": sum(r.medium_count for r in all_results),
        "low": sum(r.low_count for r in all_results),
    }

    console.print(f"\n[bold]전체 LB: {totals['total']}개[/bold]")
    if totals["critical"] > 0:
        console.print(f"  [red bold]CRITICAL: {totals['critical']}건[/red bold] - 즉시 조치 필요")
    if totals["high"] > 0:
        console.print(f"  [red]HIGH: {totals['high']}건[/red]")
    if totals["medium"] > 0:
        console.print(f"  [yellow]MEDIUM: {totals['medium']}건[/yellow]")
    if totals["low"] > 0:
        console.print(f"  [blue]LOW: {totals['low']}건[/blue]")

    # 보고서 생성
    console.print("\n[#FF9900]Excel 보고서 생성 중...[/#FF9900]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("elb", "security").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
