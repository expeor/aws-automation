"""
functions/analyzers/ec2/ebs_audit.py - EBS 미사용 분석

미사용 EBS 볼륨 탐지 및 비용 절감 기회 식별

분석 기준:
- Status가 "available"인 볼륨 (아무 인스턴스에도 연결되지 않음)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import TYPE_CHECKING, Any

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.shared.aws.pricing import get_ebs_monthly_cost
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 분석 기간 (일) - AWS Trusted Advisor 기준 7일
ANALYSIS_DAYS = 7

# 미사용 기준: VolumeReadOps + VolumeWriteOps < 1 ops/day (AWS Trusted Advisor 기준)
IDLE_IOPS_THRESHOLD = 1

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeVolumes",
        "cloudwatch:GetMetricStatistics",
    ],
}


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """EBS 볼륨의 사용 상태 분류

    볼륨 연결 상태 및 CloudWatch I/O 지표 기반으로 판별됩니다.
    """

    UNUSED = "unused"  # 미사용 (available 상태)
    IDLE = "idle"  # 유휴 (연결됨 but I/O 없음)
    NORMAL = "normal"  # 정상 사용 (in-use)
    PENDING = "pending"  # 확인 필요


class Severity(Enum):
    """분석 결과의 심각도 수준

    볼륨 용량 기반으로 결정되며, 비용 영향도를 반영합니다.
    """

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class EBSInfo:
    """EBS 볼륨 정보 및 CloudWatch I/O 지표 데이터

    Attributes:
        id: EBS 볼륨 ID
        name: Name 태그 값
        state: 볼륨 상태 (available, in-use, creating 등)
        volume_type: 볼륨 타입 (gp3, gp2, io1, st1, sc1 등)
        size_gb: 볼륨 크기 (GB)
        iops: 프로비저닝된 IOPS
        throughput: 프로비저닝된 처리량 (MiB/s)
        encrypted: 암호화 여부
        kms_key_id: KMS 키 ID
        availability_zone: 가용 영역
        create_time: 볼륨 생성 시간
        snapshot_id: 원본 스냅샷 ID
        attachments: 연결 정보 리스트
        tags: 사용자 태그 딕셔너리 (aws: 접두사 제외)
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        monthly_cost: 추정 월간 비용 (USD)
        read_ops: 분석 기간 총 VolumeReadOps
        write_ops: 분석 기간 총 VolumeWriteOps
    """

    id: str
    name: str
    state: str
    volume_type: str
    size_gb: int
    iops: int
    throughput: int
    encrypted: bool
    kms_key_id: str
    availability_zone: str
    create_time: datetime
    snapshot_id: str
    attachments: list[dict[str, Any]]
    tags: dict[str, str]

    # 메타
    account_id: str
    account_name: str
    region: str

    # 비용
    monthly_cost: float = 0.0

    # CloudWatch 메트릭 (I/O 활동)
    read_ops: float = 0.0  # VolumeReadOps
    write_ops: float = 0.0  # VolumeWriteOps

    @property
    def total_ops(self) -> float:
        """분석 기간 총 I/O 작업 수 (읽기 + 쓰기)

        Returns:
            read_ops + write_ops 합계
        """
        return self.read_ops + self.write_ops

    @property
    def avg_daily_ops(self) -> float:
        """일평균 I/O 작업 수

        Returns:
            분석 기간 일평균 총 I/O 작업 수
        """
        return self.total_ops / ANALYSIS_DAYS if ANALYSIS_DAYS > 0 else 0

    @property
    def is_attached(self) -> bool:
        """EC2 인스턴스 연결 여부

        Returns:
            in-use 상태이고 attachment가 존재하면 True
        """
        return self.state == "in-use" and len(self.attachments) > 0

    @property
    def attached_instance_id(self) -> str:
        """연결된 EC2 인스턴스 ID

        Returns:
            첫 번째 attachment의 InstanceId. 연결 없으면 빈 문자열.
        """
        if self.attachments:
            return str(self.attachments[0].get("InstanceId", ""))
        return ""


@dataclass
class EBSFinding:
    """개별 EBS 볼륨 분석 결과

    Attributes:
        volume: 분석 대상 EBS 볼륨 정보
        usage_status: 분석으로 판별된 사용 상태
        severity: 심각도 수준
        description: 상태 설명
        recommendation: 권장 조치 사항
    """

    volume: EBSInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class EBSAnalysisResult:
    """EBS 볼륨 분석 결과 집계 (계정/리전별)

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        findings: 개별 볼륨 분석 결과 리스트
        total_count: 전체 볼륨 수
        unused_count: 미사용 볼륨 수 (available 상태)
        idle_count: 유휴 볼륨 수 (연결됨, I/O 없음)
        normal_count: 정상 사용 볼륨 수
        pending_count: 확인 필요 볼륨 수
        total_size_gb: 전체 볼륨 용량 (GB)
        unused_size_gb: 미사용 볼륨 용량 (GB)
        idle_size_gb: 유휴 볼륨 용량 (GB)
        unused_monthly_cost: 미사용 볼륨 추정 월간 비용 (USD)
        idle_monthly_cost: 유휴 볼륨 추정 월간 비용 (USD)
    """

    account_id: str
    account_name: str
    region: str
    findings: list[EBSFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    idle_count: int = 0
    normal_count: int = 0
    pending_count: int = 0

    # 비용
    total_size_gb: int = 0
    unused_size_gb: int = 0
    idle_size_gb: int = 0
    unused_monthly_cost: float = 0.0
    idle_monthly_cost: float = 0.0


# =============================================================================
# 수집
# =============================================================================


def collect_ebs(session, account_id: str, account_name: str, region: str) -> list[EBSInfo]:
    """EBS 볼륨 목록 수집 및 CloudWatch I/O 메트릭 조회

    모든 EBS 볼륨을 페이지네이션으로 수집하고, in-use 볼륨에 대해
    CloudWatch GetMetricStatistics로 VolumeReadOps/VolumeWriteOps를 조회합니다.

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        EBS 볼륨 정보 리스트 (I/O 메트릭 포함)
    """
    from datetime import timedelta, timezone

    from botocore.exceptions import ClientError

    volumes = []
    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        ec2 = get_client(session, "ec2", region_name=region)
        cloudwatch = get_client(session, "cloudwatch", region_name=region)
        paginator = ec2.get_paginator("describe_volumes")

        for page in paginator.paginate():
            for data in page.get("Volumes", []):
                # 태그 파싱
                tags = {
                    t.get("Key", ""): t.get("Value", "")
                    for t in data.get("Tags", [])
                    if not t.get("Key", "").startswith("aws:")
                }

                # 월간 비용 계산
                volume_type = data.get("VolumeType", "")
                size_gb = data.get("Size", 0)
                monthly_cost = get_ebs_monthly_cost(volume_type, size_gb, region)

                volume = EBSInfo(
                    id=data.get("VolumeId", ""),
                    name=tags.get("Name", ""),
                    state=data.get("State", ""),
                    volume_type=volume_type,
                    size_gb=size_gb,
                    iops=data.get("Iops", 0),
                    throughput=data.get("Throughput", 0),
                    encrypted=data.get("Encrypted", False),
                    kms_key_id=data.get("KmsKeyId", ""),
                    availability_zone=data.get("AvailabilityZone", ""),
                    create_time=data.get("CreateTime"),
                    snapshot_id=data.get("SnapshotId", ""),
                    attachments=data.get("Attachments", []),
                    tags=tags,
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    monthly_cost=monthly_cost,
                )

                # in-use 볼륨에 대해 CloudWatch 메트릭 수집 (AWS Trusted Advisor 기준)
                if volume.state == "in-use":
                    try:
                        # VolumeReadOps
                        read_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/EBS",
                            MetricName="VolumeReadOps",
                            Dimensions=[{"Name": "VolumeId", "Value": volume.id}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if read_resp.get("Datapoints"):
                            volume.read_ops = sum(d["Sum"] for d in read_resp["Datapoints"])

                        # VolumeWriteOps
                        write_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/EBS",
                            MetricName="VolumeWriteOps",
                            Dimensions=[{"Name": "VolumeId", "Value": volume.id}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if write_resp.get("Datapoints"):
                            volume.write_ops = sum(d["Sum"] for d in write_resp["Datapoints"])
                    except ClientError:
                        pass

                volumes.append(volume)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} EBS 수집 오류: {error_code}[/yellow]")

    return volumes


# =============================================================================
# 분석
# =============================================================================


def analyze_ebs(volumes: list[EBSInfo], account_id: str, account_name: str, region: str) -> EBSAnalysisResult:
    """EBS 볼륨 사용 상태 분석

    각 볼륨을 연결 상태 및 I/O 지표 기반으로 미사용/유휴/정상/확인필요로
    분류하고, 상태별 용량과 비용을 집계합니다.

    Args:
        volumes: 수집된 EBS 볼륨 정보 리스트
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        계정/리전별 분석 결과 (상태별 볼륨 수, 용량, 비용 포함)
    """
    result = EBSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for volume in volumes:
        finding = _analyze_single_volume(volume)
        result.findings.append(finding)
        result.total_size_gb += volume.size_gb

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
            result.unused_size_gb += volume.size_gb
            result.unused_monthly_cost += volume.monthly_cost
        elif finding.usage_status == UsageStatus.IDLE:
            result.idle_count += 1
            result.idle_size_gb += volume.size_gb
            result.idle_monthly_cost += volume.monthly_cost
        elif finding.usage_status == UsageStatus.NORMAL:
            result.normal_count += 1
        elif finding.usage_status == UsageStatus.PENDING:
            result.pending_count += 1

    result.total_count = len(volumes)
    return result


def _analyze_single_volume(volume: EBSInfo) -> EBSFinding:
    """개별 EBS 볼륨 사용 상태 분석

    연결 상태, I/O 활동, 용량을 기반으로 사용 상태와 심각도를 결정합니다.

    Args:
        volume: 분석 대상 EBS 볼륨 정보

    Returns:
        개별 볼륨 분석 결과 (사용 상태, 심각도, 권장 조치 포함)
    """

    # 1. 연결됨
    if volume.is_attached:
        # 연결되었지만 I/O 없음 = 유휴 (AWS Trusted Advisor 기준)
        if volume.avg_daily_ops < IDLE_IOPS_THRESHOLD:
            # 용량에 따른 심각도
            if volume.size_gb >= 500:
                severity = Severity.HIGH
            elif volume.size_gb >= 100:
                severity = Severity.MEDIUM
            else:
                severity = Severity.LOW

            cost_info = f"월 ${volume.monthly_cost:.2f}" if volume.monthly_cost > 0 else ""
            return EBSFinding(
                volume=volume,
                usage_status=UsageStatus.IDLE,
                severity=severity,
                description=f"유휴 볼륨 ({volume.size_gb}GB, 일평균 {volume.avg_daily_ops:.1f} ops) {cost_info}",
                recommendation=f"인스턴스 {volume.attached_instance_id}에 연결되었으나 {ANALYSIS_DAYS}일간 I/O 없음 - 사용 여부 확인",
            )

        # 정상 사용
        return EBSFinding(
            volume=volume,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"사용 중 (인스턴스: {volume.attached_instance_id})",
            recommendation="정상 사용 중",
        )

    # 2. Available = 미사용
    if volume.state == "available":
        # 용량에 따른 심각도
        if volume.size_gb >= 500:
            severity = Severity.HIGH
        elif volume.size_gb >= 100:
            severity = Severity.MEDIUM
        else:
            severity = Severity.LOW

        cost_info = f"월 ${volume.monthly_cost:.2f}" if volume.monthly_cost > 0 else ""
        return EBSFinding(
            volume=volume,
            usage_status=UsageStatus.UNUSED,
            severity=severity,
            description=f"미사용 볼륨 ({volume.size_gb}GB, {volume.volume_type}) {cost_info}",
            recommendation="스냅샷 생성 후 삭제 검토",
        )

    # 3. 기타 (creating, deleting, error 등)
    return EBSFinding(
        volume=volume,
        usage_status=UsageStatus.PENDING,
        severity=Severity.INFO,
        description=f"상태: {volume.state}",
        recommendation="상태 안정화 대기",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[EBSAnalysisResult], output_dir: str) -> str:
    """EBS 미사용 분석 Excel 보고서 생성

    Summary 시트(통계 요약)와 Findings 시트(개별 볼륨 상세)를 포함하는
    Excel 파일을 생성합니다.

    Args:
        results: 분석 결과 리스트
        output_dir: 출력 디렉토리 경로

    Returns:
        생성된 Excel 파일 경로
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # Summary sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("EBS 미사용 분석 보고서")
    summary.add_item("생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()

    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "idle": sum(r.idle_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "pending": sum(r.pending_count for r in results),
        "total_size": sum(r.total_size_gb for r in results),
        "unused_size": sum(r.unused_size_gb for r in results),
        "idle_size": sum(r.idle_size_gb for r in results),
        "unused_cost": sum(r.unused_monthly_cost for r in results),
        "idle_cost": sum(r.idle_monthly_cost for r in results),
    }

    summary.add_section("통계")
    summary.add_item("전체 볼륨", totals["total"])
    summary.add_item("미사용 (연결안됨)", totals["unused"], highlight="danger" if totals["unused"] > 0 else None)
    summary.add_item("유휴 (I/O 없음)", totals["idle"], highlight="warning" if totals["idle"] > 0 else None)
    summary.add_item("정상 사용", totals["normal"])
    summary.add_item("확인 필요", totals["pending"], highlight="warning" if totals["pending"] > 0 else None)
    summary.add_item("전체 용량 (GB)", totals["total_size"])
    summary.add_item("미사용 용량 (GB)", totals["unused_size"])
    summary.add_item("유휴 용량 (GB)", totals["idle_size"])
    summary.add_item("미사용 월 비용 ($)", f"${totals['unused_cost']:.2f}")
    summary.add_item("유휴 월 비용 ($)", f"${totals['idle_cost']:.2f}")

    # Findings sheet
    columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Volume ID", width=22),
        ColumnDef(header="Name", width=25),
        ColumnDef(header="State", width=12),
        ColumnDef(header="Usage", width=12),
        ColumnDef(header="Severity", width=10),
        ColumnDef(header="Type", width=10),
        ColumnDef(header="Size (GB)", width=12, style="number"),
        ColumnDef(header="Monthly Cost ($)", width=15, style="number"),
        ColumnDef(header="IOPS", width=10, style="number"),
        ColumnDef(header="Encrypted", width=10),
        ColumnDef(header="AZ", width=18),
        ColumnDef(header="Created", width=12),
        ColumnDef(header="Description", width=40),
        ColumnDef(header="Recommendation", width=30),
    ]
    sheet = wb.new_sheet("Findings", columns)

    # 상태별 스타일
    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.IDLE: PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        UsageStatus.PENDING: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # 미사용/유휴/확인필요 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.UNUSED, UsageStatus.IDLE, UsageStatus.PENDING):
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.volume.monthly_cost, reverse=True)

    for f in all_findings:
        vol = f.volume
        if f.usage_status == UsageStatus.UNUSED:
            style = Styles.danger()
        elif f.usage_status == UsageStatus.IDLE:
            style = Styles.warning()
        else:
            style = None
        row_num = sheet.add_row(
            [
                vol.account_name,
                vol.region,
                vol.id,
                vol.name,
                vol.state,
                f.usage_status.value,
                f.severity.value,
                vol.volume_type,
                vol.size_gb,
                round(vol.monthly_cost, 2),
                vol.iops,
                "Yes" if vol.encrypted else "No",
                vol.availability_zone,
                vol.create_time.strftime("%Y-%m-%d") if vol.create_time else "",
                f.description,
                f.recommendation,
            ],
            style=style,
        )

        # Usage 컬럼에 상태별 색상 적용
        fill = status_fills.get(f.usage_status)
        if fill:
            sheet._ws.cell(row=row_num, column=6).fill = fill

    return str(wb.save_as(output_dir, "EBS_Unused"))


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EBSAnalysisResult:
    """단일 계정/리전의 EBS 볼륨 수집 및 분석 (parallel_collect 콜백)

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        계정/리전별 EBS 분석 결과
    """
    volumes = collect_ebs(session, account_id, account_name, region)
    return analyze_ebs(volumes, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """EBS 미사용 분석 실행

    멀티 계정/리전에서 EBS 볼륨을 병렬 수집하고,
    미사용(available)/유휴(I/O 없음) 상태를 분석하여
    Excel 보고서를 생성합니다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 계정/리전 선택, 출력 설정 포함)
    """
    console.print("[bold]EBS 미사용 분석 시작...[/bold]")

    # 병렬 수집
    result = parallel_collect(
        ctx,
        _collect_and_analyze,
        max_workers=20,
        service="ec2",
    )

    # 결과 처리
    all_results: list[EBSAnalysisResult] = result.get_data()

    # 진행 상황 출력
    console.print(f"  [dim]수집 완료: 성공 {result.success_count}, 실패 {result.error_count}[/dim]")

    # 에러 요약
    if result.error_count > 0:
        console.print(f"\n[yellow]{result.get_error_summary()}[/yellow]")

    if not all_results:
        console.print("[yellow]분석할 EBS 없음[/yellow]")
        return

    # 개별 결과 요약
    for r in all_results:
        if r.unused_count > 0 or r.idle_count > 0:
            unused_str = f"미사용 {r.unused_count}개" if r.unused_count > 0 else ""
            idle_str = f"유휴 {r.idle_count}개" if r.idle_count > 0 else ""
            total_cost = r.unused_monthly_cost + r.idle_monthly_cost
            cost_str = f" (${total_cost:.2f}/월)" if total_cost > 0 else ""
            parts = [p for p in [unused_str, idle_str] if p]
            console.print(f"  {r.account_name}/{r.region}: [red]{' / '.join(parts)}{cost_str}[/red]")
        elif r.pending_count > 0:
            console.print(f"  {r.account_name}/{r.region}: [yellow]확인 필요 {r.pending_count}개[/yellow]")
        elif r.total_count > 0:
            console.print(f"  {r.account_name}/{r.region}: [green]정상 {r.normal_count}개[/green]")

    # 전체 통계
    totals = {
        "total": sum(r.total_count for r in all_results),
        "unused": sum(r.unused_count for r in all_results),
        "idle": sum(r.idle_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "pending": sum(r.pending_count for r in all_results),
        "total_size": sum(r.total_size_gb for r in all_results),
        "unused_size": sum(r.unused_size_gb for r in all_results),
        "idle_size": sum(r.idle_size_gb for r in all_results),
        "unused_cost": sum(r.unused_monthly_cost for r in all_results),
        "idle_cost": sum(r.idle_monthly_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 EBS: {totals['total']}개 ({totals['total_size']}GB)[/bold]")
    if totals["unused"] > 0:
        console.print(
            f"  [red bold]미사용 (연결안됨): {totals['unused']}개 ({totals['unused_size']}GB, ${totals['unused_cost']:.2f}/월)[/red bold]"
        )
    if totals["idle"] > 0:
        console.print(
            f"  [yellow bold]유휴 (I/O없음): {totals['idle']}개 ({totals['idle_size']}GB, ${totals['idle_cost']:.2f}/월)[/yellow bold]"
        )
    if totals["pending"] > 0:
        console.print(f"  [yellow]확인 필요: {totals['pending']}개[/yellow]")
    console.print(f"  [green]정상: {totals['normal']}개[/green]")

    # 보고서
    console.print("\n[#FF9900]Excel 보고서 생성 중...[/#FF9900]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("ebs", "inventory").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
