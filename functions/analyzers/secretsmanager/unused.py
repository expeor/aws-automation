"""
functions/analyzers/secretsmanager/unused.py - Secrets Manager 미사용 분석

미사용 시크릿 탐지 및 비용 분석

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.aws.pricing import get_secret_price
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "secretsmanager:ListSecrets",
    ],
}

# 미사용 기준: 90일 이상 액세스 없음
UNUSED_DAYS_THRESHOLD = 90


class SecretStatus(Enum):
    """Secrets Manager 시크릿 분석 상태.

    액세스 이력 및 삭제 예정 여부 기반으로 분류한다.

    Attributes:
        NORMAL: 정상 사용 중인 시크릿.
        UNUSED: 기준 기간(90일) 동안 액세스가 없는 미사용 시크릿.
        PENDING_DELETE: 삭제 예정인 시크릿.
    """

    NORMAL = "normal"
    UNUSED = "unused"
    PENDING_DELETE = "pending_delete"


@dataclass
class SecretInfo:
    """Secrets Manager 시크릿 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 시크릿이 위치한 리전.
        arn: 시크릿 ARN.
        name: 시크릿 이름.
        description: 시크릿 설명.
        created_date: 시크릿 생성일.
        last_accessed_date: 마지막 액세스일.
        last_changed_date: 마지막 변경일.
        rotation_enabled: 자동 교체 활성화 여부.
        deleted_date: 삭제 예정일 (삭제 예약된 경우).
    """

    account_id: str
    account_name: str
    region: str
    arn: str
    name: str
    description: str
    created_date: datetime | None
    last_accessed_date: datetime | None
    last_changed_date: datetime | None
    rotation_enabled: bool
    deleted_date: datetime | None = None

    @property
    def monthly_cost(self) -> float:
        """시크릿 월간 비용 (USD).

        Returns:
            리전별 시크릿 호스팅 가격.
        """
        return get_secret_price(self.region)

    @property
    def days_since_access(self) -> int | None:
        """마지막 액세스 이후 경과 일수.

        Returns:
            경과 일수. 액세스 기록이 없으면 None.
        """
        if self.last_accessed_date:
            return (datetime.now(timezone.utc) - self.last_accessed_date).days
        return None


@dataclass
class SecretFinding:
    """개별 시크릿에 대한 분석 결과.

    Attributes:
        secret: 분석 대상 시크릿 정보.
        status: 분석 결과 상태.
        recommendation: 권장 조치 사항 (한글).
    """

    secret: SecretInfo
    status: SecretStatus
    recommendation: str


@dataclass
class SecretAnalysisResult:
    """단일 계정/리전의 시크릿 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.
        total_count: 전체 시크릿 수.
        unused_count: 미사용 시크릿 수.
        pending_delete_count: 삭제 예정 시크릿 수.
        normal_count: 정상 시크릿 수.
        total_monthly_cost: 전체 시크릿 월간 비용 합계 (USD).
        unused_monthly_cost: 미사용 시크릿 월간 비용 합계 (USD).
        findings: 개별 시크릿 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    unused_count: int = 0
    pending_delete_count: int = 0
    normal_count: int = 0
    total_monthly_cost: float = 0.0
    unused_monthly_cost: float = 0.0
    findings: list[SecretFinding] = field(default_factory=list)


def collect_secrets(session, account_id: str, account_name: str, region: str) -> list[SecretInfo]:
    """지정된 계정/리전의 시크릿을 수집한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        수집된 시크릿 정보 목록.
    """
    sm = get_client(session, "secretsmanager", region_name=region)
    secrets = []

    paginator = sm.get_paginator("list_secrets")
    for page in paginator.paginate():
        for secret in page.get("SecretList", []):
            secrets.append(
                SecretInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    arn=secret.get("ARN", ""),
                    name=secret.get("Name", ""),
                    description=secret.get("Description", ""),
                    created_date=secret.get("CreatedDate"),
                    last_accessed_date=secret.get("LastAccessedDate"),
                    last_changed_date=secret.get("LastChangedDate"),
                    rotation_enabled=secret.get("RotationEnabled", False),
                    deleted_date=secret.get("DeletedDate"),
                )
            )

    return secrets


def analyze_secrets(secrets: list[SecretInfo], account_id: str, account_name: str, region: str) -> SecretAnalysisResult:
    """수집된 시크릿을 분석하여 미사용/삭제 예정 시크릿을 식별한다.

    90일 이상 액세스가 없거나, 생성 후 한 번도 액세스되지 않은 시크릿을 미사용으로 분류한다.

    Args:
        secrets: 분석 대상 시크릿 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.

    Returns:
        시크릿 분석 결과 집계 객체.
    """
    result = SecretAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(secrets),
    )

    for secret in secrets:
        result.total_monthly_cost += secret.monthly_cost

        if secret.deleted_date:
            result.pending_delete_count += 1
            result.findings.append(
                SecretFinding(
                    secret=secret,
                    status=SecretStatus.PENDING_DELETE,
                    recommendation="삭제 예정됨",
                )
            )
            continue

        if secret.days_since_access and secret.days_since_access > UNUSED_DAYS_THRESHOLD:
            result.unused_count += 1
            result.unused_monthly_cost += secret.monthly_cost
            result.findings.append(
                SecretFinding(
                    secret=secret,
                    status=SecretStatus.UNUSED,
                    recommendation=f"{secret.days_since_access}일간 액세스 없음 - 삭제 검토",
                )
            )
            continue

        if secret.last_accessed_date is None and secret.created_date:
            age_days = (datetime.now(timezone.utc) - secret.created_date).days
            if age_days > UNUSED_DAYS_THRESHOLD:
                result.unused_count += 1
                result.unused_monthly_cost += secret.monthly_cost
                result.findings.append(
                    SecretFinding(
                        secret=secret,
                        status=SecretStatus.UNUSED,
                        recommendation=f"생성 후 {age_days}일간 액세스 없음",
                    )
                )
                continue

        result.normal_count += 1
        result.findings.append(
            SecretFinding(
                secret=secret,
                status=SecretStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[SecretAnalysisResult], output_dir: str) -> str:
    """Secrets Manager 분석 결과를 Excel 보고서로 생성한다.

    Summary 시트(계정/리전별 통계)와 Detail 시트(비정상 시크릿 상세)를 포함한다.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    wb = Workbook()

    # Summary Sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("Secrets Manager 분석 보고서")

    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="삭제예정", width=10, style="number"),
        ColumnDef(header="총 비용", width=15),
        ColumnDef(header="낭비 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_count,
                r.unused_count,
                r.pending_delete_count,
                f"${r.total_monthly_cost:,.2f}",
                f"${r.unused_monthly_cost:,.2f}",
            ]
        )
        if r.unused_count > 0:
            summary_sheet._ws.cell(row=row_num, column=4).fill = red_fill

    # Detail Sheet
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Name", width=40),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="마지막 액세스", width=15),
        ColumnDef(header="Rotation", width=10),
        ColumnDef(header="월간 비용", width=15),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Secrets", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != SecretStatus.NORMAL:
                s = f.secret
                detail_sheet.add_row(
                    [
                        s.account_name,
                        s.region,
                        s.name,
                        f.status.value,
                        s.last_accessed_date.strftime("%Y-%m-%d") if s.last_accessed_date else "없음",
                        "예" if s.rotation_enabled else "아니오",
                        f"${s.monthly_cost:,.2f}",
                        f.recommendation,
                    ]
                )

    return str(wb.save_as(output_dir, "Secrets_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SecretAnalysisResult | None:
    """단일 계정/리전의 시크릿을 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 시크릿이 없으면 None을 반환한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 대상 리전.

    Returns:
        분석 결과 객체. 시크릿이 없으면 None.
    """
    secrets = collect_secrets(session, account_id, account_name, region)
    if not secrets:
        return None
    return analyze_secrets(secrets, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Secrets Manager 미사용 시크릿 분석 도구의 메인 실행 함수.

    멀티 계정/리전 병렬 수집 후 결과를 집계하고, Excel 보고서를 생성하여 출력 디렉토리에 저장한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 계정/리전 목록, 옵션 등 포함).
    """
    console.print("[bold]Secrets Manager 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="secretsmanager")
    results: list[SecretAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total = sum(r.total_count for r in results)
    unused = sum(r.unused_count for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체: {total}개 / 미사용: [yellow]{unused}개[/yellow] (${unused_cost:,.2f}/월)")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("secretsmanager", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
