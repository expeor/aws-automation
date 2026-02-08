"""
functions/analyzers/sagemaker/unused.py - SageMaker Endpoint 미사용 분석

유휴/미사용 SageMaker Endpoint 탐지 (CloudWatch 지표 기반)

탐지 기준:
- 미사용: 7일간 Invocations 0회

비용 영향:
- SageMaker Endpoint는 유지만 해도 인스턴스 비용 발생
- 미사용 Endpoint는 월 $100-$5,000+ 낭비 가능

최적화:
- CloudWatch GetMetricData API 사용 (배치 조회)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer
from functions.analyzers.cost.pricing import get_sagemaker_monthly_cost

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 분석 기간 (일) - 고비용 리소스이므로 짧은 기간으로 빠르게 탐지
ANALYSIS_DAYS = 7

# 미사용 기준: Invocations 0회
UNUSED_INVOCATION_THRESHOLD = 0

# 저사용 기준: 하루 평균 10회 미만
LOW_USAGE_INVOCATIONS_PER_DAY = 10

# 유휴 기준: CPU 5% 미만 (AWS Cost Optimization Pillar)
IDLE_CPU_THRESHOLD = 5.0

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "sagemaker:ListEndpoints",
        "sagemaker:DescribeEndpoint",
        "cloudwatch:GetMetricData",
    ],
}


class EndpointStatus(Enum):
    """SageMaker Endpoint 사용 상태 분류.

    CloudWatch 지표 기반으로 Endpoint의 활용도를 분류한다.

    Attributes:
        NORMAL: 정상적으로 호출되고 리소스가 활용되는 Endpoint.
        UNUSED: 7일간 Invocations가 0인 미사용 Endpoint.
        IDLE: 호출은 있으나 CPU 사용률이 5% 미만인 유휴 Endpoint.
        LOW_USAGE: 일평균 호출 10회 미만인 저사용 Endpoint.
    """

    NORMAL = "normal"
    UNUSED = "unused"
    IDLE = "idle"  # 호출은 있으나 CPU 유휴
    LOW_USAGE = "low_usage"


@dataclass
class SageMakerEndpointInfo:
    """SageMaker Endpoint 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 리전.
        endpoint_name: Endpoint 이름.
        endpoint_arn: Endpoint ARN.
        status: Endpoint 상태 (InService, Creating, Updating, Failed 등).
        creation_time: Endpoint 생성 시각.
        instance_type: 인스턴스 유형 (ml.m5.large 등).
        instance_count: 인스턴스 수.
        variant_name: Production Variant 이름.
        total_invocations: 분석 기간 중 총 호출 수 (CloudWatch Invocations Sum).
        avg_invocations_per_day: 일평균 호출 수.
        model_latency_avg_ms: 평균 모델 레이턴시 (밀리초).
        cpu_utilization_avg: 평균 CPU 사용률 (%).
        memory_utilization_avg: 평균 메모리 사용률 (%).
    """

    account_id: str
    account_name: str
    region: str
    endpoint_name: str
    endpoint_arn: str
    status: str  # InService, Creating, Updating, Failed, etc.
    creation_time: datetime | None
    instance_type: str
    instance_count: int
    variant_name: str

    # CloudWatch 지표
    total_invocations: int = 0
    avg_invocations_per_day: float = 0.0
    model_latency_avg_ms: float = 0.0
    cpu_utilization_avg: float = 0.0  # CPUUtilization (%)
    memory_utilization_avg: float = 0.0  # MemoryUtilization (%)

    @property
    def estimated_monthly_cost(self) -> float:
        """AWS Pricing API 기반 추정 월간 비용 (USD).

        Returns:
            추정 월간 비용 (USD).
        """
        return get_sagemaker_monthly_cost(self.instance_type, self.region, self.instance_count)

    @property
    def age_days(self) -> int:
        """Endpoint 생성 후 경과 일수.

        Returns:
            경과 일수. creation_time이 없으면 0.
        """
        if not self.creation_time:
            return 0
        now = datetime.now(timezone.utc)
        return (now - self.creation_time.replace(tzinfo=timezone.utc)).days


@dataclass
class EndpointFinding:
    """개별 SageMaker Endpoint에 대한 분석 결과.

    Attributes:
        endpoint: 분석 대상 Endpoint 정보.
        status: 분석된 Endpoint 상태 (NORMAL, UNUSED, IDLE, LOW_USAGE).
        recommendation: 권장 조치 메시지.
    """

    endpoint: SageMakerEndpointInfo
    status: EndpointStatus
    recommendation: str


@dataclass
class SageMakerAnalysisResult:
    """계정/리전별 SageMaker Endpoint 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 리전.
        total_endpoints: 전체 Endpoint 수.
        unused_endpoints: 미사용 Endpoint 수 (호출 0회).
        idle_endpoints: 유휴 Endpoint 수 (CPU < 5%).
        low_usage_endpoints: 저사용 Endpoint 수 (일평균 호출 < 10회).
        normal_endpoints: 정상 Endpoint 수.
        unused_monthly_cost: 미사용 Endpoint 합산 월간 추정 비용 (USD).
        idle_monthly_cost: 유휴 Endpoint 합산 월간 추정 비용 (USD).
        low_usage_monthly_cost: 저사용 Endpoint 합산 월간 추정 비용 (USD).
        findings: 개별 Endpoint 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_endpoints: int = 0
    unused_endpoints: int = 0
    idle_endpoints: int = 0
    low_usage_endpoints: int = 0
    normal_endpoints: int = 0
    unused_monthly_cost: float = 0.0
    idle_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[EndpointFinding] = field(default_factory=list)


def collect_sagemaker_endpoints(
    session, account_id: str, account_name: str, region: str
) -> list[SageMakerEndpointInfo]:
    """SageMaker InService Endpoint를 수집하고 CloudWatch 지표를 배치 조회한다.

    Endpoint 목록 조회 후 describe_endpoint로 상세 정보(인스턴스 유형, 수)를
    수집하고 배치 메트릭으로 Invocations, Latency, CPU, Memory 지표를 수집한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        수집된 SageMakerEndpointInfo 목록.
    """
    from botocore.exceptions import ClientError

    sagemaker = get_client(session, "sagemaker", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    endpoints = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: Endpoint 목록 수집
        paginator = sagemaker.get_paginator("list_endpoints")
        for page in paginator.paginate(StatusEquals="InService"):
            for ep in page.get("Endpoints", []):
                endpoint_name = ep.get("EndpointName", "")

                # Endpoint 상세 정보 조회
                try:
                    detail = sagemaker.describe_endpoint(EndpointName=endpoint_name)
                    production_variants = detail.get("ProductionVariants", [])

                    # 첫 번째 variant 정보 사용
                    if production_variants:
                        variant = production_variants[0]
                        instance_type = variant.get("CurrentInstanceType", variant.get("InstanceType", "unknown"))
                        instance_count = variant.get("CurrentWeight", variant.get("DesiredInstanceCount", 1))
                        if isinstance(instance_count, float):
                            instance_count = int(instance_count)
                        variant_name = variant.get("VariantName", "")
                    else:
                        instance_type = "unknown"
                        instance_count = 1
                        variant_name = ""

                    endpoint = SageMakerEndpointInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        endpoint_name=endpoint_name,
                        endpoint_arn=detail.get("EndpointArn", ""),
                        status=detail.get("EndpointStatus", ""),
                        creation_time=detail.get("CreationTime"),
                        instance_type=instance_type,
                        instance_count=instance_count,
                        variant_name=variant_name,
                    )
                    endpoints.append(endpoint)

                except ClientError:
                    # 상세 조회 실패 시 기본 정보로 추가
                    endpoint = SageMakerEndpointInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        endpoint_name=endpoint_name,
                        endpoint_arn=ep.get("EndpointArn", ""),
                        status=ep.get("EndpointStatus", ""),
                        creation_time=ep.get("CreationTime"),
                        instance_type="unknown",
                        instance_count=1,
                        variant_name="",
                    )
                    endpoints.append(endpoint)

        # 2단계: 메트릭 배치 조회
        if endpoints:
            _collect_sagemaker_metrics_batch(cloudwatch, endpoints, start_time, now)

    except ClientError:
        pass

    return endpoints


def _collect_sagemaker_metrics_batch(
    cloudwatch,
    endpoints: list[SageMakerEndpointInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """CloudWatch GetMetricData API로 SageMaker Endpoint 지표를 배치 수집한다.

    Invocations(Sum), ModelLatency(Average), CPUUtilization(Average),
    MemoryUtilization(Average) 네 가지 지표를 배치 조회하여 각 Endpoint에 설정한다.

    Args:
        cloudwatch: CloudWatch boto3 클라이언트.
        endpoints: 지표를 수집할 SageMakerEndpointInfo 목록.
        start_time: 지표 조회 시작 시각.
        end_time: 지표 조회 종료 시각.
    """
    from botocore.exceptions import ClientError

    # 쿼리 생성
    queries = []
    for endpoint in endpoints:
        safe_id = sanitize_metric_id(endpoint.endpoint_name)

        # Invocations (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_invocations",
                namespace="AWS/SageMaker",
                metric_name="Invocations",
                dimensions={
                    "EndpointName": endpoint.endpoint_name,
                    "VariantName": endpoint.variant_name if endpoint.variant_name else "AllTraffic",
                },
                stat="Sum",
            )
        )

        # ModelLatency (Average) - 밀리초
        queries.append(
            MetricQuery(
                id=f"{safe_id}_latency",
                namespace="AWS/SageMaker",
                metric_name="ModelLatency",
                dimensions={
                    "EndpointName": endpoint.endpoint_name,
                    "VariantName": endpoint.variant_name if endpoint.variant_name else "AllTraffic",
                },
                stat="Average",
            )
        )

        # CPUUtilization (Average) - AWS Cost Optimization Pillar 기준
        queries.append(
            MetricQuery(
                id=f"{safe_id}_cpu",
                namespace="/aws/sagemaker/Endpoints",
                metric_name="CPUUtilization",
                dimensions={
                    "EndpointName": endpoint.endpoint_name,
                    "VariantName": endpoint.variant_name if endpoint.variant_name else "AllTraffic",
                },
                stat="Average",
            )
        )

        # MemoryUtilization (Average)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_memory",
                namespace="/aws/sagemaker/Endpoints",
                metric_name="MemoryUtilization",
                dimensions={
                    "EndpointName": endpoint.endpoint_name,
                    "VariantName": endpoint.variant_name if endpoint.variant_name else "AllTraffic",
                },
                stat="Average",
            )
        )

    try:
        # 배치 조회
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        # 결과 매핑
        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for endpoint in endpoints:
            safe_id = sanitize_metric_id(endpoint.endpoint_name)

            endpoint.total_invocations = int(results.get(f"{safe_id}_invocations", 0))
            endpoint.avg_invocations_per_day = endpoint.total_invocations / days
            # 레이턴시는 마이크로초로 반환되므로 밀리초로 변환
            endpoint.model_latency_avg_ms = results.get(f"{safe_id}_latency", 0.0) / 1000
            # CPU/Memory Utilization
            endpoint.cpu_utilization_avg = results.get(f"{safe_id}_cpu", 0.0)
            endpoint.memory_utilization_avg = results.get(f"{safe_id}_memory", 0.0)

    except ClientError:
        # 실패 시 무시 (기본값 0 유지)
        pass


def analyze_endpoints(
    endpoints: list[SageMakerEndpointInfo], account_id: str, account_name: str, region: str
) -> SageMakerAnalysisResult:
    """수집된 SageMaker Endpoint를 분석하여 미사용/유휴/저사용을 판별한다.

    미사용(Invocations 0), 유휴(CPU < 5%), 저사용(일평균 호출 < 10회),
    정상으로 분류하고 각 카테고리별 비용을 합산한다.

    Args:
        endpoints: 수집된 SageMakerEndpointInfo 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 리전.

    Returns:
        분석 결과를 담은 SageMakerAnalysisResult 객체.
    """
    result = SageMakerAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_endpoints=len(endpoints),
    )

    for endpoint in endpoints:
        # 미사용: Invocations 0회
        if endpoint.total_invocations == UNUSED_INVOCATION_THRESHOLD:
            result.unused_endpoints += 1
            result.unused_monthly_cost += endpoint.estimated_monthly_cost
            result.findings.append(
                EndpointFinding(
                    endpoint=endpoint,
                    status=EndpointStatus.UNUSED,
                    recommendation=f"미사용 - {ANALYSIS_DAYS}일간 호출 0회, {endpoint.age_days}일 경과 - 삭제 검토 (${endpoint.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 유휴: 호출은 있으나 CPU 5% 미만 (AWS Cost Optimization Pillar)
        if endpoint.cpu_utilization_avg > 0 and endpoint.cpu_utilization_avg < IDLE_CPU_THRESHOLD:
            result.idle_endpoints += 1
            result.idle_monthly_cost += endpoint.estimated_monthly_cost
            result.findings.append(
                EndpointFinding(
                    endpoint=endpoint,
                    status=EndpointStatus.IDLE,
                    recommendation=f"유휴 (CPU {endpoint.cpu_utilization_avg:.1f}%) - 인스턴스 축소 또는 Serverless 전환 검토 (${endpoint.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: 하루 평균 10회 미만
        if endpoint.avg_invocations_per_day < LOW_USAGE_INVOCATIONS_PER_DAY:
            result.low_usage_endpoints += 1
            result.low_usage_monthly_cost += endpoint.estimated_monthly_cost
            result.findings.append(
                EndpointFinding(
                    endpoint=endpoint,
                    status=EndpointStatus.LOW_USAGE,
                    recommendation=f"저사용 (일 평균 {endpoint.avg_invocations_per_day:.1f}회) - Serverless Endpoint 전환 검토",
                )
            )
            continue

        result.normal_endpoints += 1
        result.findings.append(
            EndpointFinding(
                endpoint=endpoint,
                status=EndpointStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[SageMakerAnalysisResult], output_dir: str) -> str:
    """SageMaker Endpoint 미사용 분석 결과를 Excel 보고서로 생성한다.

    Summary(계정별 통계)와 Endpoints(비정상 Endpoint 상세) 시트를 포함한다.

    Args:
        results: 계정/리전별 SageMaker 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_endpoints,
                r.unused_endpoints,
                r.low_usage_endpoints,
                r.normal_endpoints,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.unused_endpoints > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_endpoints > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Endpoint Name", width=35),
        ColumnDef(header="Status", width=12),
        ColumnDef(header="Instance Type", width=18),
        ColumnDef(header="Count", width=8, style="number"),
        ColumnDef(header="Total Invocations", width=15, style="number"),
        ColumnDef(header="Avg/Day", width=10),
        ColumnDef(header="Latency (ms)", width=12),
        ColumnDef(header="Age (days)", width=12, style="number"),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=45),
    ]
    detail_sheet = wb.new_sheet("Endpoints", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != EndpointStatus.NORMAL:
                ep = f.endpoint
                style = Styles.danger() if f.status == EndpointStatus.UNUSED else Styles.warning()

                detail_sheet.add_row(
                    [
                        ep.account_name,
                        ep.region,
                        ep.endpoint_name,
                        ep.status,
                        ep.instance_type,
                        ep.instance_count,
                        ep.total_invocations,
                        f"{ep.avg_invocations_per_day:.1f}",
                        f"{ep.model_latency_avg_ms:.2f}",
                        ep.age_days,
                        f"${ep.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "SageMaker_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SageMakerAnalysisResult | None:
    """parallel_collect 콜백: 단일 계정/리전의 SageMaker Endpoint를 수집 및 분석한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        SageMaker 분석 결과. Endpoint가 없으면 None.
    """
    endpoints = collect_sagemaker_endpoints(session, account_id, account_name, region)
    if not endpoints:
        return None
    return analyze_endpoints(endpoints, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """SageMaker Endpoint 미사용 분석 도구의 메인 실행 함수.

    CloudWatch 지표 기반으로 7일간 미사용/유휴/저사용 SageMaker Endpoint를
    탐지하고 비용 절감 기회를 보고서로 생성한다.

    Args:
        ctx: 실행 컨텍스트. 계정 정보, 리전, 프로파일 등을 포함한다.
    """
    console.print("[bold]SageMaker Endpoint 미사용 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="sagemaker")
    results: list[SageMakerAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_endpoints for r in results)
    total_low = sum(r.low_usage_endpoints for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월)"
    )

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("sagemaker", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
