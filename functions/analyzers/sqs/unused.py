"""
functions/analyzers/sqs/unused.py - SQS 미사용 큐 분석

유휴/미사용 SQS 큐 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from core.shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 분석 기간 (일) - AWS 권장 14일 이상
ANALYSIS_DAYS = 14

# 저사용 기준: 하루 평균 메시지 10개 미만
LOW_USAGE_MESSAGES_PER_DAY = 10

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "sqs:ListQueues",
        "sqs:GetQueueAttributes",
        "cloudwatch:GetMetricStatistics",
    ],
}


class QueueStatus(Enum):
    """SQS 큐 분석 상태.

    CloudWatch 메시지 활동 지표 및 DLQ 여부 기반으로 분류한다.

    Attributes:
        NORMAL: 정상 사용 중인 큐.
        UNUSED: 분석 기간 동안 메시지 활동이 없는 미사용 큐.
        LOW_USAGE: 일 평균 메시지 수가 기준치 미만인 저사용 큐.
        EMPTY_DLQ: 메시지가 없고 활동도 없는 빈 Dead Letter Queue.
    """

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"
    EMPTY_DLQ = "empty_dlq"


@dataclass
class SQSQueueInfo:
    """SQS 큐 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 큐가 위치한 리전.
        queue_name: 큐 이름.
        queue_url: 큐 URL.
        queue_arn: 큐 ARN.
        is_fifo: FIFO 큐 여부.
        is_dlq: Dead Letter Queue 여부.
        approximate_messages: 현재 대기 중인 메시지 수 (근사값).
        approximate_messages_delayed: 지연 전달 중인 메시지 수 (근사값).
        approximate_messages_not_visible: 처리 중(invisible) 메시지 수 (근사값).
        created_timestamp: 큐 생성 시각.
        messages_sent: 분석 기간 동안 전송된 메시지 수 (CloudWatch Sum).
        messages_received: 분석 기간 동안 수신된 메시지 수 (CloudWatch Sum).
        messages_deleted: 분석 기간 동안 삭제된 메시지 수 (CloudWatch Sum).
    """

    account_id: str
    account_name: str
    region: str
    queue_name: str
    queue_url: str
    queue_arn: str
    is_fifo: bool
    is_dlq: bool
    approximate_messages: int
    approximate_messages_delayed: int
    approximate_messages_not_visible: int
    created_timestamp: datetime | None
    # CloudWatch 지표
    messages_sent: float = 0.0
    messages_received: float = 0.0
    messages_deleted: float = 0.0


@dataclass
class QueueFinding:
    """개별 SQS 큐에 대한 분석 결과.

    Attributes:
        queue: 분석 대상 큐 정보.
        status: 분석 결과 상태.
        recommendation: 권장 조치 사항 (한글).
    """

    queue: SQSQueueInfo
    status: QueueStatus
    recommendation: str


@dataclass
class SQSAnalysisResult:
    """단일 계정/리전의 SQS 큐 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.
        total_queues: 전체 큐 수.
        unused_queues: 미사용 큐 수.
        low_usage_queues: 저사용 큐 수.
        empty_dlqs: 빈 DLQ 수.
        normal_queues: 정상 큐 수.
        findings: 개별 큐 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_queues: int = 0
    unused_queues: int = 0
    low_usage_queues: int = 0
    empty_dlqs: int = 0
    normal_queues: int = 0
    findings: list[QueueFinding] = field(default_factory=list)


def collect_sqs_queues(session, account_id: str, account_name: str, region: str) -> list[SQSQueueInfo]:
    """지정된 계정/리전의 SQS 큐를 수집하고 CloudWatch 메트릭을 조회한다.

    1단계에서 큐 목록과 속성(메시지 수, FIFO 여부, DLQ 여부)을 수집하고,
    2단계에서 batch_get_metrics를 통해 메시지 전송/수신/삭제 지표를 배치 조회한다.
    기존 큐당 3 API 호출을 전체 1-2 호출로 최적화했다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        수집된 SQS 큐 정보 목록. CloudWatch 메트릭이 포함된다.
    """
    from botocore.exceptions import ClientError

    sqs = get_client(session, "sqs", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    queues: list[SQSQueueInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 큐 목록 수집
        paginator = sqs.get_paginator("list_queues")
        for page in paginator.paginate():
            for queue_url in page.get("QueueUrls", []):
                try:
                    attrs = sqs.get_queue_attributes(
                        QueueUrl=queue_url,
                        AttributeNames=["All"],
                    ).get("Attributes", {})

                    queue_arn = attrs.get("QueueArn", "")
                    queue_name = queue_url.split("/")[-1]

                    # DLQ 여부 확인
                    is_dlq = "dlq" in queue_name.lower() or "dead" in queue_name.lower()

                    # 생성 시간
                    created_ts = attrs.get("CreatedTimestamp")
                    created_at = None
                    if created_ts:
                        created_at = datetime.fromtimestamp(int(created_ts), tz=timezone.utc)

                    info = SQSQueueInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        queue_name=queue_name,
                        queue_url=queue_url,
                        queue_arn=queue_arn,
                        is_fifo=queue_name.endswith(".fifo"),
                        is_dlq=is_dlq,
                        approximate_messages=int(attrs.get("ApproximateNumberOfMessages", 0)),
                        approximate_messages_delayed=int(attrs.get("ApproximateNumberOfMessagesDelayed", 0)),
                        approximate_messages_not_visible=int(attrs.get("ApproximateNumberOfMessagesNotVisible", 0)),
                        created_timestamp=created_at,
                    )
                    queues.append(info)

                except ClientError as e:
                    category = categorize_error(e)
                    if category != ErrorCategory.NOT_FOUND:
                        logger.debug(f"SQS 큐 속성 조회 실패: {queue_url} ({get_error_code(e)})")
                    continue

        # 2단계: 배치 메트릭 조회
        if queues:
            _collect_sqs_metrics_batch(cloudwatch, queues, start_time, now)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"SQS 권한 없음: {account_name}/{region}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"SQS 조회 오류: {get_error_code(e)}")

    return queues


def _collect_sqs_metrics_batch(
    cloudwatch,
    queues: list[SQSQueueInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """SQS 큐의 CloudWatch 메트릭을 배치로 수집하여 큐 객체에 반영한다.

    NumberOfMessagesSent, NumberOfMessagesReceived, NumberOfMessagesDeleted
    세 가지 메트릭의 Sum 값을 일별(86400초) 기간으로 조회한다.

    Args:
        cloudwatch: CloudWatch boto3 클라이언트.
        queues: 메트릭을 수집할 SQS 큐 목록 (결과가 각 객체에 직접 반영됨).
        start_time: 메트릭 조회 시작 시각 (UTC).
        end_time: 메트릭 조회 종료 시각 (UTC).
    """
    from botocore.exceptions import ClientError

    queries: list[MetricQuery] = []

    for queue in queues:
        safe_id = sanitize_metric_id(queue.queue_name)
        dimensions = {"QueueName": queue.queue_name}

        queries.extend(
            [
                MetricQuery(
                    id=f"{safe_id}_sent",
                    namespace="AWS/SQS",
                    metric_name="NumberOfMessagesSent",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_received",
                    namespace="AWS/SQS",
                    metric_name="NumberOfMessagesReceived",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_deleted",
                    namespace="AWS/SQS",
                    metric_name="NumberOfMessagesDeleted",
                    dimensions=dimensions,
                    stat="Sum",
                ),
            ]
        )

    if not queries:
        return

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        for queue in queues:
            safe_id = sanitize_metric_id(queue.queue_name)
            queue.messages_sent = results.get(f"{safe_id}_sent", 0.0)
            queue.messages_received = results.get(f"{safe_id}_received", 0.0)
            queue.messages_deleted = results.get(f"{safe_id}_deleted", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_queues(queues: list[SQSQueueInfo], account_id: str, account_name: str, region: str) -> SQSAnalysisResult:
    """수집된 SQS 큐를 분석하여 미사용/저사용 큐를 식별한다.

    메시지 활동(전송+수신+삭제) 합계, DLQ 여부, 일 평균 메시지 수를 기반으로
    EMPTY_DLQ, UNUSED, LOW_USAGE, NORMAL 상태로 분류한다.

    Args:
        queues: 분석 대상 SQS 큐 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.

    Returns:
        SQS 큐 분석 결과 집계 객체.
    """
    result = SQSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_queues=len(queues),
    )

    for queue in queues:
        total_activity = queue.messages_sent + queue.messages_received + queue.messages_deleted

        # 빈 DLQ
        if queue.is_dlq and queue.approximate_messages == 0 and total_activity == 0:
            result.empty_dlqs += 1
            result.findings.append(
                QueueFinding(
                    queue=queue,
                    status=QueueStatus.EMPTY_DLQ,
                    recommendation="빈 DLQ - 정리 검토",
                )
            )
            continue

        # 미사용 큐
        if total_activity == 0:
            result.unused_queues += 1
            result.findings.append(
                QueueFinding(
                    queue=queue,
                    status=QueueStatus.UNUSED,
                    recommendation=f"{ANALYSIS_DAYS}일간 활동 없음 - 삭제 검토",
                )
            )
            continue

        # 저사용 큐 (일평균 메시지 수 기준)
        avg_messages_per_day = (queue.messages_sent + queue.messages_received) / ANALYSIS_DAYS
        if avg_messages_per_day < LOW_USAGE_MESSAGES_PER_DAY:
            result.low_usage_queues += 1
            result.findings.append(
                QueueFinding(
                    queue=queue,
                    status=QueueStatus.LOW_USAGE,
                    recommendation=f"저사용 (일평균 {avg_messages_per_day:.1f}건) - 통합 검토",
                )
            )
            continue

        result.normal_queues += 1
        result.findings.append(
            QueueFinding(
                queue=queue,
                status=QueueStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[SQSAnalysisResult], output_dir: str) -> str:
    """SQS 미사용 큐 분석 결과를 Excel 보고서로 생성한다.

    Summary 시트(계정/리전별 통계)와 Queues 시트(비정상 큐 상세)를 포함한다.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary Sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("SQS 미사용 분석 보고서")

    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="빈DLQ", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_queues,
                r.unused_queues,
                r.empty_dlqs,
                r.normal_queues,
            ]
        )
        ws = summary_sheet._ws
        if r.unused_queues > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.empty_dlqs > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail Sheet
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Queue Name", width=50),
        ColumnDef(header="Type", width=15),
        ColumnDef(header="Messages", width=12, style="number"),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="Sent", width=12, style="number"),
        ColumnDef(header="Received", width=12, style="number"),
        ColumnDef(header="Deleted", width=12, style="number"),
        ColumnDef(header="권장 조치", width=30),
    ]
    detail_sheet = wb.new_sheet("Queues", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != QueueStatus.NORMAL:
                q = f.queue
                queue_type = "FIFO" if q.is_fifo else "Standard"
                if q.is_dlq:
                    queue_type += " (DLQ)"
                detail_sheet.add_row(
                    [
                        q.account_name,
                        q.region,
                        q.queue_name,
                        queue_type,
                        q.approximate_messages,
                        f.status.value,
                        int(q.messages_sent),
                        int(q.messages_received),
                        int(q.messages_deleted),
                        f.recommendation,
                    ]
                )

    return str(wb.save_as(output_dir, "SQS_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SQSAnalysisResult | None:
    """단일 계정/리전의 SQS 큐를 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 큐가 없으면 None을 반환한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 대상 리전.

    Returns:
        분석 결과 객체. 큐가 없으면 None.
    """
    queues = collect_sqs_queues(session, account_id, account_name, region)
    if not queues:
        return None
    return analyze_queues(queues, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """SQS 미사용 큐 분석 도구의 메인 실행 함수.

    멀티 계정/리전 병렬 수집 후 결과를 집계하고, Excel 보고서를 생성하여 출력 디렉토리에 저장한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 계정/리전 목록, 옵션 등 포함).
    """
    console.print("[bold]SQS 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="sqs")
    results: list[SQSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_queues for r in results)
    total_empty_dlq = sum(r.empty_dlqs for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미사용 큐: [red]{total_unused}개[/red] / 빈 DLQ: [yellow]{total_empty_dlq}개[/yellow]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("sqs", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
