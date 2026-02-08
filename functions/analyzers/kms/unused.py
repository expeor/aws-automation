"""
functions/analyzers/kms/unused.py - KMS 키 미사용 분석

미사용/비활성화 고객 관리 키 (CMK) 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.aws.pricing import get_kms_key_price
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "kms:ListKeys",
        "kms:DescribeKey",
        "kms:ListAliases",
    ],
}


class KMSKeyStatus(Enum):
    """KMS 키 분석 상태.

    고객 관리 키(CMK)의 활성화/비활성화/삭제 예정 상태를 분류한다.

    Attributes:
        NORMAL: 정상 활성화된 키.
        DISABLED: 비활성화된 키 (비용은 계속 발생).
        PENDING_DELETE: 삭제 예정인 키.
    """

    NORMAL = "normal"
    DISABLED = "disabled"
    PENDING_DELETE = "pending_delete"


@dataclass
class KMSKeyInfo:
    """KMS 키 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 키가 위치한 리전.
        key_id: KMS 키 ID.
        arn: 키 ARN.
        description: 키 설명.
        key_state: 키 상태 (Enabled, Disabled, PendingDeletion 등).
        key_manager: 키 관리자 (AWS 또는 CUSTOMER).
        creation_date: 키 생성일.
        deletion_date: 삭제 예정일 (PendingDeletion 상태인 경우).
        alias: 키 별칭 (aws/ 접두사 제외).
    """

    account_id: str
    account_name: str
    region: str
    key_id: str
    arn: str
    description: str
    key_state: str
    key_manager: str  # AWS or CUSTOMER
    creation_date: datetime | None
    deletion_date: datetime | None = None
    alias: str = ""

    @property
    def is_customer_managed(self) -> bool:
        """고객 관리 키(CMK) 여부.

        Returns:
            CUSTOMER 관리 키이면 True.
        """
        return self.key_manager == "CUSTOMER"

    @property
    def monthly_cost(self) -> float:
        """키 월간 비용 (USD).

        Returns:
            리전 및 키 관리자 유형에 따른 월 비용.
        """
        return get_kms_key_price(self.region, self.key_manager)


@dataclass
class KMSKeyFinding:
    """개별 KMS 키에 대한 분석 결과.

    Attributes:
        key: 분석 대상 키 정보.
        status: 분석 결과 상태.
        recommendation: 권장 조치 사항 (한글).
    """

    key: KMSKeyInfo
    status: KMSKeyStatus
    recommendation: str


@dataclass
class KMSKeyAnalysisResult:
    """단일 계정/리전의 KMS 키 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.
        total_count: 전체 키 수.
        customer_managed_count: 고객 관리 키(CMK) 수.
        aws_managed_count: AWS 관리 키 수.
        disabled_count: 비활성화된 CMK 수.
        pending_delete_count: 삭제 예정 키 수.
        normal_count: 정상 키 수.
        disabled_monthly_cost: 비활성화 키의 월간 비용 합계 (USD).
        findings: 개별 키 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    customer_managed_count: int = 0
    aws_managed_count: int = 0
    disabled_count: int = 0
    pending_delete_count: int = 0
    normal_count: int = 0
    disabled_monthly_cost: float = 0.0
    findings: list[KMSKeyFinding] = field(default_factory=list)


def collect_kms_keys(session, account_id: str, account_name: str, region: str) -> list[KMSKeyInfo]:
    """지정된 계정/리전의 KMS 키를 수집한다.

    모든 키(AWS 관리 + 고객 관리)를 조회하고 별칭 정보도 함께 수집한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        수집된 KMS 키 정보 목록.
    """
    from botocore.exceptions import ClientError

    kms = get_client(session, "kms", region_name=region)
    keys = []

    paginator = kms.get_paginator("list_keys")
    for page in paginator.paginate():
        for key in page.get("Keys", []):
            try:
                key_info = kms.describe_key(KeyId=key["KeyId"])["KeyMetadata"]

                alias = ""
                try:
                    aliases_resp = kms.list_aliases(KeyId=key["KeyId"])
                    for a in aliases_resp.get("Aliases", []):
                        if not a["AliasName"].startswith("alias/aws/"):
                            alias = a["AliasName"]
                            break
                except ClientError:
                    pass

                keys.append(
                    KMSKeyInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        key_id=key["KeyId"],
                        arn=key_info.get("Arn", ""),
                        description=key_info.get("Description", ""),
                        key_state=key_info.get("KeyState", ""),
                        key_manager=key_info.get("KeyManager", ""),
                        creation_date=key_info.get("CreationDate"),
                        deletion_date=key_info.get("DeletionDate"),
                        alias=alias,
                    )
                )
            except ClientError:
                continue

    return keys


def analyze_kms_keys(keys: list[KMSKeyInfo], account_id: str, account_name: str, region: str) -> KMSKeyAnalysisResult:
    """수집된 KMS 키를 분석하여 비활성화/삭제 예정 키를 식별한다.

    AWS 관리 키는 무료이므로 정상 처리하고, CMK 중 비활성화/삭제 예정인 키를 분류한다.

    Args:
        keys: 분석 대상 KMS 키 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.

    Returns:
        KMS 키 분석 결과 집계 객체.
    """
    result = KMSKeyAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(keys),
    )

    for key in keys:
        if key.is_customer_managed:
            result.customer_managed_count += 1
        else:
            result.aws_managed_count += 1
            result.normal_count += 1
            result.findings.append(
                KMSKeyFinding(
                    key=key,
                    status=KMSKeyStatus.NORMAL,
                    recommendation="AWS 관리 키 (무료)",
                )
            )
            continue

        if key.key_state == "PendingDeletion":
            result.pending_delete_count += 1
            result.findings.append(
                KMSKeyFinding(
                    key=key,
                    status=KMSKeyStatus.PENDING_DELETE,
                    recommendation=f"삭제 예정: {key.deletion_date.strftime('%Y-%m-%d') if key.deletion_date else 'N/A'}",
                )
            )
            continue

        if key.key_state == "Disabled":
            result.disabled_count += 1
            result.disabled_monthly_cost += key.monthly_cost
            result.findings.append(
                KMSKeyFinding(
                    key=key,
                    status=KMSKeyStatus.DISABLED,
                    recommendation="비활성화됨 - 삭제 검토",
                )
            )
            continue

        result.normal_count += 1
        result.findings.append(
            KMSKeyFinding(
                key=key,
                status=KMSKeyStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[KMSKeyAnalysisResult], output_dir: str) -> str:
    """KMS 미사용 키 분석 결과를 Excel 보고서로 생성한다.

    Summary 시트(계정/리전별 통계)와 Detail 시트(CMK 상세)를 포함한다.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="CMK", width=10, style="number"),
        ColumnDef(header="AWS관리", width=10, style="number"),
        ColumnDef(header="비활성화", width=10, style="number"),
        ColumnDef(header="삭제예정", width=10, style="number"),
        ColumnDef(header="비활성화 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_style = Styles.warning() if r.disabled_count > 0 else None
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_count,
                r.customer_managed_count,
                r.aws_managed_count,
                r.disabled_count,
                r.pending_delete_count,
                f"${r.disabled_monthly_cost:,.2f}",
            ],
            style=row_style,
        )
        # Cell-level highlighting for disabled count
        if r.disabled_count > 0:
            ws = summary_sheet._ws
            ws.cell(row=row_num, column=6).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Key ID", width=25),
        ColumnDef(header="Alias", width=25),
        ColumnDef(header="Type", width=12, style="center"),
        ColumnDef(header="상태", width=15, style="center"),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=30),
    ]
    detail_sheet = wb.new_sheet("KMS Keys", detail_columns)

    for r in results:
        for f in r.findings:
            if f.key.is_customer_managed:
                k = f.key
                row_style = None
                if f.status == KMSKeyStatus.DISABLED:
                    row_style = Styles.warning()
                elif f.status == KMSKeyStatus.PENDING_DELETE:
                    row_style = Styles.danger()
                detail_sheet.add_row(
                    [
                        k.account_name,
                        k.region,
                        k.key_id[:20] + "...",
                        k.alias or "-",
                        k.key_manager,
                        k.key_state,
                        f"${k.monthly_cost:,.2f}",
                        f.recommendation,
                    ],
                    style=row_style,
                )

    return str(wb.save_as(output_dir, "KMS_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> KMSKeyAnalysisResult | None:
    """단일 계정/리전의 KMS 키를 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 키가 없으면 None을 반환한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 대상 리전.

    Returns:
        분석 결과 객체. 키가 없으면 None.
    """
    keys = collect_kms_keys(session, account_id, account_name, region)
    if not keys:
        return None
    return analyze_kms_keys(keys, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """KMS 미사용/비활성화 키 분석 도구의 메인 실행 함수.

    멀티 계정/리전 병렬 수집 후 결과를 집계하고, Excel 보고서를 생성하여 출력 디렉토리에 저장한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 계정/리전 목록, 옵션 등 포함).
    """
    console.print("[bold]KMS 키 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="kms")
    results: list[KMSKeyAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_cmk = sum(r.customer_managed_count for r in results)
    total_disabled = sum(r.disabled_count for r in results)
    disabled_cost = sum(r.disabled_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"CMK: {total_cmk}개 / 비활성화: [yellow]{total_disabled}개[/yellow] (${disabled_cost:,.2f}/월)")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("kms", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
