"""
functions/analyzers/cloudwatch/alarm_orphan.py - CloudWatch 고아 알람 분석

모니터링 대상 리소스가 없는 알람 탐지
- 지표 데이터 존재 여부로 판단 (모든 namespace 지원)
- INSUFFICIENT_DATA 상태 = 고아 알람

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "cloudwatch:DescribeAlarms",
        "cloudwatch:GetMetricStatistics",
    ],
}

# 지표 데이터 확인 기간 (일)
METRIC_CHECK_DAYS = 7


class AlarmStatus(Enum):
    """CloudWatch 알람 분석 상태.

    알람의 정상/고아/무액션 상태를 분류한다.

    Attributes:
        NORMAL: 정상 동작 중인 알람.
        ORPHAN: 모니터링 대상 리소스가 삭제되어 지표 데이터가 없는 고아 알람.
        NO_ACTIONS: 알람 트리거 시 실행할 액션이 설정되지 않은 알람.
    """

    NORMAL = "normal"
    ORPHAN = "orphan"  # 지표 데이터 없음 (리소스 삭제됨)
    NO_ACTIONS = "no_actions"  # 알람 액션 없음


@dataclass
class AlarmInfo:
    """CloudWatch 알람 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 리전.
        alarm_name: 알람 이름.
        alarm_arn: 알람 ARN.
        namespace: CloudWatch 네임스페이스 (AWS/EC2, AWS/RDS 등).
        metric_name: 지표 이름.
        dimensions: Dimension 문자열 (표시용, "Name=Value" 형식).
        dimensions_list: Dimension 딕셔너리 목록 (API 호출용).
        state: 알람 상태 (OK, ALARM, INSUFFICIENT_DATA).
        state_reason: 상태 사유 (최대 100자).
        actions_enabled: 알람 액션 활성화 여부.
        alarm_actions: ALARM 상태 시 실행할 액션 ARN 목록.
        ok_actions: OK 상태 시 실행할 액션 ARN 목록.
        insufficient_data_actions: INSUFFICIENT_DATA 상태 시 실행할 액션 ARN 목록.
        has_metric_data: 최근 7일간 지표 데이터 존재 여부.
    """

    account_id: str
    account_name: str
    region: str
    alarm_name: str
    alarm_arn: str
    namespace: str
    metric_name: str
    dimensions: str
    dimensions_list: list[dict]
    state: str
    state_reason: str
    actions_enabled: bool
    alarm_actions: list[str]
    ok_actions: list[str]
    insufficient_data_actions: list[str]
    has_metric_data: bool = True


@dataclass
class AlarmFinding:
    """개별 알람에 대한 분석 결과.

    Attributes:
        alarm: 분석 대상 알람 정보.
        status: 분석된 알람 상태 (NORMAL, ORPHAN, NO_ACTIONS).
        recommendation: 권장 조치 메시지.
    """

    alarm: AlarmInfo
    status: AlarmStatus
    recommendation: str


@dataclass
class AlarmAnalysisResult:
    """계정/리전별 알람 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 리전.
        total_alarms: 전체 알람 수.
        orphan_alarms: 고아 알람 수 (지표 데이터 없음).
        no_actions: 액션 미설정 알람 수.
        normal_alarms: 정상 알람 수.
        findings: 개별 알람 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_alarms: int = 0
    orphan_alarms: int = 0
    no_actions: int = 0
    normal_alarms: int = 0
    findings: list[AlarmFinding] = field(default_factory=list)


def check_metric_has_data(cloudwatch, namespace: str, metric_name: str, dimensions: list[dict]) -> bool:
    """지표에 최근 데이터가 있는지 확인한다.

    get_metric_statistics로 최근 METRIC_CHECK_DAYS일간 데이터포인트를 조회하여
    리소스 존재 여부를 판단한다. 모든 CloudWatch 네임스페이스를 지원한다.

    Args:
        cloudwatch: CloudWatch boto3 클라이언트.
        namespace: CloudWatch 네임스페이스 (AWS/EC2 등).
        metric_name: 지표 이름.
        dimensions: Dimension 딕셔너리 목록.

    Returns:
        데이터포인트가 하나라도 있으면 True.
    """
    from botocore.exceptions import ClientError

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=METRIC_CHECK_DAYS)

    try:
        resp = cloudwatch.get_metric_statistics(
            Namespace=namespace,
            MetricName=metric_name,
            Dimensions=dimensions,
            StartTime=start_time,
            EndTime=now,
            Period=86400,  # 1일
            Statistics=["Sum", "Average", "SampleCount"],
        )
        # 데이터포인트가 하나라도 있으면 정상
        return len(resp.get("Datapoints", [])) > 0
    except ClientError:
        # API 오류는 데이터 없음으로 간주
        return False


def collect_alarms(session, account_id: str, account_name: str, region: str) -> list[AlarmInfo]:
    """CloudWatch MetricAlarm을 수집하고 지표 데이터 존재 여부를 확인한다.

    INSUFFICIENT_DATA 상태의 알람은 즉시 고아로 판단하고, 그 외 알람은
    get_metric_statistics로 실제 데이터 존재를 확인한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        수집된 AlarmInfo 목록.
    """
    from botocore.exceptions import ClientError

    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    alarms = []

    try:
        paginator = cloudwatch.get_paginator("describe_alarms")
        for page in paginator.paginate():
            for alarm in page.get("MetricAlarms", []):
                dimensions_list = alarm.get("Dimensions", [])
                dim_str_parts = [f"{d['Name']}={d['Value']}" for d in dimensions_list]

                namespace = alarm.get("Namespace", "")
                metric_name = alarm.get("MetricName", "")
                state = alarm.get("StateValue", "")

                info = AlarmInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    alarm_name=alarm.get("AlarmName", ""),
                    alarm_arn=alarm.get("AlarmArn", ""),
                    namespace=namespace,
                    metric_name=metric_name,
                    dimensions=", ".join(dim_str_parts) if dim_str_parts else "-",
                    dimensions_list=dimensions_list,
                    state=state,
                    state_reason=alarm.get("StateReason", "")[:100],
                    actions_enabled=alarm.get("ActionsEnabled", False),
                    alarm_actions=alarm.get("AlarmActions", []),
                    ok_actions=alarm.get("OKActions", []),
                    insufficient_data_actions=alarm.get("InsufficientDataActions", []),
                )

                # INSUFFICIENT_DATA 상태 = 고아 알람 가능성 높음
                # 추가로 지표 데이터 존재 여부 확인
                if state == "INSUFFICIENT_DATA":
                    info.has_metric_data = False
                elif namespace and metric_name:
                    # 모든 namespace에 대해 지표 데이터 확인
                    info.has_metric_data = check_metric_has_data(cloudwatch, namespace, metric_name, dimensions_list)

                alarms.append(info)

    except ClientError:
        pass

    return alarms


def analyze_alarms(alarms: list[AlarmInfo], account_id: str, account_name: str, region: str) -> AlarmAnalysisResult:
    """수집된 CloudWatch 알람을 분석한다.

    고아 알람(지표 데이터 없음), 액션 미설정 알람, 정상 알람으로 분류하고
    각 알람에 대한 권장 조치를 생성한다.

    Args:
        alarms: 수집된 AlarmInfo 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 리전.

    Returns:
        분석 결과를 담은 AlarmAnalysisResult 객체.
    """
    result = AlarmAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_alarms=len(alarms),
    )

    for alarm in alarms:
        # 고아 알람: 지표 데이터 없음 (INSUFFICIENT_DATA 포함)
        if not alarm.has_metric_data:
            result.orphan_alarms += 1
            reason = "INSUFFICIENT_DATA" if alarm.state == "INSUFFICIENT_DATA" else "지표 데이터 없음"
            result.findings.append(
                AlarmFinding(
                    alarm=alarm,
                    status=AlarmStatus.ORPHAN,
                    recommendation=f"{reason} - 리소스 삭제됨, 알람 삭제 검토",
                )
            )
            continue

        # 액션 없는 알람
        if not alarm.actions_enabled or (
            not alarm.alarm_actions and not alarm.ok_actions and not alarm.insufficient_data_actions
        ):
            result.no_actions += 1
            result.findings.append(
                AlarmFinding(
                    alarm=alarm,
                    status=AlarmStatus.NO_ACTIONS,
                    recommendation="알람 액션 없음 - 알림 설정 검토",
                )
            )
            continue

        result.normal_alarms += 1
        result.findings.append(
            AlarmFinding(
                alarm=alarm,
                status=AlarmStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[AlarmAnalysisResult], output_dir: str) -> str:
    """고아 알람 분석 결과를 Excel 보고서로 생성한다.

    Summary(계정별 통계)와 Alarms(비정상 알람 상세) 시트를 포함한다.

    Args:
        results: 계정/리전별 알람 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="고아", width=10, style="number"),
        ColumnDef(header="액션없음", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_alarms,
                r.orphan_alarms,
                r.no_actions,
                r.normal_alarms,
            ]
        )
        ws = summary_sheet._ws
        if r.orphan_alarms > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.no_actions > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Alarm Name", width=30),
        ColumnDef(header="Namespace", width=20),
        ColumnDef(header="Metric", width=20),
        ColumnDef(header="Dimensions", width=30),
        ColumnDef(header="State", width=15),
        ColumnDef(header="분석상태", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Alarms", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != AlarmStatus.NORMAL:
                a = f.alarm
                style = Styles.danger() if f.status == AlarmStatus.ORPHAN else Styles.warning()
                detail_sheet.add_row(
                    [
                        a.account_name,
                        a.region,
                        a.alarm_name,
                        a.namespace,
                        a.metric_name,
                        a.dimensions,
                        a.state,
                        f.status.value,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "CloudWatch_Alarm_Orphan"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> AlarmAnalysisResult | None:
    """parallel_collect 콜백: 단일 계정/리전의 알람을 수집 및 분석한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        알람 분석 결과. 알람이 없으면 None.
    """
    alarms = collect_alarms(session, account_id, account_name, region)
    if not alarms:
        return None
    return analyze_alarms(alarms, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """CloudWatch 고아 알람 분석 도구의 메인 실행 함수.

    모니터링 대상 리소스가 없는 고아 알람과 액션 미설정 알람을 탐지한다.
    최근 7일간 지표 데이터 존재 여부로 고아 여부를 판단한다.

    Args:
        ctx: 실행 컨텍스트. 계정 정보, 리전, 프로파일 등을 포함한다.
    """
    console.print("[bold]CloudWatch 알람 분석 시작...[/bold]\n")
    console.print(f"[dim]* 고아 알람 기준: {METRIC_CHECK_DAYS}일간 지표 데이터 없음[/dim]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="cloudwatch")
    results: list[AlarmAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_orphan = sum(r.orphan_alarms for r in results)
    total_no_action = sum(r.no_actions for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"고아 알람: [red]{total_orphan}개[/red] / 액션없음: [yellow]{total_no_action}개[/yellow]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("cloudwatch", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
