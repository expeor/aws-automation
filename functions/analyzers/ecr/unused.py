"""
functions/analyzers/ecr/unused.py - ECR 미사용 이미지 분석

오래된/미사용 ECR 이미지 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.aws.pricing import get_ecr_storage_price
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ecr:DescribeRepositories",
        "ecr:GetLifecyclePolicy",
        "ecr:DescribeImages",
    ],
}

# 미사용 기준: 90일 이상 pull 없음
UNUSED_DAYS_THRESHOLD = 90


class ECRRepoStatus(Enum):
    """ECR 리포지토리 사용 상태 분류.

    이미지 유무, 오래된 이미지 존재 여부, 라이프사이클 정책 유무를 기반으로 분류한다.
    """

    NORMAL = "normal"
    EMPTY = "empty"
    OLD_IMAGES = "old_images"
    NO_LIFECYCLE = "no_lifecycle"


@dataclass
class ECRRepoInfo:
    """ECR 리포지토리 메타데이터 및 이미지 통계 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        name: 리포지토리 이름.
        arn: 리포지토리 ARN.
        uri: 리포지토리 URI.
        created_at: 리포지토리 생성 시각.
        image_count: 전체 이미지 수.
        total_size_bytes: 전체 이미지 크기 (bytes).
        has_lifecycle_policy: 라이프사이클 정책 존재 여부.
        old_image_count: 오래된 이미지 수 (UNUSED_DAYS_THRESHOLD 기준).
        old_images_size_bytes: 오래된 이미지 크기 (bytes).
    """

    account_id: str
    account_name: str
    region: str
    name: str
    arn: str
    uri: str
    created_at: datetime | None
    image_count: int = 0
    total_size_bytes: int = 0
    has_lifecycle_policy: bool = False
    old_image_count: int = 0
    old_images_size_bytes: int = 0

    @property
    def total_size_gb(self) -> float:
        """전체 이미지 크기를 GB 단위로 반환한다.

        Returns:
            전체 이미지 크기 (GB).
        """
        return self.total_size_bytes / (1024**3)

    @property
    def old_images_size_gb(self) -> float:
        """오래된 이미지 크기를 GB 단위로 반환한다.

        Returns:
            오래된 이미지 크기 (GB).
        """
        return self.old_images_size_bytes / (1024**3)

    @property
    def monthly_cost(self) -> float:
        """전체 이미지의 월간 스토리지 비용.

        Returns:
            월간 스토리지 비용 (USD).
        """
        return self.total_size_gb * get_ecr_storage_price(self.region)

    @property
    def old_images_monthly_cost(self) -> float:
        """오래된 이미지의 월간 스토리지 비용.

        Returns:
            오래된 이미지 월간 스토리지 비용 (USD).
        """
        return self.old_images_size_gb * get_ecr_storage_price(self.region)


@dataclass
class ECRRepoFinding:
    """개별 ECR 리포지토리의 분석 결과.

    Attributes:
        repo: 분석 대상 리포지토리 정보.
        status: 분석된 리포지토리 상태.
        recommendation: 권장 조치 사항 문자열.
    """

    repo: ECRRepoInfo
    status: ECRRepoStatus
    recommendation: str


@dataclass
class ECRAnalysisResult:
    """단일 계정/리전의 ECR 미사용 이미지 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        total_repos: 전체 리포지토리 수.
        empty_repos: 빈 리포지토리 수.
        repos_with_old_images: 오래된 이미지가 있는 리포지토리 수.
        no_lifecycle_repos: 라이프사이클 정책이 없는 리포지토리 수.
        total_images: 전체 이미지 수.
        old_images: 오래된 이미지 수.
        total_size_gb: 전체 이미지 크기 (GB).
        old_images_size_gb: 오래된 이미지 크기 (GB).
        old_images_monthly_cost: 오래된 이미지 월간 스토리지 비용 (USD).
        findings: 개별 리포지토리별 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_repos: int = 0
    empty_repos: int = 0
    repos_with_old_images: int = 0
    no_lifecycle_repos: int = 0
    total_images: int = 0
    old_images: int = 0
    total_size_gb: float = 0.0
    old_images_size_gb: float = 0.0
    old_images_monthly_cost: float = 0.0
    findings: list[ECRRepoFinding] = field(default_factory=list)


def collect_ecr_repos(session, account_id: str, account_name: str, region: str) -> list[ECRRepoInfo]:
    """ECR 리포지토리 및 이미지 정보를 수집한다.

    DescribeRepositories로 리포지토리 목록을 가져온 후,
    각 리포지토리의 라이프사이클 정책과 이미지 상세 정보를 조회한다.
    UNUSED_DAYS_THRESHOLD 이상 pull 없는 이미지를 오래된 것으로 분류한다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        ECRRepoInfo 목록.
    """
    from botocore.exceptions import ClientError

    ecr = get_client(session, "ecr", region_name=region)
    repos = []

    paginator = ecr.get_paginator("describe_repositories")
    for page in paginator.paginate():
        for repo in page.get("repositories", []):
            repo_info = ECRRepoInfo(
                account_id=account_id,
                account_name=account_name,
                region=region,
                name=repo.get("repositoryName", ""),
                arn=repo.get("repositoryArn", ""),
                uri=repo.get("repositoryUri", ""),
                created_at=repo.get("createdAt"),
            )

            # 라이프사이클 정책 확인
            try:
                ecr.get_lifecycle_policy(repositoryName=repo_info.name)
                repo_info.has_lifecycle_policy = True
            except ecr.exceptions.LifecyclePolicyNotFoundException:
                repo_info.has_lifecycle_policy = False
            except ClientError:
                pass

            # 이미지 정보 수집
            try:
                img_paginator = ecr.get_paginator("describe_images")
                now = datetime.now(timezone.utc)
                threshold_date = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

                for img_page in img_paginator.paginate(repositoryName=repo_info.name):
                    for img in img_page.get("imageDetails", []):
                        repo_info.image_count += 1
                        size = img.get("imageSizeInBytes", 0)
                        repo_info.total_size_bytes += size

                        pushed_at = img.get("imagePushedAt")
                        last_pull = img.get("lastRecordedPullTime")

                        # 마지막 pull이 90일 이상 전이거나 push 후 한번도 pull 안 된 경우
                        is_old = False
                        if (
                            last_pull
                            and last_pull < threshold_date
                            or not last_pull
                            and pushed_at
                            and pushed_at < threshold_date
                        ):
                            is_old = True

                        if is_old:
                            repo_info.old_image_count += 1
                            repo_info.old_images_size_bytes += size

            except ClientError:
                pass

            repos.append(repo_info)

    return repos


def analyze_ecr_repos(repos: list[ECRRepoInfo], account_id: str, account_name: str, region: str) -> ECRAnalysisResult:
    """ECR 리포지토리를 이미지 사용 현황 기준으로 분석한다.

    빈 리포지토리, 오래된 이미지 보유, 라이프사이클 정책 미설정을 탐지한다.

    Args:
        repos: 분석할 ECR 리포지토리 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        분석 결과 집계 (ECRAnalysisResult).
    """
    result = ECRAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_repos=len(repos),
    )

    for repo in repos:
        result.total_images += repo.image_count
        result.old_images += repo.old_image_count
        result.total_size_gb += repo.total_size_gb
        result.old_images_size_gb += repo.old_images_size_gb
        result.old_images_monthly_cost += repo.old_images_monthly_cost

        # 빈 리포지토리
        if repo.image_count == 0:
            result.empty_repos += 1
            result.findings.append(
                ECRRepoFinding(
                    repo=repo,
                    status=ECRRepoStatus.EMPTY,
                    recommendation="빈 리포지토리 - 삭제 검토",
                )
            )
            continue

        # 오래된 이미지가 있는 경우
        if repo.old_image_count > 0:
            result.repos_with_old_images += 1
            result.findings.append(
                ECRRepoFinding(
                    repo=repo,
                    status=ECRRepoStatus.OLD_IMAGES,
                    recommendation=f"{repo.old_image_count}개 오래된 이미지 ({repo.old_images_size_gb:.2f} GB)",
                )
            )

            # 라이프사이클 정책 없는 경우 추가 경고
            if not repo.has_lifecycle_policy:
                result.no_lifecycle_repos += 1
            continue

        # 라이프사이클 정책 없는 경우
        if not repo.has_lifecycle_policy:
            result.no_lifecycle_repos += 1
            result.findings.append(
                ECRRepoFinding(
                    repo=repo,
                    status=ECRRepoStatus.NO_LIFECYCLE,
                    recommendation="라이프사이클 정책 없음 - 설정 권장",
                )
            )
            continue

        result.findings.append(
            ECRRepoFinding(
                repo=repo,
                status=ECRRepoStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[ECRAnalysisResult], output_dir: str) -> str:
    """ECR 미사용 이미지 분석 Excel 보고서를 생성한다.

    Summary 시트(계정/리전별 집계)와 Repositories 시트(문제 리포지토리 상세)를 포함.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        저장된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Repos", width=10, style="number"),
        ColumnDef(header="빈 Repo", width=10, style="number"),
        ColumnDef(header="오래된 이미지", width=12, style="number"),
        ColumnDef(header="총 크기", width=12),
        ColumnDef(header="낭비 비용", width=12),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_repos,
                r.empty_repos,
                r.old_images,
                f"{r.total_size_gb:.2f} GB",
                f"${r.old_images_monthly_cost:,.2f}",
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.empty_repos > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.old_images > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Repository", width=35),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="이미지수", width=10, style="number"),
        ColumnDef(header="오래된", width=10, style="number"),
        ColumnDef(header="크기", width=12),
        ColumnDef(header="낭비", width=12),
        ColumnDef(header="Lifecycle", width=10),
        ColumnDef(header="권장 조치", width=35),
    ]
    detail_sheet = wb.new_sheet("Repositories", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != ECRRepoStatus.NORMAL:
                repo = f.repo
                style = Styles.danger() if f.status == ECRRepoStatus.EMPTY else Styles.warning()

                detail_sheet.add_row(
                    [
                        repo.account_name,
                        repo.region,
                        repo.name,
                        f.status.value,
                        repo.image_count,
                        repo.old_image_count,
                        f"{repo.total_size_gb:.2f} GB",
                        f"${repo.old_images_monthly_cost:.2f}",
                        "있음" if repo.has_lifecycle_policy else "없음",
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "ECR_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ECRAnalysisResult | None:
    """단일 계정/리전의 ECR 리포지토리를 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 병렬로 실행된다.

    Args:
        session: boto3 Session.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        분석 결과. 리포지토리가 없으면 None.
    """
    repos = collect_ecr_repos(session, account_id, account_name, region)
    if not repos:
        return None
    return analyze_ecr_repos(repos, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """ECR 미사용 이미지 분석 도구의 진입점.

    멀티 계정/리전 병렬 수집 후 콘솔 요약 출력, Excel 보고서를 생성한다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 리전, 출력 설정 포함).
    """
    console.print("[bold]ECR 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="ecr")
    results: list[ECRAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_old = sum(r.old_images for r in results)
    total_cost = sum(r.old_images_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"오래된 이미지: [yellow]{total_old}개[/yellow] (${total_cost:,.2f}/월)")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("ecr", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
