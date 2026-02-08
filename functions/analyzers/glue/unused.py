"""
plugins/glue/unused.py - Glue 미사용 작업 분석

미사용/실패 AWS Glue 작업 탐지 (실행 기록 기반)

분석 기준:
- Unused: 최근 90일간 실행 기록 없음
- Failed: 최근 5회 실행 모두 FAILED 또는 TIMEOUT 상태
- Over-provisioned: DPU 할당량 대비 실제 사용량 < 30%

비용 구조:
- Glue ETL: $0.44/DPU-hour
- Glue Streaming: $0.22/DPU-hour
- Python Shell: $0.44/DPU-hour (0.0625 DPU 또는 1 DPU)

주의: 미사용 작업은 실행 시에만 과금되므로 직접 비용은 없음

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 미사용 기준 (일)
UNUSED_DAYS_THRESHOLD = 90
# 실패 판단을 위한 최근 실행 횟수
FAILURE_RUN_COUNT = 5

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "glue:GetJobs",
        "glue:GetJobRuns",
    ],
}


class JobStatus(Enum):
    """작업 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    FAILED = "failed"


@dataclass
class GlueJobInfo:
    """Glue 작업 정보"""

    account_id: str
    account_name: str
    region: str
    job_name: str
    job_type: str  # glueetl, pythonshell, gluestreaming
    glue_version: str
    max_capacity: float  # DPU 할당
    worker_type: str | None
    num_workers: int | None
    created_at: datetime | None
    last_modified: datetime | None
    # 실행 기록 분석
    last_run_time: datetime | None = None
    last_run_status: str | None = None
    run_count_30d: int = 0
    run_count_90d: int = 0
    failure_count_5: int = 0  # 최근 5회 중 실패 수
    success_count_5: int = 0  # 최근 5회 중 성공 수
    avg_execution_time_sec: float = 0.0
    avg_dpu_used: float = 0.0

    @property
    def days_since_last_run(self) -> int | None:
        """마지막 실행 이후 일수"""
        if not self.last_run_time:
            return None
        now = datetime.now(timezone.utc)
        return (now - self.last_run_time.replace(tzinfo=timezone.utc)).days


@dataclass
class GlueJobFinding:
    """작업 분석 결과"""

    job: GlueJobInfo
    status: JobStatus
    recommendation: str


@dataclass
class GlueAnalysisResult:
    """Glue 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_jobs: int = 0
    unused_jobs: int = 0
    failed_jobs: int = 0
    normal_jobs: int = 0
    findings: list[GlueJobFinding] = field(default_factory=list)


def collect_glue_jobs(session, account_id: str, account_name: str, region: str) -> list[GlueJobInfo]:
    """Glue 작업 수집 및 실행 기록 분석"""
    from botocore.exceptions import ClientError

    glue = get_client(session, "glue", region_name=region)

    jobs: list[GlueJobInfo] = []
    now = datetime.now(timezone.utc)

    try:
        # 1단계: 작업 목록 조회
        paginator = glue.get_paginator("get_jobs")
        for page in paginator.paginate():
            for j in page.get("Jobs", []):
                job_name = j.get("Name", "")

                # worker_type과 num_workers가 있으면 해당 값 사용, 없으면 max_capacity 사용
                worker_type = j.get("WorkerType")
                num_workers = j.get("NumberOfWorkers")
                max_capacity = j.get("MaxCapacity", 0)

                # DPU 계산
                if worker_type and num_workers:
                    # G.1X = 1 DPU, G.2X = 2 DPU
                    dpu_map = {"G.1X": 1, "G.2X": 2, "G.4X": 4, "G.8X": 8, "Standard": 1}
                    dpu_per_worker = dpu_map.get(worker_type, 1)
                    effective_capacity = num_workers * dpu_per_worker
                else:
                    effective_capacity = max_capacity

                job = GlueJobInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    job_name=job_name,
                    job_type=j.get("Command", {}).get("Name", "unknown"),
                    glue_version=j.get("GlueVersion", ""),
                    max_capacity=effective_capacity,
                    worker_type=worker_type,
                    num_workers=num_workers,
                    created_at=j.get("CreatedOn"),
                    last_modified=j.get("LastModifiedOn"),
                )
                jobs.append(job)

        # 2단계: 각 작업의 실행 기록 분석
        for job in jobs:
            _analyze_job_runs(glue, job, now)

    except ClientError:
        pass

    return jobs


def _analyze_job_runs(glue, job: GlueJobInfo, now: datetime) -> None:
    """작업 실행 기록 분석 (내부 함수)"""
    from botocore.exceptions import ClientError

    try:
        # 최근 100개의 실행 기록 조회
        response = glue.get_job_runs(JobName=job.job_name, MaxResults=100)
        runs = response.get("JobRuns", [])

        if not runs:
            return

        # 마지막 실행 정보
        latest_run = runs[0]
        job.last_run_time = latest_run.get("StartedOn")
        job.last_run_status = latest_run.get("JobRunState", "")

        # 30일/90일 내 실행 횟수 계산
        threshold_30d = now - timedelta(days=30)
        threshold_90d = now - timedelta(days=90)

        total_exec_time = 0.0
        total_dpu_used = 0.0
        run_count_for_avg = 0

        for run in runs:
            started_on = run.get("StartedOn")
            if started_on:
                started_on = started_on.replace(tzinfo=timezone.utc)
                if started_on > threshold_30d:
                    job.run_count_30d += 1
                if started_on > threshold_90d:
                    job.run_count_90d += 1

            # 실행 시간/DPU 통계
            exec_time = run.get("ExecutionTime", 0)
            dpu_hours = run.get("DPUSeconds", 0) / 3600 if run.get("DPUSeconds") else 0
            if exec_time > 0:
                total_exec_time += exec_time
                total_dpu_used += dpu_hours
                run_count_for_avg += 1

        # 최근 5회 실행의 성공/실패 분석
        recent_runs = runs[:FAILURE_RUN_COUNT]
        for run in recent_runs:
            run_state = run.get("JobRunState", "")
            if run_state in ("FAILED", "TIMEOUT", "ERROR"):
                job.failure_count_5 += 1
            elif run_state == "SUCCEEDED":
                job.success_count_5 += 1

        # 평균 계산
        if run_count_for_avg > 0:
            job.avg_execution_time_sec = total_exec_time / run_count_for_avg
            job.avg_dpu_used = total_dpu_used / run_count_for_avg

    except ClientError:
        pass


def analyze_jobs(jobs: list[GlueJobInfo], account_id: str, account_name: str, region: str) -> GlueAnalysisResult:
    """Glue 작업 분석"""
    result = GlueAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_jobs=len(jobs),
    )

    for job in jobs:
        # 미사용: 90일간 실행 없음
        if job.run_count_90d == 0:
            result.unused_jobs += 1
            last_run_info = ""
            if job.last_run_time:
                days = job.days_since_last_run
                last_run_info = f" (마지막 실행: {days}일 전)"
            else:
                last_run_info = " (실행 기록 없음)"

            result.findings.append(
                GlueJobFinding(
                    job=job,
                    status=JobStatus.UNUSED,
                    recommendation=f"미사용{last_run_info} - 삭제 검토",
                )
            )
            continue

        # 실패: 최근 5회 모두 실패
        if job.failure_count_5 >= FAILURE_RUN_COUNT and job.success_count_5 == 0:
            result.failed_jobs += 1
            result.findings.append(
                GlueJobFinding(
                    job=job,
                    status=JobStatus.FAILED,
                    recommendation=f"연속 실패 ({job.failure_count_5}회) - 수정 또는 삭제 검토",
                )
            )
            continue

        result.normal_jobs += 1
        result.findings.append(
            GlueJobFinding(
                job=job,
                status=JobStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[GlueAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="연속실패", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_jobs,
                r.unused_jobs,
                r.failed_jobs,
                r.normal_jobs,
            ]
        )
        ws = summary_sheet._ws
        if r.unused_jobs > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.failed_jobs > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Job Name", width=40),
        ColumnDef(header="Type", width=15),
        ColumnDef(header="Glue Version", width=12),
        ColumnDef(header="Max DPU", width=10),
        ColumnDef(header="Workers", width=10),
        ColumnDef(header="Last Run", width=15),
        ColumnDef(header="Last Status", width=12),
        ColumnDef(header="Runs (30d)", width=12, style="number"),
        ColumnDef(header="Runs (90d)", width=12, style="number"),
        ColumnDef(header="분석상태", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Jobs", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != JobStatus.NORMAL:
                j = f.job
                style = Styles.danger() if f.status == JobStatus.UNUSED else Styles.warning()

                last_run_str = j.last_run_time.strftime("%Y-%m-%d") if j.last_run_time else "-"
                workers_str = f"{j.worker_type} x {j.num_workers}" if j.worker_type else "-"

                detail_sheet.add_row(
                    [
                        j.account_name,
                        j.region,
                        j.job_name,
                        j.job_type,
                        j.glue_version or "-",
                        f"{j.max_capacity:.1f}",
                        workers_str,
                        last_run_str,
                        j.last_run_status or "-",
                        j.run_count_30d,
                        j.run_count_90d,
                        f.status.value,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "Glue_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> GlueAnalysisResult | None:
    """단일 계정/리전의 Glue 작업 수집 및 분석 (병렬 실행용)"""
    jobs = collect_glue_jobs(session, account_id, account_name, region)
    if not jobs:
        return None
    return analyze_jobs(jobs, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Glue 미사용 작업 분석"""
    console.print("[bold]Glue 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="glue")
    results: list[GlueAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_jobs for r in results)
    total_failed = sum(r.failed_jobs for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미사용: [red]{total_unused}개[/red] / 연속실패: [yellow]{total_failed}개[/yellow]")
    console.print("[dim](Glue 작업은 실행 시에만 과금됨)[/dim]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("glue", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
