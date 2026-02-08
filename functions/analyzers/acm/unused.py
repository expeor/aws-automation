"""
functions/analyzers/acm/unused.py - ACM 미사용 인증서 분석

미사용/만료 임박 인증서 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "acm:ListCertificates",
        "acm:DescribeCertificate",
    ],
}

# 만료 임박 기준 (AWS Security Hub 기준)
CRITICAL_EXPIRING_DAYS = 7  # 7일 이내 = 긴급
EXPIRING_DAYS_THRESHOLD = 30  # 30일 이내 = 주의
WARNING_EXPIRING_DAYS = 60  # 60일 이내 = 경고


class CertStatus(Enum):
    """ACM 인증서 분석 상태.

    인증서 만료 임박 기준은 AWS Security Hub 권장 기준을 따른다.

    Attributes:
        NORMAL: 정상 사용 중인 인증서.
        UNUSED: 어떤 리소스에도 연결되지 않은 미사용 인증서.
        CRITICAL_EXPIRING: 7일 이내 만료 예정 (긴급).
        EXPIRING: 30일 이내 만료 예정 (주의).
        WARNING_EXPIRING: 60일 이내 만료 예정 (경고).
        EXPIRED: 이미 만료된 인증서.
        PENDING: 검증 대기 중인 인증서.
    """

    NORMAL = "normal"
    UNUSED = "unused"
    CRITICAL_EXPIRING = "critical_expiring"  # 7일 이내
    EXPIRING = "expiring"  # 30일 이내
    WARNING_EXPIRING = "warning_expiring"  # 60일 이내
    EXPIRED = "expired"
    PENDING = "pending"


@dataclass
class CertInfo:
    """ACM 인증서 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 인증서가 위치한 리전.
        certificate_arn: 인증서 ARN.
        domain_name: 인증서 도메인 이름.
        status: ACM 인증서 상태 (예: ISSUED, PENDING_VALIDATION).
        cert_type: 인증서 유형 (AMAZON_ISSUED 또는 IMPORTED).
        key_algorithm: 키 알고리즘 (예: RSA_2048, EC_prime256v1).
        in_use_by: 인증서를 사용 중인 리소스 ARN 목록.
        not_before: 인증서 유효 시작일.
        not_after: 인증서 만료일.
        renewal_eligibility: 갱신 가능 여부.
    """

    account_id: str
    account_name: str
    region: str
    certificate_arn: str
    domain_name: str
    status: str
    cert_type: str  # AMAZON_ISSUED, IMPORTED
    key_algorithm: str
    in_use_by: list[str]
    not_before: datetime | None
    not_after: datetime | None
    renewal_eligibility: str

    @property
    def is_in_use(self) -> bool:
        """인증서가 하나 이상의 리소스에 연결되어 있는지 여부.

        Returns:
            리소스에 연결되어 있으면 True.
        """
        return len(self.in_use_by) > 0

    @property
    def days_until_expiry(self) -> int | None:
        """인증서 만료까지 남은 일수.

        Returns:
            만료까지 남은 일수. 만료일이 없으면 None.
        """
        if self.not_after:
            now = datetime.now(timezone.utc)
            delta = self.not_after - now
            return delta.days
        return None


@dataclass
class CertFinding:
    """개별 인증서에 대한 분석 결과.

    Attributes:
        cert: 분석 대상 인증서 정보.
        status: 분석 결과 상태.
        recommendation: 권장 조치 사항 (한글).
    """

    cert: CertInfo
    status: CertStatus
    recommendation: str


@dataclass
class ACMAnalysisResult:
    """단일 계정/리전의 ACM 인증서 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.
        total_certs: 전체 인증서 수.
        unused_certs: 미사용 인증서 수.
        critical_expiring_certs: 7일 이내 만료 예정 인증서 수.
        expiring_certs: 30일 이내 만료 예정 인증서 수.
        warning_expiring_certs: 60일 이내 만료 예정 인증서 수.
        expired_certs: 만료된 인증서 수.
        pending_certs: 검증 대기 중인 인증서 수.
        normal_certs: 정상 인증서 수.
        findings: 개별 인증서 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_certs: int = 0
    unused_certs: int = 0
    critical_expiring_certs: int = 0  # 7일 이내
    expiring_certs: int = 0  # 30일 이내
    warning_expiring_certs: int = 0  # 60일 이내
    expired_certs: int = 0
    pending_certs: int = 0
    normal_certs: int = 0
    findings: list[CertFinding] = field(default_factory=list)


def collect_certificates(session, account_id: str, account_name: str, region: str) -> list[CertInfo]:
    """지정된 계정/리전의 ACM 인증서를 수집한다.

    모든 키 유형의 인증서를 조회하고 상세 정보(도메인, 만료일, 사용 현황 등)를 수집한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        수집된 인증서 정보 목록.
    """
    from botocore.exceptions import ClientError

    acm = get_client(session, "acm", region_name=region)
    certs = []

    try:
        paginator = acm.get_paginator("list_certificates")
        for page in paginator.paginate(
            Includes={
                "keyTypes": [
                    "RSA_1024",
                    "RSA_2048",
                    "RSA_3072",
                    "RSA_4096",
                    "EC_prime256v1",
                    "EC_secp384r1",
                    "EC_secp521r1",
                ]
            }
        ):
            for cert_summary in page.get("CertificateSummaryList", []):
                cert_arn = cert_summary.get("CertificateArn", "")

                try:
                    cert_detail = acm.describe_certificate(CertificateArn=cert_arn).get("Certificate", {})

                    info = CertInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        certificate_arn=cert_arn,
                        domain_name=cert_detail.get("DomainName", ""),
                        status=cert_detail.get("Status", ""),
                        cert_type=cert_detail.get("Type", ""),
                        key_algorithm=cert_detail.get("KeyAlgorithm", ""),
                        in_use_by=cert_detail.get("InUseBy", []),
                        not_before=cert_detail.get("NotBefore"),
                        not_after=cert_detail.get("NotAfter"),
                        renewal_eligibility=cert_detail.get("RenewalEligibility", ""),
                    )
                    certs.append(info)

                except ClientError:
                    continue

    except ClientError:
        pass

    return certs


def analyze_certificates(certs: list[CertInfo], account_id: str, account_name: str, region: str) -> ACMAnalysisResult:
    """수집된 ACM 인증서를 분석하여 미사용/만료 임박 인증서를 식별한다.

    AWS Security Hub 기준에 따라 만료 임박 등급을 분류한다:
    7일 이내(CRITICAL), 30일 이내(EXPIRING), 60일 이내(WARNING).

    Args:
        certs: 분석 대상 인증서 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.

    Returns:
        인증서 분석 결과 집계 객체.
    """
    result = ACMAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_certs=len(certs),
    )

    now = datetime.now(timezone.utc)

    for cert in certs:
        # 발급 대기 중
        if cert.status == "PENDING_VALIDATION":
            result.pending_certs += 1
            result.findings.append(
                CertFinding(
                    cert=cert,
                    status=CertStatus.PENDING,
                    recommendation="검증 대기 중 - 오래된 경우 삭제 검토",
                )
            )
            continue

        # 만료됨
        if cert.not_after and cert.not_after < now:
            result.expired_certs += 1
            result.findings.append(
                CertFinding(
                    cert=cert,
                    status=CertStatus.EXPIRED,
                    recommendation="만료됨 - 삭제 검토",
                )
            )
            continue

        # 만료 임박 체크 (AWS Security Hub 기준)
        days_left = cert.days_until_expiry
        if days_left is not None:
            # 7일 이내 = 긴급
            if days_left <= CRITICAL_EXPIRING_DAYS:
                result.critical_expiring_certs += 1
                result.findings.append(
                    CertFinding(
                        cert=cert,
                        status=CertStatus.CRITICAL_EXPIRING,
                        recommendation=f"긴급! 만료 {days_left}일 남음 - 즉시 갱신 필요",
                    )
                )
                continue

            # 30일 이내 = 주의
            if days_left <= EXPIRING_DAYS_THRESHOLD:
                result.expiring_certs += 1
                result.findings.append(
                    CertFinding(
                        cert=cert,
                        status=CertStatus.EXPIRING,
                        recommendation=f"만료 임박 ({days_left}일 남음) - 갱신 필요",
                    )
                )
                continue

            # 60일 이내 = 경고
            if days_left <= WARNING_EXPIRING_DAYS:
                result.warning_expiring_certs += 1
                result.findings.append(
                    CertFinding(
                        cert=cert,
                        status=CertStatus.WARNING_EXPIRING,
                        recommendation=f"만료 예정 ({days_left}일 남음) - 갱신 준비",
                    )
                )
                continue

        # 미사용
        if not cert.is_in_use:
            result.unused_certs += 1
            result.findings.append(
                CertFinding(
                    cert=cert,
                    status=CertStatus.UNUSED,
                    recommendation="미사용 - 삭제 검토",
                )
            )
            continue

        result.normal_certs += 1
        result.findings.append(
            CertFinding(
                cert=cert,
                status=CertStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[ACMAnalysisResult], output_dir: str) -> str:
    """ACM 분석 결과를 Excel 보고서로 생성한다.

    Summary 시트(계정/리전별 통계)와 Detail 시트(비정상 인증서 상세)를 포함한다.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="만료임박", width=10, style="number"),
        ColumnDef(header="만료됨", width=10, style="number"),
        ColumnDef(header="대기중", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_certs,
                r.unused_certs,
                r.expiring_certs,
                r.expired_certs,
                r.pending_certs,
                r.normal_certs,
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.unused_certs > 0:
            ws.cell(row=row_num, column=4).fill = yellow_fill
        if r.expiring_certs > 0:
            ws.cell(row=row_num, column=5).fill = orange_fill
        if r.expired_certs > 0:
            ws.cell(row=row_num, column=6).fill = red_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Domain", width=30),
        ColumnDef(header="Type", width=15),
        ColumnDef(header="Status", width=15),
        ColumnDef(header="Expiry", width=12),
        ColumnDef(header="Days Left", width=10, style="number"),
        ColumnDef(header="In Use", width=10, style="number"),
        ColumnDef(header="분석상태", width=12),
        ColumnDef(header="권장 조치", width=25),
    ]
    detail_sheet = wb.new_sheet("Certificates", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != CertStatus.NORMAL:
                c = f.cert
                style = None
                if f.status == CertStatus.EXPIRED:
                    style = Styles.danger()
                elif f.status == CertStatus.EXPIRING:
                    style = Styles.warning()

                detail_sheet.add_row(
                    [
                        c.account_name,
                        c.region,
                        c.domain_name,
                        c.cert_type,
                        c.status,
                        c.not_after.strftime("%Y-%m-%d") if c.not_after else "-",
                        c.days_until_expiry if c.days_until_expiry else "-",
                        len(c.in_use_by),
                        f.status.value,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "ACM_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ACMAnalysisResult | None:
    """단일 계정/리전의 ACM 인증서를 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 인증서가 없으면 None을 반환한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 대상 리전.

    Returns:
        분석 결과 객체. 인증서가 없으면 None.
    """
    certs = collect_certificates(session, account_id, account_name, region)
    if not certs:
        return None
    return analyze_certificates(certs, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """ACM 미사용/만료 임박 인증서 분석 도구의 메인 실행 함수.

    멀티 계정/리전 병렬 수집 후 결과를 집계하고, Excel 보고서를 생성하여 출력 디렉토리에 저장한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 계정/리전 목록, 옵션 등 포함).
    """
    console.print("[bold]ACM 인증서 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="acm")
    results: list[ACMAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_certs for r in results)
    total_expiring = sum(r.expiring_certs for r in results)
    total_expired = sum(r.expired_certs for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [yellow]{total_unused}개[/yellow] / "
        f"만료임박: [orange1]{total_expiring}개[/orange1] / "
        f"만료: [red]{total_expired}개[/red]"
    )

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("acm", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
