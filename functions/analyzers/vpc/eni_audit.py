"""
functions/analyzers/vpc/eni_audit.py - ENI 미사용 분석

미사용 ENI (Elastic Network Interface) 탐지

분석 기준:
- Status가 "available"인 ENI (아무것도 연결되지 않음)
- AWS 관리형 ENI는 제외 (NAT Gateway, Lambda, VPC Endpoint 등)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeNetworkInterfaces",
    ],
}


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """ENI의 사용 상태 분류

    연결 상태 및 인터페이스 유형으로 판별됩니다.
    """

    UNUSED = "unused"  # 미사용 (available 상태)
    NORMAL = "normal"  # 정상 사용 (in-use)
    PENDING = "pending"  # 확인 필요
    AWS_MANAGED = "aws_managed"  # AWS 관리형 (삭제 불가)


class Severity(Enum):
    """분석 결과의 심각도 수준"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class ENIInfo:
    """Elastic Network Interface 정보

    Attributes:
        id: ENI ID
        description: ENI 설명
        status: 연결 상태 (available, in-use, associated)
        vpc_id: VPC ID
        subnet_id: 서브넷 ID
        availability_zone: 가용 영역
        private_ip: 프라이빗 IP 주소
        public_ip: 퍼블릭 IP 주소
        interface_type: 인터페이스 유형 (interface, nat_gateway, lambda 등)
        requester_id: 요청자 ID (AWS 서비스)
        owner_id: 소유자 계정 ID
        instance_id: 연결된 EC2 인스턴스 ID
        attachment_status: attachment 상태
        security_groups: 연결된 Security Group ID 리스트
        tags: 사용자 태그 딕셔너리 (aws: 접두사 제외)
        name: Name 태그 값
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
    """

    id: str
    description: str
    status: str
    vpc_id: str
    subnet_id: str
    availability_zone: str
    private_ip: str
    public_ip: str
    interface_type: str
    requester_id: str
    owner_id: str
    instance_id: str
    attachment_status: str
    security_groups: list[str]
    tags: dict[str, str]
    name: str

    # 메타
    account_id: str
    account_name: str
    region: str

    @property
    def is_attached(self) -> bool:
        """EC2 인스턴스 연결 여부

        Returns:
            status가 in-use이면 True
        """
        return self.status == "in-use"

    @property
    def is_aws_managed(self) -> bool:
        """AWS 관리형 ENI 여부

        Returns:
            requester_id가 다른 서비스이거나 관리형 인터페이스 유형이면 True
        """
        if self.requester_id and self.requester_id != self.owner_id:
            return True
        aws_managed_types = {
            "nat_gateway",
            "gateway_load_balancer",
            "gateway_load_balancer_endpoint",
            "vpc_endpoint",
            "efa",
            "trunk",
            "load_balancer",
            "lambda",
        }
        return self.interface_type in aws_managed_types


@dataclass
class ENIFinding:
    """개별 ENI 분석 결과

    Attributes:
        eni: 분석 대상 ENI 정보
        usage_status: 분석으로 판별된 사용 상태
        severity: 심각도 수준
        description: 상태 설명
        recommendation: 권장 조치 사항
    """

    eni: ENIInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class ENIAnalysisResult:
    """ENI 분석 결과 집계 (계정/리전별)

    Attributes:
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전
        findings: 개별 ENI 분석 결과 리스트
        total_count: 전체 ENI 수
        unused_count: 미사용 ENI 수
        normal_count: 정상 사용 ENI 수
        aws_managed_count: AWS 관리형 ENI 수
        pending_count: 확인 필요 ENI 수
    """

    account_id: str
    account_name: str
    region: str
    findings: list[ENIFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    normal_count: int = 0
    aws_managed_count: int = 0
    pending_count: int = 0


# =============================================================================
# 수집
# =============================================================================


def collect_enis(session, account_id: str, account_name: str, region: str) -> list[ENIInfo]:
    """ENI 목록 수집

    DescribeNetworkInterfaces API로 모든 ENI를 페이지네이션으로 수집합니다.

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        ENI 정보 리스트
    """
    from botocore.exceptions import ClientError

    enis = []

    try:
        ec2 = get_client(session, "ec2", region_name=region)
        paginator = ec2.get_paginator("describe_network_interfaces")

        for page in paginator.paginate():
            for data in page.get("NetworkInterfaces", []):
                attachment = data.get("Attachment")

                # 태그 파싱
                tags = {
                    t.get("Key", ""): t.get("Value", "")
                    for t in data.get("TagSet", [])
                    if not t.get("Key", "").startswith("aws:")
                }

                eni = ENIInfo(
                    id=data.get("NetworkInterfaceId", ""),
                    description=data.get("Description", ""),
                    status=data.get("Status", ""),
                    vpc_id=data.get("VpcId", ""),
                    subnet_id=data.get("SubnetId", ""),
                    availability_zone=data.get("AvailabilityZone", ""),
                    private_ip=data.get("PrivateIpAddress", ""),
                    public_ip=data.get("Association", {}).get("PublicIp", "") if data.get("Association") else "",
                    interface_type=data.get("InterfaceType", ""),
                    requester_id=data.get("RequesterId", ""),
                    owner_id=data.get("OwnerId", ""),
                    instance_id=attachment.get("InstanceId", "") if attachment else "",
                    attachment_status=attachment.get("Status", "") if attachment else "",
                    security_groups=[g.get("GroupId", "") for g in data.get("Groups", [])],
                    tags=tags,
                    name=tags.get("Name", ""),
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                )
                enis.append(eni)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} ENI 수집 오류: {error_code}[/yellow]")

    return enis


# =============================================================================
# 분석
# =============================================================================


def analyze_enis(enis: list[ENIInfo], account_id: str, account_name: str, region: str) -> ENIAnalysisResult:
    """ENI 사용 상태 분석

    각 ENI의 연결 상태 및 인터페이스 유형을 기반으로
    미사용/정상/AWS관리형/확인필요로 분류합니다.

    Args:
        enis: 수집된 ENI 정보 리스트
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        계정/리전별 분석 결과 (상태별 ENI 수 포함)
    """
    result = ENIAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for eni in enis:
        finding = _analyze_single_eni(eni)
        result.findings.append(finding)

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
        elif finding.usage_status == UsageStatus.NORMAL:
            result.normal_count += 1
        elif finding.usage_status == UsageStatus.AWS_MANAGED:
            result.aws_managed_count += 1
        elif finding.usage_status == UsageStatus.PENDING:
            result.pending_count += 1

    result.total_count = len(enis)
    return result


def _analyze_single_eni(eni: ENIInfo) -> ENIFinding:
    """개별 ENI 사용 상태 분석

    연결 상태, 인터페이스 유형, description 패턴을 기반으로
    ENI의 사용 상태와 심각도를 판별합니다.

    Args:
        eni: 분석 대상 ENI 정보

    Returns:
        ENI 분석 결과 (상태, 심각도, 권장 조치 포함)
    """

    # 1. AWS 관리형
    if eni.is_aws_managed:
        return ENIFinding(
            eni=eni,
            usage_status=UsageStatus.AWS_MANAGED,
            severity=Severity.INFO,
            description=f"AWS 관리형 ({eni.interface_type})",
            recommendation="삭제하지 마세요.",
        )

    # 2. 연결됨 = 정상
    if eni.is_attached:
        return ENIFinding(
            eni=eni,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"사용 중 ({eni.status})",
            recommendation="정상 사용 중",
        )

    # 3. Available = 미사용
    if eni.status == "available":
        desc_lower = eni.description.lower()

        # 특정 패턴은 주의
        if "efs" in desc_lower:
            return ENIFinding(
                eni=eni,
                usage_status=UsageStatus.PENDING,
                severity=Severity.LOW,
                description="EFS 관련 ENI",
                recommendation="EFS 마운트 타겟 확인",
            )

        if "elb" in desc_lower or "load" in desc_lower:
            return ENIFinding(
                eni=eni,
                usage_status=UsageStatus.PENDING,
                severity=Severity.LOW,
                description="Load Balancer 관련 ENI",
                recommendation="LB 상태 확인",
            )

        # 일반 미사용
        return ENIFinding(
            eni=eni,
            usage_status=UsageStatus.UNUSED,
            severity=Severity.HIGH,
            description="미사용 ENI (available, 연결 없음)",
            recommendation="삭제 검토",
        )

    # 4. 기타
    return ENIFinding(
        eni=eni,
        usage_status=UsageStatus.PENDING,
        severity=Severity.INFO,
        description=f"상태: {eni.status}",
        recommendation="상태 안정화 대기",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[ENIAnalysisResult], output_dir: str) -> str:
    """ENI 미사용 분석 Excel 보고서 생성

    Summary 시트와 미사용/확인필요 ENI Findings 시트를 포함합니다.

    Args:
        results: 계정/리전별 분석 결과 리스트
        output_dir: 출력 디렉토리 경로

    Returns:
        생성된 Excel 파일 경로
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 상태별 색상 정의 (cell-level 적용용)
    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.PENDING: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
        UsageStatus.AWS_MANAGED: PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid"),
    }

    # Summary 시트
    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "aws_managed": sum(r.aws_managed_count for r in results),
        "pending": sum(r.pending_count for r in results),
    }

    summary = wb.new_summary_sheet("Summary")
    summary.add_title("ENI 미사용 분석 보고서")
    summary.add_section("분석 결과")
    summary.add_item("전체 ENI", totals["total"])
    summary.add_item("미사용", totals["unused"], highlight="danger" if totals["unused"] > 0 else None)
    summary.add_item("정상 사용", totals["normal"])
    summary.add_item("AWS 관리형", totals["aws_managed"])
    summary.add_item("확인 필요", totals["pending"], highlight="warning" if totals["pending"] > 0 else None)

    # Findings 시트
    columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="ENI ID", width=25),
        ColumnDef(header="Name", width=20),
        ColumnDef(header="Status", width=12, style="center"),
        ColumnDef(header="Usage", width=12, style="center"),
        ColumnDef(header="Severity", width=10, style="center"),
        ColumnDef(header="Description", width=30),
        ColumnDef(header="Recommendation", width=20),
        ColumnDef(header="VPC ID", width=22),
        ColumnDef(header="Subnet ID", width=25),
        ColumnDef(header="Private IP", width=15),
        ColumnDef(header="Type", width=15),
    ]
    sheet = wb.new_sheet("Findings", columns)

    # 미사용/확인필요만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.UNUSED, UsageStatus.PENDING):
                all_findings.append(f)

    # 심각도순
    severity_order = {
        Severity.HIGH: 0,
        Severity.MEDIUM: 1,
        Severity.LOW: 2,
        Severity.INFO: 3,
    }
    all_findings.sort(key=lambda x: severity_order.get(x.severity, 9))

    for f in all_findings:
        eni = f.eni
        # 행 스타일 결정
        style = Styles.danger() if f.usage_status == UsageStatus.UNUSED else Styles.warning()

        row_num = sheet.add_row(
            [
                eni.account_name,
                eni.region,
                eni.id,
                eni.name,
                eni.status,
                f.usage_status.value,
                f.severity.value,
                f.description,
                f.recommendation,
                eni.vpc_id,
                eni.subnet_id,
                eni.private_ip,
                eni.interface_type,
            ],
            style=style,
        )

        # Usage 컬럼에 상태별 색상 적용
        fill = status_fills.get(f.usage_status)
        if fill:
            sheet._ws.cell(row=row_num, column=6).fill = fill

    return str(wb.save_as(output_dir, "ENI_Unused"))


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ENIAnalysisResult:
    """단일 계정/리전의 ENI 수집 및 분석 (parallel_collect 콜백)

    Args:
        session: boto3 Session 객체
        account_id: AWS 계정 ID
        account_name: AWS 계정 이름
        region: AWS 리전

    Returns:
        ENI 분석 결과 (상태별 ENI 수 포함)
    """
    enis = collect_enis(session, account_id, account_name, region)
    return analyze_enis(enis, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """ENI 미사용 분석 실행

    멀티 계정/리전에서 ENI를 병렬 수집하고,
    미사용 ENI를 식별하여 Excel 보고서를 생성합니다.

    Args:
        ctx: CLI 실행 컨텍스트 (인증, 계정/리전 선택, 출력 설정 포함)
    """
    console.print("[bold]ENI 미사용 분석 시작...[/bold]")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="ec2")

    all_results: list[ENIAnalysisResult] = result.get_data()

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not all_results:
        console.print("[yellow]분석할 ENI 없음[/yellow]")
        return

    # 요약
    totals = {
        "total": sum(r.total_count for r in all_results),
        "unused": sum(r.unused_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "aws_managed": sum(r.aws_managed_count for r in all_results),
        "pending": sum(r.pending_count for r in all_results),
    }

    console.print(f"\n[bold]전체 ENI: {totals['total']}개[/bold]")
    if totals["unused"] > 0:
        console.print(f"  [red bold]미사용: {totals['unused']}개[/red bold]")
    if totals["pending"] > 0:
        console.print(f"  [yellow]확인 필요: {totals['pending']}개[/yellow]")
    console.print(f"  [green]정상: {totals['normal']}개[/green]")
    console.print(f"  [dim]AWS 관리형: {totals['aws_managed']}개[/dim]")

    # 보고서
    console.print("\n[#FF9900]Excel 보고서 생성 중...[/#FF9900]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("vpc", "network").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
