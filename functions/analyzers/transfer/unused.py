"""
functions/analyzers/transfer/unused.py - AWS Transfer Family 미사용 서버 분석

유휴/미사용 Transfer Family 서버 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from core.shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from core.shared.aws.pricing.transfer import get_transfer_monthly_cost
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "transfer:ListServers",
        "transfer:DescribeServer",
        "transfer:ListUsers",
        "cloudwatch:GetMetricStatistics",
    ],
}

# 분석 기간 (일) - CloudFix 권장 30일
ANALYSIS_DAYS = 30


class ServerStatus(Enum):
    """Transfer Family 서버 분석 상태.

    CloudWatch 파일 전송 지표, 사용자 수, 서버 상태 기반으로 분류한다.

    Attributes:
        NORMAL: 정상 사용 중인 서버.
        UNUSED: 분석 기간 동안 파일 전송이 없는 미사용 서버.
        IDLE: 일 평균 파일 전송이 1건 미만인 저사용 서버.
        NO_USERS: 등록된 사용자가 없는 서버.
        STOPPED: 중지 상태인 서버 (OFFLINE, STOPPING, STOP_FAILED).
    """

    NORMAL = "normal"
    UNUSED = "unused"
    IDLE = "idle"
    NO_USERS = "no_users"
    STOPPED = "stopped"


@dataclass
class TransferServerInfo:
    """Transfer Family 서버 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 서버가 위치한 리전.
        server_id: Transfer 서버 ID.
        endpoint_type: 엔드포인트 유형 (PUBLIC, VPC, VPC_ENDPOINT).
        protocols: 지원 프로토콜 목록 (SFTP, FTPS, FTP, AS2).
        state: 서버 상태 (ONLINE, OFFLINE, STARTING, STOPPING, START_FAILED, STOP_FAILED).
        identity_provider_type: 인증 제공자 유형.
        user_count: 등록된 사용자 수.
        files_in: 분석 기간 동안 수신 파일 수 (CloudWatch Sum).
        files_out: 분석 기간 동안 송신 파일 수 (CloudWatch Sum).
        bytes_in: 분석 기간 동안 수신 바이트 수 (CloudWatch Sum).
        bytes_out: 분석 기간 동안 송신 바이트 수 (CloudWatch Sum).
        estimated_monthly_cost: 예상 월간 비용 (USD).
    """

    account_id: str
    account_name: str
    region: str
    server_id: str
    endpoint_type: str  # PUBLIC, VPC, VPC_ENDPOINT
    protocols: list[str]  # SFTP, FTPS, FTP, AS2
    state: str  # ONLINE, OFFLINE, STARTING, STOPPING, START_FAILED, STOP_FAILED
    identity_provider_type: str
    user_count: int = 0
    # CloudWatch 지표
    files_in: float = 0.0
    files_out: float = 0.0
    bytes_in: float = 0.0
    bytes_out: float = 0.0
    # 비용
    estimated_monthly_cost: float = 0.0

    @property
    def total_files(self) -> float:
        """분석 기간 동안 총 파일 전송 수 (수신 + 송신).

        Returns:
            총 파일 수.
        """
        return self.files_in + self.files_out

    @property
    def total_bytes(self) -> float:
        """분석 기간 동안 총 전송 바이트 수 (수신 + 송신).

        Returns:
            총 바이트 수.
        """
        return self.bytes_in + self.bytes_out

    @property
    def is_active(self) -> bool:
        """파일 전송 활동이 있는지 여부.

        Returns:
            파일 전송 또는 바이트 전송이 있으면 True.
        """
        return self.total_files > 0 or self.total_bytes > 0


@dataclass
class TransferFinding:
    """개별 Transfer Family 서버에 대한 분석 결과.

    Attributes:
        server: 분석 대상 서버 정보.
        status: 분석 결과 상태.
        recommendation: 권장 조치 사항 (한글).
    """

    server: TransferServerInfo
    status: ServerStatus
    recommendation: str


@dataclass
class TransferAnalysisResult:
    """단일 계정/리전의 Transfer Family 서버 분석 결과 집계.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.
        total_servers: 전체 서버 수.
        unused_servers: 미사용 서버 수.
        idle_servers: 저사용 서버 수.
        no_users_servers: 사용자 없는 서버 수.
        stopped_servers: 중지된 서버 수.
        normal_servers: 정상 서버 수.
        total_monthly_waste: 미사용/저사용 서버의 월간 낭비 비용 합계 (USD).
        findings: 개별 서버 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_servers: int = 0
    unused_servers: int = 0
    idle_servers: int = 0
    no_users_servers: int = 0
    stopped_servers: int = 0
    normal_servers: int = 0
    total_monthly_waste: float = 0.0
    findings: list[TransferFinding] = field(default_factory=list)


def collect_servers(session, account_id: str, account_name: str, region: str) -> list[TransferServerInfo]:
    """지정된 계정/리전의 Transfer Family 서버를 수집하고 CloudWatch 메트릭을 조회한다.

    1단계에서 서버 목록, 상세 정보, 사용자 수를 수집하고,
    2단계에서 ONLINE 서버에 대해 batch_get_metrics를 통해 파일/바이트 전송 지표를 배치 조회한다.
    기존 서버당 4 API 호출을 전체 1-2 호출로 최적화했다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 조회 대상 리전.

    Returns:
        수집된 Transfer Family 서버 정보 목록. CloudWatch 메트릭과 비용 정보가 포함된다.
    """
    from botocore.exceptions import ClientError

    transfer = get_client(session, "transfer", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    servers: list[TransferServerInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 서버 목록 수집
        paginator = transfer.get_paginator("list_servers")
        for page in paginator.paginate():
            for server_summary in page.get("Servers", []):
                server_id = server_summary.get("ServerId", "")

                try:
                    # 서버 상세 정보
                    server_detail = transfer.describe_server(ServerId=server_id).get("Server", {})

                    protocols = server_detail.get("Protocols", ["SFTP"])
                    endpoint_type = server_detail.get("EndpointType", "PUBLIC")

                    # 사용자 수 확인
                    user_count = 0
                    try:
                        users_resp = transfer.list_users(ServerId=server_id)
                        user_count = len(users_resp.get("Users", []))
                    except ClientError as e:
                        category = categorize_error(e)
                        if category != ErrorCategory.NOT_FOUND:
                            logger.debug(f"Transfer 사용자 조회 실패: {server_id} ({get_error_code(e)})")

                    # 월 비용 계산 (pricing 모듈 사용)
                    monthly_cost = get_transfer_monthly_cost(region, protocols=protocols, session=session)

                    info = TransferServerInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        server_id=server_id,
                        endpoint_type=endpoint_type,
                        protocols=protocols,
                        state=server_detail.get("State", ""),
                        identity_provider_type=server_detail.get("IdentityProviderType", ""),
                        user_count=user_count,
                        estimated_monthly_cost=monthly_cost,
                    )
                    servers.append(info)

                except ClientError as e:
                    category = categorize_error(e)
                    if category != ErrorCategory.NOT_FOUND:
                        logger.debug(f"Transfer 서버 상세 조회 실패: {server_id} ({get_error_code(e)})")
                    continue

        # 2단계: ONLINE 상태인 서버만 배치 메트릭 조회
        online_servers = [s for s in servers if s.state == "ONLINE"]
        if online_servers:
            _collect_transfer_metrics_batch(cloudwatch, online_servers, start_time, now)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"Transfer 권한 없음: {account_name}/{region}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"Transfer 조회 오류: {get_error_code(e)}")

    return servers


def _collect_transfer_metrics_batch(
    cloudwatch,
    servers: list[TransferServerInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """Transfer Family 서버의 CloudWatch 메트릭을 배치로 수집하여 서버 객체에 반영한다.

    FilesIn, FilesOut, BytesIn, BytesOut 네 가지 메트릭의 Sum 값을
    일별(86400초) 기간으로 조회한다.

    Args:
        cloudwatch: CloudWatch boto3 클라이언트.
        servers: 메트릭을 수집할 서버 목록 (결과가 각 객체에 직접 반영됨).
        start_time: 메트릭 조회 시작 시각 (UTC).
        end_time: 메트릭 조회 종료 시각 (UTC).
    """
    from botocore.exceptions import ClientError

    queries: list[MetricQuery] = []

    for server in servers:
        safe_id = sanitize_metric_id(server.server_id)
        dimensions = {"ServerId": server.server_id}

        queries.extend(
            [
                MetricQuery(
                    id=f"{safe_id}_files_in",
                    namespace="AWS/Transfer",
                    metric_name="FilesIn",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_files_out",
                    namespace="AWS/Transfer",
                    metric_name="FilesOut",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_bytes_in",
                    namespace="AWS/Transfer",
                    metric_name="BytesIn",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_bytes_out",
                    namespace="AWS/Transfer",
                    metric_name="BytesOut",
                    dimensions=dimensions,
                    stat="Sum",
                ),
            ]
        )

    if not queries:
        return

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        for server in servers:
            safe_id = sanitize_metric_id(server.server_id)
            server.files_in = results.get(f"{safe_id}_files_in", 0.0)
            server.files_out = results.get(f"{safe_id}_files_out", 0.0)
            server.bytes_in = results.get(f"{safe_id}_bytes_in", 0.0)
            server.bytes_out = results.get(f"{safe_id}_bytes_out", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_servers(
    servers: list[TransferServerInfo], account_id: str, account_name: str, region: str
) -> TransferAnalysisResult:
    """수집된 Transfer Family 서버를 분석하여 미사용/저사용 서버를 식별한다.

    서버 상태, 사용자 수, 파일 전송 활동을 기반으로 STOPPED, NO_USERS,
    UNUSED, IDLE, NORMAL 상태로 분류한다. 비용 낭비는 미사용/사용자없음은 100%,
    저사용은 50%로 추정한다.

    Args:
        servers: 분석 대상 Transfer Family 서버 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.

    Returns:
        Transfer Family 서버 분석 결과 집계 객체.
    """
    result = TransferAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_servers=len(servers),
    )

    for server in servers:
        # 중지됨
        if server.state in ("OFFLINE", "STOPPING", "STOP_FAILED"):
            result.stopped_servers += 1
            result.findings.append(
                TransferFinding(
                    server=server,
                    status=ServerStatus.STOPPED,
                    recommendation="서버 중지됨 - 불필요시 삭제 검토",
                )
            )
            continue

        # 사용자 없음
        if server.user_count == 0:
            result.no_users_servers += 1
            result.total_monthly_waste += server.estimated_monthly_cost
            result.findings.append(
                TransferFinding(
                    server=server,
                    status=ServerStatus.NO_USERS,
                    recommendation="사용자 없음 - 삭제 검토",
                )
            )
            continue

        # 미사용 (파일 전송 없음)
        if not server.is_active:
            result.unused_servers += 1
            result.total_monthly_waste += server.estimated_monthly_cost
            result.findings.append(
                TransferFinding(
                    server=server,
                    status=ServerStatus.UNUSED,
                    recommendation=f"파일 전송 없음 ({ANALYSIS_DAYS}일간) - 삭제 검토",
                )
            )
            continue

        # 저사용 (하루 평균 1개 미만)
        avg_daily_files = server.total_files / ANALYSIS_DAYS
        if avg_daily_files < 1:
            result.idle_servers += 1
            result.total_monthly_waste += server.estimated_monthly_cost * 0.5  # 50% 낭비 추정
            result.findings.append(
                TransferFinding(
                    server=server,
                    status=ServerStatus.IDLE,
                    recommendation=f"저사용 (일 평균 {avg_daily_files:.1f}건) - 통합 검토",
                )
            )
            continue

        result.normal_servers += 1
        result.findings.append(
            TransferFinding(
                server=server,
                status=ServerStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[TransferAnalysisResult], output_dir: str) -> str:
    """Transfer Family 미사용 서버 분석 결과를 Excel 보고서로 생성한다.

    Summary 시트(계정/리전별 통계)와 Servers 시트(비정상 서버 상세)를 포함한다.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="사용자없음", width=12, style="number"),
        ColumnDef(header="중지됨", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="월 낭비(USD)", width=12, style="currency"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_servers,
                r.unused_servers,
                r.idle_servers,
                r.no_users_servers,
                r.stopped_servers,
                r.normal_servers,
                r.total_monthly_waste,
            ]
        )
        ws = summary_sheet._ws
        if r.unused_servers > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.idle_servers > 0 or r.no_users_servers > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Server ID", width=25),
        ColumnDef(header="Protocols", width=15),
        ColumnDef(header="Endpoint", width=15),
        ColumnDef(header="State", width=12),
        ColumnDef(header="Users", width=10, style="number"),
        ColumnDef(header="Files In", width=12, style="number"),
        ColumnDef(header="Files Out", width=12, style="number"),
        ColumnDef(header="상태", width=12),
        ColumnDef(header="월 비용(USD)", width=12, style="currency"),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Servers", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != ServerStatus.NORMAL:
                s = f.server
                style = Styles.danger() if f.status == ServerStatus.UNUSED else Styles.warning()
                detail_sheet.add_row(
                    [
                        s.account_name,
                        s.region,
                        s.server_id,
                        ", ".join(s.protocols),
                        s.endpoint_type,
                        s.state,
                        s.user_count,
                        int(s.files_in),
                        int(s.files_out),
                        f.status.value,
                        s.estimated_monthly_cost,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "Transfer_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> TransferAnalysisResult | None:
    """단일 계정/리전의 Transfer Family 서버를 수집하고 분석한다.

    parallel_collect 콜백으로 사용되며, 서버가 없으면 None을 반환한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 대상 리전.

    Returns:
        분석 결과 객체. 서버가 없으면 None.
    """
    servers = collect_servers(session, account_id, account_name, region)
    if not servers:
        return None
    return analyze_servers(servers, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Transfer Family 미사용 서버 분석 도구의 메인 실행 함수.

    멀티 계정/리전 병렬 수집 후 결과를 집계하고, Excel 보고서를 생성하여 출력 디렉토리에 저장한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 계정/리전 목록, 옵션 등 포함).
    """
    console.print("[bold]Transfer Family 서버 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="transfer")
    results: list[TransferAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_servers for r in results)
    total_idle = sum(r.idle_servers for r in results)
    total_no_users = sum(r.no_users_servers for r in results)
    total_waste = sum(r.total_monthly_waste for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] / "
        f"저사용: [yellow]{total_idle}개[/yellow] / "
        f"사용자없음: [orange1]{total_no_users}개[/orange1]"
    )
    console.print(f"예상 월 낭비: [red]${total_waste:,.2f}[/red]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("transfer", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
