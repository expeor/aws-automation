"""
functions/analyzers/backup/audit.py - AWS Backup 현황 분석

Backup Vault, Plan, 최근 작업 현황을 분석합니다.

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.shared.io.output import OutputPath, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 최근 작업 조회 기간
JOB_DAYS = 30

REQUIRED_PERMISSIONS = {
    "read": [
        "backup:ListBackupVaults",
        "backup:ListRecoveryPointsByBackupVault",
        "backup:ListBackupPlans",
        "backup:GetBackupPlan",
        "backup:ListBackupJobs",
    ],
}


class JobStatus(Enum):
    """Backup 작업 상태.

    Attributes:
        COMPLETED: 정상 완료.
        FAILED: 실패.
        RUNNING: 실행 중.
        ABORTED: 중단됨.
        PARTIAL: 부분 성공.
    """

    COMPLETED = "COMPLETED"
    FAILED = "FAILED"
    RUNNING = "RUNNING"
    ABORTED = "ABORTED"
    PARTIAL = "PARTIAL"


@dataclass
class VaultInfo:
    """AWS Backup Vault 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: Vault가 위치한 리전.
        vault_name: Vault 이름.
        vault_arn: Vault ARN.
        creation_date: Vault 생성일.
        encryption_key_arn: 암호화 KMS 키 ARN.
        recovery_point_count: 복구 지점 수.
        total_size_bytes: 전체 백업 크기 (바이트).
        locked: Vault 잠금 여부.
    """

    account_id: str
    account_name: str
    region: str
    vault_name: str
    vault_arn: str
    creation_date: datetime | None
    encryption_key_arn: str
    recovery_point_count: int = 0
    total_size_bytes: int = 0
    locked: bool = False

    @property
    def size_gb(self) -> float:
        """전체 백업 크기 (GB).

        Returns:
            바이트를 GB로 변환한 값.
        """
        return self.total_size_bytes / (1024**3)

    @property
    def is_encrypted(self) -> bool:
        """Vault 암호화 여부.

        Returns:
            KMS 키가 설정되어 있으면 True.
        """
        return bool(self.encryption_key_arn)


@dataclass
class PlanInfo:
    """AWS Backup Plan 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: Plan이 위치한 리전.
        plan_id: Backup Plan ID.
        plan_name: Plan 이름.
        version_id: Plan 버전 ID.
        creation_date: Plan 생성일.
        rule_count: 백업 규칙 수.
        resource_count: 리소스 선택 수.
    """

    account_id: str
    account_name: str
    region: str
    plan_id: str
    plan_name: str
    version_id: str
    creation_date: datetime | None
    rule_count: int = 0
    resource_count: int = 0


@dataclass
class JobInfo:
    """AWS Backup Job 상세 정보.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 작업이 실행된 리전.
        job_id: Backup Job ID.
        vault_name: 백업 대상 Vault 이름.
        resource_arn: 백업 대상 리소스 ARN.
        resource_type: 리소스 타입.
        status: 작업 상태.
        status_message: 상태 메시지.
        creation_date: 작업 생성일.
        completion_date: 작업 완료일.
        backup_size_bytes: 백업 크기 (바이트).
    """

    account_id: str
    account_name: str
    region: str
    job_id: str
    vault_name: str
    resource_arn: str
    resource_type: str
    status: str
    status_message: str
    creation_date: datetime | None
    completion_date: datetime | None
    backup_size_bytes: int = 0

    @property
    def is_failed(self) -> bool:
        return self.status in ("FAILED", "ABORTED", "PARTIAL")


@dataclass
class BackupAnalysisResult:
    """단일 계정/리전의 AWS Backup 분석 결과.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 분석 대상 리전.
        vaults: Backup Vault 목록.
        plans: Backup Plan 목록.
        jobs: 최근 Backup Job 목록.
    """

    account_id: str
    account_name: str
    region: str
    vaults: list[VaultInfo] = field(default_factory=list)
    plans: list[PlanInfo] = field(default_factory=list)
    jobs: list[JobInfo] = field(default_factory=list)

    @property
    def total_vaults(self) -> int:
        return len(self.vaults)

    @property
    def total_recovery_points(self) -> int:
        return sum(v.recovery_point_count for v in self.vaults)

    @property
    def total_plans(self) -> int:
        return len(self.plans)

    @property
    def job_failed_count(self) -> int:
        return sum(1 for j in self.jobs if j.is_failed)

    @property
    def failed_jobs(self) -> list:
        return [j for j in self.jobs if j.is_failed]


def _collect_backup_data(session, account_id: str, account_name: str, region: str) -> BackupAnalysisResult | None:
    """단일 계정/리전의 Backup Vault, Plan, 최근 Job을 수집한다.

    parallel_collect 콜백으로 사용되며, 데이터가 없으면 None을 반환한다.

    Args:
        session: boto3 Session 객체.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: 대상 리전.

    Returns:
        Backup 분석 결과 객체. 데이터가 없으면 None.
    """
    from botocore.exceptions import ClientError

    backup = get_client(session, "backup", region_name=region)
    result = BackupAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    # 1. Vault 수집
    try:
        paginator = backup.get_paginator("list_backup_vaults")
        for page in paginator.paginate():
            for vault in page.get("BackupVaultList", []):
                vault_name = vault.get("BackupVaultName", "")

                # 복구 지점 수 및 총 크기 계산
                rp_count = 0
                total_size = 0
                try:
                    rp_paginator = backup.get_paginator("list_recovery_points_by_backup_vault")
                    for rp_page in rp_paginator.paginate(BackupVaultName=vault_name):
                        for rp in rp_page.get("RecoveryPoints", []):
                            rp_count += 1
                            total_size += rp.get("BackupSizeInBytes", 0)
                except ClientError:
                    pass

                result.vaults.append(
                    VaultInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        vault_name=vault_name,
                        vault_arn=vault.get("BackupVaultArn", ""),
                        creation_date=vault.get("CreationDate"),
                        encryption_key_arn=vault.get("EncryptionKeyArn", ""),
                        recovery_point_count=rp_count,
                        total_size_bytes=total_size,
                        locked=vault.get("Locked", False),
                    )
                )
    except ClientError:
        pass

    # 2. Plan 수집
    try:
        paginator = backup.get_paginator("list_backup_plans")
        for page in paginator.paginate():
            for plan in page.get("BackupPlansList", []):
                plan_id = plan.get("BackupPlanId", "")

                # 플랜 상세 (규칙 수)
                rule_count = 0
                try:
                    plan_detail = backup.get_backup_plan(BackupPlanId=plan_id)
                    rules = plan_detail.get("BackupPlan", {}).get("Rules", [])
                    rule_count = len(rules)
                except ClientError:
                    pass

                # 리소스 선택 수
                resource_count = 0
                try:
                    selections = backup.list_backup_selections(BackupPlanId=plan_id)
                    resource_count = len(selections.get("BackupSelectionsList", []))
                except ClientError:
                    pass

                result.plans.append(
                    PlanInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        plan_id=plan_id,
                        plan_name=plan.get("BackupPlanName", ""),
                        version_id=plan.get("VersionId", ""),
                        creation_date=plan.get("CreationDate"),
                        rule_count=rule_count,
                        resource_count=resource_count,
                    )
                )
    except ClientError:
        pass

    # 3. 최근 Job 수집
    try:
        now = datetime.now(timezone.utc)
        start_date = now - timedelta(days=JOB_DAYS)

        paginator = backup.get_paginator("list_backup_jobs")
        for page in paginator.paginate(
            ByCreatedAfter=start_date,
            ByCreatedBefore=now,
        ):
            for job in page.get("BackupJobs", []):
                result.jobs.append(
                    JobInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        job_id=job.get("BackupJobId", ""),
                        vault_name=job.get("BackupVaultName", ""),
                        resource_arn=job.get("ResourceArn", ""),
                        resource_type=job.get("ResourceType", ""),
                        status=job.get("State", ""),
                        status_message=job.get("StatusMessage", ""),
                        creation_date=job.get("CreationDate"),
                        completion_date=job.get("CompletionDate"),
                        backup_size_bytes=job.get("BackupSizeInBytes", 0),
                    )
                )
    except ClientError:
        pass

    # 데이터가 하나라도 있으면 반환
    if result.vaults or result.plans or result.jobs:
        return result
    return None


def generate_report(results: list[BackupAnalysisResult], output_dir: str) -> str:
    """AWS Backup 현황 분석 결과를 Excel 보고서로 생성한다.

    Summary, Vaults, Plans, Failed Jobs 시트를 포함한다.

    Args:
        results: 계정/리전별 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # ========== Summary 시트 ==========
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Vault 수", width=12, style="number"),
        ColumnDef(header="복구지점 수", width=12, style="number"),
        ColumnDef(header="Plan 수", width=12, style="number"),
        ColumnDef(header="실패 작업", width=12, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_vaults,
                r.total_recovery_points,
                r.total_plans,
                r.job_failed_count,
            ]
        )
        if r.job_failed_count > 0:
            summary_sheet._ws.cell(row=row_num, column=6).fill = red_fill

    # ========== Vaults 시트 ==========
    vault_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Vault 이름", width=30),
        ColumnDef(header="복구지점 수", width=12, style="number"),
        ColumnDef(header="총 크기", width=15),
        ColumnDef(header="암호화", width=10, style="center"),
        ColumnDef(header="잠금", width=10, style="center"),
    ]
    vault_sheet = wb.new_sheet("Vaults", vault_columns)

    for r in results:
        for v in r.vaults:
            row_num = vault_sheet.add_row(
                [
                    v.account_name,
                    v.region,
                    v.vault_name,
                    v.recovery_point_count,
                    f"{v.size_gb:.2f} GB",
                    "Yes" if v.is_encrypted else "No",
                    "Yes" if v.locked else "No",
                ]
            )
            if not v.is_encrypted:
                vault_sheet._ws.cell(row=row_num, column=6).fill = yellow_fill

    # ========== Plans 시트 ==========
    plan_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Plan 이름", width=30),
        ColumnDef(header="ID", width=40),
        ColumnDef(header="규칙 수", width=10, style="number"),
        ColumnDef(header="리소스 선택 수", width=15, style="number"),
        ColumnDef(header="생성일", width=12),
    ]
    plan_sheet = wb.new_sheet("Plans", plan_columns)

    for r in results:
        for p in r.plans:
            row_num = plan_sheet.add_row(
                [
                    p.account_name,
                    p.region,
                    p.plan_name,
                    p.plan_id,
                    p.rule_count,
                    p.resource_count,
                    p.creation_date.strftime("%Y-%m-%d") if p.creation_date else "-",
                ]
            )
            # 리소스 없는 플랜 경고
            if p.resource_count == 0:
                plan_sheet._ws.cell(row=row_num, column=6).fill = yellow_fill

    # ========== Failed Jobs 시트 (실패 작업만) ==========
    failed_jobs: list[JobInfo] = []
    for r in results:
        failed_jobs.extend(r.failed_jobs)

    if failed_jobs:
        job_columns = [
            ColumnDef(header="Account", width=20),
            ColumnDef(header="Region", width=15),
            ColumnDef(header="상태", width=12, style="center"),
            ColumnDef(header="리소스 타입", width=15),
            ColumnDef(header="리소스 ARN", width=50),
            ColumnDef(header="Vault", width=25),
            ColumnDef(header="생성일", width=18),
            ColumnDef(header="메시지", width=40),
        ]
        job_sheet = wb.new_sheet("Failed Jobs", job_columns)

        failed_jobs.sort(key=lambda j: j.creation_date or datetime.min, reverse=True)

        for j in failed_jobs:
            row_num = job_sheet.add_row(
                [
                    j.account_name,
                    j.region,
                    j.status,
                    j.resource_type,
                    j.resource_arn,
                    j.vault_name,
                    j.creation_date.strftime("%Y-%m-%d %H:%M") if j.creation_date else "-",
                    j.status_message or "-",
                ],
                style=Styles.danger(),
            )
            job_sheet._ws.cell(row=row_num, column=3).fill = red_fill

    return str(wb.save_as(output_dir, "AWS_Backup"))


def run(ctx: ExecutionContext) -> None:
    """AWS Backup 현황 분석 도구의 메인 실행 함수.

    멀티 계정/리전 병렬 수집 후 결과를 집계하고, Excel 보고서를 생성하여 출력 디렉토리에 저장한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 계정/리전 목록, 옵션 등 포함).
    """
    console.print("[bold]AWS Backup 현황 분석 시작...[/bold]\n")
    console.print(f"[dim]최근 {JOB_DAYS}일 작업 현황 포함[/dim]\n")

    result = parallel_collect(ctx, _collect_backup_data, max_workers=10, service="backup")
    results: list[BackupAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        for err in result.get_errors()[:5]:
            console.print(f"[dim]  - {err.identifier}/{err.region}: {err.message}[/dim]")

    if not results:
        console.print("\n[yellow]Backup 데이터 없음[/yellow]")
        console.print("[dim]Backup Vault가 없거나 접근 권한이 없습니다.[/dim]")
        return

    # 통계 계산
    total_vaults = sum(r.total_vaults for r in results)
    total_rps = sum(r.total_recovery_points for r in results)
    total_plans = sum(r.total_plans for r in results)
    total_failed = sum(r.job_failed_count for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"Vault: {total_vaults}개, 복구 지점: {total_rps}개")
    console.print(f"Plan: {total_plans}개")

    if total_failed > 0:
        console.print(f"[red]최근 {JOB_DAYS}일 실패 작업: {total_failed}건[/red]")

    # 보고서 생성
    identifier = ctx.accounts[0].id if ctx.accounts else ctx.profile_name or "default"
    output_path = OutputPath(identifier).sub("backup", "inventory").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
