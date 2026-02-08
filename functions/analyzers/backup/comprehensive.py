"""
functions/analyzers/backup/comprehensive.py - 통합 백업 현황 분석

AWS Backup + 각 서비스별 자체 자동 백업 현황을 통합 조회합니다.

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import parallel_collect

from .collectors import JOB_DAYS, _collect_comprehensive_backup_data
from .models import (
    BackupPlanTagCondition,
    BackupStatus,
    ComprehensiveBackupResult,
    FailedBackupJob,
    ServiceBackupSummary,
)

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

REQUIRED_PERMISSIONS = {
    "read": [
        "backup:ListBackupJobs",
        "backup:ListBackupPlans",
        "backup:ListBackupSelections",
        "backup:GetBackupSelection",
        "backup:ListProtectedResources",
        "rds:DescribeDBInstances",
        "rds:DescribeDBClusters",
        "rds:DescribeDBSnapshots",
        "rds:DescribeDBClusterSnapshots",
        "dynamodb:ListTables",
        "dynamodb:DescribeContinuousBackups",
        "elasticfilesystem:DescribeFileSystems",
        "elasticfilesystem:DescribeBackupPolicy",
        "fsx:DescribeFileSystems",
        "fsx:DescribeBackups",
        "ec2:DescribeInstances",
        "docdb:DescribeDBClusters",
        "neptune:DescribeDBClusters",
        "redshift:DescribeClusters",
        "redshift-serverless:ListNamespaces",
        "elasticache:DescribeReplicationGroups",
        "elasticache:DescribeCacheClusters",
        "memorydb:DescribeClusters",
        "es:ListDomainNames",
        "es:DescribeDomain",
        "tag:GetResources",
    ],
}


# Re-export for backwards compatibility
__all__ = [
    "BackupStatus",
    "ServiceBackupSummary",
    "BackupPlanTagCondition",
    "FailedBackupJob",
    "ComprehensiveBackupResult",
    "generate_report",
    "run",
]


def generate_report(results: list[ComprehensiveBackupResult], output_dir: str) -> str:
    """통합 백업 현황 분석 결과를 Excel 보고서로 생성한다.

    Summary, Backup Status, Backup Plan Tags, Failed Jobs 시트를 포함한다.
    Backup Status 시트에는 보호 상태/백업 지연/실패 횟수에 따라 셀 색상을 적용한다.

    Args:
        results: 계정/리전별 통합 백업 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 스타일
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    green_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")

    # ========== Summary 시트 ==========
    # 서비스별 요약
    all_summaries: dict[str, ServiceBackupSummary] = {}
    for r in results:
        for s in r.get_service_summary():
            if s.service not in all_summaries:
                all_summaries[s.service] = ServiceBackupSummary(
                    service=s.service,
                    total_resources=0,
                    backup_enabled=0,
                    backup_disabled=0,
                    failed_count=0,
                    warning_count=0,
                )
            ss = all_summaries[s.service]
            ss.total_resources += s.total_resources
            ss.backup_enabled += s.backup_enabled
            ss.backup_disabled += s.backup_disabled
            ss.failed_count += s.failed_count
            ss.warning_count += s.warning_count

    summary = wb.new_summary_sheet("Summary")
    summary.add_title("통합 백업 현황 보고서")
    summary.add_item("생성일", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()
    summary.add_section("서비스별 요약")

    for service in [
        "RDS",
        "Aurora",
        "DynamoDB",
        "EFS",
        "FSx",
        "EC2",
        "DocumentDB",
        "Neptune",
        "Redshift",
        "Redshift Serverless",
        "ElastiCache",
        "MemoryDB",
        "OpenSearch",
    ]:
        if service in all_summaries:
            s = all_summaries[service]
            disabled_rate = (s.backup_disabled / s.total_resources * 100) if s.total_resources > 0 else 0
            highlight = "warning" if s.backup_disabled > 0 else None
            summary.add_item(
                f"{s.service} (총 {s.total_resources}개)",
                f"보호: {s.backup_enabled}, 미보호: {s.backup_disabled} ({disabled_rate:.1f}%)",
                highlight=highlight,
            )

    # 전체 통계
    total_failed_jobs = sum(len(r.failed_jobs) for r in results)
    summary.add_blank_row()
    highlight = "danger" if total_failed_jobs > 0 else None
    summary.add_item("실패한 Backup 작업 (최근 30일)", total_failed_jobs, highlight=highlight)

    # ========== Backup Status 시트 (서비스별) ==========
    all_statuses: list[BackupStatus] = []
    for r in results:
        all_statuses.extend(r.backup_statuses)

    # 실패 이력 집계 (리소스 ARN별)
    all_failed_jobs: list[FailedBackupJob] = []
    for r in results:
        all_failed_jobs.extend(r.failed_jobs)

    # 리소스별 실패 횟수 및 최근 실패 시간
    failure_by_resource: dict[str, dict] = {}  # {arn: {count: int, last_failure: datetime}}
    for job in all_failed_jobs:
        arn = job.resource_arn
        if arn not in failure_by_resource:
            failure_by_resource[arn] = {"count": 0, "last_failure": None}
        failure_by_resource[arn]["count"] += 1
        if job.creation_date:
            if failure_by_resource[arn]["last_failure"] is None:
                failure_by_resource[arn]["last_failure"] = job.creation_date
            else:
                failure_by_resource[arn]["last_failure"] = max(
                    failure_by_resource[arn]["last_failure"], job.creation_date
                )

    if all_statuses:
        status_columns = [
            ColumnDef(header="Account", width=15),
            ColumnDef(header="Region", width=15),
            ColumnDef(header="서비스", width=15),
            ColumnDef(header="리소스 타입", width=20),
            ColumnDef(header="리소스 ID", width=25),
            ColumnDef(header="리소스 이름", width=25),
            ColumnDef(header="자체 백업", width=10, style="center"),
            ColumnDef(header="AWS Backup", width=10, style="center"),
            ColumnDef(header="보호 상태", width=15, style="center"),
            ColumnDef(header="백업 방식", width=18),
            ColumnDef(header="보존 기간(일)", width=12, style="number"),
            ColumnDef(header="자체 백업 시간", width=18),
            ColumnDef(header="AWS Backup 시간", width=18),
            ColumnDef(header="백업 지연", width=12),
            ColumnDef(header="실패 횟수", width=10, style="number"),
            ColumnDef(header="최근 실패", width=18),
            ColumnDef(header="Backup Plan", width=25),
            ColumnDef(header="상태", width=12, style="center"),
            ColumnDef(header="메시지", width=40),
        ]
        status_sheet = wb.new_sheet("Backup Status", status_columns)

        now = datetime.now(timezone.utc)
        two_days_ago = now - timedelta(days=2)

        for s in all_statuses:
            # 백업 지연 체크 (보호 중인데 마지막 백업이 2일 이상 전이면 지연)
            latest_backup = None
            if s.last_backup_time and s.aws_backup_last_time:
                latest_backup = max(s.last_backup_time, s.aws_backup_last_time)
            elif s.last_backup_time:
                latest_backup = s.last_backup_time
            elif s.aws_backup_last_time:
                latest_backup = s.aws_backup_last_time

            backup_delayed = ""
            if (s.backup_enabled or s.aws_backup_protected) and latest_backup:
                if latest_backup < two_days_ago:
                    days_delayed = (now - latest_backup).days
                    backup_delayed = f"{days_delayed}일 지연"

            # 실패 이력 (최근 30일)
            failure_info = failure_by_resource.get(s.resource_arn, {"count": 0, "last_failure": None})
            failure_count = failure_info["count"]
            last_failure = failure_info["last_failure"]

            # Backup Plan 표시
            backup_plan_str = ", ".join(s.backup_plan_names) if s.backup_plan_names else "-"

            row_num = status_sheet.add_row(
                [
                    s.account_name,
                    s.region,
                    s.service,
                    s.resource_type,
                    s.resource_id,
                    s.resource_name,
                    "Yes" if s.backup_enabled else "No",
                    "Yes" if s.aws_backup_protected else "No",
                    s.protection_summary,
                    s.backup_method,
                    s.backup_retention_days,
                    s.last_backup_time.strftime("%Y-%m-%d %H:%M") if s.last_backup_time else "-",
                    s.aws_backup_last_time.strftime("%Y-%m-%d %H:%M") if s.aws_backup_last_time else "-",
                    backup_delayed,
                    failure_count if failure_count > 0 else "-",
                    last_failure.strftime("%Y-%m-%d %H:%M") if last_failure else "-",
                    backup_plan_str,
                    s.status,
                    s.message,
                ]
            )

            ws = status_sheet._ws

            if backup_delayed:
                ws.cell(row=row_num, column=14).fill = yellow_fill

            if failure_count > 0:
                ws.cell(row=row_num, column=15).fill = red_fill
            if last_failure:
                ws.cell(row=row_num, column=16).fill = red_fill

            # 보호 상태별 색상 (서비스 특성 고려)
            if s.protection_summary == "미보호":
                ws.cell(row=row_num, column=9).fill = red_fill
            elif s.is_fully_protected:
                ws.cell(row=row_num, column=9).fill = green_fill
            else:
                ws.cell(row=row_num, column=9).fill = yellow_fill

            # 상태별 색상
            if s.status == "DISABLED" and not s.aws_backup_protected:
                ws.cell(row=row_num, column=18).fill = yellow_fill
            elif s.status == "FAILED":
                ws.cell(row=row_num, column=18).fill = red_fill
            elif s.status == "OK":
                ws.cell(row=row_num, column=18).fill = green_fill

    # ========== Backup Plan Tags 시트 ==========
    all_tag_conditions: list[BackupPlanTagCondition] = []
    for r in results:
        all_tag_conditions.extend(r.tag_conditions)

    if all_tag_conditions:
        tag_columns = [
            ColumnDef(header="Account", width=15),
            ColumnDef(header="Region", width=15),
            ColumnDef(header="Plan ID", width=40),
            ColumnDef(header="Plan 이름", width=25),
            ColumnDef(header="Selection ID", width=40),
            ColumnDef(header="Selection 이름", width=25),
            ColumnDef(header="조건 타입", width=15),
            ColumnDef(header="태그 키", width=20),
            ColumnDef(header="태그 값", width=20),
            ColumnDef(header="대상 리소스 타입", width=20),
        ]
        tag_sheet = wb.new_sheet("Backup Plan Tags", tag_columns)

        for t in all_tag_conditions:
            tag_sheet.add_row(
                [
                    t.account_name,
                    t.region,
                    t.plan_id,
                    t.plan_name,
                    t.selection_id,
                    t.selection_name,
                    t.condition_type,
                    t.tag_key,
                    t.tag_value,
                    ", ".join(t.resource_types) if t.resource_types else "*",
                ]
            )

    # ========== Failed Jobs 시트 ==========
    all_failed_jobs = []
    for r in results:
        all_failed_jobs.extend(r.failed_jobs)

    if all_failed_jobs:
        failed_columns = [
            ColumnDef(header="Account", width=15),
            ColumnDef(header="Region", width=15),
            ColumnDef(header="Job ID", width=40),
            ColumnDef(header="상태", width=12, style="center"),
            ColumnDef(header="리소스 타입", width=15),
            ColumnDef(header="리소스 ARN", width=50),
            ColumnDef(header="Vault", width=25),
            ColumnDef(header="생성일", width=18),
            ColumnDef(header="완료일", width=18),
            ColumnDef(header="메시지", width=40),
        ]
        failed_sheet = wb.new_sheet("Failed Jobs", failed_columns)

        # 최신순 정렬
        all_failed_jobs.sort(key=lambda j: j.creation_date or datetime.min, reverse=True)

        for j in all_failed_jobs:
            row_num = failed_sheet.add_row(
                [
                    j.account_name,
                    j.region,
                    j.job_id,
                    j.status,
                    j.resource_type,
                    j.resource_arn,
                    j.vault_name,
                    j.creation_date.strftime("%Y-%m-%d %H:%M") if j.creation_date else "-",
                    j.completion_date.strftime("%Y-%m-%d %H:%M") if j.completion_date else "-",
                    j.status_message or "-",
                ],
                style=Styles.danger(),
            )
            failed_sheet._ws.cell(row=row_num, column=4).fill = red_fill

    return str(wb.save_as(output_dir, "Comprehensive_Backup"))


# ============================================================================
# 실행 함수
# ============================================================================


def run(ctx: ExecutionContext) -> None:
    """통합 백업 현황 분석 도구의 메인 실행 함수.

    RDS/Aurora, DynamoDB, EFS, FSx, EC2, DocumentDB, Neptune, Redshift,
    ElastiCache, MemoryDB, OpenSearch 등 13개 서비스의 백업 현황을 통합 분석한다.
    AWS Backup과 서비스 자체 백업을 모두 조회하고, 최근 실패 작업을 포함하여 보고한다.

    Args:
        ctx: 실행 컨텍스트. 계정 정보, 리전, 프로파일 등을 포함한다.
    """
    from core.shared.io.output import OutputPath, open_in_explorer

    console.print("[bold]통합 백업 현황 분석 시작...[/bold]\n")
    console.print(
        "[dim]분석 대상: RDS/Aurora, DynamoDB, EFS, FSx, EC2, DocumentDB, Neptune, "
        "Redshift, ElastiCache, MemoryDB, OpenSearch[/dim]"
    )
    console.print(f"[dim]최근 {JOB_DAYS}일 실패 작업 포함[/dim]\n")

    result = parallel_collect(ctx, _collect_comprehensive_backup_data, max_workers=10, service="backup")
    results: list[ComprehensiveBackupResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        for err in result.get_errors()[:5]:
            console.print(f"[dim]  - {err.identifier}/{err.region}: {err.message}[/dim]")

    if not results:
        console.print("\n[yellow]백업 데이터 없음[/yellow]")
        console.print("[dim]리소스가 없거나 접근 권한이 없습니다.[/dim]")
        return

    # 통계 계산
    total_resources = sum(len(r.backup_statuses) for r in results)
    disabled_resources = sum(len(r.get_disabled_resources()) for r in results)
    failed_jobs = sum(len(r.failed_jobs) for r in results)
    tag_conditions = sum(len(r.tag_conditions) for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"분석 리소스: {total_resources}개")

    if disabled_resources > 0:
        console.print(f"[yellow]백업 비활성화: {disabled_resources}개[/yellow]")
    else:
        console.print("[green]백업 비활성화: 0개[/green]")

    if failed_jobs > 0:
        console.print(f"[red]최근 {JOB_DAYS}일 실패 작업: {failed_jobs}건[/red]")

    if tag_conditions > 0:
        console.print(f"Backup Plan 태그 조건: {tag_conditions}개")

    # 보고서 생성
    identifier = ctx.accounts[0].id if ctx.accounts else ctx.profile_name or "default"
    output_path = OutputPath(identifier).sub("backup", "comprehensive").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
