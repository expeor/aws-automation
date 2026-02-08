"""
analyzers/fn/runtime_deprecated.py - Lambda 런타임 지원 종료 분석

Lambda 함수 런타임 지원 종료 현황 분석:
- Deprecated: 이미 지원 종료된 런타임
- Deprecated Soon: 365일 이내 지원 종료 예정
- Safe: 365일 이상 남았거나 종료일 미정
- Container: 컨테이너 이미지 기반 (런타임 해당 없음)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console
from rich.table import Table

from core.parallel import parallel_collect, quiet_mode
from core.shared.aws.lambda_ import (
    EOLStatus,
    LambdaFunctionInfo,
    collect_functions,
    get_runtime_info,
)
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 분류 기준: 365일 이내 -> SOON
SOON_THRESHOLD_DAYS = 365

# EOLStatus 표시 라벨 (범례)
EOL_STATUS_LABELS: dict[str, str] = {
    "deprecated": "Deprecated (지원 종료)",
    "critical": "Critical (30일 이내)",
    "high": "High (90일 이내)",
    "medium": "Medium (180일 이내)",
    "low": "Low (365일 이내)",
    "supported": "Supported (EOL 미정)",
}


def _eol_label(status_value: str) -> str:
    """EOLStatus value를 사람이 읽기 쉬운 라벨로 변환"""
    return EOL_STATUS_LABELS.get(status_value, status_value)


# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:ListTags",
        "lambda:ListProvisionedConcurrencyConfigs",
        "lambda:GetFunctionConcurrency",
    ],
}


# =============================================================================
# 데이터 구조
# =============================================================================


class DeprecationCategory(Enum):
    """런타임 지원 종료 분류"""

    DEPRECATED = "deprecated"  # 이미 지원 종료
    SOON = "soon"  # 365일 이내 종료 예정
    SAFE = "safe"  # 365일 이상 남음 또는 종료일 미정
    CONTAINER = "container"  # 컨테이너 이미지 (런타임 없음)


@dataclass
class RuntimeDeprecationFinding:
    """개별 함수 분석 결과"""

    function: LambdaFunctionInfo
    category: DeprecationCategory
    runtime_name: str
    os_version: str
    deprecation_date: date | None
    days_remaining: int | None
    recommended_upgrade: str
    eol_status: EOLStatus


@dataclass
class RuntimeDeprecationResult:
    """리전별 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_functions: int = 0
    deprecated_count: int = 0
    soon_count: int = 0
    safe_count: int = 0
    container_count: int = 0
    findings: list[RuntimeDeprecationFinding] = field(default_factory=list)


# =============================================================================
# 분류 로직
# =============================================================================


def _classify_function(func: LambdaFunctionInfo) -> RuntimeDeprecationFinding:
    """Lambda 함수의 런타임 지원 종료 상태 분류"""
    runtime = func.runtime

    # 컨테이너 이미지 함수 (PackageType=Image -> Runtime="unknown")
    if runtime in ("", "unknown", None):
        return RuntimeDeprecationFinding(
            function=func,
            category=DeprecationCategory.CONTAINER,
            runtime_name="Container Image",
            os_version="",
            deprecation_date=None,
            days_remaining=None,
            recommended_upgrade="",
            eol_status=EOLStatus.SUPPORTED,
        )

    info = get_runtime_info(runtime)

    # 알 수 없는 런타임 -> SAFE (보수적 분류)
    if info is None:
        return RuntimeDeprecationFinding(
            function=func,
            category=DeprecationCategory.SAFE,
            runtime_name=runtime,
            os_version="",
            deprecation_date=None,
            days_remaining=None,
            recommended_upgrade="",
            eol_status=EOLStatus.SUPPORTED,
        )

    days = info.days_until_deprecation
    upgrade = info.recommended_upgrade or ""

    # 이미 지원 종료
    if info.is_deprecated:
        return RuntimeDeprecationFinding(
            function=func,
            category=DeprecationCategory.DEPRECATED,
            runtime_name=info.name,
            os_version=info.os_version,
            deprecation_date=info.deprecation_date,
            days_remaining=days,
            recommended_upgrade=upgrade,
            eol_status=info.status,
        )

    # 365일 이내 종료 예정
    if days is not None and days <= SOON_THRESHOLD_DAYS:
        return RuntimeDeprecationFinding(
            function=func,
            category=DeprecationCategory.SOON,
            runtime_name=info.name,
            os_version=info.os_version,
            deprecation_date=info.deprecation_date,
            days_remaining=days,
            recommended_upgrade=upgrade,
            eol_status=info.status,
        )

    # 안전
    return RuntimeDeprecationFinding(
        function=func,
        category=DeprecationCategory.SAFE,
        runtime_name=info.name,
        os_version=info.os_version,
        deprecation_date=info.deprecation_date,
        days_remaining=days,
        recommended_upgrade=upgrade,
        eol_status=info.status,
    )


def _analyze_region(
    functions: list[LambdaFunctionInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> RuntimeDeprecationResult:
    """리전의 모든 함수 분석"""
    result = RuntimeDeprecationResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_functions=len(functions),
    )

    for func in functions:
        finding = _classify_function(func)
        result.findings.append(finding)

        if finding.category == DeprecationCategory.DEPRECATED:
            result.deprecated_count += 1
        elif finding.category == DeprecationCategory.SOON:
            result.soon_count += 1
        elif finding.category == DeprecationCategory.CONTAINER:
            result.container_count += 1
        else:
            result.safe_count += 1

    return result


# =============================================================================
# 보고서
# =============================================================================


def _build_runtime_distribution(
    results: list[RuntimeDeprecationResult],
) -> list[dict]:
    """런타임 분포 데이터 생성"""
    runtime_counter: Counter[str] = Counter()
    for r in results:
        for f in r.findings:
            if f.category != DeprecationCategory.CONTAINER:
                runtime_counter[f.function.runtime] += 1
            else:
                runtime_counter["container"] += 1

    distribution = []
    for runtime_id, count in runtime_counter.most_common():
        if runtime_id == "container":
            distribution.append(
                {
                    "runtime": "Container Image",
                    "count": count,
                    "status": "N/A",
                    "os_version": "",
                    "deprecation_date": "",
                    "days_remaining": "",
                    "recommended_upgrade": "",
                }
            )
            continue

        info = get_runtime_info(runtime_id)
        if info:
            days = info.days_until_deprecation
            distribution.append(
                {
                    "runtime": info.name,
                    "count": count,
                    "status": info.status.value,
                    "os_version": info.os_version,
                    "deprecation_date": str(info.deprecation_date) if info.deprecation_date else "-",
                    "days_remaining": str(days) if days is not None else "-",
                    "recommended_upgrade": info.recommended_upgrade or "-",
                }
            )
        else:
            distribution.append(
                {
                    "runtime": runtime_id,
                    "count": count,
                    "status": "unknown",
                    "os_version": "",
                    "deprecation_date": "-",
                    "days_remaining": "-",
                    "recommended_upgrade": "-",
                }
            )

    return distribution


def generate_excel_report(results: list[RuntimeDeprecationResult], output_dir: str) -> str:
    """Excel 보고서 생성 (6 시트)"""
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    wb = Workbook()

    # --- Sheet 1: Summary ---
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("Lambda 런타임 지원 종료 분석 보고서")
    summary.add_item("생성", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_item("분류 기준", f"SOON = {SOON_THRESHOLD_DAYS}일 이내")
    summary.add_blank_row()

    total_functions = sum(r.total_functions for r in results)
    total_deprecated = sum(r.deprecated_count for r in results)
    total_soon = sum(r.soon_count for r in results)
    total_safe = sum(r.safe_count for r in results)
    total_container = sum(r.container_count for r in results)

    summary.add_item("전체 함수", str(total_functions))
    summary.add_item("지원 종료됨 (Deprecated)", str(total_deprecated))
    summary.add_item("곧 종료 (Soon)", str(total_soon))
    summary.add_item("안전 (Safe)", str(total_safe))
    summary.add_item("컨테이너 (Container)", str(total_container))

    # --- Sheet 2: Summary Data ---
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="종료됨", width=10, style="number"),
        ColumnDef(header="곧 종료", width=10, style="number"),
        ColumnDef(header="안전", width=10, style="number"),
        ColumnDef(header="컨테이너", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)
    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_functions,
                r.deprecated_count,
                r.soon_count,
                r.safe_count,
                r.container_count,
            ]
        )
        ws = summary_sheet._ws
        if r.deprecated_count > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.soon_count > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    summary_sheet.add_summary_row(
        ["합계", "-", total_functions, total_deprecated, total_soon, total_safe, total_container]
    )

    # --- Sheet 3: Deprecated ---
    dep_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function Name", width=40),
        ColumnDef(header="Runtime", width=20),
        ColumnDef(header="OS Version", width=12, style="center"),
        ColumnDef(header="Deprecated Date", width=15),
        ColumnDef(header="Days Past", width=12, style="number"),
        ColumnDef(header="Recommended Upgrade", width=20),
        ColumnDef(header="Last Modified", width=15),
    ]
    dep_sheet = wb.new_sheet("Deprecated", dep_columns)
    for r in results:
        for f in r.findings:
            if f.category == DeprecationCategory.DEPRECATED:
                fn = f.function
                days_past = abs(f.days_remaining) if f.days_remaining is not None else "-"
                row_num = dep_sheet.add_row(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        f.runtime_name,
                        f.os_version or "-",
                        str(f.deprecation_date) if f.deprecation_date else "-",
                        days_past,
                        f.recommended_upgrade or "-",
                        fn.last_modified.strftime("%Y-%m-%d") if fn.last_modified else "-",
                    ]
                )
                dep_sheet._ws.cell(row=row_num, column=4).fill = red_fill

    # --- Sheet 4: Deprecated Soon ---
    soon_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function Name", width=40),
        ColumnDef(header="Runtime", width=20),
        ColumnDef(header="OS Version", width=12, style="center"),
        ColumnDef(header="Deprecation Date", width=15),
        ColumnDef(header="Days Remaining", width=14, style="number"),
        ColumnDef(header="EOL Status", width=28),
        ColumnDef(header="Recommended Upgrade", width=20),
        ColumnDef(header="Last Modified", width=15),
    ]
    soon_sheet = wb.new_sheet("Deprecated Soon", soon_columns)
    for r in results:
        for f in r.findings:
            if f.category == DeprecationCategory.SOON:
                fn = f.function
                row_num = soon_sheet.add_row(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        f.runtime_name,
                        f.os_version or "-",
                        str(f.deprecation_date) if f.deprecation_date else "-",
                        f.days_remaining if f.days_remaining is not None else "-",
                        _eol_label(f.eol_status.value),
                        f.recommended_upgrade or "-",
                        fn.last_modified.strftime("%Y-%m-%d") if fn.last_modified else "-",
                    ]
                )
                soon_sheet._ws.cell(row=row_num, column=4).fill = yellow_fill

    # --- Sheet 5: Safe ---
    safe_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function Name", width=40),
        ColumnDef(header="Runtime", width=20),
        ColumnDef(header="OS Version", width=12, style="center"),
        ColumnDef(header="EOL Status", width=28),
        ColumnDef(header="Deprecation Date", width=15),
        ColumnDef(header="Days Remaining", width=14, style="number"),
    ]
    safe_sheet = wb.new_sheet("Safe", safe_columns)
    for r in results:
        for f in r.findings:
            if f.category in (DeprecationCategory.SAFE, DeprecationCategory.CONTAINER):
                fn = f.function
                safe_sheet.add_row(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        f.runtime_name,
                        f.os_version or "-",
                        _eol_label(f.eol_status.value),
                        str(f.deprecation_date) if f.deprecation_date else "-",
                        f.days_remaining if f.days_remaining is not None else "-",
                    ]
                )

    # --- Sheet 6: Runtime Distribution ---
    dist_columns = [
        ColumnDef(header="Runtime", width=25),
        ColumnDef(header="Functions", width=12, style="number"),
        ColumnDef(header="EOL Status", width=28),
        ColumnDef(header="OS Version", width=12, style="center"),
        ColumnDef(header="Deprecation Date", width=15),
        ColumnDef(header="Days Remaining", width=14),
        ColumnDef(header="Recommended Upgrade", width=20),
    ]
    dist_sheet = wb.new_sheet("Runtime Distribution", dist_columns)
    distribution = _build_runtime_distribution(results)
    for d in distribution:
        row_num = dist_sheet.add_row(
            [
                d["runtime"],
                d["count"],
                _eol_label(d["status"]),
                d["os_version"] or "-",
                d["deprecation_date"],
                d["days_remaining"],
                d["recommended_upgrade"],
            ]
        )
        ws = dist_sheet._ws
        status = d["status"]
        if status == "deprecated":
            ws.cell(row=row_num, column=3).fill = red_fill
        elif status in ("critical", "high", "medium"):
            ws.cell(row=row_num, column=3).fill = yellow_fill
        elif status in ("low", "supported"):
            ws.cell(row=row_num, column=3).fill = green_fill

    return str(wb.save_as(output_dir, "Lambda_Runtime_Deprecation"))


def generate_html_report(
    results: list[RuntimeDeprecationResult],
    output_dir: str,
) -> str:
    """HTML 보고서 생성"""
    from core.shared.io.html import HTMLReport

    total_functions = sum(r.total_functions for r in results)
    total_deprecated = sum(r.deprecated_count for r in results)
    total_soon = sum(r.soon_count for r in results)
    total_safe = sum(r.safe_count for r in results)
    total_container = sum(r.container_count for r in results)

    report = HTMLReport("Lambda 런타임 지원 종료 분석", f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # Summary cards
    report.add_summary(
        [
            ("전체 함수", total_functions, None),
            ("지원 종료됨", total_deprecated, "danger" if total_deprecated > 0 else None),
            ("곧 종료", total_soon, "warning" if total_soon > 0 else None),
            ("안전", total_safe, "success"),
            ("컨테이너", total_container, None),
        ]
    )

    # Chart 1: Deprecation category pie chart
    report.add_section_title("개요")
    category_data: list[tuple[str, int]] = []
    if total_deprecated > 0:
        category_data.append(("지원 종료됨", total_deprecated))
    if total_soon > 0:
        category_data.append(("곧 종료", total_soon))
    if total_safe > 0:
        category_data.append(("안전", total_safe))
    if total_container > 0:
        category_data.append(("컨테이너", total_container))

    if category_data:
        report.add_pie_chart("런타임 지원 종료 분류", category_data, doughnut=True)

    # Chart 2: Runtime distribution bar chart
    distribution = _build_runtime_distribution(results)
    if distribution:
        runtime_names = [d["runtime"] for d in distribution]
        runtime_counts = [d["count"] for d in distribution]
        report.add_bar_chart(
            "런타임별 함수 분포",
            categories=runtime_names,
            series=[("함수 수", runtime_counts)],
            top_n=15,
        )

    # Chart 3: OS version doughnut chart
    os_counter: Counter[str] = Counter()
    for r in results:
        for f in r.findings:
            if f.os_version:
                os_counter[f.os_version] += 1
            elif f.category == DeprecationCategory.CONTAINER:
                os_counter["Container"] += 1
            else:
                os_counter["Unknown"] += 1

    os_data = [(os_name, count) for os_name, count in os_counter.most_common()]
    if os_data:
        report.add_pie_chart("Amazon Linux 버전 분포", os_data, doughnut=True)

    # Chart 4: Monthly deprecation timeline (SOON category)
    soon_findings = []
    for r in results:
        for f in r.findings:
            if f.category == DeprecationCategory.SOON and f.deprecation_date:
                soon_findings.append(f)

    if soon_findings:
        # Group by month
        monthly: dict[str, int] = defaultdict(int)
        for f in soon_findings:
            assert f.deprecation_date is not None
            month_key = f.deprecation_date.strftime("%Y-%m")
            monthly[month_key] += 1

        months = sorted(monthly.keys())
        counts = [monthly[m] for m in months]
        if months:
            report.add_section_title("지원 종료 예정 타임라인")
            report.add_bar_chart(
                "월별 지원 종료 예정 함수 수",
                categories=months,
                series=[("함수 수", counts)],
            )

    # Table: Deprecated + Soon functions
    at_risk_rows: list[list] = []
    for r in results:
        for f in r.findings:
            if f.category in (DeprecationCategory.DEPRECATED, DeprecationCategory.SOON):
                fn = f.function
                at_risk_rows.append(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        f.runtime_name,
                        f.os_version or "-",
                        _eol_label(f.eol_status.value),
                        str(f.deprecation_date) if f.deprecation_date else "-",
                        f.days_remaining if f.days_remaining is not None else "-",
                        f.recommended_upgrade or "-",
                    ]
                )

    if at_risk_rows:
        report.add_section_title("조치 필요 함수 목록")
        report.add_table(
            "지원 종료 / 곧 종료 함수",
            headers=[
                "Account",
                "Region",
                "Function",
                "Runtime",
                "OS",
                "EOL Status",
                "Deprecation Date",
                "Days Remaining",
                "Recommended Upgrade",
            ],
            rows=at_risk_rows,
        )

    # Save
    from pathlib import Path

    html_path = Path(output_dir) / "Lambda_Runtime_Deprecation.html"
    report.save(str(html_path), auto_open=False)
    return str(html_path)


# =============================================================================
# 콘솔 출력
# =============================================================================


def _print_console_summary(results: list[RuntimeDeprecationResult]) -> None:
    """콘솔에 요약 출력"""
    total_functions = sum(r.total_functions for r in results)
    total_deprecated = sum(r.deprecated_count for r in results)
    total_soon = sum(r.soon_count for r in results)
    total_safe = sum(r.safe_count for r in results)
    total_container = sum(r.container_count for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체 Lambda 함수: {total_functions}개")
    if total_deprecated > 0:
        console.print(f"  [red]지원 종료됨: {total_deprecated}개[/red]")
    if total_soon > 0:
        console.print(f"  [yellow]곧 종료 (365일 이내): {total_soon}개[/yellow]")
    if total_safe > 0:
        console.print(f"  [green]안전: {total_safe}개[/green]")
    if total_container > 0:
        console.print(f"  [dim]컨테이너: {total_container}개[/dim]")

    # 위험 런타임별 함수 수 테이블
    runtime_counter: Counter[str] = Counter()
    runtime_meta: dict[str, RuntimeDeprecationFinding] = {}
    for r in results:
        for f in r.findings:
            if f.category in (DeprecationCategory.DEPRECATED, DeprecationCategory.SOON):
                runtime_counter[f.function.runtime] += 1
                if f.function.runtime not in runtime_meta:
                    runtime_meta[f.function.runtime] = f

    if runtime_counter:
        console.print()
        table = Table(title="조치 필요 런타임", show_lines=False)
        table.add_column("Runtime", style="bold")
        table.add_column("Functions", justify="right")
        table.add_column("EOL Status")
        table.add_column("Days", justify="right")
        table.add_column("Upgrade To")

        for runtime_id, count in runtime_counter.most_common():
            meta = runtime_meta[runtime_id]
            cat_style = "red" if meta.category == DeprecationCategory.DEPRECATED else "yellow"
            days_str = str(meta.days_remaining) if meta.days_remaining is not None else "-"
            status_label = _eol_label(meta.eol_status.value)
            table.add_row(
                meta.runtime_name,
                str(count),
                f"[{cat_style}]{status_label}[/{cat_style}]",
                days_str,
                meta.recommended_upgrade or "-",
            )

        console.print(table)


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> RuntimeDeprecationResult:
    """단일 계정/리전의 Lambda 수집 및 분석 (병렬 실행용)"""
    functions = collect_functions(session, account_id, account_name, region)
    return _analyze_region(functions, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Lambda 런타임 지원 종료 분석 실행"""
    # 병렬 수집 및 분석
    # timeline이 ctx에 있으면 parallel_collect가 자동으로 프로그레스 연결
    with quiet_mode():
        result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")

    results: list[RuntimeDeprecationResult] = result.get_data()

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 콘솔 요약
    _print_console_summary(results)

    # 출력 경로
    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("lambda", "runtime_deprecated").with_date().build()

    # Excel 보고서
    filepath = generate_excel_report(results, output_path)
    console.print(f"\n[bold green]Excel:[/bold green] {filepath}")

    # HTML 보고서
    html_path = generate_html_report(results, output_path)
    console.print(f"[bold green]HTML:[/bold green] {html_path}")

    open_in_explorer(output_path)
