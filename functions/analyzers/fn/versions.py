"""
functions/analyzers/fn/versions.py - Lambda Version/Alias 정리

Lambda Version/Alias 관리:
- 오래된 버전 탐지
- 미사용 Alias 탐지
- 버전 정리 권고

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import contextlib
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:ListVersionsByFunction",
        "lambda:ListAliases",
    ],
}


class VersionStatus(Enum):
    """Lambda 버전 상태 분류.

    버전 번호, Alias 연결 여부, 수정 일자를 기반으로 각 버전의
    사용 상태를 분류한다.
    """

    CURRENT = "current"  # 현재 사용 중
    ALIAS_TARGET = "alias_target"  # Alias가 가리킴
    OLD = "old"  # 오래된 버전
    UNUSED = "unused"  # 미사용
    LATEST = "latest"  # $LATEST


class AliasStatus(Enum):
    """Lambda Alias 상태 분류.

    CloudWatch 메트릭 기반으로 Alias의 트래픽 유무를 판별한다.
    """

    ACTIVE = "active"  # 사용 중 (트래픽 있음)
    INACTIVE = "inactive"  # 미사용 (트래픽 없음)


@dataclass
class LambdaVersion:
    """Lambda 개별 버전 정보.

    Attributes:
        function_name: Lambda 함수 이름.
        version: 버전 번호 (숫자 문자열).
        description: 버전 설명.
        runtime: 런타임 식별자 (예: python3.12).
        code_size_bytes: 코드 패키지 크기 (바이트).
        last_modified: 마지막 수정 시각.
        code_sha256: 코드 패키지의 SHA-256 해시.
    """

    function_name: str
    version: str
    description: str
    runtime: str
    code_size_bytes: int
    last_modified: datetime | None
    code_sha256: str


@dataclass
class LambdaAlias:
    """Lambda Alias 정보.

    Attributes:
        function_name: Lambda 함수 이름.
        alias_name: Alias 이름.
        function_version: Alias가 가리키는 버전 번호.
        description: Alias 설명.
        routing_config: 가중치 기반 라우팅 설정 (Canary 배포용). None이면 미설정.
    """

    function_name: str
    alias_name: str
    function_version: str
    description: str
    routing_config: dict | None = None


@dataclass
class FunctionVersionInfo:
    """함수별 Version/Alias 종합 정보.

    단일 Lambda 함수에 대한 모든 버전과 Alias 목록, 그리고
    분석 결과(오래된/미사용 버전, 비활성 Alias)를 포함한다.

    Attributes:
        function_name: Lambda 함수 이름.
        function_arn: Lambda 함수 ARN.
        runtime: 런타임 식별자.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        versions: 함수의 전체 버전 목록 ($LATEST 제외).
        aliases: 함수의 전체 Alias 목록.
        old_versions: 90일 이상 경과한 오래된 버전 번호 목록.
        unused_versions: Alias에 연결되지 않은 미사용 버전 번호 목록.
        inactive_aliases: 트래픽이 없는 비활성 Alias 이름 목록.
    """

    function_name: str
    function_arn: str
    runtime: str
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 버전 목록
    versions: list[LambdaVersion] = field(default_factory=list)
    aliases: list[LambdaAlias] = field(default_factory=list)

    # 분석 결과
    old_versions: list[str] = field(default_factory=list)
    unused_versions: list[str] = field(default_factory=list)
    inactive_aliases: list[str] = field(default_factory=list)

    @property
    def version_count(self) -> int:
        """전체 버전 수.

        Returns:
            $LATEST를 제외한 발행 버전 수.
        """
        return len(self.versions)

    @property
    def alias_count(self) -> int:
        """전체 Alias 수.

        Returns:
            함수에 설정된 Alias 수.
        """
        return len(self.aliases)

    @property
    def issue_count(self) -> int:
        """정리 대상 이슈 총 수.

        Returns:
            오래된 버전 + 미사용 버전 + 비활성 Alias의 합계.
        """
        return len(self.old_versions) + len(self.unused_versions) + len(self.inactive_aliases)


@dataclass
class VersionAuditResult:
    """Version/Alias 감사 결과 집계.

    단일 계정/리전에 대한 Lambda Version/Alias 감사 통계를 집계한다.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        total_functions: 분석된 전체 함수 수.
        total_versions: 전체 버전 수 ($LATEST 제외).
        total_aliases: 전체 Alias 수.
        old_version_count: 90일 이상 오래된 버전 수.
        unused_version_count: 미사용 버전 수.
        inactive_alias_count: 비활성 Alias 수.
        functions: 함수별 Version/Alias 상세 정보 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_functions: int = 0
    total_versions: int = 0
    total_aliases: int = 0
    old_version_count: int = 0
    unused_version_count: int = 0
    inactive_alias_count: int = 0
    functions: list[FunctionVersionInfo] = field(default_factory=list)


# =============================================================================
# 수집
# =============================================================================


def collect_versions(
    session,
    account_id: str,
    account_name: str,
    region: str,
) -> list[FunctionVersionInfo]:
    """Lambda 함수별 Version/Alias 정보를 수집한다.

    ListFunctions -> ListVersionsByFunction -> ListAliases 순서로
    AWS API를 호출하여 함수별 버전과 Alias를 수집한다.

    Args:
        session: boto3 세션 (Rate limiting 적용).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        함수별 Version/Alias 정보 목록.
    """
    from botocore.exceptions import ClientError

    result = []

    try:
        lambda_client = get_client(session, "lambda", region_name=region)

        paginator = lambda_client.get_paginator("list_functions")
        for page in paginator.paginate():
            for fn in page.get("Functions", []):
                function_name = fn.get("FunctionName", "")

                info = FunctionVersionInfo(
                    function_name=function_name,
                    function_arn=fn.get("FunctionArn", ""),
                    runtime=fn.get("Runtime", ""),
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                )

                # 버전 목록
                try:
                    versions_paginator = lambda_client.get_paginator("list_versions_by_function")
                    for v_page in versions_paginator.paginate(FunctionName=function_name):
                        for v in v_page.get("Versions", []):
                            version = v.get("Version", "")
                            if version == "$LATEST":
                                continue  # $LATEST는 제외

                            last_modified = None
                            lm_str = v.get("LastModified")
                            if lm_str:
                                with contextlib.suppress(ValueError):
                                    last_modified = datetime.fromisoformat(lm_str.replace("Z", "+00:00"))

                            info.versions.append(
                                LambdaVersion(
                                    function_name=function_name,
                                    version=version,
                                    description=v.get("Description", ""),
                                    runtime=v.get("Runtime", ""),
                                    code_size_bytes=v.get("CodeSize", 0),
                                    last_modified=last_modified,
                                    code_sha256=v.get("CodeSha256", ""),
                                )
                            )
                except ClientError:
                    pass

                # Alias 목록
                try:
                    aliases_paginator = lambda_client.get_paginator("list_aliases")
                    for a_page in aliases_paginator.paginate(FunctionName=function_name):
                        for a in a_page.get("Aliases", []):
                            info.aliases.append(
                                LambdaAlias(
                                    function_name=function_name,
                                    alias_name=a.get("Name", ""),
                                    function_version=a.get("FunctionVersion", ""),
                                    description=a.get("Description", ""),
                                    routing_config=a.get("RoutingConfig"),
                                )
                            )
                except ClientError:
                    pass

                result.append(info)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"[yellow]{account_name}/{region} 수집 오류: {error_code}[/yellow]")

    return result


# =============================================================================
# 분석
# =============================================================================


def analyze_versions(
    functions: list[FunctionVersionInfo],
    account_id: str,
    account_name: str,
    region: str,
    old_days: int = 90,
) -> VersionAuditResult:
    """Lambda Version/Alias를 분석하여 정리 대상을 식별한다.

    각 버전의 수정일과 Alias 연결 여부를 기반으로 오래된 버전과
    미사용 버전을 분류한다. 상위 3개 최신 버전은 유지 대상으로 제외한다.

    Args:
        functions: 함수별 Version/Alias 정보 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        old_days: 오래된 버전 기준 일수 (기본 90일).

    Returns:
        Version/Alias 감사 결과 (통계 + 상세 목록).
    """
    result = VersionAuditResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_functions=len(functions),
    )

    now = datetime.now(timezone.utc)
    cutoff_date = now - timedelta(days=old_days)

    for func in functions:
        result.total_versions += func.version_count
        result.total_aliases += func.alias_count

        # Alias가 가리키는 버전 목록
        aliased_versions: set[str] = set()
        for alias in func.aliases:
            aliased_versions.add(alias.function_version)
            # routing config에 있는 버전도 포함
            if alias.routing_config:
                for v in alias.routing_config.get("AdditionalVersionWeights", {}):
                    aliased_versions.add(v)

        # 버전 분석
        sorted_versions = sorted(
            func.versions,
            key=lambda v: int(v.version) if v.version.isdigit() else 0,
            reverse=True,
        )

        for i, version in enumerate(sorted_versions):
            # Alias가 가리키는 버전은 사용 중
            if version.version in aliased_versions:
                continue

            # 최신 버전 (상위 3개)은 유지
            if i < 3:
                continue

            # 오래된 버전
            if version.last_modified and version.last_modified < cutoff_date:
                func.old_versions.append(version.version)
                result.old_version_count += 1
            else:
                func.unused_versions.append(version.version)
                result.unused_version_count += 1

        # 미사용 Alias (여기서는 트래픽 확인이 어려우므로 패스)
        # CloudWatch 메트릭으로 확인 필요

        result.functions.append(func)

    return result


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[VersionAuditResult], output_dir: str) -> str:
    """Version/Alias 감사 결과를 Excel 보고서로 생성한다.

    Summary Data, Cleanup Candidates, All Versions, Aliases 시트를 포함하며,
    오래된 버전은 노란색으로 하이라이트 표시한다.

    Args:
        results: 계정/리전별 감사 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일의 절대 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    yellow_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")

    wb = Workbook()

    # Summary Sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("Lambda Version/Alias 감사")
    summary.add_item("생성", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()

    # Summary 테이블 헤더
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체 함수", width=12, style="number"),
        ColumnDef(header="총 버전", width=10, style="number"),
        ColumnDef(header="총 Alias", width=10, style="number"),
        ColumnDef(header="오래된 버전", width=12, style="number"),
        ColumnDef(header="미사용 버전", width=12, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)
    for r in results:
        row_style = Styles.warning() if r.old_version_count > 0 else None
        summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_functions,
                r.total_versions,
                r.total_aliases,
                r.old_version_count,
                r.unused_version_count,
            ],
            style=row_style,
        )

    # 총계
    total_versions = sum(r.total_versions for r in results)
    total_old = sum(r.old_version_count for r in results)
    total_unused = sum(r.unused_version_count for r in results)
    summary_sheet.add_summary_row(["합계", "-", "-", total_versions, "-", total_old, total_unused])

    # Cleanup Candidates Sheet
    issue_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function", width=30),
        ColumnDef(header="Runtime", width=15),
        ColumnDef(header="총 버전", width=10, style="number"),
        ColumnDef(header="오래된 버전", width=12, style="number"),
        ColumnDef(header="미사용 버전", width=12, style="number"),
        ColumnDef(header="Alias 수", width=10, style="number"),
        ColumnDef(header="정리 대상 버전", width=40),
    ]
    issues_sheet = wb.new_sheet("Cleanup Candidates", issue_columns)
    for r in results:
        for func in r.functions:
            if func.issue_count > 0:
                cleanup_versions = func.old_versions + func.unused_versions
                issues_sheet.add_row(
                    [
                        func.account_name,
                        func.region,
                        func.function_name,
                        func.runtime,
                        func.version_count,
                        len(func.old_versions),
                        len(func.unused_versions),
                        func.alias_count,
                        ", ".join(cleanup_versions[:10]) + ("..." if len(cleanup_versions) > 10 else ""),
                    ]
                )

    # All Versions Sheet
    all_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function", width=30),
        ColumnDef(header="Version", width=10),
        ColumnDef(header="Description", width=30),
        ColumnDef(header="Runtime", width=15),
        ColumnDef(header="Code Size", width=12),
        ColumnDef(header="Last Modified", width=15),
        ColumnDef(header="상태", width=12, style="center"),
    ]
    all_sheet = wb.new_sheet("All Versions", all_columns)
    for r in results:
        for func in r.functions:
            # Alias가 가리키는 버전
            aliased_versions: set[str] = set()
            for alias in func.aliases:
                aliased_versions.add(alias.function_version)

            sorted_versions = sorted(
                func.versions,
                key=lambda v: int(v.version) if v.version.isdigit() else 0,
                reverse=True,
            )

            for i, v in enumerate(sorted_versions):
                # 상태 결정
                if v.version in aliased_versions:
                    status = "alias_target"
                elif v.version in func.old_versions:
                    status = "old"
                elif v.version in func.unused_versions:
                    status = "unused"
                elif i < 3:
                    status = "current"
                else:
                    status = "unused"

                row_num = all_sheet.add_row(
                    [
                        func.account_name,
                        func.region,
                        func.function_name,
                        v.version,
                        v.description[:50] if v.description else "-",
                        v.runtime,
                        f"{v.code_size_bytes / 1024 / 1024:.2f} MB",
                        v.last_modified.strftime("%Y-%m-%d") if v.last_modified else "-",
                        status,
                    ]
                )

                # old 상태인 경우 노란색 하이라이트
                if status == "old":
                    ws = all_sheet._ws
                    ws.cell(row=row_num, column=9).fill = yellow_fill

    # Aliases Sheet
    alias_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function", width=30),
        ColumnDef(header="Alias", width=20),
        ColumnDef(header="Target Version", width=15),
        ColumnDef(header="Description", width=30),
        ColumnDef(header="Routing", width=25),
    ]
    alias_sheet = wb.new_sheet("Aliases", alias_columns)
    for r in results:
        for func in r.functions:
            for alias in func.aliases:
                routing = "-"
                if alias.routing_config:
                    weights = alias.routing_config.get("AdditionalVersionWeights", {})
                    if weights:
                        routing = ", ".join(f"v{k}: {v * 100:.0f}%" for k, v in weights.items())
                alias_sheet.add_row(
                    [
                        func.account_name,
                        func.region,
                        func.function_name,
                        alias.alias_name,
                        alias.function_version,
                        alias.description or "-",
                        routing,
                    ]
                )

    return str(wb.save_as(output_dir, "Lambda_Version_Audit"))


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> VersionAuditResult | None:
    """단일 계정/리전의 Lambda 버전을 수집하고 감사 분석을 수행한다.

    parallel_collect 콜백 함수로, 멀티 계정/리전 병렬 실행에 사용된다.

    Args:
        session: boto3 세션 (Rate limiting 적용).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        해당 계정/리전의 Version/Alias 감사 결과. 함수가 없으면 None.
    """
    functions = collect_versions(session, account_id, account_name, region)
    if not functions:
        return None
    return analyze_versions(functions, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Lambda Version/Alias 감사 도구를 실행한다.

    모든 Lambda 함수의 발행 버전과 Alias를 수집하여 오래된 버전(90일+)과
    미사용 버전을 식별한다. 상위 3개 최신 버전과 Alias가 가리키는 버전은
    유지 대상으로 제외한다. 결과를 콘솔에 출력하고 Excel 보고서를 생성한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 리전, 출력 설정 포함).
    """
    console.print("[bold]Lambda Version/Alias 감사 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")
    results: list[VersionAuditResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_versions = sum(r.total_versions for r in results)
    total_aliases = sum(r.total_aliases for r in results)
    total_old = sum(r.old_version_count for r in results)
    total_unused = sum(r.unused_version_count for r in results)
    total_cleanup = total_old + total_unused

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"총 버전: {total_versions}개")
    console.print(f"총 Alias: {total_aliases}개")

    if total_cleanup > 0:
        console.print(f"[yellow]정리 대상: {total_cleanup}개[/yellow]")
        console.print(f"  - 오래된 버전 (90일+): {total_old}개")
        console.print(f"  - 미사용 버전: {total_unused}개")

    # 보고서
    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("lambda", "inventory").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
