"""
functions/analyzers/fn/comprehensive.py - Lambda 종합 분석 보고서

Lambda 함수 종합 분석:
- 런타임 EOL 분석
- 메모리 사용량 최적화
- 비용 분석
- 성능 지표 (에러율, Throttle)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import parallel_collect
from core.shared.aws.lambda_ import (
    EOLStatus,
    LambdaFunctionInfo,
    collect_functions_with_metrics,
    get_recommended_upgrade,
    get_runtime_info,
)
from core.shared.aws.pricing import (
    get_lambda_monthly_cost,
    get_lambda_provisioned_monthly_cost,
)
from core.shared.io.compat import generate_dual_report
from core.shared.io.output import open_in_explorer, print_report_complete
from core.shared.io.output.helpers import create_output_path

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:ListTags",
        "lambda:ListProvisionedConcurrencyConfigs",
        "lambda:GetFunctionConcurrency",
        "cloudwatch:GetMetricStatistics",
    ],
}


class IssueType(Enum):
    """Lambda 종합 분석 이슈 유형 분류.

    런타임 EOL, 메모리, 에러, Throttle, 미사용, Timeout 등
    다양한 카테고리의 이슈를 식별한다.
    """

    RUNTIME_EOL = "runtime_eol"
    MEMORY_OVERSIZED = "memory_oversized"
    MEMORY_UNDERSIZED = "memory_undersized"
    HIGH_ERROR_RATE = "high_error_rate"
    THROTTLED = "throttled"
    UNUSED = "unused"
    TIMEOUT_RISK = "timeout_risk"


class Severity(Enum):
    """이슈 심각도 분류.

    CRITICAL > HIGH > MEDIUM > LOW > INFO 순서로 긴급도를 나타낸다.
    """

    CRITICAL = "critical"
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class LambdaIssue:
    """Lambda 개별 이슈 정보.

    Attributes:
        issue_type: 이슈 유형 (런타임 EOL, 메모리, 에러 등).
        severity: 심각도.
        description: 이슈 설명 (한글).
        recommendation: 권장 조치 설명 (한글).
        potential_savings: 이슈 해결 시 절감 가능 비용 (USD). 해당 없으면 0.
    """

    issue_type: IssueType
    severity: Severity
    description: str
    recommendation: str
    potential_savings: float = 0.0


@dataclass
class LambdaComprehensiveResult:
    """Lambda 개별 함수 종합 분석 결과.

    Attributes:
        function: Lambda 함수 정보.
        issues: 발견된 이슈 목록.
        estimated_monthly_cost: 추정 월간 비용 (USD, PC 비용 포함).
        memory_recommendation: 권장 메모리 크기 (MB). None이면 변경 불필요.
        potential_savings: 전체 이슈 해결 시 절감 가능 총액 (USD).
    """

    function: LambdaFunctionInfo
    issues: list[LambdaIssue] = field(default_factory=list)
    estimated_monthly_cost: float = 0.0
    memory_recommendation: int | None = None
    potential_savings: float = 0.0

    @property
    def has_critical_issues(self) -> bool:
        """CRITICAL 심각도 이슈 존재 여부.

        Returns:
            CRITICAL 이슈가 1개 이상이면 True.
        """
        return any(i.severity == Severity.CRITICAL for i in self.issues)

    @property
    def has_high_issues(self) -> bool:
        """HIGH 심각도 이슈 존재 여부.

        Returns:
            HIGH 이슈가 1개 이상이면 True.
        """
        return any(i.severity == Severity.HIGH for i in self.issues)

    @property
    def issue_count(self) -> int:
        """발견된 이슈 총 수.

        Returns:
            이슈 목록의 길이.
        """
        return len(self.issues)


@dataclass
class ComprehensiveAnalysisResult:
    """종합 분석 결과 집계.

    단일 계정/리전의 Lambda 종합 분석 통계를 집계한다.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        total_functions: 분석된 전체 함수 수.
        functions_with_issues: 이슈가 1개 이상인 함수 수.
        runtime_eol_count: 런타임 EOL 이슈 수.
        memory_issue_count: 메모리 이슈 수 (과다 + 부족).
        error_issue_count: 에러율 이슈 수.
        total_monthly_cost: 전체 함수 월간 비용 합계 (USD).
        potential_savings: 모든 이슈 해결 시 절감 가능 총액 (USD).
        results: 함수별 종합 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_functions: int = 0
    functions_with_issues: int = 0
    runtime_eol_count: int = 0
    memory_issue_count: int = 0
    error_issue_count: int = 0
    total_monthly_cost: float = 0.0
    potential_savings: float = 0.0
    results: list[LambdaComprehensiveResult] = field(default_factory=list)


# =============================================================================
# 분석
# =============================================================================


def analyze_function_comprehensive(
    func: LambdaFunctionInfo,
    region: str,
    memory_stats: dict | None = None,
) -> LambdaComprehensiveResult:
    """단일 Lambda 함수에 대해 종합 분석을 수행한다.

    런타임 EOL, 메모리, 에러율, Throttle, 미사용, Timeout 위험을
    순차적으로 점검하고, 비용과 절감 가능액을 계산한다.

    Args:
        func: Lambda 함수 정보 (메트릭 포함).
        region: AWS 리전 코드 (비용 계산용).
        memory_stats: CloudWatch Logs Insights 메모리 통계. None이면 휴리스틱 사용.

    Returns:
        함수의 종합 분석 결과 (이슈 목록, 비용, 절감액).
    """
    result = LambdaComprehensiveResult(function=func)
    metrics = func.metrics

    # 비용 계산
    if metrics and metrics.invocations > 0:
        result.estimated_monthly_cost = get_lambda_monthly_cost(
            region=region,
            invocations=metrics.invocations,
            avg_duration_ms=metrics.duration_avg_ms,
            memory_mb=func.memory_mb,
        )

    # PC 비용 추가
    if func.provisioned_concurrency > 0:
        pc_cost = get_lambda_provisioned_monthly_cost(
            region=region,
            memory_mb=func.memory_mb,
            provisioned_concurrency=func.provisioned_concurrency,
        )
        result.estimated_monthly_cost += pc_cost

    # 1. 런타임 EOL 분석
    _analyze_runtime_eol(func, result)

    # 2. 메모리 분석
    _analyze_memory(func, result, memory_stats)

    # 3. 에러율 분석
    _analyze_errors(func, result)

    # 4. Throttle 분석
    _analyze_throttles(func, result)

    # 5. 미사용 분석
    _analyze_usage(func, result)

    # 6. Timeout 위험 분석
    _analyze_timeout_risk(func, result)

    # 총 잠재 절감액
    result.potential_savings = sum(i.potential_savings for i in result.issues)

    return result


def _analyze_runtime_eol(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """런타임 EOL 상태를 분석하여 이슈를 추가한다.

    Args:
        func: Lambda 함수 정보.
        result: 이슈를 추가할 종합 분석 결과 객체.
    """
    runtime_info = get_runtime_info(func.runtime)
    if not runtime_info:
        return

    status = runtime_info.status

    if status == EOLStatus.DEPRECATED:
        upgrade = get_recommended_upgrade(func.runtime) or "최신 버전"
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.RUNTIME_EOL,
                severity=Severity.CRITICAL,
                description=f"런타임 지원 종료됨: {runtime_info.name}",
                recommendation=f"{upgrade}로 업그레이드 필요",
            )
        )
    elif status == EOLStatus.CRITICAL:
        days = runtime_info.days_until_deprecation
        upgrade = get_recommended_upgrade(func.runtime) or "최신 버전"
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.RUNTIME_EOL,
                severity=Severity.CRITICAL,
                description=f"런타임 {days}일 내 지원 종료: {runtime_info.name}",
                recommendation=f"{upgrade}로 즉시 업그레이드 권장",
            )
        )
    elif status == EOLStatus.HIGH:
        days = runtime_info.days_until_deprecation
        upgrade = get_recommended_upgrade(func.runtime) or "최신 버전"
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.RUNTIME_EOL,
                severity=Severity.HIGH,
                description=f"런타임 {days}일 내 지원 종료: {runtime_info.name}",
                recommendation=f"{upgrade}로 업그레이드 계획 수립",
            )
        )
    elif status == EOLStatus.MEDIUM:
        days = runtime_info.days_until_deprecation
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.RUNTIME_EOL,
                severity=Severity.MEDIUM,
                description=f"런타임 {days}일 내 지원 종료 예정: {runtime_info.name}",
                recommendation="업그레이드 계획 수립 권장",
            )
        )


def _analyze_memory(
    func: LambdaFunctionInfo,
    result: LambdaComprehensiveResult,
    memory_stats: dict | None = None,
):
    """메모리 사용량을 분석하여 과다/부족 이슈를 추가한다.

    memory_stats가 제공되면 실측 데이터 기반으로, 없으면 Duration 기반
    휴리스틱으로 판단한다.

    Args:
        func: Lambda 함수 정보.
        result: 이슈를 추가할 종합 분석 결과 객체.
        memory_stats: CloudWatch Logs Insights 메모리 통계. None이면 휴리스틱 사용.
    """
    metrics = func.metrics
    if not metrics or metrics.invocations == 0:
        return

    # 메모리 통계가 있으면 사용 (CloudWatch Logs Insights 결과)
    if memory_stats:
        max_used = memory_stats.get("max_memory_used_mb", 0)
        memory_stats.get("avg_memory_used_mb", 0)

        if max_used > 0:
            utilization = max_used / func.memory_mb * 100

            # 과다 할당 (최대 사용량이 50% 미만)
            if utilization < 50 and func.memory_mb > 128:
                recommended = max(128, int(max_used * 1.5))  # 50% 여유
                recommended = (recommended // 64) * 64  # 64MB 단위로 올림
                if recommended < func.memory_mb:
                    # 절감액 계산
                    current_cost = get_lambda_monthly_cost(
                        region=result.function.region,
                        invocations=metrics.invocations,
                        avg_duration_ms=metrics.duration_avg_ms,
                        memory_mb=func.memory_mb,
                    )
                    new_cost = get_lambda_monthly_cost(
                        region=result.function.region,
                        invocations=metrics.invocations,
                        avg_duration_ms=metrics.duration_avg_ms,
                        memory_mb=recommended,
                    )
                    savings = current_cost - new_cost

                    result.memory_recommendation = recommended
                    result.issues.append(
                        LambdaIssue(
                            issue_type=IssueType.MEMORY_OVERSIZED,
                            severity=Severity.MEDIUM,
                            description=f"메모리 과다 할당 (사용률 {utilization:.0f}%, {max_used:.0f}MB/{func.memory_mb}MB)",
                            recommendation=f"{recommended}MB로 축소 권장",
                            potential_savings=savings,
                        )
                    )

            # 부족 (최대 사용량이 90% 이상)
            elif utilization >= 90:
                recommended = int(max_used * 1.3)  # 30% 여유
                recommended = ((recommended // 64) + 1) * 64  # 64MB 단위로 올림
                result.memory_recommendation = recommended
                result.issues.append(
                    LambdaIssue(
                        issue_type=IssueType.MEMORY_UNDERSIZED,
                        severity=Severity.HIGH,
                        description=f"메모리 부족 위험 (사용률 {utilization:.0f}%)",
                        recommendation=f"{recommended}MB로 증가 권장 (OOM 방지)",
                    )
                )

        return

    # 메모리 통계가 없으면 휴리스틱 사용
    # Duration이 매우 높고 메모리가 낮으면 메모리 부족 가능성
    if metrics.duration_max_ms > 10000 and func.memory_mb <= 256:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.MEMORY_UNDERSIZED,
                severity=Severity.LOW,
                description=f"실행 시간 {metrics.duration_max_ms:.0f}ms, 메모리 {func.memory_mb}MB",
                recommendation="메모리 증가 시 성능 개선 가능성 검토",
            )
        )


def _analyze_errors(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """에러율을 분석하여 이슈를 추가한다.

    에러율 10% 이상이면 CRITICAL, 5% 이상이면 HIGH, 1% 이상이면 MEDIUM으로 분류한다.

    Args:
        func: Lambda 함수 정보.
        result: 이슈를 추가할 종합 분석 결과 객체.
    """
    metrics = func.metrics
    if not metrics or metrics.invocations == 0:
        return

    error_rate = metrics.errors / metrics.invocations * 100

    if error_rate >= 10:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.HIGH_ERROR_RATE,
                severity=Severity.CRITICAL,
                description=f"높은 에러율: {error_rate:.1f}% ({metrics.errors:,}/{metrics.invocations:,})",
                recommendation="에러 원인 분석 및 수정 필요",
            )
        )
    elif error_rate >= 5:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.HIGH_ERROR_RATE,
                severity=Severity.HIGH,
                description=f"에러율: {error_rate:.1f}% ({metrics.errors:,}/{metrics.invocations:,})",
                recommendation="에러 로그 분석 권장",
            )
        )
    elif error_rate >= 1:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.HIGH_ERROR_RATE,
                severity=Severity.MEDIUM,
                description=f"에러율: {error_rate:.1f}%",
                recommendation="에러 모니터링 권장",
            )
        )


def _analyze_throttles(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """Throttle 발생을 분석하여 이슈를 추가한다.

    Args:
        func: Lambda 함수 정보.
        result: 이슈를 추가할 종합 분석 결과 객체.
    """
    metrics = func.metrics
    if not metrics or metrics.throttles == 0:
        return

    throttle_rate = metrics.throttles / (metrics.invocations + metrics.throttles) * 100

    if throttle_rate >= 5:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.THROTTLED,
                severity=Severity.HIGH,
                description=f"Throttle 발생: {metrics.throttles:,}회 ({throttle_rate:.1f}%)",
                recommendation="Reserved Concurrency 또는 Provisioned Concurrency 설정 검토",
            )
        )
    elif metrics.throttles >= 100:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.THROTTLED,
                severity=Severity.MEDIUM,
                description=f"Throttle 발생: {metrics.throttles:,}회",
                recommendation="동시성 설정 검토 권장",
            )
        )


def _analyze_usage(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """미사용 상태를 분석하여 이슈를 추가한다.

    30일간 호출이 없는 함수를 탐지하며, PC가 설정된 미사용 함수는
    CRITICAL로 분류한다.

    Args:
        func: Lambda 함수 정보.
        result: 이슈를 추가할 종합 분석 결과 객체.
    """
    metrics = func.metrics
    if not metrics:
        return

    if metrics.invocations == 0:
        # PC 설정된 미사용 함수
        if func.provisioned_concurrency > 0:
            pc_cost = get_lambda_provisioned_monthly_cost(
                region=func.region,
                memory_mb=func.memory_mb,
                provisioned_concurrency=func.provisioned_concurrency,
            )
            result.issues.append(
                LambdaIssue(
                    issue_type=IssueType.UNUSED,
                    severity=Severity.CRITICAL,
                    description=f"미사용 함수 (PC {func.provisioned_concurrency}개 설정됨)",
                    recommendation="PC 해제 또는 함수 삭제",
                    potential_savings=pc_cost,
                )
            )
        else:
            result.issues.append(
                LambdaIssue(
                    issue_type=IssueType.UNUSED,
                    severity=Severity.MEDIUM,
                    description="30일간 호출 없음",
                    recommendation="필요 여부 검토 후 삭제 고려",
                )
            )


def _analyze_timeout_risk(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """Timeout 위험을 분석하여 이슈를 추가한다.

    최대 실행 시간이 Timeout의 80% 이상이면 HIGH, 60% 이상이면 LOW로 분류한다.

    Args:
        func: Lambda 함수 정보.
        result: 이슈를 추가할 종합 분석 결과 객체.
    """
    metrics = func.metrics
    if not metrics or metrics.invocations == 0:
        return

    timeout_ms = func.timeout_seconds * 1000
    max_duration = metrics.duration_max_ms

    # 최대 실행 시간이 Timeout의 80% 이상
    if max_duration >= timeout_ms * 0.8:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.TIMEOUT_RISK,
                severity=Severity.HIGH,
                description=f"Timeout 위험: 최대 {max_duration:.0f}ms (Timeout: {timeout_ms}ms)",
                recommendation=f"Timeout 증가 권장 (현재 {func.timeout_seconds}초)",
            )
        )
    elif max_duration >= timeout_ms * 0.6:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.TIMEOUT_RISK,
                severity=Severity.LOW,
                description=f"Timeout 여유 부족: 최대 {max_duration:.0f}ms (Timeout: {timeout_ms}ms)",
                recommendation="Timeout 설정 검토",
            )
        )


def analyze_comprehensive(
    functions: list[LambdaFunctionInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> ComprehensiveAnalysisResult:
    """Lambda 함수 목록에 대해 종합 분석을 수행하고 결과를 집계한다.

    Args:
        functions: Lambda 함수 정보 목록 (메트릭 포함).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        종합 분석 결과 (통계 + 함수별 상세 결과).
    """
    result = ComprehensiveAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_functions=len(functions),
    )

    for func in functions:
        comp_result = analyze_function_comprehensive(func, region)
        result.results.append(comp_result)

        result.total_monthly_cost += comp_result.estimated_monthly_cost
        result.potential_savings += comp_result.potential_savings

        if comp_result.issue_count > 0:
            result.functions_with_issues += 1

        for issue in comp_result.issues:
            if issue.issue_type == IssueType.RUNTIME_EOL:
                result.runtime_eol_count += 1
            elif issue.issue_type in (
                IssueType.MEMORY_OVERSIZED,
                IssueType.MEMORY_UNDERSIZED,
            ):
                result.memory_issue_count += 1
            elif issue.issue_type == IssueType.HIGH_ERROR_RATE:
                result.error_issue_count += 1

    return result


# =============================================================================
# 보고서
# =============================================================================


def _build_excel(results: list[ComprehensiveAnalysisResult]):
    """종합 분석 결과를 Excel Workbook으로 구성한다.

    Summary Data, Issues, All Functions 시트를 포함하며, 심각도별로
    셀 색상을 다르게 적용한다.

    Args:
        results: 계정/리전별 종합 분석 결과 목록.

    Returns:
        저장 전 Workbook 객체.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    severity_fills = {
        Severity.CRITICAL: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        Severity.HIGH: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        Severity.MEDIUM: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        Severity.LOW: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }

    wb = Workbook()

    # Summary Sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("Lambda 종합 분석 보고서")
    summary.add_item("생성", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()

    # Summary Data Sheet
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체 함수", width=12, style="number"),
        ColumnDef(header="이슈 함수", width=12, style="number"),
        ColumnDef(header="런타임 EOL", width=12, style="number"),
        ColumnDef(header="메모리 이슈", width=12, style="number"),
        ColumnDef(header="에러 이슈", width=12, style="number"),
        ColumnDef(header="월 비용", width=15),
        ColumnDef(header="절감 가능", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)
    for r in results:
        row_style = Styles.warning() if r.functions_with_issues > 0 else None
        summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_functions,
                r.functions_with_issues,
                r.runtime_eol_count,
                r.memory_issue_count,
                r.error_issue_count,
                f"${r.total_monthly_cost:,.2f}",
                f"${r.potential_savings:,.2f}",
            ],
            style=row_style,
        )

    # 총계
    total_functions = sum(r.total_functions for r in results)
    total_issues = sum(r.functions_with_issues for r in results)
    total_cost = sum(r.total_monthly_cost for r in results)
    total_savings = sum(r.potential_savings for r in results)
    summary_sheet.add_summary_row(
        [
            "합계",
            "-",
            total_functions,
            total_issues,
            "-",
            "-",
            "-",
            f"${total_cost:,.2f}",
            f"${total_savings:,.2f}",
        ]
    )

    # Issues Sheet
    issue_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function", width=30),
        ColumnDef(header="Runtime", width=15),
        ColumnDef(header="Memory", width=10, style="number"),
        ColumnDef(header="이슈 유형", width=15),
        ColumnDef(header="심각도", width=12, style="center"),
        ColumnDef(header="설명", width=40),
        ColumnDef(header="권장 조치", width=40),
        ColumnDef(header="절감 가능", width=15),
    ]
    issues_sheet = wb.new_sheet("Issues", issue_columns)
    for r in results:
        for comp in r.results:
            for issue in comp.issues:
                fn = comp.function
                row_num = issues_sheet.add_row(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        fn.runtime,
                        fn.memory_mb,
                        issue.issue_type.value,
                        issue.severity.value,
                        issue.description,
                        issue.recommendation,
                        f"${issue.potential_savings:,.2f}" if issue.potential_savings > 0 else "-",
                    ]
                )

                # 심각도에 따른 셀 하이라이트
                fill = severity_fills.get(issue.severity)
                if fill:
                    ws = issues_sheet._ws
                    ws.cell(row=row_num, column=7).fill = fill

    # All Functions Sheet
    all_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function", width=30),
        ColumnDef(header="Runtime", width=15),
        ColumnDef(header="Memory", width=10, style="number"),
        ColumnDef(header="Timeout", width=10, style="number"),
        ColumnDef(header="Invocations (30d)", width=18, style="number"),
        ColumnDef(header="Avg Duration", width=15),
        ColumnDef(header="Errors", width=10, style="number"),
        ColumnDef(header="Throttles", width=10, style="number"),
        ColumnDef(header="이슈 수", width=10, style="number"),
        ColumnDef(header="월 비용", width=15),
    ]
    all_sheet = wb.new_sheet("All Functions", all_columns)
    for r in results:
        for comp in r.results:
            fn = comp.function
            metrics = fn.metrics
            all_sheet.add_row(
                [
                    fn.account_name,
                    fn.region,
                    fn.function_name,
                    fn.runtime,
                    fn.memory_mb,
                    fn.timeout_seconds,
                    metrics.invocations if metrics else 0,
                    f"{metrics.duration_avg_ms:.1f}ms" if metrics else "-",
                    metrics.errors if metrics else 0,
                    metrics.throttles if metrics else 0,
                    comp.issue_count,
                    f"${comp.estimated_monthly_cost:,.4f}",
                ]
            )

    return wb


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ComprehensiveAnalysisResult:
    """단일 계정/리전의 Lambda 함수를 수집하고 종합 분석을 수행한다.

    parallel_collect 콜백 함수로, 멀티 계정/리전 병렬 실행에 사용된다.

    Args:
        session: boto3 세션 (Rate limiting 적용).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        해당 계정/리전의 Lambda 종합 분석 결과.
    """
    functions = collect_functions_with_metrics(session, account_id, account_name, region)
    return analyze_comprehensive(functions, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Lambda 종합 분석 도구를 실행한다.

    모든 Lambda 함수에 대해 런타임 EOL, 메모리, 에러율, Throttle,
    미사용, Timeout 위험을 종합적으로 분석한다. 결과를 콘솔에 출력하고
    Excel + HTML 듀얼 보고서를 생성한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 리전, 출력 설정 포함).
    """
    console.print("[bold]Lambda 종합 분석 시작...[/bold]\n")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")

    results: list[ComprehensiveAnalysisResult] = result.get_data()

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_functions = sum(r.total_functions for r in results)
    total_issues = sum(r.functions_with_issues for r in results)
    total_runtime = sum(r.runtime_eol_count for r in results)
    total_memory = sum(r.memory_issue_count for r in results)
    total_error = sum(r.error_issue_count for r in results)
    total_cost = sum(r.total_monthly_cost for r in results)
    total_savings = sum(r.potential_savings for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체 Lambda 함수: {total_functions}개")
    console.print(f"이슈 함수: {total_issues}개")
    if total_runtime > 0:
        console.print(f"  [red]런타임 EOL: {total_runtime}개[/red]")
    if total_memory > 0:
        console.print(f"  [yellow]메모리 이슈: {total_memory}개[/yellow]")
    if total_error > 0:
        console.print(f"  [yellow]에러 이슈: {total_error}개[/yellow]")
    console.print(f"총 월간 비용: ${total_cost:,.2f}")
    if total_savings > 0:
        console.print(f"[green]절감 가능: ${total_savings:,.2f}[/green]")

    # HTML용 flat 데이터
    flat_data = []
    for r in results:
        for comp in r.results:
            for issue in comp.issues:
                fn = comp.function
                flat_data.append(
                    {
                        "account_id": fn.account_id,
                        "account_name": fn.account_name,
                        "region": fn.region,
                        "resource_id": fn.function_name,
                        "resource_name": fn.function_name,
                        "status": issue.severity.value,
                        "reason": f"[{issue.issue_type.value}] {issue.description}",
                        "cost": issue.potential_savings,
                    }
                )

    # 보고서
    output_path = create_output_path(ctx, "lambda", "comprehensive")
    report_paths = generate_dual_report(
        ctx,
        data=flat_data,
        output_dir=output_path,
        prefix="Lambda_Comprehensive",
        excel_builder=lambda: _build_excel(results),
        html_config={
            "title": "Lambda 종합 분석",
            "service": "Lambda",
            "tool_name": "comprehensive",
            "total": total_functions,
            "found": total_issues,
            "savings": total_savings,
        },
    )

    print_report_complete(report_paths)
    open_in_explorer(output_path)
