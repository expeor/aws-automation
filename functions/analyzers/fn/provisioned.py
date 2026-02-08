"""
functions/analyzers/fn/provisioned.py - Provisioned Concurrency 분석

Lambda Provisioned Concurrency 최적화 분석:
- 실제 동시 실행 대비 PC 과다 설정 탐지
- 미사용 PC 탐지 (비용 낭비)
- PC 부족 탐지 (Throttle 발생)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.shared.aws.pricing import get_lambda_provisioned_monthly_cost
from core.shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from core.cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:ListProvisionedConcurrencyConfigs",
        "lambda:GetFunctionConcurrency",
        "cloudwatch:GetMetricStatistics",
    ],
}


class PCStatus(Enum):
    """Provisioned Concurrency 최적화 상태 분류.

    활용률과 Throttle 발생 여부를 기준으로 PC 설정의 적정성을 분류한다.
    """

    UNUSED = "unused"  # PC 설정됐으나 미사용
    OVERSIZED = "oversized"  # 과다 설정
    OPTIMAL = "optimal"  # 적정
    UNDERSIZED = "undersized"  # 부족 (Throttle 발생)
    NO_PC = "no_pc"  # PC 미설정


@dataclass
class PCConfig:
    """Provisioned Concurrency 개별 설정 정보.

    Attributes:
        qualifier: PC가 적용된 대상 (버전 번호, Alias 이름, 또는 $LATEST).
        allocated: 할당된 PC 수.
        available: 사용 가능한 PC 수.
        status: PC 프로비저닝 상태 (InProgress, Ready, Failed).
    """

    qualifier: str  # $LATEST, version, alias
    allocated: int
    available: int
    status: str  # InProgress, Ready, Failed


@dataclass
class LambdaPCInfo:
    """Lambda 함수의 Provisioned Concurrency 상세 정보.

    Attributes:
        function_name: Lambda 함수 이름.
        function_arn: Lambda 함수 ARN.
        runtime: 런타임 식별자.
        memory_mb: 메모리 크기 (MB).
        pc_configs: PC 개별 설정 목록 (버전/Alias별).
        total_provisioned: 총 할당 PC 수.
        reserved_concurrency: Reserved Concurrency 설정값. None이면 미설정.
        invocations_30d: 30일간 총 호출 수.
        max_concurrent: 30일간 최대 동시 실행 수.
        avg_concurrent: 30일간 평균 동시 실행 수.
        throttles_30d: 30일간 Throttle 발생 수.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        monthly_cost: PC 월간 비용 (USD).
    """

    function_name: str
    function_arn: str
    runtime: str
    memory_mb: int

    # PC 설정
    pc_configs: list[PCConfig] = field(default_factory=list)
    total_provisioned: int = 0
    reserved_concurrency: int | None = None

    # 메트릭
    invocations_30d: int = 0
    max_concurrent: int = 0
    avg_concurrent: float = 0.0
    throttles_30d: int = 0

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def has_pc(self) -> bool:
        """Provisioned Concurrency 설정 여부.

        Returns:
            PC가 1개 이상 설정되어 있으면 True.
        """
        return self.total_provisioned > 0

    @property
    def utilization_pct(self) -> float:
        """PC 활용률 (%).

        최대 동시 실행 수를 할당된 PC 수로 나누어 계산한다.

        Returns:
            활용률 (0~100%). PC 미설정 시 0.0.
        """
        if self.total_provisioned == 0:
            return 0.0
        return min(100.0, self.max_concurrent / self.total_provisioned * 100)


@dataclass
class PCFinding:
    """개별 함수의 PC 분석 결과.

    Attributes:
        function: Lambda PC 정보.
        status: PC 최적화 상태.
        recommendation: 권장 조치 설명 (한글).
        recommended_pc: 권장 PC 수. None이면 변경 불필요.
        monthly_waste: 완전 낭비 비용 (미사용 PC, USD).
        monthly_savings: 축소 시 절감 가능 비용 (과다 설정, USD).
    """

    function: LambdaPCInfo
    status: PCStatus
    recommendation: str
    recommended_pc: int | None = None
    monthly_waste: float = 0.0
    monthly_savings: float = 0.0


@dataclass
class PCAnalysisResult:
    """PC 분석 결과 집계.

    단일 계정/리전의 Provisioned Concurrency 분석 통계를 집계한다.

    Attributes:
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.
        total_functions: 분석된 전체 함수 수.
        functions_with_pc: PC가 설정된 함수 수.
        unused_pc_count: 미사용 PC 함수 수.
        oversized_pc_count: PC 과다 설정 함수 수.
        undersized_pc_count: PC 부족 (Throttle 발생) 함수 수.
        total_pc_cost: 전체 PC 월간 비용 (USD).
        potential_savings: 절감 가능한 총 비용 (USD).
        findings: 개별 함수별 PC 분석 결과 목록.
    """

    account_id: str
    account_name: str
    region: str
    total_functions: int = 0
    functions_with_pc: int = 0
    unused_pc_count: int = 0
    oversized_pc_count: int = 0
    undersized_pc_count: int = 0
    total_pc_cost: float = 0.0
    potential_savings: float = 0.0
    findings: list[PCFinding] = field(default_factory=list)


# =============================================================================
# 수집
# =============================================================================


def collect_pc_info(
    session,
    account_id: str,
    account_name: str,
    region: str,
) -> list[LambdaPCInfo]:
    """Lambda 함수별 Provisioned Concurrency 정보와 CloudWatch 메트릭을 수집한다.

    ListFunctions -> ListProvisionedConcurrencyConfigs -> GetFunctionConcurrency ->
    CloudWatch 메트릭 순서로 AWS API를 호출한다.

    Args:
        session: boto3 세션 (Rate limiting 적용).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        함수별 PC 상세 정보 목록 (메트릭 포함).
    """
    from botocore.exceptions import ClientError

    functions = []

    try:
        lambda_client = get_client(session, "lambda", region_name=region)
        cloudwatch = get_client(session, "cloudwatch", region_name=region)

        end_time = datetime.now(timezone.utc)
        start_time = end_time - timedelta(days=30)

        paginator = lambda_client.get_paginator("list_functions")
        for page in paginator.paginate():
            for fn in page.get("Functions", []):
                function_name = fn.get("FunctionName", "")

                info = LambdaPCInfo(
                    function_name=function_name,
                    function_arn=fn.get("FunctionArn", ""),
                    runtime=fn.get("Runtime", ""),
                    memory_mb=fn.get("MemorySize", 128),
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                )

                # PC 설정 조회
                try:
                    pc_response = lambda_client.list_provisioned_concurrency_configs(FunctionName=function_name)
                    for pc in pc_response.get("ProvisionedConcurrencyConfigs", []):
                        qualifier = (
                            pc.get("FunctionArn", "").split(":")[-1] if ":" in pc.get("FunctionArn", "") else "$LATEST"
                        )
                        config = PCConfig(
                            qualifier=qualifier,
                            allocated=pc.get("AllocatedProvisionedConcurrentExecutions", 0),
                            available=pc.get("AvailableProvisionedConcurrentExecutions", 0),
                            status=pc.get("Status", ""),
                        )
                        info.pc_configs.append(config)
                        info.total_provisioned += config.allocated
                except ClientError:
                    pass

                # Reserved Concurrency 조회
                try:
                    concurrency = lambda_client.get_function_concurrency(FunctionName=function_name)
                    info.reserved_concurrency = concurrency.get("ReservedConcurrentExecutions")
                except ClientError:
                    pass

                # CloudWatch 메트릭
                dimensions = [{"Name": "FunctionName", "Value": function_name}]

                # Invocations
                try:
                    inv_response = cloudwatch.get_metric_statistics(
                        Namespace="AWS/Lambda",
                        MetricName="Invocations",
                        Dimensions=dimensions,
                        StartTime=start_time,
                        EndTime=end_time,
                        Period=86400 * 30,
                        Statistics=["Sum"],
                    )
                    for dp in inv_response.get("Datapoints", []):
                        info.invocations_30d += int(dp.get("Sum", 0))
                except ClientError:
                    pass

                # ConcurrentExecutions
                try:
                    conc_response = cloudwatch.get_metric_statistics(
                        Namespace="AWS/Lambda",
                        MetricName="ConcurrentExecutions",
                        Dimensions=dimensions,
                        StartTime=start_time,
                        EndTime=end_time,
                        Period=3600,  # 1시간 단위
                        Statistics=["Maximum", "Average"],
                    )
                    for dp in conc_response.get("Datapoints", []):
                        max_val = int(dp.get("Maximum", 0))
                        if max_val > info.max_concurrent:
                            info.max_concurrent = max_val
                        info.avg_concurrent = max(info.avg_concurrent, dp.get("Average", 0))
                except ClientError:
                    pass

                # Throttles
                try:
                    throttle_response = cloudwatch.get_metric_statistics(
                        Namespace="AWS/Lambda",
                        MetricName="Throttles",
                        Dimensions=dimensions,
                        StartTime=start_time,
                        EndTime=end_time,
                        Period=86400 * 30,
                        Statistics=["Sum"],
                    )
                    for dp in throttle_response.get("Datapoints", []):
                        info.throttles_30d += int(dp.get("Sum", 0))
                except ClientError:
                    pass

                # 비용 계산
                if info.total_provisioned > 0:
                    info.monthly_cost = get_lambda_provisioned_monthly_cost(
                        region=region,
                        memory_mb=info.memory_mb,
                        provisioned_concurrency=info.total_provisioned,
                    )

                functions.append(info)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"[yellow]{account_name}/{region} 수집 오류: {error_code}[/yellow]")

    return functions


# =============================================================================
# 분석
# =============================================================================


def analyze_pc(
    functions: list[LambdaPCInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> PCAnalysisResult:
    """Lambda Provisioned Concurrency 최적화 상태를 분석한다.

    활용률, Throttle 발생 여부, 호출 유무를 기반으로 각 함수의
    PC 설정 적정성을 평가하고 권장 조치를 제시한다.

    Args:
        functions: Lambda PC 정보 목록.
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        PC 분석 결과 (통계 + 개별 함수 분석).
    """
    result = PCAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_functions=len(functions),
    )

    for func in functions:
        finding = _analyze_single_pc(func, region)
        result.findings.append(finding)

        if func.has_pc:
            result.functions_with_pc += 1
            result.total_pc_cost += func.monthly_cost

        if finding.status == PCStatus.UNUSED:
            result.unused_pc_count += 1
            result.potential_savings += finding.monthly_waste
        elif finding.status == PCStatus.OVERSIZED:
            result.oversized_pc_count += 1
            result.potential_savings += finding.monthly_savings
        elif finding.status == PCStatus.UNDERSIZED:
            result.undersized_pc_count += 1

    return result


def _analyze_single_pc(func: LambdaPCInfo, region: str) -> PCFinding:
    """개별 Lambda 함수의 PC 설정을 분석한다.

    Args:
        func: Lambda PC 정보.
        region: AWS 리전 코드 (비용 계산용).

    Returns:
        함수의 PC 분석 결과 (상태, 권장 조치, 절감액).
    """

    # PC 미설정
    if not func.has_pc:
        # Throttle 발생했으면 PC 설정 권장
        if func.throttles_30d > 0:
            recommended = max(1, int(func.max_concurrent * 1.2))
            return PCFinding(
                function=func,
                status=PCStatus.UNDERSIZED,
                recommendation=f"Throttle {func.throttles_30d:,}회 발생 - PC {recommended}개 설정 권장",
                recommended_pc=recommended,
            )
        return PCFinding(
            function=func,
            status=PCStatus.NO_PC,
            recommendation="PC 미설정",
        )

    # 미사용 PC
    if func.invocations_30d == 0:
        return PCFinding(
            function=func,
            status=PCStatus.UNUSED,
            recommendation=f"미사용 함수에 PC {func.total_provisioned}개 설정됨 - PC 해제 권장",
            monthly_waste=func.monthly_cost,
        )

    # 활용률 분석
    utilization = func.utilization_pct

    # 과다 설정 (활용률 30% 미만)
    if utilization < 30 and func.max_concurrent < func.total_provisioned:
        recommended = max(1, int(func.max_concurrent * 1.3))  # 30% 여유
        current_cost = func.monthly_cost
        new_cost = get_lambda_provisioned_monthly_cost(
            region=region,
            memory_mb=func.memory_mb,
            provisioned_concurrency=recommended,
        )
        savings = current_cost - new_cost

        return PCFinding(
            function=func,
            status=PCStatus.OVERSIZED,
            recommendation=f"PC 과다 설정 (활용률 {utilization:.0f}%) - {recommended}개로 축소 권장",
            recommended_pc=recommended,
            monthly_savings=savings,
        )

    # 부족 (Throttle 발생)
    if func.throttles_30d > 0:
        recommended = max(func.total_provisioned, int(func.max_concurrent * 1.2))
        return PCFinding(
            function=func,
            status=PCStatus.UNDERSIZED,
            recommendation=f"Throttle {func.throttles_30d:,}회 발생 - PC {recommended}개로 증가 권장",
            recommended_pc=recommended,
        )

    # 적정
    return PCFinding(
        function=func,
        status=PCStatus.OPTIMAL,
        recommendation=f"적정 (활용률 {utilization:.0f}%)",
    )


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[PCAnalysisResult], output_dir: str) -> str:
    """PC 분석 결과를 Excel 보고서로 생성한다.

    Summary Data, PC Functions 시트를 포함하며, 상태별로
    셀 색상을 다르게 적용한다 (빨강=미사용, 노랑=부족, 초록=적정).

    Args:
        results: 계정/리전별 PC 분석 결과 목록.
        output_dir: 보고서 저장 디렉토리 경로.

    Returns:
        생성된 Excel 파일의 절대 경로.
    """
    from openpyxl.styles import PatternFill

    from core.shared.io.excel import ColumnDef, Styles, Workbook

    status_fills = {
        PCStatus.UNUSED: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        PCStatus.OVERSIZED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        PCStatus.UNDERSIZED: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        PCStatus.OPTIMAL: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }

    wb = Workbook()

    # Summary Sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("Lambda Provisioned Concurrency 분석")
    summary.add_item("생성", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()

    # Summary Data Sheet
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체 함수", width=12, style="number"),
        ColumnDef(header="PC 설정", width=10, style="number"),
        ColumnDef(header="미사용 PC", width=12, style="number"),
        ColumnDef(header="과다 설정", width=12, style="number"),
        ColumnDef(header="부족", width=10, style="number"),
        ColumnDef(header="PC 월 비용", width=15),
        ColumnDef(header="절감 가능", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)
    for r in results:
        row_style = Styles.warning() if r.unused_pc_count > 0 or r.oversized_pc_count > 0 else None
        summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_functions,
                r.functions_with_pc,
                r.unused_pc_count,
                r.oversized_pc_count,
                r.undersized_pc_count,
                f"${r.total_pc_cost:,.2f}",
                f"${r.potential_savings:,.2f}",
            ],
            style=row_style,
        )

    # PC Functions Sheet
    pc_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function", width=30),
        ColumnDef(header="Runtime", width=15),
        ColumnDef(header="Memory", width=10, style="number"),
        ColumnDef(header="PC 설정", width=10, style="number"),
        ColumnDef(header="최대 동시성", width=12, style="number"),
        ColumnDef(header="활용률", width=10),
        ColumnDef(header="Throttles", width=10, style="number"),
        ColumnDef(header="월 비용", width=12),
        ColumnDef(header="상태", width=12, style="center"),
        ColumnDef(header="권장 PC", width=10),
        ColumnDef(header="절감 가능", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    pc_sheet = wb.new_sheet("PC Functions", pc_columns)
    for r in results:
        for f in r.findings:
            if f.function.has_pc or f.status == PCStatus.UNDERSIZED:
                fn = f.function
                savings = f.monthly_waste + f.monthly_savings
                row_num = pc_sheet.add_row(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        fn.runtime,
                        fn.memory_mb,
                        fn.total_provisioned,
                        fn.max_concurrent,
                        f"{fn.utilization_pct:.0f}%",
                        fn.throttles_30d,
                        f"${fn.monthly_cost:,.2f}",
                        f.status.value,
                        f.recommended_pc if f.recommended_pc else "-",
                        f"${savings:,.2f}" if savings > 0 else "-",
                        f.recommendation,
                    ]
                )

                # 상태에 따른 셀 하이라이트
                fill = status_fills.get(f.status)
                if fill:
                    ws = pc_sheet._ws
                    ws.cell(row=row_num, column=11).fill = fill

    return str(wb.save_as(output_dir, "Lambda_PC_Analysis"))


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> PCAnalysisResult | None:
    """단일 계정/리전의 Lambda PC 정보를 수집하고 최적화 분석을 수행한다.

    parallel_collect 콜백 함수로, 멀티 계정/리전 병렬 실행에 사용된다.

    Args:
        session: boto3 세션 (Rate limiting 적용).
        account_id: AWS 계정 ID.
        account_name: AWS 계정 이름.
        region: AWS 리전 코드.

    Returns:
        해당 계정/리전의 PC 분석 결과. 함수가 없으면 None.
    """
    functions = collect_pc_info(session, account_id, account_name, region)
    if not functions:
        return None
    return analyze_pc(functions, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Lambda Provisioned Concurrency 분석 도구를 실행한다.

    모든 Lambda 함수의 PC 설정을 수집하고, 활용률/Throttle 기반으로
    미사용, 과다 설정, 부족 상태를 분류한다. PC 비용과 절감 가능액을
    계산하여 콘솔에 출력하고 Excel 보고서를 생성한다.

    Args:
        ctx: 실행 컨텍스트 (인증 정보, 리전, 출력 설정 포함).
    """
    console.print("[bold]Lambda Provisioned Concurrency 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")
    results: list[PCAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_with_pc = sum(r.functions_with_pc for r in results)
    total_unused = sum(r.unused_pc_count for r in results)
    total_oversized = sum(r.oversized_pc_count for r in results)
    total_undersized = sum(r.undersized_pc_count for r in results)
    total_cost = sum(r.total_pc_cost for r in results)
    total_savings = sum(r.potential_savings for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"PC 설정된 함수: {total_with_pc}개")
    console.print(f"총 PC 비용: ${total_cost:,.2f}/월")

    if total_unused > 0:
        console.print(f"[red]미사용 PC: {total_unused}개[/red]")
    if total_oversized > 0:
        console.print(f"[yellow]과다 설정: {total_oversized}개[/yellow]")
    if total_undersized > 0:
        console.print(f"[yellow]부족 (Throttle): {total_undersized}개[/yellow]")
    if total_savings > 0:
        console.print(f"[green]절감 가능: ${total_savings:,.2f}/월[/green]")

    # 보고서
    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("lambda", "cost").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
