"""
plugins/apigateway/unused.py - API Gateway 미사용 분석

유휴/미사용 API 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "apigateway:GET",
        "cloudwatch:GetMetricStatistics",
    ],
}

# 분석 기간 (일) - AWS 권장 14일 이상
ANALYSIS_DAYS = 14

# 저사용 기준: 하루 평균 1회 미만 (AWS Trusted Advisor 기준)
LOW_USAGE_REQUESTS_PER_DAY = 1


class APIStatus(Enum):
    """API 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    NO_STAGES = "no_stages"
    LOW_USAGE = "low_usage"


@dataclass
class APIInfo:
    """API Gateway 정보"""

    account_id: str
    account_name: str
    region: str
    api_id: str
    api_name: str
    api_type: str  # REST, HTTP, WEBSOCKET
    protocol_type: str
    endpoint_type: str
    stage_count: int
    created_date: datetime | None
    # CloudWatch 지표
    total_requests: float = 0.0
    error_4xx: float = 0.0
    error_5xx: float = 0.0


@dataclass
class APIFinding:
    """API 분석 결과"""

    api: APIInfo
    status: APIStatus
    recommendation: str


@dataclass
class APIGatewayAnalysisResult:
    """API Gateway 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_apis: int = 0
    unused_apis: int = 0
    no_stages: int = 0
    low_usage: int = 0
    normal_apis: int = 0
    findings: list[APIFinding] = field(default_factory=list)


def collect_rest_apis(session, account_id: str, account_name: str, region: str) -> list[APIInfo]:
    """REST API 수집 (메트릭 제외)"""
    from botocore.exceptions import ClientError

    apigw = get_client(session, "apigateway", region_name=region)
    apis = []

    try:
        paginator = apigw.get_paginator("get_rest_apis")
        for page in paginator.paginate():
            for api in page.get("items", []):
                api_id = api.get("id", "")
                api_name = api.get("name", "")

                # 스테이지 수 확인
                stage_count = 0
                try:
                    stages = apigw.get_stages(restApiId=api_id)
                    stage_count = len(stages.get("item", []))
                except ClientError as e:
                    category = categorize_error(e)
                    if category != ErrorCategory.NOT_FOUND:
                        logger.debug(f"REST API 스테이지 조회 실패: {api_id} ({get_error_code(e)})")

                # 엔드포인트 타입
                endpoint_config = api.get("endpointConfiguration", {})
                endpoint_types = endpoint_config.get("types", [])
                endpoint_type = ", ".join(endpoint_types) if endpoint_types else "EDGE"

                info = APIInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    api_id=api_id,
                    api_name=api_name,
                    api_type="REST",
                    protocol_type="REST",
                    endpoint_type=endpoint_type,
                    stage_count=stage_count,
                    created_date=api.get("createdDate"),
                )
                apis.append(info)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"REST API 권한 없음: {account_name}/{region}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"REST API 조회 오류: {get_error_code(e)}")

    return apis


def collect_http_apis(session, account_id: str, account_name: str, region: str) -> list[APIInfo]:
    """HTTP API (API Gateway v2) 수집 (메트릭 제외)"""
    from botocore.exceptions import ClientError

    apigwv2 = get_client(session, "apigatewayv2", region_name=region)
    apis = []

    try:
        paginator = apigwv2.get_paginator("get_apis")
        for page in paginator.paginate():
            for api in page.get("Items", []):
                api_id = api.get("ApiId", "")
                api_name = api.get("Name", "")
                protocol = api.get("ProtocolType", "HTTP")

                # 스테이지 수 확인
                stage_count = 0
                try:
                    stages = apigwv2.get_stages(ApiId=api_id)
                    stage_count = len(stages.get("Items", []))
                except ClientError as e:
                    category = categorize_error(e)
                    if category != ErrorCategory.NOT_FOUND:
                        logger.debug(f"HTTP API 스테이지 조회 실패: {api_id} ({get_error_code(e)})")

                info = APIInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    api_id=api_id,
                    api_name=api_name,
                    api_type="HTTP" if protocol == "HTTP" else "WEBSOCKET",
                    protocol_type=protocol,
                    endpoint_type="REGIONAL",
                    stage_count=stage_count,
                    created_date=api.get("CreatedDate"),
                )
                apis.append(info)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"HTTP API 권한 없음: {account_name}/{region}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"HTTP API 조회 오류: {get_error_code(e)}")

    return apis


def collect_apis(session, account_id: str, account_name: str, region: str) -> list[APIInfo]:
    """모든 API Gateway 수집 (배치 메트릭 최적화)

    최적화:
    - 기존: API당 1-3 API 호출 → 최적화: 전체 1-2 API 호출
    """
    rest_apis = collect_rest_apis(session, account_id, account_name, region)
    http_apis = collect_http_apis(session, account_id, account_name, region)

    all_apis = rest_apis + http_apis

    # 스테이지가 있는 API만 메트릭 조회
    apis_with_stages = [api for api in all_apis if api.stage_count > 0]
    if apis_with_stages:
        cloudwatch = get_client(session, "cloudwatch", region_name=region)
        now = datetime.now(timezone.utc)
        start_time = now - timedelta(days=ANALYSIS_DAYS)
        _collect_apigateway_metrics_batch(cloudwatch, apis_with_stages, start_time, now)

    return all_apis


def _collect_apigateway_metrics_batch(
    cloudwatch,
    apis: list[APIInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """API Gateway 메트릭 배치 수집 (내부 함수)

    최적화:
    - REST API: Count, 4XXError, 5XXError (ApiName 차원)
    - HTTP/WebSocket API: Count (ApiId 차원)
    """
    from botocore.exceptions import ClientError

    queries: list[MetricQuery] = []

    for api in apis:
        safe_id = sanitize_metric_id(api.api_id)

        if api.api_type == "REST":
            # REST API는 ApiName 차원 사용
            queries.extend(
                [
                    MetricQuery(
                        id=f"{safe_id}_count",
                        namespace="AWS/ApiGateway",
                        metric_name="Count",
                        dimensions={"ApiName": api.api_name},
                        stat="Sum",
                    ),
                    MetricQuery(
                        id=f"{safe_id}_4xx",
                        namespace="AWS/ApiGateway",
                        metric_name="4XXError",
                        dimensions={"ApiName": api.api_name},
                        stat="Sum",
                    ),
                    MetricQuery(
                        id=f"{safe_id}_5xx",
                        namespace="AWS/ApiGateway",
                        metric_name="5XXError",
                        dimensions={"ApiName": api.api_name},
                        stat="Sum",
                    ),
                ]
            )
        else:
            # HTTP/WebSocket API는 ApiId 차원 사용
            queries.append(
                MetricQuery(
                    id=f"{safe_id}_count",
                    namespace="AWS/ApiGateway",
                    metric_name="Count",
                    dimensions={"ApiId": api.api_id},
                    stat="Sum",
                )
            )

    if not queries:
        return

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        for api in apis:
            safe_id = sanitize_metric_id(api.api_id)
            api.total_requests = results.get(f"{safe_id}_count", 0.0)
            if api.api_type == "REST":
                api.error_4xx = results.get(f"{safe_id}_4xx", 0.0)
                api.error_5xx = results.get(f"{safe_id}_5xx", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_apis(apis: list[APIInfo], account_id: str, account_name: str, region: str) -> APIGatewayAnalysisResult:
    """API Gateway 분석"""
    result = APIGatewayAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_apis=len(apis),
    )

    for api in apis:
        # 스테이지 없음
        if api.stage_count == 0:
            result.no_stages += 1
            result.findings.append(
                APIFinding(
                    api=api,
                    status=APIStatus.NO_STAGES,
                    recommendation="스테이지 없음 - 삭제 검토",
                )
            )
            continue

        # 미사용
        if api.total_requests == 0:
            result.unused_apis += 1
            result.findings.append(
                APIFinding(
                    api=api,
                    status=APIStatus.UNUSED,
                    recommendation="요청 없음 - 삭제 검토",
                )
            )
            continue

        # 저사용 (AWS Trusted Advisor 기준: 하루 평균 1회 미만)
        avg_daily = api.total_requests / ANALYSIS_DAYS
        if avg_daily < LOW_USAGE_REQUESTS_PER_DAY:
            result.low_usage += 1
            result.findings.append(
                APIFinding(
                    api=api,
                    status=APIStatus.LOW_USAGE,
                    recommendation=f"저사용 (일 평균 {avg_daily:.1f}회) - 통합 검토",
                )
            )
            continue

        result.normal_apis += 1
        result.findings.append(
            APIFinding(
                api=api,
                status=APIStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[APIGatewayAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from shared.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="스테이지없음", width=12, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_apis,
                r.unused_apis,
                r.no_stages,
                r.low_usage,
                r.normal_apis,
            ]
        )
        ws = summary_sheet._ws
        if r.unused_apis > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.no_stages > 0 or r.low_usage > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="API Name", width=30),
        ColumnDef(header="Type", width=12),
        ColumnDef(header="Endpoint", width=15),
        ColumnDef(header="Stages", width=10, style="number"),
        ColumnDef(header="Requests", width=12, style="number"),
        ColumnDef(header="상태", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("APIs", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != APIStatus.NORMAL:
                a = f.api
                style = Styles.danger() if f.status == APIStatus.UNUSED else Styles.warning()
                detail_sheet.add_row(
                    [
                        a.account_name,
                        a.region,
                        a.api_name,
                        a.api_type,
                        a.endpoint_type,
                        a.stage_count,
                        int(a.total_requests),
                        f.status.value,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "APIGateway_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> APIGatewayAnalysisResult | None:
    """단일 계정/리전의 API Gateway 수집 및 분석 (병렬 실행용)"""
    apis = collect_apis(session, account_id, account_name, region)
    if not apis:
        return None
    return analyze_apis(apis, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """API Gateway 미사용 분석"""
    console.print("[bold]API Gateway 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="apigateway")
    results: list[APIGatewayAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_apis for r in results)
    total_no_stages = sum(r.no_stages for r in results)
    total_low = sum(r.low_usage for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] / "
        f"스테이지없음: [yellow]{total_no_stages}개[/yellow] / "
        f"저사용: {total_low}개"
    )

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("apigateway", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
