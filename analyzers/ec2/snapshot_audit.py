"""
plugins/ec2/snapshot_audit.py - EBS Snapshot 미사용 분석

오래되거나 고아 상태인 EBS Snapshot 탐지

분석 기준:
- 오래된 스냅샷: 90일 이상 경과 (설정 가능)
- 고아 스냅샷: 연결된 AMI가 삭제됨
- 대용량 스냅샷: 100GB 이상

비용:
- $0.05/GB-월 (리전별 차이 있음)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from shared.aws.pricing import get_snapshot_monthly_cost
from shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeSnapshots",
        "ec2:DescribeImages",
    ],
}


# =============================================================================
# 상수
# =============================================================================

# 오래된 스냅샷 기준 (일)
OLD_SNAPSHOT_DAYS = 90


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """사용 상태"""

    ORPHAN = "orphan"  # 고아 (AMI 삭제됨)
    OLD = "old"  # 오래됨
    NORMAL = "normal"  # 정상


class Severity(Enum):
    """심각도"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class SnapshotInfo:
    """EBS Snapshot 정보"""

    id: str
    name: str
    description: str
    state: str
    volume_id: str
    volume_size_gb: int
    start_time: datetime | None
    encrypted: bool
    owner_id: str
    tags: dict[str, str]

    # AMI 연결
    ami_ids: list[str] = field(default_factory=list)

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def age_days(self) -> int:
        """스냅샷 나이 (일)"""
        if not self.start_time:
            return 0
        now = datetime.now(timezone.utc)
        delta = now - self.start_time
        return delta.days

    @property
    def is_old(self) -> bool:
        """오래된 스냅샷 여부"""
        return self.age_days >= OLD_SNAPSHOT_DAYS

    @property
    def has_ami(self) -> bool:
        """AMI 연결 여부"""
        return len(self.ami_ids) > 0


@dataclass
class SnapshotFinding:
    """Snapshot 분석 결과"""

    snapshot: SnapshotInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class SnapshotAnalysisResult:
    """분석 결과"""

    account_id: str
    account_name: str
    region: str
    findings: list[SnapshotFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    orphan_count: int = 0
    old_count: int = 0
    normal_count: int = 0

    # 용량/비용
    total_size_gb: int = 0
    orphan_size_gb: int = 0
    old_size_gb: int = 0
    orphan_monthly_cost: float = 0.0
    old_monthly_cost: float = 0.0


# =============================================================================
# 수집
# =============================================================================


def collect_snapshots(session, account_id: str, account_name: str, region: str) -> list[SnapshotInfo]:
    """EBS Snapshot 목록 수집 (자체 소유만)"""
    from botocore.exceptions import ClientError

    snapshots = []

    try:
        ec2 = get_client(session, "ec2", region_name=region)

        # 자체 소유 스냅샷만 조회
        paginator = ec2.get_paginator("describe_snapshots")
        for page in paginator.paginate(OwnerIds=["self"]):
            for data in page.get("Snapshots", []):
                # 태그 파싱
                tags = {
                    t.get("Key", ""): t.get("Value", "")
                    for t in data.get("Tags", [])
                    if not t.get("Key", "").startswith("aws:")
                }

                size_gb = data.get("VolumeSize", 0)
                monthly_cost = get_snapshot_monthly_cost(region, size_gb)

                snapshot = SnapshotInfo(
                    id=data.get("SnapshotId", ""),
                    name=tags.get("Name", ""),
                    description=data.get("Description", ""),
                    state=data.get("State", ""),
                    volume_id=data.get("VolumeId", ""),
                    volume_size_gb=size_gb,
                    start_time=data.get("StartTime"),
                    encrypted=data.get("Encrypted", False),
                    owner_id=data.get("OwnerId", ""),
                    tags=tags,
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    monthly_cost=monthly_cost,
                )
                snapshots.append(snapshot)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} Snapshot 수집 오류: {error_code}[/yellow]")

    return snapshots


def get_ami_snapshot_mapping(session, region: str) -> dict[str, list[str]]:
    """AMI가 사용하는 스냅샷 ID 매핑 조회

    Returns:
        {snapshot_id: [ami_id, ...]}
    """
    from botocore.exceptions import ClientError

    mapping: dict[str, list[str]] = {}

    try:
        ec2 = get_client(session, "ec2", region_name=region)

        # 자체 소유 AMI만 조회
        paginator = ec2.get_paginator("describe_images")
        for page in paginator.paginate(Owners=["self"]):
            for ami in page.get("Images", []):
                ami_id = ami.get("ImageId", "")

                # BlockDeviceMappings에서 스냅샷 ID 추출
                for bdm in ami.get("BlockDeviceMappings", []):
                    ebs = bdm.get("Ebs", {})
                    snapshot_id = ebs.get("SnapshotId", "")
                    if snapshot_id:
                        if snapshot_id not in mapping:
                            mapping[snapshot_id] = []
                        mapping[snapshot_id].append(ami_id)

    except ClientError:
        pass

    return mapping


# =============================================================================
# 분석
# =============================================================================


def analyze_snapshots(
    snapshots: list[SnapshotInfo],
    ami_mapping: dict[str, list[str]],
    account_id: str,
    account_name: str,
    region: str,
) -> SnapshotAnalysisResult:
    """Snapshot 미사용 분석"""
    result = SnapshotAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for snapshot in snapshots:
        # AMI 연결 정보 설정
        snapshot.ami_ids = ami_mapping.get(snapshot.id, [])

        finding = _analyze_single_snapshot(snapshot)
        result.findings.append(finding)
        result.total_size_gb += snapshot.volume_size_gb

        if finding.usage_status == UsageStatus.ORPHAN:
            result.orphan_count += 1
            result.orphan_size_gb += snapshot.volume_size_gb
            result.orphan_monthly_cost += snapshot.monthly_cost
        elif finding.usage_status == UsageStatus.OLD:
            result.old_count += 1
            result.old_size_gb += snapshot.volume_size_gb
            result.old_monthly_cost += snapshot.monthly_cost
        else:
            result.normal_count += 1

    result.total_count = len(snapshots)
    return result


def _analyze_single_snapshot(snapshot: SnapshotInfo) -> SnapshotFinding:
    """개별 스냅샷 분석"""

    # 완료되지 않은 스냅샷
    if snapshot.state != "completed":
        return SnapshotFinding(
            snapshot=snapshot,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"상태: {snapshot.state}",
            recommendation="완료 대기",
        )

    # 고아 스냅샷 (AMI 없음) + 오래됨
    if not snapshot.has_ami and snapshot.is_old:
        # 대용량일수록 심각도 높음
        if snapshot.volume_size_gb >= 500:
            severity = Severity.HIGH
        elif snapshot.volume_size_gb >= 100:
            severity = Severity.MEDIUM
        else:
            severity = Severity.LOW

        return SnapshotFinding(
            snapshot=snapshot,
            usage_status=UsageStatus.ORPHAN,
            severity=severity,
            description=f"고아 스냅샷 ({snapshot.age_days}일, {snapshot.volume_size_gb}GB, ${snapshot.monthly_cost:.2f}/월)",
            recommendation="AMI 삭제 후 남은 스냅샷. 삭제 검토",
        )

    # 고아이지만 최근 생성
    if not snapshot.has_ami and not snapshot.is_old:
        return SnapshotFinding(
            snapshot=snapshot,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.LOW,
            description=f"AMI 없음 ({snapshot.age_days}일)",
            recommendation="수동 생성 스냅샷. 필요성 확인",
        )

    # AMI 있지만 오래됨
    if snapshot.has_ami and snapshot.is_old:
        return SnapshotFinding(
            snapshot=snapshot,
            usage_status=UsageStatus.OLD,
            severity=Severity.LOW,
            description=f"오래된 AMI 스냅샷 ({snapshot.age_days}일, {snapshot.volume_size_gb}GB)",
            recommendation=f"AMI ({', '.join(snapshot.ami_ids[:2])}) 사용 여부 확인",
        )

    # 정상
    return SnapshotFinding(
        snapshot=snapshot,
        usage_status=UsageStatus.NORMAL,
        severity=Severity.INFO,
        description=f"정상 ({snapshot.age_days}일, AMI: {len(snapshot.ami_ids)}개)",
        recommendation="정상 유지",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[SnapshotAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # Summary sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("EBS Snapshot 미사용 분석 보고서")
    summary.add_item("생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()

    totals = {
        "total": sum(r.total_count for r in results),
        "orphan": sum(r.orphan_count for r in results),
        "old": sum(r.old_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "total_size": sum(r.total_size_gb for r in results),
        "orphan_size": sum(r.orphan_size_gb for r in results),
        "old_size": sum(r.old_size_gb for r in results),
        "orphan_cost": sum(r.orphan_monthly_cost for r in results),
        "old_cost": sum(r.old_monthly_cost for r in results),
    }

    summary.add_section("통계")
    summary.add_item("전체 스냅샷", totals["total"])
    summary.add_item("고아 스냅샷", totals["orphan"], highlight="danger" if totals["orphan"] > 0 else None)
    summary.add_item("오래된 스냅샷", totals["old"], highlight="warning" if totals["old"] > 0 else None)
    summary.add_item("정상", totals["normal"])
    summary.add_item("전체 용량 (GB)", totals["total_size"])
    summary.add_item("고아 용량 (GB)", totals["orphan_size"])
    summary.add_item("오래된 용량 (GB)", totals["old_size"])
    summary.add_item("고아 월 비용 ($)", f"${totals['orphan_cost']:.2f}")
    summary.add_item("오래된 월 비용 ($)", f"${totals['old_cost']:.2f}")

    # Findings sheet
    columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Snapshot ID", width=25),
        ColumnDef(header="Name", width=25),
        ColumnDef(header="Usage", width=12),
        ColumnDef(header="Severity", width=10),
        ColumnDef(header="Size (GB)", width=12, style="number"),
        ColumnDef(header="Age (days)", width=12, style="number"),
        ColumnDef(header="Monthly Cost ($)", width=15, style="number"),
        ColumnDef(header="AMI Count", width=10, style="number"),
        ColumnDef(header="Volume ID", width=22),
        ColumnDef(header="Encrypted", width=10),
        ColumnDef(header="Created", width=12),
        ColumnDef(header="Description", width=40),
        ColumnDef(header="Recommendation", width=30),
    ]
    sheet = wb.new_sheet("Findings", columns)

    # 상태별 스타일
    status_fills = {
        UsageStatus.ORPHAN: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.OLD: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # 고아/오래된 것만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.ORPHAN, UsageStatus.OLD):
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.snapshot.monthly_cost, reverse=True)

    for f in all_findings:
        snap = f.snapshot
        style = Styles.danger() if f.usage_status == UsageStatus.ORPHAN else Styles.warning()
        row_num = sheet.add_row(
            [
                snap.account_name,
                snap.region,
                snap.id,
                snap.name,
                f.usage_status.value,
                f.severity.value,
                snap.volume_size_gb,
                snap.age_days,
                round(snap.monthly_cost, 2),
                len(snap.ami_ids),
                snap.volume_id,
                "Yes" if snap.encrypted else "No",
                snap.start_time.strftime("%Y-%m-%d") if snap.start_time else "",
                f.description,
                f.recommendation,
            ],
            style=style,
        )

        # Usage 컬럼에 상태별 색상 적용
        fill = status_fills.get(f.usage_status)
        if fill:
            sheet._ws.cell(row=row_num, column=5).fill = fill

    return str(wb.save_as(output_dir, "Snapshot_Unused"))


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SnapshotAnalysisResult | None:
    """단일 계정/리전의 스냅샷 수집 및 분석 (병렬 실행용)"""
    snapshots = collect_snapshots(session, account_id, account_name, region)
    if not snapshots:
        return None

    ami_mapping = get_ami_snapshot_mapping(session, region)
    return analyze_snapshots(snapshots, ami_mapping, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """EBS Snapshot 미사용 분석 실행 (병렬 처리)"""
    console.print("[bold]EBS Snapshot 미사용 분석 시작...[/bold]")
    console.print(f"  [dim]기준: {OLD_SNAPSHOT_DAYS}일 이상 오래된 스냅샷[/dim]")

    # 병렬 수집
    result = parallel_collect(
        ctx,
        _collect_and_analyze,
        max_workers=20,
        service="ec2",
    )

    # 결과 처리 (None 제외)
    all_results: list[SnapshotAnalysisResult] = [r for r in result.get_data() if r is not None]

    # 진행 상황 출력
    console.print(f"  [dim]수집 완료: 성공 {result.success_count}, 실패 {result.error_count}[/dim]")

    # 에러 요약
    if result.error_count > 0:
        console.print(f"\n[yellow]{result.get_error_summary()}[/yellow]")

    if not all_results:
        console.print("[yellow]분석할 스냅샷 없음[/yellow]")
        return

    # 개별 결과 요약
    for r in all_results:
        if r.orphan_count > 0 or r.old_count > 0:
            parts = []
            if r.orphan_count > 0:
                parts.append(f"고아 {r.orphan_count}개 ({r.orphan_size_gb}GB)")
            if r.old_count > 0:
                parts.append(f"오래됨 {r.old_count}개 ({r.old_size_gb}GB)")
            total_waste = r.orphan_monthly_cost + r.old_monthly_cost
            cost_str = f" (${total_waste:.2f}/월)" if total_waste > 0 else ""
            console.print(f"  {r.account_name}/{r.region}: [red]{', '.join(parts)}{cost_str}[/red]")
        elif r.total_count > 0:
            console.print(
                f"  {r.account_name}/{r.region}: [green]정상 {r.normal_count}개 ({r.total_size_gb}GB)[/green]"
            )

    # 전체 통계
    totals = {
        "total": sum(r.total_count for r in all_results),
        "orphan": sum(r.orphan_count for r in all_results),
        "old": sum(r.old_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "total_size": sum(r.total_size_gb for r in all_results),
        "orphan_size": sum(r.orphan_size_gb for r in all_results),
        "old_size": sum(r.old_size_gb for r in all_results),
        "orphan_cost": sum(r.orphan_monthly_cost for r in all_results),
        "old_cost": sum(r.old_monthly_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 스냅샷: {totals['total']}개 ({totals['total_size']}GB)[/bold]")
    if totals["orphan"] > 0:
        console.print(
            f"  [red bold]고아: {totals['orphan']}개 ({totals['orphan_size']}GB, ${totals['orphan_cost']:.2f}/월)[/red bold]"
        )
    if totals["old"] > 0:
        console.print(
            f"  [yellow]오래됨: {totals['old']}개 ({totals['old_size']}GB, ${totals['old_cost']:.2f}/월)[/yellow]"
        )
    console.print(f"  [green]정상: {totals['normal']}개[/green]")

    total_waste = totals["orphan_cost"] + totals["old_cost"]
    if total_waste > 0:
        console.print(f"\n  [red]총 절감 가능: ${total_waste:.2f}/월[/red]")

    # 보고서
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("ebs", "snapshot").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
