"""
plugins/kinesis/unused.py - Kinesis 미사용 스트림 분석

유휴/저사용 Kinesis Data Streams 탐지 (CloudWatch 지표 기반)

분석 기준:
- Unused: IncomingRecords = 0 (7일간)
- Low Usage: IncomingRecords < 100/day AND GetRecords.Records = 0

비용 구조:
- Provisioned Mode: $0.015/shard-hour ($10.80/shard/month)
- On-Demand Mode: $0.04/hour/stream + $0.08/GB data in

최적화:
- CloudWatch GetMetricData API 사용 (배치 조회)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 분석 기간 (일)
ANALYSIS_DAYS = 7
# 저사용 기준: 일평균 레코드 < 100
LOW_USAGE_RECORDS_THRESHOLD = 100

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "kinesis:DescribeStream",
        "kinesis:ListStreams",
        "kinesis:DescribeStreamSummary",
        "cloudwatch:GetMetricData",
    ],
}


class StreamStatus(Enum):
    """스트림 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"


@dataclass
class KinesisStreamInfo:
    """Kinesis 스트림 정보"""

    account_id: str
    account_name: str
    region: str
    stream_name: str
    stream_arn: str
    stream_mode: str  # PROVISIONED or ON_DEMAND
    shard_count: int
    retention_hours: int
    stream_status: str
    created_at: datetime | None
    # CloudWatch 지표
    incoming_records: float = 0.0
    incoming_bytes: float = 0.0
    get_records: float = 0.0
    consumer_count: int = 0

    def get_estimated_monthly_cost(self, session=None) -> float:
        """월간 비용 추정 (Pricing API 사용)"""
        from analyzers.cost.pricing.kinesis import get_kinesis_monthly_cost

        return get_kinesis_monthly_cost(
            self.region, shard_count=self.shard_count, mode=self.stream_mode, session=session
        )

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (후방 호환용)"""
        return self.get_estimated_monthly_cost()


@dataclass
class KinesisStreamFinding:
    """스트림 분석 결과"""

    stream: KinesisStreamInfo
    status: StreamStatus
    recommendation: str


@dataclass
class KinesisAnalysisResult:
    """Kinesis 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_streams: int = 0
    unused_streams: int = 0
    low_usage_streams: int = 0
    normal_streams: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[KinesisStreamFinding] = field(default_factory=list)


def collect_kinesis_streams(session, account_id: str, account_name: str, region: str) -> list[KinesisStreamInfo]:
    """Kinesis 스트림 수집 (배치 메트릭 최적화)"""
    from botocore.exceptions import ClientError

    kinesis = get_client(session, "kinesis", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)

    streams: list[KinesisStreamInfo] = []
    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 스트림 목록 조회
        paginator = kinesis.get_paginator("list_streams")
        stream_names: list[str] = []
        for page in paginator.paginate():
            stream_names.extend(page.get("StreamNames", []))

        if not stream_names:
            return []

        # 2단계: 스트림 상세 정보 조회
        for stream_name in stream_names:
            try:
                response = kinesis.describe_stream_summary(StreamName=stream_name)
                desc = response.get("StreamDescriptionSummary", {})

                # Enhanced fan-out 소비자 수 조회
                consumer_count = 0
                try:
                    consumers = kinesis.list_stream_consumers(StreamARN=desc.get("StreamARN", ""))
                    consumer_count = len(consumers.get("Consumers", []))
                except ClientError:
                    pass

                stream = KinesisStreamInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    stream_name=stream_name,
                    stream_arn=desc.get("StreamARN", ""),
                    stream_mode=desc.get("StreamModeDetails", {}).get("StreamMode", "PROVISIONED"),
                    shard_count=desc.get("OpenShardCount", 0),
                    retention_hours=desc.get("RetentionPeriodHours", 24),
                    stream_status=desc.get("StreamStatus", ""),
                    created_at=desc.get("StreamCreationTimestamp"),
                    consumer_count=consumer_count,
                )
                streams.append(stream)
            except ClientError:
                pass

    except ClientError:
        pass

    # 배치 메트릭 조회
    if streams:
        _collect_kinesis_metrics_batch(cloudwatch, streams, start_time, now)

    return streams


def _collect_kinesis_metrics_batch(
    cloudwatch,
    streams: list[KinesisStreamInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """Kinesis 스트림 메트릭 배치 수집 (내부 함수)"""
    from botocore.exceptions import ClientError

    metrics_to_fetch = [
        ("IncomingRecords", "incoming_records"),
        ("IncomingBytes", "incoming_bytes"),
        ("GetRecords.Records", "get_records"),
    ]

    queries = []
    for stream in streams:
        safe_id = sanitize_metric_id(stream.stream_name)
        for metric_name, _ in metrics_to_fetch:
            metric_key = metric_name.replace(".", "_").lower()
            queries.append(
                MetricQuery(
                    id=f"{safe_id}_{metric_key}",
                    namespace="AWS/Kinesis",
                    metric_name=metric_name,
                    dimensions={"StreamName": stream.stream_name},
                    stat="Sum",
                )
            )

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for stream in streams:
            safe_id = sanitize_metric_id(stream.stream_name)
            for metric_name, attr_name in metrics_to_fetch:
                metric_key = metric_name.replace(".", "_").lower()
                # 합계를 일 평균으로 변환
                value = results.get(f"{safe_id}_{metric_key}", 0.0) / days
                setattr(stream, attr_name, value)

    except ClientError:
        pass


def analyze_streams(
    streams: list[KinesisStreamInfo], account_id: str, account_name: str, region: str
) -> KinesisAnalysisResult:
    """Kinesis 스트림 분석"""
    result = KinesisAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_streams=len(streams),
    )

    for stream in streams:
        # 미사용: 입력 레코드 = 0
        if stream.incoming_records == 0:
            result.unused_streams += 1
            result.unused_monthly_cost += stream.estimated_monthly_cost
            result.findings.append(
                KinesisStreamFinding(
                    stream=stream,
                    status=StreamStatus.UNUSED,
                    recommendation=f"미사용 (레코드 0) - 삭제 검토 (${stream.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: 입력 레코드 < 100/day AND 읽기 없음
        if stream.incoming_records < LOW_USAGE_RECORDS_THRESHOLD and stream.get_records == 0:
            result.low_usage_streams += 1
            result.low_usage_monthly_cost += stream.estimated_monthly_cost
            result.findings.append(
                KinesisStreamFinding(
                    stream=stream,
                    status=StreamStatus.LOW_USAGE,
                    recommendation=f"저사용 (레코드 {stream.incoming_records:.0f}/일, 소비 없음) - 삭제/통합 검토",
                )
            )
            continue

        result.normal_streams += 1
        result.findings.append(
            KinesisStreamFinding(
                stream=stream,
                status=StreamStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[KinesisAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_streams,
                r.unused_streams,
                r.low_usage_streams,
                r.normal_streams,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
            ]
        )
        ws = summary_sheet._ws
        if r.unused_streams > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_streams > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Stream Name", width=35),
        ColumnDef(header="Mode", width=12),
        ColumnDef(header="Shards", width=8, style="number"),
        ColumnDef(header="Retention", width=12),
        ColumnDef(header="Status", width=12),
        ColumnDef(header="Avg Records/Day", width=15),
        ColumnDef(header="Avg Bytes/Day", width=15),
        ColumnDef(header="Consumers", width=10, style="number"),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="분석상태", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Streams", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != StreamStatus.NORMAL:
                s = f.stream
                style = Styles.danger() if f.status == StreamStatus.UNUSED else Styles.warning()

                detail_sheet.add_row(
                    [
                        s.account_name,
                        s.region,
                        s.stream_name,
                        s.stream_mode,
                        s.shard_count,
                        f"{s.retention_hours}h",
                        s.stream_status,
                        f"{s.incoming_records:,.0f}",
                        f"{s.incoming_bytes / (1024 * 1024):,.2f} MB",
                        s.consumer_count,
                        f"${s.estimated_monthly_cost:.2f}",
                        f.status.value,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "Kinesis_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> KinesisAnalysisResult | None:
    """단일 계정/리전의 Kinesis 스트림 수집 및 분석 (병렬 실행용)"""
    streams = collect_kinesis_streams(session, account_id, account_name, region)
    if not streams:
        return None
    return analyze_streams(streams, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Kinesis 미사용 스트림 분석"""
    console.print("[bold]Kinesis 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="kinesis")
    results: list[KinesisAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_streams for r in results)
    total_low = sum(r.low_usage_streams for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월)"
    )

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("kinesis", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
