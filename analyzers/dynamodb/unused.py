"""
plugins/dynamodb/unused.py - DynamoDB 미사용 테이블 분석

유휴/저사용 DynamoDB 테이블 탐지 (CloudWatch 지표 기반)

탐지 기준:
- 미사용: ConsumedReadCapacityUnits = 0 AND ConsumedWriteCapacityUnits = 0 (7-30일)
- 저사용: Consumed < 30% of Provisioned (AWS nOps 기준)
- https://docs.aws.amazon.com/amazondynamodb/latest/developerguide/metrics-dimensions.html
- https://www.nops.io/underutilized-30-read-write-dynamodb-tables/

CloudWatch 메트릭:
- Namespace: AWS/DynamoDB
- 메트릭: ConsumedReadCapacityUnits, ConsumedWriteCapacityUnits, ThrottledRequests
- Dimension: TableName

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from shared.aws.pricing.dynamodb import get_dynamodb_monthly_cost

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "dynamodb:ListTables",
        "dynamodb:DescribeTable",
        "cloudwatch:GetMetricStatistics",
    ],
}

# 분석 기간 (일)
ANALYSIS_DAYS = 7
# 저사용 기준: 평균 소비 용량이 프로비저닝의 30% 미만 (AWS nOps 기준)
LOW_USAGE_THRESHOLD_PERCENT = 30


class TableStatus(Enum):
    """테이블 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"


@dataclass
class TableInfo:
    """DynamoDB 테이블 정보"""

    account_id: str
    account_name: str
    region: str
    table_name: str
    table_status: str
    billing_mode: str  # PROVISIONED or PAY_PER_REQUEST
    item_count: int
    size_bytes: int
    provisioned_read: int = 0
    provisioned_write: int = 0
    # CloudWatch 지표
    consumed_read: float = 0.0
    consumed_write: float = 0.0
    throttled_requests: float = 0.0
    created_at: datetime | None = None

    @property
    def size_mb(self) -> float:
        """테이블 크기 (MB)"""
        return self.size_bytes / (1024 * 1024)

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (pricing 모듈 사용)"""
        storage_gb = self.size_bytes / (1024**3)

        if self.billing_mode == "PAY_PER_REQUEST":
            # On-Demand: 일평균 소비량 * 30일 * 86400초
            seconds_per_month = 30 * 24 * 3600
            return get_dynamodb_monthly_cost(
                region=self.region,
                billing_mode="PAY_PER_REQUEST",
                read_requests=int(self.consumed_read * seconds_per_month),
                write_requests=int(self.consumed_write * seconds_per_month),
                storage_gb=storage_gb,
            )
        else:
            # Provisioned
            return get_dynamodb_monthly_cost(
                region=self.region,
                billing_mode="PROVISIONED",
                rcu=self.provisioned_read,
                wcu=self.provisioned_write,
                storage_gb=storage_gb,
            )


@dataclass
class TableFinding:
    """테이블 분석 결과"""

    table: TableInfo
    status: TableStatus
    recommendation: str


@dataclass
class DynamoDBAnalysisResult:
    """DynamoDB 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_tables: int = 0
    unused_tables: int = 0
    low_usage_tables: int = 0
    normal_tables: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[TableFinding] = field(default_factory=list)


def collect_dynamodb_tables(session, account_id: str, account_name: str, region: str) -> list[TableInfo]:
    """DynamoDB 테이블 수집"""
    from botocore.exceptions import ClientError

    dynamodb = get_client(session, "dynamodb", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    tables = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        paginator = dynamodb.get_paginator("list_tables")
        for page in paginator.paginate():
            for table_name in page.get("TableNames", []):
                try:
                    # 테이블 상세 정보 조회
                    desc = dynamodb.describe_table(TableName=table_name)
                    t = desc.get("Table", {})

                    billing = t.get("BillingModeSummary", {})
                    billing_mode = billing.get("BillingMode", "PROVISIONED")

                    throughput = t.get("ProvisionedThroughput", {})

                    table = TableInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        table_name=table_name,
                        table_status=t.get("TableStatus", ""),
                        billing_mode=billing_mode,
                        item_count=t.get("ItemCount", 0),
                        size_bytes=t.get("TableSizeBytes", 0),
                        provisioned_read=throughput.get("ReadCapacityUnits", 0),
                        provisioned_write=throughput.get("WriteCapacityUnits", 0),
                        created_at=t.get("CreationDateTime"),
                    )

                    # CloudWatch 지표 조회
                    try:
                        # ConsumedReadCapacityUnits
                        read_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ConsumedReadCapacityUnits",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if read_resp.get("Datapoints"):
                            table.consumed_read = sum(d["Sum"] for d in read_resp["Datapoints"]) / len(
                                read_resp["Datapoints"]
                            )

                        # ConsumedWriteCapacityUnits
                        write_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ConsumedWriteCapacityUnits",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if write_resp.get("Datapoints"):
                            table.consumed_write = sum(d["Sum"] for d in write_resp["Datapoints"]) / len(
                                write_resp["Datapoints"]
                            )

                        # ThrottledRequests
                        throttle_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ThrottledRequests",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if throttle_resp.get("Datapoints"):
                            table.throttled_requests = sum(d["Sum"] for d in throttle_resp["Datapoints"])

                    except ClientError:
                        pass

                    tables.append(table)

                except ClientError:
                    continue

    except ClientError:
        pass

    return tables


def analyze_tables(tables: list[TableInfo], account_id: str, account_name: str, region: str) -> DynamoDBAnalysisResult:
    """DynamoDB 테이블 분석"""
    result = DynamoDBAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_tables=len(tables),
    )

    for table in tables:
        # 미사용: 읽기/쓰기 모두 0
        if table.consumed_read == 0 and table.consumed_write == 0:
            result.unused_tables += 1
            result.unused_monthly_cost += table.estimated_monthly_cost
            result.findings.append(
                TableFinding(
                    table=table,
                    status=TableStatus.UNUSED,
                    recommendation=f"읽기/쓰기 없음 - 삭제 검토 (${table.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: Provisioned 모드에서 사용량이 프로비저닝의 10% 미만
        if table.billing_mode == "PROVISIONED":
            read_usage = (table.consumed_read / table.provisioned_read * 100) if table.provisioned_read > 0 else 0
            write_usage = (table.consumed_write / table.provisioned_write * 100) if table.provisioned_write > 0 else 0

            if read_usage < LOW_USAGE_THRESHOLD_PERCENT and write_usage < LOW_USAGE_THRESHOLD_PERCENT:
                result.low_usage_tables += 1
                result.low_usage_monthly_cost += table.estimated_monthly_cost
                result.findings.append(
                    TableFinding(
                        table=table,
                        status=TableStatus.LOW_USAGE,
                        recommendation=f"저사용 (R:{read_usage:.1f}%, W:{write_usage:.1f}%) - On-Demand 전환 또는 용량 축소 검토",
                    )
                )
                continue

        result.normal_tables += 1
        result.findings.append(
            TableFinding(
                table=table,
                status=TableStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[DynamoDBAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_style = None
        if r.unused_tables > 0:
            row_style = Styles.danger()
        elif r.low_usage_tables > 0:
            row_style = Styles.warning()
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_tables,
                r.unused_tables,
                r.low_usage_tables,
                r.normal_tables,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
            ],
            style=row_style,
        )
        # Cell-level highlighting
        ws = summary_sheet._ws
        if r.unused_tables > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_tables > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Table Name", width=30),
        ColumnDef(header="Status", width=12, style="center"),
        ColumnDef(header="Billing Mode", width=15),
        ColumnDef(header="Items", width=12, style="number"),
        ColumnDef(header="Size (MB)", width=12),
        ColumnDef(header="Prov. RCU", width=12, style="number"),
        ColumnDef(header="Prov. WCU", width=12, style="number"),
        ColumnDef(header="Consumed R", width=12),
        ColumnDef(header="Consumed W", width=12),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Tables", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != TableStatus.NORMAL:
                t = f.table
                row_style = Styles.danger() if f.status == TableStatus.UNUSED else Styles.warning()
                detail_sheet.add_row(
                    [
                        t.account_name,
                        t.region,
                        t.table_name,
                        f.status.value,
                        t.billing_mode,
                        t.item_count,
                        f"{t.size_mb:.2f}",
                        t.provisioned_read,
                        t.provisioned_write,
                        f"{t.consumed_read:.1f}",
                        f"{t.consumed_write:.1f}",
                        f"${t.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ],
                    style=row_style,
                )

    return str(wb.save_as(output_dir, "DynamoDB_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> DynamoDBAnalysisResult | None:
    """단일 계정/리전의 DynamoDB 테이블 수집 및 분석 (병렬 실행용)"""
    tables = collect_dynamodb_tables(session, account_id, account_name, region)
    if not tables:
        return None
    return analyze_tables(tables, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """DynamoDB 미사용 테이블 분석"""
    console.print("[bold]DynamoDB 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="dynamodb")
    results: list[DynamoDBAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_tables for r in results)
    total_low = sum(r.low_usage_tables for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월)"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("dynamodb", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
