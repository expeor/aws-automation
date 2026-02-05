"""
plugins/fsx/unused.py - Amazon FSx 미사용 파일 시스템 분석

유휴/미사용 FSx 파일 시스템 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from core.tools.output import OutputPath, open_in_explorer
from shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from shared.aws.pricing.fsx import get_fsx_monthly_cost

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "fsx:DescribeFileSystems",
        "cloudwatch:GetMetricStatistics",
    ],
}

# 분석 기간 (일) - AWS 권장 14일
ANALYSIS_DAYS = 14

# 저사용 기준: 하루 평균 I/O 작업 100회 미만
LOW_USAGE_OPS_PER_DAY = 100


class FileSystemStatus(Enum):
    """파일 시스템 상태"""

    NORMAL = "normal"
    UNUSED = "unused"  # I/O 없음
    IDLE = "idle"  # 저사용
    CREATING = "creating"
    DELETING = "deleting"
    FAILED = "failed"


@dataclass
class FSxFileSystemInfo:
    """FSx 파일 시스템 정보"""

    account_id: str
    account_name: str
    region: str
    file_system_id: str
    file_system_type: str  # WINDOWS, LUSTRE, ONTAP, OPENZFS
    lifecycle: str  # AVAILABLE, CREATING, DELETING, FAILED, etc.
    storage_capacity_gb: int
    storage_type: str  # SSD, HDD
    throughput_capacity: int  # MBps
    vpc_id: str
    subnet_ids: list[str]
    dns_name: str
    created_time: datetime | None
    # CloudWatch 지표
    data_read_ops: float = 0.0
    data_write_ops: float = 0.0
    metadata_ops: float = 0.0
    data_read_bytes: float = 0.0
    data_write_bytes: float = 0.0
    # 비용
    estimated_monthly_cost: float = 0.0

    @property
    def total_ops(self) -> float:
        return self.data_read_ops + self.data_write_ops + self.metadata_ops

    @property
    def total_bytes(self) -> float:
        return self.data_read_bytes + self.data_write_bytes

    @property
    def avg_daily_ops(self) -> float:
        return self.total_ops / ANALYSIS_DAYS if ANALYSIS_DAYS > 0 else 0

    @property
    def is_active(self) -> bool:
        return self.total_ops > 0


@dataclass
class FSxFinding:
    """파일 시스템 분석 결과"""

    filesystem: FSxFileSystemInfo
    status: FileSystemStatus
    recommendation: str


@dataclass
class FSxAnalysisResult:
    """FSx 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_filesystems: int = 0
    unused_filesystems: int = 0
    idle_filesystems: int = 0
    failed_filesystems: int = 0
    normal_filesystems: int = 0
    total_storage_gb: int = 0
    unused_storage_gb: int = 0
    total_monthly_waste: float = 0.0
    findings: list[FSxFinding] = field(default_factory=list)


def collect_filesystems(session, account_id: str, account_name: str, region: str) -> list[FSxFileSystemInfo]:
    """FSx 파일 시스템 수집 (배치 메트릭 최적화)

    최적화:
    - 기존: 파일시스템당 5 API 호출 → 최적화: 전체 1-2 API 호출
    """
    from botocore.exceptions import ClientError

    fsx = get_client(session, "fsx", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    filesystems: list[FSxFileSystemInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 파일 시스템 목록 수집
        paginator = fsx.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page.get("FileSystems", []):
                fs_id = fs.get("FileSystemId", "")
                fs_type = fs.get("FileSystemType", "")

                # 스토리지 용량
                storage_capacity = fs.get("StorageCapacity", 0)

                # 스토리지 타입
                storage_type = fs.get("StorageType", "SSD")

                # Throughput (Windows, ONTAP)
                throughput = 0
                if fs_type == "WINDOWS":
                    windows_config = fs.get("WindowsConfiguration", {})
                    throughput = windows_config.get("ThroughputCapacity", 0)
                elif fs_type == "ONTAP":
                    ontap_config = fs.get("OntapConfiguration", {})
                    throughput = ontap_config.get("ThroughputCapacity", 0)
                elif fs_type == "OPENZFS":
                    openzfs_config = fs.get("OpenZFSConfiguration", {})
                    throughput = openzfs_config.get("ThroughputCapacity", 0)
                elif fs_type == "LUSTRE":
                    lustre_config = fs.get("LustreConfiguration", {})
                    throughput = lustre_config.get("PerUnitStorageThroughput", 0)

                # 월 비용 계산 (pricing 모듈 사용)
                monthly_cost = get_fsx_monthly_cost(
                    region, fsx_type=fs_type, storage_gb=storage_capacity, storage_type=storage_type, session=session
                )

                info = FSxFileSystemInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    file_system_id=fs_id,
                    file_system_type=fs_type,
                    lifecycle=fs.get("Lifecycle", ""),
                    storage_capacity_gb=storage_capacity,
                    storage_type=storage_type,
                    throughput_capacity=throughput,
                    vpc_id=fs.get("VpcId", ""),
                    subnet_ids=fs.get("SubnetIds", []),
                    dns_name=fs.get("DNSName", ""),
                    created_time=fs.get("CreationTime"),
                    estimated_monthly_cost=monthly_cost,
                )
                filesystems.append(info)

        # 2단계: AVAILABLE 상태인 파일 시스템만 배치 메트릭 조회
        available_fs = [fs for fs in filesystems if fs.lifecycle == "AVAILABLE"]
        if available_fs:
            _collect_fsx_metrics_batch(cloudwatch, available_fs, start_time, now)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"FSx 권한 없음: {account_name}/{region}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"FSx 조회 오류: {get_error_code(e)}")

    return filesystems


def _collect_fsx_metrics_batch(
    cloudwatch,
    filesystems: list[FSxFileSystemInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """FSx 메트릭 배치 수집 (내부 함수)

    메트릭:
    - DataReadOperations (Sum)
    - DataWriteOperations (Sum)
    - MetadataOperations (Sum)
    - DataReadBytes (Sum)
    - DataWriteBytes (Sum)
    """
    from botocore.exceptions import ClientError

    queries: list[MetricQuery] = []

    for fs in filesystems:
        safe_id = sanitize_metric_id(fs.file_system_id)
        dimensions = {"FileSystemId": fs.file_system_id}

        queries.extend([
            MetricQuery(
                id=f"{safe_id}_read_ops",
                namespace="AWS/FSx",
                metric_name="DataReadOperations",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_write_ops",
                namespace="AWS/FSx",
                metric_name="DataWriteOperations",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_metadata_ops",
                namespace="AWS/FSx",
                metric_name="MetadataOperations",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_read_bytes",
                namespace="AWS/FSx",
                metric_name="DataReadBytes",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_write_bytes",
                namespace="AWS/FSx",
                metric_name="DataWriteBytes",
                dimensions=dimensions,
                stat="Sum",
            ),
        ])

    if not queries:
        return

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        for fs in filesystems:
            safe_id = sanitize_metric_id(fs.file_system_id)
            fs.data_read_ops = results.get(f"{safe_id}_read_ops", 0.0)
            fs.data_write_ops = results.get(f"{safe_id}_write_ops", 0.0)
            fs.metadata_ops = results.get(f"{safe_id}_metadata_ops", 0.0)
            fs.data_read_bytes = results.get(f"{safe_id}_read_bytes", 0.0)
            fs.data_write_bytes = results.get(f"{safe_id}_write_bytes", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_filesystems(
    filesystems: list[FSxFileSystemInfo], account_id: str, account_name: str, region: str
) -> FSxAnalysisResult:
    """FSx 파일 시스템 분석"""
    result = FSxAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_filesystems=len(filesystems),
    )

    for fs in filesystems:
        result.total_storage_gb += fs.storage_capacity_gb

        # 실패/삭제 중
        if fs.lifecycle in ("FAILED", "DELETING", "MISCONFIGURED"):
            result.failed_filesystems += 1
            result.findings.append(
                FSxFinding(
                    filesystem=fs,
                    status=FileSystemStatus.FAILED,
                    recommendation=f"상태 이상 ({fs.lifecycle}) - 확인 필요",
                )
            )
            continue

        # 생성 중
        if fs.lifecycle in ("CREATING", "UPDATING"):
            result.findings.append(
                FSxFinding(
                    filesystem=fs,
                    status=FileSystemStatus.CREATING,
                    recommendation="생성/업데이트 중",
                )
            )
            continue

        # 미사용 (I/O 없음)
        if not fs.is_active:
            result.unused_filesystems += 1
            result.unused_storage_gb += fs.storage_capacity_gb
            result.total_monthly_waste += fs.estimated_monthly_cost
            result.findings.append(
                FSxFinding(
                    filesystem=fs,
                    status=FileSystemStatus.UNUSED,
                    recommendation=f"I/O 없음 ({ANALYSIS_DAYS}일간) - 삭제 검토",
                )
            )
            continue

        # 저사용
        if fs.avg_daily_ops < LOW_USAGE_OPS_PER_DAY:
            result.idle_filesystems += 1
            result.total_monthly_waste += fs.estimated_monthly_cost * 0.5  # 50% 낭비 추정
            result.findings.append(
                FSxFinding(
                    filesystem=fs,
                    status=FileSystemStatus.IDLE,
                    recommendation=f"저사용 (일 평균 {fs.avg_daily_ops:.0f} ops) - 축소 검토",
                )
            )
            continue

        result.normal_filesystems += 1
        result.findings.append(
            FSxFinding(
                filesystem=fs,
                status=FileSystemStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[FSxAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="전체(GB)", width=12, style="number"),
        ColumnDef(header="미사용(GB)", width=12, style="number"),
        ColumnDef(header="월 낭비(USD)", width=12, style="currency"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_filesystems,
                r.unused_filesystems,
                r.idle_filesystems,
                r.normal_filesystems,
                r.total_storage_gb,
                r.unused_storage_gb,
                r.total_monthly_waste,
            ]
        )
        ws = summary_sheet._ws
        if r.unused_filesystems > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.idle_filesystems > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="File System ID", width=25),
        ColumnDef(header="Type", width=12),
        ColumnDef(header="Storage(GB)", width=12, style="number"),
        ColumnDef(header="Throughput", width=12, style="number"),
        ColumnDef(header="Lifecycle", width=12),
        ColumnDef(header="Total Ops", width=15, style="number"),
        ColumnDef(header="Daily Avg", width=12, style="number"),
        ColumnDef(header="상태", width=12),
        ColumnDef(header="월 비용(USD)", width=12, style="currency"),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("FileSystems", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status not in (FileSystemStatus.NORMAL, FileSystemStatus.CREATING):
                fs = f.filesystem
                style = Styles.danger() if f.status == FileSystemStatus.UNUSED else Styles.warning()
                detail_sheet.add_row(
                    [
                        fs.account_name,
                        fs.region,
                        fs.file_system_id,
                        fs.file_system_type,
                        fs.storage_capacity_gb,
                        fs.throughput_capacity,
                        fs.lifecycle,
                        int(fs.total_ops),
                        int(fs.avg_daily_ops),
                        f.status.value,
                        fs.estimated_monthly_cost,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "FSx_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> FSxAnalysisResult | None:
    """단일 계정/리전의 FSx 파일 시스템 수집 및 분석 (병렬 실행용)"""
    filesystems = collect_filesystems(session, account_id, account_name, region)
    if not filesystems:
        return None
    return analyze_filesystems(filesystems, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """FSx 미사용 파일 시스템 분석"""
    console.print("[bold]FSx 파일 시스템 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="fsx")
    results: list[FSxAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_filesystems for r in results)
    total_idle = sum(r.idle_filesystems for r in results)
    total_unused_gb = sum(r.unused_storage_gb for r in results)
    total_waste = sum(r.total_monthly_waste for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] ({total_unused_gb:,} GB) / 저사용: [yellow]{total_idle}개[/yellow]"
    )
    console.print(f"예상 월 낭비: [red]${total_waste:,.2f}[/red]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("fsx", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
