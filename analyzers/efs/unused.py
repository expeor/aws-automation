"""
plugins/efs/unused.py - EFS 미사용 파일시스템 분석

유휴/미사용 EFS 파일시스템 탐지 (마운트 타겟 및 CloudWatch 지표 기반)

탐지 기준 (CloudFix 기반):
- 미사용: ClientConnections = 0 AND MeteredIOBytes = 0 (30일)
- 마운트없음: mount_target_count = 0
- 빈파일: size < 1MB
- https://docs.aws.amazon.com/efs/latest/ug/efs-metrics.html
- https://cloudfix.com/blog/delete-idle-volumes-save-on-efs/

CloudWatch 메트릭:
- Namespace: AWS/EFS
- 메트릭: ClientConnections, MeteredIOBytes, DataReadIOBytes, DataWriteIOBytes
- Dimension: FileSystemId

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from core.tools.output import OutputPath, open_in_explorer
from shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 분석 기간 (일) - CloudFix 권장 30일
ANALYSIS_DAYS = 30

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticfilesystem:DescribeFileSystems",
        "elasticfilesystem:DescribeMountTargets",
        "cloudwatch:GetMetricStatistics",
    ],
}


class FileSystemStatus(Enum):
    """파일시스템 상태"""

    NORMAL = "normal"
    NO_MOUNT_TARGET = "no_mount_target"
    NO_IO = "no_io"
    EMPTY = "empty"


@dataclass
class EFSInfo:
    """EFS 파일시스템 정보"""

    account_id: str
    account_name: str
    region: str
    file_system_id: str
    name: str
    lifecycle_state: str
    performance_mode: str
    throughput_mode: str
    size_bytes: int
    mount_target_count: int
    created_at: datetime | None
    # CloudWatch 지표
    avg_client_connections: float = 0.0
    metered_io_bytes: float = 0.0  # 과금 대상 I/O (더 정확)
    data_read_bytes: float = 0.0
    data_write_bytes: float = 0.0

    @property
    def size_gb(self) -> float:
        return self.size_bytes / (1024**3)

    def get_estimated_monthly_cost(self, session=None) -> float:
        """월간 비용 추정 (Pricing API 사용)"""
        from analyzers.cost.pricing.efs import get_efs_monthly_cost

        return get_efs_monthly_cost(self.region, self.size_gb, session=session)

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (후방 호환용)"""
        return self.get_estimated_monthly_cost()


@dataclass
class EFSFinding:
    """EFS 분석 결과"""

    efs: EFSInfo
    status: FileSystemStatus
    recommendation: str


@dataclass
class EFSAnalysisResult:
    """EFS 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_filesystems: int = 0
    no_mount_target: int = 0
    no_io: int = 0
    empty: int = 0
    normal: int = 0
    unused_monthly_cost: float = 0.0
    findings: list[EFSFinding] = field(default_factory=list)


def collect_efs_filesystems(session, account_id: str, account_name: str, region: str) -> list[EFSInfo]:
    """EFS 파일시스템 수집 (배치 메트릭 최적화)

    최적화:
    - 기존: 파일시스템당 4 API 호출 → 최적화: 전체 1-2 API 호출
    """
    from botocore.exceptions import ClientError

    efs_client = get_client(session, "efs", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    filesystems: list[EFSInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 파일시스템 목록 수집
        paginator = efs_client.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page.get("FileSystems", []):
                fs_id = fs.get("FileSystemId", "")

                # 이름 태그 찾기
                name = ""
                for tag in fs.get("Tags", []):
                    if tag.get("Key") == "Name":
                        name = tag.get("Value", "")
                        break

                # 마운트 타겟 수 확인
                mount_target_count = 0
                try:
                    mt_resp = efs_client.describe_mount_targets(FileSystemId=fs_id)
                    mount_target_count = len(mt_resp.get("MountTargets", []))
                except ClientError as e:
                    category = categorize_error(e)
                    if category != ErrorCategory.NOT_FOUND:
                        logger.debug(f"EFS 마운트 타겟 조회 실패: {fs_id} ({get_error_code(e)})")

                info = EFSInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    file_system_id=fs_id,
                    name=name,
                    lifecycle_state=fs.get("LifeCycleState", ""),
                    performance_mode=fs.get("PerformanceMode", ""),
                    throughput_mode=fs.get("ThroughputMode", ""),
                    size_bytes=fs.get("SizeInBytes", {}).get("Value", 0),
                    mount_target_count=mount_target_count,
                    created_at=fs.get("CreationTime"),
                )
                filesystems.append(info)

        # 2단계: 배치 메트릭 조회
        if filesystems:
            _collect_efs_metrics_batch(cloudwatch, filesystems, start_time, now)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"EFS 권한 없음: {account_name}/{region}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"EFS 조회 오류: {get_error_code(e)}")

    return filesystems


def _collect_efs_metrics_batch(
    cloudwatch,
    filesystems: list[EFSInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """EFS 메트릭 배치 수집 (내부 함수)

    최적화:
    - ClientConnections (Average)
    - MeteredIOBytes (Sum)
    - DataReadIOBytes (Sum)
    - DataWriteIOBytes (Sum)
    """
    from botocore.exceptions import ClientError

    queries: list[MetricQuery] = []
    days = (end_time - start_time).days or 1

    for fs in filesystems:
        safe_id = sanitize_metric_id(fs.file_system_id)
        dimensions = {"FileSystemId": fs.file_system_id}

        queries.extend([
            MetricQuery(
                id=f"{safe_id}_conn",
                namespace="AWS/EFS",
                metric_name="ClientConnections",
                dimensions=dimensions,
                stat="Average",
            ),
            MetricQuery(
                id=f"{safe_id}_metered",
                namespace="AWS/EFS",
                metric_name="MeteredIOBytes",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_read",
                namespace="AWS/EFS",
                metric_name="DataReadIOBytes",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_write",
                namespace="AWS/EFS",
                metric_name="DataWriteIOBytes",
                dimensions=dimensions,
                stat="Sum",
            ),
        ])

    if not queries:
        return

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        for fs in filesystems:
            safe_id = sanitize_metric_id(fs.file_system_id)
            # Average는 일별 평균의 합을 일수로 나눔
            fs.avg_client_connections = results.get(f"{safe_id}_conn", 0.0) / days
            fs.metered_io_bytes = results.get(f"{safe_id}_metered", 0.0)
            fs.data_read_bytes = results.get(f"{safe_id}_read", 0.0)
            fs.data_write_bytes = results.get(f"{safe_id}_write", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_filesystems(
    filesystems: list[EFSInfo], account_id: str, account_name: str, region: str
) -> EFSAnalysisResult:
    """EFS 파일시스템 분석"""
    result = EFSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_filesystems=len(filesystems),
    )

    for fs in filesystems:
        # 마운트 타겟 없음
        if fs.mount_target_count == 0:
            result.no_mount_target += 1
            result.unused_monthly_cost += fs.estimated_monthly_cost
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.NO_MOUNT_TARGET,
                    recommendation=f"마운트 타겟 없음 - 삭제 검토 (${fs.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 빈 파일시스템 (거의 0 바이트)
        if fs.size_bytes < 1024 * 1024:  # 1MB 미만
            result.empty += 1
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.EMPTY,
                    recommendation="빈 파일시스템 - 삭제 검토",
                )
            )
            continue

        # I/O 없음 (MeteredIOBytes 사용 - 과금 대상 I/O)
        if fs.metered_io_bytes == 0 and fs.avg_client_connections == 0:
            result.no_io += 1
            result.unused_monthly_cost += fs.estimated_monthly_cost
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.NO_IO,
                    recommendation=f"30일간 I/O 없음 - 삭제 검토 (${fs.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        result.normal += 1
        result.findings.append(
            EFSFinding(
                efs=fs,
                status=FileSystemStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[EFSAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="마운트없음", width=12, style="number"),
        ColumnDef(header="I/O없음", width=10, style="number"),
        ColumnDef(header="빈FS", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=12),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_filesystems,
                r.no_mount_target,
                r.no_io,
                r.empty,
                r.normal,
                f"${r.unused_monthly_cost:,.2f}",
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.no_mount_target > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.no_io > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="ID", width=22),
        ColumnDef(header="Name", width=25),
        ColumnDef(header="Size", width=12),
        ColumnDef(header="Mount Targets", width=12, style="number"),
        ColumnDef(header="Mode", width=12),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="Avg Conn", width=10),
        ColumnDef(header="Metered I/O", width=12),
        ColumnDef(header="Read/Write", width=15),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=35),
    ]
    detail_sheet = wb.new_sheet("FileSystems", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != FileSystemStatus.NORMAL:
                fs = f.efs
                style = Styles.danger() if f.status == FileSystemStatus.NO_MOUNT_TARGET else Styles.warning()

                detail_sheet.add_row(
                    [
                        fs.account_name,
                        fs.region,
                        fs.file_system_id,
                        fs.name or "-",
                        f"{fs.size_gb:.2f} GB",
                        fs.mount_target_count,
                        fs.throughput_mode,
                        f.status.value,
                        f"{fs.avg_client_connections:.1f}",
                        f"{fs.metered_io_bytes / (1024**2):.1f} MB",
                        f"R:{fs.data_read_bytes / (1024**2):.1f}/W:{fs.data_write_bytes / (1024**2):.1f} MB",
                        f"${fs.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "EFS_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EFSAnalysisResult | None:
    """단일 계정/리전의 EFS 수집 및 분석 (병렬 실행용)"""
    filesystems = collect_efs_filesystems(session, account_id, account_name, region)
    if not filesystems:
        return None
    return analyze_filesystems(filesystems, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """EFS 미사용 파일시스템 분석"""
    console.print("[bold]EFS 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="efs")
    results: list[EFSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.no_mount_target + r.no_io + r.empty for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월)")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("efs", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
