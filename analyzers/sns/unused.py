"""
plugins/sns/unused.py - SNS 미사용 토픽 분석

유휴/미사용 SNS 토픽 탐지 (구독자 및 CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 분석 기간 (일) - AWS 권장 14일 이상
ANALYSIS_DAYS = 14

# 저사용 기준: 하루 평균 메시지 10개 미만
LOW_USAGE_MESSAGES_PER_DAY = 10

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "sns:ListTopics",
        "sns:ListSubscriptionsByTopic",
        "cloudwatch:GetMetricStatistics",
    ],
}


class TopicStatus(Enum):
    """토픽 상태"""

    NORMAL = "normal"
    NO_SUBSCRIBERS = "no_subscribers"
    NO_MESSAGES = "no_messages"
    LOW_USAGE = "low_usage"
    UNUSED = "unused"


@dataclass
class SNSTopicInfo:
    """SNS 토픽 정보"""

    account_id: str
    account_name: str
    region: str
    topic_name: str
    topic_arn: str
    subscription_count: int
    # CloudWatch 지표
    messages_published: float = 0.0
    notifications_delivered: float = 0.0
    notifications_failed: float = 0.0


@dataclass
class TopicFinding:
    """토픽 분석 결과"""

    topic: SNSTopicInfo
    status: TopicStatus
    recommendation: str


@dataclass
class SNSAnalysisResult:
    """SNS 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_topics: int = 0
    no_subscribers: int = 0
    no_messages: int = 0
    low_usage: int = 0
    unused_topics: int = 0
    normal_topics: int = 0
    findings: list[TopicFinding] = field(default_factory=list)


def collect_sns_topics(session, account_id: str, account_name: str, region: str) -> list[SNSTopicInfo]:
    """SNS 토픽 수집"""
    from botocore.exceptions import ClientError

    sns = get_client(session, "sns", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    topics = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        paginator = sns.get_paginator("list_topics")
        for page in paginator.paginate():
            for topic in page.get("Topics", []):
                topic_arn = topic.get("TopicArn", "")
                topic_name = topic_arn.split(":")[-1]

                # 구독자 수 확인
                subscription_count = 0
                try:
                    sub_paginator = sns.get_paginator("list_subscriptions_by_topic")
                    for sub_page in sub_paginator.paginate(TopicArn=topic_arn):
                        subscription_count += len(sub_page.get("Subscriptions", []))
                except ClientError:
                    pass

                info = SNSTopicInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    topic_name=topic_name,
                    topic_arn=topic_arn,
                    subscription_count=subscription_count,
                )

                # CloudWatch 지표 조회
                try:
                    # NumberOfMessagesPublished
                    pub_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/SNS",
                        MetricName="NumberOfMessagesPublished",
                        Dimensions=[{"Name": "TopicName", "Value": topic_name}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Sum"],
                    )
                    if pub_resp.get("Datapoints"):
                        info.messages_published = sum(d["Sum"] for d in pub_resp["Datapoints"])

                    # NumberOfNotificationsDelivered
                    del_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/SNS",
                        MetricName="NumberOfNotificationsDelivered",
                        Dimensions=[{"Name": "TopicName", "Value": topic_name}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Sum"],
                    )
                    if del_resp.get("Datapoints"):
                        info.notifications_delivered = sum(d["Sum"] for d in del_resp["Datapoints"])

                    # NumberOfNotificationsFailed
                    fail_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/SNS",
                        MetricName="NumberOfNotificationsFailed",
                        Dimensions=[{"Name": "TopicName", "Value": topic_name}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Sum"],
                    )
                    if fail_resp.get("Datapoints"):
                        info.notifications_failed = sum(d["Sum"] for d in fail_resp["Datapoints"])

                except ClientError:
                    pass

                topics.append(info)

    except ClientError:
        pass

    return topics


def analyze_topics(topics: list[SNSTopicInfo], account_id: str, account_name: str, region: str) -> SNSAnalysisResult:
    """SNS 토픽 분석"""
    result = SNSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_topics=len(topics),
    )

    for topic in topics:
        # 구독자 없고 메시지도 없음 = 미사용
        if topic.subscription_count == 0 and topic.messages_published == 0:
            result.unused_topics += 1
            result.findings.append(
                TopicFinding(
                    topic=topic,
                    status=TopicStatus.UNUSED,
                    recommendation="구독자 없음 + 메시지 없음 - 삭제 검토",
                )
            )
            continue

        # 구독자 없음
        if topic.subscription_count == 0:
            result.no_subscribers += 1
            result.findings.append(
                TopicFinding(
                    topic=topic,
                    status=TopicStatus.NO_SUBSCRIBERS,
                    recommendation="구독자 없음 - 구독 추가 또는 삭제 검토",
                )
            )
            continue

        # 메시지 발행 없음
        if topic.messages_published == 0:
            result.no_messages += 1
            result.findings.append(
                TopicFinding(
                    topic=topic,
                    status=TopicStatus.NO_MESSAGES,
                    recommendation=f"{ANALYSIS_DAYS}일간 메시지 발행 없음 - 사용 여부 확인",
                )
            )
            continue

        # 저사용 (일평균 메시지 수 기준)
        avg_messages_per_day = topic.messages_published / ANALYSIS_DAYS
        if avg_messages_per_day < LOW_USAGE_MESSAGES_PER_DAY:
            result.low_usage += 1
            result.findings.append(
                TopicFinding(
                    topic=topic,
                    status=TopicStatus.LOW_USAGE,
                    recommendation=f"저사용 (일평균 {avg_messages_per_day:.1f}건) - 통합 검토",
                )
            )
            continue

        result.normal_topics += 1
        result.findings.append(
            TopicFinding(
                topic=topic,
                status=TopicStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[SNSAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="구독자없음", width=12, style="number"),
        ColumnDef(header="메시지없음", width=12, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_topics,
                r.unused_topics,
                r.no_subscribers,
                r.no_messages,
                r.normal_topics,
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.unused_topics > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.no_subscribers > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Topic Name", width=40),
        ColumnDef(header="Subscribers", width=12, style="number"),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="Published", width=12, style="number"),
        ColumnDef(header="Delivered", width=12, style="number"),
        ColumnDef(header="Failed", width=10, style="number"),
        ColumnDef(header="권장 조치", width=35),
    ]
    detail_sheet = wb.new_sheet("Topics", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != TopicStatus.NORMAL:
                t = f.topic
                style = Styles.danger() if f.status == TopicStatus.UNUSED else Styles.warning()

                detail_sheet.add_row(
                    [
                        t.account_name,
                        t.region,
                        t.topic_name,
                        t.subscription_count,
                        f.status.value,
                        int(t.messages_published),
                        int(t.notifications_delivered),
                        int(t.notifications_failed),
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "SNS_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SNSAnalysisResult | None:
    """단일 계정/리전의 SNS 토픽 수집 및 분석 (병렬 실행용)"""
    topics = collect_sns_topics(session, account_id, account_name, region)
    if not topics:
        return None
    return analyze_topics(topics, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """SNS 미사용 토픽 분석"""
    console.print("[bold]SNS 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="sns")
    results: list[SNSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_topics for r in results)
    total_no_sub = sum(r.no_subscribers for r in results)
    total_no_msg = sum(r.no_messages for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] / "
        f"구독자없음: [yellow]{total_no_sub}개[/yellow] / "
        f"메시지없음: {total_no_msg}개"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("sns", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
