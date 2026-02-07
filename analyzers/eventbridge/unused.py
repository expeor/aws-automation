"""
plugins/eventbridge/unused.py - EventBridge 미사용 규칙 분석

비활성화/미사용 규칙 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from shared.io.output import OutputPath, get_context_identifier, open_in_explorer

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 미사용 기준: 7일간 트리거 0
UNUSED_DAYS_THRESHOLD = 7

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "events:ListEventBuses",
        "events:ListRules",
        "events:ListTargetsByRule",
        "cloudwatch:GetMetricStatistics",
    ],
}


class RuleStatus(Enum):
    """규칙 상태"""

    NORMAL = "normal"
    DISABLED = "disabled"
    NO_TARGETS = "no_targets"
    UNUSED = "unused"


@dataclass
class RuleInfo:
    """EventBridge 규칙 정보"""

    account_id: str
    account_name: str
    region: str
    rule_name: str
    rule_arn: str
    event_bus_name: str
    state: str
    schedule_expression: str
    event_pattern: str
    target_count: int
    # CloudWatch 지표
    invocations: float = 0.0
    failed_invocations: float = 0.0
    triggered_rules: float = 0.0


@dataclass
class RuleFinding:
    """규칙 분석 결과"""

    rule: RuleInfo
    status: RuleStatus
    recommendation: str


@dataclass
class EventBridgeAnalysisResult:
    """EventBridge 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_rules: int = 0
    disabled_rules: int = 0
    no_targets: int = 0
    unused_rules: int = 0
    normal_rules: int = 0
    findings: list[RuleFinding] = field(default_factory=list)


def collect_rules(session, account_id: str, account_name: str, region: str) -> list[RuleInfo]:
    """EventBridge 규칙 수집 (배치 메트릭 최적화)

    최적화:
    - 기존: 규칙당 3 API 호출 → 최적화: 전체 1-2 API 호출
    """
    from botocore.exceptions import ClientError

    events = get_client(session, "events", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    rules: list[RuleInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    # 기본 이벤트 버스 + 커스텀 이벤트 버스
    event_buses = ["default"]
    try:
        resp = events.list_event_buses()
        for bus in resp.get("EventBuses", []):
            bus_name = bus.get("Name", "")
            if bus_name and bus_name != "default":
                event_buses.append(bus_name)
    except ClientError as e:
        category = categorize_error(e)
        if category != ErrorCategory.NOT_FOUND:
            logger.debug(f"EventBridge 버스 목록 조회 실패: {get_error_code(e)}")

    # 1단계: 규칙 목록 수집
    for bus_name in event_buses:
        try:
            paginator = events.get_paginator("list_rules")
            paginate_kwargs = {}
            if bus_name != "default":
                paginate_kwargs["EventBusName"] = bus_name

            for page in paginator.paginate(**paginate_kwargs):
                for rule in page.get("Rules", []):
                    rule_name = rule.get("Name", "")
                    rule_arn = rule.get("Arn", "")

                    # 타겟 수 확인
                    target_count = 0
                    try:
                        targets_kwargs = {"Rule": rule_name}
                        if bus_name != "default":
                            targets_kwargs["EventBusName"] = bus_name
                        targets = events.list_targets_by_rule(**targets_kwargs)
                        target_count = len(targets.get("Targets", []))
                    except ClientError as e:
                        category = categorize_error(e)
                        if category != ErrorCategory.NOT_FOUND:
                            logger.debug(f"EventBridge 타겟 조회 실패: {rule_name} ({get_error_code(e)})")

                    info = RuleInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        rule_name=rule_name,
                        rule_arn=rule_arn,
                        event_bus_name=bus_name,
                        state=rule.get("State", ""),
                        schedule_expression=rule.get("ScheduleExpression", ""),
                        event_pattern=rule.get("EventPattern", "")[:100] if rule.get("EventPattern") else "",
                        target_count=target_count,
                    )
                    rules.append(info)

        except ClientError as e:
            category = categorize_error(e)
            if category == ErrorCategory.ACCESS_DENIED:
                logger.info(f"EventBridge 권한 없음: {account_name}/{region}/{bus_name}")
            elif category != ErrorCategory.NOT_FOUND:
                logger.warning(f"EventBridge 규칙 조회 오류: {get_error_code(e)}")
            continue

    # 2단계: 활성화된 규칙만 배치 메트릭 조회
    enabled_rules = [r for r in rules if r.state == "ENABLED"]
    if enabled_rules:
        _collect_eventbridge_metrics_batch(cloudwatch, enabled_rules, start_time, now)

    return rules


def _collect_eventbridge_metrics_batch(
    cloudwatch,
    rules: list[RuleInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """EventBridge 메트릭 배치 수집 (내부 함수)

    메트릭:
    - TriggeredRules (Sum)
    - Invocations (Sum)
    - FailedInvocations (Sum)
    """
    from botocore.exceptions import ClientError

    queries: list[MetricQuery] = []

    for rule in rules:
        safe_id = sanitize_metric_id(rule.rule_name)
        dimensions = {"RuleName": rule.rule_name}

        queries.extend(
            [
                MetricQuery(
                    id=f"{safe_id}_triggered",
                    namespace="AWS/Events",
                    metric_name="TriggeredRules",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_invocations",
                    namespace="AWS/Events",
                    metric_name="Invocations",
                    dimensions=dimensions,
                    stat="Sum",
                ),
                MetricQuery(
                    id=f"{safe_id}_failed",
                    namespace="AWS/Events",
                    metric_name="FailedInvocations",
                    dimensions=dimensions,
                    stat="Sum",
                ),
            ]
        )

    if not queries:
        return

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        for rule in rules:
            safe_id = sanitize_metric_id(rule.rule_name)
            rule.triggered_rules = results.get(f"{safe_id}_triggered", 0.0)
            rule.invocations = results.get(f"{safe_id}_invocations", 0.0)
            rule.failed_invocations = results.get(f"{safe_id}_failed", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_rules(rules: list[RuleInfo], account_id: str, account_name: str, region: str) -> EventBridgeAnalysisResult:
    """EventBridge 규칙 분석"""
    result = EventBridgeAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_rules=len(rules),
    )

    for rule in rules:
        # 비활성화된 규칙
        if rule.state == "DISABLED":
            result.disabled_rules += 1
            result.findings.append(
                RuleFinding(
                    rule=rule,
                    status=RuleStatus.DISABLED,
                    recommendation="비활성화됨 - 삭제 검토",
                )
            )
            continue

        # 타겟 없음
        if rule.target_count == 0:
            result.no_targets += 1
            result.findings.append(
                RuleFinding(
                    rule=rule,
                    status=RuleStatus.NO_TARGETS,
                    recommendation="타겟 없음 - 삭제 검토",
                )
            )
            continue

        # 미사용 (트리거 없음)
        if rule.triggered_rules == 0 and rule.invocations == 0:
            result.unused_rules += 1
            result.findings.append(
                RuleFinding(
                    rule=rule,
                    status=RuleStatus.UNUSED,
                    recommendation="트리거 없음 - 사용 여부 확인",
                )
            )
            continue

        result.normal_rules += 1
        result.findings.append(
            RuleFinding(
                rule=rule,
                status=RuleStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[EventBridgeAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="비활성화", width=10, style="number"),
        ColumnDef(header="타겟없음", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_rules,
                r.disabled_rules,
                r.no_targets,
                r.unused_rules,
                r.normal_rules,
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.disabled_rules > 0:
            ws.cell(row=row_num, column=4).fill = gray_fill
        if r.no_targets > 0:
            ws.cell(row=row_num, column=5).fill = red_fill
        if r.unused_rules > 0:
            ws.cell(row=row_num, column=6).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Rule Name", width=30),
        ColumnDef(header="Event Bus", width=15),
        ColumnDef(header="State", width=12),
        ColumnDef(header="Schedule", width=20),
        ColumnDef(header="Targets", width=10, style="number"),
        ColumnDef(header="Triggers", width=10, style="number"),
        ColumnDef(header="상태", width=12),
        ColumnDef(header="권장 조치", width=25),
    ]
    detail_sheet = wb.new_sheet("Rules", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != RuleStatus.NORMAL:
                rule = f.rule
                style = None
                if f.status == RuleStatus.NO_TARGETS:
                    style = Styles.danger()
                elif f.status in (RuleStatus.UNUSED, RuleStatus.DISABLED):
                    style = Styles.warning()

                detail_sheet.add_row(
                    [
                        rule.account_name,
                        rule.region,
                        rule.rule_name,
                        rule.event_bus_name,
                        rule.state,
                        rule.schedule_expression or "-",
                        rule.target_count,
                        int(rule.triggered_rules),
                        f.status.value,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "EventBridge_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EventBridgeAnalysisResult | None:
    """단일 계정/리전의 EventBridge 규칙 수집 및 분석 (병렬 실행용)"""
    rules = collect_rules(session, account_id, account_name, region)
    if not rules:
        return None
    return analyze_rules(rules, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """EventBridge 미사용 규칙 분석"""
    console.print("[bold]EventBridge 규칙 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="events")
    results: list[EventBridgeAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_disabled = sum(r.disabled_rules for r in results)
    total_no_targets = sum(r.no_targets for r in results)
    total_unused = sum(r.unused_rules for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"비활성화: {total_disabled}개 / "
        f"타겟없음: [red]{total_no_targets}개[/red] / "
        f"미사용: [yellow]{total_unused}개[/yellow]"
    )

    identifier = get_context_identifier(ctx)

    output_path = OutputPath(identifier).sub("eventbridge", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
