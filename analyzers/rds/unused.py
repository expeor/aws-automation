"""
plugins/rds/unused.py - RDS 유휴 인스턴스 분석

유휴/저사용 RDS 인스턴스 탐지 (CloudWatch 지표 기반)

탐지 기준 (AWS Trusted Advisor 기반):
- 미사용: DatabaseConnections < 1 (7일 평균) AND IOPS 낮음 (< 20/일)
- 저사용: CPU < 5% AND 낮은 IOPS
- 정지됨: stopped 상태

참고:
- AWS Trusted Advisor: No connections (7 days) → Idle DB
- https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/rds-metrics.html

CloudWatch 메트릭:
- Namespace: AWS/RDS
- 메트릭: DatabaseConnections, CPUUtilization, ReadIOPS, WriteIOPS
- Dimension: DBInstanceIdentifier

최적화:
- CloudWatch GetMetricData API 사용 (배치 조회)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from shared.io.compat import generate_dual_report
from shared.io.output import open_in_explorer, print_report_complete
from shared.io.output.helpers import create_output_path

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 미사용 기준: 7일간 (AWS Trusted Advisor 기준)
ANALYSIS_DAYS = 7
# 미사용 기준: 연결 수 평균 < 1
UNUSED_CONNECTION_THRESHOLD = 1
# 미사용 기준: IOPS 평균 < 20/일
UNUSED_IOPS_THRESHOLD = 20
# 저사용 기준: CPU 평균 5% 미만
LOW_USAGE_CPU_THRESHOLD = 5.0
# 저사용 기준: IOPS 평균 < 100/일
LOW_USAGE_IOPS_THRESHOLD = 100

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "rds:DescribeDBInstances",
        "cloudwatch:GetMetricStatistics",
    ],
}


class InstanceStatus(Enum):
    """인스턴스 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"
    STOPPED = "stopped"


@dataclass
class RDSInstanceInfo:
    """RDS 인스턴스 정보"""

    account_id: str
    account_name: str
    region: str
    db_instance_id: str
    db_instance_class: str
    engine: str
    engine_version: str
    status: str
    multi_az: bool
    storage_type: str
    allocated_storage: int  # GB
    created_at: datetime | None
    # CloudWatch 지표
    avg_connections: float = 0.0
    avg_cpu: float = 0.0
    avg_read_iops: float = 0.0
    avg_write_iops: float = 0.0

    def get_estimated_monthly_cost(self, session=None) -> float:
        """월간 비용 추정 (Pricing API 사용)"""
        from analyzers.cost.pricing.rds import get_rds_monthly_cost

        return get_rds_monthly_cost(
            region=self.region,
            instance_class=self.db_instance_class,
            engine=self.engine,
            storage_gb=self.allocated_storage,
            storage_type=self.storage_type if self.storage_type else "gp3",
            multi_az=self.multi_az,
            session=session,
        )

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (후방 호환용)"""
        return self.get_estimated_monthly_cost()


@dataclass
class InstanceFinding:
    """인스턴스 분석 결과"""

    instance: RDSInstanceInfo
    status: InstanceStatus
    recommendation: str


@dataclass
class RDSAnalysisResult:
    """RDS 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_instances: int = 0
    unused_instances: int = 0
    low_usage_instances: int = 0
    stopped_instances: int = 0
    normal_instances: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[InstanceFinding] = field(default_factory=list)


def collect_rds_instances(session, account_id: str, account_name: str, region: str) -> list[RDSInstanceInfo]:
    """RDS 인스턴스 수집 (배치 메트릭 최적화)

    최적화:
    - 기존: 인스턴스당 4 API 호출 → 최적화: 전체 1-2 API 호출
    - 예: 20개 인스턴스 × 4 메트릭 = 80 API → 1 API
    """
    from botocore.exceptions import ClientError

    rds = get_client(session, "rds", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    instances = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 인스턴스 목록 수집
        paginator = rds.get_paginator("describe_db_instances")
        for page in paginator.paginate():
            for db in page.get("DBInstances", []):
                db_id = db.get("DBInstanceIdentifier", "")

                instance = RDSInstanceInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    db_instance_id=db_id,
                    db_instance_class=db.get("DBInstanceClass", ""),
                    engine=db.get("Engine", ""),
                    engine_version=db.get("EngineVersion", ""),
                    status=db.get("DBInstanceStatus", ""),
                    multi_az=db.get("MultiAZ", False),
                    storage_type=db.get("StorageType", ""),
                    allocated_storage=db.get("AllocatedStorage", 0),
                    created_at=db.get("InstanceCreateTime"),
                )
                instances.append(instance)

        # 2단계: 메트릭이 필요한 인스턴스 필터링 (stopped 제외)
        active_instances = [i for i in instances if i.status != "stopped"]

        if active_instances:
            # 3단계: 배치 메트릭 조회
            _collect_rds_metrics_batch(cloudwatch, active_instances, start_time, now)

    except ClientError:
        pass

    return instances


def _collect_rds_metrics_batch(
    cloudwatch,
    instances: list[RDSInstanceInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """RDS 인스턴스 메트릭 배치 수집 (내부 함수)"""
    from botocore.exceptions import ClientError

    metrics_to_fetch = [
        ("DatabaseConnections", "avg_connections"),
        ("CPUUtilization", "avg_cpu"),
        ("ReadIOPS", "avg_read_iops"),
        ("WriteIOPS", "avg_write_iops"),
    ]

    # 모든 인스턴스에 대한 쿼리 생성
    queries = []
    for instance in instances:
        safe_id = sanitize_metric_id(instance.db_instance_id)
        for metric_name, _ in metrics_to_fetch:
            metric_key = metric_name.lower()
            queries.append(
                MetricQuery(
                    id=f"{safe_id}_{metric_key}",
                    namespace="AWS/RDS",
                    metric_name=metric_name,
                    dimensions={"DBInstanceIdentifier": instance.db_instance_id},
                    stat="Average",
                )
            )

    try:
        # 배치 조회
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        # 결과 매핑
        # Note: GetMetricData는 평균값을 직접 반환하지 않고 datapoint 합계를 반환
        # 따라서 기간 내 데이터포인트 수로 나눠 평균 계산
        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for instance in instances:
            safe_id = sanitize_metric_id(instance.db_instance_id)
            for metric_name, attr_name in metrics_to_fetch:
                metric_key = metric_name.lower()
                # GetMetricData with Average stat returns sum of averages
                # We need to divide by number of periods to get true average
                value = results.get(f"{safe_id}_{metric_key}", 0.0) / days
                setattr(instance, attr_name, value)

    except ClientError:
        # 실패 시 무시 (기본값 0 유지)
        pass


def analyze_instances(
    instances: list[RDSInstanceInfo], account_id: str, account_name: str, region: str
) -> RDSAnalysisResult:
    """RDS 인스턴스 분석"""
    result = RDSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_instances=len(instances),
    )

    for instance in instances:
        # 정지된 인스턴스
        if instance.status == "stopped":
            result.stopped_instances += 1
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.STOPPED,
                    recommendation="정지됨 - 장기 미사용 시 스냅샷 후 삭제 검토",
                )
            )
            continue

        total_iops = instance.avg_read_iops + instance.avg_write_iops

        # 미사용: 연결 수 < 1 AND IOPS 낮음 (AWS Trusted Advisor 기준)
        if instance.avg_connections < UNUSED_CONNECTION_THRESHOLD and total_iops < UNUSED_IOPS_THRESHOLD:
            result.unused_instances += 1
            result.unused_monthly_cost += instance.estimated_monthly_cost
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.UNUSED,
                    recommendation=f"미사용 (연결 {instance.avg_connections:.1f}, IOPS {total_iops:.1f}) - 삭제 검토 (${instance.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: CPU < 5% AND IOPS 낮음
        if instance.avg_cpu < LOW_USAGE_CPU_THRESHOLD and total_iops < LOW_USAGE_IOPS_THRESHOLD:
            result.low_usage_instances += 1
            result.low_usage_monthly_cost += instance.estimated_monthly_cost
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.LOW_USAGE,
                    recommendation=f"저사용 (CPU {instance.avg_cpu:.1f}%, IOPS {total_iops:.1f}) - 다운사이징 검토",
                )
            )
            continue

        result.normal_instances += 1
        result.findings.append(
            InstanceFinding(
                instance=instance,
                status=InstanceStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def _build_excel(results: list[RDSAnalysisResult]):
    """Excel Workbook 빌더 (저장하지 않고 반환)"""
    from openpyxl.styles import PatternFill

    from shared.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정지", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_instances,
                r.unused_instances,
                r.low_usage_instances,
                r.stopped_instances,
                r.normal_instances,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.unused_instances > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_instances > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill
        if r.stopped_instances > 0:
            ws.cell(row=row_num, column=6).fill = gray_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Instance ID", width=30),
        ColumnDef(header="Engine", width=15),
        ColumnDef(header="Class", width=18),
        ColumnDef(header="Storage", width=12),
        ColumnDef(header="Multi-AZ", width=10),
        ColumnDef(header="상태", width=12),
        ColumnDef(header="Avg Conn", width=10),
        ColumnDef(header="Avg CPU", width=10),
        ColumnDef(header="Avg IOPS", width=10),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Instances", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != InstanceStatus.NORMAL:
                inst = f.instance
                style = None
                if f.status == InstanceStatus.UNUSED:
                    style = Styles.danger()
                elif f.status == InstanceStatus.LOW_USAGE:
                    style = Styles.warning()

                total_iops = inst.avg_read_iops + inst.avg_write_iops
                detail_sheet.add_row(
                    [
                        inst.account_name,
                        inst.region,
                        inst.db_instance_id,
                        inst.engine,
                        inst.db_instance_class,
                        f"{inst.allocated_storage} GB",
                        "Yes" if inst.multi_az else "No",
                        f.status.value,
                        f"{inst.avg_connections:.1f}",
                        f"{inst.avg_cpu:.1f}%",
                        f"{total_iops:.1f}",
                        f"${inst.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ],
                    style=style,
                )

    return wb


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> RDSAnalysisResult | None:
    """단일 계정/리전의 RDS 인스턴스 수집 및 분석 (병렬 실행용)"""
    instances = collect_rds_instances(session, account_id, account_name, region)
    if not instances:
        return None
    return analyze_instances(instances, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """RDS 유휴 인스턴스 분석"""
    console.print("[bold]RDS 유휴 인스턴스 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="rds")
    results: list[RDSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_instances = sum(r.total_instances for r in results)
    total_unused = sum(r.unused_instances for r in results)
    total_low = sum(r.low_usage_instances for r in results)
    total_stopped = sum(r.stopped_instances for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월) / "
        f"정지: {total_stopped}개"
    )

    # HTML용 flat 데이터
    flat_data = []
    for r in results:
        for f in r.findings:
            if f.status != InstanceStatus.NORMAL:
                inst = f.instance
                flat_data.append(
                    {
                        "account_id": inst.account_id,
                        "account_name": inst.account_name,
                        "region": inst.region,
                        "resource_id": inst.db_instance_id,
                        "resource_name": inst.db_instance_id,
                        "status": f.status.value,
                        "reason": f.recommendation,
                        "cost": inst.estimated_monthly_cost,
                    }
                )

    output_path = create_output_path(ctx, "rds", "unused")
    report_paths = generate_dual_report(
        ctx,
        data=flat_data,
        output_dir=output_path,
        prefix="RDS_Unused",
        excel_builder=lambda: _build_excel(results),
        html_config={
            "title": "RDS 유휴 인스턴스 분석",
            "service": "RDS",
            "tool_name": "unused",
            "total": total_instances,
            "found": total_unused + total_low + total_stopped,
            "savings": unused_cost + low_cost,
        },
    )

    print_report_complete(report_paths)
    open_in_explorer(output_path)
