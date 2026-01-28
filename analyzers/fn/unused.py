"""
plugins/fn/unused.py - 미사용 Lambda 함수 분석

30일 이상 호출되지 않은 Lambda 함수 탐지

분석 기준:
- 30일간 Invocations 메트릭이 0인 함수
- Provisioned Concurrency가 설정된 미사용 함수 (비용 낭비)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from shared.aws.pricing import (
    get_lambda_monthly_cost,
    get_lambda_provisioned_monthly_cost,
)

from shared.aws.lambda_ import LambdaFunctionInfo, collect_functions_with_metrics

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:GetFunction",
        "lambda:ListProvisionedConcurrencyConfigs",
        "cloudwatch:GetMetricStatistics",
    ],
}


class UsageStatus(Enum):
    """사용 상태"""

    UNUSED = "unused"  # 30일간 호출 없음
    UNUSED_PROVISIONED = "unused_provisioned"  # 미사용 + PC 설정됨
    LOW_USAGE = "low_usage"  # 저사용 (월 100회 미만)
    NORMAL = "normal"  # 정상 사용


@dataclass
class LambdaFinding:
    """Lambda 분석 결과"""

    function: LambdaFunctionInfo
    status: UsageStatus
    recommendation: str
    monthly_waste: float = 0.0


@dataclass
class LambdaAnalysisResult:
    """Lambda 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    unused_count: int = 0
    low_usage_count: int = 0
    normal_count: int = 0
    unused_monthly_cost: float = 0.0
    findings: list[LambdaFinding] = field(default_factory=list)


# =============================================================================
# 분석
# =============================================================================


def analyze_functions(
    functions: list[LambdaFunctionInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> LambdaAnalysisResult:
    """Lambda 함수 미사용 분석"""
    result = LambdaAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(functions),
    )

    for func in functions:
        finding = _analyze_single_function(func, region)
        result.findings.append(finding)

        if finding.status == UsageStatus.UNUSED or finding.status == UsageStatus.UNUSED_PROVISIONED:
            result.unused_count += 1
            result.unused_monthly_cost += finding.monthly_waste
        elif finding.status == UsageStatus.LOW_USAGE:
            result.low_usage_count += 1
        else:
            result.normal_count += 1

    return result


def _analyze_single_function(func: LambdaFunctionInfo, region: str) -> LambdaFinding:
    """개별 Lambda 함수 분석"""

    # 메트릭이 없으면 알 수 없음
    if func.metrics is None:
        return LambdaFinding(
            function=func,
            status=UsageStatus.NORMAL,
            recommendation="메트릭 조회 실패",
        )

    invocations = func.metrics.invocations

    # 미사용 (30일간 호출 없음)
    if invocations == 0:
        # Provisioned Concurrency가 있으면 더 심각
        if func.provisioned_concurrency > 0:
            waste = get_lambda_provisioned_monthly_cost(
                region=region,
                memory_mb=func.memory_mb,
                provisioned_concurrency=func.provisioned_concurrency,
            )
            return LambdaFinding(
                function=func,
                status=UsageStatus.UNUSED_PROVISIONED,
                recommendation=f"미사용 + PC {func.provisioned_concurrency}개 설정됨 - 즉시 삭제 또는 PC 해제 권장",
                monthly_waste=waste,
            )

        return LambdaFinding(
            function=func,
            status=UsageStatus.UNUSED,
            recommendation="30일간 호출 없음 - 삭제 또는 비활성화 검토",
            monthly_waste=0.0,  # 미호출 시 비용 없음 (PC 제외)
        )

    # 저사용 (월 100회 미만)
    if invocations < 100:
        return LambdaFinding(
            function=func,
            status=UsageStatus.LOW_USAGE,
            recommendation=f"저사용 (30일간 {invocations}회) - 통합 또는 삭제 검토",
        )

    # 정상 사용
    estimated_cost = get_lambda_monthly_cost(
        region=region,
        invocations=invocations,
        avg_duration_ms=func.metrics.duration_avg_ms,
        memory_mb=func.memory_mb,
    )

    # PC가 있으면 비용 추가
    if func.provisioned_concurrency > 0:
        pc_cost = get_lambda_provisioned_monthly_cost(
            region=region,
            memory_mb=func.memory_mb,
            provisioned_concurrency=func.provisioned_concurrency,
        )
        estimated_cost += pc_cost

    func.estimated_monthly_cost = estimated_cost

    return LambdaFinding(
        function=func,
        status=UsageStatus.NORMAL,
        recommendation=f"정상 사용 (30일간 {invocations:,}회)",
    )


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[LambdaAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")

    wb = Workbook()

    # Summary Sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("Lambda 미사용 분석 보고서")
    summary.add_item("생성", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="월간 낭비", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_count,
                r.unused_count,
                r.low_usage_count,
                r.normal_count,
                f"${r.unused_monthly_cost:,.2f}",
            ]
        )
        if r.unused_count > 0:
            summary_sheet._ws.cell(row=row_num, column=4).fill = red_fill

    # 총계
    total_functions = sum(r.total_count for r in results)
    total_unused = sum(r.unused_count for r in results)
    total_waste = sum(r.unused_monthly_cost for r in results)
    summary_sheet.add_summary_row(
        [
            "합계",
            "-",
            total_functions,
            total_unused,
            "-",
            "-",
            f"${total_waste:,.2f}",
        ]
    )

    # Unused Functions Sheet
    unused_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function Name", width=40),
        ColumnDef(header="Runtime", width=15),
        ColumnDef(header="Memory (MB)", width=12, style="number"),
        ColumnDef(header="Timeout (s)", width=12, style="number"),
        ColumnDef(header="Code Size (MB)", width=14, style="number"),
        ColumnDef(header="Last Modified", width=15),
        ColumnDef(header="Provisioned Concurrency", width=20, style="number"),
        ColumnDef(header="상태", width=20),
        ColumnDef(header="월간 낭비", width=15),
        ColumnDef(header="권장 조치", width=40),
    ]
    unused_sheet = wb.new_sheet("Unused", unused_columns)

    for r in results:
        for f in r.findings:
            if f.status in (
                UsageStatus.UNUSED,
                UsageStatus.UNUSED_PROVISIONED,
                UsageStatus.LOW_USAGE,
            ):
                fn = f.function
                row_num = unused_sheet.add_row(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        fn.runtime,
                        fn.memory_mb,
                        fn.timeout_seconds,
                        round(fn.code_size_mb, 2),
                        fn.last_modified.strftime("%Y-%m-%d") if fn.last_modified else "-",
                        fn.provisioned_concurrency or "-",
                        f.status.value,
                        f"${f.monthly_waste:,.2f}" if f.monthly_waste > 0 else "-",
                        f.recommendation,
                    ]
                )

                # 상태별 색상
                ws = unused_sheet._ws
                if f.status == UsageStatus.UNUSED_PROVISIONED or f.status == UsageStatus.UNUSED:
                    ws.cell(row=row_num, column=10).fill = red_fill
                elif f.status == UsageStatus.LOW_USAGE:
                    ws.cell(row=row_num, column=10).fill = yellow_fill

    # All Functions Sheet
    all_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Function Name", width=40),
        ColumnDef(header="Runtime", width=15),
        ColumnDef(header="Memory (MB)", width=12, style="number"),
        ColumnDef(header="Timeout (s)", width=12, style="number"),
        ColumnDef(header="Code Size (MB)", width=14, style="number"),
        ColumnDef(header="Invocations (30d)", width=16, style="number"),
        ColumnDef(header="Avg Duration (ms)", width=16, style="number"),
        ColumnDef(header="Errors", width=10, style="number"),
        ColumnDef(header="Throttles", width=10, style="number"),
        ColumnDef(header="PC", width=10, style="number"),
        ColumnDef(header="Reserved", width=10, style="number"),
        ColumnDef(header="상태", width=20),
        ColumnDef(header="추정 월 비용", width=15),
    ]
    all_sheet = wb.new_sheet("All Functions", all_columns)

    for r in results:
        for f in r.findings:
            fn = f.function
            metrics = fn.metrics

            all_sheet.add_row(
                [
                    fn.account_name,
                    fn.region,
                    fn.function_name,
                    fn.runtime,
                    fn.memory_mb,
                    fn.timeout_seconds,
                    round(fn.code_size_mb, 2),
                    metrics.invocations if metrics else 0,
                    round(metrics.duration_avg_ms, 2) if metrics else 0,
                    metrics.errors if metrics else 0,
                    metrics.throttles if metrics else 0,
                    fn.provisioned_concurrency or "-",
                    fn.reserved_concurrency if fn.reserved_concurrency is not None else "-",
                    f.status.value,
                    f"${fn.estimated_monthly_cost:,.4f}" if fn.estimated_monthly_cost > 0 else "-",
                ]
            )

    return str(wb.save_as(output_dir, "Lambda_Unused"))


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> LambdaAnalysisResult:
    """단일 계정/리전의 Lambda 수집 및 분석 (병렬 실행용)"""
    functions = collect_functions_with_metrics(session, account_id, account_name, region)
    return analyze_functions(functions, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Lambda 미사용 분석 실행"""
    console.print("[bold]Lambda 미사용 분석 시작...[/bold]\n")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")

    results: list[LambdaAnalysisResult] = result.get_data()

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_functions = sum(r.total_count for r in results)
    total_unused = sum(r.unused_count for r in results)
    total_low = sum(r.low_usage_count for r in results)
    total_waste = sum(r.unused_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체 Lambda 함수: {total_functions}개")
    if total_unused > 0:
        console.print(f"[red]미사용: {total_unused}개[/red]")
    if total_low > 0:
        console.print(f"[yellow]저사용: {total_low}개[/yellow]")
    if total_waste > 0:
        console.print(f"[red]월간 낭비 비용: ${total_waste:,.2f}[/red]")

    # 보고서
    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("lambda", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
