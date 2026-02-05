"""
plugins/transfer/unused.py - AWS Transfer Family 미사용 서버 분석

유휴/미사용 Transfer Family 서버 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.parallel.decorators import categorize_error, get_error_code
from core.parallel.types import ErrorCategory
from core.tools.output import OutputPath, open_in_explorer
from shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id
from shared.aws.pricing.transfer import get_transfer_monthly_cost

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()
logger = logging.getLogger(__name__)

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "transfer:ListServers",
        "transfer:DescribeServer",
        "transfer:ListUsers",
        "cloudwatch:GetMetricStatistics",
    ],
}

# 분석 기간 (일) - CloudFix 권장 30일
ANALYSIS_DAYS = 30


class ServerStatus(Enum):
    """서버 상태"""

    NORMAL = "normal"
    UNUSED = "unused"  # 파일 전송 없음
    IDLE = "idle"  # 저사용
    NO_USERS = "no_users"  # 사용자 없음
    STOPPED = "stopped"  # 중지됨


@dataclass
class TransferServerInfo:
    """Transfer Family 서버 정보"""

    account_id: str
    account_name: str
    region: str
    server_id: str
    endpoint_type: str  # PUBLIC, VPC, VPC_ENDPOINT
    protocols: list[str]  # SFTP, FTPS, FTP, AS2
    state: str  # ONLINE, OFFLINE, STARTING, STOPPING, START_FAILED, STOP_FAILED
    identity_provider_type: str
    user_count: int = 0
    # CloudWatch 지표
    files_in: float = 0.0
    files_out: float = 0.0
    bytes_in: float = 0.0
    bytes_out: float = 0.0
    # 비용
    estimated_monthly_cost: float = 0.0

    @property
    def total_files(self) -> float:
        return self.files_in + self.files_out

    @property
    def total_bytes(self) -> float:
        return self.bytes_in + self.bytes_out

    @property
    def is_active(self) -> bool:
        return self.total_files > 0 or self.total_bytes > 0


@dataclass
class TransferFinding:
    """서버 분석 결과"""

    server: TransferServerInfo
    status: ServerStatus
    recommendation: str


@dataclass
class TransferAnalysisResult:
    """Transfer Family 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_servers: int = 0
    unused_servers: int = 0
    idle_servers: int = 0
    no_users_servers: int = 0
    stopped_servers: int = 0
    normal_servers: int = 0
    total_monthly_waste: float = 0.0
    findings: list[TransferFinding] = field(default_factory=list)


def collect_servers(session, account_id: str, account_name: str, region: str) -> list[TransferServerInfo]:
    """Transfer Family 서버 수집 (배치 메트릭 최적화)

    최적화:
    - 기존: 서버당 4 API 호출 → 최적화: 전체 1-2 API 호출
    """
    from botocore.exceptions import ClientError

    transfer = get_client(session, "transfer", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    servers: list[TransferServerInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 서버 목록 수집
        paginator = transfer.get_paginator("list_servers")
        for page in paginator.paginate():
            for server_summary in page.get("Servers", []):
                server_id = server_summary.get("ServerId", "")

                try:
                    # 서버 상세 정보
                    server_detail = transfer.describe_server(ServerId=server_id).get("Server", {})

                    protocols = server_detail.get("Protocols", ["SFTP"])
                    endpoint_type = server_detail.get("EndpointType", "PUBLIC")

                    # 사용자 수 확인
                    user_count = 0
                    try:
                        users_resp = transfer.list_users(ServerId=server_id)
                        user_count = len(users_resp.get("Users", []))
                    except ClientError as e:
                        category = categorize_error(e)
                        if category != ErrorCategory.NOT_FOUND:
                            logger.debug(f"Transfer 사용자 조회 실패: {server_id} ({get_error_code(e)})")

                    # 월 비용 계산 (pricing 모듈 사용)
                    monthly_cost = get_transfer_monthly_cost(region, protocols=protocols, session=session)

                    info = TransferServerInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        server_id=server_id,
                        endpoint_type=endpoint_type,
                        protocols=protocols,
                        state=server_detail.get("State", ""),
                        identity_provider_type=server_detail.get("IdentityProviderType", ""),
                        user_count=user_count,
                        estimated_monthly_cost=monthly_cost,
                    )
                    servers.append(info)

                except ClientError as e:
                    category = categorize_error(e)
                    if category != ErrorCategory.NOT_FOUND:
                        logger.debug(f"Transfer 서버 상세 조회 실패: {server_id} ({get_error_code(e)})")
                    continue

        # 2단계: ONLINE 상태인 서버만 배치 메트릭 조회
        online_servers = [s for s in servers if s.state == "ONLINE"]
        if online_servers:
            _collect_transfer_metrics_batch(cloudwatch, online_servers, start_time, now)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"Transfer 권한 없음: {account_name}/{region}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"Transfer 조회 오류: {get_error_code(e)}")

    return servers


def _collect_transfer_metrics_batch(
    cloudwatch,
    servers: list[TransferServerInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """Transfer Family 메트릭 배치 수집 (내부 함수)

    메트릭:
    - FilesIn (Sum)
    - FilesOut (Sum)
    - BytesIn (Sum)
    - BytesOut (Sum)
    """
    from botocore.exceptions import ClientError

    queries: list[MetricQuery] = []

    for server in servers:
        safe_id = sanitize_metric_id(server.server_id)
        dimensions = {"ServerId": server.server_id}

        queries.extend([
            MetricQuery(
                id=f"{safe_id}_files_in",
                namespace="AWS/Transfer",
                metric_name="FilesIn",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_files_out",
                namespace="AWS/Transfer",
                metric_name="FilesOut",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_bytes_in",
                namespace="AWS/Transfer",
                metric_name="BytesIn",
                dimensions=dimensions,
                stat="Sum",
            ),
            MetricQuery(
                id=f"{safe_id}_bytes_out",
                namespace="AWS/Transfer",
                metric_name="BytesOut",
                dimensions=dimensions,
                stat="Sum",
            ),
        ])

    if not queries:
        return

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        for server in servers:
            safe_id = sanitize_metric_id(server.server_id)
            server.files_in = results.get(f"{safe_id}_files_in", 0.0)
            server.files_out = results.get(f"{safe_id}_files_out", 0.0)
            server.bytes_in = results.get(f"{safe_id}_bytes_in", 0.0)
            server.bytes_out = results.get(f"{safe_id}_bytes_out", 0.0)

    except ClientError as e:
        category = categorize_error(e)
        if category == ErrorCategory.ACCESS_DENIED:
            logger.info(f"CloudWatch 권한 없음: {get_error_code(e)}")
        elif category == ErrorCategory.THROTTLING:
            logger.warning(f"CloudWatch API 쓰로틀링: {get_error_code(e)}")
        elif category != ErrorCategory.NOT_FOUND:
            logger.warning(f"CloudWatch 메트릭 조회 오류: {get_error_code(e)}")


def analyze_servers(
    servers: list[TransferServerInfo], account_id: str, account_name: str, region: str
) -> TransferAnalysisResult:
    """Transfer Family 서버 분석"""
    result = TransferAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_servers=len(servers),
    )

    for server in servers:
        # 중지됨
        if server.state in ("OFFLINE", "STOPPING", "STOP_FAILED"):
            result.stopped_servers += 1
            result.findings.append(
                TransferFinding(
                    server=server,
                    status=ServerStatus.STOPPED,
                    recommendation="서버 중지됨 - 불필요시 삭제 검토",
                )
            )
            continue

        # 사용자 없음
        if server.user_count == 0:
            result.no_users_servers += 1
            result.total_monthly_waste += server.estimated_monthly_cost
            result.findings.append(
                TransferFinding(
                    server=server,
                    status=ServerStatus.NO_USERS,
                    recommendation="사용자 없음 - 삭제 검토",
                )
            )
            continue

        # 미사용 (파일 전송 없음)
        if not server.is_active:
            result.unused_servers += 1
            result.total_monthly_waste += server.estimated_monthly_cost
            result.findings.append(
                TransferFinding(
                    server=server,
                    status=ServerStatus.UNUSED,
                    recommendation=f"파일 전송 없음 ({ANALYSIS_DAYS}일간) - 삭제 검토",
                )
            )
            continue

        # 저사용 (하루 평균 1개 미만)
        avg_daily_files = server.total_files / ANALYSIS_DAYS
        if avg_daily_files < 1:
            result.idle_servers += 1
            result.total_monthly_waste += server.estimated_monthly_cost * 0.5  # 50% 낭비 추정
            result.findings.append(
                TransferFinding(
                    server=server,
                    status=ServerStatus.IDLE,
                    recommendation=f"저사용 (일 평균 {avg_daily_files:.1f}건) - 통합 검토",
                )
            )
            continue

        result.normal_servers += 1
        result.findings.append(
            TransferFinding(
                server=server,
                status=ServerStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[TransferAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="사용자없음", width=12, style="number"),
        ColumnDef(header="중지됨", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="월 낭비(USD)", width=12, style="currency"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_servers,
                r.unused_servers,
                r.idle_servers,
                r.no_users_servers,
                r.stopped_servers,
                r.normal_servers,
                r.total_monthly_waste,
            ]
        )
        ws = summary_sheet._ws
        if r.unused_servers > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.idle_servers > 0 or r.no_users_servers > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Server ID", width=25),
        ColumnDef(header="Protocols", width=15),
        ColumnDef(header="Endpoint", width=15),
        ColumnDef(header="State", width=12),
        ColumnDef(header="Users", width=10, style="number"),
        ColumnDef(header="Files In", width=12, style="number"),
        ColumnDef(header="Files Out", width=12, style="number"),
        ColumnDef(header="상태", width=12),
        ColumnDef(header="월 비용(USD)", width=12, style="currency"),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Servers", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != ServerStatus.NORMAL:
                s = f.server
                style = Styles.danger() if f.status == ServerStatus.UNUSED else Styles.warning()
                detail_sheet.add_row(
                    [
                        s.account_name,
                        s.region,
                        s.server_id,
                        ", ".join(s.protocols),
                        s.endpoint_type,
                        s.state,
                        s.user_count,
                        int(s.files_in),
                        int(s.files_out),
                        f.status.value,
                        s.estimated_monthly_cost,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "Transfer_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> TransferAnalysisResult | None:
    """단일 계정/리전의 Transfer Family 서버 수집 및 분석 (병렬 실행용)"""
    servers = collect_servers(session, account_id, account_name, region)
    if not servers:
        return None
    return analyze_servers(servers, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """Transfer Family 미사용 서버 분석"""
    console.print("[bold]Transfer Family 서버 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="transfer")
    results: list[TransferAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_servers for r in results)
    total_idle = sum(r.idle_servers for r in results)
    total_no_users = sum(r.no_users_servers for r in results)
    total_waste = sum(r.total_monthly_waste for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] / "
        f"저사용: [yellow]{total_idle}개[/yellow] / "
        f"사용자없음: [orange1]{total_no_users}개[/orange1]"
    )
    console.print(f"예상 월 낭비: [red]${total_waste:,.2f}[/red]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("transfer", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
