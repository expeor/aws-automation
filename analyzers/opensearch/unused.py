"""
plugins/opensearch/unused.py - OpenSearch 미사용 도메인 분석

유휴/저사용 OpenSearch 도메인 탐지 (CloudWatch 지표 기반)

분석 기준:
- Unused: CPUUtilization < 2% (7일간) AND SearchableDocuments = 0
- Low Usage: CPUUtilization < 5% AND IndexingRate < 1/min
- Over-provisioned: JVMMemoryPressure < 30% consistently

최적화:
- CloudWatch GetMetricData API 사용 (배치 조회)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from shared.aws.metrics import MetricQuery, batch_get_metrics, sanitize_metric_id

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 분석 기간 (일)
ANALYSIS_DAYS = 7
# 미사용 기준: CPU 평균 < 2%
UNUSED_CPU_THRESHOLD = 2.0
# 저사용 기준: CPU 평균 < 5%
LOW_USAGE_CPU_THRESHOLD = 5.0
# 저사용 기준: 인덱싱 레이트 < 1/min
LOW_USAGE_INDEXING_THRESHOLD = 1.0

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "es:DescribeDomains",
        "es:ListDomainNames",
        "cloudwatch:GetMetricData",
    ],
}


class DomainStatus(Enum):
    """도메인 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"


@dataclass
class OpenSearchDomainInfo:
    """OpenSearch 도메인 정보"""

    account_id: str
    account_name: str
    region: str
    domain_name: str
    domain_id: str
    instance_type: str
    instance_count: int
    storage_type: str
    storage_gb: float
    engine_version: str
    created_at: datetime | None
    # CloudWatch 지표
    avg_cpu: float = 0.0
    searchable_docs: float = 0.0
    indexing_rate: float = 0.0
    jvm_memory_pressure: float = 0.0

    def get_estimated_monthly_cost(self, session=None) -> float:
        """월간 비용 추정 (Pricing API 사용)"""
        from analyzers.cost.pricing.opensearch import get_opensearch_monthly_cost

        return get_opensearch_monthly_cost(
            region=self.region,
            instance_type=self.instance_type,
            instance_count=self.instance_count,
            storage_gb=int(self.storage_gb),
            session=session,
        )

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (후방 호환용)"""
        return self.get_estimated_monthly_cost()


@dataclass
class OpenSearchDomainFinding:
    """도메인 분석 결과"""

    domain: OpenSearchDomainInfo
    status: DomainStatus
    recommendation: str


@dataclass
class OpenSearchAnalysisResult:
    """OpenSearch 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_domains: int = 0
    unused_domains: int = 0
    low_usage_domains: int = 0
    normal_domains: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[OpenSearchDomainFinding] = field(default_factory=list)


def collect_opensearch_domains(session, account_id: str, account_name: str, region: str) -> list[OpenSearchDomainInfo]:
    """OpenSearch 도메인 수집 (배치 메트릭 최적화)"""
    from botocore.exceptions import ClientError

    opensearch = get_client(session, "opensearch", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)

    domains: list[OpenSearchDomainInfo] = []
    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 도메인 이름 목록 조회
        response = opensearch.list_domain_names()
        domain_names = [d.get("DomainName", "") for d in response.get("DomainNames", [])]

        if not domain_names:
            return []

        # 2단계: 도메인 상세 정보 조회 (배치로 최대 5개씩)
        for i in range(0, len(domain_names), 5):
            batch = domain_names[i : i + 5]
            try:
                details = opensearch.describe_domains(DomainNames=batch)
                for d in details.get("DomainStatusList", []):
                    cluster_config = d.get("ClusterConfig", {})
                    ebs_options = d.get("EBSOptions", {})

                    domain = OpenSearchDomainInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        domain_name=d.get("DomainName", ""),
                        domain_id=d.get("DomainId", ""),
                        instance_type=cluster_config.get("InstanceType", "unknown"),
                        instance_count=cluster_config.get("InstanceCount", 1),
                        storage_type="ebs" if ebs_options.get("EBSEnabled") else "instance",
                        storage_gb=float(ebs_options.get("VolumeSize", 0)),
                        engine_version=d.get("EngineVersion", ""),
                        created_at=d.get("Created"),
                    )
                    domains.append(domain)
            except ClientError:
                pass

    except ClientError:
        pass

    # 배치 메트릭 조회
    if domains:
        _collect_opensearch_metrics_batch(cloudwatch, account_id, domains, start_time, now)

    return domains


def _collect_opensearch_metrics_batch(
    cloudwatch,
    account_id: str,
    domains: list[OpenSearchDomainInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """OpenSearch 도메인 메트릭 배치 수집 (내부 함수)"""
    from botocore.exceptions import ClientError

    metrics_to_fetch = [
        ("CPUUtilization", "avg_cpu"),
        ("SearchableDocuments", "searchable_docs"),
        ("IndexingRate", "indexing_rate"),
        ("JVMMemoryPressure", "jvm_memory_pressure"),
    ]

    queries = []
    for domain in domains:
        safe_id = sanitize_metric_id(domain.domain_name)
        for metric_name, _ in metrics_to_fetch:
            metric_key = metric_name.lower()
            queries.append(
                MetricQuery(
                    id=f"{safe_id}_{metric_key}",
                    namespace="AWS/ES",
                    metric_name=metric_name,
                    dimensions={
                        "DomainName": domain.domain_name,
                        "ClientId": account_id,
                    },
                    stat="Average",
                )
            )

    try:
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for domain in domains:
            safe_id = sanitize_metric_id(domain.domain_name)
            for metric_name, attr_name in metrics_to_fetch:
                metric_key = metric_name.lower()
                value = results.get(f"{safe_id}_{metric_key}", 0.0)
                # SearchableDocuments는 평균이 아닌 최신값
                if metric_name == "SearchableDocuments":
                    setattr(domain, attr_name, value)
                else:
                    setattr(domain, attr_name, value / days)

    except ClientError:
        pass


def analyze_domains(
    domains: list[OpenSearchDomainInfo], account_id: str, account_name: str, region: str
) -> OpenSearchAnalysisResult:
    """OpenSearch 도메인 분석"""
    result = OpenSearchAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_domains=len(domains),
    )

    for domain in domains:
        # 미사용: CPU < 2% AND 문서 수 = 0
        if domain.avg_cpu < UNUSED_CPU_THRESHOLD and domain.searchable_docs == 0:
            result.unused_domains += 1
            result.unused_monthly_cost += domain.estimated_monthly_cost
            result.findings.append(
                OpenSearchDomainFinding(
                    domain=domain,
                    status=DomainStatus.UNUSED,
                    recommendation=f"미사용 (CPU {domain.avg_cpu:.1f}%, 문서 0) - 삭제 검토 (${domain.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: CPU < 5% AND 인덱싱 레이트 < 1/min
        if domain.avg_cpu < LOW_USAGE_CPU_THRESHOLD and domain.indexing_rate < LOW_USAGE_INDEXING_THRESHOLD:
            result.low_usage_domains += 1
            result.low_usage_monthly_cost += domain.estimated_monthly_cost
            result.findings.append(
                OpenSearchDomainFinding(
                    domain=domain,
                    status=DomainStatus.LOW_USAGE,
                    recommendation=f"저사용 (CPU {domain.avg_cpu:.1f}%, 인덱싱 {domain.indexing_rate:.2f}/min) - 다운사이징 검토",
                )
            )
            continue

        result.normal_domains += 1
        result.findings.append(
            OpenSearchDomainFinding(
                domain=domain,
                status=DomainStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[OpenSearchAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_domains,
                r.unused_domains,
                r.low_usage_domains,
                r.normal_domains,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
            ]
        )
        ws = summary_sheet._ws
        if r.unused_domains > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_domains > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Domain Name", width=30),
        ColumnDef(header="Instance Type", width=18),
        ColumnDef(header="Instances", width=10, style="number"),
        ColumnDef(header="Storage (GB)", width=12),
        ColumnDef(header="Engine", width=15),
        ColumnDef(header="Avg CPU", width=10),
        ColumnDef(header="Docs", width=12),
        ColumnDef(header="Index Rate", width=12),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="분석상태", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Domains", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != DomainStatus.NORMAL:
                d = f.domain
                style = Styles.danger() if f.status == DomainStatus.UNUSED else Styles.warning()

                detail_sheet.add_row(
                    [
                        d.account_name,
                        d.region,
                        d.domain_name,
                        d.instance_type,
                        d.instance_count,
                        f"{d.storage_gb:.0f}",
                        d.engine_version,
                        f"{d.avg_cpu:.1f}%",
                        f"{d.searchable_docs:,.0f}",
                        f"{d.indexing_rate:.2f}/min",
                        f"${d.estimated_monthly_cost:.2f}",
                        f.status.value,
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "OpenSearch_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> OpenSearchAnalysisResult | None:
    """단일 계정/리전의 OpenSearch 도메인 수집 및 분석 (병렬 실행용)"""
    domains = collect_opensearch_domains(session, account_id, account_name, region)
    if not domains:
        return None
    return analyze_domains(domains, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """OpenSearch 미사용 도메인 분석"""
    console.print("[bold]OpenSearch 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="opensearch")
    results: list[OpenSearchAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_domains for r in results)
    total_low = sum(r.low_usage_domains for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월)"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("opensearch", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
