"""SLA (Service Level Agreement) compliance sheet writer."""

from __future__ import annotations

import logging
from typing import Any

from .base import BaseSheetWriter, get_column_letter
from .config import SheetConfig

logger = logging.getLogger(__name__)

# SLA Sheet configuration
SHEET_NAME_SLA = "SLA 준수 현황"
HEADERS_SLA = (
    "임계값",
    "준수 요청",
    "미준수 요청",
    "준수율 (%)",
    "SLO 목표 (%)",
    "상태",
)

SLA_COLUMN_WIDTHS = {
    "임계값": 15,
    "준수 요청": 15,
    "미준수 요청": 15,
    "준수율 (%)": 12,
    "SLO 목표 (%)": 12,
    "상태": 10,
}


class SLASheetWriter(BaseSheetWriter):
    """Creates SLA compliance analysis sheet."""

    def write(self) -> None:
        """Create SLA compliance sheet."""
        try:
            sla_data = self.data.get("sla_compliance", {})
            if not sla_data:
                return

            ws = self.create_sheet(SHEET_NAME_SLA)
            headers = list(HEADERS_SLA)
            self.write_header_row(ws, headers)

            # Write SLA data rows
            self._write_sla_rows(ws, sla_data, headers)

            # Finalize sheet
            self._apply_sla_column_widths(ws, headers)
            ws.row_dimensions[1].height = SheetConfig.HEADER_ROW_HEIGHT

            ws.freeze_panes = ws.cell(row=2, column=1)  # pyright: ignore[reportAttributeAccessIssue]
            ws.sheet_view.zoomScale = SheetConfig.ZOOM_SCALE

        except Exception as e:
            logger.error(f"SLA 시트 생성 중 오류: {e}")

    def _write_sla_rows(
        self,
        ws,
        sla_data: dict[str, dict[str, Any]],
        headers: list[str],
    ) -> None:
        """Write SLA compliance data rows."""
        border = self.styles.thin_border

        # Order: 100ms, 500ms, 1s
        sla_order = ["under_100ms", "under_500ms", "under_1s"]

        row_idx = 2
        for sla_key in sla_order:
            if sla_key not in sla_data:
                continue

            item = sla_data[sla_key]
            rate = item.get("rate", 0)
            target = item.get("slo_target", 99.0)

            # Determine status
            if rate >= target:
                status = "Pass"
            elif rate >= target - 1:
                status = "Warning"
            else:
                status = "Fail"

            values = [
                item.get("threshold", ""),
                item.get("compliant", 0),
                item.get("non_compliant", 0),
                rate,
                target,
                status,
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                # Column-specific formatting
                if col_idx == 1:  # Threshold
                    cell.alignment = self.styles.align_center
                elif col_idx in (2, 3):  # Compliant/Non-compliant counts
                    cell.alignment = self.styles.align_right
                    cell.number_format = "#,##0"
                elif col_idx in (4, 5):  # Rates
                    cell.alignment = self.styles.align_right
                    cell.number_format = "0.00"
                elif col_idx == 6:  # Status
                    cell.alignment = self.styles.align_center
                    if value == "Pass":
                        from openpyxl.styles import PatternFill

                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "Warning":
                        cell.fill = self.styles.warning_fill
                    elif value == "Fail":
                        cell.fill = self.styles.danger_fill

            row_idx += 1

    def _apply_sla_column_widths(self, ws, headers: list[str]) -> None:
        """Apply column widths for SLA sheet."""
        for col, header in enumerate(headers, 1):
            width = SLA_COLUMN_WIDTHS.get(header, 15)
            ws.column_dimensions[get_column_letter(col)].width = width
