"""
pkg/io/excel/styles.py - Excel 스타일 상수 및 유틸리티

모든 도구에서 일관된 Excel 출력을 위한 색상, 정렬, 숫자 포맷 상수
"""

import logging
from typing import Any, Dict

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

# =============================================================================
# 색상 상수 (RGB Hex)
# =============================================================================

# 헤더/구조 색상
COLOR_HEADER_BG = "4472C4"  # 헤더 배경 (파란색)
COLOR_HEADER_FG = "FFFFFF"  # 헤더 글자 (흰색)
COLOR_SUMMARY_BG = "FFF2CC"  # 요약 배경 (연한 노랑)

# 상태 색상 (표준)
COLOR_SUCCESS = "C6EFCE"  # 성공 (연한 초록)
COLOR_SUCCESS_FG = "006100"  # 성공 글자 (진한 초록)
COLOR_WARNING = "FFEB9C"  # 경고 (연한 노랑)
COLOR_WARNING_FG = "9C5700"  # 경고 글자 (진한 노랑)
COLOR_DANGER = "FFC7CE"  # 위험 (연한 빨강)
COLOR_DANGER_FG = "9C0006"  # 위험 글자 (진한 빨강)

# 하이라이트 색상 (강조)
COLOR_ABUSE = "FF8080"  # 악성/차단 (진한 빨강)
COLOR_ERROR = "FFCCCC"  # 에러 (연한 빨강)
COLOR_INFO = "CCE5FF"  # 정보 (연한 파랑)

# 일반 색상
COLOR_DATA_BG = "FFFFFF"  # 데이터 배경 (흰색)
COLOR_ALT_ROW_BG = "F2F2F2"  # 교대 행 배경 (연한 회색)

# =============================================================================
# 숫자 포맷 상수
# =============================================================================

NUMBER_FORMAT_INTEGER = "#,##0"  # 정수 (1,234)
NUMBER_FORMAT_DECIMAL = "#,##0.00"  # 소수점 2자리 (1,234.56)
NUMBER_FORMAT_CURRENCY = "#,##0.00"  # 금액 (1,234.56)
NUMBER_FORMAT_PERCENT = "0.00%"  # 퍼센트 (12.34%)
NUMBER_FORMAT_DATE = "YYYY-MM-DD"  # 날짜
NUMBER_FORMAT_DATETIME = "YYYY-MM-DD HH:MM"  # 날짜시간
NUMBER_FORMAT_TEXT = "@"  # 텍스트 (문자열로 강제)
NUMBER_FORMAT_STATUS = "0"  # HTTP 상태코드 등 정수

# 별칭 (ALB 스타일 호환)
FMT_COUNT = NUMBER_FORMAT_INTEGER
FMT_STATUS = NUMBER_FORMAT_STATUS
FMT_TEXT = NUMBER_FORMAT_TEXT

# =============================================================================
# 기본 스타일 객체
# =============================================================================


def get_thin_border() -> Border:
    """얇은 테두리 스타일 반환"""
    thin_side = Side(style="thin", color="808080")
    return Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )


def get_header_font() -> Font:
    """헤더 폰트 스타일 반환"""
    return Font(
        name="맑은 고딕",
        size=10,
        bold=True,
        color=COLOR_HEADER_FG,
    )


def get_data_font() -> Font:
    """데이터 폰트 스타일 반환"""
    return Font(
        name="맑은 고딕",
        size=10,
        bold=False,
    )


def get_summary_font() -> Font:
    """요약 행 폰트 스타일 반환"""
    return Font(
        name="맑은 고딕",
        size=10,
        bold=True,
    )


# =============================================================================
# 정렬 상수
# =============================================================================

# 기본 정렬 (줄바꿈 없음)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)

# 줄바꿈 정렬 (긴 텍스트용)
ALIGN_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_WRAP = ALIGN_WRAP  # 별칭
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT_WRAP = Alignment(horizontal="right", vertical="center", wrap_text=True)


# =============================================================================
# 채우기 스타일
# =============================================================================


def get_header_fill() -> PatternFill:
    """헤더 채우기 스타일"""
    return PatternFill(
        start_color=COLOR_HEADER_BG,
        end_color=COLOR_HEADER_BG,
        fill_type="solid",
    )


def get_summary_fill() -> PatternFill:
    """요약 행 채우기 스타일"""
    return PatternFill(
        start_color=COLOR_SUMMARY_BG,
        end_color=COLOR_SUMMARY_BG,
        fill_type="solid",
    )


def get_success_fill() -> PatternFill:
    """성공 상태 채우기 스타일"""
    return PatternFill(
        start_color=COLOR_SUCCESS,
        end_color=COLOR_SUCCESS,
        fill_type="solid",
    )


def get_warning_fill() -> PatternFill:
    """경고 상태 채우기 스타일"""
    return PatternFill(
        start_color=COLOR_WARNING,
        end_color=COLOR_WARNING,
        fill_type="solid",
    )


def get_danger_fill() -> PatternFill:
    """위험 상태 채우기 스타일"""
    return PatternFill(
        start_color=COLOR_DANGER,
        end_color=COLOR_DANGER,
        fill_type="solid",
    )


def get_abuse_fill() -> PatternFill:
    """악성/차단 하이라이트 채우기 스타일 (진한 빨강)"""
    return PatternFill(
        start_color=COLOR_ABUSE,
        end_color=COLOR_ABUSE,
        fill_type="solid",
    )


def get_error_fill() -> PatternFill:
    """에러 하이라이트 채우기 스타일 (연한 빨강)"""
    return PatternFill(
        start_color=COLOR_ERROR,
        end_color=COLOR_ERROR,
        fill_type="solid",
    )


def get_info_fill() -> PatternFill:
    """정보 하이라이트 채우기 스타일 (연한 파랑)"""
    return PatternFill(
        start_color=COLOR_INFO,
        end_color=COLOR_INFO,
        fill_type="solid",
    )


# 미리 생성된 Fill 인스턴스 (ALB 스타일 호환)
FILL_ABUSE = get_abuse_fill()
FILL_WARN = get_warning_fill()
FILL_ERROR = get_error_fill()
FILL_INFO = get_info_fill()
FILL_SUCCESS = get_success_fill()
FILL_DANGER = get_danger_fill()


# =============================================================================
# 행 스타일 프리셋 (RowStyle)
# =============================================================================


class RowStyle:
    """행 수준 스타일 프리셋"""

    @staticmethod
    def data() -> dict:
        """일반 데이터 행 스타일"""
        return {
            "fill": None,
            "font": get_data_font(),
        }

    @staticmethod
    def warning() -> dict:
        """경고 행 스타일 (연한 노랑)"""
        return {
            "fill": get_warning_fill(),
            "font": get_data_font(),
        }

    @staticmethod
    def danger() -> dict:
        """위험 행 스타일 (연한 빨강)"""
        return {
            "fill": get_danger_fill(),
            "font": get_data_font(),
        }

    @staticmethod
    def success() -> dict:
        """성공 행 스타일 (연한 초록)"""
        return {
            "fill": get_success_fill(),
            "font": get_data_font(),
        }

    @staticmethod
    def summary() -> dict:
        """요약 행 스타일 (연한 노랑, 볼드)"""
        return {
            "fill": get_summary_fill(),
            "font": get_summary_font(),
        }

    @staticmethod
    def abuse() -> dict:
        """악성/차단 행 스타일 (진한 빨강)"""
        return {
            "fill": get_abuse_fill(),
            "font": get_data_font(),
        }

    @staticmethod
    def error() -> dict:
        """에러 행 스타일 (연한 빨강)"""
        return {
            "fill": get_error_fill(),
            "font": get_data_font(),
        }

    @staticmethod
    def info() -> dict:
        """정보 행 스타일 (연한 파랑)"""
        return {
            "fill": get_info_fill(),
            "font": get_data_font(),
        }


# 편의를 위한 Styles 클래스
class Styles:
    """스타일 프리셋 접근용 클래스"""

    data = RowStyle.data
    warning = RowStyle.warning
    danger = RowStyle.danger
    success = RowStyle.success
    summary = RowStyle.summary
    abuse = RowStyle.abuse
    error = RowStyle.error
    info = RowStyle.info


# =============================================================================
# 스타일 Dict 생성 함수
# =============================================================================


def get_header_style() -> Dict[str, Any]:
    """헤더 스타일 dict 반환 (셀 적용용)"""
    return {
        "font": get_header_font(),
        "fill": get_header_fill(),
        "alignment": ALIGN_CENTER_WRAP,
        "border": get_thin_border(),
    }


def get_basic_header_style() -> Dict[str, Any]:
    """기본 헤더 스타일 (Consolas 폰트, 연한 파랑 배경)"""
    return {
        "font": Font(name="Consolas", bold=True),
        "fill": PatternFill(
            start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": get_thin_border(),
    }


def get_center_header_style() -> Dict[str, Any]:
    """중앙 정렬 헤더 스타일 (12pt)"""
    return {
        "font": Font(name="Consolas", bold=True, size=12),
        "fill": PatternFill(
            start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"
        ),
        "alignment": ALIGN_CENTER,
        "border": get_thin_border(),
    }


def get_data_cell_style() -> Dict[str, Any]:
    """데이터 셀 스타일"""
    return {
        "font": Font(name="Consolas", size=11),
        "alignment": ALIGN_LEFT,
        "border": get_thin_border(),
    }


def get_summary_cell_style() -> Dict[str, Any]:
    """요약/합계 셀 스타일"""
    return {
        "font": Font(name="Consolas", bold=True),
        "fill": PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        ),
        "alignment": ALIGN_CENTER,
        "border": get_thin_border(),
    }


def get_center_alignment(wrap_text: bool = True) -> Alignment:
    """중앙 정렬 반환 (wrap_text 선택)"""
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)


def create_status_style(status: str) -> Dict[str, Any]:
    """상태에 따른 스타일 반환

    Args:
        status: 상태 문자열 (success, failed, warning, info, excellent, good, poor)
    """
    status_colors = {
        "success": COLOR_SUCCESS,
        "failed": COLOR_DANGER,
        "warning": COLOR_WARNING,
        "info": COLOR_INFO,
        "excellent": "92D050",
        "good": "FFFF00",
        "poor": "FF6B6B",
    }

    color = status_colors.get(status.lower(), "FFFFFF")
    return {
        "font": Font(bold=True),
        "alignment": ALIGN_CENTER,
        "fill": PatternFill(start_color=color, end_color=color, fill_type="solid"),
        "border": get_thin_border(),
    }


def create_summary_cell_style() -> Dict[str, Any]:
    """요약/중요 셀 스타일"""
    return {
        "font": Font(bold=True),
        "alignment": ALIGN_CENTER,
        "fill": PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        ),
        "border": get_thin_border(),
    }


def create_summary_dashboard_style() -> Dict[str, Any]:
    """대시보드 요약 섹션 스타일"""
    return {
        "font": Font(bold=True, size=12),
        "alignment": ALIGN_CENTER,
        "fill": get_header_fill(),
        "border": get_thin_border(),
    }


def create_summary_value_style() -> Dict[str, Any]:
    """요약 값 스타일"""
    return {
        "font": Font(size=11),
        "alignment": ALIGN_CENTER,
        "border": get_thin_border(),
    }
