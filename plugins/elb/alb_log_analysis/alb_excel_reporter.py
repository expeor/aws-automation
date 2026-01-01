import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.progress import (
    BarColumn,
    Progress,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
)

# 콘솔 및 로거 (aa_cli.aa.ui 또는 로컬 생성)
try:
    from cli.ui import console, logger
except ImportError:
    import logging

    console = Console()
    logger = logging.getLogger(__name__)

from core.tools.io.excel import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_LEFT_WRAP,
    ALIGN_RIGHT,
    FILL_ABUSE,
    FMT_TEXT,
    get_header_style,
    get_thin_border,
)


class ExcelReportError(Exception):
    """Excel 보고서 생성 중 발생하는 오류"""


class _SummarySheetHelper:
    """분석 요약 시트 생성을 위한 헬퍼 클래스 (pkg/io/excel.SummarySheet 스타일)"""

    # 스타일 상수
    TITLE_FONT = Font(name="Consolas", size=16, bold=True, color="1F4E79")
    HEADER_FONT = Font(name="Consolas", size=12, bold=True, color="2F5597")
    LABEL_FONT = Font(name="Consolas", size=11, bold=True)
    VALUE_FONT = Font(name="Consolas", size=11)
    TITLE_FILL = PatternFill(
        start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
    )
    HEADER_FILL = PatternFill(
        start_color="EBF1FA", end_color="EBF1FA", fill_type="solid"
    )
    DANGER_FILL = PatternFill(
        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
    )
    WARNING_FILL = PatternFill(
        start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"
    )

    def __init__(self, ws, border):
        self._ws = ws
        self._border = border
        self._current_row = 1
        # 기본 컬럼 너비
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25

    def add_title(self, title: str) -> "_SummarySheetHelper":
        """제목 추가"""
        ws = self._ws
        ws.merge_cells(f"A{self._current_row}:B{self._current_row}")
        cell = ws.cell(row=self._current_row, column=1, value=title)
        cell.font = self.TITLE_FONT
        cell.fill = self.TITLE_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = self._border
        ws.cell(row=self._current_row, column=2).border = self._border
        ws.row_dimensions[self._current_row].height = 40
        self._current_row += 2
        return self

    def add_section(self, section_name: str) -> "_SummarySheetHelper":
        """섹션 헤더 추가"""
        ws = self._ws
        ws.merge_cells(f"A{self._current_row}:B{self._current_row}")
        cell = ws.cell(row=self._current_row, column=1, value=section_name)
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = self._border
        ws.cell(row=self._current_row, column=2).border = self._border
        self._current_row += 1
        return self

    def add_item(
        self,
        label: str,
        value: Any,
        highlight: Optional[str] = None,
        number_format: Optional[str] = None,
    ) -> "_SummarySheetHelper":
        """항목 추가"""
        ws = self._ws

        label_cell = ws.cell(row=self._current_row, column=1, value=label)
        label_cell.font = self.LABEL_FONT
        label_cell.alignment = ALIGN_LEFT
        label_cell.border = self._border

        value_cell = ws.cell(row=self._current_row, column=2, value=value)
        value_cell.font = self.VALUE_FONT
        value_cell.alignment = ALIGN_LEFT
        value_cell.border = self._border

        if number_format:
            value_cell.number_format = number_format

        if highlight == "danger":
            value_cell.fill = self.DANGER_FILL
        elif highlight == "warning":
            value_cell.fill = self.WARNING_FILL

        self._current_row += 1
        return self

    def add_blank_row(self) -> "_SummarySheetHelper":
        """빈 행 추가"""
        self._current_row += 1
        return self

    def add_list_section(
        self,
        section_label: str,
        items: List[Tuple[str, Any]],
        max_items: int = 5,
        value_suffix: str = "",
    ) -> "_SummarySheetHelper":
        """순위 리스트 섹션 (예: Top 5 URL)"""
        ws = self._ws

        # 섹션 레이블
        label_cell = ws.cell(row=self._current_row, column=1, value=f"{section_label}:")
        label_cell.font = self.LABEL_FONT
        label_cell.alignment = ALIGN_LEFT
        label_cell.border = self._border
        ws.cell(row=self._current_row, column=2).border = self._border
        self._current_row += 1

        if not items:
            cell = ws.cell(row=self._current_row, column=1, value="데이터 없음")
            cell.font = self.VALUE_FONT
            cell.border = self._border
            ws.cell(row=self._current_row, column=2).border = self._border
            self._current_row += 1
        else:
            for i, (name, count) in enumerate(items[:max_items], 1):
                display_name = (
                    str(name)[:47] + "..." if len(str(name)) > 50 else str(name)
                )
                name_cell = ws.cell(
                    row=self._current_row, column=1, value=f"{i}. {display_name}"
                )
                name_cell.font = self.VALUE_FONT
                name_cell.alignment = ALIGN_LEFT
                name_cell.border = self._border

                display_value = (
                    f"{count:,}{value_suffix}"
                    if isinstance(count, (int, float))
                    else str(count)
                )
                count_cell = ws.cell(
                    row=self._current_row, column=2, value=display_value
                )
                count_cell.font = self.VALUE_FONT
                count_cell.alignment = ALIGN_RIGHT
                count_cell.border = self._border
                self._current_row += 1

        return self

    @property
    def current_row(self) -> int:
        return self._current_row

    @current_row.setter
    def current_row(self, value: int):
        self._current_row = value


# =============================================================================
# 공통 스타일 상수
# =============================================================================
FONT_BOLD = Font(name="Consolas", bold=True)
FONT_ITALIC_GRAY = Font(name="Consolas", italic=True, color="808080")

# =============================================================================
# 컬럼 스타일 정의 (헤더 → 스타일 타입 매핑)
# =============================================================================

# 스타일 타입: "number" (우측 정렬, #,##0), "center" (가운데), "data" (좌측), "decimal" (0.00), "decimal3" (0.000)
COLUMN_STYLE_MAP: Dict[str, str] = {
    # 숫자 컬럼 (우측 정렬, 천단위 구분)
    "Count": "number",
    "Unique IPs": "number",
    "IP Count": "number",
    # 소수점 컬럼
    "Percentage": "decimal",
    "Avg Response Time": "decimal3",
    "Error Rate (%)": "decimal",
    # 상태 코드 (정수)
    "Top Status": "status",
    "ELB Status Code": "status",
    "Backend Status Code": "status",
    # 가운데 정렬 컬럼
    "Client": "center",
    "Country": "center",
    "Abuse": "center",
    "Method": "center",
    "Target": "center",
    "Target group name": "center",
    "Timestamp": "center",
    "Error Reason": "center",
    # 좌측 정렬 (기본값)
    "Request": "data",
    "Redirect URL": "data",
    "User Agent": "data",
}

# 컬럼 너비 정의
COLUMN_WIDTH_MAP: Dict[str, int] = {
    "Count": 11,
    "Client": 20,
    "Country": 10,
    "Abuse": 10,
    "Method": 9,
    "Request": 80,
    "Redirect URL": 60,
    "User Agent": 80,
    "Target": 20,
    "Target group name": 20,
    "Timestamp": 22,
    "ELB Status Code": 12,
    "Backend Status Code": 12,
    "Unique IPs": 12,
    "Avg Response Time": 15,
    "Top Status": 12,
    "Error Rate (%)": 12,
    "IP Count": 15,
    "Percentage": 15,
    "ASN": 15,
    "ISP": 40,
    "IP": 20,
    "Error Reason": 40,
    "Response time": 15,
    # 바이트 관련 컬럼
    "수신 데이터 (Bytes)": 20,
    "송신 데이터 (Bytes)": 20,
    "총 데이터 (Bytes)": 20,
    "총 데이터 (변환)": 25,
}


class ALBExcelReporter:
    """ALB 로그 분석 결과를 Excel 보고서로 생성하는 클래스"""

    def __init__(
        self,
        data: Optional[Dict[str, Any]] = None,
        output_dir: str = "reports",
    ) -> None:
        self.output_dir = output_dir
        self.data: Dict[str, Any] = data or {}
        os.makedirs(output_dir, exist_ok=True)

        # 스타일 구성
        self.workbook = Workbook()
        self.header_style = get_header_style()
        self.thin_border = get_thin_border()
        self.cell_alignment = ALIGN_LEFT_WRAP
        self.abuse_fill = FILL_ABUSE

    # =========================================================================
    # 공통 헬퍼 메서드
    # =========================================================================

    def _write_header_row(self, ws, headers: List[str], row: int = 1) -> None:
        """헤더 행을 작성하고 스타일을 적용합니다."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_style["font"]
            cell.fill = self.header_style["fill"]
            cell.alignment = self.header_style["alignment"]
            cell.border = self.header_style["border"]

    def _write_empty_message(self, ws, message: str, row: int, col_count: int) -> None:
        """데이터가 없을 때 메시지를 표시합니다."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
        cell = ws.cell(row=row, column=1, value=message)
        cell.alignment = ALIGN_CENTER
        cell.font = FONT_ITALIC_GRAY
        cell.border = self.thin_border

    def _finalize_sheet(
        self,
        ws,
        headers: List[str],
        data_count: int = 0,
        header_height: int = 40,
        zoom: int = 85,
    ) -> None:
        """시트 마무리: 컬럼 스타일, 필터, freeze panes 등 적용"""
        self._apply_column_styles(ws, headers)
        ws.row_dimensions[1].height = header_height
        if data_count > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_count + 1}"
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.sheet_view.zoomScale = zoom

    def _apply_range_style(
        self,
        ws,
        cell_range: str,
        fill: Optional[PatternFill] = None,
        font: Optional[Font] = None,
        border=None,
        alignment: Optional[Alignment] = None,
    ) -> None:
        try:
            for row in ws[cell_range]:
                for cell in row:
                    if fill is not None:
                        cell.fill = fill
                    if font is not None:
                        cell.font = font
                    if border is not None:
                        cell.border = border
                    if alignment is not None:
                        cell.alignment = alignment
        except Exception as exc:  # noqa: BLE001
            logger.debug(f"범위 스타일 적용 중 오류 (무시): {cell_range} - {exc}")

    def _apply_column_styles(self, ws, headers: List[str]) -> None:
        """헤더 이름을 기반으로 컬럼별 스타일과 너비를 적용합니다.

        wrap_text가 필요한 컬럼들에 자동 줄바꿈을 적용합니다.
        """
        # wrap_text가 필요한 컬럼과 정렬 방향
        WRAP_TEXT_COLUMNS = {
            "Request": "left",
            "Redirect URL": "left",
            "User Agent": "left",
            "Client": "center",
            "Target": "center",
            "Target group name": "center",
            "Country": "center",
        }

        try:
            # 1. 컬럼 너비 적용
            self._apply_column_widths_by_map(ws, headers)

            # 2. wrap_text 적용
            for col_idx, header in enumerate(headers, start=1):
                if header not in WRAP_TEXT_COLUMNS:
                    continue

                h_align = WRAP_TEXT_COLUMNS[header]
                for row_idx in range(2, ws.max_row + 1):  # 헤더 제외
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.alignment:
                        cell.alignment = cell.alignment.copy(wrap_text=True)
                    else:
                        cell.alignment = Alignment(
                            horizontal=h_align, vertical="center", wrap_text=True
                        )
        except Exception as exc:  # noqa: BLE001
            logger.error(f"컬럼 스타일 적용 중 오류 발생: {exc}")

    def _enforce_vertical_center_alignment(self, ws) -> None:
        """모든 셀의 세로 정렬을 가운데로 맞춥니다. 가로 정렬 및 줄바꿈 설정은 보존합니다."""
        try:
            if ws.max_row == 0 or ws.max_column == 0:
                return
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    try:
                        existing = cell.alignment
                        if existing is not None:
                            if getattr(existing, "vertical", None) != "center":
                                cell.alignment = existing.copy(vertical="center")
                        else:
                            cell.alignment = Alignment(vertical="center")
                    except Exception:
                        # 개별 셀 정렬 실패는 무시 (다음 셀 계속)
                        continue
        except Exception as exc:  # noqa: BLE001
            logger.debug(f"세로 정렬 적용 중 오류 (무시): {ws.title} - {exc}")

    def _enforce_vertical_center_for_workbook(self, wb: Workbook) -> None:
        """워크북의 모든 워크시트에 세로 가운데 정렬을 강제합니다."""
        try:
            for ws in wb.worksheets:
                self._enforce_vertical_center_alignment(ws)
        except Exception as exc:  # noqa: BLE001
            logger.debug(f"워크북 세로 정렬 적용 중 오류 (무시): {exc}")

    def create_report(
        self, data: Dict[str, Any], report_name: Optional[str] = None
    ) -> str:
        try:
            wb = Workbook()
            if wb.active is not None:
                wb.remove(wb.active)

            use_progress = console.is_terminal

            if use_progress:
                progress_columns = [
                    TextColumn("[bold blue]{task.description}", justify="right"),
                    BarColumn(
                        bar_width=40, complete_style="green", finished_style="green"
                    ),
                    TaskProgressColumn(),
                    TimeElapsedColumn(),
                    TimeRemainingColumn(),
                ]
                with Progress(*progress_columns, console=console) as progress:
                    main_task = progress.add_task("[bold]Excel 보고서 생성중...", total=8)

                    def update_progress(
                        description: str, finished: bool = False
                    ) -> None:
                        if finished:
                            progress.update(main_task, advance=1)
                        progress.update(main_task, description=description)

                    output_path = self._create_report_internal(
                        wb, data, report_name, update_progress
                    )
            else:
                output_path = self._create_report_internal(wb, data, report_name, None)
            return output_path
        except Exception as exc:  # noqa: BLE001
            logger.error(f"엑셀 보고서 생성 중 오류 발생: {exc}")
            raise

    def generate_report(self, report_name: Optional[str] = None) -> str:
        return self.create_report(self.data, report_name)

    def _create_report_internal(
        self,
        wb: Workbook,
        data: Dict[str, Any],
        report_name: Optional[str] = None,
        update_progress: Optional[Any] = None,
    ) -> str:
        # Progress 헬퍼 - None일 때 no-op
        def progress(msg: str, finished: bool = False) -> None:
            if update_progress:
                update_progress(msg, finished=finished)

        # 시트 생성 작업 정의: (조건, 시트명, 생성함수, 인자)
        # 조건이 None이면 항상 실행, callable이면 호출하여 판단
        def always() -> bool:
            return True

        # 1) 전체 개요 섹션
        progress("[bold blue]분석 요약 시트 생성중...")
        self._create_summary_sheet(wb, data)
        progress("[bold blue]분석 요약 시트 생성 완료", finished=True)

        # 조건부 시트들 정의
        conditional_sheets = [
            # (조건, 시트명, 생성함수)
            (
                lambda: data.get("country_statistics"),
                "국가별 통계",
                lambda: self._add_country_statistics_sheet(wb, data),
            ),
            (
                lambda: data.get("request_url_details")
                or data.get("request_url_counts"),
                "요청 URL Top 100",
                lambda: self._add_url_request_sheet(wb, data),
            ),
            (
                lambda: data.get("client_status_statistics"),
                "Client 상태코드 통계",
                lambda: self._add_client_status_statistics_sheet(wb, data),
            ),
            (
                lambda: data.get("target_status_statistics"),
                "Target 상태코드 통계",
                lambda: self._add_target_backend_status_statistics_sheet(wb, data),
            ),
            (
                lambda: data.get("long_response_times"),
                "응답 시간",
                lambda: self._create_response_time_sheet(wb, data),
            ),
            (
                lambda: data.get("received_bytes") or data.get("sent_bytes"),
                "데이터 전송량 Top 100",
                lambda: self._create_bytes_analysis_sheet(wb, data),
            ),
        ]

        # 조건부 시트 생성 실행
        for condition, sheet_name, create_fn in conditional_sheets:
            if condition():
                progress(f"[bold blue]{sheet_name} 시트 생성중...")
                create_fn()
                progress(f"[bold blue]{sheet_name} 시트 생성 완료", finished=True)

        # 3) 보안/문제 해결 섹션 (특수 처리 - 연속 2개 시트)
        abuse_ips_list, abuse_ip_details = self._get_normalized_abuse_ips(data)
        if abuse_ips_list:
            progress("[bold blue]Abuse IP 시트 생성중...")
            self._create_abuse_ip_worksheet(wb, abuse_ips_list, abuse_ip_details, data)
            progress("[bold blue]Abuse IP 시트 생성 완료", finished=True)

            progress("[bold blue]악성 IP 요청 분석 시트 생성중...")
            self._create_abuse_requests_sheet(wb, data)
            progress("[bold blue]악성 IP 요청 분석 시트 생성 완료", finished=True)

        # 4) 상세 분석 섹션
        progress("[bold blue]상태 코드 상세 시트 생성중...")
        abuse_ips_set = self._get_matching_abuse_ips(data)
        self._add_status_code_sheets(wb, data, abuse_ips_set)
        progress("[bold blue]상태 코드 상세 시트 생성 완료", finished=True)

        # 5) 저장
        progress("[bold blue]파일 저장중...")
        output_path = self._save_workbook(wb, data, report_name)
        progress("[bold green]완료!")
        return output_path

    def _save_workbook(
        self, wb: Workbook, data: Dict[str, Any], report_name: Optional[str]
    ) -> str:
        if report_name:
            # report_name이 전체 경로인지 파일명인지를 구분하여 처리
            is_full_path = os.path.isabs(report_name) or any(
                sep in report_name for sep in ("/", "\\")
            )
            if is_full_path:
                output_path = (
                    report_name
                    if report_name.lower().endswith(".xlsx")
                    else f"{report_name}.xlsx"
                )
            else:
                output_path = os.path.join(self.output_dir, f"{report_name}.xlsx")
        else:
            from secrets import token_hex

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # 간결/고유 파일명 생성 (analyzer 기본값과 일관성 유지)
            s3_uri = data.get("s3_uri", "")
            account_id = "acct"
            region = "region"
            try:
                if "/AWSLogs/" in (s3_uri or ""):
                    path = s3_uri.replace("s3://", "")
                    parts = path.split("/AWSLogs/")[1].split("/")
                    if len(parts) >= 3:
                        account_id = parts[0]
                        region = parts[2]
            except Exception:
                pass
            alb_name = (
                str(data.get("alb_name") or "alb")
                .strip()
                .replace("/", "-")
                .replace("\\", "-")
            )
            filename = f"{account_id}_{region}_{alb_name}_reporter_{token_hex(4)}.xlsx"
            output_path = os.path.join(self.output_dir, filename)
        # 모든 워크시트의 세로 정렬을 가운데로 강제 적용
        try:
            self._enforce_vertical_center_for_workbook(wb)
        except Exception:
            # 정렬 강제 실패는 저장을 막지 않음
            pass
        wb.save(output_path)
        return output_path

    def _generate_filename_from_s3_path(
        self, data: Dict[str, Any], timestamp: str
    ) -> str:
        """AWS ALB 로그 파일 네이밍 규칙에 맞는 Excel 리포트 파일명을 생성합니다."""
        try:
            s3_uri = data.get("s3_uri", "")
            bucket_name = data.get("s3_bucket_name", "")
            prefix = data.get("s3_prefix", "")

            if not s3_uri:
                return f"ALB_Log_Analysis_{timestamp}.xlsx"

            # S3 URI에서 AWS Account ID, Region, Load Balancer ID 추출
            # 형식: s3://bucket/prefix/AWSLogs/aws-account-id/elasticloadbalancing/region/...
            path = s3_uri.replace("s3://", "")

            if "/AWSLogs/" not in path:
                return f"ALB_Log_Analysis_{timestamp}.xlsx"

            # AWSLogs 이후 경로 파싱
            aws_logs_path = path.split("/AWSLogs/")[1]
            path_parts = aws_logs_path.split("/")

            if len(path_parts) < 3:
                return f"ALB_Log_Analysis_{timestamp}.xlsx"

            aws_account_id = path_parts[0]  # aws-account-id
            service = path_parts[1]  # elasticloadbalancing
            region = path_parts[2]  # region

            # Load Balancer ID 추출 (prefix에서 ALB 관련 키워드 추출)
            load_balancer_id = "elb"
            if prefix:
                # prefix에서 로드밸런서 정보 추출
                prefix_clean = prefix.strip("/")
                # ALB나 ELB 관련 키워드가 있는 경우 추출
                prefix_parts = prefix_clean.split("/")
                for part in prefix_parts:
                    if part and (
                        "alb" in part.lower()
                        or "elb" in part.lower()
                        or "load" in part.lower()
                    ):
                        load_balancer_id = part.replace("/", "-").replace("_", "-")
                        break

                # 특별한 키워드가 없으면 마지막 부분 사용
                if load_balancer_id == "elb" and prefix_parts and prefix_parts[-1]:
                    load_balancer_id = (
                        prefix_parts[-1].replace("/", "-").replace("_", "-")
                    )

            # bucket 이름에서도 확인
            if load_balancer_id == "elb" and bucket_name:
                if (
                    "alb" in bucket_name.lower()
                    or "elb" in bucket_name.lower()
                    or "load" in bucket_name.lower()
                ):
                    load_balancer_id = bucket_name.replace("_", "-")

            # AWS ALB 로그 파일명 형식에 맞는 Excel 리포트 파일명 생성
            # 형식: aws-account-id_elasticloadbalancing_region_app.load-balancer-id_report_YYYYMMDD_HHMMSS.xlsx
            filename = f"{aws_account_id}_{service}_{region}_app.{load_balancer_id}_report_{timestamp}.xlsx"

            return filename

        except Exception as exc:  # noqa: BLE001
            logger.error(f"파일명 생성 중 오류 발생: {exc}")
            return f"ALB_Log_Analysis_{timestamp}.xlsx"

    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """분석 요약 시트 생성 - _SummarySheetHelper 활용으로 코드 간소화"""
        try:
            ws = wb.create_sheet("분석 요약", 0)
            summary = _SummarySheetHelper(ws, self.thin_border)

            # 1. 제목
            summary.add_title("ALB 로그 분석 보고서")

            # 2. 분석 정보 섹션
            s3_uri = data.get("s3_uri", "")
            bucket_name, account_id, region, service_prefix = self._parse_s3_uri(s3_uri)
            alb_name = data.get("alb_name", "N/A")

            summary.add_section("분석 정보")
            summary.add_item("계정 번호", account_id, number_format="@")
            summary.add_item("ALB 이름", alb_name)
            summary.add_item("S3 버킷", bucket_name)
            summary.add_item("리전", region)

            # 3. 분석 기간 섹션
            summary.add_blank_row()
            summary.add_section("분석 기간")
            summary.add_item("요청 시작 시간", data.get("start_time", "N/A"))
            summary.add_item("요청 종료 시간", data.get("end_time", "N/A"))
            summary.add_item("타임존", data.get("timezone", "N/A"))

            if data.get("actual_start_time") and data.get("actual_start_time") != "N/A":
                summary.add_item("실제 로그 시작", data.get("actual_start_time", "N/A"))
                summary.add_item("실제 로그 종료", data.get("actual_end_time", "N/A"))

            # 4. 데이터 통계 섹션
            summary.add_blank_row()
            summary.add_section("데이터 통계")
            summary.add_item("총 로그 라인 수", f"{data.get('log_lines_count', 0):,}개")
            summary.add_item("분석된 로그 파일 수", f"{data.get('log_files_count', 0):,}개")
            summary.add_item("고유 클라이언트 IP 수", f"{data.get('unique_client_ips', 0):,}개")
            summary.add_item(
                "총 수신 바이트",
                self._format_bytes(data.get("total_received_bytes", 0)),
            )
            summary.add_item(
                "총 송신 바이트", self._format_bytes(data.get("total_sent_bytes", 0))
            )
            summary.add_item("평균 응답 시간", self._calculate_average_response_time(data))
            summary.add_item("전체 에러율", self._calculate_error_rate(data))

            # 5. HTTP 상태 코드 통계 (별도 헬퍼로 처리)
            summary.add_blank_row()
            self._add_status_code_statistics_v2(summary, data)

            # 6. 보안 정보 섹션
            summary.add_blank_row()
            summary.add_section("보안 정보")

            abuse_count = self._get_abuse_ip_count(data)
            highlight = "danger" if abuse_count > 0 else None
            summary.add_item("탐지된 Abuse IP", f"{abuse_count:,}개", highlight=highlight)

            # Abuse IP 요청 수 계산
            abuse_total_requests = 0
            try:
                client_ip_counts = data.get("client_ip_counts", {})
                if isinstance(client_ip_counts, dict):
                    matching_abuse_ips = self._get_matching_abuse_ips(data)
                    abuse_total_requests = int(
                        sum(
                            int(client_ip_counts.get(ip, 0) or 0)
                            for ip in matching_abuse_ips
                        )
                    )
            except Exception:
                pass
            highlight = "danger" if abuse_total_requests > 0 else None
            summary.add_item(
                "전체 Abuse IP 요청 수",
                f"{abuse_total_requests:,}개",
                highlight=highlight,
            )

            # 7. 요청 패턴 분석 섹션
            summary.add_blank_row()
            summary.add_section("요청 패턴 분석")
            top_urls = self._get_top_request_urls(data, 5)
            summary.add_list_section("상위 요청 URL", top_urls, value_suffix="회")

            summary.add_blank_row()
            top_agents = self._get_top_user_agents(data, 5)
            summary.add_list_section("상위 User Agent", top_agents, value_suffix="회")

            # 8. 지리적 분석 섹션
            summary.add_blank_row()
            summary.add_section("지리적 분석")
            top_countries = self._get_top_countries(data, 5)
            summary.add_list_section("상위 국가", top_countries, value_suffix="개 IP")

            # 9. 성능 분석 섹션
            summary.add_blank_row()
            summary.add_section("성능 분석")
            response_stats = self._calculate_response_time_stats(data)
            summary.add_item("최대 응답 시간", response_stats["max"])
            summary.add_item("최소 응답 시간", response_stats["min"])
            summary.add_item("중간 응답 시간", response_stats["median"])

            # 줌 설정
            ws.sheet_view.zoomScale = 85

        except Exception as exc:  # noqa: BLE001
            logger.error(f"Summary 시트 생성 중 오류 발생: {exc}")

    def _add_status_code_statistics_v2(
        self, summary: "_SummarySheetHelper", data: Dict[str, Any]
    ) -> None:
        """HTTP 상태 코드 통계를 SummarySheetHelper에 추가"""
        summary.add_section("HTTP 상태 코드 통계")

        total_requests = (
            data.get("elb_2xx_count", 0)
            + data.get("elb_3xx_count", 0)
            + data.get("elb_4xx_count", 0)
            + data.get("elb_5xx_count", 0)
        )

        status_codes = [
            ("ELB 2xx", "elb_2xx_count", None),
            ("ELB 3xx", "elb_3xx_count", None),
            ("ELB 4xx", "elb_4xx_count", "warning"),
            ("ELB 5xx", "elb_5xx_count", "danger"),
            ("Backend 4xx", "backend_4xx_count", "warning"),
            ("Backend 5xx", "backend_5xx_count", "danger"),
        ]

        for label, key, highlight_type in status_codes:
            count = data.get(key, 0)
            if total_requests > 0 and key.startswith("elb_"):
                percentage = (count / total_requests) * 100
                display_value = f"{count:,}개 ({percentage:.1f}%)"
            else:
                display_value = f"{count:,}개"

            highlight = highlight_type if count > 0 else None
            summary.add_item(label, display_value, highlight=highlight)

    def _parse_s3_uri(self, s3_uri: str) -> Tuple[str, str, str, str]:
        bucket_name = account_id = region = service_prefix = "N/A"
        if s3_uri:
            try:
                path = s3_uri.replace("s3://", "")
                parts = path.split("/")
                if parts:
                    bucket_name = parts[0]
                if "/AWSLogs/" in path:
                    prefix_part = path.split("/AWSLogs/")[0]
                    service_prefix = (
                        prefix_part.split("/", 1)[1]
                        if "/" in prefix_part
                        else prefix_part
                    )
                    awslogs_part = path.split("/AWSLogs/")[1]
                    awslogs_parts = awslogs_part.split("/")
                    if awslogs_parts:
                        account_id = awslogs_parts[0]
                    if len(awslogs_parts) > 2:
                        region = awslogs_parts[2]
            except Exception as exc:  # noqa: BLE001
                logger.warning(f"S3 URI 파싱 중 오류: {exc}")
        return bucket_name, account_id, region, service_prefix

    def _format_bytes(self, size: Union[int, float, str, None]) -> str:
        try:
            if size is None:
                size = 0
            size = float(size)
            for unit in ["", "KB", "MB", "GB", "TB"]:
                if size < 1024.0:
                    return f"{size:.2f} {unit}"
                size /= 1024.0
            return f"{size:.2f} PB"
        except (ValueError, TypeError):
            return "N/A"

    def _calculate_average_response_time(self, data: Dict[str, Any]) -> str:
        """응답 시간 데이터로부터 평균 응답 시간을 계산합니다."""
        try:
            long_response_times = data.get("long_response_times", [])
            if not long_response_times or not isinstance(long_response_times, list):
                return "0.000초"

            total_response_time = 0.0
            valid_count = 0

            for log_entry in long_response_times:
                if isinstance(log_entry, dict) and "response_time" in log_entry:
                    response_time = log_entry.get("response_time", 0)
                    if isinstance(response_time, (int, float)) and response_time > 0:
                        total_response_time += float(response_time)
                        valid_count += 1

            if valid_count == 0:
                return "0.000초"

            avg_response_time = total_response_time / valid_count
            return f"{avg_response_time:.3f}초"

        except Exception as exc:  # noqa: BLE001
            logger.error(f"평균 응답 시간 계산 중 오류 발생: {exc}")
            return "N/A"

    def _calculate_error_rate(self, data: Dict[str, Any]) -> str:
        """전체 요청 대비 에러 요청 비율을 계산합니다."""
        try:
            total_requests = (
                data.get("elb_2xx_count", 0)
                + data.get("elb_3xx_count", 0)
                + data.get("elb_4xx_count", 0)
                + data.get("elb_5xx_count", 0)
            )

            if total_requests == 0:
                return "0.0%"

            error_requests = data.get("elb_4xx_count", 0) + data.get("elb_5xx_count", 0)

            error_rate = (error_requests / total_requests) * 100
            return f"{error_rate:.1f}%"

        except Exception as exc:  # noqa: BLE001
            logger.error(f"에러율 계산 중 오류 발생: {exc}")
            return "N/A"

    def _get_top_request_urls(
        self, data: Dict[str, Any], limit: int = 5
    ) -> List[Tuple[str, int]]:
        """상위 요청 URL 목록을 반환합니다."""
        try:
            request_url_counts = data.get("request_url_counts", {})
            if not request_url_counts:
                return []

            sorted_urls = sorted(
                request_url_counts.items(), key=lambda x: x[1], reverse=True
            )
            return sorted_urls[:limit]
        except Exception as exc:  # noqa: BLE001
            logger.error(f"상위 요청 URL 계산 중 오류 발생: {exc}")
            return []

    def _get_top_user_agents(
        self, data: Dict[str, Any], limit: int = 5
    ) -> List[Tuple[str, int]]:
        """상위 User Agent 목록을 반환합니다."""
        try:
            user_agent_counts = data.get("user_agent_counts", {})
            if not user_agent_counts:
                return []

            sorted_agents = sorted(
                user_agent_counts.items(), key=lambda x: x[1], reverse=True
            )
            return sorted_agents[:limit]
        except Exception as exc:  # noqa: BLE001
            logger.error(f"상위 User Agent 계산 중 오류 발생: {exc}")
            return []

    def _get_top_countries(
        self, data: Dict[str, Any], limit: int = 5
    ) -> List[Tuple[str, int]]:
        """상위 국가 목록을 반환합니다."""
        try:
            country_statistics = data.get("country_statistics", {})
            if not country_statistics:
                return []

            sorted_countries = sorted(
                country_statistics.items(), key=lambda x: x[1], reverse=True
            )
            return sorted_countries[:limit]
        except Exception as exc:  # noqa: BLE001
            logger.error(f"상위 국가 계산 중 오류 발생: {exc}")
            return []

    def _calculate_response_time_stats(self, data: Dict[str, Any]) -> Dict[str, str]:
        """응답 시간 통계를 계산합니다."""
        try:
            long_response_times = data.get("long_response_times", [])
            if not long_response_times or not isinstance(long_response_times, list):
                return {"max": "N/A", "min": "N/A", "median": "N/A"}

            response_times = []
            for log_entry in long_response_times:
                if isinstance(log_entry, dict) and "response_time" in log_entry:
                    response_time = log_entry.get("response_time", 0)
                    if isinstance(response_time, (int, float)) and response_time > 0:
                        response_times.append(float(response_time))

            if not response_times:
                return {"max": "N/A", "min": "N/A", "median": "N/A"}

            response_times.sort()
            max_time = response_times[-1]
            min_time = response_times[0]

            # 중간값 계산
            n = len(response_times)
            if n % 2 == 0:
                median_time = (response_times[n // 2 - 1] + response_times[n // 2]) / 2
            else:
                median_time = response_times[n // 2]

            return {
                "max": f"{max_time:.3f}초",
                "min": f"{min_time:.3f}초",
                "median": f"{median_time:.3f}초",
            }

        except Exception as exc:  # noqa: BLE001
            logger.error(f"응답 시간 통계 계산 중 오류 발생: {exc}")
            return {"max": "N/A", "min": "N/A", "median": "N/A"}

    def _convert_status_code(self, status_code: Any) -> Union[int, str]:
        """상태 코드를 적절한 형식으로 변환합니다."""
        try:
            if status_code is None or status_code == "" or status_code == "-":
                return ""

            # 이미 정수인 경우
            if isinstance(status_code, int):
                return status_code

            # 문자열인 경우 정수로 변환 시도
            if isinstance(status_code, str):
                status_code = status_code.strip()
                if status_code in ["", "-", "N/A"]:
                    return ""
                # 숫자로 변환 가능한지 확인 (정수 및 소수점 포함)
                try:
                    # float로 변환 시도 후 정수로 변환
                    return int(float(status_code))
                except (ValueError, TypeError):
                    # 변환 실패 시 빈 문자열 반환
                    return ""

            # float인 경우 정수로 변환
            if isinstance(status_code, float):
                return int(status_code)

            # 그 외의 경우 빈 문자열로 반환
            return ""

        except (ValueError, TypeError):
            return ""

    def _get_normalized_abuse_ips(
        self, data: Dict[str, Any]
    ) -> Tuple[List[str], Dict[str, Any]]:
        try:
            abuse_ips_list: List[str] = []
            abuse_ip_details: Dict[str, Any] = {}
            if data.get("abuse_ips_list"):
                abuse_ips_list = [
                    str(ip).strip()
                    for ip in data["abuse_ips_list"]
                    if ip
                    and str(ip).strip()
                    and not any(
                        k in str(ip)
                        for k in ["abuse_ips", "abuse_ip_details", "timestamp"]
                    )
                ]
            elif data.get("abuse_ips"):
                abuse_ips = data["abuse_ips"]
                if isinstance(abuse_ips, list) or isinstance(abuse_ips, set):
                    abuse_ips_list = [
                        str(ip).strip()
                        for ip in abuse_ips
                        if ip
                        and str(ip).strip()
                        and not any(
                            k in str(ip)
                            for k in ["abuse_ips", "abuse_ip_details", "timestamp"]
                        )
                    ]
                elif isinstance(abuse_ips, dict):
                    abuse_ips_list = [
                        str(ip).strip()
                        for ip in abuse_ips.keys()
                        if ip
                        and str(ip).strip()
                        and not any(
                            k in str(ip)
                            for k in ["abuse_ips", "abuse_ip_details", "timestamp"]
                        )
                    ]
                    abuse_ip_details = abuse_ips
            if data.get("abuse_ip_details"):
                abuse_ip_details.update(data["abuse_ip_details"])  # type: ignore[arg-type]
            return abuse_ips_list, abuse_ip_details
        except Exception as exc:  # noqa: BLE001
            logger.error(f"Abuse IP 정규화 중 예상치 못한 오류: {exc}")
            return [], {}

    def _get_matching_abuse_ips(self, data: Dict[str, Any]) -> Set[str]:
        try:
            actual_client_ips = set(data.get("client_ip_counts", {}).keys())
            abuse_ips_list, _ = self._get_normalized_abuse_ips(data)
            return actual_client_ips.intersection(set(abuse_ips_list))
        except Exception as exc:  # noqa: BLE001
            logger.error(f"일치하는 Abuse IP 계산 중 오류: {exc}")
            return set()

    def _get_abuse_ip_count(self, data: Dict[str, Any]) -> int:
        try:
            return len(self._get_matching_abuse_ips(data))
        except Exception as exc:  # noqa: BLE001
            logger.error(f"Abuse IP 수 계산 중 오류: {exc}")
            return 0

    def _add_country_statistics_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """국가별 통계 시트를 생성합니다."""
        try:
            country_stats = data.get("country_statistics", {})
            if not country_stats:
                return

            client_ip_counts: Dict[str, int] = data.get("client_ip_counts", {}) or {}
            ip_country_mapping: Dict[str, Optional[str]] = (
                data.get("ip_country_mapping", {}) or {}
            )

            # 국가별 총 요청 수 계산 (IP별 요청 수를 국가 코드로 합산)
            country_request_counts: Dict[str, int] = {}
            for ip, req_count in client_ip_counts.items():
                country_code = ip_country_mapping.get(ip) or "ZZ"
                if country_code in [
                    "UNKNOWN",
                    "PRIVATE",
                    "LOOPBACK",
                    "LINK_LOCAL",
                    "MULTICAST",
                ]:
                    country_code = "ZZ"
                country_request_counts[country_code] = country_request_counts.get(
                    country_code, 0
                ) + int(req_count or 0)

            total_requests_all_countries = sum(country_request_counts.values()) or 0

            # 국가별 통계를 DataFrame으로 변환
            country_data = []
            for country_code, ip_count in country_stats.items():
                total_requests = country_request_counts.get(country_code, 0)
                percentage = (
                    round((total_requests / total_requests_all_countries) * 100, 2)
                    if total_requests_all_countries > 0
                    else 0.0
                )
                country_data.append(
                    {
                        "Count": total_requests,
                        "Country": country_code,
                        "IP Count": ip_count,
                        "Percentage": percentage,
                    }
                )

            # Count 기준으로 정렬, 상위 50개 국가만 표시
            country_data_sorted = sorted(
                country_data, key=lambda x: x["Count"], reverse=True
            )[:50]

            self._add_data_sheet(
                wb,
                country_data_sorted,
                ["Count", "Country", "IP Count", "Percentage"],
                "국가별 통계",
                None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(f"국가별 통계 시트 생성 중 오류 발생: {exc}")

    def _add_url_request_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        if data.get("request_url_details"):
            url_details: Dict[str, Any] = data["request_url_details"]
            url_data: List[Dict[str, Any]] = []
            for url, details in url_details.items():
                methods = details.get("methods", {})
                top_method = (
                    max(methods.items(), key=lambda x: x[1])[0] if methods else "-"
                )
                if isinstance(top_method, str):
                    top_method = top_method.replace("-", "") or "-"
                # Unique IPs: prefer aggregated value, fallback to set/list length
                if "unique_ips" in details and isinstance(
                    details.get("unique_ips"), (int, float)
                ):
                    unique_ips = int(details.get("unique_ips") or 0)
                else:
                    client_ips_val = details.get("client_ips", set())
                    if isinstance(client_ips_val, (set, list, tuple)):
                        unique_ips = len(client_ips_val)
                    else:
                        unique_ips = 0
                # Avg response time: prefer aggregated numeric value, robust fallbacks
                avg_response_time: Any
                if "avg_response_time" in details:
                    avg_rt_val = details.get("avg_response_time")
                    # 문자열로 들어온 경우도 처리
                    if isinstance(avg_rt_val, str):
                        try:
                            avg_rt_val = float(avg_rt_val.strip())
                        except Exception:
                            avg_rt_val = None
                    if isinstance(avg_rt_val, (int, float)):
                        avg_response_time = (
                            0.0
                            if avg_rt_val is None or avg_rt_val < 0
                            else round(float(avg_rt_val), 3)
                        )
                    else:
                        # 비정상 타입이면 리스트에서 계산 시도
                        response_times = details.get("response_times", [])
                        parsed_times: List[float] = []
                        for rt in response_times or []:
                            if isinstance(rt, (int, float)):
                                if rt is not None and rt >= 0:
                                    parsed_times.append(float(rt))
                            elif isinstance(rt, str):
                                try:
                                    val = float(rt.strip())
                                    if val >= 0:
                                        parsed_times.append(val)
                                except Exception:
                                    continue
                        avg_response_time = (
                            round(sum(parsed_times) / len(parsed_times), 3)
                            if parsed_times
                            else 0.0
                        )
                else:
                    # 키가 없으면 0으로 표시 (N/A 방지)
                    response_times = details.get("response_times", [])
                    parsed_times: List[float] = []
                    for rt in response_times or []:
                        if isinstance(rt, (int, float)):
                            if rt is not None and rt >= 0:
                                parsed_times.append(float(rt))
                        elif isinstance(rt, str):
                            try:
                                val = float(rt.strip())
                                if val >= 0:
                                    parsed_times.append(val)
                            except Exception:
                                continue
                    avg_response_time = (
                        round(sum(parsed_times) / len(parsed_times), 3)
                        if parsed_times
                        else 0.0
                    )
                status_codes = details.get("status_codes", {})
                top_status = (
                    max(status_codes.items(), key=lambda x: x[1])[0]
                    if status_codes
                    else "-"
                )
                # 상태 코드를 숫자로 변환하여 엑셀에서 '텍스트 형식 숫자' 경고 제거
                top_status = self._convert_status_code(top_status)

                # 에러율 계산 (4xx, 5xx 비율)
                total_requests = sum(status_codes.values()) if status_codes else 0
                error_requests = (
                    sum(
                        count
                        for status, count in status_codes.items()
                        if str(status).startswith(("4", "5"))
                    )
                    if status_codes
                    else 0
                )
                error_rate = (
                    round((error_requests / total_requests * 100), 2)
                    if total_requests > 0
                    else 0
                )

                url_data.append(
                    {
                        "Count": details.get("count", 0),
                        "Method": top_method,
                        "Request": url,
                        "Unique IPs": unique_ips,
                        "Avg Response Time": avg_response_time,
                        "Top Status": top_status,
                        "Error Rate (%)": error_rate,
                    }
                )
            # Count 기준으로 정렬, 상위 100개
            url_data_sorted = sorted(url_data, key=lambda x: x["Count"], reverse=True)[
                :100
            ]
            try:
                abuse_ips_set = self._get_matching_abuse_ips(data)
            except Exception as exc:  # noqa: BLE001
                logger.error(f"요청 URL Top 100 시트에서 Abuse IP 처리 중 오류: {exc}")
                abuse_ips_set = set()
            headers = [
                "Count",
                "Method",
                "Request",
                "Unique IPs",
                "Avg Response Time",
                "Top Status",
                "Error Rate (%)",
            ]
            self._add_data_sheet(
                wb, url_data_sorted, headers, "요청 URL Top 100", abuse_ips_set
            )
        elif data.get("request_url_counts"):
            url_counts = data["request_url_counts"]
            # dict를 list[dict]로 변환, Count 기준 정렬, 상위 100개
            url_data_simple = [
                {"Count": count, "Request": url} for url, count in url_counts.items()
            ]
            url_data_sorted = sorted(
                url_data_simple, key=lambda x: x["Count"], reverse=True
            )[:100]
            try:
                abuse_ips_set = self._get_matching_abuse_ips(data)
            except Exception as exc:  # noqa: BLE001
                logger.error(f"요청 URL Top 100 시트에서 Abuse IP 처리 중 오류: {exc}")
                abuse_ips_set = set()
            self._add_data_sheet(
                wb,
                url_data_sorted,
                ["Count", "Request"],
                "요청 URL Top 100",
                abuse_ips_set,
            )

    def _add_data_sheet(
        self,
        wb: Workbook,
        data: List[Dict[str, Any]],
        headers: List[str],
        sheet_name: str,
        abuse_ip_set: Optional[Set[str]] = None,
    ) -> None:
        """list[dict] 데이터를 Excel 시트로 추가합니다.

        Args:
            wb: Workbook 객체
            data: 데이터 리스트 (각 dict는 한 행)
            headers: 컬럼 헤더 목록 (순서대로 출력)
            sheet_name: 시트 이름
            abuse_ip_set: Abuse IP 집합 (하이라이트용)
        """
        try:
            if not data:
                console.print(
                    f"[yellow]⚠ 데이터가 비어 있습니다. {sheet_name} 시트를 추가하지 않습니다.[/yellow]"
                )
                return

            max_rows_per_sheet = 1_000_000
            total_rows = len(data)
            num_sheets = (total_rows // max_rows_per_sheet) + 1

            for sheet_index in range(num_sheets):
                current_sheet_name = (
                    f"{sheet_name}_{sheet_index + 1}" if sheet_index > 0 else sheet_name
                )
                start_idx = sheet_index * max_rows_per_sheet
                end_idx = min((sheet_index + 1) * max_rows_per_sheet, total_rows)
                current_data = data[start_idx:end_idx]
                if not current_data:
                    continue

                ws = wb.create_sheet(current_sheet_name)

                # 헤더 행 설정
                self._write_header_row(ws, headers)

                # 데이터 행 설정 - COLUMN_STYLE_MAP 활용
                for row_idx, row_data in enumerate(current_data, start=2):
                    has_abuse_ip = (
                        abuse_ip_set
                        and "Client" in headers
                        and row_data.get("Client", "") in abuse_ip_set
                    )

                    for col_idx, header in enumerate(headers, start=1):
                        value = row_data.get(header, "")
                        if value is None:
                            value = ""

                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = self.thin_border

                        # COLUMN_STYLE_MAP 기반 스타일 적용
                        self._apply_cell_style_by_header(cell, header, value)

                        # Abuse IP 하이라이트
                        if has_abuse_ip:
                            cell.fill = self.abuse_fill

                        # Boolean 값 변환
                        if isinstance(value, bool):
                            cell.value = "Yes" if value else "No"

                # 행 높이 설정
                ws.row_dimensions[1].height = 40
                for row_idx in range(2, 2 + len(current_data)):
                    ws.row_dimensions[row_idx].height = 20

                # 컬럼 스타일 및 너비 적용
                self._apply_column_styles(ws, headers)
                self._apply_column_widths_by_map(ws, headers)

                # 요약 행 추가 (특정 시트만)
                if sheet_name.startswith(
                    (
                        "IP별 요청",
                        "IP 요청 Top 100",
                        "상위 요청 URL",
                        "요청 URL Top 100",
                        "상위 User Agent",
                    )
                ):
                    self._add_summary_row(ws, current_data, headers, sheet_name)

                # 자동 필터 및 고정
                ws.auto_filter.ref = (
                    f"A1:{get_column_letter(len(headers))}{len(current_data) + 1}"
                )
                ws.freeze_panes = ws.cell(row=2, column=1)
                ws.sheet_view.zoomScale = 85
            if num_sheets > 1:
                console.print(
                    f"[green]✓ {sheet_name} 데이터가 {num_sheets}개의 시트로 분할되었습니다 (총 {total_rows} 행).[/green]"
                )
        except Exception as exc:  # noqa: BLE001
            logger.error(f"데이터 시트 추가 중 오류 발생: {exc}")

    def _apply_cell_style_by_header(self, cell, header: str, value: Any) -> None:
        """COLUMN_STYLE_MAP을 기반으로 셀 스타일 적용"""
        style_type = COLUMN_STYLE_MAP.get(header, "data")

        if style_type == "number":
            cell.alignment = ALIGN_RIGHT
            cell.number_format = "#,##0"
        elif style_type == "decimal":
            cell.alignment = ALIGN_RIGHT
            if isinstance(value, (int, float)):
                cell.number_format = "0.00"
        elif style_type == "decimal3":
            cell.alignment = ALIGN_RIGHT
            if isinstance(value, (int, float)):
                cell.number_format = "0.000"
        elif style_type == "status":
            cell.alignment = ALIGN_RIGHT
            if isinstance(value, (int, float)):
                cell.number_format = "0"
        elif style_type == "center":
            cell.alignment = ALIGN_CENTER
        else:  # "data" - 기본값
            cell.alignment = ALIGN_LEFT

    def _apply_column_widths_by_map(self, ws, headers: List[str]) -> None:
        """COLUMN_WIDTH_MAP을 기반으로 컬럼 너비 적용"""
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            width = COLUMN_WIDTH_MAP.get(header, 15)  # 기본값 15
            ws.column_dimensions[col_letter].width = width

    def _add_summary_row(
        self, ws, data: List[Dict[str, Any]], headers: List[str], sheet_name: str
    ) -> None:
        """list[dict] 데이터의 합계 행을 추가합니다."""
        try:
            total_row = len(data) + 2
            total_count = sum(row.get("Count", 0) for row in data)

            # 헬퍼: 합계 셀 작성
            def write_sum_cell(col: int, value, is_number: bool = False) -> None:
                cell = ws.cell(row=total_row, column=col, value=value)
                cell.font = FONT_BOLD
                cell.border = self.thin_border
                if is_number:
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = "#,##0"

            # 헬퍼: 빈 셀 채우기
            def fill_empty_cells(start_col: int, end_col: int) -> None:
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=total_row, column=col_idx, value="")
                    cell.border = self.thin_border

            if sheet_name.startswith(("IP별 요청", "IP 요청 Top 100")):
                write_sum_cell(1, "합계")
                write_sum_cell(2, total_count, is_number=True)
                fill_empty_cells(3, 4)

            elif sheet_name.startswith(("상위 요청 URL", "요청 URL Top 100")):
                if len(headers) >= 7:
                    write_sum_cell(1, f"합계: {total_count:,}")
                    ws.cell(row=total_row, column=1).alignment = ALIGN_LEFT
                    fill_empty_cells(2, len(headers))
                else:
                    write_sum_cell(1, "합계")
                    write_sum_cell(2, total_count, is_number=True)

            elif sheet_name.startswith("상위 User Agent"):
                write_sum_cell(1, "합계")
                write_sum_cell(2, total_count, is_number=True)

            elif sheet_name.startswith("국가별 통계"):
                total_ip_count = sum(row.get("IP Count", 0) for row in data)
                write_sum_cell(1, "합계")
                write_sum_cell(3, total_ip_count, is_number=True)
                # Percentage 100%
                cell = ws.cell(row=total_row, column=4, value="100.00")
                cell.font = FONT_BOLD
                cell.border = self.thin_border
                cell.alignment = ALIGN_RIGHT
                cell.number_format = "0.00"

        except Exception as exc:  # noqa: BLE001
            logger.error(f"합계 행 추가 중 오류 발생: {exc}")

    def _create_abuse_ip_worksheet(
        self,
        workbook: Workbook,
        abuse_ips_list: List[str],
        abuse_ip_details: Dict[str, Any],
        data: Optional[Dict[str, Any]] = None,
    ) -> None:
        try:
            ws = workbook.create_sheet("악성 IP 목록")
            actual_client_ips: Set[str] = (
                set(data.get("client_ip_counts", {}).keys()) if data else set()
            )
            abuse_ips_set = set(abuse_ips_list) if abuse_ips_list else set()
            matching_abuse_ips = actual_client_ips.intersection(abuse_ips_set)

            headers = ["Count", "IP", "Country", "ASN", "ISP"]
            self._write_header_row(ws, headers)

            if not matching_abuse_ips:
                self._write_empty_message(
                    ws, "ALB 로그에서 탐지된 악성 IP가 없습니다.", 2, len(headers)
                )
                console.print(
                    "[green]✓ 악성 IP 목록 시트가 생성되었지만 ALB 로그에서 악성 IP가 발견되지 않았습니다.[/green]"
                )
            else:
                sorted_matching_ips = sorted(
                    matching_abuse_ips,
                    key=lambda ip: (
                        data.get("client_ip_counts", {}).get(ip, 0) if data else 0
                    ),
                    reverse=True,
                )
                for row_idx, ip in enumerate(sorted_matching_ips, start=2):
                    details = (
                        abuse_ip_details.get(ip, {})
                        if isinstance(abuse_ip_details, dict)
                        else {}
                    )
                    request_count = (
                        data.get("client_ip_counts", {}).get(ip, 0) if data else 0
                    )
                    # Count (A)
                    count_cell = ws.cell(row=row_idx, column=1, value=request_count)
                    count_cell.border = self.thin_border
                    count_cell.alignment = ALIGN_RIGHT
                    count_cell.number_format = "#,##0"
                    # IP (B)
                    ws.cell(row=row_idx, column=2, value=ip).border = self.thin_border
                    ws.cell(row=row_idx, column=2).alignment = ALIGN_CENTER
                    # Country (C) - IPDeny 데이터 우선 사용
                    country_mapping = data.get("ip_country_mapping", {}) if data else {}
                    country_value = country_mapping.get(ip, "N/A")
                    ws.cell(
                        row=row_idx, column=3, value=country_value
                    ).border = self.thin_border
                    ws.cell(row=row_idx, column=3).alignment = ALIGN_CENTER
                    # ASN (D)
                    asn_value = details.get("asn", "N/A")
                    ws.cell(
                        row=row_idx, column=4, value=asn_value
                    ).border = self.thin_border
                    ws.cell(row=row_idx, column=4).alignment = ALIGN_CENTER
                    # ISP (E)
                    isp_value = details.get("isp", "N/A")
                    ws.cell(
                        row=row_idx, column=5, value=isp_value
                    ).border = self.thin_border
                    ws.cell(row=row_idx, column=5).alignment = ALIGN_CENTER

            ws.row_dimensions[1].height = 40
            if matching_abuse_ips:
                for row_idx in range(2, 2 + len(matching_abuse_ips)):
                    ws.row_dimensions[row_idx].height = 20
                ws.auto_filter.ref = (
                    f"A1:{get_column_letter(len(headers))}{len(matching_abuse_ips) + 1}"
                )
            self._apply_column_widths_by_map(ws, headers)
            ws.freeze_panes = ws.cell(row=2, column=1)
            ws.sheet_view.zoomScale = 85
        except Exception as exc:  # noqa: BLE001
            logger.error(f"Abuse IP 시트 생성 중 오류 발생: {exc}")

    def _create_abuse_requests_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """악성 IP들의 개별 요청 내역을 시간 순으로 분석하는 시트를 생성합니다."""
        try:
            ws = wb.create_sheet("악성 IP 요청 모음")

            # Abuse IP 목록 가져오기
            abuse_ips_set = self._get_matching_abuse_ips(data)

            # 헤더 생성
            headers = [
                "Timestamp",
                "Client",
                "Country",
                "Target",
                "Target group name",
                "Method",
                "Request",
                "User Agent",
                "ELB Status Code",
                "Backend Status Code",
            ]
            self._write_header_row(ws, headers)

            if not abuse_ips_set:
                # 악성 IP가 없는 경우 빈 시트 생성
                self._write_empty_message(
                    ws, "악성 IP의 요청 데이터를 찾을 수 없습니다.", 2, len(headers)
                )
                return

            # 모든 상태 코드별 로그 데이터 수집
            all_logs = []
            status_code_types = [
                "ELB 2xx Count",
                "ELB 3xx Count",
                "ELB 4xx Count",
                "ELB 5xx Count",
                "Backend 4xx Count",
                "Backend 5xx Count",
            ]

            for status_type in status_code_types:
                if status_type in data and isinstance(data[status_type], dict):
                    full_logs = data[status_type].get("full_logs", [])
                    if isinstance(full_logs, list):
                        all_logs.extend(full_logs)

            # 악성 IP의 요청만 필터링 (집계하지 않음)
            abuse_requests = []
            for log_entry in all_logs:
                client_ip = log_entry.get("client_ip", "N/A")
                if client_ip in abuse_ips_set:
                    abuse_requests.append(log_entry)

            if not abuse_requests:
                ws.merge_cells(
                    start_row=2, start_column=1, end_row=2, end_column=len(headers)
                )
                cell = ws.cell(row=2, column=1, value="악성 IP의 요청 데이터를 찾을 수 없습니다.")
                cell.alignment = ALIGN_CENTER
                cell.font = FONT_ITALIC_GRAY
                cell.border = self.thin_border
            else:
                # 시간 순으로 정렬 (timestamp가 문자열일 수도 있으니 안전하게 처리)
                def safe_timestamp_key(log_entry):
                    timestamp = log_entry.get("timestamp")
                    if isinstance(timestamp, datetime):
                        return timestamp
                    elif isinstance(timestamp, str):
                        try:
                            # 문자열을 datetime으로 파싱 시도
                            return datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                        except ValueError:
                            # 파싱 실패시 원본 문자열 반환 (문자열 정렬)
                            return timestamp
                    else:
                        return datetime.min

                sorted_requests = sorted(abuse_requests, key=safe_timestamp_key)

                for row_idx, log_entry in enumerate(sorted_requests, start=2):
                    client_ip = log_entry.get("client_ip", "N/A")
                    target_group_name = log_entry.get("target_group_name", "")
                    target = log_entry.get("target", "")
                    target_field = "" if not target or target == "-" else target
                    target_group_field = target_group_name or ""

                    # 타임스탬프 포맷팅
                    timestamp = log_entry.get("timestamp")
                    if isinstance(timestamp, datetime):
                        timestamp_str = timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        timestamp_str = str(timestamp) if timestamp else "N/A"

                    http_method = log_entry.get("http_method", "").replace("-", "")
                    request = log_entry.get("request", "N/A")
                    user_agent = log_entry.get("user_agent", "N/A")
                    elb_status = self._convert_status_code(
                        log_entry.get("elb_status_code", "N/A")
                    )
                    backend_status = self._convert_status_code(
                        log_entry.get("target_status_code", "N/A")
                    )

                    # 국가 정보 추가
                    country_mapping = data.get("ip_country_mapping", {})
                    country_code = country_mapping.get(client_ip, "N/A")

                    fields = [
                        timestamp_str,
                        client_ip,
                        country_code,
                        target_field,
                        target_group_field,
                        http_method,
                        request,
                        user_agent,
                        elb_status,
                        backend_status,
                    ]

                    for col_idx, field_value in enumerate(fields, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=field_value)
                        cell.border = self.thin_border

                        # 정렬 규칙: Timestamp, Client, Country, Target, Target group name, Method 가운데
                        # Request, User Agent 왼쪽, 상태코드 오른쪽
                        if col_idx in [1, 2, 3, 4, 5, 6]:
                            cell.alignment = ALIGN_CENTER
                        elif col_idx in [7, 8]:  # Request, User Agent
                            cell.alignment = ALIGN_LEFT
                        elif col_idx in [9, 10]:  # Status codes
                            cell.alignment = ALIGN_RIGHT
                            if isinstance(field_value, int):
                                cell.number_format = "0"

                if sorted_requests:
                    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_requests) + 1}"

            # Apply column-specific styles
            self._apply_column_styles(ws, headers)
            self._apply_column_widths_by_map(ws, headers)

            ws.row_dimensions[1].height = 40
            ws.freeze_panes = ws.cell(row=2, column=1)
            ws.sheet_view.zoomScale = 85

        except Exception as exc:  # noqa: BLE001
            logger.error(f"악성 IP 요청 분석 시트 생성 중 오류 발생: {exc}")

    def _create_bytes_analysis_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        sheet_name = "데이터 전송량 Top 100"
        received_bytes = data.get("received_bytes", {})
        sent_bytes = data.get("sent_bytes", {})
        ws = wb.create_sheet(title=sheet_name)
        headers = [
            "Request",
            "수신 데이터 (Bytes)",
            "송신 데이터 (Bytes)",
            "총 데이터 (Bytes)",
            "총 데이터 (변환)",
        ]
        self._write_header_row(ws, headers)
        all_urls = sorted(set(list(received_bytes.keys()) + list(sent_bytes.keys())))
        total_data: Dict[str, int] = {}
        for url in all_urls:
            total_data[url] = int(received_bytes.get(url, 0)) + int(
                sent_bytes.get(url, 0)
            )
        sorted_urls = sorted(all_urls, key=lambda url: total_data[url], reverse=True)[
            :100
        ]
        for row_idx, url in enumerate(sorted_urls, 2):
            received = int(received_bytes.get(url, 0))
            sent = int(sent_bytes.get(url, 0))
            total_bytes = received + sent
            formatted_bytes = self._format_bytes(total_bytes)
            fields: List[Any] = [url, received, sent, total_bytes, formatted_bytes]
            for col_idx, field_value in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=field_value)
                if col_idx > 1:
                    cell.alignment = ALIGN_RIGHT
                    if col_idx in [2, 3, 4]:
                        cell.number_format = "#,##0"
                else:
                    cell.alignment = self.cell_alignment
                cell.border = self.thin_border
        total_row = len(sorted_urls) + 2
        total_received = sum(int(v) for v in received_bytes.values())
        total_sent = sum(int(v) for v in sent_bytes.values())
        total_bytes_all = total_received + total_sent
        fields = [
            "총계",
            total_received,
            total_sent,
            total_bytes_all,
            self._format_bytes(total_bytes_all),
        ]
        for col_idx, field_value in enumerate(fields, 1):
            cell = ws.cell(row=total_row, column=col_idx, value=field_value)
            cell.font = FONT_BOLD
            if col_idx > 1:
                cell.alignment = ALIGN_RIGHT
                if col_idx in [2, 3, 4]:
                    cell.number_format = "#,##0"
            else:
                cell.alignment = ALIGN_LEFT
            cell.border = self.thin_border

        # Apply column-specific styles
        self._apply_column_styles(ws, headers)
        self._apply_column_widths_by_map(ws, headers)

        ws.row_dimensions[1].height = 40
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.sheet_view.zoomScale = 85

    def _create_response_time_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        try:
            if not (data.get("response_time") or data.get("long_response_times")):
                return
            sheet_name = "응답 시간 Top 100"
            ws = wb.create_sheet(sheet_name)
            header_cols = [
                "Response time",
                "Timestamp",
                "Client",
                "Country",
                "Target",
                "Target group name",
                "Method",
                "Request",
                "User Agent",
                "ELB Status Code",
                "Backend Status Code",
            ]
            self._write_header_row(ws, header_cols)
            try:
                abuse_ips_list, _ = self._get_normalized_abuse_ips(data)
                abuse_ips_set = set(abuse_ips_list)
            except Exception as exc:  # noqa: BLE001
                logger.error(f"{sheet_name}에서 Abuse IP 처리 중 오류: {exc}")
                abuse_ips_set = set()
            long_response_logs = (
                data.get("long_response_times", [])
                if isinstance(data.get("long_response_times"), list)
                else []
            )
            # NULL(None) 값은 -1로 처리하여 맨 뒤로 정렬
            sorted_logs = sorted(
                long_response_logs,
                key=lambda x: (
                    x.get("response_time") if x.get("response_time") is not None else -1
                ),
                reverse=True,
            )
            # NULL이 아닌 유효한 응답시간만 필터링
            filtered_logs = [
                log for log in sorted_logs if log.get("response_time") is not None
            ][:100]
            for i, log in enumerate(filtered_logs, start=2):
                client_ip = log.get("client_ip", "N/A")
                is_abuse_ip = client_ip in abuse_ips_set
                target_group_name = log.get("target_group_name", "")
                target = log.get("target", "")
                target_field = "" if not target or target == "-" else target
                target_group_field = target_group_name or ""
                # 국가 정보 추가
                country_mapping = data.get("ip_country_mapping", {})
                country_code = country_mapping.get(client_ip, "N/A")

                fields = [
                    log.get("response_time", 0),
                    (
                        log.get("timestamp", "").strftime("%Y-%m-%d %H:%M:%S")
                        if isinstance(log.get("timestamp"), datetime)
                        else str(log.get("timestamp", "N/A"))
                    ),
                    client_ip,
                    country_code,
                    target_field,
                    target_group_field,
                    log.get("http_method", "").replace("-", ""),
                    log.get("request", "N/A"),
                    log.get("user_agent", "N/A"),
                    self._convert_status_code(log.get("elb_status_code", "N/A")),
                    self._convert_status_code(log.get("target_status_code", "N/A")),
                ]
                for col_idx, field_value in enumerate(fields, start=1):
                    cell = ws.cell(row=i, column=col_idx, value=field_value)
                    cell.border = self.thin_border

                    # 정렬 규칙 적용
                    if col_idx == 1:  # Response time
                        cell.alignment = ALIGN_RIGHT
                    elif col_idx in [
                        2,
                        3,
                        4,
                        5,
                        6,
                        7,
                    ]:  # Timestamp, Client, Country, Target, Target group name, Method
                        cell.alignment = ALIGN_CENTER
                    elif col_idx in [8, 9]:  # Request, User Agent
                        cell.alignment = ALIGN_LEFT
                    elif col_idx in [10, 11]:  # Status codes
                        cell.alignment = ALIGN_RIGHT
                        if isinstance(field_value, int):
                            cell.number_format = "0"

                    if is_abuse_ip:
                        cell.fill = self.abuse_fill

            # Apply column-specific styles
            self._apply_column_styles(ws, header_cols)
            self._apply_column_widths_by_map(ws, header_cols)

            ws.row_dimensions[1].height = 40
            if filtered_logs:
                ws.auto_filter.ref = (
                    f"A1:{get_column_letter(len(header_cols))}{len(filtered_logs) + 1}"
                )
            ws.freeze_panes = ws.cell(row=2, column=1)
            ws.sheet_view.zoomScale = 85
        except Exception as exc:  # noqa: BLE001
            logger.error(f"응답 시간 시트 생성 중 에러 발생: {exc}")

    def _add_status_code_sheets(
        self, wb: Workbook, data: Dict[str, Any], abuse_ip_set: Optional[Set[str]]
    ) -> None:
        try:
            count_sheets: Dict[str, Dict[str, Any]] = {
                "ELB 2xx Top 100": {
                    "count": data.get("elb_2xx_count", 0),
                    "full_logs": data.get("ELB 2xx Count", {}).get("full_logs", []),
                },
                "ELB 3xx Top 100": {
                    "count": data.get("elb_3xx_count", 0),
                    "full_logs": data.get("ELB 3xx Count", {}).get("full_logs", []),
                },
                "ELB 4xx Count": {
                    "count": data.get("elb_4xx_count", 0),
                    "full_logs": data.get("ELB 4xx Count", {}).get("full_logs", []),
                },
                "ELB 5xx Count": {
                    "count": data.get("elb_5xx_count", 0),
                    "full_logs": data.get("ELB 5xx Count", {}).get("full_logs", []),
                },
                "Backend 4xx Count": {
                    "count": data.get("backend_4xx_count", 0),
                    "full_logs": data.get("Backend 4xx Count", {}).get("full_logs", []),
                },
                "Backend 5xx Count": {
                    "count": data.get("backend_5xx_count", 0),
                    "full_logs": data.get("Backend 5xx Count", {}).get("full_logs", []),
                },
            }
            # 심각도 우선 순서로 생성: 5xx → 4xx → 3xx → 2xx
            ordered_count_sheet_names: List[str] = [
                "ELB 5xx Count",
                "Backend 5xx Count",
                "ELB 4xx Count",
                "Backend 4xx Count",
                "ELB 3xx Top 100",
                "ELB 2xx Top 100",
            ]

            timestamp_sheets: Dict[str, Dict[str, Any]] = {
                "ELB 4xx Timestamp": {
                    "count": data.get("elb_4xx_count", 0),
                    "full_logs": data.get("ELB 4xx Count", {}).get("full_logs", []),
                },
                "ELB 5xx Timestamp": {
                    "count": data.get("elb_5xx_count", 0),
                    "full_logs": data.get("ELB 5xx Count", {}).get("full_logs", []),
                },
                "Backend 4xx Timestamp": {
                    "count": data.get("backend_4xx_count", 0),
                    "full_logs": data.get("Backend 4xx Count", {}).get("full_logs", []),
                },
                "Backend 5xx Timestamp": {
                    "count": data.get("backend_5xx_count", 0),
                    "full_logs": data.get("Backend 5xx Count", {}).get("full_logs", []),
                },
            }
            ordered_timestamp_sheet_names: List[str] = [
                "ELB 5xx Timestamp",
                "Backend 5xx Timestamp",
                "ELB 4xx Timestamp",
                "Backend 4xx Timestamp",
            ]

            for sheet_name in ordered_count_sheet_names:
                sheet_data = count_sheets.get(sheet_name, {})
                self._create_status_code_count_sheet(
                    wb, sheet_name, sheet_data, abuse_ip_set, data
                )
            for sheet_name in ordered_timestamp_sheet_names:
                sheet_data = timestamp_sheets.get(sheet_name, {})
                self._create_status_code_timestamp_sheet(
                    wb, sheet_name, sheet_data, abuse_ip_set, data
                )
        except Exception as exc:  # noqa: BLE001
            logger.error(f"상태 코드 시트 생성 중 오류 발생: {exc}")

    def _create_status_code_count_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        status_data: Dict[str, Any],
        abuse_ip_set: Optional[Set[str]],
        data: Optional[Dict[str, Any]] = None,
    ) -> None:
        try:
            ws = wb.create_sheet(sheet_name)
            if sheet_name == "ELB 3xx Top 100":
                header_cols = [
                    "Count",
                    "Client",
                    "Country",
                    "Target",
                    "Target group name",
                    "Method",
                    "Request",
                    "Redirect URL",
                    "ELB Status Code",
                    "Backend Status Code",
                ]
            else:
                header_cols = [
                    "Count",
                    "Client",
                    "Country",
                    "Target",
                    "Target group name",
                    "Method",
                    "Request",
                    "ELB Status Code",
                    "Backend Status Code",
                ]
            self._write_header_row(ws, header_cols)
            # 시트 표시 조건: 카운트와 로그가 모두 0/없을 때만 비표시
            status_full_logs = status_data.get("full_logs") or []
            status_count_val = int(status_data.get("count", 0) or 0)
            if status_count_val == 0 and not status_full_logs:
                self._write_empty_message(
                    ws, "해당 상태 코드를 가진 요청이 없습니다.", 2, len(header_cols)
                )
                self._finalize_sheet(ws, header_cols, 0)
            else:
                full_logs = status_full_logs
                ip_request_counts: Dict[Tuple[Any, ...], int] = {}
                for log_entry in full_logs:
                    client_ip = log_entry.get("client_ip", "N/A")
                    request = log_entry.get("request", "N/A")
                    http_method = log_entry.get("http_method", "").replace("-", "")
                    elb_status = self._convert_status_code(
                        log_entry.get("elb_status_code", "N/A")
                    )
                    target_status = self._convert_status_code(
                        log_entry.get("target_status_code", "N/A")
                    )
                    redirect_url = log_entry.get("redirect_url", "").replace("-", "")
                    target_group_name = log_entry.get("target_group_name", "")
                    target = log_entry.get("target", "")
                    target_field = "" if not target or target == "-" else target
                    target_group_field = target_group_name or ""
                    if sheet_name == "ELB 3xx Top 100":
                        key: Tuple[Any, ...] = (
                            client_ip,
                            target_field,
                            target_group_field,
                            http_method,
                            request,
                            redirect_url,
                            elb_status,
                            target_status,
                        )
                    else:
                        key = (
                            client_ip,
                            target_field,
                            target_group_field,
                            http_method,
                            request,
                            elb_status,
                            target_status,
                        )
                    ip_request_counts[key] = ip_request_counts.get(key, 0) + 1
                sorted_items = sorted(
                    ip_request_counts.items(), key=lambda x: x[1], reverse=True
                )
                if sheet_name in ("ELB 2xx Top 100", "ELB 3xx Top 100"):
                    sorted_items = sorted_items[:100]
                for i, (
                    (
                        client_ip,
                        target_field,
                        target_group_field,
                        http_method,
                        request,
                        *other_fields,
                    ),
                    count,
                ) in enumerate(sorted_items, start=2):
                    is_abuse_ip = bool(abuse_ip_set and client_ip in abuse_ip_set)

                    # 국가 정보 추가
                    country_mapping = data.get("ip_country_mapping", {})
                    country_code = country_mapping.get(client_ip, "N/A")

                    fields: List[Any] = [
                        count,
                        client_ip,
                        country_code,
                        target_field,
                        target_group_field,
                        http_method,
                        request,
                    ]
                    if sheet_name == "ELB 3xx Top 100":
                        redirect_url = other_fields[0]
                        elb_status = other_fields[1]
                        target_status = other_fields[2]
                        fields.append(redirect_url)
                    else:
                        elb_status = other_fields[0]
                        target_status = other_fields[1]
                    fields.extend([elb_status, target_status])
                    for col_idx, field_value in enumerate(fields, start=1):
                        header = header_cols[col_idx - 1]
                        cell = ws.cell(row=i, column=col_idx, value=field_value)
                        cell.border = self.thin_border
                        self._apply_cell_style_by_header(cell, header, field_value)

                        if is_abuse_ip:
                            cell.fill = self.abuse_fill
                self._finalize_sheet(
                    ws, header_cols, len(sorted_items) if sorted_items else 0
                )
        except Exception as exc:  # noqa: BLE001
            logger.error(f"상태 코드 카운트 시트 생성 중 오류 발생: {exc}")

    def _create_status_code_timestamp_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        status_data: Dict[str, Any],
        abuse_ip_set: Optional[Set[str]],
        data: Optional[Dict[str, Any]] = None,
    ) -> None:
        try:
            ws = wb.create_sheet(sheet_name)
            include_user_agent = sheet_name in {
                "ELB 4xx Timestamp",
                "ELB 5xx Timestamp",
                "Backend 4xx Timestamp",
                "Backend 5xx Timestamp",
            }
            if include_user_agent:
                header_cols = [
                    "Timestamp",
                    "Client",
                    "Country",
                    "Target",
                    "Target group name",
                    "Method",
                    "Request",
                    "User Agent",
                    "ELB Status Code",
                    "Backend Status Code",
                    "Error Reason",
                ]
            else:
                header_cols = [
                    "Timestamp",
                    "Client",
                    "Country",
                    "Target",
                    "Target group name",
                    "Method",
                    "Request",
                    "ELB Status Code",
                    "Backend Status Code",
                    "Error Reason",
                ]
            self._write_header_row(ws, header_cols)
            status_full_logs = status_data.get("full_logs") or []
            has_data = bool(status_full_logs)
            status_count_val = int(status_data.get("count", 0) or 0)
            # 시트 표시 조건: 카운트와 로그가 모두 0/없을 때만 비표시
            if not has_data and status_count_val == 0:
                self._write_empty_message(
                    ws, "해당 상태 코드를 가진 요청이 없습니다.", 2, len(header_cols)
                )
                self._finalize_sheet(ws, header_cols, 0)
            else:
                full_logs = status_full_logs

                # 안전한 타임스탬프 정렬 키
                def _safe_ts_key(entry: Dict[str, Any]):
                    ts = entry.get("timestamp")
                    if isinstance(ts, datetime):
                        return ts
                    if isinstance(ts, str):
                        try:
                            return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                        except Exception:
                            return datetime.min
                    return datetime.min

                sorted_logs = sorted(full_logs, key=_safe_ts_key)
                for i, log in enumerate(sorted_logs, start=2):
                    client_ip = log.get("client_ip", "N/A")
                    is_abuse_ip = bool(abuse_ip_set and client_ip in abuse_ip_set)
                    timestamp = log.get("timestamp")
                    if isinstance(timestamp, datetime):
                        timestamp_str = timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        timestamp_str = str(timestamp) if timestamp else "N/A"
                    target_group_name = log.get("target_group_name", "")
                    target = log.get("target", "")

                    target_field = "" if not target or target == "-" else target
                    target_group_field = target_group_name or ""
                    # 국가 정보 추가
                    country_mapping = data.get("ip_country_mapping", {})
                    country_code = country_mapping.get(client_ip, "N/A")

                    if include_user_agent:
                        fields = [
                            timestamp_str,
                            client_ip,
                            country_code,
                            target_field,
                            target_group_field,
                            log.get("http_method", "-"),
                            log.get("request", "N/A"),
                            log.get("user_agent", "N/A"),
                            self._convert_status_code(
                                log.get("elb_status_code", "N/A")
                            ),
                            self._convert_status_code(
                                log.get("target_status_code", "N/A")
                            ),
                            log.get("error_reason", "-"),
                        ]
                    else:
                        fields = [
                            timestamp_str,
                            client_ip,
                            country_code,
                            target_field,
                            target_group_field,
                            log.get("http_method", "-"),
                            log.get("request", "N/A"),
                            self._convert_status_code(
                                log.get("elb_status_code", "N/A")
                            ),
                            self._convert_status_code(
                                log.get("target_status_code", "N/A")
                            ),
                            log.get("error_reason", "-"),
                        ]
                    for col_idx, field_value in enumerate(fields, start=1):
                        header = header_cols[col_idx - 1]
                        cell = ws.cell(row=i, column=col_idx, value=field_value)
                        cell.border = self.thin_border

                        # Error Reason 특수 처리: 값이 없으면 "" 처리
                        if header == "Error Reason" and field_value in (None, "-"):
                            cell.value = ""

                        self._apply_cell_style_by_header(cell, header, field_value)

                        if is_abuse_ip:
                            cell.fill = self.abuse_fill
                self._finalize_sheet(
                    ws, header_cols, len(sorted_logs) if sorted_logs else 0
                )
        except Exception as exc:  # noqa: BLE001
            logger.error(f"상태 코드 타임스탬프 시트 생성 중 오류 발생: {exc}")

    def _add_client_status_statistics_sheet(
        self, wb: Workbook, data: Dict[str, Any]
    ) -> None:
        """Client별 상태코드 통계 시트를 추가합니다."""
        try:
            client_status_stats = data.get("client_status_statistics", {})
            if not client_status_stats:
                logger.debug("Client 상태코드 통계 데이터가 없습니다.")
                return

            ws = wb.create_sheet("Client 상태코드 통계 Top 100")

            # 모든 상태코드 수집 및 정렬
            all_status_codes = set()
            for status_counts in client_status_stats.values():
                all_status_codes.update(status_counts.keys())

            # 상태코드 정렬 (숫자 우선, 문자열은 뒤에)
            sorted_status_codes = sorted(
                all_status_codes,
                key=lambda x: (int(x) if x.isdigit() else float("inf"), x),
            )

            # 헤더 구성
            headers = ["Client", "Country"] + sorted_status_codes + ["Count"]
            self._write_header_row(ws, headers)

            # IP 국가 매핑 정보 가져오기
            ip_country_mapping = data.get("ip_country_mapping", {})

            # 데이터 작성 (요청 수가 많은 순으로 정렬)
            client_total_counts = {}
            for client_ip, status_counts in client_status_stats.items():
                client_total_counts[client_ip] = sum(status_counts.values())

            sorted_clients = sorted(
                client_total_counts.items(), key=lambda x: x[1], reverse=True
            )

            row = 2
            for client_ip, total_count in sorted_clients[:100]:
                status_counts = client_status_stats[client_ip]
                country = ip_country_mapping.get(client_ip, "UNKNOWN")

                # 기본 정보 (정의된 스타일 사용)
                client_cell = ws.cell(row=row, column=1, value=client_ip)
                client_cell.alignment = ALIGN_CENTER

                country_cell = ws.cell(row=row, column=2, value=country)
                country_cell.alignment = ALIGN_CENTER

                # 각 상태코드별 데이터
                for col_idx, status_code in enumerate(sorted_status_codes, 3):
                    status_cell = ws.cell(
                        row=row, column=col_idx, value=status_counts.get(status_code, 0)
                    )
                    status_cell.number_format = "#,##0"

                # Count (총 요청 수)
                count_cell = ws.cell(row=row, column=len(headers), value=total_count)
                count_cell.alignment = ALIGN_RIGHT
                count_cell.number_format = "#,##0"

                row += 1

            # 헤더/데이터 셀 테두리 적용
            try:
                last_col_letter = get_column_letter(len(headers))
                # 헤더 행 테두리
                self._apply_range_style(
                    ws, f"A1:{last_col_letter}1", border=self.header_style["border"]
                )
                # 데이터 영역 테두리
                if ws.max_row >= 2:
                    self._apply_range_style(
                        ws,
                        f"A2:{last_col_letter}{ws.max_row}",
                        border=self.thin_border,
                    )
            except Exception:
                # 테두리 적용 실패는 보고서 생성을 막지 않음
                pass

            # 스타일 적용
            self._apply_column_styles(ws, headers)
            self._enforce_vertical_center_alignment(ws)

            # 열 너비 조정
            ws.column_dimensions["A"].width = 15  # Client IP
            ws.column_dimensions["B"].width = 10  # Country

            # 상태코드 컬럼들 (동적 개수)
            for col_idx in range(3, len(headers)):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 8

            # Count 컬럼
            total_col_letter = get_column_letter(len(headers))
            ws.column_dimensions[total_col_letter].width = 10

            ws.freeze_panes = ws.cell(row=2, column=1)
            ws.sheet_view.zoomScale = 85

        except Exception as exc:  # noqa: BLE001
            logger.error(f"Client 상태코드 통계 시트 생성 중 오류 발생: {exc}")

    def _add_target_backend_status_statistics_sheet(
        self, wb: Workbook, data: Dict[str, Any]
    ) -> None:
        """Target IP별 Target 상태코드(Backend) 통계 시트를 추가합니다."""
        try:
            target_status_stats = data.get("target_status_statistics", {})
            if not target_status_stats:
                logger.info("Target 상태코드 통계 데이터가 없습니다.")
                return

            ws = wb.create_sheet("Target 상태코드 통계")

            # Backend 상태코드만 수집
            backend_codes = set()

            for target, status_counts in target_status_stats.items():
                for key in status_counts.keys():
                    if key.startswith("Backend:"):
                        # Backend:xxx 형태에서 상태코드 부분만 추출
                        backend_code = key.split("Backend:")[1]
                        backend_codes.add(backend_code)

            # Backend 상태코드가 없어도 시트는 생성
            if not backend_codes:
                logger.warning("Backend 상태코드가 없습니다. 빈 시트를 생성합니다.")
                logger.warning("데이터에 ELB 상태코드만 있는 것 같습니다. Target에서 실제 응답이 없는 상태입니다.")
                # 기본 헤더만으로 빈 시트 생성
                sorted_backend_codes = []
            else:
                # Backend 상태코드 정렬
                sorted_backend_codes = sorted(
                    backend_codes,
                    key=lambda x: (int(x) if x.isdigit() else float("inf"), x),
                )

            # 헤더 구성
            headers = ["Target", "Target group name"] + sorted_backend_codes + ["Count"]
            self._write_header_row(ws, headers)

            # Backend 요청이 있는 Target만 필터링
            target_backend_counts = {}
            for target, status_counts in target_status_stats.items():
                backend_total = sum(
                    count
                    for key, count in status_counts.items()
                    if key.startswith("Backend:")
                )
                if backend_total > 0:
                    target_backend_counts[target] = backend_total

            # Backend 요청 수가 많은 순으로 정렬
            sorted_targets = sorted(
                target_backend_counts.items(), key=lambda x: x[1], reverse=True
            )

            row = 2
            for target_display, backend_total_count in sorted_targets:
                status_counts = target_status_stats[target_display]

                # Target / Target group name 분리
                parsed_target = target_display
                parsed_group = ""
                try:
                    if "(" in target_display and target_display.endswith(")"):
                        open_idx = target_display.rfind("(")
                        parsed_group = target_display[:open_idx]
                        parsed_target = target_display[open_idx + 1 : -1]
                except Exception:
                    parsed_target = target_display
                    parsed_group = ""

                # Target / Target group name 필드에 가운데 정렬 적용
                target_cell = ws.cell(row=row, column=1, value=parsed_target)
                target_cell.alignment = ALIGN_CENTER
                group_cell = ws.cell(row=row, column=2, value=parsed_group)
                group_cell.alignment = ALIGN_CENTER

                col_idx = 3

                # Backend 상태코드 데이터
                for backend_code in sorted_backend_codes:
                    count = self._get_target_backend_count(status_counts, backend_code)
                    backend_cell = ws.cell(row=row, column=col_idx, value=count)
                    backend_cell.number_format = "#,##0"
                    col_idx += 1

                # Count (총 Backend 요청 수)
                count_cell = ws.cell(
                    row=row, column=len(headers), value=backend_total_count
                )
                count_cell.alignment = ALIGN_RIGHT
                count_cell.number_format = "#,##0"

                row += 1

            # 헤더/데이터 셀 테두리 적용
            try:
                last_col_letter = get_column_letter(len(headers))
                # 헤더 행 테두리
                self._apply_range_style(
                    ws, f"A1:{last_col_letter}1", border=self.header_style["border"]
                )
                # 데이터 영역 테두리
                if ws.max_row >= 2:
                    self._apply_range_style(
                        ws,
                        f"A2:{last_col_letter}{ws.max_row}",
                        border=self.thin_border,
                    )
            except Exception:
                # 테두리 적용 실패는 보고서 생성을 막지 않음
                pass

            # 스타일 적용
            self._apply_column_styles(ws, headers)
            self._enforce_vertical_center_alignment(ws)

            # 열 너비 조정
            ws.column_dimensions["A"].width = 20  # Target
            ws.column_dimensions["B"].width = 20  # Target group name

            # Backend 상태코드 컬럼들 (동적 개수)
            for col_idx in range(3, len(headers)):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 10

            ws.freeze_panes = ws.cell(row=2, column=1)
            ws.sheet_view.zoomScale = 85

        except Exception as exc:  # noqa: BLE001
            logger.error(f"Target 상태코드 통계 시트 생성 중 오류 발생: {exc}")

    def _get_target_backend_count(
        self, status_counts: Dict[str, int], backend_code: str
    ) -> int:
        """Target 통계에서 특정 Backend 상태코드의 카운트를 가져옵니다."""
        backend_key = f"Backend:{backend_code}"
        return status_counts.get(backend_key, 0)
