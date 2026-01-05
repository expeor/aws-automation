"""
plugins/elb/common.py - ELB 공통 유틸리티

ALB, NLB, CLB, GWLB에서 공유하는 데이터 구조와 헬퍼 함수.

boto3 클라이언트:
    - ALB/NLB/GWLB: elbv2 클라이언트
    - CLB: elb 클라이언트
"""

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from plugins.cost.pricing import get_elb_monthly_cost


class LBType(Enum):
    """Load Balancer 타입"""

    ALB = "application"
    NLB = "network"
    GWLB = "gateway"
    CLB = "classic"

    @classmethod
    def from_string(cls, value: str) -> "LBType":
        """문자열에서 LBType 변환"""
        mapping = {
            "application": cls.ALB,
            "network": cls.NLB,
            "gateway": cls.GWLB,
            "classic": cls.CLB,
        }
        return mapping.get(value, cls.ALB)

    @property
    def display_name(self) -> str:
        """표시용 이름"""
        return {
            LBType.ALB: "ALB",
            LBType.NLB: "NLB",
            LBType.GWLB: "GWLB",
            LBType.CLB: "CLB",
        }.get(self, "Unknown")


class UsageStatus(Enum):
    """사용 상태"""

    UNUSED = "unused"  # 미사용 (타겟 없음)
    UNHEALTHY = "unhealthy"  # 모든 타겟 비정상
    NORMAL = "normal"  # 정상 사용


class Severity(Enum):
    """심각도"""

    CRITICAL = "critical"
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class TargetGroupInfo:
    """타겟 그룹 정보"""

    arn: str
    name: str
    target_type: str
    total_targets: int
    healthy_targets: int
    unhealthy_targets: int


@dataclass
class LoadBalancerInfo:
    """Load Balancer 정보"""

    arn: str
    name: str
    dns_name: str
    lb_type: str  # application, network, gateway, classic
    scheme: str  # internet-facing, internal
    state: str
    vpc_id: str
    availability_zones: list[str]
    created_time: datetime | None
    tags: dict[str, str]

    # 타겟 그룹 (ALB/NLB/GWLB)
    target_groups: list[TargetGroupInfo] = field(default_factory=list)

    # CLB 전용
    registered_instances: int = 0
    healthy_instances: int = 0

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def lb_type_enum(self) -> LBType:
        """LBType enum으로 변환"""
        return LBType.from_string(self.lb_type)

    @property
    def total_targets(self) -> int:
        """전체 타겟 수"""
        if self.lb_type == "classic":
            return self.registered_instances
        return sum(tg.total_targets for tg in self.target_groups)

    @property
    def healthy_targets(self) -> int:
        """정상 타겟 수"""
        if self.lb_type == "classic":
            return self.healthy_instances
        return sum(tg.healthy_targets for tg in self.target_groups)


@dataclass
class LBFinding:
    """LB 분석 결과"""

    lb: LoadBalancerInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class LBAnalysisResult:
    """분석 결과"""

    account_id: str
    account_name: str
    region: str
    lb_type_filter: str | None = None  # 필터링된 LB 타입 (None이면 전체)
    findings: list[LBFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    unhealthy_count: int = 0
    normal_count: int = 0

    # 비용
    unused_monthly_cost: float = 0.0


# =============================================================================
# 수집 함수 - ALB/NLB/GWLB (elbv2)
# =============================================================================


def collect_v2_load_balancers(
    session,
    account_id: str,
    account_name: str,
    region: str,
    lb_type_filter: str | None = None,
) -> list[LoadBalancerInfo]:
    """ALB/NLB/GWLB 목록 수집

    Args:
        session: boto3 session
        account_id: AWS 계정 ID
        account_name: 계정 이름
        region: 리전
        lb_type_filter: 필터링할 LB 타입 (application, network, gateway)
    """
    from botocore.exceptions import ClientError

    from core.parallel import get_client

    load_balancers = []

    try:
        elbv2 = get_client(session, "elbv2", region_name=region)

        # Load Balancers 조회
        paginator = elbv2.get_paginator("describe_load_balancers")
        for page in paginator.paginate():
            for data in page.get("LoadBalancers", []):
                lb_type = data.get("Type", "application")

                # 타입 필터
                if lb_type_filter and lb_type != lb_type_filter:
                    continue

                lb_arn = data.get("LoadBalancerArn", "")

                # 태그 조회
                tags = {}
                try:
                    tag_response = elbv2.describe_tags(ResourceArns=[lb_arn])
                    for tag_desc in tag_response.get("TagDescriptions", []):
                        for t in tag_desc.get("Tags", []):
                            key = t.get("Key", "")
                            if not key.startswith("aws:"):
                                tags[key] = t.get("Value", "")
                except ClientError:
                    pass

                lb = LoadBalancerInfo(
                    arn=lb_arn,
                    name=data.get("LoadBalancerName", ""),
                    dns_name=data.get("DNSName", ""),
                    lb_type=lb_type,
                    scheme=data.get("Scheme", ""),
                    state=data.get("State", {}).get("Code", ""),
                    vpc_id=data.get("VpcId", ""),
                    availability_zones=[az.get("ZoneName", "") for az in data.get("AvailabilityZones", [])],
                    created_time=data.get("CreatedTime"),
                    tags=tags,
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    monthly_cost=get_elb_monthly_cost(region, lb_type),
                )

                # 타겟 그룹 조회
                lb.target_groups = _get_target_groups(elbv2, lb_arn)
                load_balancers.append(lb)

    except ClientError:
        pass

    return load_balancers


def _get_target_groups(elbv2, lb_arn: str) -> list[TargetGroupInfo]:
    """LB에 연결된 타겟 그룹 조회"""
    from botocore.exceptions import ClientError

    target_groups = []

    try:
        response = elbv2.describe_target_groups(LoadBalancerArn=lb_arn)

        for tg in response.get("TargetGroups", []):
            tg_arn = tg.get("TargetGroupArn", "")

            # 타겟 헬스 조회
            healthy = 0
            unhealthy = 0
            total = 0

            try:
                health_response = elbv2.describe_target_health(TargetGroupArn=tg_arn)
                for target in health_response.get("TargetHealthDescriptions", []):
                    total += 1
                    state = target.get("TargetHealth", {}).get("State", "")
                    if state == "healthy":
                        healthy += 1
                    else:
                        unhealthy += 1
            except ClientError:
                pass

            target_groups.append(
                TargetGroupInfo(
                    arn=tg_arn,
                    name=tg.get("TargetGroupName", ""),
                    target_type=tg.get("TargetType", ""),
                    total_targets=total,
                    healthy_targets=healthy,
                    unhealthy_targets=unhealthy,
                )
            )

    except ClientError:
        pass

    return target_groups


# =============================================================================
# 수집 함수 - CLB (elb)
# =============================================================================


def collect_classic_load_balancers(
    session,
    account_id: str,
    account_name: str,
    region: str,
) -> list[LoadBalancerInfo]:
    """Classic Load Balancer 목록 수집"""
    from botocore.exceptions import ClientError

    from core.parallel import get_client

    load_balancers = []

    try:
        elb = get_client(session, "elb", region_name=region)

        response = elb.describe_load_balancers()

        for data in response.get("LoadBalancerDescriptions", []):
            lb_name = data.get("LoadBalancerName", "")

            # 태그 조회
            tags = {}
            try:
                tag_response = elb.describe_tags(LoadBalancerNames=[lb_name])
                for tag_desc in tag_response.get("TagDescriptions", []):
                    for t in tag_desc.get("Tags", []):
                        key = t.get("Key", "")
                        if not key.startswith("aws:"):
                            tags[key] = t.get("Value", "")
            except ClientError:
                pass

            # 인스턴스 헬스 조회
            instances = data.get("Instances", [])
            healthy = 0
            try:
                if instances:
                    health_response = elb.describe_instance_health(LoadBalancerName=lb_name)
                    for state in health_response.get("InstanceStates", []):
                        if state.get("State") == "InService":
                            healthy += 1
            except ClientError:
                pass

            lb = LoadBalancerInfo(
                arn=f"arn:aws:elasticloadbalancing:{region}:{account_id}:loadbalancer/{lb_name}",
                name=lb_name,
                dns_name=data.get("DNSName", ""),
                lb_type="classic",
                scheme=data.get("Scheme", ""),
                state="active",  # CLB는 state 없음
                vpc_id=data.get("VPCId", ""),
                availability_zones=data.get("AvailabilityZones", []),
                created_time=data.get("CreatedTime"),
                tags=tags,
                registered_instances=len(instances),
                healthy_instances=healthy,
                account_id=account_id,
                account_name=account_name,
                region=region,
                monthly_cost=get_elb_monthly_cost(region, "classic"),
            )
            load_balancers.append(lb)

    except ClientError:
        pass

    return load_balancers


# =============================================================================
# 분석 함수
# =============================================================================


def analyze_single_lb(lb: LoadBalancerInfo) -> LBFinding:
    """개별 LB 분석"""

    # 비활성 상태
    if lb.state not in ("active", ""):
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNUSED,
            severity=Severity.LOW,
            description=f"비활성 상태 ({lb.state})",
            recommendation="상태 확인 또는 삭제 검토",
        )

    total = lb.total_targets
    healthy = lb.healthy_targets

    # 타겟 없음
    if total == 0:
        # 타겟 그룹 자체가 없는 경우 (ALB/NLB)
        if lb.lb_type != "classic" and not lb.target_groups:
            return LBFinding(
                lb=lb,
                usage_status=UsageStatus.UNUSED,
                severity=Severity.HIGH,
                description=f"타겟 그룹 없음 (${lb.monthly_cost:.2f}/월)",
                recommendation="타겟 그룹 연결 또는 삭제 검토",
            )

        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNUSED,
            severity=Severity.HIGH,
            description=f"등록된 타겟 없음 (${lb.monthly_cost:.2f}/월)",
            recommendation="타겟 등록 또는 삭제 검토",
        )

    # 모든 타겟 unhealthy
    if healthy == 0:
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNHEALTHY,
            severity=Severity.HIGH,
            description=f"모든 타겟 비정상 ({total}개 unhealthy)",
            recommendation="타겟 헬스체크 확인",
        )

    # 일부 unhealthy
    unhealthy = total - healthy
    if unhealthy > 0:
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.MEDIUM,
            description=f"일부 타겟 비정상 ({healthy}/{total} healthy)",
            recommendation="비정상 타겟 확인",
        )

    # 정상
    return LBFinding(
        lb=lb,
        usage_status=UsageStatus.NORMAL,
        severity=Severity.INFO,
        description=f"정상 ({healthy}개 healthy)",
        recommendation="정상 운영 중",
    )


def analyze_load_balancers(
    load_balancers: list[LoadBalancerInfo],
    account_id: str,
    account_name: str,
    region: str,
    lb_type_filter: str | None = None,
) -> LBAnalysisResult:
    """Load Balancer 미사용 분석"""
    result = LBAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        lb_type_filter=lb_type_filter,
    )

    for lb in load_balancers:
        finding = analyze_single_lb(lb)
        result.findings.append(finding)

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
            result.unused_monthly_cost += lb.monthly_cost
        elif finding.usage_status == UsageStatus.UNHEALTHY:
            result.unhealthy_count += 1
            result.unused_monthly_cost += lb.monthly_cost  # unhealthy도 낭비
        else:
            result.normal_count += 1

    result.total_count = len(load_balancers)
    return result


# =============================================================================
# Excel 보고서 공통
# =============================================================================


def generate_unused_report(
    results: list[LBAnalysisResult],
    output_dir: str,
    lb_type_name: str = "ELB",
) -> str:
    """미사용 분석 Excel 보고서 생성

    Args:
        results: 분석 결과 리스트
        output_dir: 출력 디렉토리
        lb_type_name: 보고서 제목에 사용할 LB 타입 이름 (ELB, ALB, NLB, CLB)
    """
    import os

    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    # 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.UNHEALTHY: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"{lb_type_name} 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "unhealthy": sum(r.unhealthy_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "unused_cost": sum(r.unused_monthly_cost for r in results),
    }

    stats = [
        ("항목", "값"),
        (f"전체 {lb_type_name}", totals["total"]),
        ("미사용", totals["unused"]),
        ("Unhealthy", totals["unhealthy"]),
        ("정상", totals["normal"]),
        ("미사용 월 비용 ($)", f"${totals['unused_cost']:.2f}"),
    ]

    for i, (item, value) in enumerate(stats):
        row = 4 + i
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=value)
        if i == 0:
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            ws.cell(row=row, column=2).font = header_font

    # Findings
    ws2 = wb.create_sheet("Findings")
    headers = [
        "Account",
        "Region",
        "Name",
        "Type",
        "Scheme",
        "Usage",
        "Severity",
        "Targets",
        "Healthy",
        "Monthly Cost ($)",
        "DNS Name",
        "Description",
        "Recommendation",
    ]
    ws2.append(headers)

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # 미사용/unhealthy만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.UNUSED, UsageStatus.UNHEALTHY):
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.lb.monthly_cost, reverse=True)

    for f in all_findings:
        lb = f.lb
        ws2.append(
            [
                lb.account_name,
                lb.region,
                lb.name,
                lb.lb_type.upper(),
                lb.scheme,
                f.usage_status.value,
                f.severity.value,
                lb.total_targets,
                lb.healthy_targets,
                round(lb.monthly_cost, 2),
                lb.dns_name,
                f.description,
                f.recommendation,
            ]
        )

        fill = status_fills.get(f.usage_status)
        if fill:
            ws2.cell(row=ws2.max_row, column=6).fill = fill

    # 열 너비
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 50)

    ws2.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"{lb_type_name}_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath
