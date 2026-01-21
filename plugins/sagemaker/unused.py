"""
plugins/sagemaker/unused.py - SageMaker Endpoint 미사용 분석

유휴/미사용 SageMaker Endpoint 탐지 (CloudWatch 지표 기반)

탐지 기준:
- 미사용: 7일간 Invocations 0회

비용 영향:
- SageMaker Endpoint는 유지만 해도 인스턴스 비용 발생
- 미사용 Endpoint는 월 $100-$5,000+ 낭비 가능

최적화:
- CloudWatch GetMetricData API 사용 (배치 조회)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.cloudwatch import MetricQuery, batch_get_metrics, sanitize_metric_id
from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_sagemaker_monthly_cost

console = Console()

# 분석 기간 (일)
ANALYSIS_DAYS = 7
# 미사용 기준: Invocations 0회
UNUSED_INVOCATION_THRESHOLD = 0

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "sagemaker:ListEndpoints",
        "sagemaker:DescribeEndpoint",
        "cloudwatch:GetMetricData",
    ],
}


class EndpointStatus(Enum):
    """Endpoint 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"


@dataclass
class SageMakerEndpointInfo:
    """SageMaker Endpoint 정보"""

    account_id: str
    account_name: str
    region: str
    endpoint_name: str
    endpoint_arn: str
    status: str  # InService, Creating, Updating, Failed, etc.
    creation_time: datetime | None
    instance_type: str
    instance_count: int
    variant_name: str

    # CloudWatch 지표
    total_invocations: int = 0
    avg_invocations_per_day: float = 0.0
    model_latency_avg_ms: float = 0.0

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (AWS Pricing API 동적 조회)"""
        return get_sagemaker_monthly_cost(
            self.instance_type, self.region, self.instance_count
        )

    @property
    def age_days(self) -> int:
        """Endpoint 생성 후 경과 일수"""
        if not self.creation_time:
            return 0
        now = datetime.now(timezone.utc)
        return (now - self.creation_time.replace(tzinfo=timezone.utc)).days


@dataclass
class EndpointFinding:
    """Endpoint 분석 결과"""

    endpoint: SageMakerEndpointInfo
    status: EndpointStatus
    recommendation: str


@dataclass
class SageMakerAnalysisResult:
    """SageMaker 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_endpoints: int = 0
    unused_endpoints: int = 0
    low_usage_endpoints: int = 0
    normal_endpoints: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[EndpointFinding] = field(default_factory=list)


def collect_sagemaker_endpoints(session, account_id: str, account_name: str, region: str) -> list[SageMakerEndpointInfo]:
    """SageMaker Endpoint 수집 (배치 메트릭 최적화)"""
    from botocore.exceptions import ClientError

    sagemaker = get_client(session, "sagemaker", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    endpoints = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: Endpoint 목록 수집
        paginator = sagemaker.get_paginator("list_endpoints")
        for page in paginator.paginate(StatusEquals="InService"):
            for ep in page.get("Endpoints", []):
                endpoint_name = ep.get("EndpointName", "")

                # Endpoint 상세 정보 조회
                try:
                    detail = sagemaker.describe_endpoint(EndpointName=endpoint_name)
                    production_variants = detail.get("ProductionVariants", [])

                    # 첫 번째 variant 정보 사용
                    if production_variants:
                        variant = production_variants[0]
                        instance_type = variant.get("CurrentInstanceType", variant.get("InstanceType", "unknown"))
                        instance_count = variant.get("CurrentWeight", variant.get("DesiredInstanceCount", 1))
                        if isinstance(instance_count, float):
                            instance_count = int(instance_count)
                        variant_name = variant.get("VariantName", "")
                    else:
                        instance_type = "unknown"
                        instance_count = 1
                        variant_name = ""

                    endpoint = SageMakerEndpointInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        endpoint_name=endpoint_name,
                        endpoint_arn=detail.get("EndpointArn", ""),
                        status=detail.get("EndpointStatus", ""),
                        creation_time=detail.get("CreationTime"),
                        instance_type=instance_type,
                        instance_count=instance_count,
                        variant_name=variant_name,
                    )
                    endpoints.append(endpoint)

                except ClientError:
                    # 상세 조회 실패 시 기본 정보로 추가
                    endpoint = SageMakerEndpointInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        endpoint_name=endpoint_name,
                        endpoint_arn=ep.get("EndpointArn", ""),
                        status=ep.get("EndpointStatus", ""),
                        creation_time=ep.get("CreationTime"),
                        instance_type="unknown",
                        instance_count=1,
                        variant_name="",
                    )
                    endpoints.append(endpoint)

        # 2단계: 메트릭 배치 조회
        if endpoints:
            _collect_sagemaker_metrics_batch(cloudwatch, endpoints, start_time, now)

    except ClientError:
        pass

    return endpoints


def _collect_sagemaker_metrics_batch(
    cloudwatch,
    endpoints: list[SageMakerEndpointInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """SageMaker Endpoint 메트릭 배치 수집 (내부 함수)"""
    from botocore.exceptions import ClientError

    # 쿼리 생성
    queries = []
    for endpoint in endpoints:
        safe_id = sanitize_metric_id(endpoint.endpoint_name)

        # Invocations (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_invocations",
                namespace="AWS/SageMaker",
                metric_name="Invocations",
                dimensions={
                    "EndpointName": endpoint.endpoint_name,
                    "VariantName": endpoint.variant_name if endpoint.variant_name else "AllTraffic",
                },
                stat="Sum",
            )
        )

        # ModelLatency (Average) - 밀리초
        queries.append(
            MetricQuery(
                id=f"{safe_id}_latency",
                namespace="AWS/SageMaker",
                metric_name="ModelLatency",
                dimensions={
                    "EndpointName": endpoint.endpoint_name,
                    "VariantName": endpoint.variant_name if endpoint.variant_name else "AllTraffic",
                },
                stat="Average",
            )
        )

    try:
        # 배치 조회
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        # 결과 매핑
        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for endpoint in endpoints:
            safe_id = sanitize_metric_id(endpoint.endpoint_name)

            endpoint.total_invocations = int(results.get(f"{safe_id}_invocations", 0))
            endpoint.avg_invocations_per_day = endpoint.total_invocations / days
            # 레이턴시는 마이크로초로 반환되므로 밀리초로 변환
            endpoint.model_latency_avg_ms = results.get(f"{safe_id}_latency", 0.0) / 1000

    except ClientError:
        # 실패 시 무시 (기본값 0 유지)
        pass


def analyze_endpoints(
    endpoints: list[SageMakerEndpointInfo], account_id: str, account_name: str, region: str
) -> SageMakerAnalysisResult:
    """SageMaker Endpoint 분석"""
    result = SageMakerAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_endpoints=len(endpoints),
    )

    for endpoint in endpoints:
        # 미사용: Invocations 0회
        if endpoint.total_invocations == UNUSED_INVOCATION_THRESHOLD:
            result.unused_endpoints += 1
            result.unused_monthly_cost += endpoint.estimated_monthly_cost
            result.findings.append(
                EndpointFinding(
                    endpoint=endpoint,
                    status=EndpointStatus.UNUSED,
                    recommendation=f"미사용 - {ANALYSIS_DAYS}일간 호출 0회, {endpoint.age_days}일 경과 - 삭제 검토 (${endpoint.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: 하루 평균 10회 미만
        if endpoint.avg_invocations_per_day < 10:
            result.low_usage_endpoints += 1
            result.low_usage_monthly_cost += endpoint.estimated_monthly_cost
            result.findings.append(
                EndpointFinding(
                    endpoint=endpoint,
                    status=EndpointStatus.LOW_USAGE,
                    recommendation=f"저사용 (일 평균 {endpoint.avg_invocations_per_day:.1f}회) - Serverless Endpoint 전환 검토",
                )
            )
            continue

        result.normal_endpoints += 1
        result.findings.append(
            EndpointFinding(
                endpoint=endpoint,
                status=EndpointStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[SageMakerAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "SageMaker Endpoint 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "Account",
        "Region",
        "전체",
        "미사용",
        "저사용",
        "정상",
        "미사용 비용",
        "저사용 비용",
    ]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_endpoints)
        ws.cell(row=row, column=4, value=r.unused_endpoints)
        ws.cell(row=row, column=5, value=r.low_usage_endpoints)
        ws.cell(row=row, column=6, value=r.normal_endpoints)
        ws.cell(row=row, column=7, value=f"${r.unused_monthly_cost:,.2f}")
        ws.cell(row=row, column=8, value=f"${r.low_usage_monthly_cost:,.2f}")
        if r.unused_endpoints > 0:
            ws.cell(row=row, column=4).fill = red_fill
        if r.low_usage_endpoints > 0:
            ws.cell(row=row, column=5).fill = yellow_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Endpoints")
    detail_headers = [
        "Account",
        "Region",
        "Endpoint Name",
        "Status",
        "Instance Type",
        "Count",
        "Total Invocations",
        "Avg/Day",
        "Latency (ms)",
        "Age (days)",
        "월간 비용",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != EndpointStatus.NORMAL:
                detail_row += 1
                ep = f.endpoint
                ws_detail.cell(row=detail_row, column=1, value=ep.account_name)
                ws_detail.cell(row=detail_row, column=2, value=ep.region)
                ws_detail.cell(row=detail_row, column=3, value=ep.endpoint_name)
                ws_detail.cell(row=detail_row, column=4, value=ep.status)
                ws_detail.cell(row=detail_row, column=5, value=ep.instance_type)
                ws_detail.cell(row=detail_row, column=6, value=ep.instance_count)
                ws_detail.cell(row=detail_row, column=7, value=ep.total_invocations)
                ws_detail.cell(row=detail_row, column=8, value=f"{ep.avg_invocations_per_day:.1f}")
                ws_detail.cell(row=detail_row, column=9, value=f"{ep.model_latency_avg_ms:.2f}")
                ws_detail.cell(row=detail_row, column=10, value=ep.age_days)
                ws_detail.cell(row=detail_row, column=11, value=f"${ep.estimated_monthly_cost:.2f}")
                ws_detail.cell(row=detail_row, column=12, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"SageMaker_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SageMakerAnalysisResult | None:
    """단일 계정/리전의 SageMaker Endpoint 수집 및 분석 (병렬 실행용)"""
    endpoints = collect_sagemaker_endpoints(session, account_id, account_name, region)
    if not endpoints:
        return None
    return analyze_endpoints(endpoints, account_id, account_name, region)


def run(ctx) -> None:
    """SageMaker Endpoint 미사용 분석"""
    console.print("[bold]SageMaker Endpoint 미사용 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="sagemaker")
    results: list[SageMakerAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_endpoints for r in results)
    total_low = sum(r.low_usage_endpoints for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월)"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("sagemaker", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
