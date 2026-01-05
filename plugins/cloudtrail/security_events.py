"""
plugins/cloudtrail/security_events.py - CloudTrail 보안 이벤트 분석

최근 90일간 보안 관련 이벤트를 분석합니다:
- 루트 계정 로그인
- 콘솔 로그인 실패
- IAM 사용자/역할/정책 변경
- Access Key 생성/삭제
- 보안 그룹 변경

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

REQUIRED_PERMISSIONS = {
    "read": [
        "cloudtrail:LookupEvents",
    ],
}


class EventSeverity(Enum):
    """이벤트 심각도"""

    CRITICAL = "critical"
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


# 감시할 보안 이벤트 정의 (핵심 이벤트만 - 분기 감사용)
# 대형 서비스 계정에서도 노이즈 없이 사용 가능한 수준
SECURITY_EVENTS = {
    # === 인증 (루트/실패만 의미있음) ===
    "ConsoleLogin": {
        "category": "인증",
        "description": "콘솔 로그인",
        "severity": EventSeverity.INFO,  # 루트/실패시 상향됨
    },
    # === IAM 핵심 (백도어/자격증명/MFA/권한상승) ===
    "CreateUser": {
        "category": "IAM",
        "description": "IAM 사용자 생성 (백도어)",
        "severity": EventSeverity.HIGH,
    },
    "CreateAccessKey": {
        "category": "IAM",
        "description": "Access Key 생성",
        "severity": EventSeverity.HIGH,
    },
    "DeactivateMFADevice": {
        "category": "IAM",
        "description": "MFA 비활성화",
        "severity": EventSeverity.HIGH,
    },
    "DeleteVirtualMFADevice": {
        "category": "IAM",
        "description": "MFA 디바이스 삭제",
        "severity": EventSeverity.HIGH,
    },
    "UpdateAssumeRolePolicy": {
        "category": "IAM",
        "description": "역할 신뢰 정책 변경 (크로스 계정)",
        "severity": EventSeverity.HIGH,
    },
    "AttachUserPolicy": {
        "category": "IAM",
        "description": "사용자에 관리형 정책 연결",
        "severity": EventSeverity.HIGH,
    },
    "AttachRolePolicy": {
        "category": "IAM",
        "description": "역할에 관리형 정책 연결",
        "severity": EventSeverity.HIGH,
    },
    "PutUserPolicy": {
        "category": "IAM",
        "description": "사용자 인라인 정책 (권한 상승)",
        "severity": EventSeverity.HIGH,
    },
    "PutRolePolicy": {
        "category": "IAM",
        "description": "역할 인라인 정책 (권한 상승)",
        "severity": EventSeverity.HIGH,
    },
    "CreateLoginProfile": {
        "category": "IAM",
        "description": "콘솔 비밀번호 생성 (백도어)",
        "severity": EventSeverity.HIGH,
    },
    # === S3 퍼블릭 노출 ===
    "PutBucketPolicy": {
        "category": "S3",
        "description": "버킷 정책 변경",
        "severity": EventSeverity.HIGH,
    },
    "PutBucketAcl": {
        "category": "S3",
        "description": "버킷 ACL 변경",
        "severity": EventSeverity.HIGH,
    },
    "PutBucketPublicAccessBlock": {
        "category": "S3",
        "description": "퍼블릭 액세스 차단 설정 변경",
        "severity": EventSeverity.HIGH,
    },
    # === KMS ===
    "ScheduleKeyDeletion": {
        "category": "KMS",
        "description": "KMS 키 삭제 예약",
        "severity": EventSeverity.CRITICAL,
    },
    "DisableKey": {
        "category": "KMS",
        "description": "KMS 키 비활성화",
        "severity": EventSeverity.HIGH,
    },
    "PutKeyPolicy": {
        "category": "KMS",
        "description": "KMS 키 정책 변경",
        "severity": EventSeverity.HIGH,
    },
    # === 탐지/로깅 비활성화 (침해 시 필수) ===
    "StopLogging": {
        "category": "CloudTrail",
        "description": "CloudTrail 로깅 중지",
        "severity": EventSeverity.CRITICAL,
    },
    "DeleteTrail": {
        "category": "CloudTrail",
        "description": "CloudTrail 삭제",
        "severity": EventSeverity.CRITICAL,
    },
    "UpdateTrail": {
        "category": "CloudTrail",
        "description": "CloudTrail 설정 변경 (S3 버킷 등)",
        "severity": EventSeverity.HIGH,
    },
    "PutEventSelectors": {
        "category": "CloudTrail",
        "description": "이벤트 선택기 변경 (로깅 우회)",
        "severity": EventSeverity.HIGH,
    },
    "DeleteFlowLogs": {
        "category": "VPC",
        "description": "VPC Flow Logs 삭제",
        "severity": EventSeverity.CRITICAL,
    },
    "DeleteLogGroup": {
        "category": "CloudWatch",
        "description": "CloudWatch 로그 그룹 삭제 (증거 인멸)",
        "severity": EventSeverity.HIGH,
    },
    "DeleteDetector": {
        "category": "GuardDuty",
        "description": "GuardDuty 탐지기 삭제",
        "severity": EventSeverity.CRITICAL,
    },
    "StopConfigurationRecorder": {
        "category": "Config",
        "description": "Config 레코더 중지",
        "severity": EventSeverity.CRITICAL,
    },
    "DeleteConfigurationRecorder": {
        "category": "Config",
        "description": "Config 레코더 삭제",
        "severity": EventSeverity.CRITICAL,
    },
    # === VPC 피어링 (네트워크 경계 침해) ===
    "CreateVpcPeeringConnection": {
        "category": "VPC",
        "description": "VPC 피어링 연결 생성",
        "severity": EventSeverity.HIGH,
    },
    "AcceptVpcPeeringConnection": {
        "category": "VPC",
        "description": "VPC 피어링 연결 수락",
        "severity": EventSeverity.HIGH,
    },
    # === Lambda 외부 노출 ===
    "AddPermission20150331": {
        "category": "Lambda",
        "description": "Lambda 권한 추가 (크로스 계정)",
        "severity": EventSeverity.HIGH,
    },
    "CreateFunctionUrlConfig": {
        "category": "Lambda",
        "description": "Lambda Function URL 생성 (퍼블릭)",
        "severity": EventSeverity.HIGH,
    },
    # === Organizations (통제 이탈) ===
    "LeaveOrganization": {
        "category": "Organizations",
        "description": "조직 탈퇴",
        "severity": EventSeverity.CRITICAL,
    },
    "RemoveAccountFromOrganization": {
        "category": "Organizations",
        "description": "조직에서 계정 제거",
        "severity": EventSeverity.CRITICAL,
    },
    # === 데이터 유출 (스냅샷/AMI 외부 공유) ===
    "ModifySnapshotAttribute": {
        "category": "EBS",
        "description": "EBS 스냅샷 외부 공유",
        "severity": EventSeverity.CRITICAL,
    },
    "ModifyDBSnapshotAttribute": {
        "category": "RDS",
        "description": "RDS 스냅샷 외부 공유",
        "severity": EventSeverity.CRITICAL,
    },
    "ModifyImageAttribute": {
        "category": "EC2",
        "description": "AMI 외부 공유 (데이터 유출)",
        "severity": EventSeverity.CRITICAL,
    },
}


@dataclass
class SecurityEvent:
    """보안 이벤트 정보"""

    account_id: str
    account_name: str
    region: str
    event_time: datetime
    event_name: str
    event_source: str
    category: str
    description: str
    severity: EventSeverity
    user_identity: str
    user_type: str  # Root, IAMUser, AssumedRole, etc.
    source_ip: str
    error_code: str = ""
    error_message: str = ""
    resources: str = ""

    @property
    def is_root(self) -> bool:
        return self.user_type == "Root"

    @property
    def is_failed(self) -> bool:
        return bool(self.error_code)


@dataclass
class SecurityEventResult:
    """보안 이벤트 분석 결과"""

    account_id: str
    account_name: str
    region: str
    events: list[SecurityEvent] = field(default_factory=list)
    root_logins: int = 0
    failed_logins: int = 0
    iam_changes: int = 0
    critical_events: int = 0


def _parse_user_identity(user_identity: dict) -> tuple[str, str]:
    """사용자 정보 파싱"""
    user_type = user_identity.get("type", "Unknown")

    if user_type == "Root":
        return "Root", "Root"
    elif user_type == "IAMUser":
        return user_identity.get("userName", "Unknown"), "IAMUser"
    elif user_type == "AssumedRole":
        arn = user_identity.get("arn", "")
        # arn:aws:sts::123456789012:assumed-role/role-name/session-name
        if "assumed-role/" in arn:
            parts = arn.split("assumed-role/")[1].split("/")
            return f"{parts[0]}/{parts[1]}" if len(parts) > 1 else parts[0], "AssumedRole"
        return arn, "AssumedRole"
    elif user_type == "AWSService":
        return user_identity.get("invokedBy", "AWSService"), "AWSService"
    else:
        return user_identity.get("arn", "Unknown"), user_type


def _collect_security_events(
    session, account_id: str, account_name: str, region: str
) -> SecurityEventResult | None:
    """단일 계정/리전의 보안 이벤트 수집"""
    from botocore.exceptions import ClientError

    result = SecurityEventResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    cloudtrail = get_client(session, "cloudtrail", region_name=region)

    # 최근 90일
    end_time = datetime.now(timezone.utc)
    start_time = end_time - timedelta(days=90)

    # 각 이벤트 타입별로 조회 (lookup_events는 한 번에 하나의 필터만 지원)
    for event_name, event_info in SECURITY_EVENTS.items():
        try:
            paginator = cloudtrail.get_paginator("lookup_events")

            for page in paginator.paginate(
                LookupAttributes=[
                    {"AttributeKey": "EventName", "AttributeValue": event_name}
                ],
                StartTime=start_time,
                EndTime=end_time,
                MaxResults=50,
            ):
                for event in page.get("Events", []):
                    # CloudTrailEvent JSON 파싱
                    import json
                    try:
                        event_data = json.loads(event.get("CloudTrailEvent", "{}"))
                    except json.JSONDecodeError:
                        continue

                    user_identity = event_data.get("userIdentity", {})
                    user_name, user_type = _parse_user_identity(user_identity)

                    # 리소스 정보 추출
                    resources = event.get("Resources", [])
                    resource_str = ", ".join(
                        r.get("ResourceName", r.get("ResourceType", ""))
                        for r in resources[:3]
                    )

                    error_code = event_data.get("errorCode", "")
                    error_message = event_data.get("errorMessage", "")

                    # 심각도 조정 (루트 계정이면 한 단계 상승)
                    severity = event_info["severity"]
                    if user_type == "Root" and severity != EventSeverity.CRITICAL:
                        severity = EventSeverity.HIGH

                    # 로그인 실패면 심각도 상승
                    if event_name == "ConsoleLogin" and error_code:
                        severity = EventSeverity.HIGH

                    sec_event = SecurityEvent(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        event_time=event.get("EventTime", datetime.now(timezone.utc)),
                        event_name=event_name,
                        event_source=event_data.get("eventSource", ""),
                        category=event_info["category"],
                        description=event_info["description"],
                        severity=severity,
                        user_identity=user_name,
                        user_type=user_type,
                        source_ip=event_data.get("sourceIPAddress", ""),
                        error_code=error_code,
                        error_message=error_message,
                        resources=resource_str,
                    )
                    result.events.append(sec_event)

                    # 통계 업데이트
                    if event_name == "ConsoleLogin":
                        if user_type == "Root" and not error_code:
                            result.root_logins += 1
                        if error_code:
                            result.failed_logins += 1

                    if event_info["category"] == "IAM":
                        result.iam_changes += 1

                    if severity == EventSeverity.CRITICAL:
                        result.critical_events += 1

        except ClientError:
            # 개별 이벤트 타입 조회 실패는 무시
            continue

    # 결과가 없어도 반환 (스캔 정보 유지)
    return result if result.events else None


def generate_report(results: list[SecurityEventResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from collections import defaultdict

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    medium_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # 모든 이벤트 수집
    all_events: list[SecurityEvent] = []
    for r in results:
        all_events.extend(r.events)
    all_events.sort(key=lambda e: e.event_time, reverse=True)

    # ========== Summary 시트 ==========
    ws = wb.create_sheet("Summary")
    ws["A1"] = "CloudTrail 보안 이벤트 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "총 이벤트", "루트 로그인", "로그인 실패", "IAM 변경", "Critical"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=len(r.events))
        ws.cell(row=row, column=4, value=r.root_logins)
        ws.cell(row=row, column=5, value=r.failed_logins)
        ws.cell(row=row, column=6, value=r.iam_changes)
        ws.cell(row=row, column=7, value=r.critical_events)

        if r.root_logins > 0:
            ws.cell(row=row, column=4).fill = high_fill
        if r.failed_logins > 0:
            ws.cell(row=row, column=5).fill = medium_fill
        if r.critical_events > 0:
            ws.cell(row=row, column=7).fill = critical_fill

    # ========== Actors 시트 (행위자별 분석) ==========
    ws_actors = wb.create_sheet("Actors")
    actor_headers = ["행위자", "유형", "이벤트 수", "Critical", "High", "주요 활동", "IP 목록"]
    for col, h in enumerate(actor_headers, 1):
        ws_actors.cell(row=1, column=col, value=h).fill = header_fill
        ws_actors.cell(row=1, column=col).font = header_font

    # 행위자별 집계
    actor_stats: dict[str, dict] = defaultdict(lambda: {
        "type": "", "count": 0, "critical": 0, "high": 0,
        "events": defaultdict(int), "ips": set()
    })
    for e in all_events:
        key = e.user_identity
        actor_stats[key]["type"] = e.user_type
        actor_stats[key]["count"] += 1
        actor_stats[key]["events"][e.event_name] += 1
        actor_stats[key]["ips"].add(e.source_ip)
        if e.severity == EventSeverity.CRITICAL:
            actor_stats[key]["critical"] += 1
        elif e.severity == EventSeverity.HIGH:
            actor_stats[key]["high"] += 1

    actor_row = 1
    for actor, stats in sorted(actor_stats.items(), key=lambda x: -x[1]["count"]):
        actor_row += 1
        # 상위 3개 이벤트
        top_events = sorted(stats["events"].items(), key=lambda x: -x[1])[:3]
        top_events_str = ", ".join(f"{k}({v})" for k, v in top_events)
        ips_str = ", ".join(list(stats["ips"])[:5])
        if len(stats["ips"]) > 5:
            ips_str += f" 외 {len(stats['ips']) - 5}개"

        ws_actors.cell(row=actor_row, column=1, value=actor)
        ws_actors.cell(row=actor_row, column=2, value=stats["type"])
        ws_actors.cell(row=actor_row, column=3, value=stats["count"])
        ws_actors.cell(row=actor_row, column=4, value=stats["critical"])
        ws_actors.cell(row=actor_row, column=5, value=stats["high"])
        ws_actors.cell(row=actor_row, column=6, value=top_events_str)
        ws_actors.cell(row=actor_row, column=7, value=ips_str)

        # Root 계정 강조
        if stats["type"] == "Root":
            ws_actors.cell(row=actor_row, column=1).fill = high_fill
        if stats["critical"] > 0:
            ws_actors.cell(row=actor_row, column=4).fill = critical_fill
        if stats["high"] > 0:
            ws_actors.cell(row=actor_row, column=5).fill = high_fill

    # ========== IP Analysis 시트 ==========
    ws_ip = wb.create_sheet("IP Analysis")
    ip_headers = ["IP 주소", "이벤트 수", "행위자 수", "주요 행위자", "Critical", "유형"]
    for col, h in enumerate(ip_headers, 1):
        ws_ip.cell(row=1, column=col, value=h).fill = header_fill
        ws_ip.cell(row=1, column=col).font = header_font

    # IP별 집계
    ip_stats: dict[str, dict] = defaultdict(lambda: {
        "count": 0, "actors": set(), "critical": 0, "type": "Unknown"
    })
    for e in all_events:
        ip = e.source_ip
        ip_stats[ip]["count"] += 1
        ip_stats[ip]["actors"].add(e.user_identity)
        if e.severity == EventSeverity.CRITICAL:
            ip_stats[ip]["critical"] += 1
        # IP 유형 분류
        if ip.startswith("AWS Internal"):
            ip_stats[ip]["type"] = "AWS Internal"
        elif ip.endswith(".amazonaws.com"):
            ip_stats[ip]["type"] = "AWS Service"
        elif ip.startswith(("10.", "172.16.", "172.17.", "172.18.", "172.19.",
                           "172.20.", "172.21.", "172.22.", "172.23.", "172.24.",
                           "172.25.", "172.26.", "172.27.", "172.28.", "172.29.",
                           "172.30.", "172.31.", "192.168.")):
            ip_stats[ip]["type"] = "Private"
        else:
            ip_stats[ip]["type"] = "Public"

    ip_row = 1
    for ip, stats in sorted(ip_stats.items(), key=lambda x: -x[1]["count"]):
        ip_row += 1
        actors_list = list(stats["actors"])
        actors_str = ", ".join(actors_list[:3])
        if len(actors_list) > 3:
            actors_str += f" 외 {len(actors_list) - 3}명"

        ws_ip.cell(row=ip_row, column=1, value=ip)
        ws_ip.cell(row=ip_row, column=2, value=stats["count"])
        ws_ip.cell(row=ip_row, column=3, value=len(stats["actors"]))
        ws_ip.cell(row=ip_row, column=4, value=actors_str)
        ws_ip.cell(row=ip_row, column=5, value=stats["critical"])
        ws_ip.cell(row=ip_row, column=6, value=stats["type"])

        # 외부 IP + 다수 행위자 = 의심
        if stats["type"] == "Public" and len(stats["actors"]) > 3:
            ws_ip.cell(row=ip_row, column=1).fill = warning_fill
        if stats["critical"] > 0:
            ws_ip.cell(row=ip_row, column=5).fill = critical_fill

    # ========== Timeline 시트 (시간대별 분석) ==========
    ws_time = wb.create_sheet("Timeline")
    time_headers = ["시간대 (UTC)", "이벤트 수", "Critical", "High", "주요 이벤트"]
    for col, h in enumerate(time_headers, 1):
        ws_time.cell(row=1, column=col, value=h).fill = header_fill
        ws_time.cell(row=1, column=col).font = header_font

    # 시간대별 집계 (0-23시)
    hourly_stats: dict[int, dict] = {h: {"count": 0, "critical": 0, "high": 0, "events": defaultdict(int)}
                                      for h in range(24)}
    for e in all_events:
        hour = e.event_time.hour
        hourly_stats[hour]["count"] += 1
        hourly_stats[hour]["events"][e.event_name] += 1
        if e.severity == EventSeverity.CRITICAL:
            hourly_stats[hour]["critical"] += 1
        elif e.severity == EventSeverity.HIGH:
            hourly_stats[hour]["high"] += 1

    time_row = 1
    for hour in range(24):
        stats = hourly_stats[hour]
        if stats["count"] == 0:
            continue
        time_row += 1
        top_events = sorted(stats["events"].items(), key=lambda x: -x[1])[:3]
        top_events_str = ", ".join(f"{k}({v})" for k, v in top_events)

        ws_time.cell(row=time_row, column=1, value=f"{hour:02d}:00 - {hour:02d}:59")
        ws_time.cell(row=time_row, column=2, value=stats["count"])
        ws_time.cell(row=time_row, column=3, value=stats["critical"])
        ws_time.cell(row=time_row, column=4, value=stats["high"])
        ws_time.cell(row=time_row, column=5, value=top_events_str)

        if stats["critical"] > 0:
            ws_time.cell(row=time_row, column=3).fill = critical_fill

    # ========== Events 시트 ==========
    ws_events = wb.create_sheet("Events")
    event_headers = [
        "Account", "Region", "시간", "카테고리", "이벤트",
        "설명", "심각도", "사용자", "유형", "IP", "오류", "리소스"
    ]
    for col, h in enumerate(event_headers, 1):
        ws_events.cell(row=1, column=col, value=h).fill = header_fill
        ws_events.cell(row=1, column=col).font = header_font

    event_row = 1
    for e in all_events:
        event_row += 1
        ws_events.cell(row=event_row, column=1, value=e.account_name)
        ws_events.cell(row=event_row, column=2, value=e.region)
        ws_events.cell(row=event_row, column=3, value=e.event_time.strftime("%Y-%m-%d %H:%M:%S"))
        ws_events.cell(row=event_row, column=4, value=e.category)
        ws_events.cell(row=event_row, column=5, value=e.event_name)
        ws_events.cell(row=event_row, column=6, value=e.description)
        ws_events.cell(row=event_row, column=7, value=e.severity.value)
        ws_events.cell(row=event_row, column=8, value=e.user_identity)
        ws_events.cell(row=event_row, column=9, value=e.user_type)
        ws_events.cell(row=event_row, column=10, value=e.source_ip)
        ws_events.cell(row=event_row, column=11, value=e.error_code or "-")
        ws_events.cell(row=event_row, column=12, value=e.resources or "-")

        # 심각도별 색상
        if e.severity == EventSeverity.CRITICAL:
            ws_events.cell(row=event_row, column=7).fill = critical_fill
        elif e.severity == EventSeverity.HIGH:
            ws_events.cell(row=event_row, column=7).fill = high_fill
        elif e.severity == EventSeverity.MEDIUM:
            ws_events.cell(row=event_row, column=7).fill = medium_fill

    # 열 너비 조정
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            col_idx = col[0].column
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"CloudTrail_Security_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def run(ctx) -> None:
    """보안 이벤트 분석"""
    console.print("[bold]CloudTrail 보안 이벤트 분석 시작...[/bold]\n")

    # us-east-1 자동 추가 (IAM 등 글로벌 서비스 이벤트는 us-east-1에 기록됨)
    added_global = False
    if "us-east-1" not in ctx.regions:
        ctx.regions = ["us-east-1"] + list(ctx.regions)
        added_global = True

    console.print(f"[dim]감시 이벤트: {len(SECURITY_EVENTS)}개 (최근 90일)[/dim]")
    console.print("[dim]카테고리: 인증, IAM, S3, KMS, CloudTrail, CloudWatch, VPC, GuardDuty, Config, Lambda, Organizations, EBS, RDS, EC2[/dim]")

    # 스캔 리전 표시
    region_list = ", ".join(ctx.regions)
    if added_global:
        console.print(f"[dim]스캔 리전: {region_list} (us-east-1 자동 추가)[/dim]\n")
    else:
        console.print(f"[dim]스캔 리전: {region_list}[/dim]\n")

    result = parallel_collect(ctx, _collect_security_events, max_workers=5, service="cloudtrail")
    results: list[SecurityEventResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        for err in result.get_errors():
            console.print(f"[dim]  - {err.identifier}/{err.region}: {err.message}[/dim]")

    if not results:
        console.print("\n[yellow]보안 이벤트가 없습니다.[/yellow]")
        console.print("[dim]최근 90일간 감시 대상 이벤트가 발생하지 않았습니다.[/dim]")
        return

    # 통계 계산
    total_events = sum(len(r.events) for r in results)
    total_root = sum(r.root_logins for r in results)
    total_failed = sum(r.failed_logins for r in results)
    total_iam = sum(r.iam_changes for r in results)
    total_critical = sum(r.critical_events for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"총 이벤트: {total_events}건")

    if total_root > 0:
        console.print(f"[red]루트 로그인: {total_root}건[/red]")
    if total_failed > 0:
        console.print(f"[yellow]로그인 실패: {total_failed}건[/yellow]")
    if total_critical > 0:
        console.print(f"[red]Critical 이벤트: {total_critical}건[/red]")

    console.print(f"IAM 변경: {total_iam}건")

    # 카테고리별 통계
    category_counts: dict[str, int] = {}
    for r in results:
        for e in r.events:
            category_counts[e.category] = category_counts.get(e.category, 0) + 1

    if category_counts:
        console.print("\n[bold]카테고리별[/bold]")
        for cat, cnt in sorted(category_counts.items(), key=lambda x: -x[1]):
            console.print(f"  {cat}: {cnt}건")

    # 보고서 생성
    identifier = ctx.accounts[0].id if ctx.accounts else ctx.profile_name or "default"
    output_path = OutputPath(identifier).sub("cloudtrail", "security").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
