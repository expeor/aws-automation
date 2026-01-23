"""
plugins/cloudwatch/loggroup_audit.py - CloudWatch Log Group 미사용 분석

빈 로그 그룹 및 오래된 로그 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "logs:DescribeLogGroups",
        "logs:DescribeLogStreams",
    ],
}

# 비용: $0.03/GB/월 (저장)
COST_PER_GB_MONTH = 0.03
# 기준: 90일 이상 ingestion 없으면 오래된 것으로 판단
OLD_DAYS_THRESHOLD = 90


class LogGroupStatus(Enum):
    """로그 그룹 상태"""

    NORMAL = "normal"
    EMPTY = "empty"  # 로그 없음
    NO_RETENTION = "no_retention"  # 보존 기간 미설정 (무기한)
    OLD = "old"  # 오래된 (ingestion 없음)


@dataclass
class LogGroupInfo:
    """CloudWatch Log Group 정보"""

    account_id: str
    account_name: str
    region: str
    name: str
    arn: str
    creation_time: datetime
    stored_bytes: int
    retention_days: int | None  # None = 무기한
    last_ingestion_time: datetime | None
    log_stream_count: int

    @property
    def stored_gb(self) -> float:
        return self.stored_bytes / (1024**3)

    @property
    def monthly_cost(self) -> float:
        return self.stored_gb * COST_PER_GB_MONTH

    @property
    def age_days(self) -> int:
        return (datetime.now(timezone.utc) - self.creation_time).days

    @property
    def days_since_ingestion(self) -> int | None:
        if self.last_ingestion_time:
            return (datetime.now(timezone.utc) - self.last_ingestion_time).days
        return None


@dataclass
class LogGroupFinding:
    """Log Group 분석 결과"""

    log_group: LogGroupInfo
    status: LogGroupStatus
    recommendation: str


@dataclass
class LogGroupAnalysisResult:
    """Log Group 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    empty_count: int = 0
    no_retention_count: int = 0
    old_count: int = 0
    normal_count: int = 0
    total_stored_gb: float = 0.0
    empty_monthly_cost: float = 0.0
    old_monthly_cost: float = 0.0
    findings: list[LogGroupFinding] = field(default_factory=list)


# =============================================================================
# 수집
# =============================================================================


def collect_log_groups(session, account_id: str, account_name: str, region: str) -> list[LogGroupInfo]:
    """CloudWatch Log Groups 수집"""
    from botocore.exceptions import ClientError

    logs = get_client(session, "logs", region_name=region)
    log_groups = []

    paginator = logs.get_paginator("describe_log_groups")
    for page in paginator.paginate():
        for lg in page.get("logGroups", []):
            creation_ts = lg.get("creationTime", 0)
            creation_time = (
                datetime.fromtimestamp(creation_ts / 1000, tz=timezone.utc)
                if creation_ts
                else datetime.now(timezone.utc)
            )

            # 마지막 ingestion 시간 (lastIngestionTime이 없으면 스트림 확인)
            last_ingestion = None
            if "lastIngestionTime" in lg:
                last_ingestion = datetime.fromtimestamp(lg["lastIngestionTime"] / 1000, tz=timezone.utc)

            # 로그 스트림 개수 (메트릭 데이터에서 확인 불가시 0)
            stream_count = lg.get("metricFilterCount", 0)  # 대략적 추정

            # 실제 스트림 개수 조회 (성능상 limit 1로 확인)
            try:
                streams_resp = logs.describe_log_streams(
                    logGroupName=lg["logGroupName"],
                    limit=1,
                    orderBy="LastEventTime",
                    descending=True,
                )
                streams = streams_resp.get("logStreams", [])
                stream_count = len(streams)  # 최소 1개 있는지 확인

                # 마지막 ingestion 시간 업데이트
                if streams and not last_ingestion and "lastIngestionTime" in streams[0]:
                    last_ingestion = datetime.fromtimestamp(streams[0]["lastIngestionTime"] / 1000, tz=timezone.utc)
            except ClientError:
                pass

            log_groups.append(
                LogGroupInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    name=lg["logGroupName"],
                    arn=lg.get("arn", ""),
                    creation_time=creation_time,
                    stored_bytes=lg.get("storedBytes", 0),
                    retention_days=lg.get("retentionInDays"),
                    last_ingestion_time=last_ingestion,
                    log_stream_count=stream_count,
                )
            )

    return log_groups


# =============================================================================
# 분석
# =============================================================================


def analyze_log_groups(
    log_groups: list[LogGroupInfo], account_id: str, account_name: str, region: str
) -> LogGroupAnalysisResult:
    """Log Group 분석"""
    result = LogGroupAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(log_groups),
    )

    for lg in log_groups:
        result.total_stored_gb += lg.stored_gb

        # 빈 로그 그룹 (0 bytes)
        if lg.stored_bytes == 0:
            result.empty_count += 1
            result.empty_monthly_cost += lg.monthly_cost
            result.findings.append(
                LogGroupFinding(
                    log_group=lg,
                    status=LogGroupStatus.EMPTY,
                    recommendation="빈 로그 그룹 삭제 검토",
                )
            )
            continue

        # 오래된 로그 (ingestion 없음)
        if lg.days_since_ingestion and lg.days_since_ingestion > OLD_DAYS_THRESHOLD:
            result.old_count += 1
            result.old_monthly_cost += lg.monthly_cost
            result.findings.append(
                LogGroupFinding(
                    log_group=lg,
                    status=LogGroupStatus.OLD,
                    recommendation=f"{lg.days_since_ingestion}일간 로그 없음 - 보존 정책 검토",
                )
            )
            continue

        # 보존 기간 미설정 (무기한)
        if lg.retention_days is None and lg.stored_bytes > 0:
            result.no_retention_count += 1
            result.findings.append(
                LogGroupFinding(
                    log_group=lg,
                    status=LogGroupStatus.NO_RETENTION,
                    recommendation="보존 기간 설정 권장 (비용 절감)",
                )
            )
            continue

        result.normal_count += 1
        result.findings.append(
            LogGroupFinding(
                log_group=lg,
                status=LogGroupStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[LogGroupAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="빈 로그", width=10, style="number"),
        ColumnDef(header="오래된", width=10, style="number"),
        ColumnDef(header="무기한 보존", width=12, style="number"),
        ColumnDef(header="저장 (GB)", width=12, style="number"),
        ColumnDef(header="월간 비용", width=12),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row([
            r.account_name,
            r.region,
            r.total_count,
            r.empty_count,
            r.old_count,
            r.no_retention_count,
            round(r.total_stored_gb, 2),
            f"${r.empty_monthly_cost + r.old_monthly_cost:,.2f}",
        ])
        ws = summary_sheet._ws
        if r.empty_count > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.old_count > 0:
            ws.cell(row=row_num, column=5).fill = red_fill
        if r.no_retention_count > 0:
            ws.cell(row=row_num, column=6).fill = yellow_fill

    # 상세 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Log Group", width=40),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="저장 (GB)", width=12, style="number"),
        ColumnDef(header="보존 기간", width=12),
        ColumnDef(header="마지막 Ingestion", width=15),
        ColumnDef(header="월간 비용", width=12, style="number"),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Log Groups", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != LogGroupStatus.NORMAL:
                lg = f.log_group
                style = Styles.danger() if f.status in (LogGroupStatus.EMPTY, LogGroupStatus.OLD) else Styles.warning()
                detail_sheet.add_row([
                    lg.account_name,
                    lg.region,
                    lg.name,
                    f.status.value,
                    round(lg.stored_gb, 4),
                    f"{lg.retention_days}일" if lg.retention_days else "무기한",
                    lg.last_ingestion_time.strftime("%Y-%m-%d") if lg.last_ingestion_time else "-",
                    round(lg.monthly_cost, 4),
                    f.recommendation,
                ], style=style)

    return str(wb.save_as(output_dir, "LogGroup_Audit"))


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> LogGroupAnalysisResult | None:
    """단일 계정/리전의 Log Group 수집 및 분석 (병렬 실행용)"""
    log_groups = collect_log_groups(session, account_id, account_name, region)
    if not log_groups:
        return None
    return analyze_log_groups(log_groups, account_id, account_name, region)


def run(ctx) -> None:
    """CloudWatch Log Group 미사용 분석"""
    console.print("[bold]CloudWatch Log Group 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="logs")
    results: list[LogGroupAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_empty = sum(r.empty_count for r in results)
    total_old = sum(r.old_count for r in results)
    total_no_retention = sum(r.no_retention_count for r in results)
    total_cost = sum(r.empty_monthly_cost + r.old_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"빈 로그 그룹: [red]{total_empty}개[/red]")
    console.print(f"오래된 로그 그룹: [red]{total_old}개[/red]")
    console.print(f"보존 기간 미설정: [yellow]{total_no_retention}개[/yellow]")
    if total_cost > 0:
        console.print(f"월간 절감 가능: [yellow]${total_cost:,.2f}[/yellow]")

    # 보고서 생성
    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("cloudwatch", "inventory").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
