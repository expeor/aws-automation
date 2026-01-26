"""
plugins/ec2/unused.py - EC2 인스턴스 미사용 분석

유휴/저사용 EC2 인스턴스 탐지 (CloudWatch 지표 기반)

탐지 기준 (AWS Trusted Advisor / Trend Micro Conformity 기반):
- 미사용: CPU 평균 < 2%, 네트워크 트래픽 < 5MB, Disk I/O < 100 ops (14일간)
- 저사용: CPU 평균 < 10%, 네트워크/Disk 활동 있음
- 정지됨: stopped 상태

참고:
- AWS Trusted Advisor: CPU < 10%, Network < 5MB (14일)
- Trend Micro Conformity: CPU < 2%, Network < 5MB (7일)

CloudWatch 메트릭:
- Namespace: AWS/EC2
- 메트릭: CPUUtilization, NetworkIn, NetworkOut, DiskReadOps, DiskWriteOps
- Dimension: InstanceId

최적화:
- CloudWatch GetMetricData API 사용 (배치 조회)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath
from plugins.cloudwatch.common import MetricQuery, batch_get_metrics, sanitize_metric_id
from plugins.cost.pricing import get_ec2_monthly_cost

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 분석 기간 (일) - AWS Trusted Advisor 기준 14일
ANALYSIS_DAYS = 14
# 미사용 기준: CPU 평균 2% 미만 (Trend Micro Conformity 기준)
UNUSED_CPU_THRESHOLD = 2.0
# 저사용 기준: CPU 평균 10% 미만 (AWS Trusted Advisor 기준)
LOW_USAGE_CPU_THRESHOLD = 10.0
# 미사용 기준: 네트워크 트래픽 5MB 미만 (AWS Trusted Advisor 기준)
UNUSED_NETWORK_THRESHOLD = 5 * 1024 * 1024  # 5MB
# 미사용 기준: Disk I/O 100 ops 미만
UNUSED_DISK_OPS_THRESHOLD = 100

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeInstances",
        "ec2:DescribeTags",
        "cloudwatch:GetMetricData",
    ],
}


class InstanceStatus(Enum):
    """인스턴스 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"
    STOPPED = "stopped"


@dataclass
class EC2InstanceInfo:
    """EC2 인스턴스 정보"""

    account_id: str
    account_name: str
    region: str
    instance_id: str
    instance_type: str
    state: str
    name: str
    launch_time: datetime | None
    platform: str  # linux or windows
    vpc_id: str
    subnet_id: str
    private_ip: str
    public_ip: str
    tags: dict[str, str] = field(default_factory=dict)

    # CloudWatch 지표
    avg_cpu: float = 0.0
    max_cpu: float = 0.0
    total_network_in: float = 0.0
    total_network_out: float = 0.0
    total_disk_read_ops: float = 0.0
    total_disk_write_ops: float = 0.0

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (AWS Pricing API 동적 조회)"""
        cost = get_ec2_monthly_cost(self.instance_type, self.region)
        if self.platform == "windows":
            cost *= 1.5  # Windows는 약 1.5배
        return cost

    @property
    def age_days(self) -> int:
        """인스턴스 생성 후 경과 일수"""
        if not self.launch_time:
            return 0
        now = datetime.now(timezone.utc)
        return (now - self.launch_time.replace(tzinfo=timezone.utc)).days


@dataclass
class InstanceFinding:
    """인스턴스 분석 결과"""

    instance: EC2InstanceInfo
    status: InstanceStatus
    recommendation: str


@dataclass
class EC2AnalysisResult:
    """EC2 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_instances: int = 0
    unused_instances: int = 0
    low_usage_instances: int = 0
    stopped_instances: int = 0
    normal_instances: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    stopped_monthly_cost: float = 0.0
    findings: list[InstanceFinding] = field(default_factory=list)


def collect_ec2_instances(session, account_id: str, account_name: str, region: str) -> list[EC2InstanceInfo]:
    """EC2 인스턴스 수집 (배치 메트릭 최적화)

    최적화:
    - 기존: 인스턴스당 3 API 호출 → 최적화: 전체 1-2 API 호출
    """
    from botocore.exceptions import ClientError

    ec2 = get_client(session, "ec2", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    instances = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        # 1단계: 인스턴스 목록 수집
        paginator = ec2.get_paginator("describe_instances")
        for page in paginator.paginate():
            for reservation in page.get("Reservations", []):
                for inst in reservation.get("Instances", []):
                    instance_id = inst.get("InstanceId", "")
                    state = inst.get("State", {}).get("Name", "")

                    # 태그에서 Name 추출
                    tags = {t["Key"]: t["Value"] for t in inst.get("Tags", []) if not t["Key"].startswith("aws:")}
                    name = tags.get("Name", "")

                    # 플랫폼 (Windows 또는 Linux)
                    platform = "windows" if inst.get("Platform") == "windows" else "linux"

                    instance = EC2InstanceInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        instance_id=instance_id,
                        instance_type=inst.get("InstanceType", ""),
                        state=state,
                        name=name,
                        launch_time=inst.get("LaunchTime"),
                        platform=platform,
                        vpc_id=inst.get("VpcId", ""),
                        subnet_id=inst.get("SubnetId", ""),
                        private_ip=inst.get("PrivateIpAddress", ""),
                        public_ip=inst.get("PublicIpAddress", ""),
                        tags=tags,
                    )
                    instances.append(instance)

        # 2단계: 실행 중인 인스턴스만 메트릭 조회
        running_instances = [i for i in instances if i.state == "running"]

        if running_instances:
            _collect_ec2_metrics_batch(cloudwatch, running_instances, start_time, now)

    except ClientError:
        pass

    return instances


def _collect_ec2_metrics_batch(
    cloudwatch,
    instances: list[EC2InstanceInfo],
    start_time: datetime,
    end_time: datetime,
) -> None:
    """EC2 인스턴스 메트릭 배치 수집 (내부 함수)"""
    from botocore.exceptions import ClientError

    # 쿼리 생성
    queries = []
    for instance in instances:
        safe_id = sanitize_metric_id(instance.instance_id)
        dimensions = {"InstanceId": instance.instance_id}

        # CPUUtilization (Average, Maximum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_cpu_avg",
                namespace="AWS/EC2",
                metric_name="CPUUtilization",
                dimensions=dimensions,
                stat="Average",
            )
        )
        queries.append(
            MetricQuery(
                id=f"{safe_id}_cpu_max",
                namespace="AWS/EC2",
                metric_name="CPUUtilization",
                dimensions=dimensions,
                stat="Maximum",
            )
        )

        # NetworkIn (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_network_in",
                namespace="AWS/EC2",
                metric_name="NetworkIn",
                dimensions=dimensions,
                stat="Sum",
            )
        )

        # NetworkOut (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_network_out",
                namespace="AWS/EC2",
                metric_name="NetworkOut",
                dimensions=dimensions,
                stat="Sum",
            )
        )

        # DiskReadOps (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_disk_read_ops",
                namespace="AWS/EC2",
                metric_name="DiskReadOps",
                dimensions=dimensions,
                stat="Sum",
            )
        )

        # DiskWriteOps (Sum)
        queries.append(
            MetricQuery(
                id=f"{safe_id}_disk_write_ops",
                namespace="AWS/EC2",
                metric_name="DiskWriteOps",
                dimensions=dimensions,
                stat="Sum",
            )
        )

    try:
        # 배치 조회
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        # 결과 매핑
        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for instance in instances:
            safe_id = sanitize_metric_id(instance.instance_id)

            # CPU 평균 (일별 평균의 평균)
            instance.avg_cpu = results.get(f"{safe_id}_cpu_avg", 0.0) / days
            # CPU 최대값 (일별 최대값의 평균 - batch_get_metrics가 합산하므로)
            instance.max_cpu = results.get(f"{safe_id}_cpu_max", 0.0) / days
            # 네트워크 총량
            instance.total_network_in = results.get(f"{safe_id}_network_in", 0.0)
            instance.total_network_out = results.get(f"{safe_id}_network_out", 0.0)
            # Disk I/O 총량
            instance.total_disk_read_ops = results.get(f"{safe_id}_disk_read_ops", 0.0)
            instance.total_disk_write_ops = results.get(f"{safe_id}_disk_write_ops", 0.0)

    except ClientError:
        # 실패 시 무시 (기본값 0 유지)
        pass


def analyze_instances(
    instances: list[EC2InstanceInfo], account_id: str, account_name: str, region: str
) -> EC2AnalysisResult:
    """EC2 인스턴스 분석"""
    result = EC2AnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_instances=len(instances),
    )

    for instance in instances:
        # 정지된 인스턴스
        if instance.state == "stopped":
            result.stopped_instances += 1
            result.stopped_monthly_cost += instance.estimated_monthly_cost
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.STOPPED,
                    recommendation=f"정지됨 - {instance.age_days}일 경과, 장기 미사용 시 AMI 생성 후 종료 검토",
                )
            )
            continue

        # 실행 중이 아닌 경우 (pending, shutting-down, terminated, stopping)
        if instance.state != "running":
            continue

        total_network = instance.total_network_in + instance.total_network_out
        total_disk_ops = instance.total_disk_read_ops + instance.total_disk_write_ops

        # 미사용: CPU < 2% AND 네트워크 < 5MB AND Disk I/O < 100 ops
        is_cpu_idle = instance.avg_cpu < UNUSED_CPU_THRESHOLD
        is_network_idle = total_network < UNUSED_NETWORK_THRESHOLD
        is_disk_idle = total_disk_ops < UNUSED_DISK_OPS_THRESHOLD

        if is_cpu_idle and is_network_idle and is_disk_idle:
            result.unused_instances += 1
            result.unused_monthly_cost += instance.estimated_monthly_cost
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.UNUSED,
                    recommendation=f"미사용 - CPU {instance.avg_cpu:.1f}%, 네트워크 {total_network / (1024 * 1024):.2f}MB, Disk {total_disk_ops:.0f}ops - 종료 검토 (${instance.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: CPU < 10% (AWS Trusted Advisor 기준)
        if instance.avg_cpu < LOW_USAGE_CPU_THRESHOLD:
            result.low_usage_instances += 1
            result.low_usage_monthly_cost += instance.estimated_monthly_cost
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.LOW_USAGE,
                    recommendation=f"저사용 (CPU {instance.avg_cpu:.1f}%) - 다운사이징 검토",
                )
            )
            continue

        result.normal_instances += 1
        result.findings.append(
            InstanceFinding(
                instance=instance,
                status=InstanceStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[EC2AnalysisResult], output_dir: str, ctx=None) -> dict[str, str]:
    """Excel + HTML 보고서 생성

    Args:
        results: 분석 결과 리스트
        output_dir: 출력 디렉토리
        ctx: 실행 컨텍스트 (출력 설정 및 HTML 생성용)

    Returns:
        생성된 파일 경로 딕셔너리 {"excel": "...", "html": "..."}
    """
    from core.tools.io.compat import generate_reports

    # HTML 생성용 flat 데이터
    flat_data = []
    for r in results:
        for f in r.findings:
            if f.status != InstanceStatus.NORMAL:
                inst = f.instance
                flat_data.append(
                    {
                        "account_id": inst.account_id,
                        "account_name": inst.account_name,
                        "region": inst.region,
                        "resource_id": inst.instance_id,
                        "resource_name": inst.name,
                        "resource_type": inst.instance_type,
                        "status": f.status.value,
                        "reason": f.recommendation,
                        "cost": inst.estimated_monthly_cost,
                        # Extra fields
                        "state": inst.state,
                        "platform": inst.platform,
                        "avg_cpu": f"{inst.avg_cpu:.1f}%",
                        "max_cpu": f"{inst.max_cpu:.1f}%",
                        "disk_io": f"{inst.total_disk_read_ops + inst.total_disk_write_ops:,.0f} ops",
                        "age_days": inst.age_days,
                    }
                )

    # 집계
    total_instances = sum(r.total_instances for r in results)
    total_unused = sum(r.unused_instances for r in results)
    total_low = sum(r.low_usage_instances for r in results)
    total_stopped = sum(r.stopped_instances for r in results)
    total_savings = sum(r.unused_monthly_cost + r.low_usage_monthly_cost for r in results)

    return generate_reports(
        ctx,
        data=flat_data,
        excel_generator=lambda d: _save_excel(results, d),
        html_config={
            "title": "EC2 미사용 인스턴스 분석",
            "service": "EC2",
            "tool_name": "unused",
            "total": total_instances,
            "found": total_unused + total_low + total_stopped,
            "savings": total_savings,
        },
        output_dir=output_dir,
    )


def _save_excel(results: list[EC2AnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성 (내부 함수)"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정지", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
        ColumnDef(header="정지 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_instances,
                r.unused_instances,
                r.low_usage_instances,
                r.stopped_instances,
                r.normal_instances,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
                f"${r.stopped_monthly_cost:,.2f}",
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.unused_instances > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_instances > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill
        if r.stopped_instances > 0:
            ws.cell(row=row_num, column=6).fill = gray_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Instance ID", width=22),
        ColumnDef(header="Name", width=25),
        ColumnDef(header="Type", width=15),
        ColumnDef(header="State", width=12),
        ColumnDef(header="Platform", width=10),
        ColumnDef(header="Avg CPU", width=10),
        ColumnDef(header="Max CPU", width=10),
        ColumnDef(header="Network In", width=15),
        ColumnDef(header="Network Out", width=15),
        ColumnDef(header="Disk I/O", width=12),
        ColumnDef(header="Age (days)", width=12, style="number"),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("Instances", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != InstanceStatus.NORMAL:
                inst = f.instance
                style = None
                if f.status == InstanceStatus.UNUSED:
                    style = Styles.danger()
                elif f.status == InstanceStatus.LOW_USAGE:
                    style = Styles.warning()

                total_disk_ops = inst.total_disk_read_ops + inst.total_disk_write_ops
                detail_sheet.add_row(
                    [
                        inst.account_name,
                        inst.region,
                        inst.instance_id,
                        inst.name,
                        inst.instance_type,
                        inst.state,
                        inst.platform,
                        f"{inst.avg_cpu:.1f}%",
                        f"{inst.max_cpu:.1f}%",
                        f"{inst.total_network_in / (1024 * 1024):.2f} MB",
                        f"{inst.total_network_out / (1024 * 1024):.2f} MB",
                        f"{total_disk_ops:,.0f} ops",
                        inst.age_days,
                        f"${inst.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "EC2_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EC2AnalysisResult | None:
    """단일 계정/리전의 EC2 인스턴스 수집 및 분석 (병렬 실행용)"""
    instances = collect_ec2_instances(session, account_id, account_name, region)
    if not instances:
        return None
    return analyze_instances(instances, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """EC2 인스턴스 미사용 분석"""
    console.print("[bold]EC2 인스턴스 미사용 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="ec2")
    results: list[EC2AnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_instances for r in results)
    total_low = sum(r.low_usage_instances for r in results)
    total_stopped = sum(r.stopped_instances for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)
    stopped_cost = sum(r.stopped_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월) / "
        f"정지: {total_stopped}개 (${stopped_cost:,.2f}/월)"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("ec2", "unused").with_date().build()

    # Excel + HTML 동시 생성 (ctx.output_config 설정에 따라)
    report_paths = generate_report(results, output_path, ctx)

    from core.tools.output import print_report_complete

    print_report_complete(report_paths, "분석 완료!")
