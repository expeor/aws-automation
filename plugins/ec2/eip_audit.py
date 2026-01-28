"""
plugins/ec2/eip_audit.py - EIP 미사용 분석

미사용 Elastic IP 탐지

분석 기준:
- AssociationId가 없는 EIP (아무 리소스에도 연결되지 않음)

참고:
- 미연결 EIP는 시간당 $0.005 = 월 ~$3.60 비용 발생

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from shared.aws.pricing import get_eip_monthly_cost

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeAddresses",
    ],
}


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """사용 상태"""

    UNUSED = "unused"  # 미사용 (미연결)
    NORMAL = "normal"  # 정상 사용 (연결됨)
    PENDING = "pending"  # 확인 필요


class Severity(Enum):
    """심각도"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class EIPInfo:
    """Elastic IP 정보"""

    allocation_id: str
    public_ip: str
    domain: str  # vpc or standard
    instance_id: str
    association_id: str
    network_interface_id: str
    private_ip: str
    network_border_group: str
    tags: dict[str, str]
    name: str

    # 메타
    account_id: str
    account_name: str
    region: str

    # 비용
    monthly_cost: float = 0.0

    @property
    def is_associated(self) -> bool:
        """연결 여부"""
        return bool(self.association_id)


@dataclass
class EIPFinding:
    """EIP 분석 결과"""

    eip: EIPInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class EIPAnalysisResult:
    """분석 결과"""

    account_id: str
    account_name: str
    region: str
    findings: list[EIPFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    normal_count: int = 0
    pending_count: int = 0

    # 비용
    unused_monthly_cost: float = 0.0


# =============================================================================
# 수집
# =============================================================================


def collect_eips(session, account_id: str, account_name: str, region: str) -> list[EIPInfo]:
    """EIP 목록 수집"""
    from botocore.exceptions import ClientError

    eips = []

    try:
        ec2 = get_client(session, "ec2", region_name=region)
        response = ec2.describe_addresses()

        for data in response.get("Addresses", []):
            # 태그 파싱
            tags = {
                t.get("Key", ""): t.get("Value", "")
                for t in data.get("Tags", [])
                if not t.get("Key", "").startswith("aws:")
            }

            # 미연결 시 비용 계산
            is_associated = bool(data.get("AssociationId"))
            monthly_cost = 0.0 if is_associated else get_eip_monthly_cost(region)

            eip = EIPInfo(
                allocation_id=data.get("AllocationId", ""),
                public_ip=data.get("PublicIp", ""),
                domain=data.get("Domain", ""),
                instance_id=data.get("InstanceId", ""),
                association_id=data.get("AssociationId", ""),
                network_interface_id=data.get("NetworkInterfaceId", ""),
                private_ip=data.get("PrivateIpAddress", ""),
                network_border_group=data.get("NetworkBorderGroup", ""),
                tags=tags,
                name=tags.get("Name", ""),
                account_id=account_id,
                account_name=account_name,
                region=region,
                monthly_cost=monthly_cost,
            )
            eips.append(eip)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} EIP 수집 오류: {error_code}[/yellow]")

    return eips


# =============================================================================
# 분석
# =============================================================================


def analyze_eips(eips: list[EIPInfo], account_id: str, account_name: str, region: str) -> EIPAnalysisResult:
    """EIP 미사용 분석"""
    result = EIPAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for eip in eips:
        finding = _analyze_single_eip(eip)
        result.findings.append(finding)

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
            result.unused_monthly_cost += eip.monthly_cost
        elif finding.usage_status == UsageStatus.NORMAL:
            result.normal_count += 1
        elif finding.usage_status == UsageStatus.PENDING:
            result.pending_count += 1

    result.total_count = len(eips)
    return result


def _analyze_single_eip(eip: EIPInfo) -> EIPFinding:
    """개별 EIP 분석"""

    # 1. 연결됨 = 정상
    if eip.is_associated:
        attached_to = eip.instance_id or eip.network_interface_id or "unknown"
        return EIPFinding(
            eip=eip,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"사용 중 ({attached_to})",
            recommendation="정상 사용 중",
        )

    # 2. 미연결 = 미사용
    unused_cost = get_eip_monthly_cost(eip.region)
    return EIPFinding(
        eip=eip,
        usage_status=UsageStatus.UNUSED,
        severity=Severity.HIGH,  # EIP 미사용은 항상 비용 발생
        description=f"미연결 EIP (월 ${unused_cost:.2f} 비용 발생)",
        recommendation="사용하지 않으면 릴리스 검토",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[EIPAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # Summary sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("EIP 미사용 분석 보고서")
    summary.add_item("생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()

    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "pending": sum(r.pending_count for r in results),
        "unused_cost": sum(r.unused_monthly_cost for r in results),
    }

    summary.add_section("통계")
    summary.add_item("전체 EIP", totals["total"])
    summary.add_item("미사용", totals["unused"], highlight="danger" if totals["unused"] > 0 else None)
    summary.add_item("정상 사용", totals["normal"])
    summary.add_item("확인 필요", totals["pending"], highlight="warning" if totals["pending"] > 0 else None)
    summary.add_item("미사용 월 비용 ($)", f"${totals['unused_cost']:.2f}")

    # Findings sheet
    columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Allocation ID", width=25),
        ColumnDef(header="Public IP", width=16),
        ColumnDef(header="Name", width=25),
        ColumnDef(header="Usage", width=12),
        ColumnDef(header="Severity", width=10),
        ColumnDef(header="Instance ID", width=22),
        ColumnDef(header="ENI ID", width=22),
        ColumnDef(header="Private IP", width=16),
        ColumnDef(header="Monthly Cost ($)", width=15, style="number"),
        ColumnDef(header="Description", width=40),
        ColumnDef(header="Recommendation", width=30),
    ]
    sheet = wb.new_sheet("Findings", columns)

    # 상태별 스타일
    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.PENDING: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # 미사용/확인필요만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.UNUSED, UsageStatus.PENDING):
                all_findings.append(f)

    # 심각도순
    severity_order = {
        Severity.HIGH: 0,
        Severity.MEDIUM: 1,
        Severity.LOW: 2,
        Severity.INFO: 3,
    }
    all_findings.sort(key=lambda x: severity_order.get(x.severity, 9))

    for f in all_findings:
        eip = f.eip
        style = Styles.danger() if f.usage_status == UsageStatus.UNUSED else Styles.warning()
        row_num = sheet.add_row(
            [
                eip.account_name,
                eip.region,
                eip.allocation_id,
                eip.public_ip,
                eip.name,
                f.usage_status.value,
                f.severity.value,
                eip.instance_id,
                eip.network_interface_id,
                eip.private_ip,
                round(eip.monthly_cost, 2),
                f.description,
                f.recommendation,
            ],
            style=style,
        )

        # Usage 컬럼에 상태별 색상 적용
        fill = status_fills.get(f.usage_status)
        if fill:
            sheet._ws.cell(row=row_num, column=6).fill = fill

    return str(wb.save_as(output_dir, "EIP_Unused"))


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EIPAnalysisResult:
    """단일 계정/리전의 EIP 수집 및 분석 (병렬 실행용)"""
    eips = collect_eips(session, account_id, account_name, region)
    return analyze_eips(eips, account_id, account_name, region)


def run(ctx: ExecutionContext) -> None:
    """EIP 미사용 분석 실행 (병렬 처리)"""
    console.print("[bold]EIP 미사용 분석 시작...[/bold]")

    # 병렬 수집
    result = parallel_collect(
        ctx,
        _collect_and_analyze,
        max_workers=20,
        service="ec2",
    )

    # 결과 처리
    all_results: list[EIPAnalysisResult] = result.get_data()

    # 진행 상황 출력
    console.print(f"  [dim]수집 완료: 성공 {result.success_count}, 실패 {result.error_count}[/dim]")

    # 에러 요약
    if result.error_count > 0:
        console.print(f"\n[yellow]{result.get_error_summary()}[/yellow]")

    if not all_results:
        console.print("[yellow]분석할 EIP 없음[/yellow]")
        return

    # 개별 결과 요약
    for r in all_results:
        if r.unused_count > 0:
            cost_str = f" (${r.unused_monthly_cost:.2f}/월)" if r.unused_monthly_cost > 0 else ""
            console.print(f"  {r.account_name}/{r.region}: [red]미사용 {r.unused_count}개{cost_str}[/red]")
        elif r.pending_count > 0:
            console.print(f"  {r.account_name}/{r.region}: [yellow]확인 필요 {r.pending_count}개[/yellow]")
        elif r.total_count > 0:
            console.print(f"  {r.account_name}/{r.region}: [green]정상 {r.normal_count}개[/green]")

    # 전체 통계
    totals = {
        "total": sum(r.total_count for r in all_results),
        "unused": sum(r.unused_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "pending": sum(r.pending_count for r in all_results),
        "unused_cost": sum(r.unused_monthly_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 EIP: {totals['total']}개[/bold]")
    if totals["unused"] > 0:
        console.print(f"  [red bold]미사용: {totals['unused']}개 (${totals['unused_cost']:.2f}/월)[/red bold]")
    if totals["pending"] > 0:
        console.print(f"  [yellow]확인 필요: {totals['pending']}개[/yellow]")
    console.print(f"  [green]정상: {totals['normal']}개[/green]")

    # 보고서
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("eip", "inventory").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
