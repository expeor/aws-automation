"""
plugins/vpc/network_analysis.py - Consolidated Network Resource Analysis

Analyzes unused/underutilized network resources:
- NAT Gateways (unused/low-usage)
- VPC Endpoints (Interface endpoints with issues)
- ENIs (unused Elastic Network Interfaces)

Consolidates: nat_audit + endpoint_audit + eni_audit

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from rich.console import Console
from rich.table import Table

from core.parallel import parallel_collect
from core.tools.output import OutputPath, open_in_explorer

# Import sub-modules for NAT analysis
from .nat_audit_analysis import NATAnalyzer, NATCollector, NATExcelReporter

# Import types and functions from endpoint_audit
from .endpoint_audit import (
    EndpointAnalysisResult,
    EndpointStatus,
    analyze_endpoints,
    collect_endpoints,
    generate_report as generate_endpoint_report,
)

# Import types and functions from eni_audit
from .eni_audit import (
    ENIAnalysisResult,
    UsageStatus as ENIUsageStatus,
    analyze_enis,
    collect_enis,
    generate_report as generate_eni_report,
)

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        # NAT Gateway
        "ec2:DescribeNatGateways",
        "cloudwatch:GetMetricStatistics",
        # VPC Endpoint
        "ec2:DescribeVpcEndpoints",
        # ENI
        "ec2:DescribeNetworkInterfaces",
    ],
}


@dataclass
class NetworkAnalysisSummary:
    """Combined network analysis summary"""

    account_id: str
    account_name: str
    region: str

    # NAT Gateway
    nat_total: int = 0
    nat_unused: int = 0
    nat_low_usage: int = 0
    nat_monthly_waste: float = 0.0

    # VPC Endpoint
    endpoint_total: int = 0
    endpoint_interface: int = 0
    endpoint_unused: int = 0
    endpoint_monthly_waste: float = 0.0

    # ENI
    eni_total: int = 0
    eni_unused: int = 0
    eni_pending: int = 0


@dataclass
class NetworkAnalysisResult:
    """Combined network analysis result"""

    summaries: list[NetworkAnalysisSummary] = field(default_factory=list)

    # Detailed results
    nat_results: list[Any] = field(default_factory=list)
    nat_stats: list[dict[str, Any]] = field(default_factory=list)
    endpoint_results: list[EndpointAnalysisResult] = field(default_factory=list)
    eni_results: list[ENIAnalysisResult] = field(default_factory=list)

    # Errors
    errors: list[str] = field(default_factory=list)


def _collect_nat(session, account_id: str, account_name: str, region: str) -> tuple[Any, dict[str, Any]] | None:
    """Collect and analyze NAT Gateways"""
    collector = NATCollector()
    audit_data = collector.collect(session, account_id, account_name, region)

    if not audit_data.nat_gateways:
        return None

    analyzer = NATAnalyzer(audit_data)
    analysis_result = analyzer.analyze()
    stats = analyzer.get_summary_stats()

    return (analysis_result, stats)


def _collect_endpoint(session, account_id: str, account_name: str, region: str) -> EndpointAnalysisResult | None:
    """Collect and analyze VPC Endpoints"""
    endpoints = collect_endpoints(session, account_id, account_name, region)
    if not endpoints:
        return None
    return analyze_endpoints(endpoints, session, account_id, account_name, region)


def _collect_eni(session, account_id: str, account_name: str, region: str) -> ENIAnalysisResult:
    """Collect and analyze ENIs"""
    enis = collect_enis(session, account_id, account_name, region)
    return analyze_enis(enis, account_id, account_name, region)


def _collect_all(session, account_id: str, account_name: str, region: str) -> dict[str, Any]:
    """Collect all network resources for a single account/region"""
    return {
        "nat": _collect_nat(session, account_id, account_name, region),
        "endpoint": _collect_endpoint(session, account_id, account_name, region),
        "eni": _collect_eni(session, account_id, account_name, region),
        "account_id": account_id,
        "account_name": account_name,
        "region": region,
    }


def generate_combined_report(result: NetworkAnalysisResult, output_dir: str) -> str:
    """Generate combined Excel report for all network resources"""
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Network Resource Analysis Report"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Summary statistics
    total_nat = sum(s.nat_total for s in result.summaries)
    total_nat_unused = sum(s.nat_unused for s in result.summaries)
    total_nat_waste = sum(s.nat_monthly_waste for s in result.summaries)

    total_endpoint = sum(s.endpoint_total for s in result.summaries)
    total_endpoint_interface = sum(s.endpoint_interface for s in result.summaries)
    total_endpoint_unused = sum(s.endpoint_unused for s in result.summaries)
    total_endpoint_waste = sum(s.endpoint_monthly_waste for s in result.summaries)

    total_eni = sum(s.eni_total for s in result.summaries)
    total_eni_unused = sum(s.eni_unused for s in result.summaries)

    summary_data = [
        ("Resource", "Total", "Unused/Issue", "Monthly Waste ($)"),
        ("NAT Gateway", total_nat, total_nat_unused, f"${total_nat_waste:,.2f}"),
        ("VPC Endpoint (Interface)", total_endpoint_interface, total_endpoint_unused, f"${total_endpoint_waste:,.2f}"),
        ("ENI", total_eni, total_eni_unused, "-"),
    ]

    for i, row_data in enumerate(summary_data):
        row = 4 + i
        for col, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font

    # Per-region breakdown
    ws_summary["A10"] = "Per-Region Summary"
    ws_summary["A10"].font = Font(bold=True, size=12)

    region_headers = ["Account", "Region", "NAT Unused", "Endpoint Unused", "ENI Unused"]
    for col, h in enumerate(region_headers, 1):
        cell = ws_summary.cell(row=11, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row = 12
    for s in result.summaries:
        ws_summary.cell(row=row, column=1, value=s.account_name)
        ws_summary.cell(row=row, column=2, value=s.region)
        ws_summary.cell(row=row, column=3, value=s.nat_unused)
        ws_summary.cell(row=row, column=4, value=s.endpoint_unused)
        ws_summary.cell(row=row, column=5, value=s.eni_unused)
        row += 1

    # Column widths
    for col in ws_summary.columns:
        max_len = max(len(str(c.value) if c.value else "") for c in col)
        ws_summary.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 40)

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Network_Analysis_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


def run(ctx) -> None:
    """Run consolidated network resource analysis"""
    console.print("[bold]Network Resource Analysis (NAT/Endpoint/ENI)[/bold]\n")

    # Parallel collection
    console.print("[cyan]Step 1: Collecting network resources...[/cyan]")
    result_data = parallel_collect(ctx, _collect_all, max_workers=20, service="ec2")

    # Process results
    analysis_result = NetworkAnalysisResult()

    for data in result_data.get_data():
        if data is None:
            continue

        account_id = data["account_id"]
        account_name = data["account_name"]
        region = data["region"]

        summary = NetworkAnalysisSummary(
            account_id=account_id,
            account_name=account_name,
            region=region,
        )

        # NAT results
        nat_data = data.get("nat")
        if nat_data is not None:
            nat_result, nat_stats = nat_data
            analysis_result.nat_results.append(nat_result)
            analysis_result.nat_stats.append(nat_stats)
            summary.nat_total = nat_stats.get("total_nat_count", 0)
            summary.nat_unused = nat_stats.get("unused_count", 0)
            summary.nat_low_usage = nat_stats.get("low_usage_count", 0)
            summary.nat_monthly_waste = nat_stats.get("total_monthly_waste", 0)

        # Endpoint results
        endpoint_data = data.get("endpoint")
        if endpoint_data is not None:
            analysis_result.endpoint_results.append(endpoint_data)
            summary.endpoint_total = endpoint_data.total_count
            summary.endpoint_interface = endpoint_data.interface_count
            summary.endpoint_unused = endpoint_data.unused_count
            summary.endpoint_monthly_waste = endpoint_data.unused_monthly_cost

        # ENI results
        eni_data = data.get("eni")
        if eni_data is not None:
            analysis_result.eni_results.append(eni_data)
            summary.eni_total = eni_data.total_count
            summary.eni_unused = eni_data.unused_count
            summary.eni_pending = eni_data.pending_count

        analysis_result.summaries.append(summary)

    # Handle errors
    if result_data.error_count > 0:
        console.print(f"[yellow]Errors: {result_data.error_count}[/yellow]")
        console.print(f"[dim]{result_data.get_error_summary()}[/dim]")

    if not analysis_result.summaries:
        console.print("[yellow]No resources found.[/yellow]")
        return

    # Print summary
    console.print("\n[cyan]Step 2: Analysis Summary[/cyan]")
    _print_summary(analysis_result)

    # Generate reports
    console.print("\n[cyan]Step 3: Generating reports...[/cyan]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("vpc", "network").with_date().build()

    # Combined summary report
    combined_filepath = generate_combined_report(analysis_result, output_path)
    console.print(f"  Combined report: {combined_filepath}")

    # Individual detailed reports
    if analysis_result.nat_results:
        nat_reporter = NATExcelReporter(analysis_result.nat_results, analysis_result.nat_stats)
        nat_filepath = nat_reporter.generate(output_path)
        console.print(f"  NAT Gateway report: {nat_filepath}")

    if analysis_result.endpoint_results:
        endpoint_filepath = generate_endpoint_report(analysis_result.endpoint_results, output_path)
        console.print(f"  VPC Endpoint report: {endpoint_filepath}")

    if analysis_result.eni_results:
        eni_filepath = generate_eni_report(analysis_result.eni_results, output_path)
        console.print(f"  ENI report: {eni_filepath}")

    console.print("\n[bold green]Analysis complete![/bold green]")
    open_in_explorer(output_path)


def _print_summary(result: NetworkAnalysisResult) -> None:
    """Print analysis summary"""
    # NAT Gateway
    total_nat = sum(s.nat_total for s in result.summaries)
    unused_nat = sum(s.nat_unused for s in result.summaries)
    low_usage_nat = sum(s.nat_low_usage for s in result.summaries)
    nat_waste = sum(s.nat_monthly_waste for s in result.summaries)

    console.print(f"\n[bold]NAT Gateway:[/bold] {total_nat} total")
    if unused_nat > 0:
        console.print(f"  [red]Unused: {unused_nat}[/red]")
    if low_usage_nat > 0:
        console.print(f"  [yellow]Low usage: {low_usage_nat}[/yellow]")
    if nat_waste > 0:
        console.print(f"  [red]Monthly waste: ${nat_waste:,.2f}[/red]")

    # VPC Endpoint
    total_endpoint = sum(s.endpoint_total for s in result.summaries)
    interface_endpoint = sum(s.endpoint_interface for s in result.summaries)
    unused_endpoint = sum(s.endpoint_unused for s in result.summaries)
    endpoint_waste = sum(s.endpoint_monthly_waste for s in result.summaries)

    console.print(f"\n[bold]VPC Endpoint:[/bold] {total_endpoint} total ({interface_endpoint} Interface)")
    if unused_endpoint > 0:
        console.print(f"  [red]Unused/Issues: {unused_endpoint}[/red]")
    if endpoint_waste > 0:
        console.print(f"  [red]Monthly waste: ${endpoint_waste:,.2f}[/red]")

    # ENI
    total_eni = sum(s.eni_total for s in result.summaries)
    unused_eni = sum(s.eni_unused for s in result.summaries)
    pending_eni = sum(s.eni_pending for s in result.summaries)

    console.print(f"\n[bold]ENI:[/bold] {total_eni} total")
    if unused_eni > 0:
        console.print(f"  [red]Unused: {unused_eni}[/red]")
    if pending_eni > 0:
        console.print(f"  [yellow]Pending review: {pending_eni}[/yellow]")
