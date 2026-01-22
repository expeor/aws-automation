"""
plugins/vpc/ip_search/private_ip/export.py - Export Utilities

Export search results to CSV, Excel, or Clipboard.
"""

import csv
import os
from datetime import datetime
from typing import TYPE_CHECKING

from core.tools.output.builder import OutputPath

if TYPE_CHECKING:
    from .cache import PrivateIPResult


def _get_output_dir(session_name: str) -> str:
    """Get output directory path."""
    return OutputPath(session_name).sub("vpc", "private_ip_search").with_date("daily").build()


def get_headers(lang: str = "ko") -> list[str]:
    """Get column headers based on language."""
    if lang == "en":
        return [
            "IP Address",
            "Profile",
            "Account ID",
            "Account Name",
            "Region",
            "Resource",
            "ENI ID",
            "VPC ID",
            "Subnet ID",
            "Private IP",
            "Public IP",
            "Interface Type",
            "Status",
            "Description",
            "Security Groups",
            "Name",
            "Managed",
            "Managed By",
        ]
    else:
        return [
            "IP 주소",
            "프로파일",
            "계정 ID",
            "계정명",
            "리전",
            "리소스",
            "ENI ID",
            "VPC ID",
            "Subnet ID",
            "Private IP",
            "Public IP",
            "인터페이스 타입",
            "상태",
            "설명",
            "Security Groups",
            "이름",
            "관리형",
            "관리자",
        ]


def result_to_row(r: "PrivateIPResult") -> list[str]:
    """Convert a PrivateIPResult to a row of values."""
    return [
        r.ip_address,
        r.profile_name,
        r.account_id,
        r.account_name,
        r.region,
        r.mapped_resource,
        r.eni_id,
        r.vpc_id,
        r.subnet_id,
        r.private_ip,
        r.public_ip,
        r.interface_type,
        r.status,
        r.description,
        ", ".join(r.security_groups),
        r.name,
        "Yes" if r.is_managed else "No",
        r.managed_by,
    ]


def export_csv(
    results: list["PrivateIPResult"],
    session_name: str,
    lang: str = "ko",
) -> str:
    """
    Export results to CSV file.

    Args:
        results: List of search results
        session_name: Session name for output path
        lang: Language for headers ("ko" or "en")

    Returns:
        Filepath of created CSV, or empty string on failure
    """
    if not results:
        return ""

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"private_ip_{timestamp}.csv"
    filepath = os.path.join(_get_output_dir(session_name), filename)

    headers = get_headers(lang)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for r in results:
            writer.writerow(result_to_row(r))

    return filepath


def export_excel(
    results: list["PrivateIPResult"],
    session_name: str,
    lang: str = "ko",
) -> str:
    """
    Export results to Excel file.

    Args:
        results: List of search results
        session_name: Session name for output path
        lang: Language for headers ("ko" or "en")

    Returns:
        Filepath of created Excel file, or empty string on failure
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return ""

    if not results:
        return ""

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"private_ip_{timestamp}.xlsx"
    filepath = os.path.join(_get_output_dir(session_name), filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Private IP Search"

    # Header styling
    headers = get_headers(lang)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_idx, r in enumerate(results, 2):
        row_values = result_to_row(r)
        for col_idx, value in enumerate(row_values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    wb.save(filepath)
    return filepath


def copy_to_clipboard(
    results: list["PrivateIPResult"],
    lang: str = "ko",
) -> bool:
    """
    Copy results to clipboard as TSV.

    Args:
        results: List of search results
        lang: Language for headers ("ko" or "en")

    Returns:
        True if successful, False otherwise
    """
    try:
        import pyperclip
    except ImportError:
        return False

    if not results:
        return False

    headers = get_headers(lang)
    lines = ["\t".join(headers)]

    for r in results:
        row = result_to_row(r)
        lines.append("\t".join(str(v) for v in row))

    pyperclip.copy("\n".join(lines))
    return True


def copy_to_clipboard_simple(
    results: list["PrivateIPResult"],
    lang: str = "ko",
) -> bool:
    """
    Copy simplified results to clipboard (IP, Resource, ENI, Region only).

    Args:
        results: List of search results
        lang: Language for headers ("ko" or "en")

    Returns:
        True if successful, False otherwise
    """
    try:
        import pyperclip
    except ImportError:
        return False

    if not results:
        return False

    if lang == "en":
        headers = ["IP Address", "Resource", "ENI ID", "Region", "Account"]
    else:
        headers = ["IP 주소", "리소스", "ENI ID", "리전", "계정"]

    lines = ["\t".join(headers)]

    for r in results:
        row = [
            r.ip_address,
            r.mapped_resource or r.interface_type,
            r.eni_id,
            r.region,
            r.account_name or r.account_id,
        ]
        lines.append("\t".join(str(v) for v in row))

    pyperclip.copy("\n".join(lines))
    return True
