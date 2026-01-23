"""
plugins/backup/comprehensive.py - 통합 백업 현황 분석

AWS Backup + 각 서비스별 자체 자동 백업 현황을 통합 조회합니다.

지원 서비스:
- AWS Backup (Backup Jobs)
- RDS/Aurora (자동 백업, 스냅샷)
- DynamoDB (PITR)
- EFS (자동 백업 정책)
- FSx (자동 백업)
- EC2 (AWS Backup 보호 여부)
- DocumentDB (자동 백업)
- Neptune (자동 백업)
- Redshift (Provisioned + Serverless)
- ElastiCache (Redis/Valkey 스냅샷, Memcached 미지원)
- MemoryDB (스냅샷)
- OpenSearch (자동 스냅샷 - 항상 활성화)

추가 기능:
- Backup Plan의 태그 기반 선택 조건 분석
- 리소스별 백업 태그 현황 조사 및 누락 리소스 식별

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from typing import Any

from botocore.exceptions import ClientError
from rich.console import Console

from core.parallel import get_client, parallel_collect

console = Console()

# 최근 작업 조회 기간
JOB_DAYS = 30

REQUIRED_PERMISSIONS = {
    "read": [
        # AWS Backup
        "backup:ListBackupJobs",
        "backup:ListBackupPlans",
        "backup:ListBackupSelections",
        "backup:GetBackupSelection",
        "backup:ListProtectedResources",
        # RDS
        "rds:DescribeDBInstances",
        "rds:DescribeDBClusters",
        "rds:DescribeDBSnapshots",
        "rds:DescribeDBClusterSnapshots",
        # DynamoDB
        "dynamodb:ListTables",
        "dynamodb:DescribeContinuousBackups",
        # EFS
        "elasticfilesystem:DescribeFileSystems",
        "elasticfilesystem:DescribeBackupPolicy",
        # FSx
        "fsx:DescribeFileSystems",
        "fsx:DescribeBackups",
        # EC2
        "ec2:DescribeInstances",
        # DocumentDB
        "docdb:DescribeDBClusters",
        # Neptune
        "neptune:DescribeDBClusters",
        # Redshift (Provisioned + Serverless)
        "redshift:DescribeClusters",
        "redshift-serverless:ListNamespaces",
        # ElastiCache
        "elasticache:DescribeReplicationGroups",
        "elasticache:DescribeCacheClusters",
        # MemoryDB
        "memorydb:DescribeClusters",
        # OpenSearch
        "es:ListDomainNames",
        "es:DescribeDomain",
        # 태그 조회 (모든 리소스)
        "tag:GetResources",
    ],
}


# ============================================================================
# 데이터 클래스 정의
# ============================================================================


@dataclass
class BackupStatus:
    """개별 리소스의 백업 상태"""

    account_id: str
    account_name: str
    region: str
    service: str  # rds, dynamodb, efs, fsx, aws-backup
    resource_type: str  # db-instance, db-cluster, table, filesystem 등
    resource_id: str
    resource_name: str
    backup_enabled: bool  # 서비스 자체 자동 백업 활성화 여부
    backup_method: str  # automated, pitr, aws-backup
    last_backup_time: datetime | None  # 자체 백업 마지막 시간
    backup_retention_days: int
    status: str  # OK, WARNING, FAILED, DISABLED
    message: str  # 상세 메시지
    resource_arn: str = ""
    aws_backup_protected: bool = False  # AWS Backup으로 보호되는지 여부
    aws_backup_last_time: datetime | None = None  # AWS Backup 마지막 백업 시간
    has_native_backup: bool = True  # 해당 서비스가 자체 백업 기능을 가지는지 (EC2는 False)
    backup_plan_names: list[str] = field(default_factory=list)  # 할당된 Backup Plan 이름들

    @property
    def is_ok(self) -> bool:
        return self.status == "OK"

    @property
    def is_disabled(self) -> bool:
        return self.status == "DISABLED"

    @property
    def protection_summary(self) -> str:
        """보호 상태 요약"""
        if not self.has_native_backup:
            # EC2 등 자체 백업 기능 없는 서비스
            return "AWS Backup" if self.aws_backup_protected else "미보호"
        else:
            # 자체 백업 기능 있는 서비스 (RDS, DynamoDB 등)
            if self.backup_enabled and self.aws_backup_protected:
                return "자체+AWS Backup"
            elif self.aws_backup_protected:
                return "AWS Backup만"
            elif self.backup_enabled:
                return "자체 백업만"
            else:
                return "미보호"

    @property
    def is_fully_protected(self) -> bool:
        """완전 보호 여부 (서비스 특성 고려)"""
        if not self.has_native_backup:
            # 자체 백업 없는 서비스: AWS Backup만 있으면 완전 보호
            return self.aws_backup_protected
        else:
            # 자체 백업 있는 서비스: 둘 다 있어야 완전 보호
            return self.backup_enabled and self.aws_backup_protected


@dataclass
class ServiceBackupSummary:
    """서비스별 백업 요약"""

    service: str
    total_resources: int
    backup_enabled: int
    backup_disabled: int
    failed_count: int
    warning_count: int


@dataclass
class BackupPlanTagCondition:
    """Backup Plan의 태그 선택 조건"""

    account_id: str
    account_name: str
    region: str
    plan_id: str
    plan_name: str
    selection_id: str
    selection_name: str
    condition_type: str  # STRINGEQUALS, STRINGLIKE 등
    tag_key: str
    tag_value: str
    resource_types: list[str] = field(default_factory=list)  # 대상 리소스 타입


@dataclass
class FailedBackupJob:
    """실패한 AWS Backup 작업"""

    account_id: str
    account_name: str
    region: str
    job_id: str
    vault_name: str
    resource_arn: str
    resource_type: str
    status: str
    status_message: str
    creation_date: datetime | None
    completion_date: datetime | None


@dataclass
class ComprehensiveBackupResult:
    """통합 백업 분석 결과 (단일 계정/리전)"""

    account_id: str
    account_name: str
    region: str
    backup_statuses: list[BackupStatus] = field(default_factory=list)
    tag_conditions: list[BackupPlanTagCondition] = field(default_factory=list)
    failed_jobs: list[FailedBackupJob] = field(default_factory=list)

    def get_service_summary(self) -> list[ServiceBackupSummary]:
        """서비스별 요약 통계 계산"""
        summaries: dict[str, ServiceBackupSummary] = {}

        for status in self.backup_statuses:
            key = status.service
            if key not in summaries:
                summaries[key] = ServiceBackupSummary(
                    service=key,
                    total_resources=0,
                    backup_enabled=0,
                    backup_disabled=0,
                    failed_count=0,
                    warning_count=0,
                )

            s = summaries[key]
            s.total_resources += 1

            # 보호 여부: 자체 백업 OR AWS Backup 중 하나라도 있으면 보호됨
            is_protected = status.backup_enabled or status.aws_backup_protected
            if is_protected:
                s.backup_enabled += 1
            else:
                s.backup_disabled += 1

            if status.status == "FAILED":
                s.failed_count += 1
            elif status.status == "WARNING":
                s.warning_count += 1

        return list(summaries.values())

    def get_disabled_resources(self) -> list[BackupStatus]:
        """백업 비활성화된 리소스 목록"""
        return [s for s in self.backup_statuses if s.is_disabled]


# ============================================================================
# 수집 함수들
# ============================================================================


def _collect_rds_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """RDS/Aurora 백업 상태 수집"""
    results: list[BackupStatus] = []
    rds = get_client(session, "rds", region_name=region)

    # 먼저 Aurora 클러스터 정보 수집 (멤버 인스턴스 참조용)
    cluster_info: dict[str, dict] = {}  # {cluster_id: {retention, latest_restorable}}
    try:
        paginator = rds.get_paginator("describe_db_clusters")
        for page in paginator.paginate():
            for cluster in page.get("DBClusters", []):
                cluster_id = cluster.get("DBClusterIdentifier", "")
                cluster_arn = cluster.get("DBClusterArn", "")
                retention = cluster.get("BackupRetentionPeriod", 0)
                backup_enabled = retention > 0
                latest_restorable = cluster.get("LatestRestorableTime")

                cluster_info[cluster_id] = {
                    "retention": retention,
                    "latest_restorable": latest_restorable,
                }

                if backup_enabled:
                    message = f"자동 백업 활성화 (보존 {retention}일)"
                else:
                    message = "자동 백업 비활성화"

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="Aurora",
                        resource_type="DB Cluster",
                        resource_id=cluster_id,
                        resource_name=cluster_id,
                        backup_enabled=backup_enabled,
                        backup_method="automated",
                        last_backup_time=latest_restorable,
                        backup_retention_days=retention,
                        status=status,
                        message=message,
                        resource_arn=cluster_arn,
                    )
                )
    except ClientError:
        pass

    # RDS DB Instances (Aurora 클러스터 멤버 제외 - 클러스터 단위로 백업됨)
    try:
        paginator = rds.get_paginator("describe_db_instances")
        for page in paginator.paginate():
            for db in page.get("DBInstances", []):
                # Aurora 클러스터 멤버는 스킵 (클러스터에서 이미 표시)
                cluster_id = db.get("DBClusterIdentifier", "")
                if cluster_id:
                    continue

                db_id = db.get("DBInstanceIdentifier", "")
                db_arn = db.get("DBInstanceArn", "")
                retention = db.get("BackupRetentionPeriod", 0)
                latest_restorable = db.get("LatestRestorableTime")

                # 일반 RDS
                backup_enabled = retention > 0
                if backup_enabled:
                    message = f"자동 백업 활성화 (보존 {retention}일)"
                else:
                    message = "자동 백업 비활성화"

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="RDS",
                        resource_type="DB Instance",
                        resource_id=db_id,
                        resource_name=db_id,
                        backup_enabled=backup_enabled,
                        backup_method="automated",
                        last_backup_time=latest_restorable,
                        backup_retention_days=retention,
                        status=status,
                        message=message,
                        resource_arn=db_arn,
                    )
                )
    except ClientError:
        pass

    return results


def _collect_dynamodb_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """DynamoDB PITR 상태 수집"""
    results: list[BackupStatus] = []
    dynamodb = get_client(session, "dynamodb", region_name=region)

    try:
        paginator = dynamodb.get_paginator("list_tables")
        for page in paginator.paginate():
            for table_name in page.get("TableNames", []):
                try:
                    backup_info = dynamodb.describe_continuous_backups(TableName=table_name)
                    cb_desc = backup_info.get("ContinuousBackupsDescription", {})
                    cb_status = cb_desc.get("ContinuousBackupsStatus", "DISABLED")
                    pitr_desc = cb_desc.get("PointInTimeRecoveryDescription", {})
                    pitr_status = pitr_desc.get("PointInTimeRecoveryStatus", "DISABLED")

                    backup_enabled = pitr_status == "ENABLED"
                    earliest_restorable = pitr_desc.get("EarliestRestorableDateTime")
                    latest_restorable = pitr_desc.get("LatestRestorableDateTime")

                    if backup_enabled:
                        message = "PITR 활성화"
                        if earliest_restorable and latest_restorable:
                            message += f" (복원 가능: {earliest_restorable.strftime('%Y-%m-%d')} ~ {latest_restorable.strftime('%Y-%m-%d')})"
                    else:
                        message = "PITR 비활성화"

                    status = "OK" if backup_enabled else "DISABLED"

                    # ARN 생성
                    table_arn = f"arn:aws:dynamodb:{region}:{account_id}:table/{table_name}"

                    results.append(
                        BackupStatus(
                            account_id=account_id,
                            account_name=account_name,
                            region=region,
                            service="DynamoDB",
                            resource_type="Table",
                            resource_id=table_name,
                            resource_name=table_name,
                            backup_enabled=backup_enabled,
                            backup_method="pitr",
                            last_backup_time=latest_restorable,
                            backup_retention_days=35 if backup_enabled else 0,  # PITR은 35일 고정
                            status=status,
                            message=message,
                            resource_arn=table_arn,
                        )
                    )
                except ClientError:
                    # 테이블 접근 권한 없음
                    pass
    except ClientError:
        pass

    return results


def _collect_efs_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """EFS 백업 정책 상태 수집"""
    results: list[BackupStatus] = []
    efs = get_client(session, "efs", region_name=region)

    try:
        paginator = efs.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page.get("FileSystems", []):
                fs_id = fs.get("FileSystemId", "")
                fs_arn = fs.get("FileSystemArn", "")
                name = fs.get("Name", fs_id)

                # 백업 정책 조회
                backup_enabled = False
                message = "백업 정책 조회 실패"
                try:
                    backup_policy = efs.describe_backup_policy(FileSystemId=fs_id)
                    policy_status = backup_policy.get("BackupPolicy", {}).get("Status", "DISABLED")
                    backup_enabled = policy_status == "ENABLED"
                    message = "자동 백업 활성화" if backup_enabled else "자동 백업 비활성화"
                except ClientError as e:
                    if e.response.get("Error", {}).get("Code") == "PolicyNotFound":
                        message = "백업 정책 없음"
                    # 기타 에러는 무시

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="EFS",
                        resource_type="FileSystem",
                        resource_id=fs_id,
                        resource_name=name,
                        backup_enabled=backup_enabled,
                        backup_method="aws-backup",
                        last_backup_time=None,
                        backup_retention_days=0,
                        status=status,
                        message=message,
                        resource_arn=fs_arn,
                        has_native_backup=False,  # EFS는 네이티브 스냅샷 없음, AWS Backup만 가능
                    )
                )
    except ClientError:
        pass

    return results


def _collect_fsx_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """FSx 자동 백업 상태 수집"""
    results: list[BackupStatus] = []
    fsx = get_client(session, "fsx", region_name=region)

    # 먼저 모든 백업 조회하여 파일시스템별 최신 백업 시간 추출
    # - Windows/Lustre/OpenZFS: FileSystem 레벨 백업
    # - ONTAP: Volume 레벨 백업 (Volume.FileSystemId로 매핑)
    fs_latest_backup: dict[str, datetime] = {}
    try:
        backup_paginator = fsx.get_paginator("describe_backups")
        for page in backup_paginator.paginate():
            for backup in page.get("Backups", []):
                backup_time = backup.get("CreationTime")
                if not backup_time:
                    continue

                # FileSystem 레벨 백업 (Windows, Lustre, OpenZFS)
                backup_fs_id = backup.get("FileSystem", {}).get("FileSystemId", "")

                # Volume 레벨 백업 (ONTAP) - Volume에서 FileSystemId 추출
                if not backup_fs_id:
                    backup_fs_id = backup.get("Volume", {}).get("FileSystemId", "")

                if backup_fs_id:
                    if backup_fs_id not in fs_latest_backup:
                        fs_latest_backup[backup_fs_id] = backup_time
                    else:
                        fs_latest_backup[backup_fs_id] = max(fs_latest_backup[backup_fs_id], backup_time)
    except ClientError:
        pass

    try:
        paginator = fsx.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page.get("FileSystems", []):
                fs_id = fs.get("FileSystemId", "")
                fs_arn = fs.get("ResourceARN", "")
                fs_type = fs.get("FileSystemType", "")

                # 이름은 태그에서 가져오기
                tags = fs.get("Tags", [])
                name = next((t["Value"] for t in tags if t["Key"] == "Name"), fs_id)

                # FSx 타입별 백업 설정 확인
                retention_days = 0

                if fs_type == "LUSTRE":
                    lustre_config = fs.get("LustreConfiguration", {})
                    retention_days = lustre_config.get("AutomaticBackupRetentionDays", 0)
                elif fs_type == "WINDOWS":
                    windows_config = fs.get("WindowsConfiguration", {})
                    retention_days = windows_config.get("AutomaticBackupRetentionDays", 0)
                elif fs_type == "ONTAP":
                    ontap_config = fs.get("OntapConfiguration", {})
                    retention_days = ontap_config.get("AutomaticBackupRetentionDays", 0)
                elif fs_type == "OPENZFS":
                    zfs_config = fs.get("OpenZFSConfiguration", {})
                    retention_days = zfs_config.get("AutomaticBackupRetentionDays", 0)

                backup_enabled = retention_days > 0
                latest_backup_time = fs_latest_backup.get(fs_id)

                if backup_enabled:
                    message = f"자동 백업 활성화 (보존 {retention_days}일)"
                else:
                    message = "자동 백업 비활성화"

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="FSx",
                        resource_type=f"FileSystem ({fs_type})",
                        resource_id=fs_id,
                        resource_name=name,
                        backup_enabled=backup_enabled,
                        backup_method="automated",
                        last_backup_time=latest_backup_time,
                        backup_retention_days=retention_days,
                        status=status,
                        message=message,
                        resource_arn=fs_arn,
                    )
                )
    except ClientError:
        pass

    return results


def _collect_ec2_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """EC2 인스턴스 백업 상태 수집 (AWS Backup 보호 여부만 체크)"""
    results: list[BackupStatus] = []
    ec2 = get_client(session, "ec2", region_name=region)

    try:
        paginator = ec2.get_paginator("describe_instances")
        for page in paginator.paginate():
            for reservation in page.get("Reservations", []):
                for instance in reservation.get("Instances", []):
                    instance_id = instance.get("InstanceId", "")
                    instance_state = instance.get("State", {}).get("Name", "")

                    # terminated 인스턴스는 제외
                    if instance_state == "terminated":
                        continue

                    # 이름은 태그에서 가져오기
                    tags = instance.get("Tags", [])
                    name = next((t["Value"] for t in tags if t["Key"] == "Name"), instance_id)

                    # ARN 생성
                    instance_arn = f"arn:aws:ec2:{region}:{account_id}:instance/{instance_id}"

                    # EC2는 자체 백업 기능 없음 (AWS Backup으로만 보호)
                    # aws_backup_protected는 나중에 매핑됨
                    message = f"상태: {instance_state}"

                    results.append(
                        BackupStatus(
                            account_id=account_id,
                            account_name=account_name,
                            region=region,
                            service="EC2",
                            resource_type="Instance",
                            resource_id=instance_id,
                            resource_name=name,
                            backup_enabled=False,  # EC2는 자체 백업 없음
                            backup_method="aws-backup",
                            last_backup_time=None,
                            backup_retention_days=0,
                            status="DISABLED",  # AWS Backup 매핑 후 업데이트됨
                            message=message,
                            resource_arn=instance_arn,
                            has_native_backup=False,  # EC2는 자체 백업 기능 없음
                        )
                    )
    except ClientError:
        pass

    return results


def _collect_documentdb_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """DocumentDB 클러스터 백업 상태 수집"""
    results: list[BackupStatus] = []

    # DocumentDB는 RDS API를 사용하지만 별도 엔드포인트
    # docdb 서비스로 호출
    try:
        docdb = get_client(session, "docdb", region_name=region)
        paginator = docdb.get_paginator("describe_db_clusters")

        for page in paginator.paginate():
            for cluster in page.get("DBClusters", []):
                # DocumentDB 클러스터만 필터링 (Engine이 docdb)
                if cluster.get("Engine") != "docdb":
                    continue

                cluster_id = cluster.get("DBClusterIdentifier", "")
                cluster_arn = cluster.get("DBClusterArn", "")
                retention = cluster.get("BackupRetentionPeriod", 0)
                backup_enabled = retention > 0
                latest_restorable = cluster.get("LatestRestorableTime")

                if backup_enabled:
                    message = f"자동 백업 활성화 (보존 {retention}일)"
                else:
                    message = "자동 백업 비활성화"

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="DocumentDB",
                        resource_type="Cluster",
                        resource_id=cluster_id,
                        resource_name=cluster_id,
                        backup_enabled=backup_enabled,
                        backup_method="automated",
                        last_backup_time=latest_restorable,
                        backup_retention_days=retention,
                        status=status,
                        message=message,
                        resource_arn=cluster_arn,
                    )
                )
    except ClientError:
        pass

    return results


def _collect_neptune_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """Neptune 클러스터 백업 상태 수집"""
    results: list[BackupStatus] = []

    try:
        neptune = get_client(session, "neptune", region_name=region)
        paginator = neptune.get_paginator("describe_db_clusters")

        for page in paginator.paginate():
            for cluster in page.get("DBClusters", []):
                # Neptune 클러스터만 필터링 (Engine이 neptune)
                if cluster.get("Engine") != "neptune":
                    continue

                cluster_id = cluster.get("DBClusterIdentifier", "")
                cluster_arn = cluster.get("DBClusterArn", "")
                retention = cluster.get("BackupRetentionPeriod", 0)
                backup_enabled = retention > 0
                latest_restorable = cluster.get("LatestRestorableTime")

                if backup_enabled:
                    message = f"자동 백업 활성화 (보존 {retention}일)"
                else:
                    message = "자동 백업 비활성화"

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="Neptune",
                        resource_type="Cluster",
                        resource_id=cluster_id,
                        resource_name=cluster_id,
                        backup_enabled=backup_enabled,
                        backup_method="automated",
                        last_backup_time=latest_restorable,
                        backup_retention_days=retention,
                        status=status,
                        message=message,
                        resource_arn=cluster_arn,
                    )
                )
    except ClientError:
        pass

    return results


def _collect_redshift_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """Redshift 백업 상태 수집 (Provisioned + Serverless)"""
    results: list[BackupStatus] = []

    # 1. Provisioned Clusters
    try:
        redshift = get_client(session, "redshift", region_name=region)
        paginator = redshift.get_paginator("describe_clusters")
        for page in paginator.paginate():
            for cluster in page.get("Clusters", []):
                cluster_id = cluster.get("ClusterIdentifier", "")
                # ARN 생성
                cluster_arn = f"arn:aws:redshift:{region}:{account_id}:cluster:{cluster_id}"
                retention = cluster.get("AutomatedSnapshotRetentionPeriod", 0)
                backup_enabled = retention > 0

                if backup_enabled:
                    message = f"자동 스냅샷 활성화 (보존 {retention}일)"
                else:
                    message = "자동 스냅샷 비활성화"

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="Redshift",
                        resource_type="Cluster (Provisioned)",
                        resource_id=cluster_id,
                        resource_name=cluster_id,
                        backup_enabled=backup_enabled,
                        backup_method="snapshot",
                        last_backup_time=None,
                        backup_retention_days=retention,
                        status=status,
                        message=message,
                        resource_arn=cluster_arn,
                    )
                )
    except ClientError:
        pass

    # 2. Serverless Namespaces (스냅샷 기본 활성화)
    try:
        rs_serverless = get_client(session, "redshift-serverless", region_name=region)
        namespaces_resp = rs_serverless.list_namespaces()
        for ns in namespaces_resp.get("namespaces", []):
            ns_name = ns.get("namespaceName", "")
            ns_arn = ns.get("namespaceArn", "")
            ns_status = ns.get("status", "")

            # Serverless는 스냅샷이 기본 활성화됨 (recovery point 자동 생성)
            # 실제로 recovery points가 있는지 확인
            backup_enabled = True
            message = "Serverless 스냅샷 활성화 (기본)"

            # 상태가 AVAILABLE이 아니면 체크
            if ns_status != "AVAILABLE":
                message = f"Namespace 상태: {ns_status}"

            results.append(
                BackupStatus(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    service="Redshift Serverless",
                    resource_type="Namespace",
                    resource_id=ns_name,
                    resource_name=ns_name,
                    backup_enabled=backup_enabled,
                    backup_method="snapshot",
                    last_backup_time=None,
                    backup_retention_days=0,  # Serverless는 별도 설정
                    status="OK",
                    message=message,
                    resource_arn=ns_arn,
                )
            )
    except ClientError:
        pass  # Serverless 미사용 또는 권한 없음

    return results


def _collect_elasticache_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """ElastiCache 백업 상태 수집 (Redis/Valkey + Memcached)"""
    results: list[BackupStatus] = []
    elasticache = get_client(session, "elasticache", region_name=region)

    # Redis/Valkey 클러스터 ID 추적 (중복 방지)
    redis_cluster_ids: set[str] = set()

    # 1. Redis/Valkey Replication Groups (스냅샷 지원)
    try:
        paginator = elasticache.get_paginator("describe_replication_groups")
        for page in paginator.paginate():
            for group in page.get("ReplicationGroups", []):
                group_id = group.get("ReplicationGroupId", "")
                group_arn = group.get("ARN", "")
                retention = group.get("SnapshotRetentionLimit", 0)
                backup_enabled = retention > 0

                # 멤버 클러스터 추적
                for member in group.get("MemberClusters", []):
                    redis_cluster_ids.add(member)

                if backup_enabled:
                    message = f"자동 스냅샷 활성화 (보존 {retention}일)"
                else:
                    message = "자동 스냅샷 비활성화"

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="ElastiCache",
                        resource_type="Replication Group (Redis/Valkey)",
                        resource_id=group_id,
                        resource_name=group_id,
                        backup_enabled=backup_enabled,
                        backup_method="snapshot",
                        last_backup_time=None,
                        backup_retention_days=retention,
                        status=status,
                        message=message,
                        resource_arn=group_arn,
                    )
                )
    except ClientError:
        pass

    # 2. Memcached 클러스터 (스냅샷 미지원)
    try:
        paginator = elasticache.get_paginator("describe_cache_clusters")
        for page in paginator.paginate():
            for cluster in page.get("CacheClusters", []):
                cluster_id = cluster.get("CacheClusterId", "")
                engine = cluster.get("Engine", "")

                # Replication Group 멤버인 Redis 클러스터는 스킵 (이미 위에서 처리)
                if cluster_id in redis_cluster_ids:
                    continue

                cluster_arn = cluster.get("ARN", "")

                if engine == "memcached":
                    # Memcached는 스냅샷 미지원
                    results.append(
                        BackupStatus(
                            account_id=account_id,
                            account_name=account_name,
                            region=region,
                            service="ElastiCache",
                            resource_type="Cluster (Memcached)",
                            resource_id=cluster_id,
                            resource_name=cluster_id,
                            backup_enabled=False,
                            backup_method="not_supported",
                            last_backup_time=None,
                            backup_retention_days=0,
                            status="NOT_SUPPORTED",
                            message="Memcached는 스냅샷 미지원",
                            resource_arn=cluster_arn,
                        )
                    )
                elif engine in ("redis", "valkey"):
                    # 단독 Redis/Valkey 클러스터 (Replication Group 없음)
                    retention = cluster.get("SnapshotRetentionLimit", 0)
                    backup_enabled = retention > 0

                    if backup_enabled:
                        message = f"자동 스냅샷 활성화 (보존 {retention}일)"
                    else:
                        message = "자동 스냅샷 비활성화"

                    status = "OK" if backup_enabled else "DISABLED"

                    results.append(
                        BackupStatus(
                            account_id=account_id,
                            account_name=account_name,
                            region=region,
                            service="ElastiCache",
                            resource_type=f"Cluster ({engine.capitalize()})",
                            resource_id=cluster_id,
                            resource_name=cluster_id,
                            backup_enabled=backup_enabled,
                            backup_method="snapshot",
                            last_backup_time=None,
                            backup_retention_days=retention,
                            status=status,
                            message=message,
                            resource_arn=cluster_arn,
                        )
                    )
    except ClientError:
        pass

    return results


def _collect_memorydb_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """MemoryDB 백업 상태 수집"""
    results: list[BackupStatus] = []

    try:
        memorydb = get_client(session, "memorydb", region_name=region)
        paginator = memorydb.get_paginator("describe_clusters")
        for page in paginator.paginate():
            for cluster in page.get("Clusters", []):
                cluster_name = cluster.get("Name", "")
                cluster_arn = cluster.get("ARN", "")
                retention = cluster.get("SnapshotRetentionLimit", 0)
                backup_enabled = retention > 0

                if backup_enabled:
                    message = f"자동 스냅샷 활성화 (보존 {retention}일)"
                else:
                    message = "자동 스냅샷 비활성화"

                status = "OK" if backup_enabled else "DISABLED"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="MemoryDB",
                        resource_type="Cluster",
                        resource_id=cluster_name,
                        resource_name=cluster_name,
                        backup_enabled=backup_enabled,
                        backup_method="snapshot",
                        last_backup_time=None,
                        backup_retention_days=retention,
                        status=status,
                        message=message,
                        resource_arn=cluster_arn,
                    )
                )
    except ClientError:
        pass

    return results


def _collect_opensearch_backup_status(
    session, account_id: str, account_name: str, region: str
) -> list[BackupStatus]:
    """OpenSearch 백업 상태 수집 (자동 스냅샷 항상 활성화)"""
    results: list[BackupStatus] = []

    try:
        opensearch = get_client(session, "opensearch", region_name=region)
        domains_resp = opensearch.list_domain_names()

        for domain_info in domains_resp.get("DomainNames", []):
            domain_name = domain_info.get("DomainName", "")

            try:
                domain_resp = opensearch.describe_domain(DomainName=domain_name)
                domain = domain_resp.get("DomainStatus", {})

                domain_arn = domain.get("ARN", "")
                snapshot_options = domain.get("SnapshotOptions", {})
                auto_hour = snapshot_options.get("AutomatedSnapshotStartHour")

                # OpenSearch는 자동 스냅샷이 항상 활성화 (14일 보존)
                # 시작 시간만 설정 가능
                if auto_hour is not None:
                    message = f"자동 스냅샷 활성화 (14일 보존, 시작: {auto_hour}:00 UTC)"
                else:
                    message = "자동 스냅샷 활성화 (14일 보존)"

                results.append(
                    BackupStatus(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        service="OpenSearch",
                        resource_type="Domain",
                        resource_id=domain_name,
                        resource_name=domain_name,
                        backup_enabled=True,  # 항상 활성화
                        backup_method="snapshot",
                        last_backup_time=None,
                        backup_retention_days=14,  # 고정
                        status="ALWAYS_ON",
                        message=message,
                        resource_arn=domain_arn,
                    )
                )
            except ClientError:
                pass  # 도메인 상세 조회 실패
    except ClientError:
        pass

    return results


def _collect_aws_backup_protected_resources(
    session, account_id: str, account_name: str, region: str
) -> dict[str, datetime | None]:
    """AWS Backup으로 보호되는 리소스 ARN과 마지막 백업 시간 수집

    Returns:
        {resource_arn: last_backup_time} - 마지막 백업 시간 (없으면 None)
    """
    protected_resources: dict[str, datetime | None] = {}
    backup = get_client(session, "backup", region_name=region)

    try:
        paginator = backup.get_paginator("list_protected_resources")
        for page in paginator.paginate():
            for resource in page.get("Results", []):
                arn = resource.get("ResourceArn", "")
                last_backup = resource.get("LastBackupTime")  # datetime or None
                if arn:
                    protected_resources[arn] = last_backup
    except ClientError:
        pass

    return protected_resources


def _collect_resource_backup_plan_mapping(
    session, account_id: str, account_name: str, region: str
) -> dict[str, set[str]]:
    """리소스별 할당된 Backup Plan 이름 수집

    최근 완료된 Backup Job에서 리소스와 Plan 매핑을 추출합니다.

    Returns:
        {resource_arn: {plan_name1, plan_name2, ...}}
    """
    resource_to_plans: dict[str, set[str]] = {}
    backup = get_client(session, "backup", region_name=region)

    # 1. Backup Plan ID -> Name 매핑 생성
    plan_id_to_name: dict[str, str] = {}
    try:
        plans_paginator = backup.get_paginator("list_backup_plans")
        for page in plans_paginator.paginate():
            for plan in page.get("BackupPlansList", []):
                plan_id = plan.get("BackupPlanId", "")
                plan_name = plan.get("BackupPlanName", "")
                if plan_id and plan_name:
                    plan_id_to_name[plan_id] = plan_name
    except ClientError:
        pass

    if not plan_id_to_name:
        return resource_to_plans

    # 2. 최근 완료된 Backup Job에서 리소스 -> Plan 매핑 추출
    now = datetime.now(timezone.utc)
    start_date = now - timedelta(days=JOB_DAYS)

    try:
        paginator = backup.get_paginator("list_backup_jobs")
        for page in paginator.paginate(
            ByCreatedAfter=start_date,
            ByCreatedBefore=now,
            ByState="COMPLETED",
        ):
            for job in page.get("BackupJobs", []):
                resource_arn = job.get("ResourceArn", "")
                plan_id = job.get("BackupPlanId", "")

                if resource_arn and plan_id and plan_id in plan_id_to_name:
                    plan_name = plan_id_to_name[plan_id]
                    if resource_arn not in resource_to_plans:
                        resource_to_plans[resource_arn] = set()
                    resource_to_plans[resource_arn].add(plan_name)
    except ClientError:
        pass

    return resource_to_plans


def _collect_backup_plan_tag_conditions(
    session, account_id: str, account_name: str, region: str
) -> list[BackupPlanTagCondition]:
    """Backup Plan의 태그 기반 선택 조건 수집

    AWS 자동 생성 Plan은 제외:
    - aws/efs/automatic-backup-plan (EFS 자동 백업용)
    """
    results: list[BackupPlanTagCondition] = []
    backup = get_client(session, "backup", region_name=region)

    # AWS 관리형 Plan 이름 패턴 (제외 대상)
    aws_managed_plan_prefixes = ("aws/efs/", "aws/")

    try:
        # Backup Plans 조회
        plans_paginator = backup.get_paginator("list_backup_plans")
        for plans_page in plans_paginator.paginate():
            for plan in plans_page.get("BackupPlansList", []):
                plan_id = plan.get("BackupPlanId", "")
                plan_name = plan.get("BackupPlanName", "")

                # AWS 관리형 Plan 스킵
                if plan_name.startswith(aws_managed_plan_prefixes):
                    continue

                # 각 Plan의 Selection 조회
                try:
                    selections = backup.list_backup_selections(BackupPlanId=plan_id)
                    for sel in selections.get("BackupSelectionsList", []):
                        sel_id = sel.get("SelectionId", "")
                        sel_name = sel.get("SelectionName", "")

                        # Selection 상세 조회
                        try:
                            sel_detail = backup.get_backup_selection(
                                BackupPlanId=plan_id, SelectionId=sel_id
                            )
                            selection = sel_detail.get("BackupSelection", {})
                            list_of_tags = selection.get("ListOfTags", [])
                            resources = selection.get("Resources", [])
                            conditions = selection.get("Conditions", {})

                            # 리소스 타입 추출 (ARN에서)
                            resource_types = []
                            for r in resources:
                                if "::" in r:
                                    # arn:aws:ec2:*:*:volume/* 형태
                                    parts = r.split(":")
                                    if len(parts) >= 6:
                                        resource_types.append(parts[2])  # 서비스명

                            # ListOfTags에서 조건 추출
                            for tag_cond in list_of_tags:
                                cond_type = tag_cond.get("ConditionType", "STRINGEQUALS")
                                cond_key = tag_cond.get("ConditionKey", "")
                                cond_value = tag_cond.get("ConditionValue", "")

                                results.append(
                                    BackupPlanTagCondition(
                                        account_id=account_id,
                                        account_name=account_name,
                                        region=region,
                                        plan_id=plan_id,
                                        plan_name=plan_name,
                                        selection_id=sel_id,
                                        selection_name=sel_name,
                                        condition_type=cond_type,
                                        tag_key=cond_key,
                                        tag_value=cond_value,
                                        resource_types=resource_types,
                                    )
                                )

                            # Conditions.StringEquals에서도 추출
                            for str_eq in conditions.get("StringEquals", []):
                                results.append(
                                    BackupPlanTagCondition(
                                        account_id=account_id,
                                        account_name=account_name,
                                        region=region,
                                        plan_id=plan_id,
                                        plan_name=plan_name,
                                        selection_id=sel_id,
                                        selection_name=sel_name,
                                        condition_type="STRINGEQUALS",
                                        tag_key=str_eq.get("ConditionKey", ""),
                                        tag_value=str_eq.get("ConditionValue", ""),
                                        resource_types=resource_types,
                                    )
                                )

                            # Conditions.StringLike에서도 추출
                            for str_like in conditions.get("StringLike", []):
                                results.append(
                                    BackupPlanTagCondition(
                                        account_id=account_id,
                                        account_name=account_name,
                                        region=region,
                                        plan_id=plan_id,
                                        plan_name=plan_name,
                                        selection_id=sel_id,
                                        selection_name=sel_name,
                                        condition_type="STRINGLIKE",
                                        tag_key=str_like.get("ConditionKey", ""),
                                        tag_value=str_like.get("ConditionValue", ""),
                                        resource_types=resource_types,
                                    )
                                )

                        except ClientError:
                            pass
                except ClientError:
                    pass
    except ClientError:
        pass

    return results


def _collect_failed_backup_jobs(
    session, account_id: str, account_name: str, region: str
) -> list[FailedBackupJob]:
    """실패한 AWS Backup 작업 수집"""
    results: list[FailedBackupJob] = []
    backup = get_client(session, "backup", region_name=region)

    now = datetime.now(timezone.utc)
    start_date = now - timedelta(days=JOB_DAYS)

    try:
        paginator = backup.get_paginator("list_backup_jobs")

        # 실패 상태별로 조회
        for state in ["FAILED", "ABORTED", "PARTIAL"]:
            try:
                for page in paginator.paginate(
                    ByCreatedAfter=start_date,
                    ByCreatedBefore=now,
                    ByState=state,
                ):
                    for job in page.get("BackupJobs", []):
                        results.append(
                            FailedBackupJob(
                                account_id=account_id,
                                account_name=account_name,
                                region=region,
                                job_id=job.get("BackupJobId", ""),
                                vault_name=job.get("BackupVaultName", ""),
                                resource_arn=job.get("ResourceArn", ""),
                                resource_type=job.get("ResourceType", ""),
                                status=job.get("State", ""),
                                status_message=job.get("StatusMessage", ""),
                                creation_date=job.get("CreationDate"),
                                completion_date=job.get("CompletionDate"),
                            )
                        )
            except ClientError:
                pass
    except ClientError:
        pass

    return results


def _collect_resource_tags_by_arn(
    session, region: str, resource_arns: list[str], tag_keys: set[str]
) -> dict[str, dict[str, str]]:
    """리소스 ARN별 태그 조회 (필요한 태그 키만)

    Returns:
        {resource_arn: {tag_key: tag_value}}
    """
    if not resource_arns or not tag_keys:
        return {}

    result: dict[str, dict[str, str]] = {}
    tagging = get_client(session, "resourcegroupstaggingapi", region_name=region)

    try:
        paginator = tagging.get_paginator("get_resources")
        for page in paginator.paginate(
            ResourceTypeFilters=[
                "rds:db",
                "rds:cluster",
                "dynamodb:table",
                "elasticfilesystem:file-system",
                "fsx:file-system",
                "ec2:instance",
                "docdb:cluster",
                "neptune:cluster",
                "redshift:cluster",
                "elasticache:cluster",
                "elasticache:replicationgroup",
                "memorydb:cluster",
                "es:domain",
            ]
        ):
            for resource in page.get("ResourceTagMappingList", []):
                arn = resource.get("ResourceARN", "")
                if arn in resource_arns:
                    tags = {t["Key"]: t["Value"] for t in resource.get("Tags", [])}
                    # 필요한 태그 키만 필터링
                    filtered_tags = {k: v for k, v in tags.items() if k in tag_keys}
                    if filtered_tags:
                        result[arn] = filtered_tags
    except ClientError:
        pass

    return result


def _collect_comprehensive_backup_data(
    session, account_id: str, account_name: str, region: str
) -> ComprehensiveBackupResult | None:
    """단일 계정/리전의 통합 백업 데이터 수집"""
    result = ComprehensiveBackupResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    # AWS Backup으로 보호되는 리소스 목록 수집
    protected_resources = _collect_aws_backup_protected_resources(session, account_id, account_name, region)

    # 리소스별 Backup Plan 매핑 수집
    resource_plan_mapping = _collect_resource_backup_plan_mapping(session, account_id, account_name, region)

    # 각 서비스별 백업 상태 수집
    result.backup_statuses.extend(_collect_rds_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_dynamodb_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_efs_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_fsx_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_ec2_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_documentdb_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_neptune_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_redshift_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_elasticache_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_memorydb_backup_status(session, account_id, account_name, region))
    result.backup_statuses.extend(_collect_opensearch_backup_status(session, account_id, account_name, region))

    # Backup Plan 태그 조건 수집
    result.tag_conditions = _collect_backup_plan_tag_conditions(session, account_id, account_name, region)

    # AWS Backup 보호 여부, 마지막 백업 시간, Backup Plan 매핑
    for status in result.backup_statuses:
        if status.resource_arn:
            # Backup Plan 이름 매핑
            if status.resource_arn in resource_plan_mapping:
                status.backup_plan_names = sorted(resource_plan_mapping[status.resource_arn])

            # AWS Backup 보호 여부 및 마지막 백업 시간
            if status.resource_arn in protected_resources:
                status.aws_backup_protected = True
                aws_backup_time = protected_resources[status.resource_arn]

                # AWS Backup 마지막 백업 시간 저장
                status.aws_backup_last_time = aws_backup_time

                # 메시지 업데이트 (자체 백업 유무에 따라 다르게 처리)
                if status.has_native_backup:
                    # 자체 백업 기능 있는 서비스 (RDS, DynamoDB, FSx 등)
                    if status.backup_enabled:
                        # 자체 백업 + AWS Backup 둘 다 있음
                        status.message += " + AWS Backup"
                        status.backup_method = f"{status.backup_method}+aws-backup"
                    else:
                        # AWS Backup만
                        status.status = "OK"
                        status.message += " (AWS Backup으로 보호됨)"
                        status.backup_method = "aws-backup"
                else:
                    # 자체 백업 기능 없는 서비스 (EFS, EC2)
                    # backup_enabled가 True여도 그것도 AWS Backup임
                    status.status = "OK"
                    if not status.backup_enabled:
                        status.message += " (AWS Backup으로 보호됨)"
                    # backup_method는 이미 "aws-backup"이므로 변경 불필요

    # 실패한 작업 수집
    result.failed_jobs = _collect_failed_backup_jobs(session, account_id, account_name, region)

    # 데이터가 하나라도 있으면 반환
    if result.backup_statuses or result.tag_conditions or result.failed_jobs:
        return result
    return None


# ============================================================================
# 보고서 생성
# ============================================================================


def generate_report(results: list[ComprehensiveBackupResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 스타일
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    green_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")

    # ========== Summary 시트 ==========
    # 서비스별 요약
    all_summaries: dict[str, ServiceBackupSummary] = {}
    for r in results:
        for s in r.get_service_summary():
            if s.service not in all_summaries:
                all_summaries[s.service] = ServiceBackupSummary(
                    service=s.service,
                    total_resources=0,
                    backup_enabled=0,
                    backup_disabled=0,
                    failed_count=0,
                    warning_count=0,
                )
            ss = all_summaries[s.service]
            ss.total_resources += s.total_resources
            ss.backup_enabled += s.backup_enabled
            ss.backup_disabled += s.backup_disabled
            ss.failed_count += s.failed_count
            ss.warning_count += s.warning_count

    summary = wb.new_summary_sheet("Summary")
    summary.add_title("통합 백업 현황 보고서")
    summary.add_item("생성일", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary.add_blank_row()
    summary.add_section("서비스별 요약")

    for service in [
        "RDS",
        "Aurora",
        "DynamoDB",
        "EFS",
        "FSx",
        "EC2",
        "DocumentDB",
        "Neptune",
        "Redshift",
        "Redshift Serverless",
        "ElastiCache",
        "MemoryDB",
        "OpenSearch",
    ]:
        if service in all_summaries:
            s = all_summaries[service]
            disabled_rate = (s.backup_disabled / s.total_resources * 100) if s.total_resources > 0 else 0
            highlight = "warning" if s.backup_disabled > 0 else None
            summary.add_item(
                f"{s.service} (총 {s.total_resources}개)",
                f"보호: {s.backup_enabled}, 미보호: {s.backup_disabled} ({disabled_rate:.1f}%)",
                highlight=highlight,
            )

    # 전체 통계
    total_failed_jobs = sum(len(r.failed_jobs) for r in results)
    summary.add_blank_row()
    highlight = "danger" if total_failed_jobs > 0 else None
    summary.add_item("실패한 Backup 작업 (최근 30일)", total_failed_jobs, highlight=highlight)

    # ========== Backup Status 시트 (서비스별) ==========
    all_statuses: list[BackupStatus] = []
    for r in results:
        all_statuses.extend(r.backup_statuses)

    # 실패 이력 집계 (리소스 ARN별)
    all_failed_jobs: list[FailedBackupJob] = []
    for r in results:
        all_failed_jobs.extend(r.failed_jobs)

    # 리소스별 실패 횟수 및 최근 실패 시간
    failure_by_resource: dict[str, dict] = {}  # {arn: {count: int, last_failure: datetime}}
    for job in all_failed_jobs:
        arn = job.resource_arn
        if arn not in failure_by_resource:
            failure_by_resource[arn] = {"count": 0, "last_failure": None}
        failure_by_resource[arn]["count"] += 1
        if job.creation_date:
            if failure_by_resource[arn]["last_failure"] is None:
                failure_by_resource[arn]["last_failure"] = job.creation_date
            else:
                failure_by_resource[arn]["last_failure"] = max(
                    failure_by_resource[arn]["last_failure"], job.creation_date
                )

    if all_statuses:
        status_columns = [
            ColumnDef(header="Account", width=15),
            ColumnDef(header="Region", width=15),
            ColumnDef(header="서비스", width=15),
            ColumnDef(header="리소스 타입", width=20),
            ColumnDef(header="리소스 ID", width=25),
            ColumnDef(header="리소스 이름", width=25),
            ColumnDef(header="자체 백업", width=10, style="center"),
            ColumnDef(header="AWS Backup", width=10, style="center"),
            ColumnDef(header="보호 상태", width=15, style="center"),
            ColumnDef(header="백업 방식", width=18),
            ColumnDef(header="보존 기간(일)", width=12, style="number"),
            ColumnDef(header="자체 백업 시간", width=18),
            ColumnDef(header="AWS Backup 시간", width=18),
            ColumnDef(header="백업 지연", width=12),
            ColumnDef(header="실패 횟수", width=10, style="number"),
            ColumnDef(header="최근 실패", width=18),
            ColumnDef(header="Backup Plan", width=25),
            ColumnDef(header="상태", width=12, style="center"),
            ColumnDef(header="메시지", width=40),
        ]
        status_sheet = wb.new_sheet("Backup Status", status_columns)

        now = datetime.now(timezone.utc)
        two_days_ago = now - timedelta(days=2)

        for s in all_statuses:
            # 백업 지연 체크 (보호 중인데 마지막 백업이 2일 이상 전이면 지연)
            latest_backup = None
            if s.last_backup_time and s.aws_backup_last_time:
                latest_backup = max(s.last_backup_time, s.aws_backup_last_time)
            elif s.last_backup_time:
                latest_backup = s.last_backup_time
            elif s.aws_backup_last_time:
                latest_backup = s.aws_backup_last_time

            backup_delayed = ""
            if (s.backup_enabled or s.aws_backup_protected) and latest_backup:
                if latest_backup < two_days_ago:
                    days_delayed = (now - latest_backup).days
                    backup_delayed = f"{days_delayed}일 지연"

            # 실패 이력 (최근 30일)
            failure_info = failure_by_resource.get(s.resource_arn, {"count": 0, "last_failure": None})
            failure_count = failure_info["count"]
            last_failure = failure_info["last_failure"]

            # Backup Plan 표시
            backup_plan_str = ", ".join(s.backup_plan_names) if s.backup_plan_names else "-"

            row_num = status_sheet.add_row([
                s.account_name,
                s.region,
                s.service,
                s.resource_type,
                s.resource_id,
                s.resource_name,
                "Yes" if s.backup_enabled else "No",
                "Yes" if s.aws_backup_protected else "No",
                s.protection_summary,
                s.backup_method,
                s.backup_retention_days,
                s.last_backup_time.strftime("%Y-%m-%d %H:%M") if s.last_backup_time else "-",
                s.aws_backup_last_time.strftime("%Y-%m-%d %H:%M") if s.aws_backup_last_time else "-",
                backup_delayed,
                failure_count if failure_count > 0 else "-",
                last_failure.strftime("%Y-%m-%d %H:%M") if last_failure else "-",
                backup_plan_str,
                s.status,
                s.message,
            ])

            ws = status_sheet._ws

            if backup_delayed:
                ws.cell(row=row_num, column=14).fill = yellow_fill

            if failure_count > 0:
                ws.cell(row=row_num, column=15).fill = red_fill
            if last_failure:
                ws.cell(row=row_num, column=16).fill = red_fill

            # 보호 상태별 색상 (서비스 특성 고려)
            if s.protection_summary == "미보호":
                ws.cell(row=row_num, column=9).fill = red_fill
            elif s.is_fully_protected:
                ws.cell(row=row_num, column=9).fill = green_fill
            else:
                ws.cell(row=row_num, column=9).fill = yellow_fill

            # 상태별 색상
            if s.status == "DISABLED" and not s.aws_backup_protected:
                ws.cell(row=row_num, column=18).fill = yellow_fill
            elif s.status == "FAILED":
                ws.cell(row=row_num, column=18).fill = red_fill
            elif s.status == "OK":
                ws.cell(row=row_num, column=18).fill = green_fill

    # ========== Backup Plan Tags 시트 ==========
    all_tag_conditions: list[BackupPlanTagCondition] = []
    for r in results:
        all_tag_conditions.extend(r.tag_conditions)

    if all_tag_conditions:
        tag_columns = [
            ColumnDef(header="Account", width=15),
            ColumnDef(header="Region", width=15),
            ColumnDef(header="Plan ID", width=40),
            ColumnDef(header="Plan 이름", width=25),
            ColumnDef(header="Selection ID", width=40),
            ColumnDef(header="Selection 이름", width=25),
            ColumnDef(header="조건 타입", width=15),
            ColumnDef(header="태그 키", width=20),
            ColumnDef(header="태그 값", width=20),
            ColumnDef(header="대상 리소스 타입", width=20),
        ]
        tag_sheet = wb.new_sheet("Backup Plan Tags", tag_columns)

        for t in all_tag_conditions:
            tag_sheet.add_row([
                t.account_name,
                t.region,
                t.plan_id,
                t.plan_name,
                t.selection_id,
                t.selection_name,
                t.condition_type,
                t.tag_key,
                t.tag_value,
                ", ".join(t.resource_types) if t.resource_types else "*",
            ])

    # ========== Failed Jobs 시트 ==========
    all_failed_jobs = []
    for r in results:
        all_failed_jobs.extend(r.failed_jobs)

    if all_failed_jobs:
        failed_columns = [
            ColumnDef(header="Account", width=15),
            ColumnDef(header="Region", width=15),
            ColumnDef(header="Job ID", width=40),
            ColumnDef(header="상태", width=12, style="center"),
            ColumnDef(header="리소스 타입", width=15),
            ColumnDef(header="리소스 ARN", width=50),
            ColumnDef(header="Vault", width=25),
            ColumnDef(header="생성일", width=18),
            ColumnDef(header="완료일", width=18),
            ColumnDef(header="메시지", width=40),
        ]
        failed_sheet = wb.new_sheet("Failed Jobs", failed_columns)

        # 최신순 정렬
        all_failed_jobs.sort(key=lambda j: j.creation_date or datetime.min, reverse=True)

        for j in all_failed_jobs:
            row_num = failed_sheet.add_row([
                j.account_name,
                j.region,
                j.job_id,
                j.status,
                j.resource_type,
                j.resource_arn,
                j.vault_name,
                j.creation_date.strftime("%Y-%m-%d %H:%M") if j.creation_date else "-",
                j.completion_date.strftime("%Y-%m-%d %H:%M") if j.completion_date else "-",
                j.status_message or "-",
            ], style=Styles.danger())
            failed_sheet._ws.cell(row=row_num, column=4).fill = red_fill

    return str(wb.save_as(output_dir, "Comprehensive_Backup"))


# ============================================================================
# 실행 함수
# ============================================================================


def run(ctx) -> None:
    """통합 백업 현황 분석"""
    from core.tools.output import OutputPath, open_in_explorer

    console.print("[bold]통합 백업 현황 분석 시작...[/bold]\n")
    console.print(
        "[dim]분석 대상: RDS/Aurora, DynamoDB, EFS, FSx, EC2, DocumentDB, Neptune, "
        "Redshift, ElastiCache, MemoryDB, OpenSearch[/dim]"
    )
    console.print(f"[dim]최근 {JOB_DAYS}일 실패 작업 포함[/dim]\n")

    result = parallel_collect(ctx, _collect_comprehensive_backup_data, max_workers=10, service="backup")
    results: list[ComprehensiveBackupResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        for err in result.get_errors()[:5]:
            console.print(f"[dim]  - {err.identifier}/{err.region}: {err.message}[/dim]")

    if not results:
        console.print("\n[yellow]백업 데이터 없음[/yellow]")
        console.print("[dim]리소스가 없거나 접근 권한이 없습니다.[/dim]")
        return

    # 통계 계산
    total_resources = sum(len(r.backup_statuses) for r in results)
    disabled_resources = sum(len(r.get_disabled_resources()) for r in results)
    failed_jobs = sum(len(r.failed_jobs) for r in results)
    tag_conditions = sum(len(r.tag_conditions) for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"분석 리소스: {total_resources}개")

    if disabled_resources > 0:
        console.print(f"[yellow]백업 비활성화: {disabled_resources}개[/yellow]")
    else:
        console.print(f"[green]백업 비활성화: 0개[/green]")

    if failed_jobs > 0:
        console.print(f"[red]최근 {JOB_DAYS}일 실패 작업: {failed_jobs}건[/red]")

    if tag_conditions > 0:
        console.print(f"Backup Plan 태그 조건: {tag_conditions}개")

    # 보고서 생성
    identifier = ctx.accounts[0].id if ctx.accounts else ctx.profile_name or "default"
    output_path = OutputPath(identifier).sub("backup", "comprehensive").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
