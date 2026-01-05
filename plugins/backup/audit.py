"""
plugins/backup/audit.py - AWS Backup 현황 분석

Backup Vault, Plan, 최근 작업 현황을 분석합니다.

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 최근 작업 조회 기간
JOB_DAYS = 7

REQUIRED_PERMISSIONS = {
    "read": [
        "backup:ListBackupVaults",
        "backup:ListRecoveryPointsByBackupVault",
        "backup:ListBackupPlans",
        "backup:GetBackupPlan",
        "backup:ListBackupJobs",
    ],
}


class JobStatus(Enum):
    """백업 작업 상태"""

    COMPLETED = "COMPLETED"
    FAILED = "FAILED"
    RUNNING = "RUNNING"
    ABORTED = "ABORTED"
    PARTIAL = "PARTIAL"


@dataclass
class VaultInfo:
    """Backup Vault 정보"""

    account_id: str
    account_name: str
    region: str
    vault_name: str
    vault_arn: str
    creation_date: datetime | None
    encryption_key_arn: str
    recovery_point_count: int = 0
    total_size_bytes: int = 0
    locked: bool = False

    @property
    def size_gb(self) -> float:
        return self.total_size_bytes / (1024**3)

    @property
    def is_encrypted(self) -> bool:
        return bool(self.encryption_key_arn)


@dataclass
class PlanInfo:
    """Backup Plan 정보"""

    account_id: str
    account_name: str
    region: str
    plan_id: str
    plan_name: str
    version_id: str
    creation_date: datetime | None
    rule_count: int = 0
    resource_count: int = 0


@dataclass
class JobInfo:
    """Backup Job 정보"""

    account_id: str
    account_name: str
    region: str
    job_id: str
    vault_name: str
    resource_arn: str
    resource_type: str
    status: str
    status_message: str
    creation_date: datetime | None
    completion_date: datetime | None
    backup_size_bytes: int = 0

    @property
    def is_failed(self) -> bool:
        return self.status in ("FAILED", "ABORTED", "PARTIAL")


@dataclass
class BackupAnalysisResult:
    """Backup 분석 결과"""

    account_id: str
    account_name: str
    region: str
    vaults: list[VaultInfo] = field(default_factory=list)
    plans: list[PlanInfo] = field(default_factory=list)
    jobs: list[JobInfo] = field(default_factory=list)

    @property
    def total_vaults(self) -> int:
        return len(self.vaults)

    @property
    def total_recovery_points(self) -> int:
        return sum(v.recovery_point_count for v in self.vaults)

    @property
    def total_plans(self) -> int:
        return len(self.plans)

    @property
    def job_failed_count(self) -> int:
        return sum(1 for j in self.jobs if j.is_failed)

    @property
    def failed_jobs(self) -> list:
        return [j for j in self.jobs if j.is_failed]


def _collect_backup_data(
    session, account_id: str, account_name: str, region: str
) -> BackupAnalysisResult | None:
    """단일 계정/리전의 Backup 데이터 수집"""
    from botocore.exceptions import ClientError

    backup = get_client(session, "backup", region_name=region)
    result = BackupAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    # 1. Vault 수집
    try:
        paginator = backup.get_paginator("list_backup_vaults")
        for page in paginator.paginate():
            for vault in page.get("BackupVaultList", []):
                vault_name = vault.get("BackupVaultName", "")

                # 복구 지점 수 및 총 크기 계산
                rp_count = 0
                total_size = 0
                try:
                    rp_paginator = backup.get_paginator("list_recovery_points_by_backup_vault")
                    for rp_page in rp_paginator.paginate(BackupVaultName=vault_name):
                        for rp in rp_page.get("RecoveryPoints", []):
                            rp_count += 1
                            total_size += rp.get("BackupSizeInBytes", 0)
                except ClientError:
                    pass

                result.vaults.append(
                    VaultInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        vault_name=vault_name,
                        vault_arn=vault.get("BackupVaultArn", ""),
                        creation_date=vault.get("CreationDate"),
                        encryption_key_arn=vault.get("EncryptionKeyArn", ""),
                        recovery_point_count=rp_count,
                        total_size_bytes=total_size,
                        locked=vault.get("Locked", False),
                    )
                )
    except ClientError:
        pass

    # 2. Plan 수집
    try:
        paginator = backup.get_paginator("list_backup_plans")
        for page in paginator.paginate():
            for plan in page.get("BackupPlansList", []):
                plan_id = plan.get("BackupPlanId", "")

                # 플랜 상세 (규칙 수)
                rule_count = 0
                try:
                    plan_detail = backup.get_backup_plan(BackupPlanId=plan_id)
                    rules = plan_detail.get("BackupPlan", {}).get("Rules", [])
                    rule_count = len(rules)
                except ClientError:
                    pass

                # 리소스 선택 수
                resource_count = 0
                try:
                    selections = backup.list_backup_selections(BackupPlanId=plan_id)
                    resource_count = len(selections.get("BackupSelectionsList", []))
                except ClientError:
                    pass

                result.plans.append(
                    PlanInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        plan_id=plan_id,
                        plan_name=plan.get("BackupPlanName", ""),
                        version_id=plan.get("VersionId", ""),
                        creation_date=plan.get("CreationDate"),
                        rule_count=rule_count,
                        resource_count=resource_count,
                    )
                )
    except ClientError:
        pass

    # 3. 최근 Job 수집
    try:
        now = datetime.now(timezone.utc)
        start_date = now - timedelta(days=JOB_DAYS)

        paginator = backup.get_paginator("list_backup_jobs")
        for page in paginator.paginate(
            ByCreatedAfter=start_date,
            ByCreatedBefore=now,
        ):
            for job in page.get("BackupJobs", []):
                result.jobs.append(
                    JobInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        job_id=job.get("BackupJobId", ""),
                        vault_name=job.get("BackupVaultName", ""),
                        resource_arn=job.get("ResourceArn", ""),
                        resource_type=job.get("ResourceType", ""),
                        status=job.get("State", ""),
                        status_message=job.get("StatusMessage", ""),
                        creation_date=job.get("CreationDate"),
                        completion_date=job.get("CompletionDate"),
                        backup_size_bytes=job.get("BackupSizeInBytes", 0),
                    )
                )
    except ClientError:
        pass

    # 데이터가 하나라도 있으면 반환
    if result.vaults or result.plans or result.jobs:
        return result
    return None


def generate_report(results: list[BackupAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # ========== Summary 시트 ==========
    ws = wb.create_sheet("Summary")
    ws["A1"] = "AWS Backup 현황 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "Vault 수", "복구지점 수", "Plan 수", "실패 작업"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_vaults)
        ws.cell(row=row, column=4, value=r.total_recovery_points)
        ws.cell(row=row, column=5, value=r.total_plans)
        ws.cell(row=row, column=6, value=r.job_failed_count)

        if r.job_failed_count > 0:
            ws.cell(row=row, column=6).fill = red_fill

    # ========== Vaults 시트 ==========
    ws_vaults = wb.create_sheet("Vaults")
    vault_headers = ["Account", "Region", "Vault 이름", "복구지점 수", "총 크기", "암호화", "잠금"]
    for col, h in enumerate(vault_headers, 1):
        ws_vaults.cell(row=1, column=col, value=h).fill = header_fill
        ws_vaults.cell(row=1, column=col).font = header_font

    vault_row = 1
    for r in results:
        for v in r.vaults:
            vault_row += 1
            ws_vaults.cell(row=vault_row, column=1, value=v.account_name)
            ws_vaults.cell(row=vault_row, column=2, value=v.region)
            ws_vaults.cell(row=vault_row, column=3, value=v.vault_name)
            ws_vaults.cell(row=vault_row, column=4, value=v.recovery_point_count)
            ws_vaults.cell(row=vault_row, column=5, value=f"{v.size_gb:.2f} GB")
            ws_vaults.cell(row=vault_row, column=6, value="Yes" if v.is_encrypted else "No")
            ws_vaults.cell(row=vault_row, column=7, value="Yes" if v.locked else "No")

            if not v.is_encrypted:
                ws_vaults.cell(row=vault_row, column=6).fill = yellow_fill

    # ========== Plans 시트 ==========
    ws_plans = wb.create_sheet("Plans")
    plan_headers = ["Account", "Region", "Plan 이름", "ID", "규칙 수", "리소스 선택 수", "생성일"]
    for col, h in enumerate(plan_headers, 1):
        ws_plans.cell(row=1, column=col, value=h).fill = header_fill
        ws_plans.cell(row=1, column=col).font = header_font

    plan_row = 1
    for r in results:
        for p in r.plans:
            plan_row += 1
            ws_plans.cell(row=plan_row, column=1, value=p.account_name)
            ws_plans.cell(row=plan_row, column=2, value=p.region)
            ws_plans.cell(row=plan_row, column=3, value=p.plan_name)
            ws_plans.cell(row=plan_row, column=4, value=p.plan_id)
            ws_plans.cell(row=plan_row, column=5, value=p.rule_count)
            ws_plans.cell(row=plan_row, column=6, value=p.resource_count)
            ws_plans.cell(
                row=plan_row,
                column=7,
                value=p.creation_date.strftime("%Y-%m-%d") if p.creation_date else "-",
            )

            # 리소스 없는 플랜 경고
            if p.resource_count == 0:
                ws_plans.cell(row=plan_row, column=6).fill = yellow_fill

    # ========== Failed Jobs 시트 (실패 작업만) ==========
    failed_jobs: list[JobInfo] = []
    for r in results:
        failed_jobs.extend(r.failed_jobs)

    if failed_jobs:
        ws_jobs = wb.create_sheet("Failed Jobs")
        job_headers = ["Account", "Region", "상태", "리소스 타입", "리소스 ARN", "Vault", "생성일", "메시지"]
        for col, h in enumerate(job_headers, 1):
            ws_jobs.cell(row=1, column=col, value=h).fill = header_fill
            ws_jobs.cell(row=1, column=col).font = header_font

        failed_jobs.sort(key=lambda j: j.creation_date or datetime.min, reverse=True)

        job_row = 1
        for j in failed_jobs:
            job_row += 1
            ws_jobs.cell(row=job_row, column=1, value=j.account_name)
            ws_jobs.cell(row=job_row, column=2, value=j.region)
            ws_jobs.cell(row=job_row, column=3, value=j.status)
            ws_jobs.cell(row=job_row, column=4, value=j.resource_type)
            ws_jobs.cell(row=job_row, column=5, value=j.resource_arn)
            ws_jobs.cell(row=job_row, column=6, value=j.vault_name)
            ws_jobs.cell(
                row=job_row,
                column=7,
                value=j.creation_date.strftime("%Y-%m-%d %H:%M") if j.creation_date else "-",
            )
            ws_jobs.cell(row=job_row, column=8, value=j.status_message or "-")
            ws_jobs.cell(row=job_row, column=3).fill = red_fill

    # 열 너비 조정
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            col_idx = col[0].column
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"AWS_Backup_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def run(ctx) -> None:
    """AWS Backup 현황 분석"""
    console.print("[bold]AWS Backup 현황 분석 시작...[/bold]\n")
    console.print(f"[dim]최근 {JOB_DAYS}일 작업 현황 포함[/dim]\n")

    result = parallel_collect(ctx, _collect_backup_data, max_workers=10, service="backup")
    results: list[BackupAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        for err in result.get_errors()[:5]:
            console.print(f"[dim]  - {err.identifier}/{err.region}: {err.message}[/dim]")

    if not results:
        console.print("\n[yellow]Backup 데이터 없음[/yellow]")
        console.print("[dim]Backup Vault가 없거나 접근 권한이 없습니다.[/dim]")
        return

    # 통계 계산
    total_vaults = sum(r.total_vaults for r in results)
    total_rps = sum(r.total_recovery_points for r in results)
    total_plans = sum(r.total_plans for r in results)
    total_failed = sum(r.job_failed_count for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"Vault: {total_vaults}개, 복구 지점: {total_rps}개")
    console.print(f"Plan: {total_plans}개")

    if total_failed > 0:
        console.print(f"[red]최근 {JOB_DAYS}일 실패 작업: {total_failed}건[/red]")

    # 보고서 생성
    identifier = ctx.accounts[0].id if ctx.accounts else ctx.profile_name or "default"
    output_path = OutputPath(identifier).sub("backup", "inventory").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
