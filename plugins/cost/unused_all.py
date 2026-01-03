"""
plugins/cost/unused_all.py - 미사용 리소스 종합 분석 (병렬 처리)

모든 미사용 리소스 종합 보고서:
- NAT Gateway, ENI, EBS, EIP, ELB, Target Group
- EBS Snapshot, AMI, RDS Snapshot
- CloudWatch Log Group
- VPC Endpoint, Secrets Manager, KMS
- ECR, Route53, S3
- Lambda

병렬 처리 전략:
1. 계정/리전 레벨: parallel_collect로 병렬 처리
2. 리소스 타입 레벨: ThreadPoolExecutor로 병렬 수집

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from typing import TYPE_CHECKING, Any, Callable, Dict, List, Optional, Tuple

from rich.console import Console
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
)

from core.parallel import get_client, is_quiet, parallel_collect, quiet_mode, set_quiet
from core.tools.output import OutputPath, open_in_explorer

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

# 각 도구에서 수집/분석 함수 import
from plugins.cloudwatch.loggroup_audit import (
    LogGroupAnalysisResult,
    analyze_log_groups,
    collect_log_groups,
)
from plugins.ec2.ami_audit import (
    AMIAnalysisResult,
    analyze_amis,
    collect_amis,
    get_used_ami_ids,
)
from plugins.ec2.ebs_audit import EBSAnalysisResult, analyze_ebs, collect_ebs
from plugins.ec2.eip_audit import EIPAnalysisResult, analyze_eips, collect_eips
from plugins.ec2.snapshot_audit import (
    SnapshotAnalysisResult,
    analyze_snapshots,
    collect_snapshots,
    get_ami_snapshot_mapping,
)
from plugins.ecr.unused import ECRAnalysisResult, analyze_ecr_repos, collect_ecr_repos
from plugins.elb.target_group_audit import (
    TargetGroupAnalysisResult,
    analyze_target_groups,
    collect_target_groups,
)
from plugins.elb.unused import (
    LBAnalysisResult,
    analyze_load_balancers,
    collect_classic_load_balancers,
    collect_v2_load_balancers,
)
from plugins.fn.common.collector import collect_functions_with_metrics
from plugins.fn.unused import LambdaAnalysisResult
from plugins.fn.unused import analyze_functions as analyze_lambda_functions
from plugins.kms.unused import KMSKeyAnalysisResult, analyze_kms_keys, collect_kms_keys
from plugins.rds.snapshot_audit import (
    RDSSnapshotAnalysisResult,
    analyze_rds_snapshots,
    collect_rds_snapshots,
)
from plugins.route53.empty_zone import (
    Route53AnalysisResult,
    analyze_hosted_zones,
    collect_hosted_zones,
)
from plugins.s3.empty_bucket import S3AnalysisResult, analyze_buckets, collect_buckets
from plugins.secretsmanager.unused import (
    SecretAnalysisResult,
    analyze_secrets,
    collect_secrets,
)
from plugins.vpc.endpoint_audit import (
    EndpointAnalysisResult,
    analyze_endpoints,
    collect_endpoints,
)
from plugins.vpc.eni_audit import ENIAnalysisResult, analyze_enis, collect_enis
from plugins.vpc.nat_audit_analysis import NATAnalyzer, NATCollector

# 신규 추가된 미사용 리소스 분석 모듈
from plugins.elasticache.unused import (
    ElastiCacheAnalysisResult,
    analyze_clusters as analyze_elasticache_clusters,
    collect_elasticache_clusters,
)
from plugins.rds.unused import (
    RDSAnalysisResult as RDSInstanceAnalysisResult,
    analyze_instances as analyze_rds_instances,
    collect_rds_instances,
)
from plugins.efs.unused import (
    EFSAnalysisResult,
    analyze_filesystems as analyze_efs_filesystems,
    collect_efs_filesystems,
)
from plugins.sqs.unused import (
    SQSAnalysisResult,
    analyze_queues as analyze_sqs_queues,
    collect_sqs_queues,
)
from plugins.sns.unused import (
    SNSAnalysisResult,
    analyze_topics as analyze_sns_topics,
    collect_sns_topics,
)
from plugins.acm.unused import (
    ACMAnalysisResult,
    analyze_certificates as analyze_acm_certificates,
    collect_certificates as collect_acm_certificates,
)
from plugins.apigateway.unused import (
    APIGatewayAnalysisResult,
    analyze_apis as analyze_apigateway_apis,
    collect_apis as collect_apigateway_apis,
)
from plugins.eventbridge.unused import (
    EventBridgeAnalysisResult,
    analyze_rules as analyze_eventbridge_rules,
    collect_rules as collect_eventbridge_rules,
)
from plugins.cloudwatch.alarm_orphan import (
    AlarmAnalysisResult,
    analyze_alarms as analyze_cw_alarms,
    collect_alarms as collect_cw_alarms,
)
from plugins.dynamodb.unused import (
    DynamoDBAnalysisResult,
    analyze_tables as analyze_dynamodb_tables,
    collect_dynamodb_tables,
)

console = Console()


# =============================================================================
# 리소스 설정 (if/elif 제거용 매핑)
# =============================================================================

# 리소스별 필드 매핑
# - display: 표시 이름
# - total/unused/waste: summary 필드명
# - data_unused: collector 반환 데이터의 unused 키 (unused, old, issue, orphan 등)
# - session/final: 세션/최종 결과의 필드명
# - data_key: collector 반환 데이터의 결과 키 (result, findings)
RESOURCE_FIELD_MAP: Dict[str, Dict[str, Any]] = {
    # Regional resources
    "nat": {
        "display": "NAT Gateway",
        "total": "nat_total",
        "unused": "nat_unused",
        "waste": "nat_monthly_waste",
        "data_unused": "unused",
        "session": "nat_findings",
        "final": "nat_findings",
        "data_key": "findings",
    },
    "eni": {
        "display": "ENI",
        "total": "eni_total",
        "unused": "eni_unused",
        "waste": None,
        "data_unused": "unused",
        "session": "eni_result",
        "final": "eni_results",
        "data_key": "result",
    },
    "ebs": {
        "display": "EBS",
        "total": "ebs_total",
        "unused": "ebs_unused",
        "waste": "ebs_monthly_waste",
        "data_unused": "unused",
        "session": "ebs_result",
        "final": "ebs_results",
        "data_key": "result",
    },
    "eip": {
        "display": "EIP",
        "total": "eip_total",
        "unused": "eip_unused",
        "waste": "eip_monthly_waste",
        "data_unused": "unused",
        "session": "eip_result",
        "final": "eip_results",
        "data_key": "result",
    },
    "elb": {
        "display": "ELB",
        "total": "elb_total",
        "unused": "elb_unused",
        "waste": "elb_monthly_waste",
        "data_unused": "unused",
        "session": "elb_result",
        "final": "elb_results",
        "data_key": "result",
    },
    "snapshot": {
        "display": "EBS Snapshot",
        "total": "snap_total",
        "unused": "snap_unused",
        "waste": "snap_monthly_waste",
        "data_unused": "unused",
        "session": "snap_result",
        "final": "snap_results",
        "data_key": "result",
    },
    "ami": {
        "display": "AMI",
        "total": "ami_total",
        "unused": "ami_unused",
        "waste": "ami_monthly_waste",
        "data_unused": "unused",
        "session": "ami_result",
        "final": "ami_results",
        "data_key": "result",
    },
    "rds_snapshot": {
        "display": "RDS Snapshot",
        "total": "rds_snap_total",
        "unused": "rds_snap_old",
        "waste": "rds_snap_monthly_waste",
        "data_unused": "old",
        "session": "rds_snap_result",
        "final": "rds_snap_results",
        "data_key": "result",
    },
    "loggroup": {
        "display": "Log Group",
        "total": "loggroup_total",
        "unused": "loggroup_issue",
        "waste": "loggroup_monthly_waste",
        "data_unused": "issue",
        "session": "loggroup_result",
        "final": "loggroup_results",
        "data_key": "result",
    },
    "target_group": {
        "display": "Target Group",
        "total": "tg_total",
        "unused": "tg_issue",
        "waste": None,
        "data_unused": "issue",
        "session": "tg_result",
        "final": "tg_results",
        "data_key": "result",
    },
    "endpoint": {
        "display": "VPC Endpoint",
        "total": "endpoint_total",
        "unused": "endpoint_unused",
        "waste": "endpoint_monthly_waste",
        "data_unused": "unused",
        "session": "endpoint_result",
        "final": "endpoint_results",
        "data_key": "result",
    },
    "secret": {
        "display": "Secrets Manager",
        "total": "secret_total",
        "unused": "secret_unused",
        "waste": "secret_monthly_waste",
        "data_unused": "unused",
        "session": "secret_result",
        "final": "secret_results",
        "data_key": "result",
    },
    "kms": {
        "display": "KMS",
        "total": "kms_total",
        "unused": "kms_unused",
        "waste": "kms_monthly_waste",
        "data_unused": "unused",
        "session": "kms_result",
        "final": "kms_results",
        "data_key": "result",
    },
    "ecr": {
        "display": "ECR",
        "total": "ecr_total",
        "unused": "ecr_issue",
        "waste": "ecr_monthly_waste",
        "data_unused": "issue",
        "session": "ecr_result",
        "final": "ecr_results",
        "data_key": "result",
    },
    "lambda": {
        "display": "Lambda",
        "total": "lambda_total",
        "unused": "lambda_unused",
        "waste": "lambda_monthly_waste",
        "data_unused": "unused",
        "session": "lambda_result",
        "final": "lambda_results",
        "data_key": "result",
    },
    "elasticache": {
        "display": "ElastiCache",
        "total": "elasticache_total",
        "unused": "elasticache_unused",
        "waste": "elasticache_monthly_waste",
        "data_unused": "unused",
        "session": "elasticache_result",
        "final": "elasticache_results",
        "data_key": "result",
    },
    "rds_instance": {
        "display": "RDS Instance",
        "total": "rds_instance_total",
        "unused": "rds_instance_unused",
        "waste": "rds_instance_monthly_waste",
        "data_unused": "unused",
        "session": "rds_instance_result",
        "final": "rds_instance_results",
        "data_key": "result",
    },
    "efs": {
        "display": "EFS",
        "total": "efs_total",
        "unused": "efs_unused",
        "waste": "efs_monthly_waste",
        "data_unused": "unused",
        "session": "efs_result",
        "final": "efs_results",
        "data_key": "result",
    },
    "sqs": {
        "display": "SQS",
        "total": "sqs_total",
        "unused": "sqs_unused",
        "waste": None,
        "data_unused": "unused",
        "session": "sqs_result",
        "final": "sqs_results",
        "data_key": "result",
    },
    "sns": {
        "display": "SNS",
        "total": "sns_total",
        "unused": "sns_unused",
        "waste": None,
        "data_unused": "unused",
        "session": "sns_result",
        "final": "sns_results",
        "data_key": "result",
    },
    "acm": {
        "display": "ACM",
        "total": "acm_total",
        "unused": "acm_unused",
        "waste": None,
        "data_unused": "unused",
        "session": "acm_result",
        "final": "acm_results",
        "data_key": "result",
    },
    "apigateway": {
        "display": "API Gateway",
        "total": "apigateway_total",
        "unused": "apigateway_unused",
        "waste": None,
        "data_unused": "unused",
        "session": "apigateway_result",
        "final": "apigateway_results",
        "data_key": "result",
    },
    "eventbridge": {
        "display": "EventBridge",
        "total": "eventbridge_total",
        "unused": "eventbridge_unused",
        "waste": None,
        "data_unused": "unused",
        "session": "eventbridge_result",
        "final": "eventbridge_results",
        "data_key": "result",
    },
    "cw_alarm": {
        "display": "CloudWatch Alarm",
        "total": "cw_alarm_total",
        "unused": "cw_alarm_orphan",
        "waste": None,
        "data_unused": "orphan",
        "session": "cw_alarm_result",
        "final": "cw_alarm_results",
        "data_key": "result",
    },
    "dynamodb": {
        "display": "DynamoDB",
        "total": "dynamodb_total",
        "unused": "dynamodb_unused",
        "waste": "dynamodb_monthly_waste",
        "data_unused": "unused",
        "session": "dynamodb_result",
        "final": "dynamodb_results",
        "data_key": "result",
    },
    # Global resources
    "route53": {
        "display": "Route53",
        "total": "route53_total",
        "unused": "route53_empty",
        "waste": "route53_monthly_waste",
        "data_unused": "empty",
        "session": "route53_result",
        "final": "route53_results",
        "data_key": "result",
        "is_global": True,
    },
    "s3": {
        "display": "S3",
        "total": "s3_total",
        "unused": "s3_empty",
        "waste": None,
        "data_unused": "empty",
        "session": "s3_result",
        "final": "s3_results",
        "data_key": "result",
        "is_global": True,
    },
}

# 비용 추정 가능한 리소스 필드 목록 (total_waste 계산용)
WASTE_FIELDS = [cfg["waste"] for cfg in RESOURCE_FIELD_MAP.values() if cfg.get("waste")]


# =============================================================================
# 전역 서비스 수집 동기화 (Thread-safe)
# =============================================================================

_global_lock = threading.Lock()
_global_collected: set = set()


def _reset_global_tracking() -> None:
    """전역 서비스 추적 초기화 (새 실행 시 호출)"""
    global _global_collected
    with _global_lock:
        _global_collected = set()


def _should_collect_global(account_id: str) -> bool:
    """해당 계정의 전역 서비스를 수집해야 하는지 확인 (thread-safe)"""
    with _global_lock:
        if account_id in _global_collected:
            return False
        _global_collected.add(account_id)
        return True


# =============================================================================
# 종합 결과 데이터 구조
# =============================================================================


@dataclass
class UnusedResourceSummary:
    """미사용 리소스 종합 요약"""

    account_id: str
    account_name: str
    region: str

    # NAT Gateway
    nat_total: int = 0
    nat_unused: int = 0
    nat_monthly_waste: float = 0.0

    # ENI
    eni_total: int = 0
    eni_unused: int = 0

    # EBS
    ebs_total: int = 0
    ebs_unused: int = 0
    ebs_monthly_waste: float = 0.0

    # EIP
    eip_total: int = 0
    eip_unused: int = 0
    eip_monthly_waste: float = 0.0

    # ELB
    elb_total: int = 0
    elb_unused: int = 0
    elb_monthly_waste: float = 0.0

    # EBS Snapshot
    snap_total: int = 0
    snap_unused: int = 0
    snap_monthly_waste: float = 0.0

    # AMI
    ami_total: int = 0
    ami_unused: int = 0
    ami_monthly_waste: float = 0.0

    # RDS Snapshot
    rds_snap_total: int = 0
    rds_snap_old: int = 0
    rds_snap_monthly_waste: float = 0.0

    # CloudWatch Log Group
    loggroup_total: int = 0
    loggroup_issue: int = 0
    loggroup_monthly_waste: float = 0.0

    # Target Group
    tg_total: int = 0
    tg_issue: int = 0

    # VPC Endpoint
    endpoint_total: int = 0
    endpoint_unused: int = 0
    endpoint_monthly_waste: float = 0.0

    # Secrets Manager
    secret_total: int = 0
    secret_unused: int = 0
    secret_monthly_waste: float = 0.0

    # KMS
    kms_total: int = 0
    kms_unused: int = 0
    kms_monthly_waste: float = 0.0

    # ECR
    ecr_total: int = 0
    ecr_issue: int = 0
    ecr_monthly_waste: float = 0.0

    # Route53
    route53_total: int = 0
    route53_empty: int = 0
    route53_monthly_waste: float = 0.0

    # S3
    s3_total: int = 0
    s3_empty: int = 0

    # Lambda
    lambda_total: int = 0
    lambda_unused: int = 0
    lambda_monthly_waste: float = 0.0

    # ElastiCache
    elasticache_total: int = 0
    elasticache_unused: int = 0
    elasticache_monthly_waste: float = 0.0

    # RDS Instance
    rds_instance_total: int = 0
    rds_instance_unused: int = 0
    rds_instance_monthly_waste: float = 0.0

    # EFS
    efs_total: int = 0
    efs_unused: int = 0
    efs_monthly_waste: float = 0.0

    # SQS
    sqs_total: int = 0
    sqs_unused: int = 0

    # SNS
    sns_total: int = 0
    sns_unused: int = 0

    # ACM
    acm_total: int = 0
    acm_unused: int = 0

    # API Gateway
    apigateway_total: int = 0
    apigateway_unused: int = 0

    # EventBridge
    eventbridge_total: int = 0
    eventbridge_unused: int = 0

    # CloudWatch Alarm
    cw_alarm_total: int = 0
    cw_alarm_orphan: int = 0

    # DynamoDB
    dynamodb_total: int = 0
    dynamodb_unused: int = 0
    dynamodb_monthly_waste: float = 0.0


@dataclass
class SessionCollectionResult:
    """단일 세션(계정/리전)의 수집 결과"""

    summary: UnusedResourceSummary

    # 상세 결과 (각 리소스 타입별)
    nat_findings: List[Any] = field(default_factory=list)
    eni_result: Optional[ENIAnalysisResult] = None
    ebs_result: Optional[EBSAnalysisResult] = None
    eip_result: Optional[EIPAnalysisResult] = None
    elb_result: Optional[LBAnalysisResult] = None
    snap_result: Optional[SnapshotAnalysisResult] = None
    ami_result: Optional[AMIAnalysisResult] = None
    rds_snap_result: Optional[RDSSnapshotAnalysisResult] = None
    loggroup_result: Optional[LogGroupAnalysisResult] = None
    tg_result: Optional[TargetGroupAnalysisResult] = None
    endpoint_result: Optional[EndpointAnalysisResult] = None
    secret_result: Optional[SecretAnalysisResult] = None
    kms_result: Optional[KMSKeyAnalysisResult] = None
    ecr_result: Optional[ECRAnalysisResult] = None
    route53_result: Optional[Route53AnalysisResult] = None
    s3_result: Optional[S3AnalysisResult] = None
    lambda_result: Optional[LambdaAnalysisResult] = None

    # 신규 추가 리소스
    elasticache_result: Optional[ElastiCacheAnalysisResult] = None
    rds_instance_result: Optional[RDSInstanceAnalysisResult] = None
    efs_result: Optional[EFSAnalysisResult] = None
    sqs_result: Optional[SQSAnalysisResult] = None
    sns_result: Optional[SNSAnalysisResult] = None
    acm_result: Optional[ACMAnalysisResult] = None
    apigateway_result: Optional[APIGatewayAnalysisResult] = None
    eventbridge_result: Optional[EventBridgeAnalysisResult] = None
    cw_alarm_result: Optional[AlarmAnalysisResult] = None
    dynamodb_result: Optional[DynamoDBAnalysisResult] = None

    # 에러 목록
    errors: List[str] = field(default_factory=list)


@dataclass
class UnusedAllResult:
    """종합 분석 결과"""

    summaries: List[UnusedResourceSummary] = field(default_factory=list)

    # 상세 결과
    nat_findings: List[Any] = field(default_factory=list)
    eni_results: List[ENIAnalysisResult] = field(default_factory=list)
    ebs_results: List[EBSAnalysisResult] = field(default_factory=list)
    eip_results: List[EIPAnalysisResult] = field(default_factory=list)
    elb_results: List[LBAnalysisResult] = field(default_factory=list)
    snap_results: List[SnapshotAnalysisResult] = field(default_factory=list)
    ami_results: List[AMIAnalysisResult] = field(default_factory=list)
    rds_snap_results: List[RDSSnapshotAnalysisResult] = field(default_factory=list)
    loggroup_results: List[LogGroupAnalysisResult] = field(default_factory=list)
    tg_results: List[TargetGroupAnalysisResult] = field(default_factory=list)
    endpoint_results: List[EndpointAnalysisResult] = field(default_factory=list)
    secret_results: List[SecretAnalysisResult] = field(default_factory=list)
    kms_results: List[KMSKeyAnalysisResult] = field(default_factory=list)
    ecr_results: List[ECRAnalysisResult] = field(default_factory=list)
    route53_results: List[Route53AnalysisResult] = field(default_factory=list)
    s3_results: List[S3AnalysisResult] = field(default_factory=list)
    lambda_results: List[LambdaAnalysisResult] = field(default_factory=list)

    # 신규 추가 리소스
    elasticache_results: List[ElastiCacheAnalysisResult] = field(default_factory=list)
    rds_instance_results: List[RDSInstanceAnalysisResult] = field(default_factory=list)
    efs_results: List[EFSAnalysisResult] = field(default_factory=list)
    sqs_results: List[SQSAnalysisResult] = field(default_factory=list)
    sns_results: List[SNSAnalysisResult] = field(default_factory=list)
    acm_results: List[ACMAnalysisResult] = field(default_factory=list)
    apigateway_results: List[APIGatewayAnalysisResult] = field(default_factory=list)
    eventbridge_results: List[EventBridgeAnalysisResult] = field(default_factory=list)
    cw_alarm_results: List[AlarmAnalysisResult] = field(default_factory=list)
    dynamodb_results: List[DynamoDBAnalysisResult] = field(default_factory=list)


# =============================================================================
# 개별 리소스 수집/분석 함수 (병렬 실행용)
# =============================================================================


def _collect_nat(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """NAT Gateway 수집 및 분석"""
    try:
        collector = NATCollector()
        nat_data = collector.collect(session, account_id, account_name, region)
        if not nat_data.nat_gateways:
            return {"total": 0, "unused": 0, "waste": 0.0, "findings": []}

        analyzer = NATAnalyzer(nat_data)
        nat_result = analyzer.analyze()
        stats = analyzer.get_summary_stats()

        return {
            "total": stats.get("total_nat_count", 0),
            "unused": stats.get("unused_count", 0) + stats.get("low_usage_count", 0),
            "waste": stats.get("total_monthly_waste", 0),
            "findings": [nat_result],
        }
    except Exception as e:
        return {"error": f"NAT Gateway: {e}"}


def _collect_eni(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """ENI 수집 및 분석"""
    try:
        enis = collect_enis(session, account_id, account_name, region)
        if not enis:
            return {"total": 0, "unused": 0, "result": None}

        result = analyze_enis(enis, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.unused_count,
            "result": result,
        }
    except Exception as e:
        return {"error": f"ENI: {e}"}


def _collect_ebs(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """EBS 수집 및 분석"""
    try:
        volumes = collect_ebs(session, account_id, account_name, region)
        if not volumes:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_ebs(volumes, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.unused_count,
            "waste": result.unused_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"EBS: {e}"}


def _collect_eip(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """EIP 수집 및 분석"""
    try:
        eips = collect_eips(session, account_id, account_name, region)
        if not eips:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_eips(eips, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.unused_count,
            "waste": result.unused_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"EIP: {e}"}


def _collect_elb(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """ELB 수집 및 분석"""
    try:
        v2_lbs = collect_v2_load_balancers(session, account_id, account_name, region)
        classic_lbs = collect_classic_load_balancers(
            session, account_id, account_name, region
        )
        all_lbs = v2_lbs + classic_lbs
        if not all_lbs:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_load_balancers(all_lbs, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.unused_count + result.unhealthy_count,
            "waste": result.unused_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"ELB: {e}"}


def _collect_snapshot(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """EBS Snapshot 수집 및 분석"""
    try:
        snapshots = collect_snapshots(session, account_id, account_name, region)
        if not snapshots:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        ami_mapping = get_ami_snapshot_mapping(session, region)
        result = analyze_snapshots(
            snapshots, ami_mapping, account_id, account_name, region
        )
        return {
            "total": result.total_count,
            "unused": result.orphan_count + result.old_count,
            "waste": result.orphan_monthly_cost + result.old_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"EBS Snapshot: {e}"}


def _collect_ami(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """AMI 수집 및 분석"""
    try:
        amis = collect_amis(session, account_id, account_name, region)
        if not amis:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        used_ami_ids = get_used_ami_ids(session, region)
        result = analyze_amis(amis, used_ami_ids, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.unused_count,
            "waste": result.unused_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"AMI: {e}"}


def _collect_rds_snapshot(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """RDS Snapshot 수집 및 분석"""
    try:
        rds_snaps = collect_rds_snapshots(session, account_id, account_name, region)
        if not rds_snaps:
            return {"total": 0, "old": 0, "waste": 0.0, "result": None}

        result = analyze_rds_snapshots(rds_snaps, account_id, account_name, region)
        return {
            "total": result.total_count,
            "old": result.old_count,
            "waste": result.old_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"RDS Snapshot: {e}"}


def _collect_loggroup(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """CloudWatch Log Group 수집 및 분석"""
    try:
        log_groups = collect_log_groups(session, account_id, account_name, region)
        if not log_groups:
            return {"total": 0, "issue": 0, "waste": 0.0, "result": None}

        result = analyze_log_groups(log_groups, account_id, account_name, region)
        return {
            "total": result.total_count,
            "issue": result.empty_count + result.old_count,
            "waste": result.empty_monthly_cost + result.old_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"Log Group: {e}"}


def _collect_target_group(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """Target Group 수집 및 분석"""
    try:
        tgs = collect_target_groups(session, account_id, account_name, region)
        if not tgs:
            return {"total": 0, "issue": 0, "result": None}

        result = analyze_target_groups(tgs, account_id, account_name, region)
        return {
            "total": result.total_count,
            "issue": result.unattached_count + result.no_targets_count,
            "result": result,
        }
    except Exception as e:
        return {"error": f"Target Group: {e}"}


def _collect_endpoint(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """VPC Endpoint 수집 및 분석"""
    try:
        endpoints = collect_endpoints(session, account_id, account_name, region)
        if not endpoints:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_endpoints(endpoints, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.unused_count,
            "waste": result.unused_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"VPC Endpoint: {e}"}


def _collect_secret(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """Secrets Manager 수집 및 분석"""
    try:
        secrets = collect_secrets(session, account_id, account_name, region)
        if not secrets:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_secrets(secrets, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.unused_count,
            "waste": result.unused_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"Secrets Manager: {e}"}


def _collect_kms(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """KMS 수집 및 분석"""
    try:
        kms_keys = collect_kms_keys(session, account_id, account_name, region)
        if not kms_keys:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_kms_keys(kms_keys, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.disabled_count + result.pending_delete_count,
            "waste": result.disabled_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"KMS: {e}"}


def _collect_ecr(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """ECR 수집 및 분석"""
    try:
        repos = collect_ecr_repos(session, account_id, account_name, region)
        if not repos:
            return {"total": 0, "issue": 0, "waste": 0.0, "result": None}

        result = analyze_ecr_repos(repos, account_id, account_name, region)
        return {
            "total": result.total_repos,
            "issue": result.empty_repos + result.repos_with_old_images,
            "waste": result.old_images_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"ECR: {e}"}


def _collect_lambda(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """Lambda 수집 및 분석"""
    try:
        functions = collect_functions_with_metrics(
            session, account_id, account_name, region
        )
        if not functions:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_lambda_functions(functions, account_id, account_name, region)
        return {
            "total": result.total_count,
            "unused": result.unused_count,
            "waste": result.unused_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"Lambda: {e}"}


def _collect_elasticache(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """ElastiCache 수집 및 분석"""
    try:
        clusters = collect_elasticache_clusters(
            session, account_id, account_name, region
        )
        if not clusters:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_elasticache_clusters(
            clusters, account_id, account_name, region
        )
        return {
            "total": result.total_clusters,
            "unused": result.unused_clusters + result.low_usage_clusters,
            "waste": result.unused_monthly_cost + result.low_usage_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"ElastiCache: {e}"}


def _collect_rds_instance(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """RDS Instance 수집 및 분석"""
    try:
        instances = collect_rds_instances(session, account_id, account_name, region)
        if not instances:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_rds_instances(instances, account_id, account_name, region)
        return {
            "total": result.total_instances,
            "unused": result.unused_instances
            + result.low_usage_instances
            + result.stopped_instances,
            "waste": result.unused_monthly_cost + result.low_usage_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"RDS Instance: {e}"}


def _collect_efs(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """EFS 수집 및 분석"""
    try:
        filesystems = collect_efs_filesystems(session, account_id, account_name, region)
        if not filesystems:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_efs_filesystems(filesystems, account_id, account_name, region)
        return {
            "total": result.total_filesystems,
            "unused": result.no_mount_target + result.no_io + result.empty,
            "waste": result.unused_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"EFS: {e}"}


def _collect_sqs(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """SQS 수집 및 분석"""
    try:
        queues = collect_sqs_queues(session, account_id, account_name, region)
        if not queues:
            return {"total": 0, "unused": 0, "result": None}

        result = analyze_sqs_queues(queues, account_id, account_name, region)
        return {
            "total": result.total_queues,
            "unused": result.unused_queues + result.empty_dlqs,
            "result": result,
        }
    except Exception as e:
        return {"error": f"SQS: {e}"}


def _collect_sns(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """SNS 수집 및 분석"""
    try:
        topics = collect_sns_topics(session, account_id, account_name, region)
        if not topics:
            return {"total": 0, "unused": 0, "result": None}

        result = analyze_sns_topics(topics, account_id, account_name, region)
        return {
            "total": result.total_topics,
            "unused": result.unused_topics + result.no_subscribers + result.no_messages,
            "result": result,
        }
    except Exception as e:
        return {"error": f"SNS: {e}"}


def _collect_acm(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """ACM 수집 및 분석"""
    try:
        certs = collect_acm_certificates(session, account_id, account_name, region)
        if not certs:
            return {"total": 0, "unused": 0, "result": None}

        result = analyze_acm_certificates(certs, account_id, account_name, region)
        return {
            "total": result.total_certs,
            "unused": result.unused_certs + result.expired_certs,
            "result": result,
        }
    except Exception as e:
        return {"error": f"ACM: {e}"}


def _collect_apigateway(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """API Gateway 수집 및 분석"""
    try:
        apis = collect_apigateway_apis(session, account_id, account_name, region)
        if not apis:
            return {"total": 0, "unused": 0, "result": None}

        result = analyze_apigateway_apis(apis, account_id, account_name, region)
        return {
            "total": result.total_apis,
            "unused": result.unused_apis + result.no_stages + result.low_usage,
            "result": result,
        }
    except Exception as e:
        return {"error": f"API Gateway: {e}"}


def _collect_eventbridge(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """EventBridge 수집 및 분석"""
    try:
        rules = collect_eventbridge_rules(session, account_id, account_name, region)
        if not rules:
            return {"total": 0, "unused": 0, "result": None}

        result = analyze_eventbridge_rules(rules, account_id, account_name, region)
        return {
            "total": result.total_rules,
            "unused": result.disabled_rules + result.no_targets + result.unused_rules,
            "result": result,
        }
    except Exception as e:
        return {"error": f"EventBridge: {e}"}


def _collect_cw_alarm(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """CloudWatch Alarm 수집 및 분석"""
    try:
        alarms = collect_cw_alarms(session, account_id, account_name, region)
        if not alarms:
            return {"total": 0, "orphan": 0, "result": None}

        result = analyze_cw_alarms(alarms, account_id, account_name, region)
        return {
            "total": result.total_alarms,
            "orphan": result.orphan_alarms + result.no_actions,
            "result": result,
        }
    except Exception as e:
        return {"error": f"CloudWatch Alarm: {e}"}


def _collect_dynamodb(
    session, account_id: str, account_name: str, region: str
) -> Dict[str, Any]:
    """DynamoDB 수집 및 분석"""
    try:
        tables = collect_dynamodb_tables(session, account_id, account_name, region)
        if not tables:
            return {"total": 0, "unused": 0, "waste": 0.0, "result": None}

        result = analyze_dynamodb_tables(tables, account_id, account_name, region)
        return {
            "total": result.total_tables,
            "unused": result.unused_tables + result.low_usage_tables,
            "waste": result.unused_monthly_cost + result.low_usage_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"DynamoDB: {e}"}


def _collect_route53(session, account_id: str, account_name: str) -> Dict[str, Any]:
    """Route53 수집 및 분석 (글로벌 서비스)"""
    try:
        zones = collect_hosted_zones(session, account_id, account_name)
        if not zones:
            return {"total": 0, "empty": 0, "waste": 0.0, "result": None}

        result = analyze_hosted_zones(zones, account_id, account_name)
        return {
            "total": result.total_zones,
            "empty": result.empty_zones + result.ns_soa_only_zones,
            "waste": result.wasted_monthly_cost,
            "result": result,
        }
    except Exception as e:
        return {"error": f"Route53: {e}"}


def _collect_s3(session, account_id: str, account_name: str) -> Dict[str, Any]:
    """S3 수집 및 분석 (글로벌 서비스)"""
    try:
        buckets = collect_buckets(session, account_id, account_name)
        if not buckets:
            return {"total": 0, "empty": 0, "result": None}

        result = analyze_buckets(buckets, account_id, account_name)
        return {
            "total": result.total_buckets,
            "empty": result.empty_buckets + result.versioning_only_buckets,
            "result": result,
        }
    except Exception as e:
        return {"error": f"S3: {e}"}


# =============================================================================
# 세션별 병렬 수집
# =============================================================================

# 리전별 리소스 수집기 정의
REGIONAL_COLLECTORS: Dict[str, Callable] = {
    "nat": _collect_nat,
    "eni": _collect_eni,
    "ebs": _collect_ebs,
    "eip": _collect_eip,
    "elb": _collect_elb,
    "snapshot": _collect_snapshot,
    "ami": _collect_ami,
    "rds_snapshot": _collect_rds_snapshot,
    "loggroup": _collect_loggroup,
    "target_group": _collect_target_group,
    "endpoint": _collect_endpoint,
    "secret": _collect_secret,
    "kms": _collect_kms,
    "ecr": _collect_ecr,
    "lambda": _collect_lambda,
    "elasticache": _collect_elasticache,
    "rds_instance": _collect_rds_instance,
    "efs": _collect_efs,
    "sqs": _collect_sqs,
    "sns": _collect_sns,
    "acm": _collect_acm,
    "apigateway": _collect_apigateway,
    "eventbridge": _collect_eventbridge,
    "cw_alarm": _collect_cw_alarm,
    "dynamodb": _collect_dynamodb,
}


def _apply_result(
    summary: UnusedResourceSummary,
    session_result: SessionCollectionResult,
    resource_type: str,
    data: Dict[str, Any],
) -> None:
    """리소스 결과를 요약 및 세션 결과에 적용 (매핑 기반)"""
    if "error" in data:
        session_result.errors.append(data["error"])
        return

    cfg = RESOURCE_FIELD_MAP.get(resource_type)
    if not cfg:
        return

    # summary 필드 설정
    setattr(summary, cfg["total"], data.get("total", 0))
    setattr(summary, cfg["unused"], data.get(cfg["data_unused"], 0))
    if cfg["waste"]:
        setattr(summary, cfg["waste"], data.get("waste", 0.0))

    # session_result 필드 설정
    result_data = data.get(cfg["data_key"])
    if result_data:
        setattr(session_result, cfg["session"], result_data)


def _run_collector_quiet(collector, session, account_id, account_name, region, quiet):
    """워커 스레드에서 quiet 상태를 설정하고 collector 실행"""
    set_quiet(quiet)
    return collector(session, account_id, account_name, region)


def _run_global_collector_quiet(collector, session, account_id, account_name, quiet):
    """워커 스레드에서 quiet 상태를 설정하고 글로벌 collector 실행"""
    set_quiet(quiet)
    return collector(session, account_id, account_name)


def collect_session_resources(
    session,
    account_id: str,
    account_name: str,
    region: str,
    selected_resources: Optional[set] = None,
) -> SessionCollectionResult:
    """단일 세션의 모든 리소스를 병렬로 수집

    Args:
        session: boto3 Session
        account_id: AWS 계정 ID
        account_name: 계정 이름
        region: AWS 리전
        selected_resources: 스캔할 리소스 set (None이면 전체)

    Returns:
        SessionCollectionResult: 수집 결과
    """
    summary = UnusedResourceSummary(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )
    result = SessionCollectionResult(summary=summary)

    # 부모 스레드의 quiet 상태를 가져와서 워커에 전파
    parent_quiet = is_quiet()

    # 선택적 스캔: 선택된 리소스만 수집
    collectors_to_run = REGIONAL_COLLECTORS
    if selected_resources:
        collectors_to_run = {
            k: v for k, v in REGIONAL_COLLECTORS.items() if k in selected_resources
        }

    # 리전별 리소스 병렬 수집 (최대 10개 동시 실행)
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {}
        for name, collector in collectors_to_run.items():
            future = executor.submit(
                _run_collector_quiet,
                collector,
                session,
                account_id,
                account_name,
                region,
                parent_quiet,
            )
            futures[future] = name

        for future in as_completed(futures):
            resource_type = futures[future]
            try:
                data = future.result()
                _apply_result(summary, result, resource_type, data)
            except Exception as e:
                result.errors.append(f"{resource_type}: {e}")

    # 글로벌 서비스 (계정당 한 번만 수집)
    # 선택적 스캔: route53, s3가 선택되었는지 확인
    collect_route53 = not selected_resources or "route53" in selected_resources
    collect_s3 = not selected_resources or "s3" in selected_resources

    if _should_collect_global(account_id) and (collect_route53 or collect_s3):
        with ThreadPoolExecutor(max_workers=2) as executor:
            futures = {}

            if collect_route53:
                futures["route53"] = executor.submit(
                    _run_global_collector_quiet,
                    _collect_route53,
                    session,
                    account_id,
                    account_name,
                    parent_quiet,
                )

            if collect_s3:
                futures["s3"] = executor.submit(
                    _run_global_collector_quiet,
                    _collect_s3,
                    session,
                    account_id,
                    account_name,
                    parent_quiet,
                )

            for resource_type, future in futures.items():
                try:
                    data = future.result()
                    _apply_result(summary, result, resource_type, data)
                except Exception as e:
                    result.errors.append(f"{resource_type}: {e}")

    return result


# =============================================================================
# 결과 집계
# =============================================================================


def _merge_session_result(
    final: UnusedAllResult, session_result: SessionCollectionResult
) -> None:
    """세션 결과를 최종 결과에 병합 (매핑 기반)"""
    final.summaries.append(session_result.summary)

    for cfg in RESOURCE_FIELD_MAP.values():
        session_field = cfg["session"]
        final_field = cfg["final"]
        session_data = getattr(session_result, session_field, None)

        if session_data:
            final_list = getattr(final, final_field)
            # findings는 리스트이므로 extend, 나머지는 단일 결과이므로 append
            if session_field.endswith("_findings"):
                final_list.extend(session_data)
            else:
                final_list.append(session_data)


# =============================================================================
# 옵션 수집 (CLI 프롬프트)
# =============================================================================


def collect_options(ctx: "ExecutionContext") -> None:
    """미사용 리소스 분석 옵션 수집

    사용자에게 전체 스캔 또는 선택 스캔 옵션을 제공합니다.

    Args:
        ctx: ExecutionContext
    """
    import questionary

    console.print("\n[bold cyan]미사용 리소스 종합 분석 설정[/bold cyan]")

    # 스캔 모드 선택
    scan_mode = questionary.select(
        "스캔 모드를 선택하세요:",
        choices=[
            questionary.Choice("전체 스캔 (모든 리소스)", value="all"),
            questionary.Choice("선택 스캔 (리소스 직접 선택)", value="select"),
        ],
        style=questionary.Style([("highlighted", "bold")]),
    ).ask()

    if scan_mode == "select":
        # 리소스 선택을 위한 체크박스
        resource_choices = [
            questionary.Choice(f"{cfg['display']} ({key})", value=key)
            for key, cfg in RESOURCE_FIELD_MAP.items()
        ]

        selected = questionary.checkbox(
            "분석할 리소스를 선택하세요 (Space로 선택, Enter로 확정):",
            choices=resource_choices,
            style=questionary.Style([("highlighted", "bold")]),
        ).ask()

        if selected:
            ctx.options["resources"] = selected
            console.print(f"[green]선택됨: {', '.join(selected)}[/green]")
        else:
            console.print("[yellow]선택 없음 - 전체 스캔으로 진행합니다.[/yellow]")
            ctx.options["resources"] = None
    else:
        ctx.options["resources"] = None
        console.print("[green]전체 리소스 스캔으로 진행합니다.[/green]")


# =============================================================================
# 메인 실행
# =============================================================================


def run(ctx: "ExecutionContext", resources: Optional[List[str]] = None) -> None:
    """미사용 리소스 종합 분석 실행 (병렬 처리)

    Args:
        ctx: 실행 컨텍스트
        resources: 스캔할 리소스 목록 (None이면 전체)
                   예: ["nat", "ebs", "eip", "lambda"]
    """
    # ctx.options에서 리소스 목록 가져오기 (collect_options에서 설정)
    if resources is None:
        resources = ctx.options.get("resources")

    # 선택적 스캔 설정
    if resources:
        selected = set(resources)
        invalid = selected - set(RESOURCE_FIELD_MAP.keys())
        if invalid:
            console.print(f"[red]알 수 없는 리소스: {', '.join(invalid)}[/red]")
            console.print(f"[dim]사용 가능: {', '.join(RESOURCE_FIELD_MAP.keys())}[/dim]")
            return
        console.print(f"[bold]선택된 리소스 분석: {', '.join(resources)}[/bold]\n")
    else:
        selected = None
        console.print("[bold]미사용 리소스 종합 분석 시작 (병렬 처리)...[/bold]\n")

    # 전역 서비스 추적 초기화
    _reset_global_tracking()

    # 병렬 수집 실행 (quiet_mode로 콘솔 출력 억제)
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("[cyan]리소스 수집 중...", total=None)

        # 선택적 스캔을 위한 래퍼 함수
        def collect_wrapper(session, account_id, account_name, region):
            return collect_session_resources(
                session, account_id, account_name, region, selected_resources=selected
            )

        with quiet_mode():
            parallel_result = parallel_collect(
                ctx,
                collect_wrapper,
                max_workers=20,
                service="multi",
            )

        progress.update(task, description="[green]수집 완료", completed=True)

    # 결과 집계
    final_result = UnusedAllResult()
    all_errors: List[str] = []

    for task_result in parallel_result.results:
        if task_result.success and task_result.data:
            session_result = task_result.data
            _merge_session_result(final_result, session_result)

            # 세션별 에러 수집
            if session_result.errors:
                for err in session_result.errors:
                    all_errors.append(
                        f"{task_result.identifier}/{task_result.region}: {err}"
                    )
        elif task_result.error:
            all_errors.append(str(task_result.error))

    if not final_result.summaries:
        console.print("[yellow]분석 결과 없음[/yellow]")
        return

    # 총 절감 가능 금액 계산 (WASTE_FIELDS 활용)
    total_waste = sum(
        sum(getattr(s, field, 0) for field in WASTE_FIELDS)
        for s in final_result.summaries
    )

    # 요약 출력 (RESOURCE_FIELD_MAP 활용)
    console.print("\n" + "=" * 50)
    console.print("[bold]종합 결과[/bold]")
    console.print("=" * 50)

    for resource_key, cfg in RESOURCE_FIELD_MAP.items():
        # 선택적 스캔인 경우 선택된 리소스만 출력
        if selected and resource_key not in selected:
            continue
        _print_summary(
            cfg["display"],
            final_result.summaries,
            cfg["total"],
            cfg["unused"],
            cfg["waste"],
        )

    if total_waste > 0:
        console.print(f"\n[bold yellow]총 월간 절감 가능: ${total_waste:,.2f}[/bold yellow]")

    # 실행 통계
    console.print(
        f"\n[dim]계정/리전: {parallel_result.success_count}개 성공, {parallel_result.error_count}개 실패[/dim]"
    )

    # 보고서 생성
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("unused-all").with_date().build()
    filepath = generate_report(final_result, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")

    if all_errors:
        _print_error_summary(all_errors)

    open_in_explorer(output_path)


def _print_error_summary(errors: List[str]) -> None:
    """오류 요약 출력 (유형별 그룹화)"""
    from collections import defaultdict
    import re

    # 오류 코드별로 그룹화
    # 형식: "{profile}/{region}: {resource}: {message}"
    error_groups: Dict[str, List[str]] = defaultdict(list)

    for err in errors:
        # 오류 코드 추출 시도
        # 예: "UnrecognizedClientException", "InvalidClientTokenId", "AuthFailure"
        code_match = re.search(
            r"(UnrecognizedClientException|InvalidClientTokenId|AuthFailure|"
            r"AccessDenied|ExpiredToken|InvalidAccessKeyId|SignatureDoesNotMatch|"
            r"ThrottlingException|ServiceUnavailable|InternalError|"
            r"UnauthorizedOperation|OptInRequired)",
            err,
        )

        if code_match:
            error_code = code_match.group(1)
        else:
            # 알 수 없는 오류는 "기타"로 분류
            error_code = "기타"

        # 리전 추출
        region_match = re.match(r"[^/]+/([^:]+):", err)
        region = region_match.group(1) if region_match else "unknown"

        error_groups[error_code].append(region)

    # 출력
    console.print(f"\n[yellow]오류 {len(errors)}건[/yellow]")

    for error_code, regions in sorted(error_groups.items(), key=lambda x: -len(x[1])):
        unique_regions = sorted(set(regions))
        if len(unique_regions) <= 3:
            region_str = ", ".join(unique_regions)
        else:
            region_str = f"{', '.join(unique_regions[:3])} 외 {len(unique_regions) - 3}개"
        console.print(f"  [dim]{error_code}: {region_str} ({len(regions)}건)[/dim]")


def _print_summary(
    name: str,
    summaries: List,
    total_attr: str,
    unused_attr: str,
    waste_attr: Optional[str],
) -> None:
    """요약 출력 헬퍼"""
    total = sum(getattr(s, total_attr, 0) for s in summaries)
    unused = sum(getattr(s, unused_attr, 0) for s in summaries)
    waste = sum(getattr(s, waste_attr, 0) for s in summaries) if waste_attr else 0

    console.print(f"\n[bold]{name}[/bold]: 전체 {total}개", end="")
    if unused > 0:
        waste_str = f" (${waste:,.2f}/월)" if waste > 0 else ""
        console.print(f" / [red]미사용 {unused}개{waste_str}[/red]")
    else:
        console.print("")


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(result: UnusedAllResult, output_dir: str) -> str:
    """종합 Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    # ===== Summary =====
    ws = wb.create_sheet("Summary")
    ws["A1"] = "미사용 리소스 종합 보고서"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    row = 4
    for col, h in enumerate(["리소스", "전체", "미사용", "월간 낭비"], 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    resources = [
        ("NAT Gateway", "nat_total", "nat_unused", "nat_monthly_waste"),
        ("ENI", "eni_total", "eni_unused", None),
        ("EBS", "ebs_total", "ebs_unused", "ebs_monthly_waste"),
        ("EIP", "eip_total", "eip_unused", "eip_monthly_waste"),
        ("ELB", "elb_total", "elb_unused", "elb_monthly_waste"),
        ("EBS Snapshot", "snap_total", "snap_unused", "snap_monthly_waste"),
        ("AMI", "ami_total", "ami_unused", "ami_monthly_waste"),
        ("RDS Snapshot", "rds_snap_total", "rds_snap_old", "rds_snap_monthly_waste"),
        ("Log Group", "loggroup_total", "loggroup_issue", "loggroup_monthly_waste"),
        ("Target Group", "tg_total", "tg_issue", None),
        ("VPC Endpoint", "endpoint_total", "endpoint_unused", "endpoint_monthly_waste"),
        ("Secrets Manager", "secret_total", "secret_unused", "secret_monthly_waste"),
        ("KMS", "kms_total", "kms_unused", "kms_monthly_waste"),
        ("ECR", "ecr_total", "ecr_issue", "ecr_monthly_waste"),
        ("Route53", "route53_total", "route53_empty", "route53_monthly_waste"),
        ("S3", "s3_total", "s3_empty", None),
        ("Lambda", "lambda_total", "lambda_unused", "lambda_monthly_waste"),
        (
            "ElastiCache",
            "elasticache_total",
            "elasticache_unused",
            "elasticache_monthly_waste",
        ),
        (
            "RDS Instance",
            "rds_instance_total",
            "rds_instance_unused",
            "rds_instance_monthly_waste",
        ),
        ("EFS", "efs_total", "efs_unused", "efs_monthly_waste"),
        ("SQS", "sqs_total", "sqs_unused", None),
        ("SNS", "sns_total", "sns_unused", None),
        ("ACM", "acm_total", "acm_unused", None),
        ("API Gateway", "apigateway_total", "apigateway_unused", None),
        ("EventBridge", "eventbridge_total", "eventbridge_unused", None),
        ("CloudWatch Alarm", "cw_alarm_total", "cw_alarm_orphan", None),
    ]

    for name, total_attr, unused_attr, waste_attr in resources:
        row += 1
        total = sum(getattr(s, total_attr, 0) for s in result.summaries)
        unused = sum(getattr(s, unused_attr, 0) for s in result.summaries)
        waste = (
            sum(getattr(s, waste_attr, 0) for s in result.summaries)
            if waste_attr
            else 0
        )
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=unused)
        ws.cell(row=row, column=4, value=f"${waste:,.2f}" if waste > 0 else "-")
        if unused > 0:
            ws.cell(row=row, column=3).fill = red_fill

    # 총 절감
    total_waste = sum(
        s.nat_monthly_waste
        + s.ebs_monthly_waste
        + s.eip_monthly_waste
        + s.elb_monthly_waste
        + s.snap_monthly_waste
        + s.ami_monthly_waste
        + s.rds_snap_monthly_waste
        + s.loggroup_monthly_waste
        + s.endpoint_monthly_waste
        + s.secret_monthly_waste
        + s.kms_monthly_waste
        + s.ecr_monthly_waste
        + s.route53_monthly_waste
        + s.lambda_monthly_waste
        + s.elasticache_monthly_waste
        + s.rds_instance_monthly_waste
        + s.efs_monthly_waste
        for s in result.summaries
    )
    row += 2
    ws.cell(row=row, column=1, value="총 월간 절감 가능").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"${total_waste:,.2f}").font = Font(
        bold=True, color="FF0000"
    )

    # ===== 상세 시트들 =====
    _create_nat_sheet(wb, result.nat_findings, header_fill, header_font)
    _create_eni_sheet(wb, result.eni_results, header_fill, header_font)
    _create_ebs_sheet(wb, result.ebs_results, header_fill, header_font)
    _create_eip_sheet(wb, result.eip_results, header_fill, header_font)
    _create_elb_sheet(wb, result.elb_results, header_fill, header_font)
    _create_snap_sheet(wb, result.snap_results, header_fill, header_font)
    _create_ami_sheet(wb, result.ami_results, header_fill, header_font)
    _create_rds_snap_sheet(wb, result.rds_snap_results, header_fill, header_font)
    _create_loggroup_sheet(wb, result.loggroup_results, header_fill, header_font)
    _create_tg_sheet(wb, result.tg_results, header_fill, header_font)
    _create_endpoint_sheet(wb, result.endpoint_results, header_fill, header_font)
    _create_secret_sheet(wb, result.secret_results, header_fill, header_font)
    _create_kms_sheet(wb, result.kms_results, header_fill, header_font)
    _create_ecr_sheet(wb, result.ecr_results, header_fill, header_font)
    _create_route53_sheet(wb, result.route53_results, header_fill, header_font)
    _create_s3_sheet(wb, result.s3_results, header_fill, header_font)
    _create_lambda_sheet(wb, result.lambda_results, header_fill, header_font)
    _create_elasticache_sheet(wb, result.elasticache_results, header_fill, header_font)
    _create_rds_instance_sheet(
        wb, result.rds_instance_results, header_fill, header_font
    )
    _create_efs_sheet(wb, result.efs_results, header_fill, header_font)
    _create_sqs_sheet(wb, result.sqs_results, header_fill, header_font)
    _create_sns_sheet(wb, result.sns_results, header_fill, header_font)
    _create_acm_sheet(wb, result.acm_results, header_fill, header_font)
    _create_apigateway_sheet(wb, result.apigateway_results, header_fill, header_font)
    _create_eventbridge_sheet(wb, result.eventbridge_results, header_fill, header_font)
    _create_cw_alarm_sheet(wb, result.cw_alarm_results, header_fill, header_font)

    # 열 너비 조정
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            col_idx = col[0].column
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(
                    max(max_len + 2, 10), 40
                )
        if sheet.title != "Summary":
            sheet.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Unused_Resources_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


def _create_nat_sheet(wb, findings, header_fill, header_font):
    ws = wb.create_sheet("NAT Gateway")
    ws.append(
        [
            "Account",
            "Region",
            "NAT ID",
            "Name",
            "Usage",
            "Monthly Waste",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for nat_result in findings:
        for f in nat_result.findings:
            if f.usage_status.value in ("unused", "low_usage"):
                ws.append(
                    [
                        f.nat.account_name,
                        f.nat.region,
                        f.nat.nat_gateway_id,
                        f.nat.name,
                        f.usage_status.value,
                        f"${f.monthly_waste:,.2f}",
                        f.recommendation,
                    ]
                )


def _create_eni_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("ENI")
    ws.append(
        ["Account", "Region", "ENI ID", "Name", "Usage", "Type", "Recommendation"]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value in ("unused", "pending"):
                ws.append(
                    [
                        f.eni.account_name,
                        f.eni.region,
                        f.eni.id,
                        f.eni.name,
                        f.usage_status.value,
                        f.eni.interface_type,
                        f.recommendation,
                    ]
                )


def _create_ebs_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("EBS")
    ws.append(
        [
            "Account",
            "Region",
            "Volume ID",
            "Name",
            "Type",
            "Size (GB)",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value in ("unused", "pending"):
                ws.append(
                    [
                        f.volume.account_name,
                        f.volume.region,
                        f.volume.id,
                        f.volume.name,
                        f.volume.volume_type,
                        f.volume.size_gb,
                        round(f.volume.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_eip_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("EIP")
    ws.append(
        [
            "Account",
            "Region",
            "Allocation ID",
            "Public IP",
            "Name",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value == "unused":
                ws.append(
                    [
                        f.eip.account_name,
                        f.eip.region,
                        f.eip.allocation_id,
                        f.eip.public_ip,
                        f.eip.name,
                        round(f.eip.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_elb_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("ELB")
    ws.append(
        [
            "Account",
            "Region",
            "Name",
            "Type",
            "Usage",
            "Targets",
            "Healthy",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value in ("unused", "unhealthy"):
                ws.append(
                    [
                        f.lb.account_name,
                        f.lb.region,
                        f.lb.name,
                        f.lb.lb_type.upper(),
                        f.usage_status.value,
                        f.lb.total_targets,
                        f.lb.healthy_targets,
                        round(f.lb.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_snap_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("EBS Snapshot")
    ws.append(
        [
            "Account",
            "Region",
            "Snapshot ID",
            "Name",
            "Usage",
            "Size (GB)",
            "Age (days)",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value in ("orphan", "old"):
                ws.append(
                    [
                        f.snapshot.account_name,
                        f.snapshot.region,
                        f.snapshot.id,
                        f.snapshot.name,
                        f.usage_status.value,
                        f.snapshot.volume_size_gb,
                        f.snapshot.age_days,
                        round(f.snapshot.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_ami_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("AMI")
    ws.append(
        [
            "Account",
            "Region",
            "AMI ID",
            "Name",
            "Size (GB)",
            "Age (days)",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value == "unused":
                ws.append(
                    [
                        f.ami.account_name,
                        f.ami.region,
                        f.ami.id,
                        f.ami.name,
                        f.ami.total_size_gb,
                        f.ami.age_days,
                        round(f.ami.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_rds_snap_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("RDS Snapshot")
    ws.append(
        [
            "Account",
            "Region",
            "Snapshot ID",
            "DB Identifier",
            "Type",
            "Engine",
            "Size (GB)",
            "Age (days)",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value == "old":
                ws.append(
                    [
                        f.snapshot.account_name,
                        f.snapshot.region,
                        f.snapshot.id,
                        f.snapshot.db_identifier,
                        f.snapshot.snapshot_type.value.upper(),
                        f.snapshot.engine,
                        f.snapshot.allocated_storage_gb,
                        f.snapshot.age_days,
                        round(f.snapshot.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_loggroup_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Log Group")
    ws.append(
        [
            "Account",
            "Region",
            "Log Group",
            "상태",
            "저장 (GB)",
            "보존 기간",
            "마지막 Ingestion",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                lg = f.log_group
                ws.append(
                    [
                        lg.account_name,
                        lg.region,
                        lg.name,
                        f.status.value,
                        round(lg.stored_gb, 4),
                        f"{lg.retention_days}일" if lg.retention_days else "무기한",
                        lg.last_ingestion_time.strftime("%Y-%m-%d")
                        if lg.last_ingestion_time
                        else "-",
                        round(lg.monthly_cost, 4),
                        f.recommendation,
                    ]
                )


def _create_tg_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Target Group")
    ws.append(
        [
            "Account",
            "Region",
            "Name",
            "상태",
            "Type",
            "Protocol",
            "Port",
            "LB 연결",
            "Total Targets",
            "Healthy",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                tg = f.tg
                ws.append(
                    [
                        tg.account_name,
                        tg.region,
                        tg.name,
                        f.status.value,
                        tg.target_type,
                        tg.protocol or "-",
                        tg.port or "-",
                        len(tg.load_balancer_arns),
                        tg.total_targets,
                        tg.healthy_targets,
                        f.recommendation,
                    ]
                )


def _create_endpoint_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("VPC Endpoint")
    ws.append(
        [
            "Account",
            "Region",
            "Endpoint ID",
            "Type",
            "Service",
            "VPC",
            "State",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                ep = f.endpoint
                ws.append(
                    [
                        ep.account_name,
                        ep.region,
                        ep.endpoint_id,
                        ep.endpoint_type,
                        ep.service_name.split(".")[-1],
                        ep.vpc_id,
                        ep.state,
                        round(ep.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_secret_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Secrets Manager")
    ws.append(["Account", "Region", "Name", "상태", "마지막 액세스", "월간 비용", "권장 조치"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                sec = f.secret
                last_access = (
                    sec.last_accessed_date.strftime("%Y-%m-%d")
                    if sec.last_accessed_date
                    else "없음"
                )
                ws.append(
                    [
                        sec.account_name,
                        sec.region,
                        sec.name,
                        f.status.value,
                        last_access,
                        round(sec.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_kms_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("KMS")
    ws.append(
        [
            "Account",
            "Region",
            "Key ID",
            "Description",
            "상태",
            "Manager",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                key = f.key
                ws.append(
                    [
                        key.account_name,
                        key.region,
                        key.key_id,
                        key.description[:50] if key.description else "-",
                        f.status.value,
                        key.key_manager,
                        round(key.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_ecr_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("ECR")
    ws.append(
        [
            "Account",
            "Region",
            "Repository",
            "상태",
            "이미지 수",
            "오래된 이미지",
            "총 크기",
            "낭비 비용",
            "Lifecycle",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                repo = f.repo
                ws.append(
                    [
                        repo.account_name,
                        repo.region,
                        repo.name,
                        f.status.value,
                        repo.image_count,
                        repo.old_image_count,
                        f"{repo.total_size_gb:.2f} GB",
                        f"${repo.old_images_monthly_cost:.2f}",
                        "있음" if repo.has_lifecycle_policy else "없음",
                        f.recommendation,
                    ]
                )


def _create_route53_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Route53")
    ws.append(["Account", "Zone ID", "Domain", "Type", "상태", "레코드 수", "월간 비용", "권장 조치"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                zone = f.zone
                ws.append(
                    [
                        zone.account_name,
                        zone.zone_id,
                        zone.name,
                        "Private" if zone.is_private else "Public",
                        f.status.value,
                        zone.record_count,
                        f"${zone.monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )


def _create_s3_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("S3")
    ws.append(
        [
            "Account",
            "Bucket",
            "Region",
            "상태",
            "객체 수",
            "크기",
            "버전관리",
            "Lifecycle",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                bucket = f.bucket
                ws.append(
                    [
                        bucket.account_name,
                        bucket.name,
                        bucket.region,
                        f.status.value,
                        bucket.object_count,
                        f"{bucket.total_size_mb:.2f} MB",
                        "Enabled" if bucket.versioning_enabled else "Disabled",
                        "있음" if bucket.has_lifecycle else "없음",
                        f.recommendation,
                    ]
                )


def _create_lambda_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Lambda")
    ws.append(
        [
            "Account",
            "Region",
            "Function Name",
            "Runtime",
            "Memory (MB)",
            "상태",
            "월간 낭비",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                fn = f.function
                ws.append(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        fn.runtime,
                        fn.memory_mb,
                        f.status.value,
                        f"${f.monthly_waste:.2f}" if f.monthly_waste > 0 else "-",
                        f.recommendation,
                    ]
                )


def _create_elasticache_sheet(wb, results, header_fill, header_font):
    """ElastiCache 상세 시트 생성"""
    ws = wb.create_sheet("ElastiCache")
    ws.append(
        [
            "Account",
            "Region",
            "Cluster ID",
            "Engine",
            "Node Type",
            "Nodes",
            "상태",
            "Avg Conn",
            "Avg CPU",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                c = f.cluster
                ws.append(
                    [
                        c.account_name,
                        c.region,
                        c.cluster_id,
                        c.engine,
                        c.node_type,
                        c.num_nodes,
                        f.status.value,
                        f"{c.avg_connections:.1f}",
                        f"{c.avg_cpu:.1f}%",
                        f"${c.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )


def _create_rds_instance_sheet(wb, results, header_fill, header_font):
    """RDS Instance 상세 시트 생성"""
    ws = wb.create_sheet("RDS Instance")
    ws.append(
        [
            "Account",
            "Region",
            "Instance ID",
            "Engine",
            "Class",
            "Storage",
            "Multi-AZ",
            "상태",
            "Avg Conn",
            "Avg CPU",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                inst = f.instance
                ws.append(
                    [
                        inst.account_name,
                        inst.region,
                        inst.db_instance_id,
                        inst.engine,
                        inst.db_instance_class,
                        f"{inst.allocated_storage} GB",
                        "Yes" if inst.multi_az else "No",
                        f.status.value,
                        f"{inst.avg_connections:.1f}",
                        f"{inst.avg_cpu:.1f}%",
                        f"${inst.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )


def _create_efs_sheet(wb, results, header_fill, header_font):
    """EFS 상세 시트 생성"""
    ws = wb.create_sheet("EFS")
    ws.append(
        [
            "Account",
            "Region",
            "ID",
            "Name",
            "Size",
            "Mount Targets",
            "Mode",
            "상태",
            "Avg Conn",
            "Total I/O",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                fs = f.efs
                ws.append(
                    [
                        fs.account_name,
                        fs.region,
                        fs.file_system_id,
                        fs.name or "-",
                        f"{fs.size_gb:.2f} GB",
                        fs.mount_target_count,
                        fs.throughput_mode,
                        f.status.value,
                        f"{fs.avg_client_connections:.1f}",
                        f"{fs.total_io_bytes / (1024**2):.1f} MB",
                        f"${fs.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )


def _create_sqs_sheet(wb, results, header_fill, header_font):
    """SQS 상세 시트 생성"""
    ws = wb.create_sheet("SQS")
    ws.append(
        [
            "Account",
            "Region",
            "Queue Name",
            "Type",
            "Messages",
            "상태",
            "Sent",
            "Received",
            "Deleted",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                q = f.queue
                queue_type = "FIFO" if q.is_fifo else "Standard"
                if q.is_dlq:
                    queue_type += " (DLQ)"
                ws.append(
                    [
                        q.account_name,
                        q.region,
                        q.queue_name,
                        queue_type,
                        q.approximate_messages,
                        f.status.value,
                        int(q.messages_sent),
                        int(q.messages_received),
                        int(q.messages_deleted),
                        f.recommendation,
                    ]
                )


def _create_sns_sheet(wb, results, header_fill, header_font):
    """SNS 상세 시트 생성"""
    ws = wb.create_sheet("SNS")
    ws.append(
        [
            "Account",
            "Region",
            "Topic Name",
            "Subscribers",
            "상태",
            "Published",
            "Delivered",
            "Failed",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                t = f.topic
                ws.append(
                    [
                        t.account_name,
                        t.region,
                        t.topic_name,
                        t.subscription_count,
                        f.status.value,
                        int(t.messages_published),
                        int(t.notifications_delivered),
                        int(t.notifications_failed),
                        f.recommendation,
                    ]
                )


def _create_acm_sheet(wb, results, header_fill, header_font):
    """ACM 상세 시트 생성"""
    ws = wb.create_sheet("ACM")
    ws.append(
        [
            "Account",
            "Region",
            "Domain",
            "Type",
            "Status",
            "Expiry",
            "Days Left",
            "In Use",
            "분석상태",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                c = f.cert
                ws.append(
                    [
                        c.account_name,
                        c.region,
                        c.domain_name,
                        c.cert_type,
                        c.status,
                        c.not_after.strftime("%Y-%m-%d") if c.not_after else "-",
                        c.days_until_expiry if c.days_until_expiry else "-",
                        len(c.in_use_by),
                        f.status.value,
                        f.recommendation,
                    ]
                )


def _create_apigateway_sheet(wb, results, header_fill, header_font):
    """API Gateway 상세 시트 생성"""
    ws = wb.create_sheet("API Gateway")
    ws.append(
        [
            "Account",
            "Region",
            "API Name",
            "Type",
            "Endpoint",
            "Stages",
            "Requests",
            "상태",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                a = f.api
                ws.append(
                    [
                        a.account_name,
                        a.region,
                        a.api_name,
                        a.api_type,
                        a.endpoint_type,
                        a.stage_count,
                        int(a.total_requests),
                        f.status.value,
                        f.recommendation,
                    ]
                )


def _create_eventbridge_sheet(wb, results, header_fill, header_font):
    """EventBridge 상세 시트 생성"""
    ws = wb.create_sheet("EventBridge")
    ws.append(
        [
            "Account",
            "Region",
            "Rule Name",
            "Event Bus",
            "State",
            "Schedule",
            "Targets",
            "Triggers",
            "상태",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                rule = f.rule
                ws.append(
                    [
                        rule.account_name,
                        rule.region,
                        rule.rule_name,
                        rule.event_bus_name,
                        rule.state,
                        rule.schedule_expression or "-",
                        rule.target_count,
                        int(rule.triggered_rules),
                        f.status.value,
                        f.recommendation,
                    ]
                )


def _create_cw_alarm_sheet(wb, results, header_fill, header_font):
    """CloudWatch Alarm 상세 시트 생성"""
    ws = wb.create_sheet("CloudWatch Alarm")
    ws.append(
        [
            "Account",
            "Region",
            "Alarm Name",
            "Namespace",
            "Metric",
            "Dimensions",
            "State",
            "분석상태",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                a = f.alarm
                ws.append(
                    [
                        a.account_name,
                        a.region,
                        a.alarm_name,
                        a.namespace,
                        a.metric_name,
                        a.dimensions,
                        a.state,
                        f.status.value,
                        f.recommendation,
                    ]
                )
