"""
plugins/sqs/unused.py - SQS 미사용 큐 분석

유휴/미사용 SQS 큐 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 분석 기간 (일) - AWS 권장 14일 이상
ANALYSIS_DAYS = 14

# 저사용 기준: 하루 평균 메시지 10개 미만
LOW_USAGE_MESSAGES_PER_DAY = 10

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "sqs:ListQueues",
        "sqs:GetQueueAttributes",
        "cloudwatch:GetMetricStatistics",
    ],
}


class QueueStatus(Enum):
    """큐 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"
    EMPTY_DLQ = "empty_dlq"


@dataclass
class SQSQueueInfo:
    """SQS 큐 정보"""

    account_id: str
    account_name: str
    region: str
    queue_name: str
    queue_url: str
    queue_arn: str
    is_fifo: bool
    is_dlq: bool
    approximate_messages: int
    approximate_messages_delayed: int
    approximate_messages_not_visible: int
    created_timestamp: datetime | None
    # CloudWatch 지표
    messages_sent: float = 0.0
    messages_received: float = 0.0
    messages_deleted: float = 0.0


@dataclass
class QueueFinding:
    """큐 분석 결과"""

    queue: SQSQueueInfo
    status: QueueStatus
    recommendation: str


@dataclass
class SQSAnalysisResult:
    """SQS 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_queues: int = 0
    unused_queues: int = 0
    low_usage_queues: int = 0
    empty_dlqs: int = 0
    normal_queues: int = 0
    findings: list[QueueFinding] = field(default_factory=list)


def collect_sqs_queues(session, account_id: str, account_name: str, region: str) -> list[SQSQueueInfo]:
    """SQS 큐 수집"""
    from botocore.exceptions import ClientError

    sqs = get_client(session, "sqs", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    queues = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        paginator = sqs.get_paginator("list_queues")
        for page in paginator.paginate():
            for queue_url in page.get("QueueUrls", []):
                try:
                    attrs = sqs.get_queue_attributes(
                        QueueUrl=queue_url,
                        AttributeNames=["All"],
                    ).get("Attributes", {})

                    queue_arn = attrs.get("QueueArn", "")
                    queue_name = queue_url.split("/")[-1]

                    # DLQ 여부 확인
                    is_dlq = "dlq" in queue_name.lower() or "dead" in queue_name.lower()

                    # 생성 시간
                    created_ts = attrs.get("CreatedTimestamp")
                    created_at = None
                    if created_ts:
                        created_at = datetime.fromtimestamp(int(created_ts), tz=timezone.utc)

                    info = SQSQueueInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        queue_name=queue_name,
                        queue_url=queue_url,
                        queue_arn=queue_arn,
                        is_fifo=queue_name.endswith(".fifo"),
                        is_dlq=is_dlq,
                        approximate_messages=int(attrs.get("ApproximateNumberOfMessages", 0)),
                        approximate_messages_delayed=int(attrs.get("ApproximateNumberOfMessagesDelayed", 0)),
                        approximate_messages_not_visible=int(attrs.get("ApproximateNumberOfMessagesNotVisible", 0)),
                        created_timestamp=created_at,
                    )

                    # CloudWatch 지표 조회
                    try:
                        # NumberOfMessagesSent
                        sent_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/SQS",
                            MetricName="NumberOfMessagesSent",
                            Dimensions=[{"Name": "QueueName", "Value": queue_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if sent_resp.get("Datapoints"):
                            info.messages_sent = sum(d["Sum"] for d in sent_resp["Datapoints"])

                        # NumberOfMessagesReceived
                        recv_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/SQS",
                            MetricName="NumberOfMessagesReceived",
                            Dimensions=[{"Name": "QueueName", "Value": queue_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if recv_resp.get("Datapoints"):
                            info.messages_received = sum(d["Sum"] for d in recv_resp["Datapoints"])

                        # NumberOfMessagesDeleted
                        del_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/SQS",
                            MetricName="NumberOfMessagesDeleted",
                            Dimensions=[{"Name": "QueueName", "Value": queue_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if del_resp.get("Datapoints"):
                            info.messages_deleted = sum(d["Sum"] for d in del_resp["Datapoints"])

                    except ClientError:
                        pass

                    queues.append(info)

                except ClientError:
                    continue

    except ClientError:
        pass

    return queues


def analyze_queues(queues: list[SQSQueueInfo], account_id: str, account_name: str, region: str) -> SQSAnalysisResult:
    """SQS 큐 분석"""
    result = SQSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_queues=len(queues),
    )

    for queue in queues:
        total_activity = queue.messages_sent + queue.messages_received + queue.messages_deleted

        # 빈 DLQ
        if queue.is_dlq and queue.approximate_messages == 0 and total_activity == 0:
            result.empty_dlqs += 1
            result.findings.append(
                QueueFinding(
                    queue=queue,
                    status=QueueStatus.EMPTY_DLQ,
                    recommendation="빈 DLQ - 정리 검토",
                )
            )
            continue

        # 미사용 큐
        if total_activity == 0:
            result.unused_queues += 1
            result.findings.append(
                QueueFinding(
                    queue=queue,
                    status=QueueStatus.UNUSED,
                    recommendation=f"{ANALYSIS_DAYS}일간 활동 없음 - 삭제 검토",
                )
            )
            continue

        # 저사용 큐 (일평균 메시지 수 기준)
        avg_messages_per_day = (queue.messages_sent + queue.messages_received) / ANALYSIS_DAYS
        if avg_messages_per_day < LOW_USAGE_MESSAGES_PER_DAY:
            result.low_usage_queues += 1
            result.findings.append(
                QueueFinding(
                    queue=queue,
                    status=QueueStatus.LOW_USAGE,
                    recommendation=f"저사용 (일평균 {avg_messages_per_day:.1f}건) - 통합 검토",
                )
            )
            continue

        result.normal_queues += 1
        result.findings.append(
            QueueFinding(
                queue=queue,
                status=QueueStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[SQSAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary Sheet
    summary = wb.new_summary_sheet("Summary")
    summary.add_title("SQS 미사용 분석 보고서")

    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="빈DLQ", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary Data", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_queues,
                r.unused_queues,
                r.empty_dlqs,
                r.normal_queues,
            ]
        )
        ws = summary_sheet._ws
        if r.unused_queues > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.empty_dlqs > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail Sheet
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Queue Name", width=50),
        ColumnDef(header="Type", width=15),
        ColumnDef(header="Messages", width=12, style="number"),
        ColumnDef(header="상태", width=15),
        ColumnDef(header="Sent", width=12, style="number"),
        ColumnDef(header="Received", width=12, style="number"),
        ColumnDef(header="Deleted", width=12, style="number"),
        ColumnDef(header="권장 조치", width=30),
    ]
    detail_sheet = wb.new_sheet("Queues", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != QueueStatus.NORMAL:
                q = f.queue
                queue_type = "FIFO" if q.is_fifo else "Standard"
                if q.is_dlq:
                    queue_type += " (DLQ)"
                detail_sheet.add_row(
                    [
                        q.account_name,
                        q.region,
                        q.queue_name,
                        queue_type,
                        q.approximate_messages,
                        f.status.value,
                        int(q.messages_sent),
                        int(q.messages_received),
                        int(q.messages_deleted),
                        f.recommendation,
                    ]
                )

    return str(wb.save_as(output_dir, "SQS_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SQSAnalysisResult | None:
    """단일 계정/리전의 SQS 큐 수집 및 분석 (병렬 실행용)"""
    queues = collect_sqs_queues(session, account_id, account_name, region)
    if not queues:
        return None
    return analyze_queues(queues, account_id, account_name, region)


def run(ctx) -> None:
    """SQS 미사용 큐 분석"""
    console.print("[bold]SQS 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="sqs")
    results: list[SQSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_queues for r in results)
    total_empty_dlq = sum(r.empty_dlqs for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미사용 큐: [red]{total_unused}개[/red] / 빈 DLQ: [yellow]{total_empty_dlq}개[/yellow]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("sqs", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
