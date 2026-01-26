"""
plugins/kms/audit.py - KMS 감사 보고서

KMS 키 보안 감사:
- 자동 로테이션 비활성화 CMK 탐지
- 위험한 키 정책 분석 (외부 계정 접근, * Principal 등)
- 권한 부여(Grants) 현황

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import TYPE_CHECKING

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

if TYPE_CHECKING:
    from cli.flow.context import ExecutionContext

console = Console()

REQUIRED_PERMISSIONS = {
    "read": [
        "kms:ListKeys",
        "kms:DescribeKey",
        "kms:ListAliases",
        "kms:GetKeyRotationStatus",
        "kms:GetKeyPolicy",
        "kms:ListGrants",
    ],
}


class RiskLevel(Enum):
    """위험도"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class PolicyFinding:
    """정책 분석 결과"""

    risk_level: RiskLevel
    issue: str
    detail: str


@dataclass
class GrantInfo:
    """권한 부여 정보"""

    grantee_principal: str
    operations: list[str]
    constraints: str


@dataclass
class KMSKeyAudit:
    """KMS 키 감사 결과"""

    key_id: str
    arn: str
    alias: str
    description: str
    key_state: str
    creation_date: datetime | None
    # 로테이션
    rotation_enabled: bool
    # 정책 분석
    policy_findings: list[PolicyFinding] = field(default_factory=list)
    # Grants
    grants: list[GrantInfo] = field(default_factory=list)

    @property
    def has_rotation_issue(self) -> bool:
        return not self.rotation_enabled

    @property
    def has_policy_issue(self) -> bool:
        return any(f.risk_level in (RiskLevel.HIGH, RiskLevel.MEDIUM) for f in self.policy_findings)

    @property
    def max_risk_level(self) -> RiskLevel:
        if not self.policy_findings:
            return RiskLevel.INFO
        levels = [f.risk_level for f in self.policy_findings]
        if RiskLevel.HIGH in levels:
            return RiskLevel.HIGH
        if RiskLevel.MEDIUM in levels:
            return RiskLevel.MEDIUM
        if RiskLevel.LOW in levels:
            return RiskLevel.LOW
        return RiskLevel.INFO


@dataclass
class KMSAuditResult:
    """KMS 감사 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_cmk: int = 0
    rotation_disabled: int = 0
    high_risk: int = 0
    medium_risk: int = 0
    total_grants: int = 0
    audits: list[KMSKeyAudit] = field(default_factory=list)


def analyze_key_policy(policy_str: str, account_id: str) -> list[PolicyFinding]:
    """키 정책 분석"""
    findings = []

    try:
        policy = json.loads(policy_str)
    except json.JSONDecodeError:
        findings.append(
            PolicyFinding(
                risk_level=RiskLevel.LOW,
                issue="정책 파싱 실패",
                detail="JSON 파싱 불가",
            )
        )
        return findings

    statements = policy.get("Statement", [])

    for stmt in statements:
        effect = stmt.get("Effect", "")
        principals = stmt.get("Principal", {})
        actions = stmt.get("Action", [])
        conditions = stmt.get("Condition", {})

        if effect != "Allow":
            continue

        # Principal 분석
        if isinstance(principals, str):
            principals = {"AWS": [principals]} if principals != "*" else {"AWS": "*"}
        elif isinstance(principals, dict) and "AWS" in principals and isinstance(principals["AWS"], str):
            principals["AWS"] = [principals["AWS"]]

        aws_principals = principals.get("AWS", [])

        # * Principal 체크
        if aws_principals == "*" or "*" in (aws_principals if isinstance(aws_principals, list) else []):
            # Condition으로 제한되어 있는지 확인
            if not conditions:
                findings.append(
                    PolicyFinding(
                        risk_level=RiskLevel.HIGH,
                        issue="무제한 Principal (*)",
                        detail="Condition 없이 모든 AWS 계정에 접근 허용",
                    )
                )
            else:
                findings.append(
                    PolicyFinding(
                        risk_level=RiskLevel.MEDIUM,
                        issue="광범위 Principal (*)",
                        detail=f"Condition으로 제한됨: {list(conditions.keys())}",
                    )
                )
            continue

        # 외부 계정 접근 체크
        if isinstance(aws_principals, list):
            for principal in aws_principals:
                if isinstance(principal, str) and "arn:aws" in principal:
                    # ARN에서 계정 ID 추출
                    parts = principal.split(":")
                    if len(parts) >= 5:
                        principal_account = parts[4]
                        if principal_account and principal_account != account_id:
                            findings.append(
                                PolicyFinding(
                                    risk_level=RiskLevel.MEDIUM,
                                    issue="외부 계정 접근 허용",
                                    detail=f"계정 {principal_account}에 접근 허용",
                                )
                            )

        # Action 분석
        if isinstance(actions, str):
            actions = [actions]

        dangerous_actions = ["kms:*", "*"]
        for action in actions:
            if action in dangerous_actions:
                findings.append(
                    PolicyFinding(
                        risk_level=RiskLevel.MEDIUM,
                        issue="광범위 Action 허용",
                        detail=f"Action: {action}",
                    )
                )
                break

    return findings


def collect_kms_audit(session, account_id: str, account_name: str, region: str) -> list[KMSKeyAudit]:
    """KMS 키 감사 정보 수집"""
    from botocore.exceptions import ClientError

    kms = get_client(session, "kms", region_name=region)
    audits = []

    # 별칭 맵
    aliases_map: dict[str, str] = {}
    try:
        paginator = kms.get_paginator("list_aliases")
        for page in paginator.paginate():
            for alias in page.get("Aliases", []):
                if "TargetKeyId" in alias and not alias["AliasName"].startswith("alias/aws/"):
                    aliases_map[alias["TargetKeyId"]] = alias["AliasName"]
    except ClientError:
        pass

    # 키 수집
    try:
        paginator = kms.get_paginator("list_keys")
        for page in paginator.paginate():
            for key in page.get("Keys", []):
                key_id = key["KeyId"]

                try:
                    key_info = kms.describe_key(KeyId=key_id)["KeyMetadata"]

                    # CMK만 분석
                    if key_info.get("KeyManager") != "CUSTOMER":
                        continue

                    # 활성 키만
                    if key_info.get("KeyState") != "Enabled":
                        continue

                    # 로테이션 상태
                    rotation_enabled = False
                    try:
                        rotation = kms.get_key_rotation_status(KeyId=key_id)
                        rotation_enabled = rotation.get("KeyRotationEnabled", False)
                    except ClientError:
                        pass

                    # 키 정책 분석
                    policy_findings = []
                    try:
                        policy_resp = kms.get_key_policy(KeyId=key_id, PolicyName="default")
                        policy_str = policy_resp.get("Policy", "{}")
                        policy_findings = analyze_key_policy(policy_str, account_id)
                    except ClientError:
                        pass

                    # Grants 수집
                    grants = []
                    try:
                        grant_paginator = kms.get_paginator("list_grants")
                        for grant_page in grant_paginator.paginate(KeyId=key_id):
                            for grant in grant_page.get("Grants", []):
                                grants.append(
                                    GrantInfo(
                                        grantee_principal=grant.get("GranteePrincipal", ""),
                                        operations=grant.get("Operations", []),
                                        constraints=str(grant.get("Constraints", {}))
                                        if grant.get("Constraints")
                                        else "",
                                    )
                                )
                    except ClientError:
                        pass

                    audits.append(
                        KMSKeyAudit(
                            key_id=key_id,
                            arn=key_info.get("Arn", ""),
                            alias=aliases_map.get(key_id, ""),
                            description=key_info.get("Description", ""),
                            key_state=key_info.get("KeyState", ""),
                            creation_date=key_info.get("CreationDate"),
                            rotation_enabled=rotation_enabled,
                            policy_findings=policy_findings,
                            grants=grants,
                        )
                    )

                except ClientError:
                    continue

    except ClientError:
        pass

    return audits


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> KMSAuditResult | None:
    """병렬 실행용"""
    audits = collect_kms_audit(session, account_id, account_name, region)

    if not audits:
        return None

    result = KMSAuditResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_cmk=len(audits),
        audits=audits,
    )

    for audit in audits:
        if audit.has_rotation_issue:
            result.rotation_disabled += 1
        if audit.max_risk_level == RiskLevel.HIGH:
            result.high_risk += 1
        elif audit.max_risk_level == RiskLevel.MEDIUM:
            result.medium_risk += 1
        result.total_grants += len(audit.grants)

    return result


def generate_report(results: list[KMSAuditResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="CMK 수", width=10, style="number"),
        ColumnDef(header="로테이션 미설정", width=15, style="number"),
        ColumnDef(header="High Risk", width=12, style="number"),
        ColumnDef(header="Medium Risk", width=12, style="number"),
        ColumnDef(header="Grants", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_style = None
        if r.high_risk > 0:
            row_style = Styles.danger()
        elif r.medium_risk > 0 or r.rotation_disabled > 0:
            row_style = Styles.warning()
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_cmk,
                r.rotation_disabled,
                r.high_risk,
                r.medium_risk,
                r.total_grants,
            ],
            style=row_style,
        )
        # Cell-level highlighting
        ws = summary_sheet._ws
        if r.rotation_disabled > 0:
            ws.cell(row=row_num, column=4).fill = yellow_fill
        if r.high_risk > 0:
            ws.cell(row=row_num, column=5).fill = red_fill
        if r.medium_risk > 0:
            ws.cell(row=row_num, column=6).fill = orange_fill

    # Rotation 시트
    rotation_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Key ID", width=25),
        ColumnDef(header="Alias", width=25),
        ColumnDef(header="로테이션", width=12, style="center"),
        ColumnDef(header="생성일", width=15, style="center"),
    ]
    rotation_sheet = wb.new_sheet("Rotation", rotation_columns)

    for r in results:
        for audit in r.audits:
            row_style = Styles.warning() if not audit.rotation_enabled else None
            row_num = rotation_sheet.add_row(
                [
                    r.account_name,
                    r.region,
                    audit.key_id[:20] + "...",
                    audit.alias or "-",
                    "활성화" if audit.rotation_enabled else "비활성화",
                    audit.creation_date.strftime("%Y-%m-%d") if audit.creation_date else "-",
                ],
                style=row_style,
            )
            # Cell-level highlighting for rotation status
            if not audit.rotation_enabled:
                ws = rotation_sheet._ws
                ws.cell(row=row_num, column=5).fill = yellow_fill

    # Policy Findings 시트
    policy_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Key ID", width=25),
        ColumnDef(header="Alias", width=25),
        ColumnDef(header="Risk", width=12, style="center"),
        ColumnDef(header="Issue", width=25),
        ColumnDef(header="Detail", width=50),
    ]
    policy_sheet = wb.new_sheet("Policy Issues", policy_columns)

    for r in results:
        for audit in r.audits:
            for finding in audit.policy_findings:
                row_style = None
                if finding.risk_level == RiskLevel.HIGH:
                    row_style = Styles.danger()
                elif finding.risk_level == RiskLevel.MEDIUM:
                    row_style = Styles.warning()
                row_num = policy_sheet.add_row(
                    [
                        r.account_name,
                        r.region,
                        audit.key_id[:20] + "...",
                        audit.alias or "-",
                        finding.risk_level.value.upper(),
                        finding.issue,
                        finding.detail,
                    ],
                    style=row_style,
                )
                # Cell-level highlighting for risk level
                ws = policy_sheet._ws
                if finding.risk_level == RiskLevel.HIGH:
                    ws.cell(row=row_num, column=5).fill = red_fill
                elif finding.risk_level == RiskLevel.MEDIUM:
                    ws.cell(row=row_num, column=5).fill = orange_fill

    # Grants 시트
    grants_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Key ID", width=25),
        ColumnDef(header="Alias", width=25),
        ColumnDef(header="Grantee", width=60),
        ColumnDef(header="Operations", width=40),
        ColumnDef(header="Constraints", width=30),
    ]
    grants_sheet = wb.new_sheet("Grants", grants_columns)

    for r in results:
        for audit in r.audits:
            for grant in audit.grants:
                # Grantee ARN 줄이기
                grantee = grant.grantee_principal
                if len(grantee) > 60:
                    grantee = "..." + grantee[-57:]
                grants_sheet.add_row(
                    [
                        r.account_name,
                        r.region,
                        audit.key_id[:20] + "...",
                        audit.alias or "-",
                        grantee,
                        ", ".join(grant.operations),
                        grant.constraints or "-",
                    ],
                )

    return str(wb.save_as(output_dir, "KMS_Audit"))


def run(ctx: ExecutionContext) -> None:
    """KMS 감사 보고서"""
    console.print("[bold]KMS 감사 시작...[/bold]\n")
    console.print("[dim]분석 항목: 키 로테이션, 키 정책, 권한 부여(Grants)[/dim]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=10, service="kms")
    results: list[KMSAuditResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_cmk = sum(r.total_cmk for r in results)
    total_rotation = sum(r.rotation_disabled for r in results)
    total_high = sum(r.high_risk for r in results)
    total_medium = sum(r.medium_risk for r in results)
    total_grants = sum(r.total_grants for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"CMK: {total_cmk}개")
    console.print(f"로테이션 미설정: [yellow]{total_rotation}개[/yellow]")
    console.print(f"정책 위험: [red]High {total_high}개[/red] / [yellow]Medium {total_medium}개[/yellow]")
    console.print(f"Grants: {total_grants}건")

    # 보고서 생성
    identifier = ctx.accounts[0].id if ctx.accounts else ctx.profile_name or "default"
    output_path = OutputPath(identifier).sub("kms", "security").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
