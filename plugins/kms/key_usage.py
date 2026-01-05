"""
plugins/kms/key_usage.py - CMK 사용처 분석

고객 관리 키(CMK)가 사용되는 AWS 리소스를 찾아서 매핑합니다.

지원 서비스 (26개):
- 스토리지: S3, EBS, EFS, FSx, ECR
- 데이터베이스: RDS, DynamoDB, ElastiCache, Redshift, DocumentDB, Neptune, MemoryDB
- 메시징/스트리밍: SNS, SQS, Kinesis, MSK
- 컴퓨팅: Lambda, EKS
- 보안/관리: Secrets Manager, Backup, CloudWatch Logs, SSM Parameter Store
- 분석/ML: OpenSearch, Glue, SageMaker, Athena

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

REQUIRED_PERMISSIONS = {
    "read": [
        "kms:ListKeys",
        "kms:DescribeKey",
        "kms:ListAliases",
        "s3:ListBuckets",
        "s3:GetBucketEncryption",
        "ec2:DescribeVolumes",
        "rds:DescribeDBInstances",
        "rds:DescribeDBClusters",
        "secretsmanager:ListSecrets",
        "lambda:ListFunctions",
        "sns:ListTopics",
        "sns:GetTopicAttributes",
        "sqs:ListQueues",
        "sqs:GetQueueAttributes",
        "es:ListDomainNames",
        "es:DescribeDomain",
        "elasticfilesystem:DescribeFileSystems",
        "dynamodb:ListTables",
        "dynamodb:DescribeTable",
        "elasticache:DescribeReplicationGroups",
        "kinesis:ListStreams",
        "kinesis:DescribeStream",
        "redshift:DescribeClusters",
        "fsx:DescribeFileSystems",
        "logs:DescribeLogGroups",
        "docdb-elastic:ListClusters",
        "rds:DescribeDBClusters",
        "neptune:DescribeDBClusters",
        "backup:ListBackupVaults",
        "glue:GetDataCatalogEncryptionSettings",
        "kafka:ListClusters",
        "sagemaker:ListNotebookInstances",
        "sagemaker:DescribeNotebookInstance",
        "eks:ListClusters",
        "eks:DescribeCluster",
        "memorydb:DescribeClusters",
        "ecr:DescribeRepositories",
        "athena:ListWorkGroups",
        "athena:GetWorkGroup",
        "ssm:DescribeParameters",
    ],
}


@dataclass
class KMSKeyInfo:
    """KMS 키 정보"""

    key_id: str
    arn: str
    alias: str
    key_manager: str
    key_state: str
    description: str


@dataclass
class ResourceUsage:
    """KMS 키를 사용하는 리소스"""

    service: str  # S3, EBS, RDS, SecretsManager, Lambda, SNS, SQS
    resource_type: str  # bucket, volume, instance, secret, function, topic, queue
    resource_id: str
    resource_name: str


@dataclass
class KMSKeyUsage:
    """KMS 키별 사용처"""

    key: KMSKeyInfo
    usages: list[ResourceUsage] = field(default_factory=list)

    @property
    def usage_count(self) -> int:
        return len(self.usages)

    @property
    def is_unused(self) -> bool:
        return self.usage_count == 0


@dataclass
class KMSUsageResult:
    """KMS 사용처 분석 결과"""

    account_id: str
    account_name: str
    region: str
    total_keys: int = 0
    customer_keys: int = 0
    unused_keys: int = 0
    key_usages: list[KMSKeyUsage] = field(default_factory=list)


def collect_kms_keys(kms_client) -> list[KMSKeyInfo]:
    """KMS 키 목록 수집"""
    from botocore.exceptions import ClientError

    keys = []
    aliases_map: dict[str, str] = {}

    # 별칭 먼저 수집
    try:
        paginator = kms_client.get_paginator("list_aliases")
        for page in paginator.paginate():
            for alias in page.get("Aliases", []):
                if "TargetKeyId" in alias and not alias["AliasName"].startswith("alias/aws/"):
                    aliases_map[alias["TargetKeyId"]] = alias["AliasName"]
    except ClientError:
        pass

    # 키 수집
    try:
        paginator = kms_client.get_paginator("list_keys")
        for page in paginator.paginate():
            for key in page.get("Keys", []):
                try:
                    key_info = kms_client.describe_key(KeyId=key["KeyId"])["KeyMetadata"]
                    keys.append(
                        KMSKeyInfo(
                            key_id=key["KeyId"],
                            arn=key_info.get("Arn", ""),
                            alias=aliases_map.get(key["KeyId"], ""),
                            key_manager=key_info.get("KeyManager", ""),
                            key_state=key_info.get("KeyState", ""),
                            description=key_info.get("Description", ""),
                        )
                    )
                except ClientError:
                    continue
    except ClientError:
        pass

    return keys


def find_s3_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """S3 버킷에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    s3 = get_client(session, "s3", region_name=region)

    try:
        buckets = s3.list_buckets().get("Buckets", [])
        for bucket in buckets:
            bucket_name = bucket["Name"]
            try:
                # 버킷 리전 확인 (해당 리전 버킷만)
                location = s3.get_bucket_location(Bucket=bucket_name).get("LocationConstraint")
                bucket_region = location or "us-east-1"
                if bucket_region != region:
                    continue

                encryption = s3.get_bucket_encryption(Bucket=bucket_name)
                rules = encryption.get("ServerSideEncryptionConfiguration", {}).get("Rules", [])
                for rule in rules:
                    sse = rule.get("ApplyServerSideEncryptionByDefault", {})
                    kms_key = sse.get("KMSMasterKeyID", "")
                    if kms_key and kms_key in key_arns:
                        usages[kms_key].append(
                            ResourceUsage(
                                service="S3",
                                resource_type="bucket",
                                resource_id=bucket_name,
                                resource_name=bucket_name,
                            )
                        )
            except ClientError:
                continue
    except ClientError:
        pass

    return usages


def find_ebs_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """EBS 볼륨에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    ec2 = get_client(session, "ec2", region_name=region)

    try:
        paginator = ec2.get_paginator("describe_volumes")
        for page in paginator.paginate():
            for volume in page.get("Volumes", []):
                kms_key = volume.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    vol_id = volume["VolumeId"]
                    name = ""
                    for tag in volume.get("Tags", []):
                        if tag["Key"] == "Name":
                            name = tag["Value"]
                            break
                    usages[kms_key].append(
                        ResourceUsage(
                            service="EBS",
                            resource_type="volume",
                            resource_id=vol_id,
                            resource_name=name or vol_id,
                        )
                    )
    except ClientError:
        pass

    return usages


def find_rds_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """RDS에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    rds = get_client(session, "rds", region_name=region)

    # DB 인스턴스
    try:
        paginator = rds.get_paginator("describe_db_instances")
        for page in paginator.paginate():
            for db in page.get("DBInstances", []):
                kms_key = db.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="RDS",
                            resource_type="instance",
                            resource_id=db["DBInstanceIdentifier"],
                            resource_name=db["DBInstanceIdentifier"],
                        )
                    )
    except ClientError:
        pass

    # DB 클러스터 (Aurora)
    try:
        paginator = rds.get_paginator("describe_db_clusters")
        for page in paginator.paginate():
            for cluster in page.get("DBClusters", []):
                kms_key = cluster.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="RDS",
                            resource_type="cluster",
                            resource_id=cluster["DBClusterIdentifier"],
                            resource_name=cluster["DBClusterIdentifier"],
                        )
                    )
    except ClientError:
        pass

    return usages


def find_secrets_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """Secrets Manager에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    sm = get_client(session, "secretsmanager", region_name=region)

    try:
        paginator = sm.get_paginator("list_secrets")
        for page in paginator.paginate():
            for secret in page.get("SecretList", []):
                kms_key = secret.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="SecretsManager",
                            resource_type="secret",
                            resource_id=secret.get("ARN", ""),
                            resource_name=secret.get("Name", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_lambda_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """Lambda에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    lambda_client = get_client(session, "lambda", region_name=region)

    try:
        paginator = lambda_client.get_paginator("list_functions")
        for page in paginator.paginate():
            for func in page.get("Functions", []):
                kms_key = func.get("KMSKeyArn", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="Lambda",
                            resource_type="function",
                            resource_id=func.get("FunctionArn", ""),
                            resource_name=func.get("FunctionName", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_sns_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """SNS에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    sns = get_client(session, "sns", region_name=region)

    try:
        paginator = sns.get_paginator("list_topics")
        for page in paginator.paginate():
            for topic in page.get("Topics", []):
                topic_arn = topic["TopicArn"]
                try:
                    attrs = sns.get_topic_attributes(TopicArn=topic_arn).get("Attributes", {})
                    kms_key = attrs.get("KmsMasterKeyId", "")
                    if kms_key and kms_key in key_arns:
                        topic_name = topic_arn.split(":")[-1]
                        usages[kms_key].append(
                            ResourceUsage(
                                service="SNS",
                                resource_type="topic",
                                resource_id=topic_arn,
                                resource_name=topic_name,
                            )
                        )
                except ClientError:
                    continue
    except ClientError:
        pass

    return usages


def find_sqs_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """SQS에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    sqs = get_client(session, "sqs", region_name=region)

    try:
        queues = sqs.list_queues().get("QueueUrls", [])
        for queue_url in queues:
            try:
                attrs = sqs.get_queue_attributes(QueueUrl=queue_url, AttributeNames=["KmsMasterKeyId"]).get(
                    "Attributes", {}
                )
                kms_key = attrs.get("KmsMasterKeyId", "")
                if kms_key and kms_key in key_arns:
                    queue_name = queue_url.split("/")[-1]
                    usages[kms_key].append(
                        ResourceUsage(
                            service="SQS",
                            resource_type="queue",
                            resource_id=queue_url,
                            resource_name=queue_name,
                        )
                    )
            except ClientError:
                continue
    except ClientError:
        pass

    return usages


def find_opensearch_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """OpenSearch (Elasticsearch)에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    es = get_client(session, "opensearch", region_name=region)

    try:
        domains = es.list_domain_names().get("DomainNames", [])
        for domain in domains:
            domain_name = domain.get("DomainName", "")
            try:
                info = es.describe_domain(DomainName=domain_name).get("DomainStatus", {})
                encryption = info.get("EncryptionAtRestOptions", {})
                if encryption.get("Enabled"):
                    kms_key = encryption.get("KmsKeyId", "")
                    if kms_key and kms_key in key_arns:
                        usages[kms_key].append(
                            ResourceUsage(
                                service="OpenSearch",
                                resource_type="domain",
                                resource_id=info.get("ARN", ""),
                                resource_name=domain_name,
                            )
                        )
            except ClientError:
                continue
    except ClientError:
        pass

    return usages


def find_efs_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """EFS에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    efs = get_client(session, "efs", region_name=region)

    try:
        paginator = efs.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page.get("FileSystems", []):
                kms_key = fs.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    fs_id = fs.get("FileSystemId", "")
                    name = ""
                    for tag in fs.get("Tags", []):
                        if tag.get("Key") == "Name":
                            name = tag.get("Value", "")
                            break
                    usages[kms_key].append(
                        ResourceUsage(
                            service="EFS",
                            resource_type="filesystem",
                            resource_id=fs_id,
                            resource_name=name or fs_id,
                        )
                    )
    except ClientError:
        pass

    return usages


def find_dynamodb_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """DynamoDB에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    dynamodb = get_client(session, "dynamodb", region_name=region)

    try:
        paginator = dynamodb.get_paginator("list_tables")
        for page in paginator.paginate():
            for table_name in page.get("TableNames", []):
                try:
                    table = dynamodb.describe_table(TableName=table_name).get("Table", {})
                    sse = table.get("SSEDescription", {})
                    kms_key = sse.get("KMSMasterKeyArn", "")
                    if kms_key and kms_key in key_arns:
                        usages[kms_key].append(
                            ResourceUsage(
                                service="DynamoDB",
                                resource_type="table",
                                resource_id=table.get("TableArn", ""),
                                resource_name=table_name,
                            )
                        )
                except ClientError:
                    continue
    except ClientError:
        pass

    return usages


def find_elasticache_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """ElastiCache에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    elasticache = get_client(session, "elasticache", region_name=region)

    try:
        paginator = elasticache.get_paginator("describe_replication_groups")
        for page in paginator.paginate():
            for rg in page.get("ReplicationGroups", []):
                kms_key = rg.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="ElastiCache",
                            resource_type="replication-group",
                            resource_id=rg.get("ARN", ""),
                            resource_name=rg.get("ReplicationGroupId", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_kinesis_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """Kinesis에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    kinesis = get_client(session, "kinesis", region_name=region)

    try:
        paginator = kinesis.get_paginator("list_streams")
        for page in paginator.paginate():
            for stream_name in page.get("StreamNames", []):
                try:
                    stream = kinesis.describe_stream(StreamName=stream_name).get("StreamDescription", {})
                    kms_key = stream.get("KeyId", "")
                    if kms_key and kms_key in key_arns:
                        usages[kms_key].append(
                            ResourceUsage(
                                service="Kinesis",
                                resource_type="stream",
                                resource_id=stream.get("StreamARN", ""),
                                resource_name=stream_name,
                            )
                        )
                except ClientError:
                    continue
    except ClientError:
        pass

    return usages


def find_redshift_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """Redshift에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    redshift = get_client(session, "redshift", region_name=region)

    try:
        paginator = redshift.get_paginator("describe_clusters")
        for page in paginator.paginate():
            for cluster in page.get("Clusters", []):
                kms_key = cluster.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="Redshift",
                            resource_type="cluster",
                            resource_id=cluster.get("ClusterIdentifier", ""),
                            resource_name=cluster.get("ClusterIdentifier", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_fsx_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """FSx에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    fsx = get_client(session, "fsx", region_name=region)

    try:
        paginator = fsx.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page.get("FileSystems", []):
                kms_key = fs.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    fs_id = fs.get("FileSystemId", "")
                    fs_type = fs.get("FileSystemType", "")
                    name = ""
                    for tag in fs.get("Tags", []):
                        if tag.get("Key") == "Name":
                            name = tag.get("Value", "")
                            break
                    usages[kms_key].append(
                        ResourceUsage(
                            service="FSx",
                            resource_type=f"filesystem-{fs_type}",
                            resource_id=fs_id,
                            resource_name=name or fs_id,
                        )
                    )
    except ClientError:
        pass

    return usages


def find_cloudwatch_logs_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """CloudWatch Logs에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    logs = get_client(session, "logs", region_name=region)

    try:
        paginator = logs.get_paginator("describe_log_groups")
        for page in paginator.paginate():
            for lg in page.get("logGroups", []):
                kms_key = lg.get("kmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="CloudWatch Logs",
                            resource_type="log-group",
                            resource_id=lg.get("arn", ""),
                            resource_name=lg.get("logGroupName", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_documentdb_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """DocumentDB에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    docdb = get_client(session, "docdb", region_name=region)

    try:
        paginator = docdb.get_paginator("describe_db_clusters")
        for page in paginator.paginate():
            for cluster in page.get("DBClusters", []):
                # DocumentDB 클러스터만 필터링 (Engine이 docdb인 것)
                if cluster.get("Engine") != "docdb":
                    continue
                kms_key = cluster.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="DocumentDB",
                            resource_type="cluster",
                            resource_id=cluster.get("DBClusterArn", ""),
                            resource_name=cluster.get("DBClusterIdentifier", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_neptune_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """Neptune에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    neptune = get_client(session, "neptune", region_name=region)

    try:
        paginator = neptune.get_paginator("describe_db_clusters")
        for page in paginator.paginate():
            for cluster in page.get("DBClusters", []):
                kms_key = cluster.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="Neptune",
                            resource_type="cluster",
                            resource_id=cluster.get("DBClusterArn", ""),
                            resource_name=cluster.get("DBClusterIdentifier", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_backup_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """AWS Backup에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    backup = get_client(session, "backup", region_name=region)

    try:
        paginator = backup.get_paginator("list_backup_vaults")
        for page in paginator.paginate():
            for vault in page.get("BackupVaultList", []):
                kms_key = vault.get("EncryptionKeyArn", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="Backup",
                            resource_type="vault",
                            resource_id=vault.get("BackupVaultArn", ""),
                            resource_name=vault.get("BackupVaultName", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_glue_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """Glue Data Catalog에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    glue = get_client(session, "glue", region_name=region)

    try:
        response = glue.get_data_catalog_encryption_settings()
        settings = response.get("DataCatalogEncryptionSettings", {})

        # Encryption at rest
        enc_at_rest = settings.get("EncryptionAtRest", {})
        if enc_at_rest.get("CatalogEncryptionMode") == "SSE-KMS":
            kms_key = enc_at_rest.get("SseAwsKmsKeyId", "")
            if kms_key and kms_key in key_arns:
                usages[kms_key].append(
                    ResourceUsage(
                        service="Glue",
                        resource_type="data-catalog",
                        resource_id=f"arn:aws:glue:{region}:data-catalog",
                        resource_name="Data Catalog (Encryption at Rest)",
                    )
                )

        # Connection password encryption
        conn_enc = settings.get("ConnectionPasswordEncryption", {})
        if conn_enc.get("ReturnConnectionPasswordEncrypted"):
            kms_key = conn_enc.get("AwsKmsKeyId", "")
            if kms_key and kms_key in key_arns:
                usages[kms_key].append(
                    ResourceUsage(
                        service="Glue",
                        resource_type="data-catalog",
                        resource_id=f"arn:aws:glue:{region}:data-catalog",
                        resource_name="Data Catalog (Connection Password)",
                    )
                )
    except ClientError:
        pass

    return usages


def find_msk_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """MSK (Managed Kafka)에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    kafka = get_client(session, "kafka", region_name=region)

    try:
        paginator = kafka.get_paginator("list_clusters_v2")
        for page in paginator.paginate():
            for cluster in page.get("ClusterInfoList", []):
                # Provisioned cluster
                provisioned = cluster.get("Provisioned", {})
                enc_info = provisioned.get("EncryptionInfo", {})
                enc_at_rest = enc_info.get("EncryptionAtRest", {})
                kms_key = enc_at_rest.get("DataVolumeKMSKeyId", "")

                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="MSK",
                            resource_type="cluster",
                            resource_id=cluster.get("ClusterArn", ""),
                            resource_name=cluster.get("ClusterName", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_sagemaker_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """SageMaker에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    sagemaker = get_client(session, "sagemaker", region_name=region)

    # Notebook Instances
    try:
        paginator = sagemaker.get_paginator("list_notebook_instances")
        for page in paginator.paginate():
            for nb in page.get("NotebookInstances", []):
                nb_name = nb.get("NotebookInstanceName", "")
                try:
                    detail = sagemaker.describe_notebook_instance(NotebookInstanceName=nb_name)
                    kms_key = detail.get("KmsKeyId", "")
                    if kms_key and kms_key in key_arns:
                        usages[kms_key].append(
                            ResourceUsage(
                                service="SageMaker",
                                resource_type="notebook",
                                resource_id=detail.get("NotebookInstanceArn", ""),
                                resource_name=nb_name,
                            )
                        )
                except ClientError:
                    continue
    except ClientError:
        pass

    # Training Jobs (최근 것만)
    try:
        response = sagemaker.list_training_jobs(MaxResults=100, SortBy="CreationTime", SortOrder="Descending")
        for job in response.get("TrainingJobSummaries", []):
            job_name = job.get("TrainingJobName", "")
            try:
                detail = sagemaker.describe_training_job(TrainingJobName=job_name)
                output_config = detail.get("OutputDataConfig", {})
                kms_key = output_config.get("KmsKeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="SageMaker",
                            resource_type="training-job",
                            resource_id=detail.get("TrainingJobArn", ""),
                            resource_name=job_name,
                        )
                    )
            except ClientError:
                continue
    except ClientError:
        pass

    return usages


def find_eks_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """EKS에서 KMS 키 사용처 찾기 (Secrets 암호화)"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    eks = get_client(session, "eks", region_name=region)

    try:
        paginator = eks.get_paginator("list_clusters")
        for page in paginator.paginate():
            for cluster_name in page.get("clusters", []):
                try:
                    cluster = eks.describe_cluster(name=cluster_name).get("cluster", {})
                    enc_config = cluster.get("encryptionConfig", [])
                    for enc in enc_config:
                        provider = enc.get("provider", {})
                        kms_key = provider.get("keyArn", "")
                        if kms_key and kms_key in key_arns:
                            usages[kms_key].append(
                                ResourceUsage(
                                    service="EKS",
                                    resource_type="cluster",
                                    resource_id=cluster.get("arn", ""),
                                    resource_name=cluster_name,
                                )
                            )
                except ClientError:
                    continue
    except ClientError:
        pass

    return usages


def find_memorydb_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """MemoryDB에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    memorydb = get_client(session, "memorydb", region_name=region)

    try:
        response = memorydb.describe_clusters()
        for cluster in response.get("Clusters", []):
            kms_key = cluster.get("KmsKeyId", "")
            if kms_key and kms_key in key_arns:
                usages[kms_key].append(
                    ResourceUsage(
                        service="MemoryDB",
                        resource_type="cluster",
                        resource_id=cluster.get("ARN", ""),
                        resource_name=cluster.get("Name", ""),
                    )
                )
    except ClientError:
        pass

    return usages


def find_ecr_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """ECR에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    ecr = get_client(session, "ecr", region_name=region)

    try:
        paginator = ecr.get_paginator("describe_repositories")
        for page in paginator.paginate():
            for repo in page.get("repositories", []):
                encryption = repo.get("encryptionConfiguration", {})
                kms_key = encryption.get("kmsKey", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="ECR",
                            resource_type="repository",
                            resource_id=repo.get("repositoryArn", ""),
                            resource_name=repo.get("repositoryName", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def find_athena_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """Athena WorkGroup에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    athena = get_client(session, "athena", region_name=region)

    try:
        # list_work_groups는 paginator를 지원하지 않음 - 수동 페이징
        next_token = None
        while True:
            if next_token:
                response = athena.list_work_groups(NextToken=next_token)
            else:
                response = athena.list_work_groups()

            for wg in response.get("WorkGroups", []):
                wg_name = wg.get("Name", "")
                try:
                    detail = athena.get_work_group(WorkGroup=wg_name).get("WorkGroup", {})
                    config = detail.get("Configuration", {})
                    result_config = config.get("ResultConfiguration", {})
                    encryption = result_config.get("EncryptionConfiguration", {})
                    kms_key = encryption.get("KmsKey", "")
                    if kms_key and kms_key in key_arns:
                        usages[kms_key].append(
                            ResourceUsage(
                                service="Athena",
                                resource_type="workgroup",
                                resource_id=detail.get("Name", ""),
                                resource_name=wg_name,
                            )
                        )
                except ClientError:
                    continue

            next_token = response.get("NextToken")
            if not next_token:
                break
    except ClientError:
        pass

    return usages


def find_ssm_parameter_usage(session, region: str, key_arns: set[str]) -> dict[str, list[ResourceUsage]]:
    """Systems Manager Parameter Store에서 KMS 키 사용처 찾기"""
    from botocore.exceptions import ClientError

    usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}
    ssm = get_client(session, "ssm", region_name=region)

    try:
        paginator = ssm.get_paginator("describe_parameters")
        for page in paginator.paginate():
            for param in page.get("Parameters", []):
                kms_key = param.get("KeyId", "")
                if kms_key and kms_key in key_arns:
                    usages[kms_key].append(
                        ResourceUsage(
                            service="SSM Parameter",
                            resource_type="parameter",
                            resource_id=param.get("ARN", ""),
                            resource_name=param.get("Name", ""),
                        )
                    )
    except ClientError:
        pass

    return usages


def merge_usages(base: dict[str, list[ResourceUsage]], *others: dict[str, list[ResourceUsage]]) -> None:
    """사용처 딕셔너리 병합"""
    for other in others:
        for key_arn, usages in other.items():
            if key_arn in base:
                base[key_arn].extend(usages)


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> KMSUsageResult:
    """단일 계정/리전의 KMS 사용처 수집 (병렬 실행용)"""
    kms = get_client(session, "kms", region_name=region)
    keys = collect_kms_keys(kms)

    # 키가 없어도 결과 반환 (스캔 정보 유지)
    if not keys:
        return KMSUsageResult(
            account_id=account_id,
            account_name=account_name,
            region=region,
            total_keys=0,
            customer_keys=0,
        )

    # 고객 관리 키만 필터링
    customer_keys = [k for k in keys if k.key_manager == "CUSTOMER" and k.key_state == "Enabled"]
    key_arns = {k.arn for k in customer_keys}

    # 각 서비스에서 사용처 찾기
    all_usages: dict[str, list[ResourceUsage]] = {arn: [] for arn in key_arns}

    merge_usages(
        all_usages,
        find_s3_usage(session, region, key_arns),
        find_ebs_usage(session, region, key_arns),
        find_rds_usage(session, region, key_arns),
        find_secrets_usage(session, region, key_arns),
        find_lambda_usage(session, region, key_arns),
        find_sns_usage(session, region, key_arns),
        find_sqs_usage(session, region, key_arns),
        find_opensearch_usage(session, region, key_arns),
        find_efs_usage(session, region, key_arns),
        find_dynamodb_usage(session, region, key_arns),
        find_elasticache_usage(session, region, key_arns),
        find_kinesis_usage(session, region, key_arns),
        find_redshift_usage(session, region, key_arns),
        find_fsx_usage(session, region, key_arns),
        find_cloudwatch_logs_usage(session, region, key_arns),
        find_documentdb_usage(session, region, key_arns),
        find_neptune_usage(session, region, key_arns),
        find_backup_usage(session, region, key_arns),
        find_glue_usage(session, region, key_arns),
        find_msk_usage(session, region, key_arns),
        find_sagemaker_usage(session, region, key_arns),
        find_eks_usage(session, region, key_arns),
        find_memorydb_usage(session, region, key_arns),
        find_ecr_usage(session, region, key_arns),
        find_athena_usage(session, region, key_arns),
        find_ssm_parameter_usage(session, region, key_arns),
    )

    # 결과 생성
    result = KMSUsageResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_keys=len(keys),
        customer_keys=len(customer_keys),
    )

    for key in customer_keys:
        key_usage = KMSKeyUsage(key=key, usages=all_usages.get(key.arn, []))
        result.key_usages.append(key_usage)
        if key_usage.is_unused:
            result.unused_keys += 1

    return result


def generate_report(results: list[KMSUsageResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "KMS 키 사용처 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "전체 키", "AWS 관리 키", "CMK", "미사용 CMK"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        aws_managed = r.total_keys - r.customer_keys
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_keys)
        ws.cell(row=row, column=4, value=aws_managed)
        ws.cell(row=row, column=5, value=r.customer_keys)
        ws.cell(row=row, column=6, value=r.unused_keys)
        if r.unused_keys > 0:
            ws.cell(row=row, column=6).fill = yellow_fill

    # Key Usage 시트
    ws_usage = wb.create_sheet("Key Usage")
    usage_headers = ["Account", "Region", "Key ID", "Alias", "사용 횟수", "Service", "Resource Type", "Resource Name"]
    for col, h in enumerate(usage_headers, 1):
        ws_usage.cell(row=1, column=col, value=h).fill = header_fill
        ws_usage.cell(row=1, column=col).font = header_font

    usage_row = 1
    for r in results:
        for ku in r.key_usages:
            if ku.usages:
                for usage in ku.usages:
                    usage_row += 1
                    ws_usage.cell(row=usage_row, column=1, value=r.account_name)
                    ws_usage.cell(row=usage_row, column=2, value=r.region)
                    ws_usage.cell(row=usage_row, column=3, value=ku.key.key_id[:20] + "...")
                    ws_usage.cell(row=usage_row, column=4, value=ku.key.alias or "-")
                    ws_usage.cell(row=usage_row, column=5, value=ku.usage_count)
                    ws_usage.cell(row=usage_row, column=6, value=usage.service)
                    ws_usage.cell(row=usage_row, column=7, value=usage.resource_type)
                    ws_usage.cell(row=usage_row, column=8, value=usage.resource_name)
            else:
                # 미사용 키
                usage_row += 1
                ws_usage.cell(row=usage_row, column=1, value=r.account_name)
                ws_usage.cell(row=usage_row, column=2, value=r.region)
                ws_usage.cell(row=usage_row, column=3, value=ku.key.key_id[:20] + "...")
                ws_usage.cell(row=usage_row, column=4, value=ku.key.alias or "-")
                cell = ws_usage.cell(row=usage_row, column=5, value=0)
                cell.fill = red_fill
                ws_usage.cell(row=usage_row, column=6, value="-")
                ws_usage.cell(row=usage_row, column=7, value="-")
                ws_usage.cell(row=usage_row, column=8, value="(미사용)")

    # 열 너비 조정
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            col_idx = col[0].column
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"KMS_Usage_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def run(ctx) -> None:
    """CMK 사용처 분석"""
    console.print("[bold]CMK 사용처 분석 시작...[/bold]\n")
    console.print(
        "[dim]스캔 대상 (26개): S3, EBS, RDS, SecretsManager, Lambda, SNS, SQS, OpenSearch, "
        "EFS, DynamoDB, ElastiCache, Kinesis, Redshift, FSx, CloudWatch Logs, "
        "DocumentDB, Neptune, Backup, Glue, MSK, SageMaker, EKS, MemoryDB, ECR, Athena, SSM Parameter[/dim]\n"
    )

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=10, service="kms")
    results: list[KMSUsageResult] = list(result.get_data())

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        # 오류 상세 출력
        for err in result.get_errors():
            console.print(f"[dim]  - {err.identifier}/{err.region}: {err.message}[/dim]")

    # 스캔 결과 분석
    scanned_regions = len(results)

    if scanned_regions == 0:
        console.print("\n[yellow]스캔된 리전이 없습니다.[/yellow]")
        if result.error_count > 0:
            console.print("[dim]위 오류 내용을 확인하세요.[/dim]")
        return

    total_keys = sum(r.total_keys for r in results)
    total_cmk = sum(r.customer_keys for r in results)
    total_unused = sum(r.unused_keys for r in results)
    total_usages = sum(sum(ku.usage_count for ku in r.key_usages) for r in results)

    # 결과 없는 경우 상세 안내
    if total_cmk == 0:
        console.print()
        if total_keys == 0:
            console.print("[yellow]KMS 키가 없습니다.[/yellow]")
            console.print(f"[dim]스캔 리전: {scanned_regions}개[/dim]")
        else:
            console.print(f"[yellow]CMK가 없습니다.[/yellow] (AWS 관리 키만 {total_keys}개 존재)")
            console.print("[dim]CMK(고객 관리 키)만 사용처를 분석합니다. AWS 관리 키는 AWS 서비스가 자동 관리합니다.[/dim]")
        return

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"CMK: {total_cmk}개 / 미사용: [yellow]{total_unused}개[/yellow] / 총 사용처: {total_usages}건")

    # 서비스별 통계
    service_counts: dict[str, int] = {}
    for r in results:
        for ku in r.key_usages:
            for usage in ku.usages:
                service_counts[usage.service] = service_counts.get(usage.service, 0) + 1

    if service_counts:
        console.print("\n[bold]서비스별 사용처[/bold]")
        for svc, cnt in sorted(service_counts.items(), key=lambda x: -x[1]):
            console.print(f"  {svc}: {cnt}개")

    # 보고서 생성
    identifier = ctx.accounts[0].id if ctx.accounts else ctx.profile_name or "default"
    output_path = OutputPath(identifier).sub("kms", "inventory").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
