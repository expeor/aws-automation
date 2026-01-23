"""
plugins/apigateway/unused.py - API Gateway 미사용 분석

유휴/미사용 API 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "apigateway:GET",
        "cloudwatch:GetMetricStatistics",
    ],
}

# 미사용 기준: 7일간 요청 0
UNUSED_DAYS_THRESHOLD = 7


class APIStatus(Enum):
    """API 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    NO_STAGES = "no_stages"
    LOW_USAGE = "low_usage"


@dataclass
class APIInfo:
    """API Gateway 정보"""

    account_id: str
    account_name: str
    region: str
    api_id: str
    api_name: str
    api_type: str  # REST, HTTP, WEBSOCKET
    protocol_type: str
    endpoint_type: str
    stage_count: int
    created_date: datetime | None
    # CloudWatch 지표
    total_requests: float = 0.0
    error_4xx: float = 0.0
    error_5xx: float = 0.0


@dataclass
class APIFinding:
    """API 분석 결과"""

    api: APIInfo
    status: APIStatus
    recommendation: str


@dataclass
class APIGatewayAnalysisResult:
    """API Gateway 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_apis: int = 0
    unused_apis: int = 0
    no_stages: int = 0
    low_usage: int = 0
    normal_apis: int = 0
    findings: list[APIFinding] = field(default_factory=list)


def collect_rest_apis(session, account_id: str, account_name: str, region: str, cloudwatch) -> list[APIInfo]:
    """REST API 수집"""
    from botocore.exceptions import ClientError

    apigw = get_client(session, "apigateway", region_name=region)
    apis = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    try:
        paginator = apigw.get_paginator("get_rest_apis")
        for page in paginator.paginate():
            for api in page.get("items", []):
                api_id = api.get("id", "")
                api_name = api.get("name", "")

                # 스테이지 수 확인
                stage_count = 0
                try:
                    stages = apigw.get_stages(restApiId=api_id)
                    stage_count = len(stages.get("item", []))
                except ClientError:
                    pass

                # 엔드포인트 타입
                endpoint_config = api.get("endpointConfiguration", {})
                endpoint_types = endpoint_config.get("types", [])
                endpoint_type = ", ".join(endpoint_types) if endpoint_types else "EDGE"

                info = APIInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    api_id=api_id,
                    api_name=api_name,
                    api_type="REST",
                    protocol_type="REST",
                    endpoint_type=endpoint_type,
                    stage_count=stage_count,
                    created_date=api.get("createdDate"),
                )

                # CloudWatch 지표 조회
                if stage_count > 0:
                    try:
                        count_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/ApiGateway",
                            MetricName="Count",
                            Dimensions=[{"Name": "ApiName", "Value": api_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if count_resp.get("Datapoints"):
                            info.total_requests = sum(d["Sum"] for d in count_resp["Datapoints"])

                        err4_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/ApiGateway",
                            MetricName="4XXError",
                            Dimensions=[{"Name": "ApiName", "Value": api_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if err4_resp.get("Datapoints"):
                            info.error_4xx = sum(d["Sum"] for d in err4_resp["Datapoints"])

                        err5_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/ApiGateway",
                            MetricName="5XXError",
                            Dimensions=[{"Name": "ApiName", "Value": api_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if err5_resp.get("Datapoints"):
                            info.error_5xx = sum(d["Sum"] for d in err5_resp["Datapoints"])

                    except ClientError:
                        pass

                apis.append(info)

    except ClientError:
        pass

    return apis


def collect_http_apis(session, account_id: str, account_name: str, region: str, cloudwatch) -> list[APIInfo]:
    """HTTP API (API Gateway v2) 수집"""
    from botocore.exceptions import ClientError

    apigwv2 = get_client(session, "apigatewayv2", region_name=region)
    apis = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    try:
        paginator = apigwv2.get_paginator("get_apis")
        for page in paginator.paginate():
            for api in page.get("Items", []):
                api_id = api.get("ApiId", "")
                api_name = api.get("Name", "")
                protocol = api.get("ProtocolType", "HTTP")

                # 스테이지 수 확인
                stage_count = 0
                try:
                    stages = apigwv2.get_stages(ApiId=api_id)
                    stage_count = len(stages.get("Items", []))
                except ClientError:
                    pass

                info = APIInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    api_id=api_id,
                    api_name=api_name,
                    api_type="HTTP" if protocol == "HTTP" else "WEBSOCKET",
                    protocol_type=protocol,
                    endpoint_type="REGIONAL",
                    stage_count=stage_count,
                    created_date=api.get("CreatedDate"),
                )

                # CloudWatch 지표 조회
                if stage_count > 0:
                    try:
                        count_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/ApiGateway",
                            MetricName="Count",
                            Dimensions=[{"Name": "ApiId", "Value": api_id}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if count_resp.get("Datapoints"):
                            info.total_requests = sum(d["Sum"] for d in count_resp["Datapoints"])

                    except ClientError:
                        pass

                apis.append(info)

    except ClientError:
        pass

    return apis


def collect_apis(session, account_id: str, account_name: str, region: str) -> list[APIInfo]:
    """모든 API Gateway 수집"""
    cloudwatch = get_client(session, "cloudwatch", region_name=region)

    rest_apis = collect_rest_apis(session, account_id, account_name, region, cloudwatch)
    http_apis = collect_http_apis(session, account_id, account_name, region, cloudwatch)

    return rest_apis + http_apis


def analyze_apis(apis: list[APIInfo], account_id: str, account_name: str, region: str) -> APIGatewayAnalysisResult:
    """API Gateway 분석"""
    result = APIGatewayAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_apis=len(apis),
    )

    for api in apis:
        # 스테이지 없음
        if api.stage_count == 0:
            result.no_stages += 1
            result.findings.append(
                APIFinding(
                    api=api,
                    status=APIStatus.NO_STAGES,
                    recommendation="스테이지 없음 - 삭제 검토",
                )
            )
            continue

        # 미사용
        if api.total_requests == 0:
            result.unused_apis += 1
            result.findings.append(
                APIFinding(
                    api=api,
                    status=APIStatus.UNUSED,
                    recommendation="요청 없음 - 삭제 검토",
                )
            )
            continue

        # 저사용 (하루 평균 10회 미만)
        avg_daily = api.total_requests / UNUSED_DAYS_THRESHOLD
        if avg_daily < 10:
            result.low_usage += 1
            result.findings.append(
                APIFinding(
                    api=api,
                    status=APIStatus.LOW_USAGE,
                    recommendation=f"저사용 (일 평균 {avg_daily:.1f}회) - 통합 검토",
                )
            )
            continue

        result.normal_apis += 1
        result.findings.append(
            APIFinding(
                api=api,
                status=APIStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[APIGatewayAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    wb = Workbook()

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="스테이지없음", width=12, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row([
            r.account_name,
            r.region,
            r.total_apis,
            r.unused_apis,
            r.no_stages,
            r.low_usage,
            r.normal_apis,
        ])
        ws = summary_sheet._ws
        if r.unused_apis > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.no_stages > 0 or r.low_usage > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="API Name", width=30),
        ColumnDef(header="Type", width=12),
        ColumnDef(header="Endpoint", width=15),
        ColumnDef(header="Stages", width=10, style="number"),
        ColumnDef(header="Requests", width=12, style="number"),
        ColumnDef(header="상태", width=12),
        ColumnDef(header="권장 조치", width=40),
    ]
    detail_sheet = wb.new_sheet("APIs", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != APIStatus.NORMAL:
                a = f.api
                style = Styles.danger() if f.status == APIStatus.UNUSED else Styles.warning()
                detail_sheet.add_row([
                    a.account_name,
                    a.region,
                    a.api_name,
                    a.api_type,
                    a.endpoint_type,
                    a.stage_count,
                    int(a.total_requests),
                    f.status.value,
                    f.recommendation,
                ], style=style)

    return str(wb.save_as(output_dir, "APIGateway_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> APIGatewayAnalysisResult | None:
    """단일 계정/리전의 API Gateway 수집 및 분석 (병렬 실행용)"""
    apis = collect_apis(session, account_id, account_name, region)
    if not apis:
        return None
    return analyze_apis(apis, account_id, account_name, region)


def run(ctx) -> None:
    """API Gateway 미사용 분석"""
    console.print("[bold]API Gateway 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="apigateway")
    results: list[APIGatewayAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_apis for r in results)
    total_no_stages = sum(r.no_stages for r in results)
    total_low = sum(r.low_usage for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] / "
        f"스테이지없음: [yellow]{total_no_stages}개[/yellow] / "
        f"저사용: {total_low}개"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("apigateway", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
