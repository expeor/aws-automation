"""
plugins/elasticache/unused.py - ElastiCache 미사용 클러스터 분석

유휴/저사용 ElastiCache 클러스터 탐지 (CloudWatch 지표 기반)

최적화:
- CloudWatch GetMetricData API 사용 (배치 조회)
- 기존: 클러스터당 2 API 호출 → 최적화: 전체 1 API 호출
- 예: 30개 클러스터 × 2 메트릭 = 60 API → 1 API

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cloudwatch.common import MetricQuery, batch_get_metrics, sanitize_metric_id

console = Console()

# 미사용 기준: 7일간 연결 수 평균 0
UNUSED_DAYS_THRESHOLD = 7
# 저사용 기준: CPU 평균 5% 미만
LOW_USAGE_CPU_THRESHOLD = 5.0

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticache:DescribeReplicationGroups",
        "elasticache:DescribeCacheClusters",
        "cloudwatch:GetMetricStatistics",
    ],
}


class ClusterStatus(Enum):
    """클러스터 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"


@dataclass
class ClusterInfo:
    """ElastiCache 클러스터 정보"""

    account_id: str
    account_name: str
    region: str
    cluster_id: str
    engine: str  # redis or memcached
    node_type: str
    num_nodes: int
    status: str
    created_at: datetime | None
    # CloudWatch 지표
    avg_connections: float = 0.0
    avg_cpu: float = 0.0
    avg_memory: float = 0.0

    @property
    def estimated_monthly_cost(self) -> float:
        """대략적인 월간 비용 추정 (노드 타입별 대략치)"""
        # 간단한 가격 맵 (실제로는 pricing 모듈 확장 필요)
        price_map = {
            "cache.t3.micro": 0.017,
            "cache.t3.small": 0.034,
            "cache.t3.medium": 0.068,
            "cache.t4g.micro": 0.016,
            "cache.t4g.small": 0.032,
            "cache.t4g.medium": 0.065,
            "cache.r6g.large": 0.206,
            "cache.r6g.xlarge": 0.413,
            "cache.r7g.large": 0.222,
            "cache.m6g.large": 0.157,
        }
        hourly = price_map.get(self.node_type, 0.10)  # 기본값
        return hourly * 730 * self.num_nodes  # 월간


@dataclass
class ClusterFinding:
    """클러스터 분석 결과"""

    cluster: ClusterInfo
    status: ClusterStatus
    recommendation: str


@dataclass
class ElastiCacheAnalysisResult:
    """ElastiCache 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_clusters: int = 0
    unused_clusters: int = 0
    low_usage_clusters: int = 0
    normal_clusters: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[ClusterFinding] = field(default_factory=list)


def collect_elasticache_clusters(session, account_id: str, account_name: str, region: str) -> list[ClusterInfo]:
    """ElastiCache 클러스터 수집 (배치 메트릭 최적화)

    최적화:
    - 기존: 클러스터당 2 API 호출 → 최적화: 전체 1-2 API 호출
    - 예: 30개 클러스터 × 2 메트릭 = 60 API → 1 API
    """
    from botocore.exceptions import ClientError

    elasticache = get_client(session, "elasticache", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)

    redis_clusters: list[ClusterInfo] = []
    memcached_clusters: list[ClusterInfo] = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    # 1단계: Redis 클러스터 (Replication Groups) 수집
    try:
        paginator = elasticache.get_paginator("describe_replication_groups")
        for page in paginator.paginate():
            for rg in page.get("ReplicationGroups", []):
                cluster_id = rg.get("ReplicationGroupId", "")
                num_nodes = len(rg.get("MemberClusters", []))
                node_type = rg.get("CacheNodeType", "unknown")

                cluster = ClusterInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    cluster_id=cluster_id,
                    engine="redis",
                    node_type=node_type,
                    num_nodes=num_nodes if num_nodes > 0 else 1,
                    status=rg.get("Status", ""),
                    created_at=None,
                )
                redis_clusters.append(cluster)
    except ClientError:
        pass

    # 2단계: Memcached 클러스터 수집
    try:
        paginator = elasticache.get_paginator("describe_cache_clusters")
        for page in paginator.paginate(ShowCacheNodeInfo=True):
            for cc in page.get("CacheClusters", []):
                # Redis replication group에 속한 클러스터는 제외
                if cc.get("ReplicationGroupId"):
                    continue

                cluster_id = cc.get("CacheClusterId", "")
                cluster = ClusterInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    cluster_id=cluster_id,
                    engine=cc.get("Engine", "memcached"),
                    node_type=cc.get("CacheNodeType", "unknown"),
                    num_nodes=cc.get("NumCacheNodes", 1),
                    status=cc.get("CacheClusterStatus", ""),
                    created_at=cc.get("CacheClusterCreateTime"),
                )
                memcached_clusters.append(cluster)
    except ClientError:
        pass

    # 3단계: 배치 메트릭 조회
    if redis_clusters:
        _collect_elasticache_metrics_batch(cloudwatch, redis_clusters, "ReplicationGroupId", start_time, now)

    if memcached_clusters:
        _collect_elasticache_metrics_batch(cloudwatch, memcached_clusters, "CacheClusterId", start_time, now)

    return redis_clusters + memcached_clusters


def _collect_elasticache_metrics_batch(
    cloudwatch,
    clusters: list[ClusterInfo],
    dimension_name: str,
    start_time: datetime,
    end_time: datetime,
) -> None:
    """ElastiCache 클러스터 메트릭 배치 수집 (내부 함수)"""
    from botocore.exceptions import ClientError

    metrics_to_fetch = [
        ("CurrConnections", "avg_connections"),
        ("CPUUtilization", "avg_cpu"),
    ]

    # 쿼리 생성
    queries = []
    for cluster in clusters:
        safe_id = sanitize_metric_id(cluster.cluster_id)
        for metric_name, _ in metrics_to_fetch:
            metric_key = metric_name.lower()
            queries.append(
                MetricQuery(
                    id=f"{safe_id}_{metric_key}",
                    namespace="AWS/ElastiCache",
                    metric_name=metric_name,
                    dimensions={dimension_name: cluster.cluster_id},
                    stat="Average",
                )
            )

    try:
        # 배치 조회
        results = batch_get_metrics(cloudwatch, queries, start_time, end_time, period=86400)

        # 결과 매핑
        days = (end_time - start_time).days
        if days <= 0:
            days = 1

        for cluster in clusters:
            safe_id = sanitize_metric_id(cluster.cluster_id)
            for metric_name, attr_name in metrics_to_fetch:
                metric_key = metric_name.lower()
                # GetMetricData with Average stat returns sum of averages
                value = results.get(f"{safe_id}_{metric_key}", 0.0) / days
                setattr(cluster, attr_name, value)

    except ClientError:
        # 실패 시 무시 (기본값 0 유지)
        pass


def analyze_clusters(
    clusters: list[ClusterInfo], account_id: str, account_name: str, region: str
) -> ElastiCacheAnalysisResult:
    """ElastiCache 클러스터 분석"""
    result = ElastiCacheAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_clusters=len(clusters),
    )

    for cluster in clusters:
        # 미사용: 연결 수 평균 0
        if cluster.avg_connections == 0:
            result.unused_clusters += 1
            result.unused_monthly_cost += cluster.estimated_monthly_cost
            result.findings.append(
                ClusterFinding(
                    cluster=cluster,
                    status=ClusterStatus.UNUSED,
                    recommendation=f"연결 없음 - 삭제 검토 (${cluster.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: CPU 5% 미만
        if cluster.avg_cpu < LOW_USAGE_CPU_THRESHOLD:
            result.low_usage_clusters += 1
            result.low_usage_monthly_cost += cluster.estimated_monthly_cost
            result.findings.append(
                ClusterFinding(
                    cluster=cluster,
                    status=ClusterStatus.LOW_USAGE,
                    recommendation=f"저사용 (CPU {cluster.avg_cpu:.1f}%) - 다운사이징 검토",
                )
            )
            continue

        result.normal_clusters += 1
        result.findings.append(
            ClusterFinding(
                cluster=cluster,
                status=ClusterStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[ElastiCacheAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl.styles import PatternFill

    from core.tools.io.excel import ColumnDef, Styles, Workbook

    wb = Workbook()

    # 조건부 셀 스타일링용 Fill
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    summary_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="전체", width=10, style="number"),
        ColumnDef(header="미사용", width=10, style="number"),
        ColumnDef(header="저사용", width=10, style="number"),
        ColumnDef(header="정상", width=10, style="number"),
        ColumnDef(header="미사용 비용", width=15),
        ColumnDef(header="저사용 비용", width=15),
    ]
    summary_sheet = wb.new_sheet("Summary", summary_columns)

    for r in results:
        row_num = summary_sheet.add_row(
            [
                r.account_name,
                r.region,
                r.total_clusters,
                r.unused_clusters,
                r.low_usage_clusters,
                r.normal_clusters,
                f"${r.unused_monthly_cost:,.2f}",
                f"${r.low_usage_monthly_cost:,.2f}",
            ]
        )
        # 셀 단위 조건부 스타일링
        ws = summary_sheet._ws
        if r.unused_clusters > 0:
            ws.cell(row=row_num, column=4).fill = red_fill
        if r.low_usage_clusters > 0:
            ws.cell(row=row_num, column=5).fill = yellow_fill

    # Detail 시트
    detail_columns = [
        ColumnDef(header="Account", width=20),
        ColumnDef(header="Region", width=15),
        ColumnDef(header="Cluster ID", width=30),
        ColumnDef(header="Engine", width=12),
        ColumnDef(header="Node Type", width=18),
        ColumnDef(header="Nodes", width=8, style="number"),
        ColumnDef(header="상태", width=12),
        ColumnDef(header="Avg Conn", width=10),
        ColumnDef(header="Avg CPU", width=10),
        ColumnDef(header="월간 비용", width=12),
        ColumnDef(header="권장 조치", width=35),
    ]
    detail_sheet = wb.new_sheet("Clusters", detail_columns)

    for r in results:
        for f in r.findings:
            if f.status != ClusterStatus.NORMAL:
                c = f.cluster
                style = Styles.danger() if f.status == ClusterStatus.UNUSED else Styles.warning()

                detail_sheet.add_row(
                    [
                        c.account_name,
                        c.region,
                        c.cluster_id,
                        c.engine,
                        c.node_type,
                        c.num_nodes,
                        f.status.value,
                        f"{c.avg_connections:.1f}",
                        f"{c.avg_cpu:.1f}%",
                        f"${c.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ],
                    style=style,
                )

    return str(wb.save_as(output_dir, "ElastiCache_Unused"))


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ElastiCacheAnalysisResult | None:
    """단일 계정/리전의 ElastiCache 클러스터 수집 및 분석 (병렬 실행용)"""
    clusters = collect_elasticache_clusters(session, account_id, account_name, region)
    if not clusters:
        return None
    return analyze_clusters(clusters, account_id, account_name, region)


def run(ctx) -> None:
    """ElastiCache 미사용 클러스터 분석"""
    console.print("[bold]ElastiCache 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="elasticache")
    results: list[ElastiCacheAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_clusters for r in results)
    total_low = sum(r.low_usage_clusters for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월)"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("elasticache", "unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
