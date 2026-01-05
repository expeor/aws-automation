"""
plugins/elasticache/common.py - ElastiCache 공통 유틸리티

Redis, Memcached에서 공유하는 데이터 구조와 헬퍼 함수.
"""

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

# 미사용 기준: 7일간 연결 수 평균 0
UNUSED_DAYS_THRESHOLD = 7
# 저사용 기준: CPU 평균 5% 미만
LOW_USAGE_CPU_THRESHOLD = 5.0


class CacheEngine(Enum):
    """캐시 엔진 타입"""

    REDIS = "redis"
    MEMCACHED = "memcached"


class ClusterStatus(Enum):
    """클러스터 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"


@dataclass
class ClusterInfo:
    """ElastiCache 클러스터 정보"""

    account_id: str
    account_name: str
    region: str
    cluster_id: str
    engine: str  # redis or memcached
    node_type: str
    num_nodes: int
    status: str
    created_at: datetime | None
    # CloudWatch 지표
    avg_connections: float = 0.0
    avg_cpu: float = 0.0
    avg_memory: float = 0.0

    @property
    def estimated_monthly_cost(self) -> float:
        """대략적인 월간 비용 추정"""
        price_map = {
            "cache.t3.micro": 0.017,
            "cache.t3.small": 0.034,
            "cache.t3.medium": 0.068,
            "cache.t4g.micro": 0.016,
            "cache.t4g.small": 0.032,
            "cache.t4g.medium": 0.065,
            "cache.r6g.large": 0.206,
            "cache.r6g.xlarge": 0.413,
            "cache.r7g.large": 0.222,
            "cache.m6g.large": 0.157,
        }
        hourly = price_map.get(self.node_type, 0.10)
        return hourly * 730 * self.num_nodes


@dataclass
class ClusterFinding:
    """클러스터 분석 결과"""

    cluster: ClusterInfo
    status: ClusterStatus
    recommendation: str


@dataclass
class ElastiCacheAnalysisResult:
    """ElastiCache 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    engine_filter: str | None = None
    total_clusters: int = 0
    unused_clusters: int = 0
    low_usage_clusters: int = 0
    normal_clusters: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[ClusterFinding] = field(default_factory=list)


# =============================================================================
# 수집 함수
# =============================================================================


def collect_redis_clusters(session, account_id: str, account_name: str, region: str) -> list[ClusterInfo]:
    """Redis 클러스터 수집 (Replication Groups)"""
    from botocore.exceptions import ClientError

    from core.parallel import get_client

    elasticache = get_client(session, "elasticache", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    clusters = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    try:
        paginator = elasticache.get_paginator("describe_replication_groups")
        for page in paginator.paginate():
            for rg in page.get("ReplicationGroups", []):
                cluster_id = rg.get("ReplicationGroupId", "")
                num_nodes = len(rg.get("MemberClusters", []))
                node_type = rg.get("CacheNodeType", "unknown")

                cluster = ClusterInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    cluster_id=cluster_id,
                    engine="redis",
                    node_type=node_type,
                    num_nodes=num_nodes if num_nodes > 0 else 1,
                    status=rg.get("Status", ""),
                    created_at=None,
                )

                # CloudWatch 지표 조회
                _fetch_cloudwatch_metrics(cloudwatch, cluster, cluster_id, "ReplicationGroupId", start_time, now)
                clusters.append(cluster)
    except ClientError:
        pass

    return clusters


def collect_memcached_clusters(session, account_id: str, account_name: str, region: str) -> list[ClusterInfo]:
    """Memcached 클러스터 수집"""
    from botocore.exceptions import ClientError

    from core.parallel import get_client

    elasticache = get_client(session, "elasticache", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    clusters = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    try:
        paginator = elasticache.get_paginator("describe_cache_clusters")
        for page in paginator.paginate(ShowCacheNodeInfo=True):
            for cc in page.get("CacheClusters", []):
                # Redis replication group에 속한 클러스터는 제외
                if cc.get("ReplicationGroupId"):
                    continue

                cluster_id = cc.get("CacheClusterId", "")
                cluster = ClusterInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    cluster_id=cluster_id,
                    engine=cc.get("Engine", "memcached"),
                    node_type=cc.get("CacheNodeType", "unknown"),
                    num_nodes=cc.get("NumCacheNodes", 1),
                    status=cc.get("CacheClusterStatus", ""),
                    created_at=cc.get("CacheClusterCreateTime"),
                )

                # CloudWatch 지표 조회
                _fetch_cloudwatch_metrics(cloudwatch, cluster, cluster_id, "CacheClusterId", start_time, now)
                clusters.append(cluster)
    except ClientError:
        pass

    return clusters


def _fetch_cloudwatch_metrics(cloudwatch, cluster: ClusterInfo, cluster_id: str, dimension_name: str, start_time, end_time):
    """CloudWatch 지표 조회"""
    from botocore.exceptions import ClientError

    try:
        # CurrConnections
        conn_resp = cloudwatch.get_metric_statistics(
            Namespace="AWS/ElastiCache",
            MetricName="CurrConnections",
            Dimensions=[{"Name": dimension_name, "Value": cluster_id}],
            StartTime=start_time,
            EndTime=end_time,
            Period=86400,
            Statistics=["Average"],
        )
        if conn_resp.get("Datapoints"):
            cluster.avg_connections = sum(d["Average"] for d in conn_resp["Datapoints"]) / len(conn_resp["Datapoints"])

        # CPUUtilization
        cpu_resp = cloudwatch.get_metric_statistics(
            Namespace="AWS/ElastiCache",
            MetricName="CPUUtilization",
            Dimensions=[{"Name": dimension_name, "Value": cluster_id}],
            StartTime=start_time,
            EndTime=end_time,
            Period=86400,
            Statistics=["Average"],
        )
        if cpu_resp.get("Datapoints"):
            cluster.avg_cpu = sum(d["Average"] for d in cpu_resp["Datapoints"]) / len(cpu_resp["Datapoints"])
    except ClientError:
        pass


# =============================================================================
# 분석 함수
# =============================================================================


def analyze_clusters(
    clusters: list[ClusterInfo],
    account_id: str,
    account_name: str,
    region: str,
    engine_filter: str | None = None,
) -> ElastiCacheAnalysisResult:
    """ElastiCache 클러스터 분석"""
    result = ElastiCacheAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        engine_filter=engine_filter,
        total_clusters=len(clusters),
    )

    for cluster in clusters:
        # 미사용: 연결 수 평균 0
        if cluster.avg_connections == 0:
            result.unused_clusters += 1
            result.unused_monthly_cost += cluster.estimated_monthly_cost
            result.findings.append(
                ClusterFinding(
                    cluster=cluster,
                    status=ClusterStatus.UNUSED,
                    recommendation=f"연결 없음 - 삭제 검토 (${cluster.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: CPU 5% 미만
        if cluster.avg_cpu < LOW_USAGE_CPU_THRESHOLD:
            result.low_usage_clusters += 1
            result.low_usage_monthly_cost += cluster.estimated_monthly_cost
            result.findings.append(
                ClusterFinding(
                    cluster=cluster,
                    status=ClusterStatus.LOW_USAGE,
                    recommendation=f"저사용 (CPU {cluster.avg_cpu:.1f}%) - 다운사이징 검토",
                )
            )
            continue

        result.normal_clusters += 1
        result.findings.append(
            ClusterFinding(
                cluster=cluster,
                status=ClusterStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


# =============================================================================
# 보고서 생성
# =============================================================================


def generate_unused_report(
    results: list[ElastiCacheAnalysisResult],
    output_dir: str,
    engine_name: str = "ElastiCache",
) -> str:
    """미사용 분석 Excel 보고서 생성"""
    import os

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"{engine_name} 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "전체", "미사용", "저사용", "정상", "미사용 비용", "저사용 비용"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_clusters)
        ws.cell(row=row, column=4, value=r.unused_clusters)
        ws.cell(row=row, column=5, value=r.low_usage_clusters)
        ws.cell(row=row, column=6, value=r.normal_clusters)
        ws.cell(row=row, column=7, value=f"${r.unused_monthly_cost:,.2f}")
        ws.cell(row=row, column=8, value=f"${r.low_usage_monthly_cost:,.2f}")
        if r.unused_clusters > 0:
            ws.cell(row=row, column=4).fill = red_fill
        if r.low_usage_clusters > 0:
            ws.cell(row=row, column=5).fill = yellow_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Clusters")
    detail_headers = [
        "Account", "Region", "Cluster ID", "Engine", "Node Type",
        "Nodes", "상태", "Avg Conn", "Avg CPU", "월간 비용", "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != ClusterStatus.NORMAL:
                detail_row += 1
                c = f.cluster
                ws_detail.cell(row=detail_row, column=1, value=c.account_name)
                ws_detail.cell(row=detail_row, column=2, value=c.region)
                ws_detail.cell(row=detail_row, column=3, value=c.cluster_id)
                ws_detail.cell(row=detail_row, column=4, value=c.engine)
                ws_detail.cell(row=detail_row, column=5, value=c.node_type)
                ws_detail.cell(row=detail_row, column=6, value=c.num_nodes)
                ws_detail.cell(row=detail_row, column=7, value=f.status.value)
                ws_detail.cell(row=detail_row, column=8, value=f"{c.avg_connections:.1f}")
                ws_detail.cell(row=detail_row, column=9, value=f"{c.avg_cpu:.1f}%")
                ws_detail.cell(row=detail_row, column=10, value=f"${c.estimated_monthly_cost:.2f}")
                ws_detail.cell(row=detail_row, column=11, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            col_idx = col[0].column
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"{engine_name}_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath
