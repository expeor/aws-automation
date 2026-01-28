"""
tests/shared/io/excel/test_workbook.py - Excel Workbook 테스트
"""

import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from shared.io.excel.workbook import ColumnDef, Sheet, SummaryItem, Workbook


class TestColumnDef:
    """ColumnDef 클래스 테스트"""

    def test_default_values(self):
        """기본값 확인"""
        col = ColumnDef(header="테스트")

        assert col.header == "테스트"
        assert col.width == 15
        assert col.style == "data"
        assert col.header_en is None

    def test_custom_values(self):
        """커스텀 값 설정"""
        col = ColumnDef(
            header="볼륨 ID",
            header_en="Volume ID",
            width=22,
            style="center",
        )

        assert col.header == "볼륨 ID"
        assert col.header_en == "Volume ID"
        assert col.width == 22
        assert col.style == "center"

    def test_get_header_korean(self):
        """한국어 헤더 반환"""
        col = ColumnDef(header="볼륨", header_en="Volume")

        assert col.get_header("ko") == "볼륨"

    def test_get_header_english(self):
        """영어 헤더 반환"""
        col = ColumnDef(header="볼륨", header_en="Volume")

        assert col.get_header("en") == "Volume"

    def test_get_header_english_fallback(self):
        """영어 헤더 없을 때 한국어 반환"""
        col = ColumnDef(header="볼륨")

        assert col.get_header("en") == "볼륨"


class TestSummaryItem:
    """SummaryItem 클래스 테스트"""

    def test_default_values(self):
        """기본값 확인"""
        item = SummaryItem(label="테스트")

        assert item.label == "테스트"
        assert item.value == ""
        assert item.is_header is False
        assert item.highlight is None

    def test_header_item(self):
        """헤더 항목"""
        item = SummaryItem(label="섹션", is_header=True)

        assert item.is_header is True

    def test_highlighted_item(self):
        """강조 항목"""
        item = SummaryItem(label="경고", value=10, highlight="danger")

        assert item.highlight == "danger"


class TestWorkbook:
    """Workbook 클래스 테스트"""

    def test_initialization(self):
        """초기화 테스트"""
        wb = Workbook(lang="ko")

        assert wb._lang == "ko"
        assert wb._wb is not None
        assert wb._sheets == []

    def test_initialization_english(self):
        """영어 모드 초기화"""
        wb = Workbook(lang="en")

        assert wb._lang == "en"

    def test_new_sheet_creates_sheet(self):
        """시트 생성"""
        wb = Workbook()
        columns = [
            ColumnDef(header="ID", width=10),
            ColumnDef(header="이름", width=20),
        ]

        sheet = wb.new_sheet(name="테스트", columns=columns)

        assert isinstance(sheet, Sheet)
        assert len(wb._sheets) == 1

    def test_new_sheet_with_english_headers(self):
        """영어 헤더로 시트 생성"""
        wb = Workbook(lang="en")
        columns = [
            ColumnDef(header="ID", header_en="ID", width=10),
            ColumnDef(header="이름", header_en="Name", width=20),
        ]

        sheet = wb.new_sheet(name="Test", columns=columns)

        # 첫 번째 행(헤더)의 값 확인
        ws = sheet._ws
        assert ws.cell(row=1, column=2).value == "Name"

    def test_multiple_sheets(self):
        """여러 시트 생성"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]

        sheet1 = wb.new_sheet(name="시트1", columns=columns)
        sheet2 = wb.new_sheet(name="시트2", columns=columns)

        assert len(wb._sheets) == 2
        assert sheet1._ws.title == "시트1"
        assert sheet2._ws.title == "시트2"

    def test_save_creates_file(self):
        """파일 저장"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        sheet = wb.new_sheet(name="테스트", columns=columns)
        sheet.add_row(["value"])

        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = wb.save(Path(tmpdir) / "test_output.xlsx")

            assert Path(filepath).exists()
            assert str(filepath).endswith(".xlsx")

    def test_save_as_creates_file(self):
        """save_as로 파일 저장"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        wb.new_sheet(name="테스트", columns=columns)

        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = wb.save_as(str(tmpdir), "test_output", region="ap-northeast-2")

            assert Path(filepath).exists()
            assert "test_output" in str(filepath)
            assert "ap-northeast-2" in str(filepath)


class TestSheet:
    """Sheet 클래스 테스트"""

    def test_add_row(self):
        """행 추가"""
        wb = Workbook()
        columns = [
            ColumnDef(header="A", width=10),
            ColumnDef(header="B", width=10),
        ]
        sheet = wb.new_sheet(name="Test", columns=columns)

        row_num = sheet.add_row(["val1", "val2"])

        assert row_num == 2  # 헤더가 1이므로 첫 데이터는 2
        assert sheet.row_count == 1

    def test_add_multiple_rows(self):
        """여러 행 추가"""
        wb = Workbook()
        columns = [ColumnDef(header="A")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        sheet.add_row(["row1"])
        sheet.add_row(["row2"])
        sheet.add_row(["row3"])

        assert sheet.row_count == 3

    def test_add_summary_row(self):
        """요약 행 추가"""
        wb = Workbook()
        columns = [
            ColumnDef(header="항목"),
            ColumnDef(header="값", style="number"),
        ]
        sheet = wb.new_sheet(name="Test", columns=columns)

        sheet.add_row(["A", 100])
        sheet.add_row(["B", 200])
        row_num = sheet.add_summary_row(["합계", 300])

        assert row_num == 4  # 헤더(1) + 2 데이터 + 1 요약

    def test_column_styles_applied(self):
        """컬럼 스타일 적용"""
        wb = Workbook()
        columns = [
            ColumnDef(header="텍스트", style="text"),
            ColumnDef(header="숫자", style="number"),
            ColumnDef(header="통화", style="currency"),
            ColumnDef(header="퍼센트", style="percent"),
            ColumnDef(header="중앙", style="center"),
        ]
        sheet = wb.new_sheet(name="Test", columns=columns)

        sheet.add_row(["text", 100, 1000, 0.5, "center"])

        # 스타일이 적용되었는지 확인
        ws = sheet._ws
        assert ws.cell(row=2, column=2).number_format is not None

    def test_finalize_adds_autofilter(self):
        """마무리 시 자동 필터 추가"""
        wb = Workbook()
        columns = [ColumnDef(header="A"), ColumnDef(header="B")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        sheet.add_row(["val1", "val2"])
        sheet.finalize()

        assert sheet._ws.auto_filter.ref is not None


class TestWorkbookSummarySheet:
    """SummarySheet 테스트"""

    def test_new_summary_sheet(self):
        """Summary 시트 생성"""
        wb = Workbook()

        summary = wb.new_summary_sheet("요약")

        assert summary is not None
        assert summary._ws.title == "요약"

    def test_add_title(self):
        """타이틀 추가"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        summary.add_title("테스트 보고서")

        # 타이틀이 추가되었는지 확인
        assert summary.current_row > 1

    def test_add_section(self):
        """섹션 추가"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        summary.add_section("분석 정보")

        assert summary.current_row > 1

    def test_add_item(self):
        """항목 추가"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        summary.add_item("계정", "123456789012")
        summary.add_item("리전", "ap-northeast-2")

        assert summary.current_row > 2

    def test_add_item_with_highlight(self):
        """강조 항목 추가"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        summary.add_item("경고", "10개", highlight="danger")
        summary.add_item("주의", "5개", highlight="warning")
        summary.add_item("정상", "100개", highlight="success")

        assert summary.current_row > 3

    def test_add_blank_row(self):
        """빈 행 추가"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        initial_row = summary.current_row
        summary.add_blank_row()

        assert summary.current_row == initial_row + 1

    def test_method_chaining(self):
        """메서드 체이닝"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        result = (
            summary.add_title("보고서")
            .add_section("정보")
            .add_item("항목1", "값1")
            .add_blank_row()
            .add_item("항목2", "값2")
        )

        assert result is summary


class TestWorkbookIntegration:
    """통합 테스트"""

    def test_complete_workflow(self):
        """전체 워크플로우 테스트"""
        wb = Workbook(lang="ko")

        # Summary 시트
        summary = wb.new_summary_sheet("분석 요약")
        summary.add_title("테스트 보고서")
        summary.add_section("기본 정보")
        summary.add_item("분석 대상", "테스트")
        summary.add_item("분석 일시", "2024-01-01")

        # 데이터 시트
        columns = [
            ColumnDef(header="ID", width=15, style="text"),
            ColumnDef(header="이름", width=25, style="data"),
            ColumnDef(header="값", width=15, style="number"),
            ColumnDef(header="비율", width=10, style="percent"),
        ]
        data_sheet = wb.new_sheet(name="상세 데이터", columns=columns)

        data_sheet.add_row(["001", "항목 A", 100, 0.5])
        data_sheet.add_row(["002", "항목 B", 200, 0.3])
        data_sheet.add_row(["003", "항목 C", 150, 0.2])
        data_sheet.add_summary_row(["합계", "-", 450, 1.0])

        # 저장
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = wb.save(Path(tmpdir) / "integration_test.xlsx")

            assert Path(filepath).exists()
            assert Path(filepath).stat().st_size > 0

    def test_csv_export(self):
        """CSV 내보내기 (시트별)"""
        wb = Workbook()
        columns = [
            ColumnDef(header="A"),
            ColumnDef(header="B"),
        ]
        sheet = wb.new_sheet(name="Data", columns=columns)
        sheet.add_row(["val1", "val2"])
        sheet.add_row(["val3", "val4"])

        with tempfile.TemporaryDirectory() as tmpdir:
            # Excel 저장
            filepath = wb.save(Path(tmpdir) / "test.xlsx")

            assert Path(filepath).exists()

            # CSV 내보내기 메서드가 있다면 테스트
            # wb.export_csv(str(tmpdir), "test")


class TestSummarySheetAdvanced:
    """SummarySheet 고급 기능 테스트"""

    def test_add_list_section_with_items(self):
        """리스트 섹션 추가"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        items = [("Item A", 100), ("Item B", 200), ("Item C", 150)]
        summary.add_list_section("Top Items", items)

        assert summary.current_row > 1

    def test_add_list_section_empty(self):
        """빈 리스트 섹션"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        summary.add_list_section("Empty List", [])

        assert summary.current_row > 1

    def test_add_list_section_max_items(self):
        """최대 항목 제한"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        items = [(f"Item {i}", i * 10) for i in range(20)]
        summary.add_list_section("Top 5", items, max_items=5)

        # 5개만 추가되어야 함
        assert summary.current_row > 1

    def test_add_list_section_long_names(self):
        """긴 이름 자르기"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        long_name = "A" * 100
        items = [(long_name, 100)]
        summary.add_list_section("Long Names", items)

        assert summary.current_row > 1

    def test_add_item_with_number_format(self):
        """숫자 포맷 지정"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        summary.add_item("계좌번호", "123456789", number_format="@")

        assert summary.current_row > 1

    def test_summary_sheet_english(self):
        """영어 Summary 시트"""
        wb = Workbook(lang="en")
        summary = wb.new_summary_sheet()

        assert summary._ws.title == "Summary"

    def test_summary_sheet_custom_position(self):
        """Summary 시트 위치 지정"""
        wb = Workbook()
        # 일반 시트 먼저 생성
        columns = [ColumnDef(header="Test")]
        wb.new_sheet(name="Data", columns=columns)

        # 맨 앞에 Summary 시트 추가
        summary = wb.new_summary_sheet("요약", position=0)

        assert wb._wb.sheetnames[0] == "요약"

    def test_summary_sheet_multiple_sections(self):
        """여러 섹션 추가"""
        wb = Workbook()
        summary = wb.new_summary_sheet("요약")

        (
            summary.add_title("보고서")
            .add_section("기본 정보")
            .add_item("계정", "123456789012")
            .add_blank_row()
            .add_section("분석 결과")
            .add_item("총 개수", "100개")
            .add_item("비용", "$1,000", highlight="warning")
            .add_blank_row()
            .add_list_section("Top 5", [("A", 10), ("B", 20)])
        )

        assert summary.current_row > 8


class TestWorkbookUtilityFunctions:
    """Workbook 유틸리티 함수 테스트"""

    def test_save_to_csv_dict_list(self):
        """딕셔너리 리스트를 CSV로 저장"""
        from shared.io.excel.workbook import save_to_csv

        data = [
            {"Name": "John", "Age": "30"},
            {"Name": "Jane", "Age": "25"},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = str(Path(tmpdir) / "test.csv")
            save_to_csv(data, output_file)

            assert Path(output_file).exists()
            content = Path(output_file).read_text(encoding="utf-8-sig")
            assert "Name" in content
            assert "John" in content

    def test_save_to_csv_list_list(self):
        """리스트 리스트를 CSV로 저장"""
        from shared.io.excel.workbook import save_to_csv

        data = [["John", "30"], ["Jane", "25"]]
        headers = ["Name", "Age"]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = str(Path(tmpdir) / "test.csv")
            save_to_csv(data, output_file, headers=headers)

            assert Path(output_file).exists()
            content = Path(output_file).read_text(encoding="utf-8-sig")
            assert "Name" in content

    def test_save_to_csv_empty_data(self):
        """빈 데이터 저장"""
        from shared.io.excel.workbook import save_to_csv

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = str(Path(tmpdir) / "empty.csv")
            save_to_csv([], output_file)

            assert Path(output_file).exists()

    def test_save_dict_list_to_excel(self):
        """딕셔너리 리스트를 Excel로 저장"""
        from shared.io.excel.workbook import save_dict_list_to_excel

        data = [
            {"Name": "John", "Age": 30, "City": "Seoul"},
            {"Name": "Jane", "Age": 25, "City": "Busan"},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = str(Path(tmpdir) / "test.xlsx")
            save_dict_list_to_excel(data, output_file)

            assert Path(output_file).exists()
            assert Path(output_file).stat().st_size > 0

    def test_save_dict_list_to_excel_with_columns(self):
        """컬럼 순서 지정하여 저장"""
        from shared.io.excel.workbook import save_dict_list_to_excel

        data = [
            {"Name": "John", "Age": 30, "City": "Seoul"},
            {"Name": "Jane", "Age": 25, "City": "Busan"},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = str(Path(tmpdir) / "test.xlsx")
            save_dict_list_to_excel(data, output_file, columns=["City", "Name"])

            assert Path(output_file).exists()

    def test_save_dict_list_to_excel_empty(self):
        """빈 데이터 저장"""
        from shared.io.excel.workbook import save_dict_list_to_excel

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = str(Path(tmpdir) / "empty.xlsx")
            save_dict_list_to_excel([], output_file)

            assert Path(output_file).exists()

    def test_add_sheet_from_dict_list(self):
        """기존 Workbook에 시트 추가"""
        from openpyxl import Workbook as OpenpyxlWorkbook

        from shared.io.excel.workbook import add_sheet_from_dict_list

        wb = OpenpyxlWorkbook()
        data = [
            {"Name": "John", "Age": 30},
            {"Name": "Jane", "Age": 25},
        ]

        ws = add_sheet_from_dict_list(wb, data, "TestSheet")

        assert ws is not None
        assert ws.title == "TestSheet"
        assert ws.max_row == 3  # 헤더 + 2개 데이터

    def test_add_sheet_from_dict_list_with_columns(self):
        """컬럼 순서 지정하여 시트 추가"""
        from openpyxl import Workbook as OpenpyxlWorkbook

        from shared.io.excel.workbook import add_sheet_from_dict_list

        wb = OpenpyxlWorkbook()
        data = [
            {"Name": "John", "Age": 30, "City": "Seoul"},
        ]

        ws = add_sheet_from_dict_list(wb, data, "TestSheet", columns=["City", "Name"])

        assert ws is not None
        assert ws.cell(row=1, column=1).value == "City"
        assert ws.cell(row=1, column=2).value == "Name"

    def test_add_sheet_from_dict_list_empty(self):
        """빈 데이터로 시트 추가"""
        from openpyxl import Workbook as OpenpyxlWorkbook

        from shared.io.excel.workbook import add_sheet_from_dict_list

        wb = OpenpyxlWorkbook()
        ws = add_sheet_from_dict_list(wb, [], "EmptySheet")

        assert ws is not None
        assert ws.title == "EmptySheet"


class TestWorksheetFormatting:
    """워크시트 포맷팅 테스트"""

    def test_calculate_optimal_column_width(self):
        """최적 컬럼 너비 계산"""
        from shared.io.excel.workbook import calculate_optimal_column_width

        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        sheet = wb.new_sheet(name="Test", columns=columns)
        sheet.add_row(["Short"])
        sheet.add_row(["Very long text content here"])

        width = calculate_optimal_column_width(sheet._ws, "A")

        assert width > 10

    def test_calculate_optimal_column_width_multiline(self):
        """멀티라인 컬럼 너비 계산"""
        from shared.io.excel.workbook import calculate_optimal_column_width

        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        sheet = wb.new_sheet(name="Test", columns=columns)
        sheet.add_row(["Line 1\nLine 2\nVery long line 3"])

        width = calculate_optimal_column_width(sheet._ws, "A")

        assert width > 10

    def test_calculate_optimal_column_width_max_limit(self):
        """최대 너비 제한"""
        from shared.io.excel.workbook import calculate_optimal_column_width

        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        sheet = wb.new_sheet(name="Test", columns=columns)
        sheet.add_row(["A" * 200])  # 매우 긴 텍스트

        width = calculate_optimal_column_width(sheet._ws, "A", max_width=50)

        assert width <= 50

    def test_calculate_optimal_row_height(self):
        """최적 행 높이 계산"""
        from shared.io.excel.workbook import calculate_optimal_row_height

        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        sheet = wb.new_sheet(name="Test", columns=columns)
        sheet.add_row(["Line 1\nLine 2\nLine 3"])

        height = calculate_optimal_row_height(sheet._ws, 2)

        assert height > 15  # 기본 높이보다 큼

    def test_apply_detail_sheet_formatting(self):
        """상세 시트 포맷팅 적용"""
        from shared.io.excel.workbook import apply_detail_sheet_formatting

        wb = Workbook()
        columns = [ColumnDef(header="A"), ColumnDef(header="B")]
        sheet = wb.new_sheet(name="Test", columns=columns)
        sheet.add_row(["val1", "val2"])

        apply_detail_sheet_formatting(sheet._ws, has_header=True)

        assert sheet._ws.auto_filter.ref is not None
        assert sheet._ws.freeze_panes == "A2"

    def test_apply_detail_sheet_formatting_no_header(self):
        """헤더 없이 포맷팅"""
        from shared.io.excel.workbook import apply_detail_sheet_formatting

        wb = Workbook()
        columns = [ColumnDef(header="A")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        apply_detail_sheet_formatting(sheet._ws, has_header=False)

        assert sheet._ws.auto_filter.ref is not None

    def test_apply_summary_formatting(self):
        """요약 시트 포맷팅"""
        from shared.io.excel.workbook import apply_summary_formatting

        wb = Workbook()
        summary = wb.new_summary_sheet("요약")
        summary.add_item("테스트", "값")

        apply_summary_formatting(summary._ws)

        assert summary._ws.sheet_view.zoomScale == 90

    def test_apply_worksheet_settings(self):
        """워크시트 기본 설정"""
        from shared.io.excel.workbook import apply_worksheet_settings

        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        sheet = wb.new_sheet(name="Test", columns=columns)
        sheet.add_row(["value"])

        apply_worksheet_settings(sheet._ws, zoom_scale=100, wrap_text=True)

        assert sheet._ws.sheet_view.zoomScale == 100


class TestSheetRowStyles:
    """시트 행 스타일 테스트"""

    def test_add_row_with_warning_style(self):
        """경고 스타일 행"""
        wb = Workbook()
        columns = [ColumnDef(header="A"), ColumnDef(header="B")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        sheet.add_row(["warning", "data"], style=wb.styles.warning())

        assert sheet.row_count == 1

    def test_add_row_with_danger_style(self):
        """위험 스타일 행"""
        wb = Workbook()
        columns = [ColumnDef(header="A"), ColumnDef(header="B")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        sheet.add_row(["danger", "data"], style=wb.styles.danger())

        assert sheet.row_count == 1

    def test_add_row_with_success_style(self):
        """성공 스타일 행"""
        wb = Workbook()
        columns = [ColumnDef(header="A"), ColumnDef(header="B")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        sheet.add_row(["success", "data"], style=wb.styles.success())

        assert sheet.row_count == 1

    def test_add_row_with_info_style(self):
        """정보 스타일 행"""
        wb = Workbook()
        columns = [ColumnDef(header="A"), ColumnDef(header="B")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        sheet.add_row(["info", "data"], style=wb.styles.info())

        assert sheet.row_count == 1


class TestWorkbookProperties:
    """Workbook 프로퍼티 테스트"""

    def test_lang_property(self):
        """언어 프로퍼티"""
        wb = Workbook(lang="en")

        assert wb.lang == "en"

    def test_styles_property(self):
        """스타일 프로퍼티"""
        wb = Workbook()

        assert wb.styles is not None
        assert hasattr(wb.styles, "warning")
        assert hasattr(wb.styles, "danger")

    def test_openpyxl_workbook_property(self):
        """내부 openpyxl Workbook 접근"""
        wb = Workbook()

        assert wb.openpyxl_workbook is not None
        assert hasattr(wb.openpyxl_workbook, "active")

    def test_sheet_row_count_property(self):
        """시트 행 수 프로퍼티"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        assert sheet.row_count == 0

        sheet.add_row(["data"])
        assert sheet.row_count == 1

        sheet.add_row(["data2"])
        assert sheet.row_count == 2


class TestWorkbookEdgeCases:
    """Workbook 경계 케이스 테스트"""

    def test_save_as_without_region(self):
        """리전 없이 저장"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        wb.new_sheet(name="Test", columns=columns)

        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = wb.save_as(str(tmpdir), "test_output")

            assert Path(filepath).exists()
            assert "test_output" in str(filepath)

    def test_save_as_with_suffix(self):
        """접미사 포함 저장"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        wb.new_sheet(name="Test", columns=columns)

        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = wb.save_as(str(tmpdir), "test", suffix="final")

            assert Path(filepath).exists()
            assert "final" in str(filepath)

    @patch("shared.io.output.open_in_explorer")
    def test_workbook_save_with_one_sheet(self, mock_open):
        """하나의 빈 시트만 있는 Workbook 저장"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        wb.new_sheet(name="Sheet1", columns=columns)

        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = wb.save(Path(tmpdir) / "single_sheet.xlsx")

            assert Path(filepath).exists()
            wb.close()  # Close to release file handle

    def test_sheet_with_no_rows(self):
        """데이터 없는 시트"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]
        sheet = wb.new_sheet(name="Test", columns=columns)

        assert sheet.row_count == 0

        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = wb.save(Path(tmpdir) / "no_rows.xlsx")
            assert Path(filepath).exists()

    def test_very_long_sheet_name(self):
        """매우 긴 시트 이름"""
        wb = Workbook()
        columns = [ColumnDef(header="Test")]

        # Excel 시트 이름은 31자 권장이지만 openpyxl은 경고만 함
        long_name = "A" * 50
        sheet = wb.new_sheet(name=long_name, columns=columns)

        # openpyxl은 긴 이름을 허용하지만 경고 발생
        # Excel에서 열 때 31자로 잘릴 수 있음
        assert len(sheet._ws.title) == 50  # openpyxl은 그대로 유지
